import hashlib
import hmac
import json
import time
from io import BytesIO
from datetime import timedelta
from pathlib import Path
from urllib.parse import urlencode
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from core.models import (
    AccessGrant,
    GroupSheetConfiguration,
    TatActionTask,
    TatActionTaskRecipient,
    TatEscalationRule,
    TatGroupExceptionStatus,
    TatNotificationProcessorRun,
    TatPrivateAlertConnection,
    TatPrivateAlertConnectionEvent,
    TatResponsibilityAssignment,
    TatResponsibilityBackup,
    TatResponsibilityEvent,
    TatTaskRerouteEvent,
    TatTrackerCase,
    UserProfile,
)
from core.services.tat_notifications import (
    begin_notification_processor_run,
    connect_private_alerts,
    disconnect_private_alerts,
    dispatch_task,
    inbox_payload,
    issue_locator,
    mark_private_alert_seen,
    process_due_tasks,
    finish_notification_processor_run,
    refresh_group_exception,
    resolve_locator,
    resolve_assignment,
    reroute_pending_task,
    send_private_alert_test,
    synchronize_case_task,
    task_access_allowed,
)
from core.services.group_config import GroupRegistry


@override_settings(
    TELEGRAM_BOT_TOKEN='test-token', TELEGRAM_BOT_USERNAME='testbot',
    TAT_TRACKER_MINI_APP_SHORT_NAME='tattracker',
)
class TatPrivateTaskTests(TestCase):
    def setUp(self):
        self.group = GroupSheetConfiguration.objects.create(
            group_id='-100-private-tat', display_name='Private TAT',
            sheet_id='sheet-test', sheet_name='TRACKER-Business',
            workflow={
                'type': 'tat_tracker', 'tat_notification_mode': 'hybrid',
                'branches': ['Nakuru'], 'products': ['business'],
                'tat_targets_minutes': {
                    'business': {'total': 1000, 'stages': {'mpesa_to_admin': 100}},
                },
            },
        )
        self.primary = self._user('primary', '101')
        self.backup = self._user('backup', '102')
        self._grant(self.primary, 'BRO')
        self._grant(self.backup, 'BRO')
        TatEscalationRule.objects.create(
            group_configuration=self.group, branch='Nakuru', threshold_percent=100,
            routing_role=TatEscalationRule.ROUTE_RESPONSIBLE,
        )
        self.assignment = TatResponsibilityAssignment.objects.create(
            group_configuration=self.group, branch='Nakuru', role='BRO',
            primary_user=self.primary,
        )
        TatResponsibilityBackup.objects.create(
            assignment=self.assignment, user=self.backup, rank=1, threshold_percent=100,
        )
        self.case = self._case('JBL-BS-2026-PRIVATE-1')

    def _user(self, username, telegram_id):
        user = get_user_model().objects.create_user(username=username, password='test-password')
        UserProfile.objects.create(user=user, telegram_id=telegram_id)
        return user

    def _grant(self, user, role, *, branch='', product=''):
        return AccessGrant.objects.create(
            user=user, workflow='tat_tracker', role=role, branch=branch,
            product=product, group_configuration=self.group,
        )

    def _case(self, case_id, *, branch='Nakuru'):
        started = timezone.now()
        return TatTrackerCase.objects.create(
            group_id=self.group.group_id, case_id=case_id, product_key='business',
            product_label='Business', client_name='SYNTHETIC APPLICANT',
            national_id='12345678', primary_phone='254700000001', branch=branch,
            bro_name='Synthetic BRO', amount='25000', status='Active',
            current_stage='mpesa_to_admin', stage_values={'created': started.isoformat()},
            stage_target_snapshots={'mpesa_to_admin': {
                'target_minutes': '100', 'started_at': started.isoformat(),
            }},
        )

    def _signed_init_data(self, telegram_id='101', username='primary'):
        pairs = {
            'auth_date': str(int(time.time())),
            'user': json.dumps({'id': int(telegram_id), 'username': username}),
        }
        check = '\n'.join(f'{key}={value}' for key, value in sorted(pairs.items()))
        secret = hmac.new(b'WebAppData', b'test-token', hashlib.sha256).digest()
        pairs['hash'] = hmac.new(secret, check.encode('utf-8'), hashlib.sha256).hexdigest()
        return urlencode(pairs)

    def test_assignment_routes_primary_and_ranked_backup_without_granting_access(self):
        task = synchronize_case_task(self.group, self.case)

        self.assertEqual(task.assignment, self.assignment)
        recipients = list(task.recipients.order_by('rank'))
        self.assertEqual([(row.user, row.kind) for row in recipients], [
            (self.primary, TatActionTaskRecipient.KIND_PRIMARY),
            (self.backup, TatActionTaskRecipient.KIND_BACKUP),
        ])
        self.assertLessEqual(recipients[0].deliver_after, timezone.now())
        self.assertGreater(recipients[1].deliver_after, timezone.now())

    def test_stage_override_derives_role_for_user_with_multiple_roles(self):
        multi_role = self._user('multi-role', '103')
        self._grant(multi_role, 'SECRETARY')
        self._grant(multi_role, 'FINANCE')

        assignment = TatResponsibilityAssignment.objects.create(
            group_configuration=self.group,
            branch='Nakuru',
            role='SECRETARY',  # A stale/manipulated form value must not win.
            product_key='business',
            stage_key='disbursement',
            primary_user=multi_role,
        )

        self.assertEqual(assignment.role, 'FINANCE')

    def test_superuser_is_not_silently_enrolled_without_tat_access_grant(self):
        superuser = get_user_model().objects.create_superuser(
            username='technical-root', password='test-password', email='root@example.test',
        )
        assignment = TatResponsibilityAssignment(
            group_configuration=self.group, branch='Nakuru', role='BRO',
            primary_user=superuser,
        )

        with self.assertRaises(ValidationError) as raised:
            assignment.full_clean()

        self.assertIn('primary_user', raised.exception.message_dict)

    def test_backup_rank_requires_strictly_later_escalation_threshold(self):
        later_backup = self._user('later-backup', '104')
        self._grant(later_backup, 'BRO')
        TatEscalationRule.objects.create(
            group_configuration=self.group, branch='Nakuru', threshold_percent=150,
            routing_role=TatEscalationRule.ROUTE_RESPONSIBLE,
        )
        duplicate_threshold = TatResponsibilityBackup(
            assignment=self.assignment, user=later_backup, rank=2, threshold_percent=100,
        )
        with self.assertRaises(ValidationError):
            duplicate_threshold.full_clean()

        later = TatResponsibilityBackup.objects.create(
            assignment=self.assignment, user=later_backup, rank=2, threshold_percent=150,
        )
        self.assertEqual(later.threshold_percent, 150)

    def test_product_stage_override_precedes_general_role_roster(self):
        specialist = self._user('specialist', '105')
        self._grant(specialist, 'BRO', product='business')
        override = TatResponsibilityAssignment.objects.create(
            group_configuration=self.group, branch='Nakuru', role='BRO',
            product_key='business', stage_key='mpesa_to_admin', primary_user=specialist,
        )

        resolved = resolve_assignment(
            group=self.group, case=self.case, role='BRO', stage_key='mpesa_to_admin',
        )

        self.assertEqual(resolved, override)

    def test_runtime_same_tier_ambiguity_fails_closed(self):
        specialist = self._user('ambiguous-specialist', '106')
        self._grant(specialist, 'BRO', product='business')
        first = TatResponsibilityAssignment.objects.create(
            group_configuration=self.group, branch='Nakuru', role='BRO',
            product_key='business', stage_key='mpesa_to_admin', primary_user=specialist,
        )
        # Simulate a legacy/corrupt case-insensitive duplicate that bypassed
        # model validation; routing must not choose one arbitrarily.
        second = TatResponsibilityAssignment.objects.create(
            group_configuration=self.group, branch='Elsewhere', role='BRO',
            product_key='business', stage_key='mpesa_to_admin', primary_user=specialist,
        )
        TatResponsibilityAssignment.objects.filter(pk=second.pk).update(branch='nakuru')
        first.refresh_from_db()

        self.assertIsNone(resolve_assignment(
            group=self.group, case=self.case, role='BRO', stage_key='mpesa_to_admin',
        ))

    def test_centralized_admin_workspace_shows_roles_stages_and_grants(self):
        superuser = get_user_model().objects.create_superuser(
            username='workspace-admin', password='test-password', email='workspace@example.test',
        )
        self.client.force_login(superuser)

        response = self.client.get(reverse('admin:core_tatresponsibilityassignment_changelist'), {
            'workspace_group': str(self.group.pk),
            'workspace_branch': 'Nakuru',
            'workspace_product': 'business',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'TAT access &amp; responsibilities')
        self.assertContains(response, 'Advanced stage overrides')
        self.assertContains(response, 'mpesa_to_admin')
        self.assertContains(response, 'primary')

    def test_workspace_projects_scheduled_active_roster_instead_of_offering_duplicate(self):
        superuser = get_user_model().objects.create_superuser(
            username='scheduled-workspace-admin', password='test-password',
            email='scheduled-workspace@example.test',
        )
        self.assignment.effective_from = timezone.now() + timedelta(days=1)
        self.assignment.save(update_fields=['effective_from'])
        self.client.force_login(superuser)

        response = self.client.get(reverse('admin:core_tatresponsibilityassignment_changelist'), {
            'workspace_group': str(self.group.pk),
            'workspace_branch': 'Nakuru',
            'workspace_product': '',
        })

        bro_row = next(row for row in response.context['responsibility_role_rows'] if row['role'] == 'BRO')
        self.assertEqual(bro_row['roster'], self.assignment)
        self.assertTrue(bro_row['roster_state'].startswith('Scheduled from'))
        self.assertContains(response, 'Edit roster')

    def test_duplicate_active_roster_has_plain_language_form_error(self):
        superuser = get_user_model().objects.create_superuser(
            username='duplicate-roster-admin', password='test-password',
            email='duplicate-roster@example.test',
        )
        self.client.force_login(superuser)
        effective = timezone.localtime()

        response = self.client.post(reverse('admin:core_tatresponsibilityassignment_add'), {
            'group_configuration': str(self.group.pk),
            'branch': 'Nakuru',
            'role': 'BRO',
            'product_key': '',
            'stage_key': '',
            'primary_user': str(self.primary.pk),
            'change_reason': 'Attempted duplicate for regression coverage.',
            'active': 'on',
            'effective_from_0': effective.date().isoformat(),
            'effective_from_1': effective.time().replace(microsecond=0).isoformat(),
            'backups-TOTAL_FORMS': '0',
            'backups-INITIAL_FORMS': '0',
            'backups-MIN_NUM_FORMS': '0',
            'backups-MAX_NUM_FORMS': '1000',
            '_save': 'Save',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'An active roster already exists for this exact role and scope.')
        self.assertEqual(TatResponsibilityAssignment.objects.filter(
            group_configuration=self.group, branch='Nakuru', role='BRO', active=True,
        ).count(), 1)

    def test_responsibility_add_form_lists_exact_scope_users(self):
        superuser = get_user_model().objects.create_superuser(
            username='form-admin', password='test-password', email='form@example.test',
        )
        self.client.force_login(superuser)

        response = self.client.get(reverse('admin:core_tatresponsibilityassignment_add'), {
            'group_configuration': str(self.group.pk),
            'branch': 'Nakuru',
            'role': 'BRO',
            'product_key': '',
        })

        self.assertEqual(response.status_code, 200)
        field = response.context['adminform'].form.fields['primary_user']
        self.assertQuerySetEqual(
            field.queryset.order_by('username'),
            [self.backup, self.primary],
            transform=lambda user: user,
        )
        self.assertContains(response, '2 eligible active users match this exact TAT scope.')
        self.assertContains(response, reverse('admin:core_tatresponsibilityassignment_eligible_users'))

    def test_superuser_can_open_guided_control_center_and_register(self):
        superuser = get_user_model().objects.create_superuser(
            username='control-admin', password='test-password', email='control@example.test',
        )
        self.client.force_login(superuser)

        setup = self.client.get(reverse('admin:core_tat_control_center', args=[self.group.pk]))
        register = self.client.get(reverse('admin:core_tat_register', args=[self.group.pk]))

        self.assertEqual(setup.status_code, 200)
        self.assertContains(setup, 'Guided TAT administration')
        self.assertContains(setup, 'Built-in register and Sheet projection')
        self.assertEqual(register.status_code, 200)
        self.assertContains(register, self.case.case_id)

    def test_non_superuser_cannot_open_guided_control_center(self):
        self.primary.is_staff = True
        self.primary.save(update_fields=['is_staff'])
        self.client.force_login(self.primary)

        response = self.client.get(reverse('admin:core_tat_control_center', args=[self.group.pk]))

        self.assertEqual(response.status_code, 403)

    def test_register_export_uses_privacy_allowlist(self):
        from openpyxl import load_workbook
        from core.models import ComplianceAuditChainState
        from core.services.tat_register import export_xlsx

        superuser = get_user_model().objects.create_superuser(
            username='export-admin', password='test-password', email='export@example.test',
        )
        ComplianceAuditChainState.objects.get_or_create(singleton=1)
        content = export_xlsx(
            group=self.group, actor=superuser, request_id='tat-export-test-1', filters={},
        )
        sheet = load_workbook(BytesIO(content), read_only=True).active
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]

        self.assertIn('Case reference', headers)
        self.assertNotIn('Applicant name', headers)
        self.assertNotIn('National ID', headers)
        self.assertNotIn('Phone', headers)
        self.assertNotIn('Remarks', headers)

    def test_eligible_user_lookup_supports_multi_role_and_filters_exact_scope(self):
        superuser = get_user_model().objects.create_superuser(
            username='lookup-admin', password='test-password', email='lookup@example.test',
        )
        multi_role = self._user('multi-scope', '107')
        self._grant(multi_role, 'BRO')
        self._grant(multi_role, 'FINANCE')
        wrong_role = self._user('finance-only', '108')
        self._grant(wrong_role, 'FINANCE')
        product_only = self._user('product-only', '109')
        self._grant(product_only, 'BRO', product='business')
        self.client.force_login(superuser)
        url = reverse('admin:core_tatresponsibilityassignment_eligible_users')

        all_products = self.client.get(url, {
            'group_configuration': str(self.group.pk),
            'branch': 'Nakuru',
            'role': 'BRO',
            'product_key': '',
        })
        self.assertEqual(all_products.status_code, 200)
        all_product_ids = [row['id'] for row in all_products.json()['users']]
        self.assertIn(str(multi_role.pk), all_product_ids)
        self.assertNotIn(str(wrong_role.pk), all_product_ids)
        self.assertNotIn(str(product_only.pk), all_product_ids)
        self.assertEqual(all_product_ids.count(str(multi_role.pk)), 1)

        business = self.client.get(url, {
            'group_configuration': str(self.group.pk),
            'branch': 'Nakuru',
            'role': 'BRO',
            'product_key': 'business',
        })
        self.assertEqual(business.status_code, 200)
        business_ids = [row['id'] for row in business.json()['users']]
        self.assertIn(str(multi_role.pk), business_ids)
        self.assertIn(str(product_only.pk), business_ids)
        self.assertNotIn(str(wrong_role.pk), business_ids)

    def test_eligible_user_lookup_explains_incomplete_scope(self):
        superuser = get_user_model().objects.create_superuser(
            username='empty-lookup-admin', password='test-password', email='empty@example.test',
        )
        self.client.force_login(superuser)
        response = self.client.get(
            reverse('admin:core_tatresponsibilityassignment_eligible_users'),
            {'group_configuration': str(self.group.pk), 'branch': 'Nakuru'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['users'], [])
        self.assertIn('role', response.json()['message'])

    def test_admin_stage_override_derives_role_and_writes_audit_event(self):
        superuser = get_user_model().objects.create_superuser(
            username='routing-admin', password='test-password', email='routing@example.test',
        )
        self.client.force_login(superuser)
        response = self.client.post(reverse('admin:core_tatresponsibilityassignment_add'), {
            'group_configuration': str(self.group.pk),
            'branch': 'Nakuru',
            'role': 'FINANCE',
            'product_key': 'business',
            'stage_key': 'mpesa_to_admin',
            'primary_user': str(self.primary.pk),
            'active': 'on',
            'effective_from_0': timezone.localdate().isoformat(),
            'effective_from_1': timezone.localtime().strftime('%H:%M:%S'),
            'effective_until_0': '',
            'effective_until_1': '',
            'change_reason': 'Assign the synthetic BRO for routing verification.',
            'backups-TOTAL_FORMS': '0',
            'backups-INITIAL_FORMS': '0',
            'backups-MIN_NUM_FORMS': '0',
            'backups-MAX_NUM_FORMS': '1000',
            '_save': 'Save',
        })

        self.assertEqual(
            response.status_code, 302,
            getattr(response.context.get('adminform'), 'form', None).errors
            if response.context else response.content[:1000],
        )
        assignment = TatResponsibilityAssignment.objects.get(
            group_configuration=self.group, product_key='business', stage_key='mpesa_to_admin',
        )
        self.assertEqual(assignment.role, 'BRO')
        event = TatResponsibilityEvent.objects.get(assignment=assignment)
        self.assertEqual(event.action, TatResponsibilityEvent.ACTION_CREATED)
        self.assertEqual(event.actor, superuser)
        self.assertEqual(event.after_snapshot['role'], 'BRO')

    @patch('core.services.tat_notifications._telegram_request', return_value={'message_id': 77})
    def test_known_unconnected_primary_immediately_soft_escalates_to_backup(self, telegram):
        TatPrivateAlertConnection.objects.create(
            user=self.primary, status=TatPrivateAlertConnection.STATUS_UNCONNECTED,
        )
        TatPrivateAlertConnection.objects.create(
            user=self.backup, status=TatPrivateAlertConnection.STATUS_CONNECTED,
        )
        task = synchronize_case_task(self.group, self.case)

        dispatch_task(task.pk)

        primary = task.recipients.get(user=self.primary)
        backup = task.recipients.get(user=self.backup)
        self.assertEqual(primary.delivery_state, TatActionTaskRecipient.DELIVERY_UNREACHABLE)
        self.assertEqual(backup.delivery_state, TatActionTaskRecipient.DELIVERY_DELIVERED)
        self.assertEqual(telegram.call_count, 1)

    @patch('core.services.tat_notifications._telegram_request', return_value={'message_id': 78})
    def test_concurrent_safe_dispatch_does_not_repeat_a_delivered_prompt(self, telegram):
        TatPrivateAlertConnection.objects.create(
            user=self.primary, status=TatPrivateAlertConnection.STATUS_CONNECTED,
        )
        task = synchronize_case_task(self.group, self.case)

        dispatch_task(task.pk)
        dispatch_task(task.pk)

        self.assertEqual(telegram.call_count, 1)
        self.assertEqual(
            task.recipients.get(user=self.primary).delivery_state,
            TatActionTaskRecipient.DELIVERY_DELIVERED,
        )
        self.assertEqual(
            task.recipients.get(user=self.primary).delivery_attempts,
            1,
        )
        self.assertEqual(
            TatPrivateAlertConnectionEvent.objects.filter(
                event_type=TatPrivateAlertConnectionEvent.EVENT_DELIVERY_SUCCEEDED,
            ).count(),
            1,
        )

    @patch('core.services.tat_notifications._telegram_request', return_value={'message_id': 781})
    def test_due_pending_recipient_is_recovered_by_notification_processor(self, telegram):
        TatPrivateAlertConnection.objects.create(
            user=self.primary, status=TatPrivateAlertConnection.STATUS_CONNECTED,
        )
        task = synchronize_case_task(self.group, self.case)

        processed = process_due_tasks(limit=100)

        recipient = task.recipients.get(user=self.primary)
        self.assertEqual(processed, 1)
        self.assertEqual(recipient.delivery_state, TatActionTaskRecipient.DELIVERY_DELIVERED)
        self.assertEqual(recipient.delivery_attempts, 1)
        self.assertEqual(telegram.call_count, 1)

    def test_notification_processor_records_success_and_releases_lock(self):
        task = synchronize_case_task(self.group, self.case)
        TatPrivateAlertConnection.objects.create(
            user=self.primary, status=TatPrivateAlertConnection.STATUS_CONNECTED,
        )

        with patch('core.services.tat_notifications._telegram_request', return_value={'message_id': 782}):
            call_command('process_tat_notifications', '--limit', '100')

        run = TatNotificationProcessorRun.objects.get(status=TatNotificationProcessorRun.STATUS_SUCCEEDED)
        self.assertIsNone(run.active_lock_key)
        self.assertEqual(run.processed_task_count, 1)
        self.assertEqual(run.overdue_recipient_count, 0)
        self.assertEqual(
            task.recipients.get(user=self.primary).delivery_state,
            TatActionTaskRecipient.DELIVERY_DELIVERED,
        )

    def test_notification_processor_skips_an_overlapping_run(self):
        active, acquired = begin_notification_processor_run()

        second, second_acquired = begin_notification_processor_run()

        self.assertTrue(acquired)
        self.assertFalse(second_acquired)
        self.assertEqual(second.status, TatNotificationProcessorRun.STATUS_SKIPPED_OVERLAP)
        finish_notification_processor_run(active)
        active.refresh_from_db()
        self.assertIsNone(active.active_lock_key)

    def test_manual_notification_run_retains_actor_and_reason_snapshots(self):
        actor = get_user_model().objects.create_superuser(
            username='manual-run-root', email='manual@example.invalid', password='password',
        )

        run, acquired = begin_notification_processor_run(
            trigger_source=TatNotificationProcessorRun.TRIGGER_ADMIN,
            actor=actor,
            reason='Verify private delivery after roster maintenance.',
            request_id='manual-run-request-0001',
        )
        replay, replay_acquired = begin_notification_processor_run(
            trigger_source=TatNotificationProcessorRun.TRIGGER_ADMIN,
            actor=actor,
            reason='Verify private delivery after roster maintenance.',
            request_id='manual-run-request-0001',
        )

        self.assertTrue(acquired)
        self.assertFalse(replay_acquired)
        self.assertEqual(replay.pk, run.pk)
        self.assertEqual(run.trigger_source, TatNotificationProcessorRun.TRIGGER_ADMIN)
        self.assertEqual(run.triggered_by_username_snapshot, actor.username)
        self.assertEqual(run.trigger_reason, 'Verify private delivery after roster maintenance.')
        finish_notification_processor_run(run)

    @patch(
        'core.management.commands.process_tat_notifications.process_due_tasks',
        side_effect=RuntimeError('synthetic runner failure'),
    )
    def test_notification_processor_records_a_privacy_safe_failure(self, _process):
        with self.assertRaises(CommandError):
            call_command('process_tat_notifications')

        run = TatNotificationProcessorRun.objects.get(status=TatNotificationProcessorRun.STATUS_FAILED)
        self.assertIsNone(run.active_lock_key)
        self.assertEqual(run.error_code, 'RuntimeError')
        self.assertNotIn('synthetic runner failure', run.error_message)

    @override_settings(TAT_NOTIFICATION_PROCESSOR_LOCK_SECONDS=60)
    def test_notification_processor_recovers_an_expired_lock(self):
        stale = TatNotificationProcessorRun.objects.create(
            status=TatNotificationProcessorRun.STATUS_RUNNING,
            active_lock_key=TatNotificationProcessorRun.LOCK_KEY,
            started_at=timezone.now() - timedelta(minutes=2),
        )

        current, acquired = begin_notification_processor_run()

        self.assertTrue(acquired)
        stale.refresh_from_db()
        self.assertEqual(stale.status, TatNotificationProcessorRun.STATUS_FAILED)
        self.assertEqual(stale.error_code, 'stale-lock-recovered')
        finish_notification_processor_run(current)

    @patch('core.services.tat_notifications._telegram_request', return_value={'message_id': 79})
    def test_connect_private_alerts_replays_same_request_without_second_message(self, telegram):
        first = connect_private_alerts(self.primary, request_id='connect-request-0001')
        second = connect_private_alerts(self.primary, request_id='connect-request-0001')

        self.assertTrue(first['connected'])
        self.assertTrue(second['connected'])
        self.assertEqual(telegram.call_count, 1)
        self.assertEqual(
            TatPrivateAlertConnectionEvent.objects.filter(
                event_type=TatPrivateAlertConnectionEvent.EVENT_CONNECTED,
            ).count(),
            1,
        )

    def test_disconnect_is_idempotent_and_session_bootstrap_does_not_reconnect(self):
        connection = TatPrivateAlertConnection.objects.create(
            user=self.primary, status=TatPrivateAlertConnection.STATUS_CONNECTED,
            connected_at=timezone.now(),
        )

        first = disconnect_private_alerts(self.primary, request_id='disconnect-request-0001')
        second = disconnect_private_alerts(self.primary, request_id='disconnect-request-0001')
        mark_private_alert_seen(self.primary, allows_write=True)

        connection.refresh_from_db()
        self.assertFalse(first['connected'])
        self.assertFalse(second['connected'])
        self.assertEqual(connection.status, TatPrivateAlertConnection.STATUS_DISCONNECTED)
        self.assertIsNotNone(connection.disconnected_at)
        self.assertEqual(
            TatPrivateAlertConnectionEvent.objects.filter(
                event_type=TatPrivateAlertConnectionEvent.EVENT_DISCONNECTED,
            ).count(),
            1,
        )

    @patch('core.services.tat_notifications._telegram_request', return_value={'message_id': 81})
    def test_explicit_connect_restores_a_disconnected_connection(self, telegram):
        TatPrivateAlertConnection.objects.create(
            user=self.primary, status=TatPrivateAlertConnection.STATUS_DISCONNECTED,
            disconnected_at=timezone.now(),
        )

        result = connect_private_alerts(self.primary, request_id='reconnect-request-0001')

        connection = TatPrivateAlertConnection.objects.get(user=self.primary)
        self.assertTrue(result['connected'])
        self.assertEqual(connection.status, TatPrivateAlertConnection.STATUS_CONNECTED)
        self.assertIsNone(connection.disconnected_at)
        self.assertEqual(telegram.call_count, 1)

    @patch('core.services.tat_notifications._telegram_request', side_effect=RuntimeError('synthetic failure'))
    def test_failed_connect_is_persisted_and_same_request_is_not_resent(self, telegram):
        with self.assertRaises(RuntimeError):
            connect_private_alerts(self.primary, request_id='failed-connect-0001')

        replay = connect_private_alerts(self.primary, request_id='failed-connect-0001')

        connection = TatPrivateAlertConnection.objects.get(user=self.primary)
        event = TatPrivateAlertConnectionEvent.objects.get(request_id='failed-connect-0001')
        self.assertFalse(replay['connected'])
        self.assertEqual(connection.status, TatPrivateAlertConnection.STATUS_TEMPORARY_FAILURE)
        self.assertEqual(event.event_type, TatPrivateAlertConnectionEvent.EVENT_CONNECT_FAILED)
        self.assertEqual(event.detail_code, 'RuntimeError')
        self.assertEqual(telegram.call_count, 1)

    @patch('core.services.tat_notifications._telegram_request', return_value={'message_id': 82})
    def test_admin_test_message_is_idempotent_and_audited(self, telegram):
        connection = TatPrivateAlertConnection.objects.create(
            user=self.primary, status=TatPrivateAlertConnection.STATUS_CONNECTED,
        )
        actor = get_user_model().objects.create_superuser(
            username='alert-test-root', email='root@example.invalid', password='password',
        )

        send_private_alert_test(
            self.primary, actor=actor, request_id='admin-alert-test-0001',
        )
        send_private_alert_test(
            self.primary, actor=actor, request_id='admin-alert-test-0001',
        )

        event = TatPrivateAlertConnectionEvent.objects.get(
            event_type=TatPrivateAlertConnectionEvent.EVENT_TEST_SUCCEEDED,
        )
        self.assertEqual(event.connection, connection)
        self.assertEqual(event.actor_username_snapshot, actor.username)
        self.assertEqual(telegram.call_count, 1)

    def test_miniapp_private_alert_button_is_a_persistent_toggle(self):
        source = Path('core/static/miniapp/tat_tracker.js').read_text(encoding='utf-8')
        template = Path('core/templates/tat_tracker/app.html').read_text(encoding='utf-8')

        self.assertIn("button.dataset.connected = data.connected ? 'true' : 'false'", source)
        self.assertIn("'Disconnect private alerts'", source)
        self.assertIn("'/api/tat-tracker/private-alerts/disconnect/'", source)
        self.assertIn('?v=48', template)

    @patch('core.services.tat_notifications._telegram_request', return_value={'message_id': 83})
    def test_superuser_can_send_confirmed_test_from_connection_admin(self, telegram):
        connection = TatPrivateAlertConnection.objects.create(
            user=self.primary, status=TatPrivateAlertConnection.STATUS_CONNECTED,
        )
        actor = get_user_model().objects.create_superuser(
            username='connection-admin-root', email='connection@example.invalid', password='password',
        )
        self.client.force_login(actor)
        url = reverse('admin:core_tatprivatealertconnection_test', args=[connection.pk])

        preview = self.client.get(url)
        response = self.client.post(url, {
            'confirmation_phrase': 'SEND TEST',
            'request_id': 'admin-test-view-0001',
        })

        self.assertEqual(preview.status_code, 200)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(telegram.call_count, 1)
        self.assertTrue(TatPrivateAlertConnectionEvent.objects.filter(
            request_id='admin-test-view-0001',
            actor_username_snapshot=actor.username,
        ).exists())

    def test_superuser_can_confirm_manual_processor_run_from_admin(self):
        actor = get_user_model().objects.create_superuser(
            username='processor-admin-root', email='processor@example.invalid', password='password',
        )
        self.client.force_login(actor)
        url = reverse('admin:core_tatnotificationprocessorrun_run_now')

        preview = self.client.get(url)
        with patch('core.services.tat_notifications.process_due_tasks', return_value=0) as run_due:
            response = self.client.post(url, {
                'confirmation_phrase': 'RUN PRIVATE ALERTS',
                'reason': 'Verify delivery after scheduler maintenance.',
                'limit': '25',
                'request_id': 'manual-admin-view-0001',
            })

        self.assertEqual(preview.status_code, 200)
        self.assertEqual(response.status_code, 302)
        run_due.assert_called_once_with(limit=25)
        run = TatNotificationProcessorRun.objects.get(
            trigger_source=TatNotificationProcessorRun.TRIGGER_ADMIN,
        )
        self.assertEqual(run.triggered_by_username_snapshot, actor.username)
        self.assertEqual(run.request_id, 'manual-admin-view-0001')
        self.assertEqual(run.status, TatNotificationProcessorRun.STATUS_SUCCEEDED)

    def test_same_case_stage_revision_is_idempotent(self):
        first = synchronize_case_task(self.group, self.case)
        second = synchronize_case_task(self.group, self.case)

        self.assertEqual(first.pk, second.pk)
        self.assertEqual(TatActionTask.objects.count(), 1)

    def test_revision_change_supersedes_old_task_and_revokes_old_locator(self):
        old = synchronize_case_task(self.group, self.case)
        token = issue_locator(old, self.primary)
        self.case.workflow_revision += 1
        self.case.save(update_fields=['workflow_revision', 'updated_at'])

        current = synchronize_case_task(self.group, self.case)

        old.refresh_from_db()
        locator = resolve_locator(token)
        self.assertEqual(old.status, TatActionTask.STATUS_SUPERSEDED)
        self.assertEqual(old.superseded_by, current)
        self.assertIsNotNone(locator.revoked_at)

    def test_explicit_reroute_revokes_removed_recipient_and_increments_generation(self):
        task = synchronize_case_task(self.group, self.case)
        token = issue_locator(task, self.primary)
        replacement = self._user('replacement', '110')
        self._grant(replacement, 'BRO')
        self.assignment.primary_user = replacement
        self.assignment.save(update_fields=['primary_user', 'updated_at'])
        superuser = get_user_model().objects.create_superuser(
            username='reroute-admin', password='test-password', email='reroute@example.test',
        )

        rerouted = reroute_pending_task(
            task=task, actor=superuser,
            reason='The approved primary responsibility owner changed.',
            request_id='reroute-test-1',
        )

        locator = resolve_locator(token)
        self.assertEqual(rerouted.routing_generation, 2)
        self.assertIsNotNone(locator.revoked_at)
        self.assertFalse(task_access_allowed(rerouted, self.primary))
        self.assertTrue(task_access_allowed(rerouted, replacement))
        self.assertTrue(TatTaskRerouteEvent.objects.filter(
            task=rerouted, generation_before=1, generation_after=2,
        ).exists())

    def test_stale_locator_redirects_authorized_recipient_to_current_revision(self):
        old = synchronize_case_task(self.group, self.case)
        token = issue_locator(old, self.primary)
        self.case.workflow_revision += 1
        self.case.save(update_fields=['workflow_revision', 'updated_at'])
        current = synchronize_case_task(self.group, self.case)
        GroupRegistry._instance = None

        response = self.client.post(
            reverse('tat_tracker_task_resolve'),
            data=json.dumps({
                'task_token': f'tt_{token}',
                'init_data': self._signed_init_data(),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['data']['link_status'], 'expired')
        self.assertEqual(response.json()['data']['task_id'], str(current.pk))

    def test_shadow_mode_records_routing_without_calling_telegram(self):
        workflow = dict(self.group.workflow)
        workflow['tat_notification_mode'] = 'shadow'
        self.group.workflow = workflow
        self.group.save(update_fields=['workflow', 'updated_at'])
        task = synchronize_case_task(self.group, self.case)

        with patch('core.services.tat_notifications._telegram_request') as telegram:
            dispatch_task(task.pk)

        self.assertFalse(telegram.called)
        self.assertEqual(
            task.recipients.get(user=self.primary).delivery_state,
            TatActionTaskRecipient.DELIVERY_SHADOW,
        )

    def test_locator_is_32_chars_hash_only_and_expires_in_72_hours(self):
        task = synchronize_case_task(self.group, self.case)
        token = issue_locator(task, self.primary)
        locator = resolve_locator(token)

        self.assertEqual(len(token), 32)
        self.assertNotEqual(locator.token_hash, token)
        self.assertGreater(locator.expires_at, timezone.now() + timedelta(hours=71))
        self.assertNotIn(token, repr(locator.__dict__))

    def test_inbox_is_personal_and_reports_unread_count(self):
        task = synchronize_case_task(self.group, self.case)

        primary = inbox_payload(self.primary, group=self.group)
        backup = inbox_payload(self.backup, group=self.group)

        self.assertEqual(primary['unread_count'], 1)
        self.assertEqual(backup['unread_count'], 1)
        self.assertEqual(primary['items'][0]['task_id'], str(task.pk))

    @patch('core.services.tat_notifications._telegram_request', return_value={'message_id': 80})
    def test_revoked_primary_scope_hides_task_and_soft_escalates_delivery(self, telegram):
        task = synchronize_case_task(self.group, self.case)
        AccessGrant.objects.filter(user=self.primary, workflow='tat_tracker').update(active=False)
        TatPrivateAlertConnection.objects.create(
            user=self.primary, status=TatPrivateAlertConnection.STATUS_CONNECTED,
        )
        TatPrivateAlertConnection.objects.create(
            user=self.backup, status=TatPrivateAlertConnection.STATUS_CONNECTED,
        )

        self.assertEqual(inbox_payload(self.primary, group=self.group)['total'], 0)
        dispatch_task(task.pk)

        self.assertEqual(
            task.recipients.get(user=self.primary).delivery_state,
            TatActionTaskRecipient.DELIVERY_UNREACHABLE,
        )
        self.assertEqual(
            task.recipients.get(user=self.backup).delivery_state,
            TatActionTaskRecipient.DELIVERY_DELIVERED,
        )
        self.assertEqual(telegram.call_count, 1)

    def test_direct_locator_endpoint_authenticates_and_opens_exact_task(self):
        task = synchronize_case_task(self.group, self.case)
        token = issue_locator(task, self.primary)
        GroupRegistry._instance = None

        response = self.client.post(
            reverse('tat_tracker_task_resolve'),
            data=json.dumps({
                'task_token': f'tt_{token}',
                'init_data': self._signed_init_data(),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['data']['case_id'], self.case.case_id)
        self.assertEqual(response.json()['data']['stage_key'], 'mpesa_to_admin')
        task.recipients.get(user=self.primary).refresh_from_db()
        self.assertEqual(
            task.recipients.get(user=self.primary).inbox_status,
            TatActionTaskRecipient.INBOX_READ,
        )

    def test_direct_locator_rejects_user_outside_recipient_and_scope(self):
        task = synchronize_case_task(self.group, self.case)
        token = issue_locator(task, self.primary)
        outsider = self._user('outsider', '999')
        self._grant(outsider, 'CA')
        GroupRegistry._instance = None

        response = self.client.post(
            reverse('tat_tracker_task_resolve'),
            data=json.dumps({
                'task_token': f'tt_{token}',
                'init_data': self._signed_init_data('999', 'outsider'),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)

    @patch('core.services.tat_notifications._telegram_request', return_value={})
    def test_unreachable_cases_remain_admin_diagnostics_without_group_message(self, telegram):
        isolated_group = GroupSheetConfiguration.objects.create(
            group_id='-100-unassigned-tat', display_name='Unassigned TAT',
            sheet_id='sheet-unassigned', sheet_name='TRACKER-Business',
            workflow={'type': 'tat_tracker', 'tat_notification_mode': 'hybrid'},
        )
        first = self._case('JBL-BS-2026-UNASSIGNED-1', branch='Mombasa')
        second = self._case('JBL-BS-2026-UNASSIGNED-2', branch='Mombasa')
        first.group_id = isolated_group.group_id
        second.group_id = isolated_group.group_id
        first.save(update_fields=['group_id'])
        second.save(update_fields=['group_id'])
        first_task = synchronize_case_task(isolated_group, first)
        second_task = synchronize_case_task(isolated_group, second)
        self.assertTrue(first_task.recipient_snapshot['delivery_exception'])
        self.assertTrue(second_task.recipient_snapshot['delivery_exception'])

        refresh_group_exception(isolated_group.pk, role='BRO')
        refresh_group_exception(isolated_group.pk, role='BRO')

        status = TatGroupExceptionStatus.objects.get(
            group_configuration=isolated_group, responsible_role='BRO',
        )
        self.assertEqual(status.unresolved_count, 2)
        self.assertEqual(status.telegram_message_id, '')
        telegram.assert_not_called()

    @patch('core.services.tat_notifications._telegram_request', return_value={})
    def test_processor_retires_legacy_group_exception_message_without_replacement(self, telegram):
        status = TatGroupExceptionStatus.objects.create(
            group_configuration=self.group,
            responsible_role='BRO',
            unresolved_count=1,
            oldest_task_at=timezone.now() - timedelta(minutes=10),
            active=True,
            telegram_message_id='501',
        )

        processed = process_due_tasks(limit=100)

        status.refresh_from_db()
        self.assertEqual(processed, 0)
        self.assertEqual(status.telegram_message_id, '')
        self.assertEqual(status.last_error, '')
        telegram.assert_called_once_with('deleteMessage', {
            'chat_id': self.group.group_id,
            'message_id': '501',
        })
