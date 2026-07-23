from datetime import date
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from django.test import TestCase
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlsxImage
from PIL import Image as PilImage

from core.services.requisition import generate_requisition_excel


class RequisitionTemplateGenerationTests(TestCase):
    def farmer(self, **overrides):
        data = {
            'customer_name': 'Mary Wanjiku',
            'primary_phone': '254712345678',
            'national_id': '12345678',
            'credit_decision': 'Approved',
            'county': 'Nakuru',
            'landmark': 'Near town centre',
            'actual_receipts': '25,000',
            'lead_source': 'HomeBiogas',
            'hb_sales_person': 'Sales One',
        }
        data.update(overrides)
        return SimpleNamespace(**data)

    def write_reconciled_shape_template(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Requisition Form'
        ws.merge_cells('C5:K5')
        ws['C5'] = 'JAWABU REQUISITION FORM'
        ws['L5'] = 'Date:'
        ws['L7'] = 'Order No:'
        ws.merge_cells('C9:M9')
        ws['C9'] = 'We hereby request a proforma invoice for the following order:'
        headers = {
            'C12': 'NO.',
            'D12': 'NAME OF THE CUSTOMER',
            'E12': 'CONTACT NO.',
            'F12': 'ID NO.',
            'G12': 'CREDIT ANALYSIS',
            'H12': 'CALLUP COMMENT',
            'I12': 'INSTALLATION LOCATION',
            'I13': 'COUNTY',
            'J13': 'LOCATION & NEAREST LANDMARK',
            'K12': 'DEPOSIT PAID TO',
            'K13': 'HBG',
            'L13': 'JBL',
            'M12': 'HB SALES PERSON',
        }
        for cell, value in headers.items():
            ws[cell] = value
        for row in range(14, 19):
            ws.cell(row=row, column=3, value=row - 13)
        ws['C22'] = 'Requisitioned by:'
        ws['F22'] = 'Signature:'
        ws['K22'] = 'Date:'
        wb.save(path)

    def generate_with_template(self, template_path: Path, farmers):
        fake_template = SimpleNamespace(file=SimpleNamespace(path=str(template_path)))
        with patch('core.models.RequisitionTemplate') as model:
            model.objects.filter.return_value.first.return_value = fake_template
            return generate_requisition_excel(farmers, 'REQ-TEST-001', date(2026, 7, 23))

    def test_reconciled_layout_preserves_id_and_writes_callup_comment_column(self):
        template_path = Path('tmp_requisition_reconciled_shape.xlsx')
        output_path = Path('tmp_requisition_reconciled_shape_output.xlsx')
        self.addCleanup(lambda: template_path.exists() and template_path.unlink())
        self.addCleanup(lambda: output_path.exists() and output_path.unlink())
        self.write_reconciled_shape_template(template_path)

        output_path.write_bytes(self.generate_with_template(template_path, [self.farmer()]))
        ws = load_workbook(output_path, data_only=False).active

        self.assertEqual(ws['L5'].value, 'Date:')
        self.assertEqual(ws['M5'].value, '23-Jul-2026')
        self.assertEqual(ws['L7'].value, 'Order No:')
        self.assertEqual(ws['M7'].value, 'REQ-TEST-001')
        self.assertTrue(ws['M5'].font.bold)
        self.assertTrue(ws['M7'].font.bold)
        self.assertEqual(ws['M5'].font.sz, ws['L5'].font.sz)
        self.assertEqual(ws['M7'].font.sz, ws['L7'].font.sz)
        self.assertEqual(ws['M5'].alignment.horizontal, 'center')
        self.assertEqual(ws['M7'].alignment.horizontal, 'center')
        self.assertEqual(ws['D14'].value, 'Mary Wanjiku')
        self.assertEqual(ws['E14'].value, '254712345678')
        self.assertEqual(ws['F14'].value, '12345678')
        self.assertEqual(ws['G14'].value, 'Approved')
        self.assertIsNone(ws['H14'].value)
        self.assertEqual(ws['I14'].value, 'Nakuru')
        self.assertEqual(ws['J14'].value, 'Near town centre')
        self.assertEqual(ws['K14'].value, 25000)
        self.assertEqual(ws['M14'].value, 'Sales One')
        self.assertIn(ws['D15'].value, (None, ''))
        self.assertIn(ws['E15'].value, (None, ''))
        self.assertNotIn('TOTAL', str(ws['D15'].value or '').upper())
        for cell_ref in ('C14', 'D14', 'E14', 'F14', 'G14', 'H14', 'I14', 'J14', 'K14', 'L14', 'M14'):
            self.assertEqual(ws[cell_ref].alignment.horizontal, 'center')

    def test_supplied_reconciled_template_is_supported_when_present(self):
        template_path = Path('requisition/JBL_Requisition_Form_Reconciled.xlsx')
        if not template_path.exists():
            self.skipTest('JBL_Requisition_Form_Reconciled.xlsx is not present in this checkout.')
        output_path = Path('tmp_requisition_reconciled_actual_output.xlsx')
        self.addCleanup(lambda: output_path.exists() and output_path.unlink())

        output_path.write_bytes(self.generate_with_template(template_path, [
            self.farmer(),
            self.farmer(
                customer_name='John Kamau',
                primary_phone='254700111222',
                national_id='87654321',
                county='Embu',
                landmark='Market road',
                actual_receipts='30000',
                lead_source='JBL referral',
                hb_sales_person='Sales Two',
            ),
        ]))
        ws = load_workbook(output_path, data_only=False).active

        self.assertEqual(ws['M5'].alignment.horizontal, 'center')
        self.assertEqual(ws['M5'].alignment.vertical, 'center')
        self.assertTrue(ws['M5'].font.bold)
        self.assertEqual(ws['M5'].font.sz, ws['L5'].font.sz)
        self.assertEqual(ws['M7'].alignment.horizontal, 'center')
        self.assertEqual(ws['M7'].alignment.vertical, 'center')
        self.assertTrue(ws['M7'].font.bold)
        self.assertEqual(ws['M7'].font.sz, ws['L7'].font.sz)
        self.assertEqual(ws['D14'].value, 'Mary Wanjiku')
        self.assertEqual(ws['F14'].value, '12345678')
        self.assertIsNone(ws['H14'].value)
        self.assertEqual(ws['K14'].value, 25000)
        self.assertEqual(ws['D15'].value, 'John Kamau')
        self.assertEqual(ws['F15'].value, '87654321')
        self.assertEqual(ws['L15'].value, 30000)
        self.assertIn(ws['D16'].value, (None, ''))
        self.assertIn(ws['E16'].value, (None, ''))
        self.assertNotIn('TOTAL', str(ws['D16'].value or '').upper())
        self.assertNotEqual(ws['K16'].value, '=SUM(K14:K15)')
        self.assertNotEqual(ws['L16'].value, '=SUM(L14:L15)')

    def test_generation_preserves_embedded_template_logos(self):
        template_path = Path('tmp_requisition_with_logo.xlsx')
        output_path = Path('tmp_requisition_with_logo_output.xlsx')
        logo_path = Path('tmp_requisition_logo.png')
        self.addCleanup(lambda: template_path.exists() and template_path.unlink())
        self.addCleanup(lambda: output_path.exists() and output_path.unlink())
        self.addCleanup(lambda: logo_path.exists() and logo_path.unlink())
        self.write_reconciled_shape_template(template_path)

        PilImage.new('RGB', (24, 24), color=(10, 80, 160)).save(logo_path)
        wb = load_workbook(template_path)
        ws = wb.active
        ws.add_image(XlsxImage(str(logo_path)), 'C3')
        wb.save(template_path)

        output_path.write_bytes(self.generate_with_template(template_path, [self.farmer()]))
        generated = load_workbook(output_path)

        self.assertEqual(len(generated.active._images), 1)
