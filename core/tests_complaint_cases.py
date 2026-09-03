import hashlib
import hmac
import json
import time
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode
from unittest.mock import patch

from django.contrib import admin
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.exceptions import ValidationError
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from core.models import (
    AccessGrant,
    CaseUpdate,
    ComplaintCaseEvidence,
    ComplaintCaseControl,
    ComplaintCaseEvent,
    ComplaintCaseImportBatch,
    ComplaintCaseImportItem,
    ComplaintCategory,
    ComplaintCaseSequence,
    GroupSheetConfiguration,
    ParsedMessage,
    ProcessedMessage,
    RawMessage,
    UserProfile,
    ComplianceAuditEvent,
)
from core.services.complaint_cases import (
    ComplaintCaseConflict,
    ComplaintCaseError,
    complete_review_details,
    create_complaint_case,
    bootstrap_data,
    evidence_filename,
    list_cases,
    list_cases_page,
    next_complaint_case_id,
    reopen_case,
    resolve_case,
    staff_actor_for_payload,
    update_case,
    ensure_case_control,
    staff_actor_for_user,
    suggest_category,
)
from core.services.complaint_register import register_overview
from core.services.complaint_imports import (
    ComplaintImportAuthorizationError,
    ComplaintImportConflict,
    associate_complaint_import_item,
    finalize_complaint_import_batch,
    mark_complaint_import_batch_failed,
    reserve_complaint_import_batch,
)
from core.services.group_config import GroupConfig, GroupRegistry
from core.services.telegram_auth import validate_telegram_init_data


class ComplaintCaseServiceTests(TestCase):
    def setUp(self):
        self.group = GroupSheetConfiguration.objects.create(
            group_id='-100100', sheet_id='test-sheet', sheet_name='Complaints', workflow={'type': 'case'}
        )
        self.config = GroupConfig(group_id=self.group.group_id, sheet_id='test-sheet', sheet_name='Complaints', workflow={'type': 'case'})
        self.category = ComplaintCategory.objects.create(
            key='product-issue', label='Product issue', default_priority='normal', default_sla_hours=72,
        )
        self.case = self.create_case('-100100', 'CASE-1')
        self.other_case = self.create_case('-100200', 'CASE-2')
        User = get_user_model()
        self.officer = User.objects.create_user(username='officer-one', first_name='Officer', last_name='One', is_active=True)
        self.officer.set_unusable_password()
        self.officer.save(update_fields=['password'])
        UserProfile.objects.create(user=self.officer, telegram_id='100', telegram_username='officer_one')
        AccessGrant.objects.create(
            user=self.officer, workflow='complaint_cases', role='OFFICER',
            group_configuration=self.group,
        )
        self.manager = User.objects.create_user(username='manager-one', first_name='Manager', last_name='One', is_active=True)
        self.manager.set_unusable_password()
        self.manager.save(update_fields=['password'])
        UserProfile.objects.create(user=self.manager, telegram_id='200', telegram_username='manager_one')
        AccessGrant.objects.create(
            user=self.manager, workflow='complaint_cases', role='MANAGER',
            group_configuration=self.group,
        )
        self.hb_staff = User.objects.create_user(username='hb-resolver', first_name='HB', last_name='Resolver', is_active=True)
        UserProfile.objects.create(user=self.hb_staff, telegram_id='300', telegram_username='hb_resolver')
        AccessGrant.objects.create(
            user=self.hb_staff, workflow='complaint_cases', role='HB_STAFF',
            group_configuration=self.group,
        )

    def create_case(self, group_id, message_id):
        raw = RawMessage.objects.create(telegram_message_id=message_id, content='raw complaint')
        processed = ProcessedMessage.objects.create(message_hash=f'hash-{message_id}', raw_message=raw)
        return ParsedMessage.objects.create(
            processed_message=processed, message_id=message_id, group_id=group_id,
            timestamp=timezone.now(), raw_message='raw complaint', customer_name='Client',
            customer_phone='0712345678', complaint_description='System is not working', complaint_status='Open',
        )

    def actor(self, telegram_id):
        return staff_actor_for_payload(self.config, {'user': json.dumps({'id': telegram_id})})

    def signed_init_data(self, telegram_id='100'):
        pairs = {'auth_date': str(int(time.time())), 'user': json.dumps({'id': int(telegram_id)})}
        check = '\n'.join(f'{key}={value}' for key, value in sorted(pairs.items()))
        secret = hmac.new(b'WebAppData', b'test-bot-token', hashlib.sha256).digest()
        pairs['hash'] = hmac.new(secret, check.encode('utf-8'), hashlib.sha256).hexdigest()
        return urlencode(pairs)

    def test_list_is_group_scoped(self):
        cases = list_cases(self.config)
        self.assertEqual([case['case_id'] for case in cases], ['CASE-1'])
        self.assertRegex(cases[0]['reference_number'], r'^CMP\d{6}$')
        self.assertTrue(cases[0]['recorded_at'])

        by_reference = list_cases(self.config, query=cases[0]['reference_number'])
        self.assertEqual([case['case_id'] for case in by_reference], ['CASE-1'])

        sequence = ComplaintCaseSequence.objects.get(group_id='__complaint_global__', year=0)
        next_number = sequence.next_number
        self.assertEqual(list_cases(self.config)[0]['reference_number'], cases[0]['reference_number'])
        sequence.refresh_from_db()
        self.assertEqual(sequence.next_number, next_number)

    def test_legacy_status_filter_remains_compatible_but_branch_filter_is_ignored(self):
        self.case.branch_region = 'Nakuru'
        self.case.save(update_fields=['branch_region'])
        embu_case = self.create_case('-100100', 'CASE-3')
        embu_case.branch_region = 'Embu'
        embu_case.complaint_status = 'In Progress'
        embu_case.save(update_fields=['branch_region', 'complaint_status'])
        nakuru_case = self.create_case('-100100', 'CASE-4')
        nakuru_case.branch_region = 'Nakuru'
        nakuru_case.complaint_status = 'In Progress'
        nakuru_case.save(update_fields=['branch_region', 'complaint_status'])

        cases = list_cases(self.config, status='In Progress', branch='Embu')

        self.assertEqual({case['case_id'] for case in cases}, {'CASE-3', 'CASE-4'})

    def test_numbered_pages_are_limited_to_ten_and_report_totals(self):
        for index in range(11):
            self.create_case('-100100', f'CASE-PAGE-{index:02d}')

        first = list_cases_page(self.config, self.actor('100'), page=1, page_size=50)
        second = list_cases_page(self.config, self.actor('100'), page=2, page_size=50)

        self.assertEqual(len(first['items']), 10)
        self.assertEqual(len(second['items']), 2)
        self.assertEqual(first['pagination'], {'page': 1, 'pages': 2, 'total': 12, 'page_size': 10})
        self.assertEqual(first['start_index'], 1)
        self.assertEqual(second['start_index'], 11)
        self.assertFalse({item['case_id'] for item in first['items']} & {item['case_id'] for item in second['items']})

    def test_two_state_counts_are_group_scoped(self):
        self.case.complaint_status = 'In Progress'
        self.case.save(update_fields=['complaint_status'])
        control = ensure_case_control(self.case)
        control.sla_due_at = timezone.now() - timedelta(hours=1)
        control.save(update_fields=['sla_due_at'])

        counts = bootstrap_data(self.config, self.actor('100'))['counts']

        self.assertEqual(counts['pending'], 1)
        self.assertEqual(counts['resolved'], 0)
        self.assertEqual(counts['total'], 1)

    def test_resolved_case_age_uses_resolution_time_and_friendly_label(self):
        reported_at = timezone.now() - timedelta(days=8)
        resolved_at = reported_at + timedelta(days=3, hours=2)
        self.case.timestamp = reported_at
        self.case.complaint_status = 'Closed'
        self.case.date_resolved = resolved_at
        self.case.save(update_fields=['timestamp', 'complaint_status', 'date_resolved'])

        item = list_cases(self.config, status='Closed')[0]

        self.assertEqual(item['days_open'], 3)
        self.assertEqual(item['age_label'], 'Resolved after 3 days')

    @override_settings(
        COMPLAINT_CASE_MAX_FILES_PER_UPDATE=4,
        COMPLAINT_CASE_MAX_FILE_SIZE_MB=7,
        COMPLAINT_CASE_MAX_TOTAL_UPLOAD_MB=18,
    )
    def test_bootstrap_exposes_evidence_limits_and_needs_details_count(self):
        self.case.complaint_status = 'Review Needed'
        self.case.save(update_fields=['complaint_status'])

        data = bootstrap_data(self.config, self.actor('100'))

        self.assertEqual(data['counts']['needs_details'], 1)
        self.assertEqual(data['evidence_limits'], {
            'max_files': 4, 'max_file_size_mb': 7, 'max_total_upload_mb': 18,
        })

    def test_branch_scoped_grant_still_sees_the_shared_group_queue(self):
        AccessGrant.objects.filter(user=self.officer, workflow='complaint_cases').update(branch='Nakuru')
        self.case.branch_region = 'Nakuru'
        self.case.save(update_fields=['branch_region'])
        embu_case = self.create_case('-100100', 'CASE-3')
        embu_case.branch_region = 'Embu'
        embu_case.save(update_fields=['branch_region'])

        cases = list_cases(self.config, self.actor('100'))

        self.assertEqual([case['case_id'] for case in cases], ['CASE-1', 'CASE-3'])

    def test_branch_scoped_manager_can_reopen_any_resolved_case_in_the_group(self):
        AccessGrant.objects.filter(user=self.manager, workflow='complaint_cases').update(branch='Nakuru')
        self.case.branch_region = 'Embu'
        self.case.complaint_status = 'Closed'
        self.case.save(update_fields=['branch_region', 'complaint_status'])

        with patch('core.services.complaint_cases.get_sheets_service') as get_service:
            get_service.return_value.update_case_row.return_value = True
            reopened = reopen_case(
                self.config, self.actor('200'), 'CASE-1',
                {'client_request_id': 'cross-branch-reopen-1', 'expected_revision': 1, 'reason': 'Resolution was insufficient.'},
            )

        self.assertEqual(reopened['status'], 'Pending')

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_list_fragment_renders_authorized_cases(self):
        self.case.branch_region = 'Nakuru'
        self.case.save(update_fields=['branch_region'])
        embu_case = self.create_case('-100100', 'CASE-3')
        embu_case.branch_region = 'Embu'
        embu_case.save(update_fields=['branch_region'])

        response = self.client.post(
            reverse('complaint_cases_list_fragment'),
            {'group_id': self.group.group_id, 'branch': 'Nakuru', 'status': 'active'},
            HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data('100'),
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'complaint_cases/partials/case_list.html')
        self.assertContains(response, 'CASE-1')
        self.assertContains(response, 'CASE-3')

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_list_endpoint_returns_numbered_pagination_contract(self):
        response = self.client.post(
            reverse('complaint_cases_list'),
            data=json.dumps({'group_id': self.group.group_id, 'status': 'all', 'page': 1}),
            content_type='application/json',
            HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data('100'),
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['pagination']['page_size'], 10)
        self.assertEqual(payload['pagination']['total'], 1)
        self.assertEqual(payload['start_index'], 1)

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_hb_resolves_and_manager_reopens_through_dedicated_endpoints(self):
        with patch('core.services.complaint_cases.get_sheets_service') as get_service:
            get_service.return_value.update_case_row.return_value = True
            resolved = self.client.post(
                reverse('complaint_cases_resolve', args=['CASE-1']),
                data=json.dumps({
                    'group_id': self.group.group_id, 'client_request_id': 'api-resolve-1',
                    'expected_revision': 1, 'resolution_text': 'Customer remedy completed.',
                }), content_type='application/json', HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data('300'),
            )
            reopened = self.client.post(
                reverse('complaint_cases_reopen', args=['CASE-1']),
                data=json.dumps({
                    'group_id': self.group.group_id, 'client_request_id': 'api-reopen-1',
                    'expected_revision': 2, 'reason': 'The remedy was incomplete.',
                }), content_type='application/json', HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data('200'),
            )

        self.assertEqual(resolved.status_code, 200)
        self.assertEqual(resolved.json()['case']['status'], 'Resolved')
        self.assertEqual(reopened.status_code, 200)
        self.assertEqual(reopened.json()['case']['status'], 'Pending')

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_category_suggestion_endpoint_is_authenticated_and_read_only(self):
        ComplaintCategory.objects.update_or_create(
            key='system-performance', defaults={'label': 'System Performance', 'active': True},
        )
        ComplaintCategory.objects.update_or_create(
            key='other-complaint', defaults={'label': 'Other Complaint', 'active': True},
        )

        response = self.client.post(
            reverse('complaint_cases_category_suggestion'),
            data=json.dumps({'group_id': self.group.group_id, 'description': 'There is no gas production.'}),
            content_type='application/json',
            HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data('100'),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['data']['suggestion']['key'], 'system-performance')
        self.case.refresh_from_db()
        self.assertEqual(self.case.complaint_category, '')

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_stale_resolution_response_includes_the_winning_resolution(self):
        with patch('core.services.complaint_cases.get_sheets_service') as get_service:
            get_service.return_value.update_case_row.return_value = True
            self.client.post(
                reverse('complaint_cases_resolve', args=['CASE-1']),
                data=json.dumps({
                    'group_id': self.group.group_id, 'client_request_id': 'api-winner-1',
                    'expected_revision': 1, 'resolution_text': 'Winning resolution.',
                }), content_type='application/json', HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data('300'),
            )
            conflict = self.client.post(
                reverse('complaint_cases_resolve', args=['CASE-1']),
                data=json.dumps({
                    'group_id': self.group.group_id, 'client_request_id': 'api-loser-1',
                    'expected_revision': 1, 'resolution_text': 'Losing draft.',
                }), content_type='application/json', HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data('300'),
            )

        self.assertEqual(conflict.status_code, 409)
        payload = conflict.json()
        self.assertEqual(payload['code'], 'revision_conflict')
        self.assertEqual(payload['current_case']['latest_resolution']['note'], 'Winning resolution.')

    def test_registry_resolves_a_group_added_after_the_initial_cache(self):
        GroupRegistry._instance = None
        GroupRegistry.get_instance()
        added = GroupSheetConfiguration.objects.create(
            group_id='-100101', sheet_id='later-sheet', sheet_name='Later', workflow={'type': 'case'},
        )

        resolved = GroupRegistry.get_instance().get_group(added.group_id)

        self.assertIsNotNone(resolved)
        self.assertEqual(resolved.group_id, added.group_id)
        self.assertEqual(resolved.sheet_id, 'later-sheet')

    def test_officer_cannot_mutate_a_pending_case(self):
        with self.assertRaisesMessage(ComplaintCaseError, 'does not permit this complaint transition'):
            update_case(self.config, self.actor('100'), 'CASE-1', {
                'client_request_id': 'request-1', 'status': 'Open',
                'resolution_text': 'Called the client.', 'expected_revision': 1,
            }, [])

    def test_officer_cannot_close_case(self):
        with self.assertRaisesMessage(ComplaintCaseError, 'does not permit this complaint transition'):
            resolve_case(
                self.config, self.actor('100'), 'CASE-1',
                {'client_request_id': 'request-2', 'expected_revision': 1, 'resolution_text': 'Done'}, [],
            )

    def test_manager_cannot_resolve_and_officer_cannot_reopen(self):
        with self.assertRaisesMessage(ComplaintCaseError, 'does not permit this complaint transition'):
            resolve_case(
                self.config, self.actor('200'), 'CASE-1',
                {'client_request_id': 'manager-resolve-denied', 'expected_revision': 1, 'resolution_text': 'Done'}, [],
            )
        self.case.complaint_status = 'Closed'
        self.case.save(update_fields=['complaint_status'])
        with self.assertRaisesMessage(ComplaintCaseError, 'does not permit this complaint transition'):
            reopen_case(
                self.config, self.actor('100'), 'CASE-1',
                {'client_request_id': 'officer-reopen-denied', 'expected_revision': 1, 'reason': 'Insufficient'},
            )

    def test_hb_resolver_cannot_create_complaints(self):
        with self.assertRaisesMessage(ComplaintCaseError, 'does not permit creating complaints'):
            create_complaint_case(
                self.config, self.actor('300'), {
                    'client_request_id': 'hb-create-denied', 'client_name': 'Denied',
                    'customer_id': '123456', 'branch_region': 'Nakuru',
                    'complaint_category': 'Product issue', 'complaint_description': 'Not allowed.',
                }, [],
            )

    @patch('core.services.complaint_cases.append_parsed_message_to_sheet', return_value=True)
    def test_officer_completes_review_details_idempotently(self, append_to_sheet):
        self.case.complaint_status = 'Review Needed'
        self.case.customer_id = ''
        self.case.complaint_category = self.category.label
        self.case.save(update_fields=['complaint_status', 'customer_id', 'complaint_category'])
        ensure_case_control(self.case, self.config)
        fields = {
            'client_request_id': 'complete-details-1',
            'expected_revision': 1,
            'customer_id': '100100',
            'complaint_category': self.category.label,
        }

        first = complete_review_details(self.config, self.actor('100'), 'CASE-1', fields)
        replay = complete_review_details(self.config, self.actor('100'), 'CASE-1', fields)

        self.case.refresh_from_db()
        self.assertEqual(self.case.complaint_status, 'Open')
        self.assertEqual(self.case.customer_id, '100100')
        self.assertFalse(first['needs_details'])
        self.assertFalse(replay['needs_details'])
        self.assertEqual(self.case.case_updates.filter(source='mini_app_review_completion').count(), 1)
        self.assertEqual(self.case.complaint_control.events.filter(action='details_completed').count(), 1)
        append_to_sheet.assert_called_once()

    def test_hb_staff_cannot_complete_review_details(self):
        self.case.complaint_status = 'Review Needed'
        self.case.customer_id = ''
        self.case.complaint_category = self.category.label
        self.case.save(update_fields=['complaint_status', 'customer_id', 'complaint_category'])

        with self.assertRaisesMessage(ComplaintCaseError, 'does not permit completing complaint details'):
            complete_review_details(self.config, self.actor('300'), 'CASE-1', {
                'client_request_id': 'complete-details-2', 'expected_revision': 1,
                'customer_id': '100101', 'complaint_category': self.category.label,
            })

    def test_category_suggestion_is_deterministic_and_never_assigns(self):
        categories = (
            ('leakage', 'Leakage'), ('pipe-connection-fault', 'Pipe/Connection Fault'),
            ('burner-knob-fault', 'Burner/Knob Fault'), ('installation-delay', 'Installation Delay'),
            ('other-complaint', 'Other Complaint'),
        )
        for key, label in categories:
            ComplaintCategory.objects.update_or_create(
                key=key, defaults={'label': label, 'active': True},
            )

        leakage = suggest_category(self.config, 'The broken pipe is leaking gas at the connection.')
        ambiguous = suggest_category(self.config, 'Installation is delayed and the burner will not ignite.')
        fallback = suggest_category(self.config, 'Customer has an unusual concern.')

        self.assertEqual(leakage['suggestion']['key'], 'leakage')
        self.assertEqual(ambiguous['state'], 'ambiguous')
        self.assertEqual({item['key'] for item in ambiguous['candidates']}, {'installation-delay', 'burner-knob-fault'})
        self.assertEqual(fallback['suggestion']['key'], 'other-complaint')
        self.assertEqual(self.case.complaint_category, '')

    def test_failed_drive_upload_is_recorded_without_losing_case_update(self):
        evidence = SimpleUploadedFile('photo.jpg', b'\xff\xd8\xff\xe0synthetic-image', content_type='image/jpeg')
        with patch('core.services.complaint_cases.get_sheets_service') as get_service, patch(
            'core.services.complaint_cases.GoogleDriveMediaStorage.upload', side_effect=RuntimeError('offline')
        ):
            get_service.return_value.update_case_row.return_value = True
            resolve_case(
                self.config, self.actor('300'), 'CASE-1',
                {'client_request_id': 'request-3', 'resolution_text': 'Resolved with photo.', 'expected_revision': 1}, [evidence],
            )
        self.assertEqual(ComplaintCaseEvidence.objects.get().upload_status, 'failed')
        self.assertEqual(self.case.case_updates.count(), 1)

    def test_sheet_failure_does_not_roll_back_local_case_update(self):
        with patch('core.services.complaint_cases.get_sheets_service') as get_service:
            get_service.return_value.update_case_row.return_value = False
            result = resolve_case(
                self.config, self.actor('300'), 'CASE-1',
                {'client_request_id': 'local-first-1', 'expected_revision': 1, 'resolution_text': 'Resolved locally.'}, [],
            )
        self.case.refresh_from_db()
        self.assertEqual(self.case.complaint_status, 'Closed')
        self.assertEqual(result['sync_status'], 'failed')
        self.assertEqual(self.case.case_updates.get().sync_status, 'failed')

    def test_spoofed_image_extension_is_rejected(self):
        evidence = SimpleUploadedFile('photo.jpg', b'not-an-image', content_type='image/jpeg')
        with self.assertRaisesMessage(ComplaintCaseError, 'genuine JPEG'):
            resolve_case(
                self.config, self.actor('300'), 'CASE-1',
                {'client_request_id': 'spoofed-file-1', 'expected_revision': 1, 'resolution_text': 'Evidence.'}, [evidence],
            )

    def test_new_docx_evidence_is_rejected(self):
        evidence = SimpleUploadedFile(
            'legacy.docx', b'PK\x03\x04legacy-docx',
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        )
        with self.assertRaisesMessage(ComplaintCaseError, 'JPEG, PNG, WebP, or PDF'):
            resolve_case(
                self.config, self.actor('300'), 'CASE-1',
                {'client_request_id': 'docx-denied-1', 'expected_revision': 1, 'resolution_text': 'Evidence.'}, [evidence],
            )

    def test_evidence_filename_does_not_expose_customer_id(self):
        self.case.customer_id = 'ID 123/456'
        self.case.save(update_fields=['customer_id'])

        filename = evidence_filename(self.case, 'site photo.jpg', 1)
        self.assertNotIn('ID_123456', filename)
        self.assertTrue(filename.endswith('-01-site_photo.jpg'))

    def test_stale_revision_is_rejected_without_overwriting_the_first_update(self):
        actor = self.actor('300')
        with patch('core.services.complaint_cases.get_sheets_service') as get_service:
            get_service.return_value.update_case_row.return_value = True
            resolve_case(self.config, actor, 'CASE-1', {
                'client_request_id': 'revision-first', 'expected_revision': 1,
                'resolution_text': 'First resolution.',
            }, [])
            with self.assertRaises(ComplaintCaseConflict):
                resolve_case(self.config, actor, 'CASE-1', {
                    'client_request_id': 'revision-stale', 'expected_revision': 1,
                    'resolution_text': 'Stale resolution.',
                }, [])
        self.case.refresh_from_db()
        self.assertEqual(self.case.complaint_status, 'Closed')
        self.assertEqual(ComplaintCaseEvent.objects.filter(case__parsed_message=self.case).count(), 1)

    def test_manager_can_reopen_a_resolved_case(self):
        with patch('core.services.complaint_cases.get_sheets_service') as get_service:
            get_service.return_value.update_case_row.return_value = True
            resolve_case(self.config, self.actor('300'), 'CASE-1', {
                'client_request_id': 'resolve-before-reopen', 'expected_revision': 1,
                'resolution_text': 'Initial remedy.',
            }, [])
            result = reopen_case(self.config, self.actor('200'), 'CASE-1', {
                'client_request_id': 'manager-reopen', 'expected_revision': 2,
                'reason': 'The remedy did not address the reported fault.',
            })
        self.case.refresh_from_db()
        self.assertEqual(self.case.complaint_status, 'Open')
        self.assertEqual(result['latest_resolution']['note'], 'Initial remedy.')
        self.assertEqual(result['latest_reopen']['note'], 'The remedy did not address the reported fault.')

    def test_resolution_retry_returns_the_first_result_without_a_second_event(self):
        fields = {
            'client_request_id': 'resolve-retry-1', 'expected_revision': 1,
            'resolution_text': 'Resolved once.',
        }
        with patch('core.services.complaint_cases.get_sheets_service') as get_service:
            get_service.return_value.update_case_row.return_value = True
            first = resolve_case(self.config, self.actor('300'), 'CASE-1', fields, [])
            replay = resolve_case(self.config, self.actor('300'), 'CASE-1', fields, [])

        self.assertEqual(first['revision'], replay['revision'])
        self.assertEqual(self.case.case_updates.filter(client_request_id='resolve-retry-1').count(), 1)
        self.assertEqual(ComplaintCaseEvent.objects.filter(case__parsed_message=self.case).count(), 1)

    def test_imported_case_exposes_audited_batch_attribution(self):
        batch = ComplaintCaseImportBatch.objects.create(
            group_id=self.config.group_id, source_telegram_message_id='telegram-batch-1',
            initiated_by=self.manager, actor_label='Manager One', telegram_user_id_snapshot='200',
            source_hash='a' * 64, status=ComplaintCaseImportBatch.STATUS_COMPLETED,
            source_count=1, created_count=1,
        )
        ComplaintCaseImportItem.objects.create(batch=batch, parsed_message=self.case, source_index=0)

        item = list_cases(self.config, self.actor('100'))[0]

        self.assertEqual(item['source_attribution']['type'], 'batch')
        self.assertEqual(item['source_attribution']['actor'], 'Manager One')

    def test_import_batch_reservation_is_authorized_idempotent_and_hash_bound(self):
        with self.assertRaises(ComplaintImportAuthorizationError):
            reserve_complaint_import_batch(
                actor=self.manager,
                group_id=self.config.group_id,
                source_telegram_message_id='telegram-batch-service-1',
                telegram_user_id='200',
                source_hash='a' * 64,
                source_count=2,
            )
        admin_user = get_user_model().objects.create_superuser(
            username='complaint-import-superuser', password='unused-test-password',
        )
        first = reserve_complaint_import_batch(
            actor=admin_user,
            group_id=self.config.group_id,
            source_telegram_message_id='telegram-batch-service-1',
            telegram_user_id='900',
            source_hash='a' * 64,
            source_count=2,
        )
        replay = reserve_complaint_import_batch(
            actor=admin_user,
            group_id=self.config.group_id,
            source_telegram_message_id='telegram-batch-service-1',
            telegram_user_id='900',
            source_hash='a' * 64,
            source_count=2,
        )

        self.assertTrue(first.created)
        self.assertFalse(replay.created)
        self.assertTrue(replay.already_processing)
        self.assertEqual(first.batch.pk, replay.batch.pk)
        self.assertEqual(ComplaintCaseImportBatch.objects.filter(
            group_id=self.config.group_id,
            source_telegram_message_id='telegram-batch-service-1',
        ).count(), 1)
        with self.assertRaises(ComplaintImportConflict):
            reserve_complaint_import_batch(
                actor=admin_user,
                group_id=self.config.group_id,
                source_telegram_message_id='telegram-batch-service-1',
                telegram_user_id='900',
                source_hash='b' * 64,
                source_count=2,
            )
        mark_complaint_import_batch_failed(batch=first.batch)
        failed_retry = reserve_complaint_import_batch(
            actor=admin_user,
            group_id=self.config.group_id,
            source_telegram_message_id='telegram-batch-service-1',
            telegram_user_id='900',
            source_hash='a' * 64,
            source_count=2,
        )
        self.assertTrue(failed_retry.retrying)
        self.assertEqual(
            failed_retry.batch.status,
            ComplaintCaseImportBatch.STATUS_QUEUED,
        )

    def test_import_item_association_and_finalization_are_idempotent(self):
        admin_user = get_user_model().objects.create_superuser(
            username='complaint-attribution-superuser', password='unused-test-password',
        )
        reservation = reserve_complaint_import_batch(
            actor=admin_user,
            group_id=self.config.group_id,
            source_telegram_message_id='telegram-batch-service-2',
            telegram_user_id='901',
            source_hash='c' * 64,
            source_count=1,
        )
        first_item, first_created = associate_complaint_import_item(
            batch=reservation.batch,
            parsed_message=self.case,
            source_index=0,
        )
        replay_item, replay_created = associate_complaint_import_item(
            batch=reservation.batch,
            parsed_message=self.case,
            source_index=0,
        )

        self.assertTrue(first_created)
        self.assertFalse(replay_created)
        self.assertEqual(first_item.pk, replay_item.pk)
        with self.assertRaises(ComplaintImportConflict):
            associate_complaint_import_item(
                batch=reservation.batch,
                parsed_message=self.other_case,
                source_index=0,
            )

        completed = finalize_complaint_import_batch(
            batch=reservation.batch,
            created_count=1,
            matched_count=0,
            rejected_count=0,
            error_count=0,
        )
        replayed_completion = finalize_complaint_import_batch(
            batch=reservation.batch,
            created_count=99,
            matched_count=99,
            rejected_count=1,
            error_count=1,
        )
        self.assertEqual(completed.status, ComplaintCaseImportBatch.STATUS_COMPLETE)
        self.assertEqual(replayed_completion.created_count, 1)
        self.assertEqual(replayed_completion.error_count, 0)

    def test_legacy_import_never_guesses_an_uploader(self):
        self.case.source = 'whatsapp_export'
        self.case.save(update_fields=['source'])

        item = list_cases(self.config, self.actor('100'))[0]

        self.assertEqual(item['source_attribution']['type'], 'legacy_batch')
        self.assertEqual(item['source_attribution']['label'], 'Imported from another system')

    @patch('core.services.complaint_cases.append_parsed_message_to_sheet', return_value=True)
    def test_officer_can_create_an_auditable_case_once_with_a_retry_identifier(self, append_to_sheet):
        def mark_case_synced(case, **_kwargs):
            case.synced_to_sheets = True
            case.last_sync_error = ''
            case.save(update_fields=['synced_to_sheets', 'last_sync_error'])
            return True

        append_to_sheet.side_effect = mark_case_synced
        fields = {
            'client_request_id': 'create-complaint-001',
            'client_name': "new o'NEIL client",
            'customer_phone': '0712345678',
            'customer_id': '',
            'branch_region': 'Nakuru',
            'complaint_category': 'Product issue',
            'complaint_description': 'The unit requires a field visit.',
            'latitude': '-1.286389',
            'longitude': '36.817223',
        }

        first = create_complaint_case(self.config, self.actor('100'), fields, [])
        second = create_complaint_case(self.config, self.actor('100'), fields, [])

        case = ParsedMessage.objects.get(message_id=first['case']['case_id'])
        self.assertEqual(first['case']['case_id'], second['case']['case_id'])
        self.assertEqual(case.customer_name, "New O'Neil Client")
        self.assertRegex(first['case']['case_id'], r'^CMP-\d{4}-001$')
        sequence = ComplaintCaseSequence.objects.get(group_id=self.config.group_id)
        self.assertEqual(sequence.next_number, 2)
        self.assertEqual(case.customer_phone, '254712345678')
        self.assertEqual(case.complaint_status, 'Open')
        self.assertEqual(case.source, 'complaint_mini_app')
        self.assertTrue(case.raw_message)
        self.assertEqual(CaseUpdate.objects.filter(parsed_message=case).count(), 1)
        append_to_sheet.assert_called_once()

    def test_case_reference_sequence_resets_for_each_calendar_year(self):
        first_2026 = next_complaint_case_id(
            self.config,
            reference_at=timezone.make_aware(datetime(2026, 12, 31, 23, 0)),
        )
        first_2027 = next_complaint_case_id(
            self.config,
            reference_at=timezone.make_aware(datetime(2027, 1, 1, 9, 0)),
        )

        self.assertEqual(first_2026, 'CMP-2026-001')
        self.assertEqual(first_2027, 'CMP-2027-001')

    def test_new_case_requires_a_phone_or_customer_id(self):
        with self.assertRaisesMessage(ComplaintCaseError, 'phone number or customer ID'):
            create_complaint_case(
                self.config,
                self.actor('100'),
                {
                    'client_request_id': 'create-complaint-002',
                    'client_name': 'New Client',
                    'branch_region': 'Nakuru',
                    'complaint_category': 'Product issue',
                    'complaint_description': 'The unit requires a field visit.',
                },
                [],
            )

    def test_customer_id_is_digits_only_and_preserves_leading_zeroes(self):
        with self.assertRaisesMessage(ComplaintCaseError, 'numbers only'):
            create_complaint_case(
                self.config, self.actor('100'), {
                    'client_request_id': 'create-invalid-id', 'client_name': 'Invalid ID',
                    'customer_phone': '0712345678', 'customer_id': 'ID-300',
                    'branch_region': 'Nakuru', 'complaint_category': 'Product issue',
                    'complaint_description': 'Invalid identifier should not be saved.',
                }, [],
            )

        with patch('core.services.complaint_cases.append_parsed_message_to_sheet', return_value=False):
            result = create_complaint_case(
                self.config, self.actor('100'), {
                    'client_request_id': 'create-leading-zero-id', 'client_name': 'Leading Zero',
                    'customer_id': '00123456', 'branch_region': 'Nakuru',
                    'complaint_category': 'Product issue',
                    'complaint_description': 'Identifier should preserve leading zeroes.',
                }, [],
            )
        self.assertEqual(ParsedMessage.objects.get(pk=result['case']['id']).customer_id, '00123456')

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_evidence_is_streamed_in_app_without_exposing_drive_url(self):
        update = CaseUpdate.objects.create(
            parsed_message=self.case, group_id=self.group.group_id,
            updated_by='Officer One', raw_update_text='Evidence uploaded.',
        )
        evidence = ComplaintCaseEvidence.objects.create(
            parsed_message=self.case, case_update=update, group_id=self.group.group_id,
            original_filename='site-photo.jpg', mime_type='image/jpeg', size=8,
            content_hash='e' * 64, drive_file_id='drive-file-id',
            drive_url='https://drive.example/private', upload_status='success',
        )

        with patch('core.api.complaint_case_views.GoogleDriveMediaStorage.download', return_value=b'jpeg-data'):
            response = self.client.post(
                reverse('complaint_cases_evidence_access', args=[evidence.pk]),
                data=json.dumps({'group_id': self.group.group_id, 'client_request_id': 'evidence-open-1'}),
                content_type='application/json',
                HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data('100'),
                HTTP_X_REQUEST_ID='evidence-open-1',
                HTTP_IDEMPOTENCY_KEY='evidence-open-1',
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content, b'jpeg-data')
        self.assertEqual(response['Content-Type'], 'image/jpeg')
        self.assertIn('no-store', response['Cache-Control'])
        self.assertNotIn('drive.example', response.content.decode('latin1'))
        self.assertTrue(ComplaintCaseEvent.objects.filter(
            case__parsed_message=self.case, action='evidence_opened',
        ).exists())

    @patch('core.services.complaint_cases.append_parsed_message_to_sheet', return_value=False)
    def test_new_case_keeps_the_audit_record_when_sheet_sync_is_deferred(self, append_to_sheet):
        result = create_complaint_case(
            self.config,
            self.actor('100'),
            {
                'client_request_id': 'create-complaint-003',
                'client_name': 'Deferred Sync Client',
                'customer_id': '00300',
                'branch_region': 'Nakuru',
                'complaint_category': 'Product issue',
                'complaint_description': 'Create locally and retry the Sheet sync later.',
            },
            [],
        )

        case = ParsedMessage.objects.get(message_id=result['case']['case_id'])
        self.assertTrue(result['created'])
        self.assertFalse(result['synced_to_sheet'])
        self.assertEqual(case.case_updates.count(), 1)
        append_to_sheet.assert_called_once()


class ComplaintCaseGlobalRegisterTests(TestCase):
    def setUp(self):
        self.group_a = GroupSheetConfiguration.objects.create(
            group_id='-100global-a', display_name='Nakuru complaints', sheet_id='sheet-a',
            sheet_name='Complaints', workflow={'type': 'case'},
        )
        self.group_b = GroupSheetConfiguration.objects.create(
            group_id='-100global-b', display_name='Embu complaints', sheet_id='sheet-b',
            sheet_name='Complaints', workflow={'type': 'case'},
        )
        self.category = ComplaintCategory.objects.create(
            key='global-product', label='Product issue', default_priority='high', default_sla_hours=24,
        )
        User = get_user_model()
        self.officer = User.objects.create_user(username='global-officer', is_active=True)
        UserProfile.objects.create(user=self.officer, telegram_id='801', telegram_username='global_officer')
        AccessGrant.objects.create(
            user=self.officer, workflow='complaint_cases', role='OFFICER',
            group_configuration=self.group_a,
        )
        AccessGrant.objects.create(
            user=self.officer, workflow='complaint_cases', role='IT',
            group_configuration=self.group_a,
        )
        self.case_a = self.create_case(self.group_a, 'CMP-2026-001', 'Alice Client')
        self.case_b = self.create_case(self.group_b, 'CMP-2026-001', '=HYPERLINK("bad")')

    def create_case(self, group, case_id, customer_name):
        raw = RawMessage.objects.create(telegram_message_id=case_id, content='private raw source')
        processed = ProcessedMessage.objects.create(
            message_hash=f'{group.group_id}-{case_id}', raw_message=raw,
        )
        case = ParsedMessage.objects.create(
            processed_message=processed, message_id=case_id, group_id=group.group_id,
            timestamp=timezone.now(), sender='Field officer', raw_message='private raw source',
            customer_name=customer_name, customer_phone='+254712345678', customer_id='ID-123',
            branch_region='Nakuru' if group == self.group_a else 'Embu',
            complaint_category=self.category.label, complaint_description='Unit is not producing gas.',
            complaint_status='Open',
        )
        ComplaintCaseControl.objects.create(
            parsed_message=case,
            reference_number=f'CMP{900001 + ComplaintCaseControl.objects.count():06d}',
            category=self.category, priority='high',
            sla_target_hours=24, sla_started_at=case.timestamp,
            sla_due_at=case.timestamp - timedelta(hours=1), sync_status='pending',
        )
        return case

    def signed_init_data(self, telegram_id=801):
        pairs = {'auth_date': str(int(time.time())), 'user': json.dumps({'id': telegram_id})}
        check = '\n'.join(f'{key}={value}' for key, value in sorted(pairs.items()))
        secret = hmac.new(b'WebAppData', b'test-bot-token', hashlib.sha256).digest()
        pairs['hash'] = hmac.new(secret, check.encode('utf-8'), hashlib.sha256).hexdigest()
        return urlencode(pairs)

    def post(self, name, payload=None, args=None):
        values = {'group_id': self.group_a.group_id, **(payload or {})}
        request_id = values.setdefault('client_request_id', f'global-test-{name}')
        return self.client.post(
            reverse(name, args=args or []), data=json.dumps(values), content_type='application/json',
            HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data(), HTTP_X_REQUEST_ID=request_id,
            HTTP_IDEMPOTENCY_KEY=request_id,
        )

    def get(self, name, params=None, *, telegram_id=801):
        values = {'group_id': self.group_a.group_id, **(params or {})}
        return self.client.get(
            reverse(name), data=values,
            HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data(telegram_id),
        )

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_management_report_user_sees_every_group(self):
        overview = self.post('complaint_cases_global_overview')
        listing = self.post('complaint_cases_global_list', {'filters': {}, 'page': 1})

        self.assertEqual(overview.status_code, 200)
        self.assertEqual(overview.json()['data']['metrics']['total'], 2)
        self.assertEqual(listing.status_code, 200)
        self.assertEqual({row['group_label'] for row in listing.json()['items']}, {'Nakuru complaints', 'Embu complaints'})
        self.assertEqual({row['case_id'] for row in listing.json()['items']}, {'CMP-2026-001'})
        self.assertEqual(
            {row['reference_number'] for row in listing.json()['items']},
            {'CMP900001', 'CMP900002'},
        )

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_ordinary_officer_and_manager_are_denied_management_reports(self):
        ordinary = get_user_model().objects.create_user(username='ordinary-officer', is_active=True)
        UserProfile.objects.create(user=ordinary, telegram_id='802')
        AccessGrant.objects.create(
            user=ordinary, workflow='complaint_cases', role='OFFICER',
            group_configuration=self.group_a,
        )
        manager = get_user_model().objects.create_user(username='ordinary-manager', is_active=True)
        UserProfile.objects.create(user=manager, telegram_id='803')
        AccessGrant.objects.create(
            user=manager, workflow='complaint_cases', role='MANAGER',
            group_configuration=self.group_a,
        )

        bootstrap = self.client.post(
            reverse('complaint_cases_bootstrap'),
            data=json.dumps({'group_id': self.group_a.group_id}), content_type='application/json',
            HTTP_X_TELEGRAM_INIT_DATA=self.signed_init_data(802),
            HTTP_X_REQUEST_ID='ordinary-bootstrap-1',
            HTTP_IDEMPOTENCY_KEY='ordinary-bootstrap-1',
        )
        self.assertEqual(bootstrap.status_code, 200)
        self.assertNotIn('complaint.reports.view', bootstrap.json()['data']['actor']['capabilities'])
        for telegram_id in (802, 803):
            self.assertEqual(self.get('complaint_reports_data', telegram_id=telegram_id).status_code, 403)
            self.assertEqual(self.get('complaint_reports_summary', telegram_id=telegram_id).status_code, 403)

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_report_data_is_allowlisted_capped_and_counts_only_successful_attachments(self):
        self.case_a.complaint_status = 'Review Needed'
        self.case_a.save(update_fields=['complaint_status'])
        update = CaseUpdate.objects.create(
            parsed_message=self.case_a, group_id=self.group_a.group_id,
            raw_update_text='attachment test', new_status='Open',
        )
        for status in ('success', 'failed', 'pending'):
            ComplaintCaseEvidence.objects.create(
                parsed_message=self.case_a, case_update=update, group_id=self.group_a.group_id,
                original_filename=f'private-{status}.jpg', mime_type='image/jpeg',
                drive_file_id=f'private-{status}', drive_url=f'https://drive.invalid/{status}',
                upload_status=status,
            )

        response = self.get('complaint_reports_data', {
            'page_size': 1000, 'search': 'CMP900001', 'sort': '-days_open',
            'branch': 'Nakuru', 'category': 'Product issue', 'status': 'pending',
            'date_from': timezone.localdate().isoformat(),
            'date_to': timezone.localdate().isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['page_size'], 100)
        self.assertEqual(payload['count'], 1)
        row = payload['results'][0]
        self.assertEqual(row['attachments'], 1)
        self.assertEqual(row['status'], 'Pending')
        self.assertTrue(row['needs_details'])
        self.assertEqual(set(row), {
            'complaint_id', 'date_reported', 'status', 'needs_details',
            'customer_name', 'customer_id', 'phone_number', 'reported_by',
            'branch_region', 'complaint_category', 'complaint_description',
            'source', 'gps_link', 'attachments', 'resolution_details',
            'date_resolved', 'days_open',
        })
        forbidden = {
            'raw_message', 'message_id', 'loan_status', 'loan_at_risk', 'risk_level',
            'drive_url', 'drive_file_id', 'original_filename', 'mime_type', 'content_hash',
        }
        self.assertFalse(forbidden.intersection(row))
        self.assertNotIn('Open', row.values())
        self.assertNotIn('private-success.jpg', json.dumps(row))

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_report_rejects_unknown_sort_and_summary_matches_database(self):
        rejected = self.get('complaint_reports_data', {'sort': 'customer_name'})
        summary = self.get('complaint_reports_summary')

        self.assertEqual(rejected.status_code, 400)
        self.assertEqual(summary.status_code, 200)
        payload = summary.json()
        self.assertEqual(payload['total'], ParsedMessage.objects.filter(complaint_control__isnull=False).count())
        self.assertEqual(payload['pending'], 2)
        self.assertEqual(payload['resolved'], 0)
        self.assertEqual(payload['time_granularity'], 'month')
        self.assertEqual(len(payload['by_time']), 1)
        self.assertEqual(set(payload['filter_options']), {'branches', 'categories'})

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_report_summary_and_table_share_filters_and_time_grouping(self):
        july = timezone.make_aware(datetime(2026, 7, 10, 9, 30))
        august = timezone.make_aware(datetime(2026, 8, 2, 9, 30))
        self.case_a.timestamp = july
        self.case_a.save(update_fields=['timestamp'])
        self.case_b.timestamp = august
        self.case_b.complaint_status = 'Closed'
        self.case_b.date_resolved = august + timedelta(days=2)
        self.case_b.save(update_fields=['timestamp', 'complaint_status', 'date_resolved'])
        filters = {
            'search': 'Alice', 'status': 'pending', 'branch': 'Nakuru',
            'category': 'Product issue', 'date_from': '2026-07-01', 'date_to': '2026-07-31',
        }

        table = self.get('complaint_reports_data', filters).json()
        summary = self.get('complaint_reports_summary', {**filters, 'granularity': 'day'}).json()

        self.assertEqual(table['count'], 1)
        self.assertEqual(summary['total'], table['count'])
        self.assertEqual(summary['pending'], 1)
        self.assertEqual(summary['resolved'], 0)
        self.assertEqual(summary['by_time'], [{'label': '2026-07-10', 'count': 1}])
        self.assertEqual(summary['time_granularity'], 'day')
        self.assertEqual(
            {item['label'] for item in summary['filter_options']['branches']},
            {'Nakuru', 'Embu'},
        )
        self.assertNotIn('customer_name', summary)
        self.assertNotIn('phone_number', json.dumps(summary))

        expected = {
            'week': '2026-07-06', 'month': '2026-07', 'year': '2026',
        }
        for granularity, label in expected.items():
            grouped = self.get(
                'complaint_reports_summary', {**filters, 'granularity': granularity},
            ).json()
            self.assertEqual(grouped['by_time'], [{'label': label, 'count': 1}])

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_report_summary_rejects_invalid_or_excessive_time_grouping(self):
        invalid = self.get('complaint_reports_summary', {'granularity': 'quarter'})
        self.case_a.timestamp = timezone.make_aware(datetime(2026, 7, 10, 9, 30))
        self.case_b.timestamp = timezone.make_aware(datetime(2026, 7, 11, 9, 30))
        self.case_a.save(update_fields=['timestamp'])
        self.case_b.save(update_fields=['timestamp'])
        with patch('core.services.complaint_register.REPORT_MAX_TIME_BUCKETS', 1):
            excessive = self.get('complaint_reports_summary', {'granularity': 'day'})

        self.assertEqual(invalid.status_code, 400)
        self.assertEqual(invalid.json()['code'], 'invalid_report_query')
        self.assertEqual(excessive.status_code, 400)
        self.assertIn('too many time periods', excessive.json()['message'])

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_report_maps_every_legacy_status_without_exposing_raw_values(self):
        for stored, expected, needs_details in (
            ('', 'Pending', False), ('Open', 'Pending', False),
            ('In Progress', 'Pending', False), ('Review Needed', 'Pending', True),
            ('Closed', 'Resolved', False),
        ):
            self.case_a.complaint_status = stored
            self.case_a.date_resolved = timezone.now() if stored == 'Closed' else None
            self.case_a.save(update_fields=['complaint_status', 'date_resolved'])
            row = self.get('complaint_reports_data', {'search': 'CMP900001'}).json()['results'][0]
            self.assertEqual(row['status'], expected)
            self.assertEqual(row['needs_details'], needs_details)
            self.assertNotEqual(row['status'], stored or 'blank')

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_cross_group_detail_is_allowlisted_and_actions_are_target_scoped(self):
        response = self.post('complaint_cases_global_detail', args=[self.case_b.pk])

        self.assertEqual(response.status_code, 200)
        item = response.json()['case']
        self.assertEqual(item['description'], 'Unit is not producing gas.')
        self.assertNotIn('raw_message', item)
        self.assertNotIn('evidence', item)
        self.assertNotIn('updates', item)
        self.assertEqual(item['actions'], {
            'close': False, 'reopen': False, 'complete_details': False, 'sync_retry': False,
        })

        AccessGrant.objects.create(
            user=self.officer, workflow='complaint_cases', role='OFFICER',
            group_configuration=self.group_b,
        )
        allowed = self.post('complaint_cases_global_detail', args=[self.case_b.pk]).json()['case']['actions']
        self.assertFalse(allowed['reopen'])
        self.assertFalse(allowed['close'])
        self.assertTrue(allowed['complete_details'])

        AccessGrant.objects.filter(
            user=self.officer, workflow='complaint_cases', group_configuration=self.group_b,
        ).update(role='MANAGER')
        manager_actions = self.post('complaint_cases_global_detail', args=[self.case_b.pk]).json()['case']['actions']
        self.assertTrue(manager_actions['complete_details'])
        self.assertTrue(manager_actions['reopen'])

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_export_requires_all_case_confirmation_and_audits_safe_workbook(self):
        denied = self.post('complaint_cases_global_export', {'confirm_all': False})
        self.assertEqual(denied.status_code, 400)

        response = self.post('complaint_cases_global_export', {
            'confirm_all': True, 'client_request_id': 'global-export-confirmed-1',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['X-Export-Row-Count'], '2')
        from openpyxl import load_workbook
        sheet = load_workbook(BytesIO(response.content), read_only=True).active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0], (
            'Complaint ID', 'Customer Name', 'Phone Number', 'Customer ID', 'Branch',
            'Category', 'Complaint', 'Status', 'Reported At', 'Resolved At',
            'Days Open', 'Resolution',
        ))
        customer_column = rows[0].index('Customer Name')
        phone_column = rows[0].index('Phone Number')
        self.assertIn("'=HYPERLINK(\"bad\")", {row[customer_column] for row in rows[1:]})
        self.assertTrue(all(str(row[phone_column]).startswith("'") for row in rows[1:]))
        self.assertEqual({row[0] for row in rows[1:]}, {'CMP900001', 'CMP900002'})
        audit = ComplianceAuditEvent.objects.get(
            workflow='complaint_cases', action='register.exported', request_id='global-export-confirmed-1',
        )
        self.assertEqual(audit.after_values['row_count'], 2)
        self.assertEqual(audit.after_values['fields'], list(rows[0]))
        self.assertNotIn('private raw source', json.dumps(audit.after_values))

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_report_advertises_and_applies_only_the_requested_filters(self):
        self.case_b.complaint_status = 'Closed'
        self.case_b.date_resolved = timezone.now()
        self.case_b.save(update_fields=['complaint_status', 'date_resolved'])
        today = timezone.localdate().isoformat()

        overview = self.post('complaint_cases_global_overview').json()['data']
        listing = self.post('complaint_cases_global_list', {
            'filters': {
                'query': 'Alice', 'status': 'pending',
                'category': 'Product issue',
                'reported_from': today, 'reported_to': today,
            },
            'page': 1,
        }).json()

        self.assertEqual(set(overview['filters']), {'categories', 'statuses'})
        self.assertEqual(listing['pagination']['total'], 1)
        self.assertEqual(listing['items'][0]['id'], str(self.case_a.pk))

        reference_search = self.post('complaint_cases_global_list', {
            'filters': {'query': 'CMP900001'}, 'page': 1,
        }).json()
        self.assertEqual(reference_search['pagination']['total'], 1)
        self.assertEqual(reference_search['items'][0]['id'], str(self.case_a.pk))

    def test_disabled_projection_keeps_backlog_visible_as_suspended(self):
        self.group_b.complaint_sheet_projection_enabled = False
        self.group_b.save(update_fields=['complaint_sheet_projection_enabled'])
        self.case_b.complaint_control.sync_status = 'failed'
        self.case_b.complaint_control.save(update_fields=['sync_status'])

        overview = register_overview()

        self.assertEqual(overview['metrics']['suspended'], 1)
        self.assertEqual(overview['metrics']['sync_attention'], 1)
        self.group_b.complaint_sheet_projection_enabled = True
        self.group_b.save(update_fields=['complaint_sheet_projection_enabled'])
        restored = register_overview()
        self.assertEqual(restored['metrics']['suspended'], 0)
        self.assertEqual(restored['metrics']['sync_attention'], 2)

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_review_needed_is_pending_with_an_explicit_detail_flag(self):
        self.case_a.complaint_status = 'Review Needed'
        self.case_a.save(update_fields=['complaint_status'])

        overview = register_overview()
        rows = self.post('complaint_cases_global_list', {
            'filters': {'status': 'pending'}, 'page': 1,
        }).json()['items']
        row = next(item for item in rows if item['id'] == str(self.case_a.pk))

        self.assertEqual(overview['metrics']['needs_details'], 1)
        self.assertEqual(row['status'], 'Pending')
        self.assertTrue(row['needs_details'])

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token', SECURE_SSL_REDIRECT=False)
    def test_global_register_filters_and_paginates_fifty_rows(self):
        for index in range(49):
            self.create_case(self.group_a, f'CMP-GLOBAL-{index:03d}', f'Client {index:03d}')
        first = self.post('complaint_cases_global_list', {'filters': {}, 'page': 1})
        second = self.post('complaint_cases_global_list', {'filters': {}, 'page': 2})
        filtered = self.post('complaint_cases_global_list', {
            'filters': {'group': self.group_b.group_id, 'branch': 'Embu'}, 'page': 1,
        })

        self.assertEqual(first.json()['pagination'], {'page': 1, 'pages': 2, 'page_size': 50, 'total': 51})
        self.assertEqual(len(first.json()['items']), 50)
        self.assertEqual(len(second.json()['items']), 1)
        self.assertEqual(filtered.json()['pagination']['total'], 1)
        self.assertEqual(filtered.json()['items'][0]['id'], str(self.case_b.pk))


class ComplaintCaseOptionalSheetTests(TestCase):
    def setUp(self):
        self.group = GroupSheetConfiguration.objects.create(
            group_id='-100django-only', display_name='Django only complaints',
            sheet_id='', sheet_name='Complaints', workflow={'type': 'case'},
            complaint_sheet_projection_enabled=False,
        )
        self.config = GroupConfig(
            group_id=self.group.group_id, sheet_id='', sheet_name='Complaints',
            workflow={'type': 'case'}, complaint_sheet_projection_enabled=False,
        )
        self.category = ComplaintCategory.objects.create(
            key='django-only-product', label='Product issue', default_priority='normal', default_sla_hours=72,
        )
        User = get_user_model()
        self.officer = User.objects.create_user(username='django-only-officer', is_active=True)
        self.manager = User.objects.create_user(username='django-only-manager', is_active=True)
        self.hb_staff = User.objects.create_user(username='django-only-hb', is_active=True)
        UserProfile.objects.create(user=self.officer, telegram_id='901')
        UserProfile.objects.create(user=self.manager, telegram_id='902')
        UserProfile.objects.create(user=self.hb_staff, telegram_id='903')
        AccessGrant.objects.create(user=self.officer, workflow='complaint_cases', role='OFFICER', group_configuration=self.group)
        AccessGrant.objects.create(user=self.manager, workflow='complaint_cases', role='MANAGER', group_configuration=self.group)
        AccessGrant.objects.create(user=self.hb_staff, workflow='complaint_cases', role='HB_STAFF', group_configuration=self.group)

    @patch('core.services.complaint_cases.get_sheets_service')
    @patch('core.services.complaint_cases.append_parsed_message_to_sheet')
    def test_disabled_projection_never_calls_sheets_for_create_or_resolution(self, append_to_sheet, get_service):
        officer = staff_actor_for_user(self.config, self.officer)
        result = create_complaint_case(
            self.config, officer, {
                'client_request_id': 'django-only-create-1', 'client_name': 'Local Client',
                'customer_id': '000001', 'branch_region': 'Nakuru',
                'complaint_category': 'Product issue', 'complaint_description': 'Saved in Django only.',
            }, [],
        )
        case = ParsedMessage.objects.get(pk=result['case']['id'])
        resolver = staff_actor_for_user(self.config, self.hb_staff)
        resolved = resolve_case(
            self.config, resolver, case.message_id, {
                'client_request_id': 'django-only-resolve-1', 'expected_revision': 1,
                'resolution_text': 'Resolved without a spreadsheet.',
            }, [],
        )

        self.assertFalse(result['sheet_projection_enabled'])
        self.assertEqual(resolved['sync_status'], 'not_required')
        self.assertEqual(case.case_updates.order_by('-created_at').first().sync_status, 'not_required')
        append_to_sheet.assert_not_called()
        get_service.assert_not_called()

    def test_disabled_projection_allows_an_enabled_group_without_sheet_id(self):
        self.group.full_clean()
        from core.services.group_config import GroupRegistry
        registry = GroupRegistry.get_instance()
        registry.reload()
        self.assertEqual(registry.get_group(self.group.group_id).group_id, self.group.group_id)

        invalid = GroupSheetConfiguration(
            group_id='-100missing-required-sheet', workflow={'type': 'case'},
            complaint_sheet_projection_enabled=True,
        )
        with self.assertRaises(ValidationError):
            invalid.full_clean()

    @patch('core.services.complaint_cases.append_parsed_message_to_sheet')
    def test_cached_create_retry_does_not_erase_a_suspended_failure(self, append_to_sheet):
        officer = staff_actor_for_user(self.config, self.officer)
        payload = {
            'client_request_id': 'django-only-retry-1', 'client_name': 'Retry Client',
            'customer_id': '000002', 'branch_region': 'Nakuru',
            'complaint_category': 'Product issue', 'complaint_description': 'Retain old sync evidence.',
        }
        first = create_complaint_case(self.config, officer, payload, [])
        control = ComplaintCaseControl.objects.get(parsed_message_id=first['case']['id'])
        control.sync_status = 'failed'
        control.sync_error = 'Pre-cutover publication failed.'
        control.save(update_fields=['sync_status', 'sync_error'])

        retried = create_complaint_case(self.config, officer, payload, [])

        control.refresh_from_db()
        self.assertFalse(retried['created'])
        self.assertEqual(control.sync_status, 'failed')
        self.assertEqual(control.sync_error, 'Pre-cutover publication failed.')
        append_to_sheet.assert_not_called()


class ComplaintCaseMiniAppAssetTests(TestCase):
    def test_compact_two_state_workspace_has_only_supported_actions(self):
        root = Path(__file__).resolve().parent
        template = (root / 'templates' / 'complaint_cases' / 'app.html').read_text(encoding='utf-8')
        icons = (root / 'templates' / 'complaint_cases' / 'lucide_icons.html').read_text(encoding='utf-8')
        script = (root / 'static' / 'miniapp' / 'complaint_cases.js').read_text(encoding='utf-8')
        styles = (root / 'static' / 'miniapp' / 'complaint_cases.css').read_text(encoding='utf-8')
        package_lock = (root.parent / 'package-lock.json').read_text(encoding='utf-8')
        ag_grid_license = (
            root / 'static' / 'miniapp' / 'vendor-ag-grid-community-36.1.0.LICENSE.txt'
        ).read_text(encoding='utf-8')

        for expected in ('class="app-top"', 'class="status-tabs"', 'id="createCaseForm"', 'name="client_name"', 'name="customer_phone"', 'name="customer_id"', 'name="branch_region"', 'name="complaint_category"', 'name="complaint_description"', 'id="createEvidenceInput"', 'data-status="pending"', 'data-status="resolved"', 'data-status="all"', 'id="completeDetailsForm"', 'id="resolveForm"', 'id="reopenForm"', 'id="conflictPanel"', 'id="queuePagination"'):
            self.assertIn(expected, template)
        for expected in ('id="globalView"', 'id="globalFilters"', 'id="complaintReportGrid"', 'id="exportConfirm"'):
            self.assertIn(expected, template)
        for expected in ('name="search"', 'name="status"', 'name="branch"', 'name="category"', 'name="date_mode"', 'name="report_month"', 'name="date_from"', 'name="date_to"'):
            self.assertIn(expected, template)
        for removed in ('name="group"', 'name="priority"', 'name="sla"', 'name="sync"', 'name="sort"'):
            self.assertNotIn(removed, template)
        self.assertNotIn('<table class="register-table">', template)
        self.assertIn('vendor-ag-grid-community-36.1.0.min.js', template)
        self.assertIn('vendor-chartjs-4.5.1.umd.min.js', template)
        self.assertIn('MIT License', ag_grid_license)
        self.assertIn('"ag-grid-community"', package_lock)
        self.assertNotIn('"ag-grid-enterprise"', package_lock)
        self.assertNotIn('"@ag-grid-enterprise/', package_lock)
        self.assertIn('type="date"', template)
        self.assertIn('type="month"', template)
        self.assertIn('Complaints over Time', template)
        self.assertIn('data-category-chart="bar"', template)
        self.assertIn('data-category-chart="pie"', template)
        self.assertIn('id="reportGranularity"', template)
        self.assertIn('inputmode="numeric" pattern="[0-9]*"', template)
        self.assertIn('id="mediaViewerOverlay"', template)
        self.assertIn('class="filter-search-control"', template)
        self.assertIn('id="downloadResult"', template)
        self.assertIn('id="openExportBtn"', template)
        self.assertIn('id="downloadAgainBtn"', template)
        self.assertIn('secure_media_viewer.js', template)
        self.assertIn('complaint_cases/lucide_icons.html', template)
        self.assertIn('href="#lucide-refresh-cw"', template)
        self.assertIn('function iconNode(name, className)', script)
        for icon in ('refresh-cw', 'clipboard-list', 'layout-dashboard', 'camera', 'eye', 'trash-2'):
            self.assertIn(f'id="lucide-{icon}"', icons)
        self.assertNotIn('unpkg.com/lucide', template)
        self.assertIn('This download includes all ${count} complaints across all complaint groups', script)
        self.assertIn('Check Downloads for ${state.exportFilename}', script)
        self.assertIn("navigator.canShare({ files: [state.exportFile] })", script)
        self.assertIn("navigator.share({ files: [state.exportFile]", script)
        self.assertIn("match[3]}-${match[2]}-${match[1].slice(-2)", script)
        self.assertIn('unSortIcon: true', script)
        self.assertIn('.report-status{font-size:11px;font-weight:850;white-space:nowrap}', styles)
        self.assertNotIn('.report-status{display:inline-flex', styles)
        self.assertIn("utils.haptic?.(error ? 'error' : 'success')", script)
        self.assertIn('grid-template-columns:minmax(0,1fr) auto minmax(0,1fr)', styles)
        for wording in (
            '<h1>Complaints</h1>', '<span>Complaints</span>', '<span>Data Overview</span>',
            'Record a New Complaint', 'Enter the customer&rsquo;s complaint details below.',
            'Complaint Type', 'What is the complaint about?', 'Use My Current Location',
            'Supporting Documents or Photos', 'Take Photos', 'Upload Files',
            'Not Saved', 'Submit Complaint',
        ):
            self.assertIn(wording, template)
        for jargon in ('Officer Intake', 'Shared queue', 'Choose the primary complaint.', 'Branch not set'):
            self.assertNotIn(jargon, template)
        self.assertIn("'Branch not provided'", script)
        service = (root / 'services' / 'complaint_cases.py').read_text(encoding='utf-8')
        self.assertIn("f'Pending for {age_days} day'", service)
        self.assertIn("f'Resolved after {age_days} day'", service)
        self.assertIn("case-age ${resolved ? 'resolved'", script)
        self.assertIn('.case-age.resolved{color:var(--success)}', styles)
        for wording in (
            'Management Report', 'Read-only organization-wide complaint data', 'Download Complaints',
            'Any Status', 'Any Category', 'Date Reported', 'Start Date', 'End Date',
            'Show Results', 'Reset Filters', 'Needs More Information',
            'Resolution History', 'Reason for Reopening', 'Attachments', 'Complaint History',
            'Include what was fixed or completed, when it was done',
        ):
            self.assertIn(wording, template)
        self.assertIn("'Resolved by'", script)
        self.assertIn("'Reopened by'", script)
        self.assertIn("'Complaint recorded by'", script)
        self.assertIn('normalizeCustomerNameInput(formNode.elements.client_name)', script)
        self.assertIn('autocapitalize="words"', template)
        self.assertIn('Sheet Sync:', script)
        self.assertNotIn('Retry Sheet Sync', template)
        self.assertNotIn('retrySyncBtn', script)
        self.assertIn('id="appHeader"', template)
        self.assertIn('function bindCollapsingHeader()', script)
        self.assertIn("header.classList.add('header-hidden')", script)
        self.assertIn('item.reference_number || item.case_id', script)
        self.assertIn('Search by complaint ID', template)
        self.assertIn('family=Plus+Jakarta+Sans', template)
        self.assertIn("can('complaint.reports.view')", script)
        self.assertIn("getJson('reports/data/'", script)
        self.assertIn("getJson('reports/summary/'", script)
        self.assertIn('function monthBoundaries(value)', script)
        self.assertIn('filter_options?.branches', script)
        self.assertIn('overflow-x:hidden', styles)
        self.assertIn('AG Grid owns horizontal scrolling', styles)
        self.assertNotIn('ag-grid-enterprise', template.casefold())
        self.assertIn("'the case is now fully resolved': 'Complaint marked as resolved'", script)
        self.assertIn("'the customer is still complaining': 'Customer reported the issue again'", script)
        self.assertIn("miniapp/utils.js", template)
        self.assertIn("data.set('expected_revision'", script)
        self.assertIn("submitTransition(event, 'resolve')", script)
        self.assertIn("submitTransition(event, 'reopen')", script)
        self.assertIn('function submitCompleteDetails(event)', script)
        self.assertIn("getUserMedia({ video: { facingMode: { ideal: 'environment' } }", script)
        self.assertIn("telegram?.onEvent?.('deactivated'", script)
        self.assertIn("document.addEventListener('visibilitychange'", script)
        self.assertIn('state.evidenceLimits.max_total_upload_mb', script)
        self.assertIn('openSelectedEvidence', script)
        self.assertIn('openPersistedEvidence', script)
        self.assertIn('id="mediaViewerPrevious"', template)
        self.assertIn('id="mediaViewerNext"', template)
        self.assertIn('id="mediaViewerDelete"', template)
        self.assertIn('id="mediaViewerRetake"', template)
        self.assertIn("setLocationCaptureState('success', 'Location Captured'", script)
        self.assertIn('`GPS: ${state.latitude}, ${state.longitude}`', script)
        self.assertIn('.location-coordinate', styles)
        self.assertIn('--secondary-text:', styles)
        self.assertIn('#mediaViewerClose{width:40px;min-width:40px;flex:0 0 40px', styles)
        self.assertIn('<span>Take Photo</span>', template)
        self.assertNotIn('>Use Photo</button>', template)
        self.assertIn("notify('Photo added. Take another photo or tap Done.')", script)
        self.assertIn('function navigateMediaViewer(offset)', script)
        self.assertIn('function deleteSelectedMediaFromViewer()', script)
        self.assertIn('function retakeSelectedMediaFromViewer()', script)
        self.assertIn("else openPersistedEvidence(target, null)", script)
        self.assertNotIn('telegram.openLink(response.url)', script)
        self.assertNotIn('.docx', template)
        self.assertIn("json('categories/suggest/'", script)
        self.assertIn('function showConflict(error)', script)
        self.assertIn("telegram?.BackButton?.onClick", script)
        for retired in ('priorityFilter', 'assignmentFilter', 'slaFilter', 'claimBtn', 'Settings', 'In Progress'):
            self.assertNotIn(retired, template)


class ComplaintCategoryCatalogueTests(TestCase):
    def test_migration_seeds_the_exact_active_category_catalogue(self):
        active = dict(ComplaintCategory.objects.filter(active=True).values_list('label', 'description'))

        self.assertEqual(active, {
            'Leakage': 'Gas, bag, pipe, connection, valve, etc.',
            'Blockage': 'Inlet or outlet blockage',
            'Burner/Knob Fault': 'Burner, knob, flame, ignition issues',
            'Pipe/Connection Fault': (
                'Physical pipe/connection problems where leakage is NOT the primary complaint'
            ),
            'System Performance': 'Low/no gas production or poor system performance',
            'Installation Delay': "Installation hasn't happened/delayed",
            'Commissioning Delay': 'Commissioning/start-up delayed',
            'Accessories Delay': 'Accessories requested but delayed',
            'Relocation Request': 'Customer wants system relocated',
            'Other Complaint': "Doesn't fit any category",
        })
        self.assertEqual(ComplaintCategory.objects.get(key='other-complaint').default_sla_hours, 72)


class ComplaintCaseAdminTests(TestCase):
    def test_group_admin_has_no_legacy_staff_inlines(self):
        complaint_group = GroupSheetConfiguration.objects.create(
            group_id='-100complaints', workflow={'type': 'case'}
        )
        tat_group = GroupSheetConfiguration.objects.create(
            group_id='-100tat', workflow={'type': 'tat_tracker'}
        )
        model_admin = admin.site._registry[GroupSheetConfiguration]
        request = RequestFactory().get('/admin/core/groupsheetconfiguration/')

        self.assertEqual(model_admin.get_inlines(request, complaint_group), [])
        self.assertEqual(model_admin.get_inlines(request, tat_group), [])


class TelegramInitDataTests(TestCase):
    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token')
    def test_valid_signed_init_data_is_accepted(self):
        pairs = {'auth_date': str(int(time.time())), 'user': json.dumps({'id': 1})}
        check = '\n'.join(f'{key}={value}' for key, value in sorted(pairs.items()))
        secret = hmac.new(b'WebAppData', b'test-bot-token', hashlib.sha256).digest()
        pairs['hash'] = hmac.new(secret, check.encode('utf-8'), hashlib.sha256).hexdigest()
        valid, error, payload = validate_telegram_init_data(urlencode(pairs))
        self.assertTrue(valid, error)
        self.assertEqual(payload['user'], json.dumps({'id': 1}))

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token')
    def test_bad_or_expired_init_data_is_rejected(self):
        valid, _, _ = validate_telegram_init_data('auth_date=1&hash=bad')
        self.assertFalse(valid)
