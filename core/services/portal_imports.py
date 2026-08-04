"""Portal-owned staging and archival for FarmUp and SysUp source files.

The Portal is intentionally a review surface only in this release.  It uses
the established FarmUp/SysUp parsers and staged-batch records, but never calls
their commit functions.  A later, separately approved release may expose a
maker-checker commit action without changing how files are staged or archived.
"""

from __future__ import annotations

import csv
import hashlib
import io
import mimetypes
import re
from pathlib import PurePath
from typing import Any

from django.conf import settings
from django.db import IntegrityError, transaction
from django.utils import timezone

from core.models import GroupSheetConfiguration, IntegrationOperation, JawabuFarmerUploadBatch
from core.services.compliance_audit import record_event
from core.services.document_sync import mark_drive_attempt, mark_drive_failure, mark_drive_success
from core.services.external_resilience import ExternalOperationError, execute_operation, reserve_operation


IMPORT_KIND_FARMUP = 'farmup'
IMPORT_KIND_SYSUP = 'sysup'
IMPORT_KINDS = frozenset({IMPORT_KIND_FARMUP, IMPORT_KIND_SYSUP})
SOURCE_MODEL = 'core.JawabuFarmerUploadBatch'
ARCHIVE_OPERATION = 'portal_import_drive_archive'
_JAWABU_IMPORT_WORKFLOW_TYPE = 'jawabu_homebiogas'
_SAFE_FILENAME = re.compile(r'[^A-Za-z0-9._ -]+')


class PortalImportError(ValueError):
    """A stable validation failure suitable for a staff-facing API response."""


def resolve_import_group(*, allowed_group_ids: set[str] | None = None) -> GroupSheetConfiguration:
    """Resolve the one configured Jawabu HomeBiogas import destination.

    FarmUp and SysUp are Portal-owned intake routes, not a multi-group import
    tool. Letting the browser choose a Telegram group made the upload route
    ambiguous and could block an authorised IT user before sending the file.
    Ordinary AccessGrant group scope is still enforced here.
    """
    configurations = [
        config
        for config in GroupSheetConfiguration.objects.filter(enabled=True).order_by('group_id')
        if str((config.workflow or {}).get('type') or '').strip() == _JAWABU_IMPORT_WORKFLOW_TYPE
    ]
    if not configurations:
        raise PortalImportError('The Jawabu HomeBiogas import workflow is not configured. Ask IT to configure it before staging files.')
    if len(configurations) > 1:
        raise PortalImportError('More than one Jawabu HomeBiogas import workflow is configured. Ask IT to correct the configuration before staging files.')
    configuration = configurations[0]
    if allowed_group_ids is not None and str(configuration.group_id) not in allowed_group_ids:
        raise PortalImportError('Your Portal import access does not cover the configured Jawabu HomeBiogas workflow.')
    return configuration


def _source_limit_bytes(kind: str) -> int:
    if kind == IMPORT_KIND_FARMUP:
        max_mb = max(1, int(getattr(settings, 'FARMUP_MAX_FILE_SIZE_MB', 5) or 5))
    else:
        max_mb = max(1, int(getattr(settings, 'SYSUP_MAX_FILE_SIZE_MB', 5) or 5))
    return max_mb * 1024 * 1024


def _validated_source(kind: str, *, filename: str, content: bytes) -> tuple[str, str, bytes]:
    normalized_kind = str(kind or '').strip().lower()
    if normalized_kind not in IMPORT_KINDS:
        raise PortalImportError('Select either FarmUp or SysUp.')
    safe_name = PurePath(str(filename or '')).name.strip()
    lower_name = safe_name.lower()
    if normalized_kind == IMPORT_KIND_FARMUP and not lower_name.endswith('.csv'):
        raise PortalImportError('FarmUp accepts a Jawabu Farmers CSV file only.')
    if normalized_kind == IMPORT_KIND_SYSUP and not lower_name.endswith(('.csv', '.xlsx')):
        raise PortalImportError('SysUp accepts a Customers Without Loans CSV or XLSX export.')
    payload = bytes(content or b'')
    if not payload:
        raise PortalImportError('Choose an import file before staging it.')
    if len(payload) > _source_limit_bytes(normalized_kind):
        raise PortalImportError('This import exceeds the configured safe upload size. Split it into smaller files and retry.')
    mime_type = mimetypes.guess_type(safe_name)[0] or 'application/octet-stream'
    return normalized_kind, safe_name, payload


def _decode_farmup_csv(content: bytes) -> str:
    for encoding in ('utf-8-sig', 'utf-8'):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise PortalImportError('FarmUp CSV must be UTF-8 encoded.')


def _decode_sysup_csv(content: bytes) -> str:
    """Decode the CSV encodings accepted by the existing SysUp parser."""
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252'):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise PortalImportError('SysUp CSV could not be read. Export it as UTF-8 CSV and retry.')


def _source_csv_table_page(
    content: bytes,
    *,
    decoder,
    page: int,
    page_size: int,
) -> dict[str, list[list[str]] | list[str]]:
    """Return one raw CSV page without exposing parser or review-only fields.

    ``parsed_rows`` deliberately contains normalized and matching metadata used
    by the command workflows.  The Portal's review-only screen instead needs
    to show the retained source exactly in its own column order, so it reads
    the original CSV values as a table of lists rather than dictionaries.
    Lists also preserve duplicate or blank source headers without inventing
    replacement column names.
    """
    reader = csv.reader(io.StringIO(decoder(content)))
    try:
        headers = next(reader)
    except StopIteration:
        return {'headers': [], 'rows': []}

    start = max(0, (max(1, page) - 1) * max(1, page_size))
    end = start + max(1, page_size)
    rows: list[list[str]] = []
    visible_position = 0
    for values in reader:
        if not any(str(value or '').strip() for value in values):
            continue
        if start <= visible_position < end:
            rows.append(list(values[:len(headers)]) + [''] * max(0, len(headers) - len(values)))
        visible_position += 1
        if visible_position >= end:
            break
    return {'headers': list(headers), 'rows': rows}


def _source_xlsx_table_page(
    content: bytes,
    *,
    page: int,
    page_size: int,
) -> dict[str, list[list[str]] | list[str]]:
    """Return a raw SysUp workbook page using the same header detection rule.

    This intentionally stops after the requested page.  It avoids loading an
    entire workbook into the Portal response merely to render the source data
    that an IT reviewer asked to inspect.
    """
    try:
        from openpyxl import load_workbook
        from core.services.system_export import REQUIRED_HEADER_KEYS, _cell_text, _header_key

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise PortalImportError('The retained SysUp workbook could not be read for review.') from exc

    try:
        sheet = workbook.active
        header_row_number = None
        headers: list[str] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
            candidate = [_cell_text(value) for value in row]
            keys = {_header_key(value) for value in candidate if value}
            if set(REQUIRED_HEADER_KEYS).issubset(keys):
                header_row_number = row_number
                headers = candidate
                break
        if header_row_number is None:
            raise PortalImportError('The retained SysUp workbook has no recognizable header row.')

        start = max(0, (max(1, page) - 1) * max(1, page_size))
        end = start + max(1, page_size)
        rows: list[list[str]] = []
        visible_position = 0
        for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
            values = [_cell_text(value) for value in row]
            if not any(values):
                continue
            if start <= visible_position < end:
                rows.append(values[:len(headers)] + [''] * max(0, len(headers) - len(values)))
            visible_position += 1
            if visible_position >= end:
                break
        return {'headers': headers, 'rows': rows}
    finally:
        workbook.close()


def source_table_page(
    batch: JawabuFarmerUploadBatch,
    *,
    page: int,
    page_size: int,
) -> dict[str, list[list[str]] | list[str]]:
    """Return the original staged import columns and values for Portal review.

    This is intentionally presentation-only.  It never re-runs matching,
    normalization, or a customer-data commit and does not expose derived
    parser fields such as Import Status, matched customer, or cleaning notes.
    """
    content = bytes(batch.source_content or b'')
    if not content:
        raise PortalImportError('The original source file is unavailable for this staged import.')
    if batch.import_kind == 'farmers':
        return _source_csv_table_page(
            content,
            decoder=_decode_farmup_csv,
            page=page,
            page_size=page_size,
        )
    if str(batch.source_filename or '').lower().endswith('.xlsx'):
        return _source_xlsx_table_page(content, page=page, page_size=page_size)
    return _source_csv_table_page(
        content,
        decoder=_decode_sysup_csv,
        page=page,
        page_size=page_size,
    )


def _existing_replay(request_id: str, *, kind: str, source_hash: str) -> JawabuFarmerUploadBatch | None:
    if not request_id:
        return None
    batch = JawabuFarmerUploadBatch.objects.filter(upload_request_id=request_id).first()
    if batch is None:
        return None
    if batch.source_content_hash != source_hash or batch.import_kind != ('farmers' if kind == IMPORT_KIND_FARMUP else 'system_export'):
        raise PortalImportError('This retry key belongs to a different import. Reload the screen and submit again.')
    return batch


def _assert_replay_is_in_scope(
    batch: JawabuFarmerUploadBatch,
    *,
    allowed_group_ids: set[str] | None,
) -> None:
    """Keep an idempotent replay inside the caller's original import scope.

    A retry key must replay exactly one source upload, not become a way to
    retrieve a staged file from a different configured Telegram group.
    """
    if allowed_group_ids is not None and str(batch.group_id) not in allowed_group_ids:
        raise PortalImportError('This staged import is unavailable in your scope.')


def archive_portal_import_working_list(
    *,
    batch_id: str,
    actor,
    request_id: str,
    allowed_group_ids: set[str] | None = None,
) -> tuple[JawabuFarmerUploadBatch, bool]:
    """Hide one retained staged import from the active Portal working list.

    This deliberately does not call Google Drive or alter the underlying
    FarmUp/SysUp workflow status.  It is safe for Mini App retries: once a
    batch is archived, all later retries return the retained batch without
    producing a second audit event.
    """
    request_id = str(request_id or '').strip()
    if not request_id:
        raise PortalImportError('This archive action needs a retry key. Reload Imports and try again.')
    with transaction.atomic():
        batch = JawabuFarmerUploadBatch.objects.select_for_update().filter(pk=batch_id).first()
        if batch is None:
            raise PortalImportError('This staged import is unavailable.')
        _assert_replay_is_in_scope(batch, allowed_group_ids=allowed_group_ids)
        if batch.is_portal_archived:
            return batch, True

        before = {
            'working_list_active': True,
            'status': batch.status,
            'drive_archive_state': import_archive_state(batch),
        }
        archived_at = timezone.now()
        batch.is_portal_archived = True
        batch.portal_archived_at = archived_at
        batch.portal_archived_by = actor
        batch.save(update_fields=[
            'is_portal_archived', 'portal_archived_at', 'portal_archived_by', 'updated_at',
        ])
        record_event(
            workflow='portal',
            action='portal.import.working_list_archived',
            category='workflow',
            subject_type='JawabuFarmerUploadBatch',
            subject_id=str(batch.pk),
            deduplication_key=f'portal-import-working-list-archive:{batch.pk}',
            actor=actor,
            request_id=request_id,
            source_model='JawabuFarmerUploadBatch',
            source_event_id=f'{batch.pk}:working-list-archive',
            before_values=before,
            after_values={
                'working_list_active': False,
                'status': batch.status,
                'drive_archive_state': import_archive_state(batch),
                'archived_at': archived_at,
            },
            metadata={'import_kind': batch.import_kind, 'group_id': batch.group_id},
            sensitive=True,
        )
    return batch, False


def reserve_import_archive(batch: JawabuFarmerUploadBatch, *, request_id: str, user=None) -> IntegrationOperation:
    """Reserve an idempotent Drive archive attempt without performing I/O."""
    operation, _ = reserve_operation(
        integration=IntegrationOperation.INTEGRATION_GOOGLE_DRIVE,
        operation_type=ARCHIVE_OPERATION,
        deduplication_key=f'portal-import-archive:{batch.pk}:{batch.source_content_hash}',
        source_model=SOURCE_MODEL,
        source_id=str(batch.pk),
        request_id=str(request_id or '')[:128],
        requested_by=user,
        requested_by_label=str(getattr(user, 'get_full_name', lambda: '')() or getattr(user, 'username', '') or '')[:255],
        operation_payload=(str(batch.pk), batch.source_content_hash),
        metadata={'import_kind': batch.import_kind, 'source_size': int(batch.source_size or 0)},
    )
    return operation


def stage_portal_import(
    *,
    kind: str,
    filename: str,
    content: bytes,
    request_id: str,
    actor,
    allowed_group_ids: set[str] | None = None,
) -> tuple[JawabuFarmerUploadBatch, IntegrationOperation, bool]:
    """Create one review-only batch and reserve archival work.

    The request key is mandatory because an Android WebView retry must return
    the original parsed batch rather than create a second import to review.
    """
    request_id = str(request_id or '').strip()
    if not request_id:
        raise PortalImportError('This upload needs a retry key. Reload the Imports screen and try again.')
    normalized_kind, safe_name, payload = _validated_source(kind, filename=filename, content=content)
    source_hash = hashlib.sha256(payload).hexdigest()
    existing = _existing_replay(request_id, kind=normalized_kind, source_hash=source_hash)
    if existing is not None:
        _assert_replay_is_in_scope(
            existing,
            allowed_group_ids=allowed_group_ids,
        )
        return existing, reserve_import_archive(existing, request_id=request_id, user=actor), True
    group_configuration = resolve_import_group(allowed_group_ids=allowed_group_ids)
    try:
        with transaction.atomic():
            # Check again inside the transaction.  The unique request key is
            # the final concurrent-retry guard.
            existing = _existing_replay(request_id, kind=normalized_kind, source_hash=source_hash)
            if existing is not None:
                _assert_replay_is_in_scope(
                    existing,
                    allowed_group_ids=allowed_group_ids,
                )
                return existing, reserve_import_archive(existing, request_id=request_id, user=actor), True
            if normalized_kind == IMPORT_KIND_FARMUP:
                from core.services.jawabu_master import create_farmup_review_batch

                batch, _stats = create_farmup_review_batch(
                    group_id=group_configuration.group_id,
                    telegram_message_id='',
                    sender=str(getattr(actor, 'get_full_name', lambda: '')() or getattr(actor, 'username', '') or ''),
                    source_filename=safe_name,
                    csv_text=_decode_farmup_csv(payload),
                    group_config=group_configuration,
                )
            else:
                from core.services.system_export import create_system_export_review_batch

                batch, _stats = create_system_export_review_batch(
                    group_id=group_configuration.group_id,
                    telegram_message_id='',
                    sender=str(getattr(actor, 'get_full_name', lambda: '')() or getattr(actor, 'username', '') or ''),
                    source_filename=safe_name,
                    content=payload,
                )
            batch.created_by = actor
            batch.upload_request_id = request_id
            batch.source_mime_type = mimetypes.guess_type(safe_name)[0] or 'application/octet-stream'
            batch.source_size = len(payload)
            batch.source_content_hash = source_hash
            batch.source_content = payload
            batch.save(update_fields=[
                'created_by', 'upload_request_id', 'source_mime_type', 'source_size',
                'source_content_hash', 'source_content', 'updated_at',
            ])
            operation = reserve_import_archive(batch, request_id=request_id, user=actor)
            return batch, operation, False
    except IntegrityError:
        # A concurrent retry may have won after parser work completed.  Do not
        # surface a duplicate-batch error or leave callers guessing its state.
        existing = _existing_replay(request_id, kind=normalized_kind, source_hash=source_hash)
        if existing is not None:
            _assert_replay_is_in_scope(
                existing,
                allowed_group_ids=allowed_group_ids,
            )
            return existing, reserve_import_archive(existing, request_id=request_id, user=actor), True
        raise


def _archive_filename(batch: JawabuFarmerUploadBatch) -> str:
    original = _SAFE_FILENAME.sub('_', PurePath(str(batch.source_filename or '')).name).strip(' ._')
    original = original[:160] or ('farmup.csv' if batch.import_kind == 'farmers' else 'system-export.csv')
    prefix = 'FarmUp' if batch.import_kind == 'farmers' else 'SysUp'
    return f'{prefix}_{str(batch.pk)[:8]}_{original}'


def attempt_import_archive(operation_id: str) -> dict[str, Any]:
    """Perform at most one bounded archive attempt for a previously staged file."""
    operation = IntegrationOperation.objects.filter(pk=operation_id).first()
    if operation is None or operation.source_model != SOURCE_MODEL or operation.operation_type != ARCHIVE_OPERATION:
        raise PortalImportError('This import archive operation is unavailable.')
    batch = JawabuFarmerUploadBatch.objects.filter(pk=operation.source_id).first()
    if batch is None:
        raise PortalImportError('The staged import is no longer available.')
    if batch.archive_file_id and batch.archive_url:
        return {'ok': True, 'batch': batch, 'operation': operation, 'replayed': True}
    folder_id = str(getattr(settings, 'GOOGLE_DRIVE_MEDIA_FOLDER_ID', '') or '').strip()
    if not folder_id:
        batch.archive_error = 'The shared Drive archive is not configured.'
        batch.save(update_fields=['archive_error', 'updated_at'])
        return {'ok': False, 'batch': batch, 'operation': operation, 'error': batch.archive_error}
    if not batch.source_content:
        batch.archive_error = 'The original source file is unavailable for Drive archival.'
        batch.save(update_fields=['archive_error', 'updated_at'])
        return {'ok': False, 'batch': batch, 'operation': operation, 'error': batch.archive_error}
    mark_drive_attempt(batch, prefix='archive')

    def archive_once() -> dict[str, str]:
        from core.services.order_approval import GoogleDriveMediaStorage

        # Imports share the approved media Drive root, but the storage gateway
        # creates their own Imports/YYYY/MM-Month/Batch_<id> path beneath it.
        file_id, url = GoogleDriveMediaStorage().upload(
            batch.source_content,
            filename=_archive_filename(batch),
            mime_type=batch.source_mime_type or 'application/octet-stream',
            id_number='portal_imports',
            received_at=batch.created_at or timezone.now(),
            workflow_key='Imports',
            record_type='Batch',
            record_key=str(batch.pk),
            attempt_budget=1,
        )
        return {'id': file_id, 'url': url}

    try:
        result = execute_operation(operation, archive_once, attempt_budget=1)
    except ExternalOperationError:
        mark_drive_failure(batch, 'Drive archive needs attention. Retry from Imports.', prefix='archive', error_field='archive_error')
        operation.refresh_from_db()
        return {'ok': False, 'batch': batch, 'operation': operation, 'error': batch.archive_error}
    if not result:
        operation.refresh_from_db()
        return {'ok': False, 'batch': batch, 'operation': operation, 'error': 'Drive archive is already being processed.'}
    file_id = str(result.get('id') or '')
    url = str(result.get('url') or result.get('webViewLink') or '')
    if not file_id or not url:
        batch.archive_error = 'Drive archive completed without a usable file reference.'
        batch.save(update_fields=['archive_error', 'updated_at'])
        return {'ok': False, 'batch': batch, 'operation': operation, 'error': batch.archive_error}
    mark_drive_success(batch, file_id=file_id, url=url, prefix='archive', error_field='archive_error')
    return {'ok': True, 'batch': batch, 'operation': operation, 'replayed': False}


def import_archive_state(batch: JawabuFarmerUploadBatch) -> str:
    if batch.archive_file_id and batch.archive_url:
        return 'archived'
    if batch.archive_error:
        return 'needs_attention'
    return 'pending'


def archive_operation_ids(batches: list[JawabuFarmerUploadBatch]) -> dict[str, str]:
    """Fetch archive-operation IDs in one query for an already scoped batch list."""
    batch_ids = [str(batch.pk) for batch in batches]
    if not batch_ids:
        return {}
    return {
        str(operation['source_id']): str(operation['id'])
        for operation in IntegrationOperation.objects.filter(
            source_model=SOURCE_MODEL,
            operation_type=ARCHIVE_OPERATION,
            source_id__in=batch_ids,
        ).values('id', 'source_id')
    }


def serialize_import_batch(
    batch: JawabuFarmerUploadBatch, *, include_rows: bool = False, archive_operation_id: str = '',
) -> dict[str, Any]:
    """Return staff-safe metadata; raw bytes and Drive URLs stay private."""
    payload: dict[str, Any] = {
        'id': str(batch.pk),
        'kind': 'farmup' if batch.import_kind == 'farmers' else 'sysup',
        'source_filename': batch.source_filename,
        'source_size': int(batch.source_size or 0),
        'group_id': batch.group_id,
        'status': batch.status,
        'total_rows': batch.total_rows,
        'review_needed': batch.review_needed,
        'committed_count': batch.committed_count,
        'skipped_count': batch.skipped_count,
        'error': batch.error,
        'created_at': batch.created_at.isoformat() if batch.created_at else None,
        'updated_at': batch.updated_at.isoformat() if batch.updated_at else None,
        'created_by': str(getattr(batch.created_by, 'get_full_name', lambda: '')() or getattr(batch.created_by, 'username', '') or batch.sender or ''),
        'archive_state': import_archive_state(batch),
        'archive_error': batch.archive_error or '',
        'archive_attempts': int(batch.archive_sync_attempts or 0),
        'archive_last_attempt_at': batch.archive_last_sync_at.isoformat() if batch.archive_last_sync_at else None,
        'archive_next_retry_at': batch.archive_next_retry_at.isoformat() if batch.archive_next_retry_at else None,
        'archive_operation_id': str(archive_operation_id or ''),
        'is_portal_archived': bool(batch.is_portal_archived),
        'portal_archived_at': batch.portal_archived_at.isoformat() if batch.portal_archived_at else None,
        'portal_archived_by': str(
            getattr(batch.portal_archived_by, 'get_full_name', lambda: '')()
            or getattr(batch.portal_archived_by, 'username', '')
            or ''
        ),
    }
    if include_rows:
        payload['mapping'] = list(batch.mapping or [])
        payload['rows'] = list(batch.parsed_rows or [])
    return payload
