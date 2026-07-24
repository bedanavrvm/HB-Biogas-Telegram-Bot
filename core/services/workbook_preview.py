"""Serialize generated workbooks for a read-only, Excel-like Mini App preview."""

from __future__ import annotations

import io
import base64
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries


def _color(value: Any) -> str:
    if value is None or getattr(value, 'type', None) != 'rgb':
        return ''
    rgb = str(getattr(value, 'rgb', '') or '')
    return f'#{rgb[-6:]}' if len(rgb) >= 6 else ''


def _value(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d-%b-%Y %H:%M')
    if isinstance(value, date):
        return value.strftime('%d-%b-%Y')
    if isinstance(value, Decimal):
        return format(value, 'f')
    return str(value)


def _print_bounds(worksheet, *, print_only: bool) -> tuple[int, int, int, int]:
    if print_only and worksheet.print_area:
        match = re.search(r'(\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)', str(worksheet.print_area))
        if match:
            return range_boundaries(match.group(1).replace('$', ''))
    return 1, 1, worksheet.max_column, worksheet.max_row


def _worksheet_images(worksheet, min_column: int, min_row: int, max_column: int, max_row: int) -> list[dict[str, Any]]:
    images = []
    for image in getattr(worksheet, '_images', []) or []:
        anchor = getattr(image, 'anchor', None)
        marker = getattr(anchor, '_from', None)
        if marker is None:
            continue
        row = int(marker.row) + 1
        column = int(marker.col) + 1
        if not (min_row <= row <= max_row and min_column <= column <= max_column):
            continue
        try:
            raw = image._data()
        except Exception:
            continue
        image_format = str(getattr(image, 'format', '') or 'png').lower()
        mime = 'image/jpeg' if image_format in {'jpg', 'jpeg'} else f'image/{image_format}'
        images.append({
            'row': row - min_row + 1,
            'column': column - min_column + 1,
            'width': float(getattr(image, 'width', 0) or 0),
            'height': float(getattr(image, 'height', 0) or 0),
            'data_url': f'data:{mime};base64,{base64.b64encode(raw).decode("ascii")}',
        })
    return images


def serialize_workbook_preview(
    data: bytes,
    *,
    max_rows: int = 150,
    max_columns: int = 40,
    print_only: bool = False,
) -> dict[str, Any]:
    """Return workbook presentation data, optionally cropped to its defined print area."""
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    sheets = []
    worksheets = [workbook.active] if print_only else workbook.worksheets
    for worksheet in worksheets:
        min_column, min_row, area_max_column, area_max_row = _print_bounds(worksheet, print_only=print_only)
        row_limit = min(area_max_row, min_row + max_rows - 1)
        column_limit = min(area_max_column, min_column + max_columns - 1)
        merged = []
        covered = set()
        merge_starts = {}
        for cell_range in worksheet.merged_cells.ranges:
            if (cell_range.min_row > row_limit or cell_range.min_col > column_limit
                    or cell_range.max_row < min_row or cell_range.max_col < min_column):
                continue
            start_row = max(cell_range.min_row, min_row)
            start_column = max(cell_range.min_col, min_column)
            row_span = min(cell_range.max_row, row_limit) - start_row + 1
            col_span = min(cell_range.max_col, column_limit) - start_column + 1
            merge_starts[(start_row, start_column)] = (row_span, col_span)
            merged.append(str(cell_range))
            for row in range(start_row, min(cell_range.max_row, row_limit) + 1):
                for column in range(start_column, min(cell_range.max_col, column_limit) + 1):
                    if (row, column) != (start_row, start_column):
                        covered.add((row, column))

        rows = []
        for row_number in range(min_row, row_limit + 1):
            cells = []
            for column_number in range(min_column, column_limit + 1):
                if (row_number, column_number) in covered:
                    continue
                cell = worksheet.cell(row=row_number, column=column_number)
                row_span, col_span = merge_starts.get((row_number, column_number), (1, 1))
                cells.append({
                    'column': column_number - min_column + 1,
                    'value': _value(cell.value),
                    'row_span': row_span,
                    'col_span': col_span,
                    'style': {
                        'background': _color(cell.fill.fgColor) if cell.fill.fill_type else '',
                        'color': _color(cell.font.color),
                        'bold': bool(cell.font.bold),
                        'italic': bool(cell.font.italic),
                        'font_size': float(cell.font.sz or 11),
                        'horizontal': cell.alignment.horizontal or '',
                        'vertical': cell.alignment.vertical or '',
                        'wrap': bool(cell.alignment.wrap_text),
                        'border_top': cell.border.top.style or '',
                        'border_right': cell.border.right.style or '',
                        'border_bottom': cell.border.bottom.style or '',
                        'border_left': cell.border.left.style or '',
                    },
                })
            rows.append({
                'number': row_number - min_row + 1,
                'height': float(worksheet.row_dimensions[row_number].height or 18),
                'hidden': bool(worksheet.row_dimensions[row_number].hidden),
                'cells': cells,
            })

        columns = []
        for column_number in range(min_column, column_limit + 1):
            letter = get_column_letter(column_number)
            dimension = worksheet.column_dimensions[letter]
            columns.append({
                'number': column_number - min_column + 1,
                'letter': letter,
                'width': float(dimension.width or 13),
                'hidden': bool(dimension.hidden),
            })
        sheets.append({
            'name': worksheet.title,
            'rows': rows,
            'columns': columns,
            'merged_ranges': merged,
            'images': _worksheet_images(worksheet, min_column, min_row, column_limit, row_limit),
            'print_area': str(worksheet.print_area or ''),
            'truncated': area_max_row > row_limit or area_max_column > column_limit,
        })
    return {'sheets': sheets, 'active_sheet': 0 if print_only else workbook.index(workbook.active), 'print_only': print_only}
