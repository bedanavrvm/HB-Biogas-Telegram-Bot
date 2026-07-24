"""Reject reusable workbook templates that contain customer/business data."""
from __future__ import annotations

import io
import re

import openpyxl
from pathlib import Path


class UnsafeTemplateError(ValueError):
    pass


def template_source_bytes(source) -> bytes:
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    source.seek(0)
    data = source.read()
    source.seek(0)
    return data


def _literal(value) -> bool:
    if value is None or value == '':
        return False
    if isinstance(value, str) and value.startswith('='):
        return False
    return True


def requisition_worksheet(workbook):
    """Return the worksheet containing the requisition customer table."""
    for sheet in workbook.worksheets:
        for row in range(1, min(sheet.max_row, 40) + 1):
            for column in range(1, min(sheet.max_column, 20) + 1):
                value = str(sheet.cell(row=row, column=column).value or '').upper()
                if 'NAME OF THE CUSTOMER' in value:
                    return sheet
    raise UnsafeTemplateError(
        'Requisition template must contain a worksheet with a "Name of the Customer" column.'
    )


def validate_template_bytes(data: bytes, kind: str) -> None:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    except Exception as exc:
        raise UnsafeTemplateError(f'Upload a valid .xlsx workbook: {exc}') from exc
    findings = []
    if kind == 'payment':
        from core.services.payment_documents import payment_template_layout
        layout = payment_template_layout(workbook)
        sheet = workbook[layout.sheet_name]
        business_columns = set(layout.columns.values()) - {layout.columns.get('no')}
        for row in range(layout.data_start_row, layout.totals_row):
            for column in business_columns:
                cell = sheet.cell(row=row, column=column)
                if _literal(cell.value):
                    findings.append(f'{sheet.title}!{cell.coordinate}')
    elif kind == 'requisition':
        sheet = requisition_worksheet(workbook)
        header_row = None
        for row in range(1, min(sheet.max_row, 40) + 1):
            for column in range(1, min(sheet.max_column, 20) + 1):
                value = str(sheet.cell(row=row, column=column).value or '').upper()
                if 'NAME OF THE CUSTOMER' in value:
                    header_row = row
            if header_row:
                break
        if header_row:
            number_col = next((
                column
                for row in (header_row, header_row + 1)
                for column in range(1, min(sheet.max_column, 20) + 1)
                if re.fullmatch(r'NO\.?', str(sheet.cell(row=row, column=column).value or '').upper().strip())
            ), 1)
            numbered_rows = []
            for row in range(header_row + 2, min(sheet.max_row, header_row + 30) + 1):
                value = sheet.cell(row=row, column=number_col).value
                try:
                    int(str(value).strip())
                except (TypeError, ValueError):
                    if numbered_rows:
                        break
                    continue
                numbered_rows.append(row)
            for row in numbered_rows:
                for column in range(1, min(sheet.max_column, 20) + 1):
                    if column == number_col:
                        continue
                    cell = sheet.cell(row=row, column=column)
                    if _literal(cell.value):
                        findings.append(f'{sheet.title}!{cell.coordinate}')
    else:
        raise UnsafeTemplateError(f'Unsupported template type: {kind}.')
    if findings:
        shown = ', '.join(findings[:20])
        suffix = ' and more' if len(findings) > 20 else ''
        raise UnsafeTemplateError(f'Template contains reusable-row data in {shown}{suffix}. Clear these cells before upload.')
