"""FCA Excel batch import workflow."""
from __future__ import annotations

import base64
import io
import json
import logging
import re
from urllib.parse import urlencode
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from django.conf import settings
from django.core import signing
from django.utils import timezone

from core.models import FcaImportRecord
from core.services.order_approval import normalize_kenyan_phone, order_record_id_prefix
from core.services.sheets import get_sheets_service

try:
    import openpyxl
except Exception:  # pragma: no cover - handled at runtime on Render if missing
    openpyxl = None


logger = logging.getLogger(__name__)

FCA_BATCH_COMMAND = '/batchfca'
FCAUP_COMMAND = '/fcaup'
DEFAULT_HEADER_ROW = 2
DEFAULT_FCA_SHEET_NAME = 'Orders'
DEFAULT_MASTER_SHEET_NAME = 'Master Data'
FCAUP_TOKEN_SALT = 'fca-section-a-upload'

FCAUP_STATUS_VALUES = (
    'Approved',
    'Awaiting Analysis',
    'JBL to Schedule Visit',
    'Rescheduled',
    'Deferred / On Hold',
    'Rejected by JBL',
    'Opted for Cash',
    'Opted for other Partner',
)
FCAUP_STATUS_LOOKUP = {
    re.sub(r'\s+', ' ', status).casefold(): status
    for status in FCAUP_STATUS_VALUES
}

FCA_FIELD_HEADERS = {
    'order_record_id': 'ORDER RECORD ID',
    'customer_name': 'CUSTOMER NAME',
    'primary_phone': 'CONTACTS / PRIMARY',
    'secondary_phone': 'CONTACTS / SECONDARY',
    'id_number': 'ID NUMBER',
    'branch': 'BRANCH',
    'county': 'COUNTY',
    'landmark': 'LOCATION AND NEAREST LANDMARK',
    'deposit_hb': 'DEPOSIT / HB',
    'fca_visit_date': 'FCA VISIT DATE',
    'fca_comment': 'FCA COMMENT',
    'fca_decision': 'FCA DECISION',
    'fca_source_file': 'FCA SOURCE FILE',
    'fca_source_sheet': 'FCA SOURCE SHEET',
    'fca_source_row': 'FCA SOURCE ROW',
    'fca_import_status': 'FCA IMPORT STATUS',
}

REQUIRED_FCA_HEADERS = (
    'customer_name',
    'fca_visit_date',
    'fca_comment',
    'fca_decision',
    'fca_import_status',
)

MONTHS = {
    month.lower(): index
    for index, month in enumerate(
        [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December',
        ],
        start=1,
    )
}
MONTHS.update({name[:3].lower(): value for name, value in list(MONTHS.items())})

PHONE_CANDIDATE_PATTERN = re.compile(
    r'(?<!\d)(?:\+?254[\s\-]?(?:7|1)\d{2}[\s\-]?\d{3}[\s\-]?\d{3}'
    r'|0(?:7|1)\d{2}[\s\-]?\d{3}[\s\-]?\d{3}'
    r'|(?:7|1)\d{2}[\s\-]?\d{3}[\s\-]?\d{3})(?!\d)'
)


@dataclass
class FcaParsedRecord:
    source_filename: str
    source_sheet: str
    source_row: int
    fields: dict[str, str]
    warnings: list[str]

    @property
    def status(self) -> str:
        import_status = str(self.fields.get('fca_import_status') or '').casefold()
        if 'review' in import_status:
            return 'review_needed'
        if 'fail' in import_status:
            return 'failed'
        if self.fields.get('fca_decision'):
            return 'pending'
        if not self.fields.get('fca_comment'):
            return 'review_needed'
        return 'pending'


def process_fcaup_files(
    group_config,
    files: list[tuple[str, bytes]],
    telegram_message_id: str,
    sender: str = '',
) -> dict[str, Any]:
    """Parse the agreed Section A FCA template and create a review batch."""
    if openpyxl is None:
        return {
            'status': 'fcaup_processed',
            'files': len(files),
            'processed': 0,
            'updated': 0,
            'created': 0,
            'review_needed': 0,
            'failed': 0,
            'errors': ['openpyxl is not installed. Add it to requirements and redeploy.'],
        }

    parsed_records: list[FcaParsedRecord] = []
    file_errors: list[str] = []
    for filename, content in files:
        try:
            parsed_records.extend(extract_fcaup_section_a_records(filename, content))
        except Exception as exc:
            logger.error("Failed to parse FCA update workbook %s: %s", filename, exc, exc_info=True)
            file_errors.append(f"{filename}: {exc}")

    if not files:
        return {
            'status': 'command',
            'reply_text': (
                "Attach the agreed FCA Section A .xlsx workbook or a .zip of workbooks and send:\n"
                "@bot /fcaup"
            ),
        }

    if not parsed_records:
        return {
            'status': 'fcaup_processed',
            'files': len(files),
            'processed': 0,
            'updated': 0,
            'created': 0,
            'review_needed': 0,
            'failed': 0,
            'errors': file_errors or ['No Section A FCA customer rows were found.'],
        }

    batch_id = str(telegram_message_id or timezone.now().timestamp())
    records = [
        save_fca_record(group_config, parsed, batch_id, sender, configured_master_sheet_name(group_config))
        for parsed in parsed_records
    ]
    review_url = build_fcaup_review_url(batch_id)
    mini_app_url = build_fcaup_mini_app_url(batch_id)
    launch_url = mini_app_url or review_url
    if not review_url:
        return {
            'status': 'fcaup_processed',
            'files': len(files),
            'processed': len(records),
            'updated': 0,
            'created': 0,
            'review_needed': len(records),
            'failed': 0,
            'errors': file_errors + ['APP_BASE_URL is not configured, so the FCA review form cannot open.'],
        }

    return {
        'status': 'fcaup_review_ready',
        'files': len(files),
        'processed': len(records),
        'review_needed': sum(1 for record in records if record.import_status == 'review_needed'),
        'batch_id': batch_id,
        'review_url': review_url,
        'mini_app_url': mini_app_url,
        'launch_url': launch_url,
        'errors': file_errors,
        'reply_markup': {
            'inline_keyboard': [[
                {'text': 'Open FCA Review', 'url': launch_url}
            ]]
        },
    }


def create_fcaup_review_token(batch_id: str) -> str:
    return signing.dumps({'batch_id': str(batch_id)}, salt=FCAUP_TOKEN_SALT)


def validate_fcaup_review_token(batch_id: str, token: str) -> tuple[bool, str]:
    try:
        payload = signing.loads(token or '', salt=FCAUP_TOKEN_SALT, max_age=7 * 24 * 3600)
    except signing.BadSignature:
        return False, 'This FCA review link is invalid or expired.'
    if str(payload.get('batch_id', '')) != str(batch_id):
        return False, 'This FCA review link does not match the upload batch.'
    return True, ''


def build_fcaup_review_url(batch_id: str) -> str:
    base_url = getattr(settings, 'APP_BASE_URL', '').rstrip('/')
    if not base_url:
        return ''
    return f"{base_url}/fca/review/?" + urlencode({
        'batch_id': str(batch_id),
        'token': create_fcaup_review_token(batch_id),
    })


def build_fcaup_mini_app_url(batch_id: str) -> str:
    bot_username = str(getattr(settings, 'TELEGRAM_BOT_USERNAME', '') or '').strip().lstrip('@')
    short_name = str(getattr(settings, 'FCAUP_MINI_APP_SHORT_NAME', '') or '').strip().strip('/')
    if not bot_username or not short_name:
        return ''
    return f"https://t.me/{bot_username}/{short_name}?startapp={create_fcaup_start_param(batch_id)}"


def create_fcaup_start_param(batch_id: str) -> str:
    payload = {
        'batch_id': str(batch_id),
        'token': create_fcaup_review_token(batch_id),
    }
    encoded = base64.urlsafe_b64encode(
        json.dumps(payload, separators=(',', ':')).encode('utf-8')
    ).decode('ascii')
    return encoded.rstrip('=')


def decode_fcaup_start_param(start_param: str) -> dict[str, str]:
    value = str(start_param or '').strip()
    if not value:
        return {}
    padding = '=' * (-len(value) % 4)
    try:
        decoded = base64.urlsafe_b64decode((value + padding).encode('ascii'))
        payload = json.loads(decoded.decode('utf-8'))
    except Exception:
        return {}
    if not isinstance(payload, dict):
        return {}
    return {
        'batch_id': str(payload.get('batch_id') or ''),
        'token': str(payload.get('token') or ''),
    }

def fcaup_review_payload(batch_id: str, token: str) -> dict[str, Any]:
    records = list(
        FcaImportRecord.objects
        .filter(telegram_message_id=str(batch_id))
        .order_by('source_filename', 'source_sheet', 'source_row', 'created_at')
    )
    rows = [fcaup_record_to_review_row(record) for record in records]
    return {
        'batch_id': str(batch_id),
        'token': token,
        'rows': rows,
        'status_values': list(FCAUP_STATUS_VALUES),
    }


def fcaup_record_to_review_row(record: FcaImportRecord) -> dict[str, Any]:
    fields = dict(record.parsed_fields or {})
    review = record.import_status == 'review_needed'
    return {
        'record_id': str(record.id),
        'approved': not review,
        'Customer Name': fields.get('customer_name') or record.customer_name or '',
        'ID Number': fields.get('id_number') or '',
        'Primary Phone': fields.get('primary_phone') or record.primary_phone or '',
        'Secondary Phone': fields.get('secondary_phone') or '',
        'Location': fields.get('landmark') or '',
        'Hub': fields.get('branch') or '',
        'Field Officer': fields.get('jbl_officer') or '',
        'HB Staff': fields.get('hb_staff') or '',
        'Deposit': fields.get('deposit_hb') or '',
        'Jawabu Visit Date': fields.get('fca_visit_date') or format_master_date(record.fca_visit_date),
        'JBL Officer': fields.get('jbl_officer') or '',
        'Status': fields.get('fca_decision') or record.fca_decision or '',
        'Comment': fields.get('fca_comment') or record.fca_comment or '',
        'Import Status': 'review_needed' if review else 'pending',
        'Review Notes': record.sync_error or '',
        'Source': f"{record.source_filename} / {record.source_sheet} row {record.source_row}",
    }


def commit_fcaup_review_batch(batch_id: str, rows: list[dict], group_config=None, sender: str = '') -> dict[str, Any]:
    record_ids = [str(row.get('record_id') or '') for row in rows if row.get('record_id')]
    records_by_id = {
        str(record.id): record
        for record in FcaImportRecord.objects.filter(telegram_message_id=str(batch_id), id__in=record_ids)
    }
    approved_records = []
    errors = []
    for index, row in enumerate(rows, start=1):
        record = records_by_id.get(str(row.get('record_id') or ''))
        if not record:
            continue
        if not row.get('approved'):
            record.delete()
            continue
        fields, row_errors = cleaned_fcaup_review_fields(row)
        if row_errors:
            record.import_status = 'review_needed'
            record.sync_error = '; '.join(row_errors)
            record.parsed_fields = {**dict(record.parsed_fields or {}), **fields, 'fca_import_status': 'Review Needed'}
            record.save(update_fields=['import_status', 'sync_error', 'parsed_fields'])
            errors.extend(f"Row {index}: {error}" for error in row_errors)
            continue
        record.customer_name = fields['customer_name']
        record.primary_phone = fields['primary_phone']
        record.fca_visit_date = parse_iso_date(fields.get('fca_visit_date'))
        record.fca_decision = fields['fca_decision']
        record.fca_comment = fields['fca_comment']
        record.parsed_fields = {**dict(record.parsed_fields or {}), **fields, 'fca_import_status': 'Pending MD Sync'}
        record.import_status = 'pending'
        record.sync_error = ''
        record.save(update_fields=[
            'customer_name', 'primary_phone', 'fca_visit_date', 'fca_decision',
            'fca_comment', 'parsed_fields', 'import_status', 'sync_error',
        ])
        approved_records.append(record)

    sync_result = sync_fcaup_records_to_master_data(group_config, approved_records) if group_config else {
        'created': 0, 'updated': 0, 'duplicates': 0, 'errors': ['Group configuration was not found.'], 'sheet_tab': '',
    }
    remaining_records = list(
        FcaImportRecord.objects
        .filter(telegram_message_id=str(batch_id))
        .exclude(import_status='imported')
        .order_by('source_filename', 'source_sheet', 'source_row', 'created_at')
    )
    return {
        'success': not errors and not sync_result.get('errors'),
        'message': 'FCA rows committed.' if not errors and not sync_result.get('errors') else 'Some FCA rows still need review or sheet sync failed.',
        'committed': sum(1 for record in approved_records if record.import_status == 'imported'),
        'review_needed': len(remaining_records),
        'errors': (errors + (sync_result.get('errors') or []))[:20],
        'sheet_sync': sync_result,
        'rows': [fcaup_record_to_review_row(record) for record in remaining_records],
    }


def cleaned_fcaup_review_fields(row: dict) -> tuple[dict[str, str], list[str]]:
    customer_name = normalize_name(row.get('Customer Name') or '')
    id_number = normalize_identifier(row.get('ID Number') or '')
    primary_phone = first_normalized_phone(row.get('Primary Phone') or '')
    secondary_phone = first_normalized_phone(row.get('Secondary Phone') or '')
    visit_date = clean_fcaup_review_date(row.get('Jawabu Visit Date') or '')
    jbl_officer = normalize_name(row.get('JBL Officer') or '')
    status = canonical_fcaup_status(row.get('Status') or '')
    comment = normalize_cell(row.get('Comment') or '')
    errors = []
    if not customer_name:
        errors.append('Customer Name is required')
    if not id_number and not primary_phone:
        errors.append('ID Number or Primary Phone is required for matching')
    if not visit_date:
        errors.append('Jawabu Visit Date is required')
    if not status:
        errors.append('Status must be one of the agreed dropdown values')
    return {
        'customer_name': customer_name,
        'id_number': id_number,
        'primary_phone': primary_phone,
        'secondary_phone': secondary_phone,
        'landmark': normalize_name(row.get('Location') or ''),
        'hb_staff': normalize_name(row.get('HB Staff') or ''),
        'deposit_hb': normalize_cell(row.get('Deposit') or ''),
        'fca_visit_date': visit_date,
        'jbl_officer': jbl_officer,
        'fca_decision': status,
        'fca_comment': comment,
    }, errors


def clean_fcaup_review_date(value: str) -> str:
    text = normalize_cell(value)
    if not text:
        return ''
    parsed = parse_fca_date_text(text)
    return format_master_date(parsed) if parsed else text


def append_error(existing: str, error: str) -> str:
    existing = normalize_cell(existing)
    if not existing:
        return error
    if error in existing:
        return existing
    return f"{existing}; {error}"


def extract_fcaup_section_a_records(filename: str, content: bytes) -> list[FcaParsedRecord]:
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    records: list[FcaParsedRecord] = []
    for worksheet in workbook.worksheets:
        rows = worksheet_rows(worksheet)
        header_row, headers = detect_fcaup_section_a_header(rows)
        if not header_row:
            continue
        column_map = fcaup_column_map(headers)
        visit_date = extract_visit_date(rows, filename)
        hub = extract_hub(rows)
        officer = extract_officer(rows)
        for row_number, values in rows:
            if row_number <= header_row:
                continue
            if looks_like_footer(values) or looks_like_next_fcaup_section(values):
                break
            if not any(values):
                continue
            parsed = parse_fcaup_section_a_row(
                filename=filename,
                sheet_name=worksheet.title,
                row_number=row_number,
                values=values,
                columns=column_map,
                visit_date=visit_date,
                hub=hub,
                officer=officer,
            )
            if parsed:
                records.append(parsed)
    return records


def process_fca_batch_files(
    group_config,
    files: list[tuple[str, bytes]],
    telegram_message_id: str,
    sender: str = '',
) -> dict[str, Any]:
    """Parse FCA Excel files and append extracted rows to the configured sheet."""
    if openpyxl is None:
        return {
            'status': 'fca_batch_processed',
            'files': len(files),
            'processed': 0,
            'imported': 0,
            'review_needed': 0,
            'failed': 0,
            'errors': ['openpyxl is not installed. Add it to requirements and redeploy.'],
        }

    parsed_records: list[FcaParsedRecord] = []
    file_errors: list[str] = []
    for filename, content in files:
        try:
            parsed_records.extend(extract_fca_workbook_records(filename, content))
        except Exception as exc:
            logger.error("Failed to parse FCA workbook %s: %s", filename, exc, exc_info=True)
            file_errors.append(f"{filename}: {exc}")

    if not files:
        return {
            'status': 'command',
            'reply_text': (
                "Attach one .xlsx file or a .zip containing FCA Excel files and send:\n"
                "@bot /batchfca"
            ),
        }

    if not parsed_records:
        return {
            'status': 'fca_batch_processed',
            'files': len(files),
            'processed': 0,
            'imported': 0,
            'review_needed': 0,
            'failed': 0,
            'errors': file_errors or ['No FCA customer rows were found in the attached workbook(s).'],
        }

    target_sheet = configured_fca_sheet_name(group_config)
    records = [
        save_fca_record(group_config, parsed, telegram_message_id, sender, target_sheet)
        for parsed in parsed_records
    ]

    duplicate_source_count = mark_duplicate_source_rows(records)
    sync_result = append_fca_records_to_sheet(group_config, records, target_sheet)
    if sync_result.get('success'):
        row_numbers = sync_result.get('row_numbers') or []
        for index, record in enumerate(records):
            record.row_number = row_numbers[index] if index < len(row_numbers) else None
            update_fields = ['row_number']
            if record.import_status == 'pending':
                record.import_status = 'imported'
                record.sync_error = ''
                update_fields.extend(['import_status', 'sync_error'])
            record.save(update_fields=update_fields)
    else:
        error = sync_result.get('error') or 'Google Sheets append failed'
        for record in records:
            record.import_status = 'failed'
            record.sync_error = error
            record.save(update_fields=['import_status', 'sync_error'])

    records = list(FcaImportRecord.objects.filter(id__in=[record.id for record in records]))
    return {
        'status': 'fca_batch_processed',
        'files': len(files),
        'processed': len(records),
        'imported': sum(1 for record in records if record.import_status == 'imported'),
        'review_needed': sum(1 for record in records if record.import_status == 'review_needed'),
        'failed': sum(1 for record in records if record.import_status == 'failed'),
        'cash': sum(1 for record in records if record.fca_decision == 'Cash'),
        'approved': sum(1 for record in records if record.fca_decision == 'Approved'),
        'rejected': sum(1 for record in records if record.fca_decision == 'Rejected'),
        'deferred': sum(1 for record in records if record.fca_decision == 'Deferred'),
        'duplicate_source_rows': duplicate_source_count,
        'sheet_tab': target_sheet,
        'errors': file_errors + ([sync_result.get('error')] if sync_result.get('error') else []),
        'review_examples': fca_review_examples(records),
    }


def extract_fca_workbook_records(filename: str, content: bytes) -> list[FcaParsedRecord]:
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    records: list[FcaParsedRecord] = []
    for worksheet in workbook.worksheets:
        if worksheet.title.strip().lower() in {'budget', 'budget '}:
            continue
        rows = worksheet_rows(worksheet)
        header_row, headers = detect_table_header(rows)
        if not header_row:
            continue

        column_map = fca_column_map(headers, worksheet.max_column)
        visit_date = extract_visit_date(rows, filename)
        hub = extract_hub(rows)
        for row_number, values in rows:
            if row_number <= header_row or not any(values):
                continue
            if looks_like_footer(values):
                break
            if looks_like_section(values):
                continue
            parsed = parse_fca_data_row(
                filename=filename,
                sheet_name=worksheet.title,
                row_number=row_number,
                values=values,
                columns=column_map,
                visit_date=visit_date,
                hub=hub,
            )
            if parsed:
                records.append(parsed)
    return records


def worksheet_rows(worksheet) -> list[tuple[int, list[str]]]:
    max_row = min(int(worksheet.max_row or 0), 200)
    max_col = min(int(worksheet.max_column or 0), 16)
    rows: list[tuple[int, list[str]]] = []
    for index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True),
        start=1,
    ):
        rows.append((index, [normalize_cell(value) for value in row]))
    return rows


def detect_table_header(rows: list[tuple[int, list[str]]]) -> tuple[int | None, list[str]]:
    for row_number, values in rows[:15]:
        lower = [value.lower() for value in values]
        if any(value == 'customer name' for value in lower) and any(value == 'contacts' for value in lower):
            return row_number, values
        if 'names' in lower and 'comment' in lower:
            return row_number, values
    return None, []


def fca_column_map(headers: list[str], max_column: int) -> dict[str, int | None]:
    lower = [header.lower() for header in headers]

    def find(*needles: str) -> int | None:
        for needle in needles:
            for index, value in enumerate(lower, start=1):
                if value == needle or needle in value:
                    return index
        return None

    basis_index = find('approval basis')
    comment_index = find('comment')
    if not comment_index and basis_index and basis_index < max_column:
        comment_index = basis_index + 1

    return {
        'customer_name': find('customer name', 'names') or 2,
        'primary_phone': find('contacts', 'mobile no') or 3,
        'id_number': find('id no', 'id number', 'national id'),
        'landmark': find('location') or 4,
        'staff': find('hb staff', 'sales person') or 5,
        'deposit_hb': find('deposit', 'amount paid') or 6,
        'approval_basis': basis_index,
        'fca_comment': comment_index,
    }


def detect_fcaup_section_a_header(rows: list[tuple[int, list[str]]]) -> tuple[int | None, list[str]]:
    for row_number, values in rows[:25]:
        lower = [value.lower().strip() for value in values]
        if (
            any(value == 'customer name' for value in lower)
            and any('id number' in value or value == 'id no' for value in lower)
            and any('status' == value.strip() for value in lower)
            and any('comment' == value.strip() for value in lower)
        ):
            return row_number, values
    return None, []


def fcaup_column_map(headers: list[str]) -> dict[str, int | None]:
    lower = [header.lower().strip() for header in headers]

    def find(*needles: str) -> int | None:
        for needle in needles:
            for index, value in enumerate(lower, start=1):
                if value == needle or needle in value:
                    return index
        return None

    return {
        'customer_name': find('customer name'),
        'id_number': find('id number', 'id no', 'national id'),
        'primary_phone': find('phone', 'contacts', 'mobile', 'tel'),
        'landmark': find('location'),
        'staff': find('hb staff', 'field officer', 'bro'),
        'deposit_hb': find('deposit'),
        'status': find('status'),
        'fca_comment': find('comment'),
    }


def parse_fcaup_section_a_row(
    filename: str,
    sheet_name: str,
    row_number: int,
    values: list[str],
    columns: dict[str, int | None],
    visit_date: date | None,
    hub: str,
    officer: str = '',
) -> FcaParsedRecord | None:
    def value(field: str) -> str:
        column = columns.get(field)
        if not column or column < 1 or column > len(values):
            return ''
        return values[column - 1]

    customer_name = normalize_name(value('customer_name'))
    if not customer_name or customer_name.lower() in {'customer name', 'names'}:
        return None

    phones = extract_phone_numbers(value('primary_phone'))
    id_number = normalize_identifier(value('id_number'))
    status = canonical_fcaup_status(value('status'))
    comment = normalize_cell(value('fca_comment'))
    warnings = []
    if not id_number and not phones:
        warnings.append('Missing ID NUMBER and PHONE; cannot match Master Data safely')
    if not status:
        warnings.append('Missing or invalid FCA STATUS dropdown value')

    fields = {
        'customer_name': customer_name,
        'primary_phone': phones[0] if phones else '',
        'secondary_phone': phones[1] if len(phones) > 1 else '',
        'id_number': id_number,
        'branch': normalize_name(hub),
        'county': normalize_name(hub),
        'landmark': normalize_name(value('landmark')),
        'hb_staff': normalize_name(value('staff')),
        'deposit_hb': normalize_cell(value('deposit_hb')),
        'fca_visit_date': format_master_date(visit_date),
        'jbl_officer': normalize_name(officer),
        'fca_comment': comment,
        'fca_decision': status,
        'fca_source_file': filename,
        'fca_source_sheet': sheet_name,
        'fca_source_row': str(row_number),
        'fca_import_status': 'Review Needed' if warnings else 'Pending MD Sync',
    }
    return FcaParsedRecord(
        source_filename=filename,
        source_sheet=sheet_name,
        source_row=row_number,
        fields=fields,
        warnings=warnings,
    )


def canonical_fcaup_status(value: str) -> str:
    text = re.sub(r'\s+', ' ', str(value or '').strip())
    if not text:
        return ''
    return FCAUP_STATUS_LOOKUP.get(text.casefold(), '')


def looks_like_next_fcaup_section(values: list[str]) -> bool:
    text = ' '.join(value.lower() for value in values if value)
    return bool(re.search(r'\bsection\s+b\b|\bcollections?\b|\badmin\b', text))


def parse_fca_data_row(
    filename: str,
    sheet_name: str,
    row_number: int,
    values: list[str],
    columns: dict[str, int | None],
    visit_date: date | None,
    hub: str,
) -> FcaParsedRecord | None:
    def value(field: str) -> str:
        column = columns.get(field)
        if not column or column < 1 or column > len(values):
            return ''
        return values[column - 1]

    customer_name = normalize_name(value('customer_name'))
    if not customer_name or customer_name.lower() in {'customer name', 'names'}:
        return None

    phones = extract_phone_numbers(value('primary_phone'))
    comment = value('fca_comment')
    if not comment and value('approval_basis') and looks_like_decision_comment(value('approval_basis')):
        comment = value('approval_basis')
    decision = derive_fca_decision(comment)
    warnings = []
    if not comment:
        warnings.append('Missing FCA comment')
    if comment and not decision:
        warnings.append('Decision could not be inferred from FCA comment')

    fields = {
        'customer_name': customer_name,
        'primary_phone': phones[0] if phones else '',
        'secondary_phone': phones[1] if len(phones) > 1 else '',
        'id_number': normalize_identifier(value('id_number')),
        'branch': normalize_name(hub),
        'county': normalize_name(hub),
        'landmark': normalize_name(value('landmark')),
        'deposit_hb': value('deposit_hb'),
        'fca_visit_date': format_sheet_date(visit_date),
        'fca_comment': comment,
        'fca_decision': decision,
        'fca_source_file': filename,
        'fca_source_sheet': sheet_name,
        'fca_source_row': str(row_number),
        'fca_import_status': 'Review Needed' if warnings else 'Imported',
    }
    return FcaParsedRecord(
        source_filename=filename,
        source_sheet=sheet_name,
        source_row=row_number,
        fields=fields,
        warnings=warnings,
    )


def save_fca_record(group_config, parsed: FcaParsedRecord, telegram_message_id: str, sender: str, sheet_tab: str) -> FcaImportRecord:
    fields = parsed.fields
    return FcaImportRecord.objects.create(
        group_id=str(group_config.group_id),
        sheet_id=str(group_config.sheet_id or ''),
        sheet_tab=sheet_tab,
        telegram_message_id=str(telegram_message_id or ''),
        source_filename=parsed.source_filename,
        source_sheet=parsed.source_sheet,
        source_row=parsed.source_row,
        sender=str(sender or ''),
        customer_name=fields.get('customer_name', ''),
        primary_phone=fields.get('primary_phone', ''),
        fca_visit_date=parse_iso_date(fields.get('fca_visit_date')),
        fca_comment=fields.get('fca_comment', ''),
        fca_decision=fields.get('fca_decision', ''),
        import_status=parsed.status,
        parsed_fields=fields,
        sync_error='; '.join(parsed.warnings),
    )


def append_fca_records_to_sheet(group_config, records: list[FcaImportRecord], sheet_name: str) -> dict[str, Any]:
    if not records:
        return {'success': True, 'row_numbers': []}

    workflow = getattr(group_config, 'workflow', None) or {}
    service = get_sheets_service(
        sheet_id=group_config.sheet_id,
        sheet_name=sheet_name,
        sheet_schema=None,
    )
    if not service.is_available():
        return {'success': False, 'error': 'Google Sheets service unavailable.'}

    try:
        header_row_number = configured_header_row(workflow)
        headers = [str(value or '').strip() for value in service._sheet.row_values(header_row_number)]
        values = service._sheet.get_all_values() if hasattr(service._sheet, 'get_all_values') else [headers]
    except Exception as exc:
        logger.error("Failed to read FCA target sheet headers: %s", exc, exc_info=True)
        return {'success': False, 'error': str(exc)}

    if not headers:
        return {'success': False, 'error': 'Header row is empty or unavailable.'}

    field_headers = configured_field_headers(workflow)
    missing_headers = [
        field_headers[field]
        for field in REQUIRED_FCA_HEADERS
        if field_headers.get(field) not in headers
    ]
    if missing_headers:
        return {
            'success': False,
            'error': 'Missing FCA import column(s): ' + ', '.join(missing_headers),
        }

    next_record_ids = next_order_record_ids(headers, values, workflow, len(records))
    rows = []
    for index, record in enumerate(records):
        fields = dict(record.parsed_fields or {})
        fields['order_record_id'] = next_record_ids[index] if index < len(next_record_ids) else ''
        fields['fca_import_status'] = sheet_import_status(record)
        row_values = ['' for _ in headers]
        for field, header in field_headers.items():
            if header in headers:
                row_values[headers.index(header)] = fields.get(field, '')
        rows.append(row_values)

    try:
        if hasattr(service._sheet, 'append_rows'):
            response = service._sheet.append_rows(rows, value_input_option='USER_ENTERED')
        else:
            responses = [
                service._sheet.append_row(row, value_input_option='USER_ENTERED')
                for row in rows
            ]
            response = responses[0] if responses else {}
        return {
            'success': True,
            'row_numbers': row_numbers_from_append_response(response, len(rows)),
        }
    except Exception as exc:
        logger.error("Failed to append FCA rows: %s", exc, exc_info=True)
        return {'success': False, 'error': str(exc)}


def sync_fcaup_records_to_master_data(group_config, records: list[FcaImportRecord]) -> dict[str, Any]:
    if not records:
        return {'created': 0, 'updated': 0, 'duplicates': 0, 'errors': [], 'sheet_tab': configured_master_sheet_name(group_config)}

    workflow = getattr(group_config, 'workflow', None) or {}
    sheet_id = str(
        workflow.get('fca_master_sheet_id')
        or workflow.get('master_sheet_id')
        or getattr(group_config, 'sheet_id', '')
        or ''
    ).strip()
    sheet_name = configured_master_sheet_name(group_config)
    header_row = configured_master_header_row(workflow)
    data_start_row = configured_master_data_start_row(workflow, header_row)
    if not sheet_id:
        error = 'FCA Master Data sync needs fca_master_sheet_id or master_sheet_id in workflow config.'
        mark_fca_records_failed(records, error)
        return {'created': 0, 'updated': 0, 'duplicates': 0, 'errors': [error], 'sheet_tab': sheet_name}

    try:
        from core.services.sheets import GoogleSheetsService
        from core.services.jawabu_master import (
            batch_update_master_sheet_rows,
            ensure_master_system_headers,
            first_existing_header,
            header_lookup_from_headers,
            next_master_append_row,
            pad_values_to_row,
            row_values_for_number,
            set_header_value,
        )

        service = GoogleSheetsService.get_instance(sheet_id=sheet_id, sheet_name=sheet_name)
        if not service.is_available():
            raise RuntimeError('Google Sheets service unavailable for Master Data sheet.')
        sheet = service._sheet
        headers = ensure_master_system_headers(sheet, header_row)
        header_lookup = header_lookup_from_headers(headers)
        required = [
            'Customer Name', 'National ID', 'Primary Phone',
            'Jawabu Visit Date', 'Jawabu Comment After visit', 'Additional Comments',
        ]
        missing = [header for header in required if not first_existing_header(header_lookup, [header])]
        if missing:
            raise RuntimeError('Missing Master Data column(s): ' + ', '.join(missing))

        values = sheet.get_all_values()
        existing = build_fcaup_master_existing_index(values, header_lookup, data_start_row)
        pending_updates = []
        created = updated = duplicates = 0
        errors = []
        now_text = timezone.now().strftime('%d-%B-%Y %H:%M')

        for record in records:
            fields = dict(record.parsed_fields or {})
            if record.import_status == 'review_needed':
                continue
            match = find_fcaup_master_match(fields, existing)
            if match.get('duplicate'):
                duplicates += 1
                record.import_status = 'review_needed'
                record.sync_error = match['message']
                record.save(update_fields=['import_status', 'sync_error'])
                continue

            row_number = int(match.get('row_number') or 0)
            created_row = False
            if row_number:
                row_values = row_values_for_number(values, row_number, len(headers))
            else:
                row_number = next_master_append_row(values, header_lookup, data_start_row)
                row_values = [''] * len(headers)
                set_header_value(row_values, header_lookup, 'No.', row_number - data_start_row + 1)
                created_row = True

            apply_fcaup_master_values(
                row_values=row_values,
                header_lookup=header_lookup,
                fields=fields,
                source_record=record,
                now_text=now_text,
            )
            pending_updates.append((row_number, row_values))
            values = pad_values_to_row(values, row_number, len(headers))
            values[row_number - 1] = row_values
            # Ensure database consistency by upserting/updating JawabuFarmerMaster model
            from django.db.models import Q
            from core.models import JawabuFarmerMaster

            queries = Q()
            if record.primary_phone:
                queries |= Q(primary_phone=record.primary_phone)
            if fields.get('id_number'):
                queries |= Q(national_id=fields['id_number'])

            farmer = None
            if queries:
                farmer = JawabuFarmerMaster.objects.filter(queries).order_by('-updated_at').first()

            visit_date = record.fca_visit_date

            if farmer:
                farmer.jbl_visit_date = visit_date
                farmer.jbl_visit_status = record.fca_decision
                farmer.jbl_visit_comment = record.fca_comment
                if fields.get('jbl_officer'):
                    farmer.jbl_officer = fields['jbl_officer']
                if not farmer.county and (fields.get('county') or fields.get('branch')):
                    farmer.county = fields.get('county') or fields.get('branch')
                if not farmer.branch and fields.get('branch'):
                    farmer.branch = fields.get('branch')
                farmer.save()
                logger.info("FCA sync: Updated existing database record for farmer %s", farmer.id)
            else:
                JawabuFarmerMaster.objects.create(
                    customer_name=record.customer_name,
                    national_id=fields.get('id_number', ''),
                    primary_phone=record.primary_phone,
                    secondary_phone=fields.get('secondary_phone', ''),
                    county=fields.get('county', '') or fields.get('branch', ''),
                    branch=fields.get('branch', ''),
                    jbl_visit_date=visit_date,
                    jbl_visit_status=record.fca_decision,
                    jbl_visit_comment=record.fca_comment,
                    jbl_officer=fields.get('jbl_officer', ''),
                    sign_date=visit_date,
                    status='active',
                )
            add_fcaup_master_index_row(existing, row_number, row_values, header_lookup)
            record.row_number = row_number
            record.import_status = 'imported'
            record.sync_error = ''
            record.save(update_fields=['row_number', 'import_status', 'sync_error'])
            if created_row:
                created += 1
            else:
                updated += 1

        if pending_updates:
            batch_update_master_sheet_rows(sheet, pending_updates, len(headers))
        return {'created': created, 'updated': updated, 'duplicates': duplicates, 'errors': errors, 'sheet_tab': sheet_name}
    except Exception as exc:
        logger.error('FCA Master Data sync failed: %s', exc, exc_info=True)
        error = f'Master Data sync failed: {exc}'
        for record in records:
            if record.import_status != 'review_needed':
                record.import_status = 'failed'
                record.sync_error = error
                record.save(update_fields=['import_status', 'sync_error'])
        return {'created': 0, 'updated': 0, 'duplicates': 0, 'errors': [error], 'sheet_tab': sheet_name}


def configured_master_sheet_name(group_config) -> str:
    workflow = getattr(group_config, 'workflow', None) or {}
    return str(
        workflow.get('fca_master_sheet_name')
        or workflow.get('master_sheet_name')
        or DEFAULT_MASTER_SHEET_NAME
    ).strip() or DEFAULT_MASTER_SHEET_NAME


def configured_master_header_row(workflow: dict) -> int:
    return configured_positive_int(
        (workflow or {}).get('fca_master_header_row') or (workflow or {}).get('master_header_row'),
        3,
    )


def configured_master_data_start_row(workflow: dict, header_row: int) -> int:
    return configured_positive_int(
        (workflow or {}).get('fca_master_data_start_row') or (workflow or {}).get('master_data_start_row'),
        header_row + 2,
    )


def configured_positive_int(value, default: int) -> int:
    try:
        return max(int(value), 1)
    except (TypeError, ValueError):
        return default


def mark_fca_records_failed(records: list[FcaImportRecord], error: str) -> None:
    for record in records:
        if record.import_status != 'review_needed':
            record.import_status = 'failed'
            record.sync_error = error
            record.save(update_fields=['import_status', 'sync_error'])


def build_fcaup_master_existing_index(values: list[list[str]], header_lookup: dict[str, int], data_start_row: int) -> dict[str, list[int]]:
    existing: dict[str, list[int]] = {}
    for row_number, row_values in enumerate(values[data_start_row - 1:], start=data_start_row):
        add_fcaup_master_index_row(existing, row_number, row_values, header_lookup)
    return existing


def add_fcaup_master_index_row(existing: dict[str, list[int]], row_number: int, row_values: list, header_lookup: dict[str, int]) -> None:
    from core.services.jawabu_master import header_row_value

    national_id = normalize_identifier(header_row_value(row_values, header_lookup, 'National ID'))
    primary_phone = first_normalized_phone(header_row_value(row_values, header_lookup, 'Primary Phone'))
    if national_id and primary_phone:
        existing.setdefault(f'id_phone:{national_id}|{primary_phone}', []).append(row_number)
    if national_id:
        existing.setdefault(f'id:{national_id}', []).append(row_number)
    if primary_phone:
        existing.setdefault(f'phone:{primary_phone}', []).append(row_number)


def find_fcaup_master_match(fields: dict, existing: dict[str, list[int]]) -> dict[str, Any]:
    national_id = fields.get('id_number') or ''
    primary_phone = first_normalized_phone(fields.get('primary_phone') or '')
    keys = [
        f'id_phone:{national_id}|{primary_phone}' if national_id and primary_phone else '',
        f'id:{national_id}' if national_id else '',
        f'phone:{primary_phone}' if primary_phone else '',
    ]
    for key in keys:
        if not key:
            continue
        rows = sorted(set(existing.get(key) or []))
        if len(rows) == 1:
            return {'row_number': rows[0]}
        if len(rows) > 1:
            return {
                'duplicate': True,
                'message': f'Multiple Master Data rows match {key.replace(":", " ")}: rows {", ".join(str(row) for row in rows[:8])}',
            }
    return {'row_number': 0}


def apply_fcaup_master_values(
    *,
    row_values: list,
    header_lookup: dict[str, int],
    fields: dict,
    source_record: FcaImportRecord,
    now_text: str,
) -> None:
    from core.services.jawabu_master import set_header_value

    fill_if_blank(row_values, header_lookup, 'Customer Name', fields.get('customer_name'))
    fill_if_blank(row_values, header_lookup, 'National ID', fields.get('id_number'))
    fill_if_blank(row_values, header_lookup, 'Primary Phone', fields.get('primary_phone'))
    fill_if_blank(row_values, header_lookup, 'Secondary Phone', fields.get('secondary_phone'))
    fill_if_blank(row_values, header_lookup, 'County', fields.get('branch'))
    fill_if_blank(row_values, header_lookup, 'HB Sales Person', fields.get('hb_staff'))
    fill_if_blank(row_values, header_lookup, 'Deposit Paid to HB', fields.get('deposit_hb'))

    set_header_value(row_values, header_lookup, 'Jawabu Visit Date', fields.get('fca_visit_date'))
    set_header_value(row_values, header_lookup, 'JBL BRO', fields.get('jbl_officer'))
    set_header_value(row_values, header_lookup, 'Jawabu Comment After visit', fields.get('fca_decision'))
    set_header_value(row_values, header_lookup, 'Additional Comments', fields.get('fca_comment'))
    set_header_value(row_values, header_lookup, 'Source Filename', source_record.source_filename)
    set_header_value(row_values, header_lookup, 'Source Row', fields.get('fca_source_row'))
    set_header_value(row_values, header_lookup, 'Import Status', 'fca_updated')
    set_header_value(row_values, header_lookup, 'Review Notes', '')
    set_header_value(row_values, header_lookup, 'Reviewed By', source_record.sender)
    set_header_value(row_values, header_lookup, 'Reviewed At', now_text)
    set_header_value(row_values, header_lookup, 'Last Updated At', now_text)


def fill_if_blank(row_values: list, header_lookup: dict[str, int], header: str, value) -> None:
    if value in (None, ''):
        return
    from core.services.jawabu_master import normalize_header

    index = header_lookup.get(normalize_header(header), 0) - 1
    if index < 0:
        return
    if index >= len(row_values):
        row_values.extend([''] * (index - len(row_values) + 1))
    if not str(row_values[index] or '').strip():
        row_values[index] = value


def first_normalized_phone(value: str) -> str:
    phones = extract_phone_numbers(value)
    return phones[0] if phones else ''


def configured_field_headers(workflow: dict) -> dict[str, str]:
    headers = dict(FCA_FIELD_HEADERS)
    for source_key in ('field_headers', 'fca_field_headers'):
        configured = (workflow or {}).get(source_key) or {}
        headers.update({
            str(field): str(header)
            for field, header in configured.items()
            if str(field).strip() and str(header).strip()
        })
    return headers


def configured_header_row(workflow: dict) -> int:
    try:
        return max(int((workflow or {}).get('header_row') or DEFAULT_HEADER_ROW), 1)
    except (TypeError, ValueError):
        return DEFAULT_HEADER_ROW


def configured_fca_sheet_name(group_config) -> str:
    workflow = getattr(group_config, 'workflow', None) or {}
    return str(
        workflow.get('fca_sheet_name')
        or workflow.get('create_sheet_name')
        or getattr(group_config, 'sheet_name', '')
        or DEFAULT_FCA_SHEET_NAME
    ).strip()


def mark_duplicate_source_rows(records: list[FcaImportRecord]) -> int:
    duplicate_count = 0
    for record in records:
        exists = FcaImportRecord.objects.filter(
            group_id=record.group_id,
            source_filename=record.source_filename,
            source_sheet=record.source_sheet,
            source_row=record.source_row,
        ).exclude(id=record.id).exclude(import_status='failed').exists()
        if exists:
            duplicate_count += 1
            record.import_status = 'review_needed'
            note = 'Possible duplicate source row from a previous FCA import.'
            record.sync_error = '; '.join(filter(None, [record.sync_error, note]))
            record.save(update_fields=['import_status', 'sync_error'])
    return duplicate_count


def next_order_record_ids(headers: list[str], values: list[list[str]], workflow: dict, count: int) -> list[str]:
    record_header = configured_field_headers(workflow).get('order_record_id')
    if not record_header or record_header not in headers:
        return ['' for _ in range(count)]
    index = headers.index(record_header)
    prefix = order_record_id_prefix(workflow)
    pattern = re.compile(rf'^{re.escape(prefix)}-(\d+)$', re.IGNORECASE)
    max_number = 0
    for row in values[configured_header_row(workflow):]:
        if index >= len(row):
            continue
        match = pattern.match(str(row[index] or '').strip())
        if match:
            max_number = max(max_number, int(match.group(1)))
    return [f'{prefix}-{max_number + offset}' for offset in range(1, count + 1)]


def row_numbers_from_append_response(response: Any, count: int) -> list[int | None]:
    if count <= 0:
        return []
    updated_range = ''
    if isinstance(response, dict):
        updated_range = str((response.get('updates') or {}).get('updatedRange') or '')
    match = re.search(r'![A-Z]+(\d+)(?::|$)', updated_range)
    if not match:
        match = re.search(r'(?:^|:|\s)(?:[A-Z]+)(\d+)(?::|$)', updated_range.split('!')[-1])
    if not match:
        return [None for _ in range(count)]
    first_row = int(match.group(1))
    return [first_row + index for index in range(count)]


def sheet_import_status(record: FcaImportRecord) -> str:
    if record.import_status == 'review_needed':
        return 'Review Needed'
    if record.import_status == 'failed':
        return 'Failed'
    return 'Imported'


def fca_review_examples(records: list[FcaImportRecord]) -> list[dict[str, str]]:
    examples = []
    for record in records:
        if record.import_status != 'review_needed':
            continue
        examples.append({
            'customer_name': record.customer_name,
            'primary_phone': record.primary_phone,
            'source': f'{record.source_filename} / {record.source_sheet} row {record.source_row}',
            'reason': record.sync_error or 'Needs manual review',
        })
        if len(examples) >= 5:
            break
    return examples


def normalize_cell(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r'\s+', ' ', str(value).strip())


def normalize_name(value: str) -> str:
    return ' '.join(str(value or '').strip().split()).upper()


def normalize_identifier(value: str) -> str:
    text = str(value or '').strip()
    if not text:
        return ''
    digits = re.sub(r'\D+', '', text)
    return digits if 5 <= len(digits) <= 12 else text


def extract_phone_numbers(value: str) -> list[str]:
    phones = []
    text = str(value or '')
    for match in PHONE_CANDIDATE_PATTERN.finditer(text):
        phone = normalize_kenyan_phone(match.group(0))
        if phone and phone not in phones:
            phones.append(phone)
    if not phones:
        for part in re.split(r'[/,;]+', text):
            phone = normalize_kenyan_phone(part)
            if phone and len(phone) >= 9 and phone not in phones:
                phones.append(phone)
    return phones[:2]


def extract_visit_date(rows: list[tuple[int, list[str]]], filename: str) -> date | None:
    top_text = ' '.join(value for _, values in rows[:4] for value in values if value)
    return parse_fca_date_text(top_text) or parse_fca_date_text(filename)


def extract_hub(rows: list[tuple[int, list[str]]]) -> str:
    for _, values in rows[:5]:
        joined = ' '.join(value for value in values if value)
        match = re.search(r'\bHUB\s*[.:?-]*\s*(.*?)(?=\b(?:field\s+officer|bro|officer|staff)\b|$)', joined, re.IGNORECASE)
        if match:
            return match.group(1).strip(' .:-?')
    return ''


def extract_officer(rows: list[tuple[int, list[str]]]) -> str:
    for _, values in rows[:6]:
        for i, val in enumerate(values):
            if val and any(needle in str(val).casefold() for needle in ('field officer', 'bro', 'officer')):
                for next_val in values[i+1:]:
                    if next_val and str(next_val).strip():
                        return str(next_val).strip(' .:-?')
                match = re.search(r'\b(?:field\s+officer|bro|officer)(?:\s*/\s*bro)?\s*[.:?-]*\s*(.+)$', str(val), re.IGNORECASE)
                if match:
                    val_part = match.group(1).strip(' .:-?')
                    if val_part:
                        return val_part
    return ''



def parse_fca_date_text(text: str) -> date | None:
    if not text:
        return None
    raw = str(text).strip()
    for pattern, formats in (
        (r'\b(20\d{2}[-/]\d{1,2}[-/]\d{1,2})\b', ('%Y-%m-%d', '%Y/%m/%d')),
        (r'\b(\d{1,2}[-/]\d{1,2}[-/]20\d{2})\b', ('%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%m-%d-%Y')),
    ):
        match = re.search(pattern, raw)
        if match:
            candidate = match.group(1)
            for fmt in formats:
                try:
                    return datetime.strptime(candidate, fmt).date()
                except ValueError:
                    continue
    cleaned = raw.replace('?', ' ').replace(':', ' ').replace(',', ' ')
    year_default = timezone.localdate().year
    patterns = (
        r'(\d{1,2})(?:st|nd|rd|th)?(?:\s*[-&]\s*\d{1,2}(?:st|nd|rd|th)?)?(?:\s+\d{1,2}(?:st|nd|rd|th)?)*\s+([A-Za-z]+)(?:\s+(20\d{2}))?',
        r'([A-Za-z]+)\s+(\d{1,2})(?:st|nd|rd|th)?(?:\s+(20\d{2}))?',
    )
    for pattern in patterns:
        match = re.search(pattern, cleaned, re.IGNORECASE)
        if not match:
            continue
        if match.group(1).isdigit():
            day = int(match.group(1))
            month = MONTHS.get(match.group(2).lower()[:3])
            year = int(match.group(3) or year_default)
        else:
            month = MONTHS.get(match.group(1).lower()[:3])
            day = int(match.group(2))
            year = int(match.group(3) or year_default)
        if month:
            try:
                return date(year, month, day)
            except ValueError:
                return None
    return None


def parse_iso_date(value: str) -> date | None:
    if not value:
        return None
    for fmt in ('%d-%b-%Y', '%d-%B-%Y'):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def format_sheet_date(value: date | None) -> str:
    return value.strftime('%d-%b-%Y') if value else ''


def format_master_date(value: date | None) -> str:
    return value.strftime('%d-%B-%Y') if value else ''


def looks_like_footer(values: list[str]) -> bool:
    text = ' '.join(values).lower()
    return any(
        marker in text
        for marker in (
            'to be visited by', 'requested by:', 'admin(after visit)',
            'updated on master data', 'approved by:', 'remarks',
        )
    )


def looks_like_section(values: list[str]) -> bool:
    text = ' '.join(value.lower() for value in values if value)
    return text in {'collections', 'collection'} or (
        sum(bool(value) for value in values) <= 2
        and any(marker in text for marker in ('collections', 'collection', 'kirinyaga'))
    )


def looks_like_decision_comment(value: str) -> bool:
    return bool(re.search(r'approve|reject|defer|await|cash|undecided|resched|not visited|not available', value or '', re.IGNORECASE))


def derive_fca_decision(comment: str) -> str:
    text = str(comment or '').strip().lower()
    if not text:
        return ''
    if re.search(r'\b(opted\s+(?:for\s+)?cash|cash)\b', text):
        return 'Cash'
    if re.search(r'\b(reject|rejected|declined|decomm?ission(?:ing)?)\b', text):
        return 'Rejected'
    if re.search(r'\b(approved|approve|appraisal)\b', text):
        return 'Approved'
    if re.search(
        r'\b(defer|deferred|undecided|awaiting|more time|requested more time|'
        r'reschedule|rescheduled|reschendule|not available|not availble|not visited|'
        r'no deposit|no commitment|revisit)\b',
        text,
    ):
        return 'Deferred'
    return ''
