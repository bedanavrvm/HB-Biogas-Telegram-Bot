"""Payment document generation using the HB payment workbook template."""
from __future__ import annotations

import io
import json
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from django.db import transaction
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter

from core.models import JawabuFarmerMaster, ParsedInvoice, PaymentDocument, PaymentDocumentTemplate
from core.services.invoice_parser import clean_amount
from core.services.requisition import copy_row_formatting
from core.services.template_storage import TemplateStorageError, workbook_source_from_template


PAYMENT_TEMPLATE_FILENAME = 'HB_PAYMENT__89__7__machine_ready (1).xlsx'
PAYMENT_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

logger = logging.getLogger(__name__)


class PaymentTemplateError(RuntimeError):
    pass


def normalize_payment_number(value: Any) -> str:
    payment_number = str(value or '').strip().lstrip('#').strip()
    if not payment_number:
        raise PaymentTemplateError('Payment number is required.')
    if not payment_number.isdigit() or len(payment_number) > 20:
        raise PaymentTemplateError('Payment number must contain digits only (for example, 89).')
    return payment_number


@dataclass(frozen=True)
class PaymentTemplateLayout:
    sheet_name: str
    header_row: int
    data_start_row: int
    totals_row: int
    signature_block_start_row: int
    columns: dict[str, int]
    sum_columns: tuple[int, ...]
    config_warnings: tuple[str, ...] = ()


def _template_source():
    active_template = PaymentDocumentTemplate.objects.filter(
        is_active=True,
    ).order_by('-updated_at', '-created_at').first()
    fallback_path = Path('requisition') / PAYMENT_TEMPLATE_FILENAME
    try:
        return workbook_source_from_template(active_template, fallback_path=fallback_path)
    except TemplateStorageError as exc:
        raise PaymentTemplateError(
            'No payment document template is available. Upload one in '
            'Django Admin > Payment document templates and confirm it was stored in Google Drive.'
        ) from exc


def _read_config_sheet(wb) -> dict[str, str]:
    if '_TEMPLATE_CONFIG' not in wb.sheetnames:
        return {}
    ws = wb['_TEMPLATE_CONFIG']
    config = {}
    for row in range(2, ws.max_row + 1):
        key = str(ws.cell(row=row, column=1).value or '').strip()
        value = ws.cell(row=row, column=2).value
        if key:
            config[key] = '' if value is None else str(value).strip()
    return config


def _words(value: Any) -> set[str]:
    return set(re.findall(r'\b\w+\b', str(value or '').upper()))


def _detect_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 50) + 1):
        row_words = set()
        for col in range(1, ws.max_column + 1):
            row_words |= _words(ws.cell(row=row, column=col).value)
        if {'CUST', 'NO', 'NAME', 'BRANCH'} <= row_words and ('INVOICE' in row_words or 'AMOUNT' in row_words):
            return row
    raise PaymentTemplateError('Could not detect payment workbook header row.')


def _column_mapping_from_headers(ws, header_row: int) -> dict[str, int]:
    mapping = {}
    for col in range(1, ws.max_column + 1):
        header_text = str(ws.cell(row=header_row, column=col).value or '')
        text = ' '.join(
            str(ws.cell(row=row, column=col).value or '')
            for row in (header_row, header_row + 1)
        )
        words = _words(text)
        header_words = _words(header_text)
        header_upper = header_text.upper().strip()
        if 'REQUISITION' in words and 'DATE' in words:
            mapping['requisition_date'] = col
        elif 'ORDER' in words:
            mapping['order_no'] = col
        elif 'CUST' in words and 'NO' in words:
            mapping['cust_no'] = col
        elif header_upper in {'NO:', 'NO', 'NO.'} or header_words == {'NO'}:
            mapping['no'] = col
        elif 'NAME' in words and 'IMAB' in words:
            mapping['name_imab'] = col
        elif words == {'NAME'} or ('NAME' in words and 'IMAB' not in words):
            mapping.setdefault('name', col)
        elif 'PRIMARY' in words and 'MOBILE' in words:
            mapping['mobile_no'] = col
        elif 'SECONDARY' in words and 'MOBILE' in words:
            mapping['secondary_mobile'] = col
        elif 'BRANCH' in words:
            mapping['branch'] = col
        elif 'LOAN' in words and 'OFFICER' in words:
            mapping['loan_officer'] = col
        elif 'HB' in words and 'INVOICE' in words and 'AMOUNT' in words:
            mapping['hb_invoice_amount'] = col
        elif 'EXPECTED' in words and 'INVOICE' in words:
            mapping['expected_invoice_amount'] = col
        elif 'DISCOUNT' in words:
            mapping['discount'] = col
        elif 'DEPOSIT' in words and 'HBG' in words:
            mapping['deposit_paid_hbg'] = col
        elif 'DEPOSIT' in words and 'JBL' in words:
            mapping['deposit_paid_jbl'] = col
        elif 'LOAN' in words and 'AMOUNT' in words:
            mapping['loan_amount'] = col
        elif 'REPAYMENT' in words:
            mapping['repayment_dates'] = col
        elif 'TENOR' in words:
            mapping['tenor'] = col
        elif 'PRODUCT' in words:
            mapping['product'] = col
        elif 'CALL' in words and 'COMMENTS' in words:
            mapping['call_up_comments'] = col
    required = {
        'requisition_date', 'order_no', 'cust_no', 'no', 'name_imab', 'name',
        'mobile_no', 'branch', 'hb_invoice_amount', 'discount',
        'deposit_paid_hbg', 'deposit_paid_jbl', 'loan_amount',
        'repayment_dates', 'tenor', 'call_up_comments',
    }
    missing = sorted(required - set(mapping))
    if missing:
        raise PaymentTemplateError(f"Payment workbook is missing required columns: {', '.join(missing)}")
    return mapping


def _first_numeric_row(ws, start_row: int, serial_col: int) -> int:
    for row in range(start_row, min(ws.max_row, 80) + 1):
        value = ws.cell(row=row, column=serial_col).value
        if value is None:
            continue
        try:
            int(str(value).strip())
            return row
        except ValueError:
            continue
    raise PaymentTemplateError('Could not detect payment workbook first data row.')


def _detect_totals_row(ws, start_row: int, sum_columns: tuple[int, ...]) -> int:
    candidate_columns = sum_columns or tuple(range(1, ws.max_column + 1))
    for row in range(start_row, min(ws.max_row, 160) + 1):
        formula_count = 0
        for col in candidate_columns:
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and value.strip().upper().startswith('=SUM('):
                formula_count += 1
        if formula_count >= 2:
            return row
    raise PaymentTemplateError('Could not detect payment workbook totals row.')


def _last_numeric_row(ws, start_row: int, serial_col: int) -> int:
    last = start_row
    for row in range(start_row, min(ws.max_row, 120) + 1):
        value = ws.cell(row=row, column=serial_col).value
        try:
            int(str(value).strip())
            last = row
        except (TypeError, ValueError):
            if row > start_row:
                break
    return last


def _sum_columns_from_config(config: dict[str, str], columns: dict[str, int]) -> tuple[int, ...]:
    value = config.get('sum_cols') or ''
    result = []
    for item in value.split(','):
        item = item.strip()
        if not item:
            continue
        try:
            result.append(column_index_from_string(item))
        except ValueError:
            pass
    if result:
        return tuple(result)
    return tuple(
        columns[key]
        for key in (
            'hb_invoice_amount', 'expected_invoice_amount', 'discount',
            'deposit_paid_hbg', 'deposit_paid_jbl', 'loan_amount',
        )
        if key in columns
    )


def payment_template_layout(wb) -> PaymentTemplateLayout:
    config = _read_config_sheet(wb)
    sheet_name = config.get('sheet_name') or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    header_row = _detect_header_row(ws)
    columns = _column_mapping_from_headers(ws, header_row)
    sum_columns = _sum_columns_from_config(config, columns)
    try:
        data_start_row = _first_numeric_row(ws, header_row + 1, columns['no'])
        last_data_row = _last_numeric_row(ws, data_start_row, columns['no'])
        totals_row = last_data_row + 1
    except PaymentTemplateError:
        totals_row = _detect_totals_row(ws, header_row + 1, sum_columns)
        data_start_row = header_row + 1
    signature_block_start_row = totals_row + 3
    warnings = []
    expected = {
        'header_row': header_row,
        'data_start_row': data_start_row,
        'totals_row': totals_row,
        'signature_block_start_row': signature_block_start_row,
    }
    for key, detected in expected.items():
        configured = config.get(key)
        if configured and str(configured) != str(detected):
            warnings.append(f'{key} config={configured} visible={detected}')
    return PaymentTemplateLayout(
        sheet_name=sheet_name,
        header_row=header_row,
        data_start_row=data_start_row,
        totals_row=totals_row,
        signature_block_start_row=signature_block_start_row,
        columns=columns,
        sum_columns=sum_columns,
        config_warnings=tuple(warnings),
    )


def _amount(value) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    return clean_amount(str(value))


def _xlsx_number(value):
    amount = _amount(value)
    if amount is None:
        return ''
    if amount == amount.to_integral_value():
        return int(amount)
    return float(amount)


def _invoice_for_farmer(farmer: JawabuFarmerMaster) -> ParsedInvoice | None:
    if not farmer.invoice_number:
        return None
    return (
        ParsedInvoice.objects
        .filter(invoice_no=farmer.invoice_number)
        .filter(
            matched_farmer=farmer,
        )
        .order_by('-updated_at')
        .first()
    )


def _row_payload(
    farmer: JawabuFarmerMaster,
    *,
    call_up_comments: str | None = None,
    case_call_up_comments: dict[str, str] | None = None,
) -> tuple[dict[str, Any], list[str], ParsedInvoice | None]:
    invoice = _invoice_for_farmer(farmer)
    missing = []
    if not farmer.customer_no:
        missing.append('Cust No')
    if not invoice:
        missing.append('Matched invoice')
    else:
        from core.services.invoice_identity import identity_gate
        identity = identity_gate(invoice, farmer)
        if identity.get('blocker') == 'invoice_name_change_pending':
            missing.append('Invoice name change pending')
        elif identity.get('blocker') == 'invoice_name_change_required':
            missing.append('Invoice name change required')
        elif identity.get('blocker'):
            missing.append('Invoice identity verification pending')
    if farmer.balance_due is None:
        missing.append('Balance Due')
    if not farmer.repayment_date:
        missing.append('Repayment Dates')
    if not farmer.repayment_tenor:
        missing.append('Tenor')
    from core.services.jawabu_approvals import JawabuApprovalError, require_effective_approval
    try:
        require_effective_approval(farmer, 'final_review')
    except JawabuApprovalError:
        missing.append('Current final approval')

    if farmer.payment_product and not farmer.product_version_id:
        missing.append('Global product mapping')
    elif farmer.product_version_id:
        from core.services.product_catalog import missing_product_requirements

        product_missing = missing_product_requirements(
            farmer.product_version,
            workflow='jawabu_portal',
            stage='payment',
            evidence=farmer.product_requirement_evidence,
        )
        missing.extend(item['label'] for item in product_missing)

    from core.services.requisition import requisition_deposit_values

    hbg_deposit, jbl_deposit = requisition_deposit_values(farmer)

    case_comment = (case_call_up_comments or {}).get(str(farmer.id))
    row = {
        'requisition_date': farmer.requisition_date,
        'order_no': farmer.order_number,
        'cust_no': farmer.customer_no,
        'name_imab': farmer.imab_customer_name,
        'name': farmer.customer_name,
        'mobile_no': farmer.primary_phone,
        'secondary_mobile': farmer.secondary_phone,
        'branch': farmer.system_branch or farmer.branch,
        'loan_officer': farmer.system_loan_officer or farmer.jbl_officer,
        # BALANCE DUE from the source invoice is the amount used for payment.
        # Expected Invoice Amount stays blank until its formula is agreed.
        'hb_invoice_amount': farmer.balance_due,
        'expected_invoice_amount': None,
        'discount': farmer.discount,
        'deposit_paid_hbg': hbg_deposit,
        'deposit_paid_jbl': jbl_deposit,
        'loan_amount': None,
        'repayment_dates': farmer.repayment_date,
        'tenor': farmer.repayment_tenor,
        'product': farmer.payment_product,
        # Payment COL is a separate Head-of-Rural checkpoint.  It must not
        # inherit the earlier order/requisition decision comment: a payment
        # draft is intentionally blank until HOR approves that batch.
        'call_up_comments': str(
            case_comment if case_comment is not None else (call_up_comments or '')
        ).strip(),
        # This is reference-only metadata for the in-app review; the payment
        # workbook layout intentionally ignores it.
        'farmer_id': str(farmer.id),
        'order_call_up_comments': str(farmer.final_decision_comment or '').strip(),
    }
    return row, missing, invoice


def payment_readiness(
    order_number: str = '',
    farmer_ids: list[str] | None = None,
    *,
    call_up_comments: str | None = None,
    case_call_up_comments: dict[str, str] | None = None,
) -> dict[str, Any]:
    queryset = JawabuFarmerMaster.objects.filter(status='active')
    if farmer_ids is not None:
        queryset = queryset.filter(id__in=farmer_ids)
    else:
        queryset = queryset.filter(order_number=order_number)
    farmers = list(queryset.order_by('customer_name'))
    ready = []
    blocked = []
    invoice_batch_ids = set()
    for farmer in farmers:
        row, missing, invoice = _row_payload(
            farmer,
            call_up_comments=call_up_comments,
            case_call_up_comments=case_call_up_comments,
        )
        item = {
            'farmer_id': str(farmer.id),
            'customer_name': farmer.customer_name,
            'national_id': farmer.national_id,
            'primary_phone': farmer.primary_phone,
            'missing': missing,
            'row': row,
            # Kept beside (not in) the payment COL so reviewers can compare
            # the earlier order decision without confusing the two comments.
            'order_call_up_comments': str(farmer.final_decision_comment or '').strip(),
            'blocker_codes': [],
        }
        if invoice:
            from core.services.invoice_identity import identity_gate
            identity_blocker = identity_gate(invoice, farmer).get('blocker')
            if identity_blocker:
                item['blocker_codes'].append(identity_blocker)
            invoice_batch_ids.add(str(invoice.batch_id))
            item['invoice_id'] = str(invoice.id)
            item['invoice_number'] = invoice.invoice_no
            item['invoice_batch_id'] = str(invoice.batch_id)
        if missing:
            blocked.append(item)
        else:
            ready.append(item)
    return {
        'order_number': order_number,
        'order_numbers': sorted({item['row'].get('order_no') for item in ready if item['row'].get('order_no')}),
        'total_clients': len(farmers),
        'ready_count': len(ready),
        'blocked_count': len(blocked),
        'invoice_batch_ids': sorted(invoice_batch_ids),
        'ready': ready,
        'blocked': blocked,
    }


def _set_cell(ws, row: int, col: int | None, value):
    if not col:
        return
    cell = ws.cell(row=row, column=col)
    if value is None:
        cell.value = ''
        return
    if isinstance(value, (date, datetime)):
        cell.value = value
        cell.number_format = 'dd-mmm-yyyy'
    elif isinstance(value, Decimal):
        cell.value = _xlsx_number(value)
    else:
        cell.value = value or ''


def _copy_payment_template_row(ws, src_row: int, dst_row: int) -> None:
    copy_row_formatting(ws, src_row, dst_row)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        if isinstance(src_cell.value, str) and src_cell.value.startswith('='):
            dst_cell = ws.cell(row=dst_row, column=col)
            dst_cell.value = Translator(src_cell.value, origin=src_cell.coordinate).translate_formula(dst_cell.coordinate)


def _write_payment_rows(ws, layout: PaymentTemplateLayout, rows: list[dict[str, Any]]) -> int:
    first_data_row = layout.data_start_row
    template_rows = max(1, layout.totals_row - layout.data_start_row)
    count = len(rows)
    if count > template_rows:
        insert_at = layout.totals_row
        ws.insert_rows(insert_at, count - template_rows)
        for row in range(insert_at, insert_at + (count - template_rows)):
            _copy_payment_template_row(ws, first_data_row, row)
    elif count < template_rows:
        ws.delete_rows(first_data_row + count, template_rows - count)

    for index, payload in enumerate(rows, start=1):
        row = first_data_row + index - 1
        _set_cell(ws, row, layout.columns.get('no'), index)
        for key, value in payload.items():
            _set_cell(ws, row, layout.columns.get(key), value)
            if key in {
                'hb_invoice_amount', 'expected_invoice_amount', 'discount',
                'deposit_paid_hbg', 'deposit_paid_jbl', 'loan_amount',
            } and layout.columns.get(key):
                ws.cell(row=row, column=layout.columns[key]).number_format = '0'

    totals_row = first_data_row + count
    for col in layout.sum_columns:
        if col == layout.columns.get('expected_invoice_amount'):
            ws.cell(row=totals_row, column=col, value='')
            ws.cell(row=totals_row, column=col).number_format = '0'
            continue
        letter = get_column_letter(col)
        if count:
            ws.cell(row=totals_row, column=col, value=f'=SUM({letter}{first_data_row}:{letter}{first_data_row + count - 1})')
        else:
            ws.cell(row=totals_row, column=col, value=0)
        ws.cell(row=totals_row, column=col).number_format = '0'
    return totals_row


def generate_payment_workbook(
    order_number: str,
    payment_number: str,
    farmer_ids: list[str] | None = None,
    *,
    call_up_comments: str | None = None,
    case_call_up_comments: dict[str, str] | None = None,
) -> tuple[bytes, dict[str, Any]]:
    payment_number = normalize_payment_number(payment_number)
    readiness = payment_readiness(
        order_number,
        farmer_ids=farmer_ids,
        call_up_comments=call_up_comments,
        case_call_up_comments=case_call_up_comments,
    )
    if farmer_ids is not None and len(readiness['ready']) + len(readiness['blocked']) != len(set(farmer_ids)):
        raise PaymentTemplateError('One or more selected payment cases was not found or is inactive.')
    if not readiness['ready']:
        raise PaymentTemplateError('Select at least one invoice-matched case for this payment batch.')
    if readiness['blocked_count']:
        raise PaymentTemplateError('Payment document has blocked rows. Resolve missing fields before generating.')
    from core.services.template_validation import template_source_bytes, validate_template_bytes, UnsafeTemplateError
    template_bytes = template_source_bytes(_template_source())
    try:
        validate_template_bytes(template_bytes, 'payment')
    except UnsafeTemplateError as exc:
        raise PaymentTemplateError(str(exc)) from exc
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    layout = payment_template_layout(wb)
    ws = wb[layout.sheet_name]
    rows = [item['row'] for item in readiness['ready']]
    totals_row = _write_payment_rows(ws, layout, rows)
    payment_label = f'#{payment_number}'
    ws['H4'] = payment_label
    ws.title = payment_label
    config = wb['_TEMPLATE_CONFIG'] if '_TEMPLATE_CONFIG' in wb.sheetnames else None
    if config:
        for row in range(2, config.max_row + 1):
            if str(config.cell(row=row, column=1).value or '').strip() == 'sheet_name':
                config.cell(row=row, column=2, value=payment_label)
                break
    out = io.BytesIO()
    wb.save(out)
    summary = {
        **{key: value for key, value in readiness.items() if key not in {'ready', 'blocked'}},
        'template_sheet': layout.sheet_name,
        'payment_number': payment_number,
        'header_row': layout.header_row,
        'data_start_row': layout.data_start_row,
        'totals_row': totals_row,
        'config_warnings': list(layout.config_warnings),
    }
    return out.getvalue(), summary


def _upload_payment_workbook(data: bytes, filename: str, actor: str, order_number: str) -> tuple[str, str]:
    from core.services.order_approval import GoogleDriveMediaStorage

    return GoogleDriveMediaStorage().upload(
        data,
        filename=filename,
        mime_type=PAYMENT_CONTENT_TYPE,
        id_number='payment_documents',
        received_at=timezone.now(),
        group_config=None,
        workflow_key='Jawabu/Payment Documents',
        record_type='Order',
        record_key=order_number,
    )


def create_payment_document(
    order_number: str,
    payment_number: str,
    actor: str = '',
    final: bool = False,
    farmer_ids: list[str] | None = None,
    status: str | None = None,
    call_up_comments: str | None = None,
    case_call_up_comments: dict[str, str] | None = None,
) -> PaymentDocument:
    """Create a preview or Head-of-Rural review artifact.

    ``final`` remains in the signature for old callers, but direct final
    generation is rejected so approval is always an explicit state transition.
    """
    payment_number = normalize_payment_number(payment_number)
    if final or status == 'final':
        raise PaymentTemplateError(
            'Direct final payment generation is disabled. Submit a payment review and approve it through Head of Rural.'
        )
    artifact_status = status or 'preview'
    if artifact_status not in {'preview', 'pending_review'}:
        raise PaymentTemplateError('Unsupported payment document status.')
    xlsx, summary = generate_payment_workbook(
        order_number,
        payment_number,
        farmer_ids=farmer_ids,
        call_up_comments=call_up_comments,
        case_call_up_comments=case_call_up_comments,
    )
    readiness_snapshot = payment_readiness(
        order_number,
        farmer_ids=farmer_ids,
        call_up_comments=call_up_comments,
        case_call_up_comments=case_call_up_comments,
    )
    printable_rows = json.loads(json.dumps(
        [item['row'] for item in readiness_snapshot['ready']], cls=DjangoJSONEncoder,
    ))
    # Reserve a monotonic artifact version before contacting Drive. Preview
    # generations are immutable snapshots too, so retries never overwrite a
    # previous preview or create an ambiguous same-name Drive file.
    with transaction.atomic():
        # Lock the selected/order farmers while reserving the version. This
        # serializes concurrent Mini App submissions for the same payment
        # scope even when no PaymentDocument row exists yet.
        from core.models import JawabuFarmerMaster
        farmer_lock = JawabuFarmerMaster.objects.select_for_update()
        if farmer_ids:
            farmer_lock = farmer_lock.filter(id__in=farmer_ids)
        else:
            farmer_lock = farmer_lock.filter(order_number=order_number)
        list(farmer_lock.values_list('id', flat=True))
        latest = (
            PaymentDocument.objects.select_for_update()
            .filter(order_number=order_number, payment_number=payment_number)
            .order_by('-version', '-created_at').first()
        )
        version = (latest.version + 1) if latest else 1
        filename_status = 'review' if artifact_status == 'pending_review' else artifact_status
        filename = f"HB_Payment_{payment_number}_{order_number}_{filename_status}_v{version}.xlsx"
        document_values = {
            'order_number': order_number,
            'payment_number': payment_number,
            'version': version,
            'filename': filename,
            'file_content': xlsx,
            'generated_by': actor,
            'reviewed_by': '',
            'reviewed_at': None,
            'call_up_comments': call_up_comments or '',
            'case_call_up_comments': case_call_up_comments or {},
            'finalized_by': '',
            'finalized_at': None,
            'row_count': summary.get('ready_count', 0),
            'farmer_ids': [item['farmer_id'] for item in readiness_snapshot['ready']],
            'invoice_batch_ids': summary.get('invoice_batch_ids', []),
            'validation_summary': {
                **summary,
                'preview_rows': printable_rows,
                'artifact_status': artifact_status,
            },
            'drive_file_id': '',
            'drive_url': '',
            'error': '',
        }
        # A failed upload remains visible as a retryable artifact instead of
        # leaving the workflow with no audit record.
        doc = PaymentDocument.objects.create(status=artifact_status, **document_values)
    from core.services.document_sync import mark_drive_attempt, mark_drive_failure, mark_drive_success
    mark_drive_attempt(doc)
    try:
        drive_file_id, drive_url = _upload_payment_workbook(xlsx, filename, actor, order_number)
    except Exception:
        logger.exception('Payment workbook upload failed: order=%s payment=%s', order_number, payment_number)
        mark_drive_failure(doc, 'Drive upload failed; retry required.', error_field='error')
        doc.status = 'failed'
        doc.save(update_fields=['status', 'updated_at'])
        raise

    try:
        from core.models import JawabuFarmerMaster
        from core.services.jawabu_case360 import record_pipeline_event
        with transaction.atomic():
            doc.status = artifact_status
            mark_drive_success(
                doc, file_id=drive_file_id, url=drive_url, error_field='error',
                update_fields=['status'],
            )
            if artifact_status == 'pending_review':
                for farmer in JawabuFarmerMaster.objects.filter(id__in=doc.farmer_ids):
                    record_pipeline_event(
                        farmer,
                        action='payment_review_submitted',
                        stage_key='payment',
                        actor=actor,
                        request_id=f'payment-review:{doc.id}:{farmer.id}',
                        source='payment_document',
                        new_values={'payment_number': payment_number, 'version': doc.version},
                        metadata={'payment_document_id': str(doc.id)},
                    )
    except Exception:
        logger.exception('Payment document finalization failed after Drive upload: order=%s payment=%s', order_number, payment_number)
        doc.status = 'failed'
        doc.drive_file_id = drive_file_id
        doc.drive_url = drive_url
        doc.error = 'Local finalization failed after Drive upload; reconciliation is required.'
        doc.save(update_fields=['status', 'drive_file_id', 'drive_url', 'error', 'updated_at'])
        raise
    return doc


def approve_payment_document(
    document_id: str,
    *,
    actor: str = '',
    actor_user=None,
    access: dict | None = None,
    call_up_comments: str = '',
    case_call_up_comments: dict[str, str] | None = None,
) -> PaymentDocument:
    """Approve a payment review snapshot and create the immutable final file.

    The review snapshot is retained as an audit record. Approval requires a
    distinct Head-of-Rural Call Up Comment for every case, writes those values
    to the corresponding COL cells, and creates a separate final artifact.
    """
    comment = str(call_up_comments or '').strip()

    from core.models import JawabuFarmerMaster

    with transaction.atomic():
        review = PaymentDocument.objects.select_for_update().get(pk=document_id)
        if review.status == 'final':
            return review
        if review.status == 'reviewed':
            final_id = str((review.validation_summary or {}).get('final_document_id') or '')
            if final_id:
                existing_final = PaymentDocument.objects.filter(pk=final_id, status='final').first()
                if existing_final:
                    return existing_final
            raise PaymentTemplateError('This payment review has already been completed.')
        if review.status != 'pending_review':
            raise PaymentTemplateError('This payment document is no longer awaiting Head of Rural review.')
        farmer_ids = list(review.farmer_ids or [])
        if not farmer_ids:
            raise PaymentTemplateError('The payment review has no selected cases.')
        comments = {
            str(key): str(value or '').strip()
            for key, value in (case_call_up_comments or {}).items()
            if str(key).strip()
        }
        # Legacy clients sent one batch comment. Preserve compatibility by
        # applying it to every case, while the current UI always sends an
        # explicit comment for each farmer.
        if not comments and comment:
            comments = {str(farmer_id): comment for farmer_id in farmer_ids}
        missing_comments = [str(farmer_id) for farmer_id in farmer_ids if not comments.get(str(farmer_id))]
        if missing_comments:
            raise PaymentTemplateError('Enter a Head of Rural Call Up Comment for every selected case.')
        readiness = payment_readiness(
            review.order_number,
            farmer_ids=farmer_ids,
            call_up_comments=comment,
            case_call_up_comments=comments,
        )
        if readiness['blocked_count']:
            raise PaymentTemplateError('Payment data changed since review. Resolve the blocked rows and regenerate the review.')

        # Reserve the next artifact version while holding the review lock.  A
        # final version is unique per order, and concurrent approvals must not
        # produce two files with the same version/name.
        latest = (
            PaymentDocument.objects.select_for_update()
            .filter(order_number=review.order_number)
            .order_by('-version', '-created_at')
            .first()
        )
        version = (latest.version + 1) if latest else 1
        final = PaymentDocument.objects.create(
            order_number=review.order_number,
            payment_number=review.payment_number,
            # Reserve the version without claiming that the final artifact is
            # complete.  The status becomes ``final`` only after Drive upload
            # and local audit updates succeed.
            status='failed',
            version=version,
            filename=f'HB_Payment_{review.payment_number}_{review.order_number}_final_v{version}.xlsx',
            generated_by=review.generated_by,
            error='Final workbook upload pending.',
            call_up_comments=comment,
            case_call_up_comments=comments,
            row_count=readiness['ready_count'],
            farmer_ids=farmer_ids,
            invoice_batch_ids=readiness.get('invoice_batch_ids', []),
            validation_summary={
                **{key: value for key, value in readiness.items() if key not in {'ready', 'blocked'}},
                'artifact_status': 'final',
                'preview_rows': json.loads(json.dumps(
                    [item['row'] for item in readiness['ready']], cls=DjangoJSONEncoder,
                )),
                'review_document_id': str(review.id),
            },
        )

    # Generate/upload outside the transaction.  The final row remains visible
    # as failed if Drive is unavailable, making reconciliation explicit.
    from core.services.document_sync import mark_drive_attempt, mark_drive_failure, mark_drive_success
    mark_drive_attempt(final)
    try:
        xlsx, _summary = generate_payment_workbook(
            review.order_number,
            review.payment_number,
            farmer_ids=farmer_ids,
            call_up_comments=comment,
            case_call_up_comments=comments,
        )
        # Preserve the exact final workbook before Drive upload.  A later
        # physical signature/stamp scan must be bound to these bytes, never a
        # mutable Drive URL or regenerated live case data.
        final.file_content = xlsx
        final.save(update_fields=['file_content', 'updated_at'])
        drive_file_id, drive_url = _upload_payment_workbook(
            xlsx, final.filename, actor, review.order_number,
        )
    except Exception:
        logger.exception(
            'Payment approval upload failed: document=%s order=%s payment=%s',
            review.id, review.order_number, review.payment_number,
        )
        mark_drive_failure(final, 'Drive upload failed; retry required.', error_field='error')
        final.status = 'failed'
        final.save(update_fields=['status', 'updated_at'])
        raise

    try:
        from core.services.jawabu_case360 import record_pipeline_event
        with transaction.atomic():
            final.status = 'final'
            mark_drive_success(
                final, file_id=drive_file_id, url=drive_url, error_field='error',
                update_fields=['status', 'finalized_by', 'finalized_at', 'reviewed_by', 'reviewed_at'],
            )
            final.finalized_by = actor
            final.finalized_at = timezone.now()
            final.reviewed_by = actor
            final.reviewed_at = timezone.now()
            final.save(update_fields=[
                'status', 'drive_file_id', 'drive_url', 'error', 'drive_next_retry_at',
                'finalized_by', 'finalized_at', 'reviewed_by', 'reviewed_at', 'updated_at',
            ])
            review.status = 'reviewed'
            review.reviewed_by = actor
            review.reviewed_at = timezone.now()
            review.call_up_comments = comment
            review.case_call_up_comments = comments
            review.validation_summary = {
                **(review.validation_summary or {}),
                'final_document_id': str(final.id),
            }
            review.save(update_fields=[
                'status', 'reviewed_by', 'reviewed_at', 'call_up_comments',
                'case_call_up_comments',
                'validation_summary', 'updated_at',
            ])
            for farmer in JawabuFarmerMaster.objects.filter(id__in=farmer_ids):
                from core.services.jawabu_approvals import record_approval
                record_approval(
                    farmer=farmer,
                    gate='payment_review',
                    decision='Approved',
                    comment=comments.get(str(farmer.id), comment),
                    actor=actor_user,
                    actor_label=actor,
                    access=access,
                    payment_document=final,
                )
                event = record_pipeline_event(
                    farmer,
                    action='payment_finalized',
                    stage_key='payment',
                    actor=actor,
                    request_id=f'payment-document:{final.id}:{farmer.id}',
                    source='payment_document',
                    new_values={
                        'order_number': farmer.order_number,
                        'payment_number': review.payment_number,
                        'version': final.version,
                        'call_up_comments': comment,
                        'case_call_up_comments': comments,
                    },
                    metadata={
                        'payment_document_id': str(final.id),
                        'review_document_id': str(review.id),
                    },
                    actor_user=actor_user,
                )
                from core.services.jawabu_comments import record_case_comment
                record_case_comment(
                    farmer=farmer,
                    stage_key='payment',
                    comment=comments.get(str(farmer.id), comment),
                    actor=actor,
                    actor_user=actor_user,
                    request_id=f'payment-document:{final.id}:{farmer.id}',
                    pipeline_event=event,
                    occurred_at=event.occurred_at,
                )
                from core.services.portal_publication import reserve_farmer_publication
                reserve_farmer_publication(
                    farmer,
                    request_id=f'payment-comment:{final.id}:{farmer.id}',
                    requested_by=actor_user,
                    requested_by_label=actor,
                )
    except Exception:
        logger.exception(
            'Payment approval finalization failed after Drive upload: document=%s', review.id,
        )
        final.status = 'failed'
        final.drive_file_id = drive_file_id
        final.drive_url = drive_url
        final.error = 'Local finalization failed after Drive upload; reconciliation is required.'
        final.save(update_fields=['status', 'drive_file_id', 'drive_url', 'error', 'updated_at'])
        raise
    return final


def serialize_payment_document(doc: PaymentDocument) -> dict[str, Any]:
    sync_error = str(doc.error or '').strip()
    if sync_error:
        sync_status = 'pending' if 'pending' in sync_error.casefold() else 'retryable_failure'
    elif str(doc.drive_url or '').strip():
        sync_status = 'succeeded'
    else:
        sync_status = 'not_requested'
    return {
        'id': str(doc.id),
        'order_number': doc.order_number,
        'payment_number': doc.payment_number,
        'status': doc.status,
        'version': doc.version,
        'filename': doc.filename,
        'drive_url': doc.drive_url,
        'sync_status': sync_status,
        'sync_error': sync_error,
        'sync_attempts': int(getattr(doc, 'drive_sync_attempts', 0) or 0),
        'next_retry_at': doc.drive_next_retry_at.isoformat() if getattr(doc, 'drive_next_retry_at', None) else None,
        'reviewed_by': doc.reviewed_by,
        'reviewed_at': doc.reviewed_at.isoformat() if doc.reviewed_at else None,
        'call_up_comments': doc.call_up_comments,
        'case_call_up_comments': doc.case_call_up_comments or {},
        'farmer_ids': [str(value) for value in (doc.farmer_ids or [])],
        'row_count': doc.row_count,
        'validation_summary': doc.validation_summary,
        'created_at': doc.created_at.isoformat() if doc.created_at else None,
        'finalized_at': doc.finalized_at.isoformat() if doc.finalized_at else None,
    }
