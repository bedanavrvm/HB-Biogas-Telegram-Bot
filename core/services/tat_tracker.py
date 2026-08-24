
"""TAT Tracker Mini App workflow."""
from __future__ import annotations

import base64
import binascii
import csv
import hmac
import hashlib
import io
import json
import logging
import re
import time
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from urllib.parse import parse_qsl, urlencode
from typing import Any

from django.conf import settings
from django.core import signing
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
import openpyxl

from core.models import TatCaseSequence, TatTrackerApprovalCertificate, TatTrackerCase, TatTrackerEvent
from core.services.access_policies import BUSINESS_ADMIN_ROLE
from core.services.branches import DEFAULT_WORKFLOW_BRANCHES, global_branch_choices, workflow_branches as configured_workflow_branches
from core.services.business_calendar import business_minutes_between
from core.services.identifiers import normalize_kenyan_phone, normalize_national_id
from core.services.sheets import get_sheets_service
from core.services.workflow_escalations import latest_escalation
from core.services.workflow_timeline import tat_case_timeline
from core.services.workflow_data_mode import (
    WORKFLOW_TAT,
    assert_record_writable,
    is_record_operational,
    mode_snapshot,
    operational_tat_cases,
    serialize_mode,
)

logger = logging.getLogger(__name__)

_TAT_HEADER_CACHE_TTL_SECONDS = 300
_TAT_HEADER_CACHE: dict[tuple[str, str, str], tuple[float, list[Any]]] = {}

TAT_TRACKER_WORKFLOW_TYPE = 'tat_tracker'
TAT_TRACKER_HEADER_ROW = 2
TAT_FORM_TOKEN_SALT = 'tat-tracker-mini-app'

BRANCHES = DEFAULT_WORKFLOW_BRANCHES
DECISION_OPTIONS = ['Approved', 'Rejected', 'Deferred']
SANCTIONS_OPTIONS = ['Pending', 'Met', 'Not Met']
REGISTER_OPTIONS = ['10:00am', '1:00pm', '3:30pm']
REGISTER_APPROVED_OPTIONS = ['Approved', 'Pending']
MINUTES_SHARED_OPTIONS = ['Yes', 'No']
BRO_APPLIED_OPTIONS = ['Pending', 'Met', 'Not Met']
STATUS_VALUES = ['Active', 'Disbursed', 'Rejected', 'Declined', 'Deferred', 'Stalled', 'Pending Docs']
TAT_BATCH_FORMAT_TEXT = (
    "TAT batch upload format\n\n"
    "Attach an Excel .xlsx or CSV file and send @bot /batch.\n\n"
    "Required headers:\n"
    "Product, Client Name, National ID, Phone, Branch, Amount\n\n"
    "Example row:\n"
    "business, Mary Wanjiku, 12345678, 254712345678, Nakuru, 25000\n\n"
    "Accepted products: business, logbook, mjengo, kilimo, micro_asset.\n"
    "The uploader must be configured as a BRO for the selected branch/product."
)
DEFAULT_TAT_TARGETS_MINUTES = {
    'business': {'total': 20160, 'stages': {}},
    'logbook': {'total': 20160, 'stages': {}},
    'mjengo': {'total': 20160, 'stages': {}},
    'kilimo': {'total': 20160, 'stages': {}},
    'micro_asset': {'total': 20160, 'stages': {}},
}
NEAR_SLA_RATIO = Decimal('0.8')
TAT_TARGET_MANAGER_ROLES = frozenset({'IT'})
TAT_CASE_CORRECTION_ROLES = frozenset({'IT', BUSINESS_ADMIN_ROLE})
TAT_HOME_PAGE_SIZE = 25
TAT_HOME_QUEUES = frozenset({'assigned', 'role', 'all'})
TAT_COMPLETED_STATUSES = frozenset({'Disbursed', 'Rejected', 'Declined'})
TAT_CREATE_INTENT_NEW_LOAN = 'new_loan'


@dataclass(frozen=True)
class StageConfig:
    key: str
    label: str
    column: int
    role: str
    kind: str = 'timestamp'
    options: tuple[str, ...] = ()
    auto_timestamp_key: str = ''
    requires_signature_certificate: bool = False


@dataclass(frozen=True)
class ProductConfig:
    key: str
    label: str
    sheet_name: str
    case_prefix: str
    min_amount: Decimal
    max_amount: Decimal | None
    remarks_col: int
    status_col: int
    tat_start_col: int
    stage_columns: dict[str, int]
    stages: tuple[StageConfig, ...]
    product_id: int | None = None
    version_id: str = ''
    stage_tat_columns: tuple['StageTatColumn', ...] = ()


@dataclass(frozen=True)
class StageTatColumn:
    stage_key: str
    fallback_col: int
    aliases: tuple[str, ...]


BASE_STAGES_OTHER = (
    StageConfig('mpesa_to_admin', 'MPESA sent to Admin', 9, 'BRO'),
    StageConfig('mpesa_verified', 'MPESA verified by Business Admin and sent to CA', 10, BUSINESS_ADMIN_ROLE),
    StageConfig('ca_analysis_sent', 'Credit analysis sent', 11, 'CA'),
    StageConfig('bro_response', 'BRO response to CA', 12, 'BRO'),
    StageConfig('bm_tat_request', 'BM TAT request sent', 13, 'BM'),
    StageConfig('tat_scheduled', 'HOCC scheduled', 14, 'SECRETARY'),
    StageConfig('tat_held', 'HOCC held', 15, 'SECRETARY'),
    StageConfig('decision', 'Decision', 16, 'CHAIR', 'dropdown', tuple(DECISION_OPTIONS), 'decision_ts'),
    StageConfig('minutes_shared', 'Minutes shared', 18, 'SECRETARY', 'dropdown', tuple(MINUTES_SHARED_OPTIONS), 'minutes_shared_ts'),
    StageConfig('sanctions', 'Sanctions', 19, 'LOAN_APPROVER', 'dropdown', tuple(SANCTIONS_OPTIONS), 'sanctions_ts'),
    StageConfig('bro_applied', 'BRO applied on system', 21, 'BRO', 'dropdown', tuple(BRO_APPLIED_OPTIONS), 'bro_applied_ts'),
    StageConfig('disbursement_register', 'Business Admin disbursement register', 22, BUSINESS_ADMIN_ROLE, 'dropdown', tuple(REGISTER_OPTIONS), 'register_ts'),
    StageConfig('register_approved', 'Register approved', 24, 'LOAN_APPROVER', 'dropdown', tuple(REGISTER_APPROVED_OPTIONS), 'register_approved_ts'),
    StageConfig('disbursement', 'Finance disbursement', 25, 'FINANCE'),
)

BASE_STAGES_LOGBOOK = (
    StageConfig('mpesa_to_admin', 'MPESA sent to Admin', 9, 'BRO'),
    StageConfig('mpesa_verified', 'MPESA verified by Business Admin and sent to CA', 10, BUSINESS_ADMIN_ROLE),
    StageConfig('ca_analysis_sent', 'Credit analysis sent', 11, 'CA'),
    StageConfig('bro_response', 'BRO response to CA', 12, 'BRO'),
    StageConfig('valuation_ready', 'Valuation ready', 13, 'BM'),
    StageConfig('bm_tat_request', 'BM TAT request sent', 14, 'BM'),
    StageConfig('tat_scheduled', 'HOCC scheduled', 15, 'SECRETARY'),
    StageConfig('tat_held', 'HOCC held', 16, 'SECRETARY'),
    StageConfig('decision', 'Decision', 17, 'CHAIR', 'dropdown', tuple(DECISION_OPTIONS), 'decision_ts'),
    StageConfig('minutes_shared', 'Minutes shared', 19, 'SECRETARY', 'dropdown', tuple(MINUTES_SHARED_OPTIONS), 'minutes_shared_ts'),
    StageConfig('sanctions', 'Sanctions', 20, 'LOAN_APPROVER', 'dropdown', tuple(SANCTIONS_OPTIONS), 'sanctions_ts'),
    StageConfig('bro_applied', 'BRO applied on system', 22, 'BRO', 'dropdown', tuple(BRO_APPLIED_OPTIONS), 'bro_applied_ts'),
    StageConfig('disbursement_register', 'Business Admin disbursement register', 23, BUSINESS_ADMIN_ROLE, 'dropdown', tuple(REGISTER_OPTIONS), 'register_ts'),
    StageConfig('register_approved', 'Register approved', 25, 'LOAN_APPROVER', 'dropdown', tuple(REGISTER_APPROVED_OPTIONS), 'register_approved_ts'),
    StageConfig('disbursement', 'Finance disbursement', 26, 'FINANCE'),
)

BASE_STAGES_BUSINESS = (
    StageConfig('mpesa_to_admin', 'MPESA sent to Admin', 9, 'BRO'),
    StageConfig('mpesa_verified', 'MPESA verified by Business Admin and sent to CA', 10, BUSINESS_ADMIN_ROLE),
    StageConfig('ca_analysis_sent', 'Credit analysis sent', 11, 'CA'),
    StageConfig('bro_response', 'BRO response to CA', 12, 'BRO'),
    StageConfig('bm_response', 'BM response to CA', 13, 'BM', requires_signature_certificate=True),
    StageConfig('bro_applied', 'BRO applied loan on system', 14, 'BRO'),
    StageConfig('disbursement_register', 'Business Admin disbursement register', 15, BUSINESS_ADMIN_ROLE, 'dropdown', tuple(REGISTER_OPTIONS), 'register_ts'),
    StageConfig('register_approved', 'Register approved', 17, 'LOAN_APPROVER', 'dropdown', tuple(REGISTER_APPROVED_OPTIONS), 'register_approved_ts'),
    StageConfig('disbursement', 'Finance disbursement', 18, 'FINANCE'),
)

PRODUCTS: dict[str, ProductConfig] = {
    'logbook': ProductConfig('logbook', 'Logbook', 'TRACKER-LOGBOOK', 'JBL-LB', Decimal('50000'), Decimal('700000'), 28, 27, 29, {'created': 8, 'decision_ts': 18, 'sanctions_ts': 21, 'register_ts': 24}, BASE_STAGES_LOGBOOK),
    'mjengo': ProductConfig('mjengo', 'Mjengo', 'TRACKER-MJENGO', 'JBL-MJ', Decimal('10000'), Decimal('500000'), 27, 26, 28, {'created': 8, 'decision_ts': 17, 'sanctions_ts': 20, 'register_ts': 23}, BASE_STAGES_OTHER),
    'kilimo': ProductConfig('kilimo', 'Kilimo', 'TRACKER-KILIMO', 'JBL-KI', Decimal('50000'), Decimal('300000'), 27, 26, 28, {'created': 8, 'decision_ts': 17, 'sanctions_ts': 20, 'register_ts': 23}, BASE_STAGES_OTHER),
    'micro_asset': ProductConfig('micro_asset', 'Micro Asset', 'TRACKER-MICRO-ASSET', 'JBL-MA', Decimal('10000'), Decimal('300000'), 27, 26, 28, {'created': 8, 'decision_ts': 17, 'sanctions_ts': 20, 'register_ts': 23}, BASE_STAGES_OTHER),
    'business': ProductConfig('business', 'Business', 'TRACKER-Business', 'JBL-BS', Decimal('5000'), None, 20, 19, 21, {'created': 8, 'register_ts': 16}, BASE_STAGES_BUSINESS),
}


def _stage_tat_aliases(stage: StageConfig) -> tuple[str, ...]:
    label = stage.label
    return (
        f'{label} TAT Minutes',
        f'{label} TAT',
        f'{label} Lag',
        f'{label} Lag Minutes',
        f'{stage.key} TAT Minutes',
    )


STAGE_TAT_COLUMNS: dict[str, tuple[StageTatColumn, ...]] = {
    'logbook': tuple(
        StageTatColumn(stage.key, 31 + index, _stage_tat_aliases(stage))
        for index, stage in enumerate(BASE_STAGES_LOGBOOK)
    ),
    'mjengo': tuple(
        StageTatColumn(stage.key, 30 + index, _stage_tat_aliases(stage))
        for index, stage in enumerate(BASE_STAGES_OTHER)
    ),
    'kilimo': tuple(
        StageTatColumn(stage.key, 30 + index, _stage_tat_aliases(stage))
        for index, stage in enumerate(BASE_STAGES_OTHER)
    ),
    'micro_asset': tuple(
        StageTatColumn(stage.key, 30 + index, _stage_tat_aliases(stage))
        for index, stage in enumerate(BASE_STAGES_OTHER)
    ),
    'business': tuple(
        StageTatColumn(stage.key, 23 + index, _stage_tat_aliases(stage))
        for index, stage in enumerate(BASE_STAGES_BUSINESS)
    ),
}


def is_tat_tracker_workflow(group_config) -> bool:
    workflow = getattr(group_config, 'workflow', None) or {}
    return str(workflow.get('type') or '') == TAT_TRACKER_WORKFLOW_TYPE


def configured_products(workflow: dict | None = None) -> list[ProductConfig]:
    workflow = workflow or {}
    keys = workflow.get('products') or []
    if not keys:
        try:
            from core.models import Product
            from core.services.product_catalog import (
                active_product_version, product_is_selectable,
            )
            keys = [
                product.code
                for product in Product.objects.filter(active=True).order_by('sort_order', 'name')
                if (version := active_product_version(product))
                and hasattr(version, 'tat_configuration')
                and product_is_selectable(
                    product=product, workflow='tat_tracker', channel='portal',
                )
            ]
        except Exception:
            keys = list(PRODUCTS.keys())
    resolved = []
    for key in keys:
        try:
            resolved.append(_database_product_by_key(key))
        except Exception:
            if key in PRODUCTS:
                resolved.append(PRODUCTS[key])
    return resolved


def _database_product_by_key(key: str) -> ProductConfig:
    """Adapt governed database configuration to the established TAT interface."""
    from core.models import ProductTatConfiguration
    from core.services.product_catalog import active_product_version, resolve_product

    product = resolve_product(key)
    version = active_product_version(product) if product else None
    if version is None:
        raise ValueError('Invalid product.')
    try:
        configuration = ProductTatConfiguration.objects.get(product_version=version)
    except ProductTatConfiguration.DoesNotExist as exc:
        raise ValueError('This product is not configured for TAT Tracker.') from exc
    stages = tuple(StageConfig(
        key=str(item.get('key') or ''), label=str(item.get('label') or item.get('key') or ''),
        column=int(item.get('column') or 0), role=str(item.get('role') or ''),
        kind=str(item.get('kind') or 'timestamp'), options=tuple(item.get('options') or ()),
        auto_timestamp_key=str(item.get('auto_timestamp_key') or ''),
        requires_signature_certificate=bool(item.get('requires_signature_certificate', False)),
    ) for item in (configuration.stages or []) if item.get('key'))
    stages_by_key = {stage.key: stage for stage in stages}
    tat_columns = tuple(StageTatColumn(
        stage_key=str(item.get('stage_key') or ''), fallback_col=int(item.get('fallback_col') or 0),
        aliases=tuple(item.get('aliases') or _stage_tat_aliases(
            stages_by_key[str(item.get('stage_key') or '')]
        )),
    ) for item in (configuration.stage_tat_columns or []) if item.get('stage_key') in stages_by_key)
    return ProductConfig(
        key=product.code, label=product.name, sheet_name=configuration.sheet_name,
        case_prefix=configuration.case_prefix, min_amount=version.min_amount,
        max_amount=version.max_amount, remarks_col=configuration.remarks_col,
        status_col=configuration.status_col, tat_start_col=configuration.tat_start_col,
        stage_columns=configuration.stage_columns or {}, stages=stages,
        product_id=product.pk, version_id=str(version.pk), stage_tat_columns=tat_columns,
    )


def workflow_branches(workflow: dict | None = None) -> list[str]:
    env_branches = str(getattr(settings, 'TAT_TRACKER_BRANCH_CHOICES', '') or '').strip()
    if env_branches:
        return configured_workflow_branches({'branches': env_branches}, default=global_branch_choices(), replace_stale_defaults=True)
    return configured_workflow_branches(workflow, default=global_branch_choices(), replace_stale_defaults=True)


def create_tat_form_token(group_id: str) -> str:
    return signing.dumps({'group_id': str(group_id)}, salt=TAT_FORM_TOKEN_SALT)


def validate_tat_form_token(token: str, group_id: str) -> tuple[bool, str]:
    if not token:
        return False, 'Form token is missing. Open the tracker again from Telegram.'
    max_age = int(getattr(settings, 'TAT_TRACKER_WEBAPP_AUTH_MAX_AGE_SECONDS', 86400))
    try:
        payload = signing.loads(token, salt=TAT_FORM_TOKEN_SALT, max_age=max_age if max_age > 0 else None)
    except signing.SignatureExpired:
        return False, 'Form token has expired. Open the tracker again from Telegram.'
    except signing.BadSignature:
        return False, 'Form token is invalid. Open the tracker again from Telegram.'
    if str(payload.get('group_id', '')) != str(group_id):
        return False, 'Form token does not match this group.'
    return True, ''


def build_tat_tracker_url(group_id: str) -> str:
    base_url = getattr(settings, 'APP_BASE_URL', '').rstrip('/')
    if not base_url:
        return ''
    return f"{base_url}/tat-tracker/?" + urlencode({'group_id': str(group_id), 'token': create_tat_form_token(group_id)})


def build_tat_tracker_mini_app_url(group_id: str) -> str:
    bot_username = str(getattr(settings, 'TELEGRAM_BOT_USERNAME', '') or '').strip().lstrip('@')
    short_name = str(getattr(settings, 'TAT_TRACKER_MINI_APP_SHORT_NAME', '') or '').strip().strip('/')
    if not bot_username or not short_name:
        return ''
    return f"https://t.me/{bot_username}/{short_name}?startapp={create_tat_start_param(group_id)}"


def build_tat_tracker_launcher_url(group_id: str) -> str:
    """Return a durable group launcher URL for a pinned Telegram message."""
    bot_username = str(getattr(settings, 'TELEGRAM_BOT_USERNAME', '') or '').strip().lstrip('@')
    short_name = str(getattr(settings, 'TAT_TRACKER_MINI_APP_SHORT_NAME', '') or '').strip().strip('/')
    if not bot_username or not short_name:
        return ''
    return f"https://t.me/{bot_username}/{short_name}?startapp={create_tat_launcher_start_param(group_id)}"


def create_tat_start_param(group_id: str) -> str:
    payload = {'group_id': str(group_id), 'token': create_tat_form_token(group_id)}
    encoded = base64.urlsafe_b64encode(json.dumps(payload, separators=(',', ':')).encode('utf-8')).decode('ascii')
    return encoded.rstrip('=')


def create_tat_launcher_start_param(group_id: str) -> str:
    """Create the non-expiring locator used only by the pinned JBL Apps message."""
    payload = {'group_id': str(group_id), 'launcher': 'jbl_apps'}
    encoded = base64.urlsafe_b64encode(json.dumps(payload, separators=(',', ':')).encode('utf-8')).decode('ascii')
    return encoded.rstrip('=')


def decode_tat_start_param(start_param: str) -> dict[str, str]:
    value = str(start_param or '').strip()
    if not value:
        return {}
    padding = '=' * (-len(value) % 4)
    try:
        payload = json.loads(base64.urlsafe_b64decode((value + padding).encode('ascii')).decode('utf-8'))
    except (binascii.Error, ValueError, json.JSONDecodeError, UnicodeDecodeError):
        return {}
    if not isinstance(payload, dict):
        return {}
    group_id = str(payload.get('group_id', '')).strip()
    token = str(payload.get('token', '')).strip()
    if str(payload.get('launcher', '')).strip() == 'jbl_apps' and group_id:
        return {'group_id': group_id, 'token': ''}
    if not group_id or not token:
        return {}
    return {'group_id': group_id, 'token': token}


def validate_tat_telegram_webapp_init_data(init_data: str) -> tuple[bool, str, dict]:
    if not getattr(settings, 'TAT_TRACKER_WEBAPP_REQUIRE_TELEGRAM_AUTH', True):
        return True, '', {}
    bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', '')
    if not bot_token:
        return False, 'TELEGRAM_BOT_TOKEN is not configured.', {}
    if not init_data:
        return False, 'Telegram Mini App authentication data is missing.', {}
    pairs = dict(parse_qsl(init_data, keep_blank_values=True))
    received_hash = pairs.pop('hash', '')
    if not received_hash:
        return False, 'Telegram Mini App hash is missing.', {}
    data_check_string = "\n".join(f"{key}={value}" for key, value in sorted(pairs.items()))
    secret_key = hmac.new(b'WebAppData', bot_token.encode('utf-8'), hashlib.sha256).digest()
    calculated_hash = hmac.new(secret_key, data_check_string.encode('utf-8'), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(calculated_hash, received_hash):
        return False, 'Telegram Mini App authentication failed.', {}
    max_age = int(getattr(settings, 'TAT_TRACKER_WEBAPP_AUTH_MAX_AGE_SECONDS', 86400))
    auth_date = pairs.get('auth_date')
    if auth_date and max_age > 0:
        try:
            if time.time() - int(auth_date) > max_age:
                return False, 'Telegram Mini App authentication expired.', {}
        except ValueError:
            return False, 'Telegram Mini App auth_date is invalid.', {}
    user_payload = {}
    if pairs.get('user'):
        try:
            user_payload = json.loads(pairs['user'])
        except json.JSONDecodeError:
            user_payload = {}
    return True, '', user_payload if isinstance(user_payload, dict) else {}


def staff_user_for_payload(group_config, user_payload: dict, fallback_name: str = '') -> dict:
    workflow = getattr(group_config, 'workflow', None) or {}
    telegram_id = str(user_payload.get('id') or '').strip()
    username = str(user_payload.get('username') or '').strip().lower().lstrip('@')
    full_name = _telegram_name(user_payload) or fallback_name or username or telegram_id or 'Unknown user'
    from core.services.telegram_identity import identity_from_user_payload, resolve_or_bind_telegram_user, user_access
    canonical_user = resolve_or_bind_telegram_user(identity_from_user_payload(user_payload)) if telegram_id else None
    access = user_access(canonical_user, 'tat_tracker', group_configuration=group_config)
    if access['authorized']:
        from core.services.workflow_capabilities import capabilities_payload
        from core.services.tat_notifications import mark_private_alert_seen
        mark_private_alert_seen(
            canonical_user,
            allows_write=bool(user_payload.get('allows_write_to_pm')),
        )
        profile = getattr(canonical_user, 'staff_profile', None)
        return {
            'authorized': True,
            'telegram_id': telegram_id,
            'username': username,
            'user_id': canonical_user.pk,
            'name': canonical_user.get_full_name() or canonical_user.get_username(),
            'roles': access['roles'] or ['BRO'],
            'branches': access['branches'],
            'products': access['products'],
            'capabilities': capabilities_payload(canonical_user, 'tat_tracker', access=access),
            '_canonical_user': canonical_user,
            '_access': access,
            '_group_configuration': group_config,
            'signing_national_id': str(getattr(profile, 'signing_national_id', '') or ''),
            'signing_phone_number': str(getattr(profile, 'signing_phone_number', '') or ''),
            'signing_email': str(getattr(profile, 'signing_email', '') or ''),
        }
    return {
        'authorized': False,
        'telegram_id': telegram_id,
        'username': username,
        'name': full_name,
        'roles': [],
        'branches': [],
        'products': [],
        'capabilities': [],
        'reason': 'Your Telegram account is not configured for the TAT Tracker. Ask an administrator to grant TAT access.',
    }


def configured_bro_users(
    workflow: dict | None,
    group_config=None,
    *,
    include_all_scopes: bool = False,
) -> list[dict]:
    """Return active users tagged with the TAT ``BRO`` role.

    The role tag is the active ``AccessGrant`` for the TAT Tracker workflow,
    rather than a legacy staff table or a free-text user field.  Returning the
    canonical user id and username keeps the dropdown stable when two staff
    members share a display name while retaining the existing name value sent
    by the form for backwards compatibility. TAT administrators/IT managers
    may see role-tagged BROs across scopes when assigning a case; ordinary
    staff remain group-scoped.
    """
    from core.models import AccessGrant
    from core.services.telegram_identity import database_group_configuration
    database_group = database_group_configuration(group_config) if group_config is not None else None
    grants = AccessGrant.objects.filter(
        workflow='tat_tracker', role__iexact='BRO', active=True,
        user__is_active=True,
    )
    if group_config is not None and not include_all_scopes:
        if database_group is not None:
            grants = grants.filter(group_configuration__in=[None, database_group])
        elif str(getattr(group_config, 'group_id', '') or '').strip() not in {'', '*'}:
            # An explicitly identified runtime group that has no matching
            # database row can only use global grants.  In legacy single-group
            # mode the registry uses ``*``; there is no group scope to apply,
            # so keep all active BRO role tags visible in that one workflow.
            grants = grants.filter(group_configuration__isnull=True)
    users = {}
    for grant in grants.select_related('user', 'user__staff_profile'):
        user = grant.user
        display_name = user.get_full_name() or user.get_username()
        users[user.pk] = {
            'id': user.pk,
            'name': display_name,
            'username': user.get_username(),
            'telegram_username': getattr(getattr(user, 'staff_profile', None), 'telegram_username', '') or '',
        }
    return sorted(users.values(), key=lambda item: str(item['name']).casefold())


def configured_bro_names(workflow: dict | None, group_config=None) -> list[str]:
    """Return active BRO display names for existing API consumers."""
    return [user['name'] for user in configured_bro_users(workflow, group_config)]


def bootstrap(group_config, user_payload: dict) -> dict:
    user = staff_user_for_payload(group_config, user_payload)
    workflow = getattr(group_config, 'workflow', None) or {}
    if not user['authorized']:
        return {'authorized': False, 'user': user, 'reason': user.get('reason', 'Unauthorized')}
    products = [serialize_product(product) for product in _allowed_products(workflow, user)]
    home = home_data(group_config, user)
    bro_users = configured_bro_users(
        workflow,
        group_config,
        include_all_scopes=bool(set(user.get('roles') or []) & {BUSINESS_ADMIN_ROLE, 'IT', 'MANAGEMENT'}),
    )
    return {
        'authorized': True,
        'workflow_mode': serialize_mode(WORKFLOW_TAT),
        'user': public_user(user),
        'products': products,
        'branches': _allowed_branches(workflow, user),
        # ``bro_names`` remains for older clients; new clients should use the
        # role-tagged ``bro_users`` records so duplicate names are unambiguous.
        'bro_names': [item['name'] for item in bro_users],
        'bro_users': bro_users,
        'statuses': STATUS_VALUES,
        'recent': home['recent'],
        'action_required': home['action_required'],
        'pagination': home['pagination'],
    }


def home_data(
    group_config,
    user: dict,
    *,
    action_offset: int = 0,
    recent_offset: int = 0,
    page_size: int = TAT_HOME_PAGE_SIZE,
    product_key: str = '',
    branch: str = '',
    queue: str = 'role',
    page: int = 1,
) -> dict:
    """Return one compact queue plus filtered dashboard metrics.

    The legacy ``action_required`` and ``recent`` collections remain in the
    response for cached Mini App clients during the responsive-home rollout.
    """
    workflow = getattr(group_config, 'workflow', None) or {}
    queryset = operational_tat_cases(
        TatTrackerCase.objects.filter(group_id=str(group_config.group_id), is_deleted=False)
    )
    queryset = _scope_tat_queryset(queryset, user, 'tat.home.view')
    allowed_keys = [p.key for p in _allowed_products(workflow, user)]
    selected_product = str(product_key or '').strip()
    if selected_product:
        if selected_product in allowed_keys:
            queryset = queryset.filter(product_key=selected_product)
        else:
            queryset = queryset.none()
    allowed_branches = _allowed_branches(workflow, user)
    selected_branch = str(branch or '').strip()
    if selected_branch:
        if selected_branch in allowed_branches:
            queryset = queryset.filter(branch=selected_branch)
        else:
            queryset = queryset.none()
    action_offset = max(0, int(action_offset or 0))
    recent_offset = max(0, int(recent_offset or 0))
    page_size = max(1, min(int(page_size or TAT_HOME_PAGE_SIZE), 50))
    queue_key = str(queue or 'role').strip().lower()
    if queue_key not in TAT_HOME_QUEUES:
        queue_key = 'role'
    current_page = max(1, int(page or 1))
    page_offset = (current_page - 1) * page_size

    cases = list(queryset.prefetch_related('approval_certificates'))
    recent_total = len(cases)
    recent_cases = sorted(cases, key=lambda item: item.updated_at, reverse=True)
    recent = [
        serialize_case_summary(case, user, workflow=workflow)
        for case in recent_cases[recent_offset:recent_offset + page_size]
    ]

    actionable_cases = []
    for case in sorted(cases, key=lambda item: item.updated_at):
        if case.status in TAT_COMPLETED_STATUSES:
            continue
        next_stage = next_action(case)
        if next_stage and can_user_edit_stage(user, case, next_stage):
            actionable_cases.append((case, next_stage))
    action_total = len(actionable_cases)
    action_required = [
        serialize_case_summary(case, user, next_stage=next_stage, workflow=workflow)
        for case, next_stage in actionable_cases[action_offset:action_offset + page_size]
    ]
    completed_total = sum(case.status in TAT_COMPLETED_STATUSES for case in cases)
    stalled_case_ids = set()
    for case in cases:
        if case.status == 'Stalled':
            stalled_case_ids.add(case.pk)
            continue
        if case.status in TAT_COMPLETED_STATUSES:
            continue
        stage = next_action(case)
        if not stage:
            continue
        product = product_by_key(case.product_key)
        target = stage_target_minutes_for_case(case, workflow, product, stage)
        elapsed = stage_tat_minutes(case, stage)
        if target is not None and target > 0 and elapsed is not None and elapsed > target:
            stalled_case_ids.add(case.pk)

    from core.services.tat_notifications import inbox_payload
    from core.services.telegram_identity import database_group_configuration

    assigned = inbox_payload(
        user.get('_canonical_user'),
        group=database_group_configuration(group_config),
        group_id=str(group_config.group_id),
        limit=page_size,
        offset=page_offset if queue_key == 'assigned' else 0,
        product_key=selected_product,
        branch=selected_branch,
    )

    if queue_key == 'assigned':
        items = assigned['items']
        selected_total = assigned['total']
    elif queue_key == 'all':
        items = [
            serialize_case_summary(case, user, workflow=workflow)
            for case in recent_cases[page_offset:page_offset + page_size]
        ]
        selected_total = recent_total
    else:
        items = [
            serialize_case_summary(case, user, next_stage=next_stage, workflow=workflow)
            for case, next_stage in actionable_cases[page_offset:page_offset + page_size]
        ]
        selected_total = action_total

    selected_pages = max(1, (selected_total + page_size - 1) // page_size)
    if current_page > selected_pages:
        current_page = selected_pages
        page_offset = (current_page - 1) * page_size
        if queue_key == 'assigned':
            assigned = inbox_payload(
                user.get('_canonical_user'),
                group=database_group_configuration(group_config),
                group_id=str(group_config.group_id),
                limit=page_size,
                offset=page_offset,
                product_key=selected_product,
                branch=selected_branch,
            )
            items = assigned['items']
        elif queue_key == 'all':
            items = [
                serialize_case_summary(case, user, workflow=workflow)
                for case in recent_cases[page_offset:page_offset + page_size]
            ]
        else:
            items = [
                serialize_case_summary(case, user, next_stage=next_stage, workflow=workflow)
                for case, next_stage in actionable_cases[page_offset:page_offset + page_size]
            ]

    selected_pagination = pagination_payload(page_offset, page_size, selected_total, len(items))
    selected_pagination.update({
        'page': current_page,
        'pages': selected_pages,
    })
    return {
        'queue': queue_key,
        'items': items,
        'metrics': {
            'assigned': assigned['total'],
            'role': action_total,
            'total': recent_total,
            'completed': completed_total,
            'stalled': len(stalled_case_ids),
        },
        'recent': recent,
        'action_required': action_required,
        'pagination': {
            **selected_pagination,
            'recent': pagination_payload(recent_offset, page_size, recent_total, len(recent)),
            'action_required': pagination_payload(action_offset, page_size, action_total, len(action_required)),
        },
    }


def pagination_payload(offset: int, page_size: int, total: int, returned: int) -> dict:
    return {
        'offset': offset,
        'page_size': page_size,
        'total': total,
        'has_more': offset + returned < total,
    }


def search_cases(group_config, user: dict, query: str) -> list[dict]:
    q = str(query or '').strip()
    if len(q) < 2:
        return []
    workflow = getattr(group_config, 'workflow', None) or {}
    normalized_id = normalize_national_id(q)
    normalized_phone = normalize_kenyan_phone(q)
    query = Q(case_id__icontains=q) | Q(client_name__icontains=q) | Q(branch__icontains=q) | Q(bro_name__icontains=q)
    if normalized_id:
        query |= Q(national_id=normalized_id)
    if normalized_phone:
        query |= Q(primary_phone=normalized_phone)
    queryset = operational_tat_cases(
        TatTrackerCase.objects.filter(group_id=str(group_config.group_id), is_deleted=False)
    ).filter(query)
    queryset = _scope_tat_queryset(queryset, user, 'tat.case.search')
    return [serialize_case_summary(case, user, workflow=workflow) for case in queryset.order_by('-updated_at')[:25]]


def get_case_detail(group_config, user: dict, case_id: str) -> dict:
    case = TatTrackerCase.objects.get(group_id=str(group_config.group_id), case_id=str(case_id), is_deleted=False)
    if not _tat_scope_allowed(user, 'tat.home.view', case):
        raise ValueError('This TAT case is outside your assigned access scope.')
    return serialize_case_detail(case, user, workflow=getattr(group_config, 'workflow', None) or {})


def record_tat_event(**values) -> TatTrackerEvent:
    """Create one native TAT event and its idempotent compliance projection."""
    event = TatTrackerEvent.objects.create(**values)
    from core.services.compliance_audit import record_event

    source = str(event.source or '')
    record_event(
        workflow='tat_tracker',
        action=f'tat.{event.transition_code or event.stage_key or "case.update"}',
        category='workflow_transition' if event.transition_code else 'workflow',
        origin='external_sync' if source == 'sheet_sync' else ('system' if source == 'workflow_transition' and not event.actor_user_id else 'human'),
        subject_type='tat_case',
        subject_id=str(event.case_id),
        customer_reference=str(event.case.case_id),
        actor=event.actor_user,
        authority_user=event.authority_user,
        actor_label=event.actor_name,
        authority_label=_user_label(event.authority_user),
        request_id=event.request_id,
        source_model='TatTrackerEvent',
        source_event_id=str(event.pk),
        deduplication_key=f'tat:TatTrackerEvent:{event.pk}',
        before_values={'value': event.old_value} if event.old_value else {},
        after_values={'value': event.new_value} if event.new_value else {},
        metadata={
            'stage_key': event.stage_key,
            'stage_label': event.stage_label,
            'source': source,
            'transition_code': event.transition_code,
            'from_state': event.from_state,
            'to_state': event.to_state,
            'reason': event.reason,
            'revision_before': event.revision_before,
            'revision_after': event.revision_after,
            'data_mode': event.case.data_mode,
            'pilot_cycle_id': str(event.case.pilot_cycle_id or ''),
            'data_scope_key': event.case.data_scope_key,
        },
        sensitive=bool(event.old_value or event.new_value),
        occurred_at=event.created_at,
    )
    return event


def _user_label(user) -> str:
    return str(user.get_full_name() or user.get_username() or '').strip() if user else ''


def soft_delete_tat_case(
    case: TatTrackerCase,
    *,
    actor_name: str = '',
    actor_telegram_id: str = '',
    actor_role: str = '',
    reason: str = '',
) -> bool:
    """Mark a TAT case as deleted while preserving the case and its events."""
    if case.is_deleted:
        return False
    deleted_at = timezone.now()
    reason = str(reason or 'Deleted from Django admin.').strip()
    case.is_deleted = True
    case.deleted_at = deleted_at
    case.deleted_by = str(actor_name or '').strip()
    case.deletion_reason = reason
    case.last_updated_by = str(actor_name or '').strip()
    case.sync_error = ''
    case.save(update_fields=[
        'is_deleted',
        'deleted_at',
        'deleted_by',
        'deletion_reason',
        'last_updated_by',
        'sync_error',
        'updated_at',
    ])
    record_tat_event(
        case=case,
        group_id=case.group_id,
        actor_name=case.deleted_by,
        actor_telegram_id=str(actor_telegram_id or ''),
        actor_role=str(actor_role or 'DJANGO_ADMIN'),
        stage_key='deleted',
        stage_label='Case Deleted',
        old_value='Active backend record',
        new_value=reason,
        source='admin_correction',
        sheet_name=case.sheet_name,
        row_number=case.row_number,
    )
    return True


@transaction.atomic
def create_case(group_config, user: dict, payload: dict) -> dict:
    _validate_tat_create_payload(payload)
    product = product_by_key(str(payload.get('product_key') or payload.get('product') or ''))
    workflow = getattr(group_config, 'workflow', None) or {}
    if product not in _allowed_products(workflow, user):
        raise ValueError('You do not have access to this product.')
    client_name = str(payload.get('client_name') or '').strip().upper()
    national_id = normalize_national_id(payload.get('national_id'))
    primary_phone = normalize_kenyan_phone(payload.get('primary_phone'))
    branch = str(payload.get('branch') or '').strip()
    bro_name = str(payload.get('bro_name') or user.get('name') or '').strip()
    amount = parse_amount(payload.get('amount'))
    if not client_name:
        raise ValueError('Client name is required.')
    if not re.fullmatch(r'\d{7,8}', national_id):
        raise ValueError('ID number must be 7 or 8 digits.')
    if not primary_phone:
        raise ValueError('Enter a valid Kenyan phone number.')
    if branch not in _allowed_branches(workflow, user):
        raise ValueError('Select a valid branch.')
    if not _tat_scope_allowed_for_values(user, 'tat.case.create', branch=branch, product=product.key):
        raise ValueError('This branch and product combination is outside your assigned access scope.')
    validate_amount(product, amount)
    product_version = None
    terms_snapshot = {}
    quote_snapshot = {}
    requirement_evidence = payload.get('product_requirement_evidence') or {}
    custom_values = payload.get('product_custom_values') or {}
    selected_fee_keys = payload.get('product_selected_fee_keys') or []
    if not isinstance(requirement_evidence, dict):
        raise ValueError('Product requirement evidence must be an object.')
    if not isinstance(custom_values, dict):
        raise ValueError('Product custom values must be an object.')
    if not isinstance(selected_fee_keys, list):
        raise ValueError('Selected product fees must be a list.')
    if product.version_id:
        from core.models import OperationalLocation, ProductVersion
        from core.services.product_catalog import (
            missing_product_requirements, product_is_available, serialize_product_version,
            validate_custom_values,
        )
        product_version = ProductVersion.objects.select_related('product').get(pk=product.version_id)
        branch_record = OperationalLocation.objects.filter(location_type='branch', name__iexact=branch, active=True).first()
        if not product_is_available(product_version.product, branch=branch_record, workflow='tat_tracker', channel='portal'):
            raise ValueError('This product is not available for the selected branch and channel.')
        missing = missing_product_requirements(
            product_version, workflow='tat_tracker', stage='created', evidence=requirement_evidence,
        )
        if missing:
            raise ValueError('Complete required product evidence: ' + ', '.join(item['label'] for item in missing))
        custom_errors = validate_custom_values(product_version, custom_values, workflow='tat_tracker')
        if custom_errors:
            raise ValueError(next(iter(custom_errors.values())))
        allowed_fees = set(product_version.fees.filter(mandatory=False).values_list('key', flat=True))
        if set(selected_fee_keys) - allowed_fees:
            raise ValueError('One or more selected fees are not available for this product version.')
        terms_snapshot = serialize_product_version(product_version)
        if payload.get('tenor') not in (None, ''):
            from core.services.product_quotes import calculate_product_quote
            quote_snapshot = calculate_product_quote(
                product_version, amount=amount, tenor=payload.get('tenor'),
                optional_fee_keys=selected_fee_keys,
            )
    scope = mode_snapshot(WORKFLOW_TAT, for_update=True)
    expected_mode_version = payload.get('workflow_mode_version')
    if expected_mode_version not in (None, '') and int(expected_mode_version) != scope.mode_version:
        from core.services.workflow_data_mode import WorkflowModeChanged
        raise WorkflowModeChanged()
    create_request_id = normalize_create_request_id(payload.get('client_request_id') or payload.get('create_request_id') or payload.get('request_id'))
    if create_request_id:
        existing = TatTrackerCase.objects.select_for_update().filter(
            group_id=str(group_config.group_id),
            create_request_id=create_request_id,
            data_scope_key=scope.data_scope_key,
        ).first()
        if existing:
            if existing.is_deleted:
                raise ValueError('This submission was previously deleted. Refresh the Mini App before trying again.')
            from core.services.tat_notifications import synchronize_case_task
            synchronize_case_task(group_config, existing)
            return serialize_case_detail(existing, user, workflow=workflow)
    case_id = next_case_id(group_config, product)
    now = timezone.now()
    stage_target_snapshots = {}
    if product.stages:
        first_stage = product.stages[0]
        target = stage_target_minutes(workflow, product, first_stage)
        if target is not None:
            stage_target_snapshots[first_stage.key] = {
                'target_minutes': str(target),
                'settings_version': int(workflow.get('settings_version') or 1),
                'started_at': now.isoformat(),
            }
    case = TatTrackerCase.objects.create(
        **scope.creation_fields(),
        group_id=str(group_config.group_id), sheet_id=str(group_config.sheet_id or ''), sheet_name=product.sheet_name,
        create_request_id=create_request_id,
        case_id=case_id, product_key=product.key, product_label=product.label, client_name=client_name,
        product_id=product.product_id, product_version=product_version,
        product_terms_snapshot=terms_snapshot, product_quote_snapshot=quote_snapshot,
        product_requirement_evidence=requirement_evidence,
        product_custom_values=custom_values,
        product_selected_fee_keys=selected_fee_keys,
        national_id=national_id, primary_phone=primary_phone,
        branch=branch, bro_name=bro_name, amount=amount, stage_values={'created': now.isoformat()},
        stage_target_snapshots=stage_target_snapshots,
        status='Active', current_stage=(product.stages[0].key if product.stages else ''),
        created_by=user.get('name', ''), created_by_telegram_id=user.get('telegram_id', ''), last_updated_by=user.get('name', ''),
    )
    record_tat_event(case=case, group_id=case.group_id, actor_name=user.get('name', ''), actor_telegram_id=user.get('telegram_id', ''), actor_role=','.join(user.get('roles') or []), actor_user_id=user.get('user_id') or None, authority_user_id=user.get('user_id') or None, stage_key='created', stage_label='Case Created', new_value=format_datetime(now), source='mini_app', sheet_name=case.sheet_name)
    from core.services.tat_notifications import synchronize_case_task
    synchronize_case_task(group_config, case)
    if payload.get('_defer_sheet_sync'):
        return serialize_case_detail(case, user, workflow=workflow)
    sync_case_to_sheet(group_config, case)
    if not case.row_number:
        raise RuntimeError('TAT tracker sheet sync did not return a row number. Case was not saved.')
    return serialize_case_detail(case, user, workflow=workflow)


def _validate_tat_create_payload(payload: dict) -> None:
    """Reject update-shaped submissions before a new loan case can be made.

    A client can legitimately open more than one TAT loan for the same person,
    so national ID and phone are context signals rather than create blockers.
    The one invariant we can enforce without blocking valid repeat loans is
    that a correction must never be accepted by the create endpoint.
    """
    if any(str(payload.get(key) or '').strip() for key in ('case_id', 'workflow_revision', 'revision')):
        raise ValueError('Case corrections must use the existing case update action, not Create new loan case.')
    if payload.get('updates'):
        raise ValueError('Case corrections must use the existing case update action, not Create new loan case.')
    intent = str(payload.get('creation_intent') or TAT_CREATE_INTENT_NEW_LOAN).strip().lower()
    if intent != TAT_CREATE_INTENT_NEW_LOAN:
        raise ValueError('Only an explicit new-loan submission may create a TAT case.')


def tat_case_identity_context(
    group_config,
    national_id: object = '',
    primary_phone: object = '',
    *,
    user: dict | None = None,
) -> dict[str, Any]:
    """Return exact same-customer TAT context without treating it as a duplicate.

    TAT tracks loans, not a single lifetime customer application. These exact
    matches help a user choose an existing case for a correction while still
    allowing a deliberate additional loan to be created.
    """
    normalized_id = normalize_national_id(national_id)
    normalized_phone = normalize_kenyan_phone(primary_phone)
    query = Q()
    matched_on: list[str] = []
    if normalized_id:
        query |= Q(national_id=normalized_id)
        matched_on.append('National ID')
    if normalized_phone:
        query |= Q(primary_phone=normalized_phone)
        matched_on.append('primary phone')
    if not query.children:
        return {'matches': [], 'matched_on': []}

    cases = operational_tat_cases(TatTrackerCase.objects.filter(
        group_id=str(group_config.group_id), is_deleted=False,
    )).filter(query)
    if user is not None:
        cases = _scope_tat_queryset(cases, user, 'tat.case.create')
    cases = cases.order_by('-updated_at')[:5]
    return {
        'matched_on': matched_on,
        'matches': [
            {
                'case_id': case.case_id,
                'client_name': case.client_name,
                'product': case.product_label or case.product_key,
                'status': case.status,
                'current_stage': case.current_stage,
                'updated_at': format_datetime(case.updated_at),
            }
            for case in cases
        ],
    }


def tat_batch_format_message() -> str:
    return TAT_BATCH_FORMAT_TEXT


def process_tat_batch_upload(
    group_config,
    batch_text: str,
    *,
    user_payload: dict,
    telegram_message_id: str,
    sender: str = '',
) -> dict:
    user = staff_user_for_payload(group_config, user_payload, fallback_name=sender)
    if not user.get('authorized'):
        return {
            'status': 'command',
            'reply_text': user.get('reason') or 'Your Telegram account is not configured for the TAT Tracker.',
        }
    if not _tat_has_capability(user, 'tat.batch.upload'):
        return {
            'status': 'command',
            'reply_text': 'Only configured BRO users or roles granted TAT batch-upload access can upload case batches.',
        }

    try:
        rows = parse_tat_batch_rows(batch_text)
    except ValueError as exc:
        return {
            'status': 'command',
            'reply_text': f"{exc}\n\n{TAT_BATCH_FORMAT_TEXT}",
        }
    if not rows:
        return {'status': 'command', 'reply_text': TAT_BATCH_FORMAT_TEXT}

    return process_tat_batch_rows(
        group_config,
        rows,
        user=user,
        telegram_message_id=telegram_message_id,
        sender=sender,
    )


def process_tat_batch_file(
    group_config,
    *,
    filename: str,
    content: bytes,
    user_payload: dict,
    telegram_message_id: str,
    sender: str = '',
) -> dict:
    user = staff_user_for_payload(group_config, user_payload, fallback_name=sender)
    if not user.get('authorized'):
        return {
            'status': 'command',
            'reply_text': user.get('reason') or 'Your Telegram account is not configured for the TAT Tracker.',
        }
    if not _tat_has_capability(user, 'tat.batch.upload'):
        return {
            'status': 'command',
            'reply_text': 'Only configured BRO users or roles granted TAT batch-upload access can upload case batches.',
        }

    try:
        rows = parse_tat_batch_file(filename, content)
    except ValueError as exc:
        return {
            'status': 'command',
            'reply_text': f"{exc}\n\n{TAT_BATCH_FORMAT_TEXT}",
        }
    if not rows:
        return {'status': 'command', 'reply_text': TAT_BATCH_FORMAT_TEXT}

    return process_tat_batch_rows(
        group_config,
        rows,
        user=user,
        telegram_message_id=telegram_message_id,
        sender=sender,
    )


def process_tat_batch_rows(
    group_config,
    rows: list[dict],
    *,
    user: dict,
    telegram_message_id: str,
    sender: str = '',
) -> dict:
    imported = 0
    duplicates = 0
    failed = 0
    errors = []
    case_ids = []
    created_cases = []
    for row in rows:
        payload = dict(row['payload'])
        payload['bro_name'] = user.get('name') or sender or payload.get('bro_name') or ''
        payload['client_request_id'] = f"tat-batch:{group_config.group_id}:{telegram_message_id}:{row['line_number']}"
        payload['_defer_sheet_sync'] = True
        try:
            before_count = TatTrackerCase.objects.filter(group_id=str(group_config.group_id), is_deleted=False).count()
            result = create_case(group_config, user, payload)
            after_count = TatTrackerCase.objects.filter(group_id=str(group_config.group_id), is_deleted=False).count()
        except Exception as exc:
            failed += 1
            errors.append(f"Line {row['line_number']}: {exc}")
            continue

        summary = result.get('summary') or {}
        case_ids.append(summary.get('case_id') or '')
        if after_count == before_count:
            duplicates += 1
        else:
            imported += 1
            case_id = summary.get('case_id') or ''
            if case_id:
                created_cases.append(TatTrackerCase.objects.get(group_id=str(group_config.group_id), case_id=case_id, is_deleted=False))

    sync_result = sync_tat_batch_created_cases(group_config, created_cases)
    if sync_result['failed']:
        errors.extend(sync_result['failed'][:8])

    reply_lines = [
        'TAT batch processed.',
        f'Rows received: {len(rows)}',
        f'Created: {imported}',
        f'Already imported: {duplicates}',
        f'Synced to sheet: {sync_result["synced"]}',
        f'Failed: {failed}',
    ]
    visible_case_ids = [case_id for case_id in case_ids if case_id][:8]
    if visible_case_ids:
        reply_lines.append('Case IDs: ' + ', '.join(visible_case_ids))
    if errors:
        reply_lines.append('')
        reply_lines.append('Issues:')
        reply_lines.extend(errors[:8])

    return {
        'status': 'tat_batch_processed',
        'total': len(rows),
        'created': imported,
        'duplicates': duplicates,
        'failed': failed,
        'errors': errors,
        'case_ids': case_ids,
        'reply_text': '\n'.join(reply_lines),
    }


def parse_tat_batch_file(filename: str, content: bytes) -> list[dict]:
    lower_filename = str(filename or '').lower()
    if lower_filename.endswith('.xlsx'):
        return parse_tat_batch_xlsx(content)
    if lower_filename.endswith('.csv'):
        return parse_tat_batch_csv(decode_tat_batch_csv(content))
    raise ValueError('TAT batch upload only supports .xlsx or .csv files.')


def parse_tat_batch_csv(csv_text: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(csv_text))
    return rows_from_tat_batch_dicts(reader, line_offset=1)


def parse_tat_batch_xlsx(content: bytes) -> list[dict]:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('Could not read the Excel workbook. Save it as .xlsx and retry.') from exc
    worksheet = workbook.worksheets[0]
    header_row = None
    headers = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        values = [str(value or '').strip() for value in row]
        if required_tat_batch_fields_present(values):
            header_row = row_number
            headers = values
            break
    if not header_row:
        raise ValueError('Excel file is missing required headers: Product, Client Name, National ID, Phone, Branch, Amount.')

    dict_rows = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(str(value or '').strip() for value in row):
            continue
        values = {
            headers[index]: row[index] if index < len(row) else ''
            for index in range(len(headers))
        }
        values['__line_number'] = row_number
        dict_rows.append(values)
    return rows_from_tat_batch_dicts(dict_rows)


def decode_tat_batch_csv(content: bytes) -> str:
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252'):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError('Could not read the CSV text encoding. Export it as UTF-8 CSV and retry.')


def rows_from_tat_batch_dicts(dict_rows, *, line_offset: int = 0) -> list[dict]:
    rows = []
    for index, row in enumerate(dict_rows, start=1):
        normalized = {normalize_tat_batch_header(key): value for key, value in dict(row or {}).items()}
        line_number = int(normalized.get('line_number') or index + line_offset)
        payload = {
            'product_key': normalize_tat_batch_product(normalized.get('product')),
            'client_name': normalized.get('client_name') or '',
            'national_id': normalized.get('national_id') or '',
            'primary_phone': normalized.get('phone') or '',
            'branch': normalized.get('branch') or '',
            'amount': normalized.get('amount') or '',
        }
        if not any(str(value or '').strip() for value in payload.values()):
            continue
        missing = [
            label
            for key, label in TAT_BATCH_REQUIRED_FIELDS.items()
            if not str(payload.get(key) or '').strip()
        ]
        if missing:
            raise ValueError(f"Line {line_number}: missing required field(s): {', '.join(missing)}.")
        rows.append({'line_number': line_number, 'payload': payload})
    return rows


TAT_BATCH_REQUIRED_FIELDS = {
    'product_key': 'Product',
    'client_name': 'Client Name',
    'national_id': 'National ID',
    'primary_phone': 'Phone',
    'branch': 'Branch',
    'amount': 'Amount',
}


def required_tat_batch_fields_present(headers: list[str]) -> bool:
    normalized = {normalize_tat_batch_header(header) for header in headers}
    return {'product', 'client_name', 'national_id', 'phone', 'branch', 'amount'}.issubset(normalized)


def normalize_tat_batch_header(value: str) -> str:
    key = re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')
    aliases = {
        'product_key': 'product',
        'product_type': 'product',
        'customer_name': 'client_name',
        'name': 'client_name',
        'client': 'client_name',
        'id': 'national_id',
        'id_number': 'national_id',
        'national_id_number': 'national_id',
        'phone_number': 'phone',
        'primary_phone': 'phone',
        'mobile': 'phone',
        'loan_amount': 'amount',
    }
    return aliases.get(key, key)


def parse_tat_batch_rows(batch_text: str) -> list[dict]:
    rows = []
    for line_number, raw_line in enumerate(str(batch_text or '').splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith('#'):
            continue
        if line.lower().startswith(('product |', 'product,', 'tat batch upload format')):
            continue
        parts = split_tat_batch_line(line)
        if len(parts) != 6:
            raise ValueError(
                f"Line {line_number}: expected 6 fields: product | client name | national id | phone | branch | amount."
            )
        product, client_name, national_id, phone, branch, amount = parts
        rows.append({
            'line_number': line_number,
            'payload': {
                'product_key': normalize_tat_batch_product(product),
                'client_name': client_name,
                'national_id': national_id,
                'primary_phone': phone,
                'branch': branch,
                'amount': amount,
            },
        })
    return rows


def split_tat_batch_line(line: str) -> list[str]:
    delimiter = '|' if '|' in line else ','
    return [part.strip() for part in line.split(delimiter)]


def normalize_tat_batch_product(value: str) -> str:
    key = str(value or '').strip().lower().replace('-', '_').replace(' ', '_')
    aliases = {
        'microasset': 'micro_asset',
        'micro': 'micro_asset',
        'sme': 'business',
    }
    return aliases.get(key, key)


def sync_tat_batch_created_cases(group_config, cases: list[TatTrackerCase]) -> dict:
    result = {'synced': 0, 'failed': []}
    if not cases:
        return result
    cases_by_product: dict[str, list[TatTrackerCase]] = {}
    for case in cases:
        cases_by_product.setdefault(case.product_key, []).append(case)

    for product_key, product_cases in cases_by_product.items():
        product = product_by_key(product_key)
        service = get_sheets_service(sheet_id=group_config.sheet_id, sheet_name=product.sheet_name)
        if not service.is_available():
            error = 'Google Sheets service unavailable.'
            for case in product_cases:
                case.sync_error = error
                case.save(update_fields=['sync_error', 'updated_at'])
                mark_case_events_sync_failed(case, error)
                result['failed'].append(f'{case.case_id}: {error}')
            continue
        sheet = service._sheet
        try:
            headers = cached_tat_sheet_headers(group_config, product, sheet)
            validate_tracker_identity_headers(headers)
            existing_case_ids = sheet.col_values(1) if hasattr(sheet, 'col_values') else None
            existing_cases = []
            new_cases = []
            for case in product_cases:
                if existing_case_ids is not None and any(
                    idx >= 5 and str(value or '').strip() == str(case.case_id).strip()
                    for idx, value in enumerate(existing_case_ids, start=1)
                ):
                    existing_cases.append(case)
                else:
                    new_cases.append(case)

            # A retry may reach this function after Google accepted the
            # append but before Django recorded row numbers. Reconcile those
            # IDs in place instead of appending duplicate rows.
            for case in existing_cases:
                sync_case_to_sheet(group_config, case)
                result['synced'] += 1

            rows = [
                build_tat_sheet_row_data(group_config, case, product, headers)
                for case in new_cases
            ]
            append_result = append_tat_batch_rows(sheet, rows) if rows else None
            start_row = row_number_from_update_result(append_result)
            now = timezone.now()
            for index, case in enumerate(new_cases):
                if start_row:
                    case.row_number = start_row + index
                case.sheet_name = product.sheet_name
                case.last_synced_at = now
                case.sync_error = ''
                case.save(update_fields=['row_number', 'sheet_name', 'last_synced_at', 'sync_error', 'updated_at'])
                mark_case_events_synced(case, synced_at=now)
                result['synced'] += 1
        except Exception as exc:
            logger.exception('TAT batch sheet sync failed for product %s', product_key)
            error = str(exc)
            for case in product_cases:
                case.sync_error = error
                case.save(update_fields=['sync_error', 'updated_at'])
                mark_case_events_sync_failed(case, error)
                result['failed'].append(f'{case.case_id}: {error}')
    return result


def append_tat_batch_rows(sheet, rows: list[list[Any]]) -> Any:
    if not rows:
        return None
    if hasattr(sheet, 'append_rows'):
        return sheet.append_rows(rows, value_input_option='USER_ENTERED')
    start_row = next_sheet_row(sheet)
    width = max(len(row) for row in rows)
    sheet.update(
        f'A{start_row}:{column_letter(width)}{start_row + len(rows) - 1}',
        rows,
        value_input_option='USER_ENTERED',
    )
    return {'updates': {'updatedRange': f'A{start_row}:{column_letter(width)}{start_row + len(rows) - 1}'}}


@transaction.atomic
def update_case(
    group_config,
    user: dict,
    case_id: str,
    updates: list[dict],
    *,
    expected_revision: int | None = None,
    expected_mode_version: int | None = None,
    request_id: str = '',
    requirement_evidence: dict | None = None,
    custom_values: dict | None = None,
    selected_fee_keys: list | None = None,
) -> dict:
    """Apply one atomic, revision-checked TAT case mutation.

    Individual field events remain useful evidence, while the receipt event
    below provides one idempotency key and explicit before/after stage state
    for the entire Mini App submission.
    """
    from core.services.workflow_transitions import next_workflow_revision, validate_workflow_revision

    workflow = getattr(group_config, 'workflow', None) or {}
    case = TatTrackerCase.objects.select_for_update().get(group_id=str(group_config.group_id), case_id=str(case_id), is_deleted=False)
    assert_record_writable(case, expected_mode_version=expected_mode_version)
    if not _tat_scope_allowed(user, 'tat.home.view', case):
        raise ValueError('This TAT case is outside your assigned access scope.')
    if request_id and case.events.filter(request_id=str(request_id), source='workflow_transition').exists():
        from core.services.tat_notifications import synchronize_case_task
        synchronize_case_task(group_config, case)
        return serialize_case_detail(case, user, workflow=workflow)
    validate_workflow_revision(case, expected_revision)
    if not updates:
        raise ValueError('No updates were submitted.')
    if requirement_evidence is not None:
        if not isinstance(requirement_evidence, dict):
            raise ValueError('Product requirement evidence must be an object.')
        case.product_requirement_evidence = {
            **(case.product_requirement_evidence or {}), **requirement_evidence,
        }
    if custom_values is not None:
        if not isinstance(custom_values, dict):
            raise ValueError('Product custom values must be an object.')
        case.product_custom_values = {**(case.product_custom_values or {}), **custom_values}
    if selected_fee_keys is not None:
        if not isinstance(selected_fee_keys, list):
            raise ValueError('Selected product fees must be a list.')
        allowed_fees = set(case.product_version.fees.filter(mandatory=False).values_list('key', flat=True)) if case.product_version_id else set()
        if set(selected_fee_keys) - allowed_fees:
            raise ValueError('One or more selected fees are not available for this product version.')
        case.product_selected_fee_keys = list(dict.fromkeys(selected_fee_keys))
    from core.services.product_catalog import missing_product_requirements, validate_custom_values
    custom_errors = validate_custom_values(case.product_version, case.product_custom_values, workflow='tat_tracker')
    if custom_errors:
        raise ValueError(next(iter(custom_errors.values())))
    for item in updates:
        stage_key = str(item.get('field') or '').strip()
        missing = missing_product_requirements(
            case.product_version, workflow='tat_tracker', stage=stage_key,
            evidence=case.product_requirement_evidence,
        )
        if missing:
            raise ValueError('Complete required product evidence: ' + ', '.join(row['label'] for row in missing))
    from_state = str(case.current_stage or '')
    revision_before, revision_after = next_workflow_revision(case)
    for item in updates:
        apply_update(case, user, item, workflow=workflow)
    next_stage = next_action(case)
    case.current_stage = next_stage.key if next_stage else ''
    if next_stage:
        snapshot_stage_target(case, workflow, product_by_key(case.product_key), next_stage)
    case.last_updated_by = user.get('name', '')
    case.save(update_fields=['stage_values', 'stage_target_snapshots', 'status', 'remarks', 'current_stage', 'last_updated_by', 'workflow_revision', 'updated_at', 'client_name', 'national_id', 'primary_phone', 'branch', 'bro_name', 'amount', 'product_requirement_evidence', 'product_custom_values', 'product_selected_fee_keys'])
    record_tat_event(
        case=case,
        group_id=case.group_id,
        actor_name=user.get('name', ''),
        actor_telegram_id=user.get('telegram_id', ''),
        actor_role=','.join(user.get('roles') or []),
        actor_user_id=user.get('user_id') or None,
        authority_user_id=user.get('user_id') or None,
        stage_key='workflow_transition',
        stage_label='Workflow transition' if from_state != case.current_stage else 'Workflow update',
        source='workflow_transition',
        request_id=str(request_id or ''),
        transition_code='tat.stage.advance' if from_state != case.current_stage else 'tat.case.update',
        from_state=from_state,
        to_state=str(case.current_stage or ''),
        revision_before=revision_before,
        revision_after=revision_after,
        sheet_name=case.sheet_name,
        row_number=case.row_number,
    )
    from core.services.tat_notifications import synchronize_case_task
    synchronize_case_task(
        group_config,
        case,
        actor_user=user.get('_canonical_user'),
    )
    sync_case_to_sheet(group_config, case)
    return serialize_case_detail(case, user, workflow=workflow)


def apply_update(case: TatTrackerCase, user: dict, item: dict, *, workflow: dict | None = None) -> None:
    field = str(item.get('field') or '').strip()
    correction = bool(item.get('correction'))
    if field in {'client_name', 'national_id', 'primary_phone', 'branch', 'bro_name', 'amount'}:
        if not correction:
            raise ValueError('Case detail changes must be submitted as corrections.')
        if not can_user_correct_case_details(user, case):
            raise ValueError('Only IT or Admin staff can correct the base case details.')
        old = getattr(case, field)
        raw_value = str(item.get('value') or '').strip()
        if field == 'client_name':
            new_value = raw_value.upper()
            if not new_value:
                raise ValueError('Client name is required.')
        elif field == 'national_id':
            new_value = normalize_national_id(raw_value)
            if not re.fullmatch(r'\d{7,8}', new_value):
                raise ValueError('ID number must be 7 or 8 digits.')
        elif field == 'primary_phone':
            new_value = normalize_kenyan_phone(raw_value)
            if not new_value:
                raise ValueError('Enter a valid Kenyan phone number.')
        elif field == 'branch':
            new_value = raw_value
            if new_value not in _allowed_branches(workflow or {}, user):
                raise ValueError('Select a valid branch.')
        elif field == 'bro_name':
            new_value = raw_value
            if not new_value:
                raise ValueError('BRO name is required.')
        else:
            product = product_by_key(case.product_key)
            new_value = parse_amount(raw_value)
            validate_amount(product, new_value)
        if str(old or '') == str(new_value or ''):
            raise ValueError(f'{field.replace("_", " ").title()} is already set to that value.')
        setattr(case, field, new_value)
        stage_key = 'case_details'
        event_label = f'Corrected {field.replace("_", " ").title()}'
        record_tat_event(
            case=case,
            group_id=case.group_id,
            actor_name=user.get('name', ''),
            actor_telegram_id=user.get('telegram_id', ''),
            actor_role=','.join(user.get('roles') or []),
            actor_user_id=user.get('user_id') or None,
            authority_user_id=user.get('user_id') or None,
            stage_key=stage_key,
            stage_label=event_label,
            old_value=str(old or ''),
            new_value=str(new_value or ''),
            source='admin_correction',
            sheet_name=case.sheet_name,
            row_number=case.row_number,
        )
        return
    product = product_by_key(case.product_key)
    if field == 'remarks':
        old = case.remarks
        case.remarks = str(item.get('value') or '').strip()
        stage_key = 'remarks'
        event_label = 'Remarks / Delays'
        new = case.remarks
    else:
        stage = stage_by_key(product, field)
        if not stage:
            raise ValueError('Invalid stage submitted.')
        old = case.stage_values.get(stage.key, '')
        if correction:
            if not old:
                raise ValueError(f'{stage.label} has not been completed yet; submit it normally.')
            if not can_user_correct_stage(user, case, stage):
                raise ValueError(f'Your role cannot correct {stage.label}.')
        else:
            if not can_user_edit_stage(user, case, stage):
                raise ValueError(f'Your role cannot update {stage.label}.')
            if not previous_stages_complete(case, stage):
                raise ValueError(f'Complete the previous stage before {stage.label}.')
            if old and stage.kind != 'dropdown':
                raise ValueError(f'{stage.label} is already completed.')
        if stage.kind == 'timestamp':
            if correction:
                parsed = parse_iso_datetime(item.get('value'))
                if not parsed:
                    raise ValueError(f'Enter {stage.label} correction as a valid date and time.')
                value = parsed.isoformat()
                new = format_datetime(parsed)
            else:
                value = timezone.now().isoformat()
                new = format_datetime(timezone.now())
        elif stage.kind == 'dropdown':
            value = str(item.get('value') or '').strip()
            if value not in stage.options:
                raise ValueError(f'Select a valid value for {stage.label}.')
            if value == old:
                raise ValueError(f'{stage.label} is already set to {value}.')
            new = value
        else:
            value = str(item.get('value') or '').strip()
            new = value
        case.stage_values[stage.key] = value
        apply_side_effects(case, product, stage, value)
        stage_key = stage.key
        event_label = stage.label
    event = record_tat_event(case=case, group_id=case.group_id, actor_name=user.get('name', ''), actor_telegram_id=user.get('telegram_id', ''), actor_role=','.join(user.get('roles') or []), actor_user_id=user.get('user_id') or None, authority_user_id=user.get('user_id') or None, stage_key=stage_key, stage_label=(f'{event_label} (Correction)' if correction else event_label), old_value=str(old or ''), new_value=str(new or ''), source=('admin_correction' if correction else 'mini_app'), sheet_name=case.sheet_name, row_number=case.row_number)
    if signatures_enabled() and field != 'remarks' and stage.requires_signature_certificate:
        create_approval_certificate(case, event, user, stage)


def create_approval_certificate(case: TatTrackerCase, event: TatTrackerEvent, user: dict, stage: StageConfig) -> None:
    if not user.get('user_id') or not user.get('signing_national_id') or not user.get('signing_phone_number'):
        raise ValueError('Your Branch Manager signing identity is incomplete. Ask an administrator to add signing details to your User profile.')
    TatTrackerApprovalCertificate.objects.get_or_create(
        event=event,
        defaults={
            'case': case,
            'staff_user_id': user['user_id'],
            'signer_name': user.get('name', ''),
            'signer_telegram_id': user.get('telegram_id', ''),
            'signer_national_id': user.get('signing_national_id', ''),
            'signer_phone_number': user.get('signing_phone_number', ''),
            'signer_email': user.get('signing_email', ''),
            'stage_key': stage.key,
            'external_reference': f'TAT-{case.id}-{stage.key}-v1',
        },
    )


def apply_side_effects(case: TatTrackerCase, product: ProductConfig, stage: StageConfig, value: str) -> None:
    now = timezone.now().isoformat()
    if stage.key == 'bro_applied' and product.stage_columns.get('sanctions_ts') and case.stage_values.get('sanctions') != 'Met':
        raise ValueError('Sanctions must be marked Met before applying on system.')
    if stage.key == 'disbursement' and case.stage_values.get('register_approved') != 'Approved':
        raise ValueError('Register must be approved before disbursement.')
    if stage.auto_timestamp_key and value:
        case.stage_values.setdefault(stage.auto_timestamp_key, now)
    if stage.key == 'decision':
        if value == 'Rejected':
            case.status = 'Rejected'
        elif value == 'Deferred':
            case.status = 'Deferred'
        elif value == 'Approved' and case.status in {'Rejected', 'Deferred'}:
            case.status = 'Active'
    if stage.key == 'sanctions' and value == 'Not Met' and 'Sanctions Not Met' not in case.remarks:
        case.remarks = f"[{format_datetime(timezone.now())}: Sanctions Not Met - conditions unfulfilled] {case.remarks}".strip()
    if stage.key == 'disbursement':
        case.status = 'Disbursed'


def normalize_create_request_id(value: Any) -> str:
    cleaned = re.sub(r'[^A-Za-z0-9_.:-]', '', str(value or '').strip())
    return cleaned[:128]


def next_case_id(group_config, product: ProductConfig) -> str:
    year = timezone.localdate().year
    prefix = f'{product.case_prefix}-{year}'
    sequence, created = TatCaseSequence.objects.select_for_update().get_or_create(
        group_id=str(group_config.group_id),
        product_key=product.key,
        year=year,
        defaults={'next_number': 1},
    )
    if created:
        existing = TatTrackerCase.objects.filter(
            group_id=str(group_config.group_id), case_id__startswith=prefix,
        ).values_list('case_id', flat=True)
        maximum = 0
        for case_id in existing:
            try:
                maximum = max(maximum, int(str(case_id).rsplit('-', 1)[-1]))
            except (TypeError, ValueError):
                continue
        sequence.next_number = maximum + 1
    number = sequence.next_number
    sequence.next_number = number + 1
    sequence.save(update_fields=['next_number', 'updated_at'])
    return f'{prefix}-{number:03d}'


def _tat_google_quota_error(exc: Exception) -> bool:
    text = str(exc or '').lower()
    return any(
        marker in text
        for marker in (
            'resource_exhausted',
            'rate_limit',
            'quota exceeded',
            '429',
            'read requests per minute',
            'write requests per minute',
        )
    )


def _tat_sheet_call(operation, *, description: str):
    """Run one gspread call with bounded quota backoff for repair jobs."""
    max_attempts = max(1, int(getattr(settings, 'GOOGLE_SHEETS_MAX_RETRIES', 4) or 4))
    base_delay = max(0.25, float(getattr(settings, 'TAT_REPAIR_RETRY_BASE_SECONDS', 2) or 2))
    for attempt in range(1, max_attempts + 1):
        try:
            return operation()
        except Exception as exc:
            if not _tat_google_quota_error(exc) or attempt >= max_attempts:
                raise
            delay = min(60.0, base_delay * (2 ** (attempt - 1)))
            logger.warning(
                'TAT Sheet quota/rate limit during %s; retrying in %.1fs (%s/%s): %s',
                description,
                delay,
                attempt,
                max_attempts,
                exc,
            )
            time.sleep(delay)


def sync_case_to_sheet(group_config, case: TatTrackerCase) -> None:
    product = product_by_key(case.product_key)
    service = get_sheets_service(sheet_id=group_config.sheet_id, sheet_name=product.sheet_name)
    if not service.is_available():
        case.sync_error = 'Google Sheets service unavailable.'
        case.save(update_fields=['sync_error', 'updated_at'])
        mark_case_events_sync_failed(case, case.sync_error)
        raise RuntimeError(case.sync_error)
    sheet = service._sheet
    try:
        # TAT values are Django-calculated display columns. Keeping them out
        # of sheet formulas avoids delayed spreadsheet recalculation.
        headers = cached_tat_sheet_headers(group_config, product, sheet)
        validate_tracker_identity_headers(headers)
        # A configured register contract turns schema drift into a safe,
        # retryable sync failure before a canonical case can be written into
        # the wrong column layout. Legacy tabs remain unchanged until an
        # operator explicitly creates and enables their contract.
        from core.services.sync_governance import assert_registered_schema_before_publish

        assert_registered_schema_before_publish(group_config, product.sheet_name, headers)
        # The persisted row number is only a hint: staff can sort or insert
        # rows in Sheets. Resolve the immutable case ID in column A before
        # writing so a stale row number cannot overwrite another customer.
        case_ids = None
        existing_row = False
        row = None
        if case.row_number:
            try:
                current_id = str(_tat_sheet_call(
                    lambda: sheet.cell(case.row_number, 1).value,
                    description=f'read case ID for {case.case_id}',
                ) or '').strip()
            except (AttributeError, IndexError, KeyError):
                current_id = ''
            if current_id == str(case.case_id).strip():
                row = case.row_number
                existing_row = True
        if row is None:
            case_ids = _tat_sheet_call(
                lambda: sheet.col_values(1),
                description=f'scan case IDs for {case.case_id}',
            ) if hasattr(sheet, 'col_values') else None
            row = resolve_case_sheet_row(sheet, case, case_ids=case_ids)
            existing_row = bool(
                case_ids is None and case.row_number and not hasattr(sheet, 'col_values')
            ) or bool(
                case_ids is not None
                and 1 <= row <= len(case_ids)
                and str(case_ids[row - 1] or '').strip() == str(case.case_id).strip()
            )
        values = _tat_sheet_call(
            lambda: sheet.row_values(row),
            description=f'read row {row} for {case.case_id}',
        ) if existing_row else []
        row_data = build_tat_sheet_row_data(group_config, case, product, headers, values)
        width = len(row_data)
        if existing_row:
            _tat_sheet_call(
                lambda: sheet.update(
                    f'A{row}:{column_letter(width)}{row}',
                    [row_data],
                    value_input_option='USER_ENTERED',
                ),
                description=f'update row {row} for {case.case_id}',
            )
        else:
            row = append_case_row(sheet, row_data)
        case.row_number = row
        case.sheet_name = product.sheet_name
        synced_at = timezone.now()
        case.last_synced_at = synced_at
        case.sync_error = ''
        case.save(update_fields=['row_number', 'sheet_name', 'last_synced_at', 'sync_error', 'updated_at'])
        mark_case_events_synced(case, synced_at=synced_at)
        if should_sync_secondary_sheets(group_config):
            try:
                sync_case_index(group_config, case)
            except Exception as exc:
                logger.warning('TAT tracker CASE_INDEX sync failed for %s: %s', case.case_id, exc, exc_info=True)
    except Exception as exc:
        case.sync_error = str(exc)
        case.save(update_fields=['sync_error', 'updated_at'])
        mark_case_events_sync_failed(case, case.sync_error)
        logger.exception('TAT tracker sheet sync failed for %s', case.case_id)
        raise


def mark_case_events_synced(case: TatTrackerCase, *, synced_at=None) -> int:
    """Mark pending events once their resulting case state reaches the primary sheet."""
    return case.events.filter(synced_to_sheet=False).update(
        synced_to_sheet=True,
        synced_at=synced_at or timezone.now(),
        sheet_name=case.sheet_name,
        row_number=case.row_number,
        sync_error='',
    )


def mark_case_events_sync_failed(case: TatTrackerCase, error: str) -> int:
    """Keep pending events retryable and expose the primary-sheet failure reason."""
    return case.events.filter(synced_to_sheet=False).update(sync_error=str(error or 'Sheet sync failed.'))


def build_tat_sheet_row_data(
    group_config,
    case: TatTrackerCase,
    product: ProductConfig,
    headers: list[Any],
    existing_values: list[Any] | None = None,
) -> list[Any]:
    del group_config
    tat_columns = resolve_tat_sheet_columns(product, headers)
    width = max([product.tat_start_col + 1, *tat_columns.values()])
    row_data = [''] * width
    for idx, value in enumerate((existing_values or [])[:width], start=1):
        row_data[idx - 1] = value
    row_data[0] = case.case_id
    row_data[1] = case.client_name
    row_data[2] = case.national_id
    row_data[3] = case.primary_phone
    row_data[4] = case.branch
    row_data[5] = case.bro_name
    row_data[6] = float(case.amount or 0) if case.amount is not None else ''
    row_data[product.stage_columns['created'] - 1] = sheet_datetime(case.stage_values.get('created'))
    for stage in product.stages:
        if stage.key in case.stage_values:
            row_data[stage.column - 1] = sheet_value_for_stage(stage, case.stage_values.get(stage.key))
        if stage.auto_timestamp_key and stage.auto_timestamp_key in case.stage_values:
            col = product.stage_columns.get(stage.auto_timestamp_key)
            if col:
                row_data[col - 1] = sheet_datetime(case.stage_values.get(stage.auto_timestamp_key))
    row_data[product.status_col - 1] = case.status
    row_data[product.remarks_col - 1] = case.remarks
    tat_minutes = calculated_tat_minutes(case)
    tat_hours = calculated_tat_hours(case) if tat_minutes is not None else None
    tat_days = calculated_tat_days(case) if tat_minutes is not None else None
    row_data[product.tat_start_col - 1] = float(tat_hours) if tat_hours is not None else ''
    row_data[product.tat_start_col] = float(tat_days) if tat_days is not None else ''
    if tat_columns.get('total_minutes'):
        row_data[tat_columns['total_minutes'] - 1] = float(tat_minutes) if tat_minutes is not None else ''
    for stage in product.stages:
        col = tat_columns.get(stage.key)
        if col:
            minutes = stage_tat_minutes(case, stage)
            row_data[col - 1] = float(minutes) if minutes is not None else ''

    # Optional operational headers are populated only when the existing
    # workbook already exposes them; no columns are created or reordered.
    header_indexes = {
        normalize_header(header): index
        for index, header in enumerate(headers)
        if str(header or '').strip()
    }

    def put_optional(value, *names):
        for name in names:
            index = header_indexes.get(normalize_header(name))
            if index is not None:
                if index >= len(row_data):
                    row_data.extend([''] * (index + 1 - len(row_data)))
                row_data[index] = value if value is not None else ''
                return

    put_optional(case.product_label or product.label, 'Product', 'Product Name')
    put_optional(case.current_stage, 'Current Stage')
    put_optional(case.status, 'Status')
    put_optional(case.created_at, 'Created At')
    put_optional(case.updated_at, 'Last Updated At')
    put_optional(case.last_updated_by, 'Last Updated By')
    return row_data


def cached_tat_sheet_headers(group_config, product: ProductConfig, sheet) -> list[Any]:
    if not hasattr(sheet, 'row_values'):
        return []
    group_key = str(getattr(group_config, 'pk', '') or getattr(group_config, 'group_id', '') or '')
    cache_key = (group_key, str(group_config.sheet_id or ''), product.sheet_name)
    now = time.monotonic()
    cached = _TAT_HEADER_CACHE.get(cache_key)
    if cached and now - cached[0] < _TAT_HEADER_CACHE_TTL_SECONDS:
        return list(cached[1])
    headers = sheet.row_values(TAT_TRACKER_HEADER_ROW)
    _TAT_HEADER_CACHE[cache_key] = (now, list(headers))
    return headers


def resync_tat_tracker_cases(
    group_config,
    *,
    product_key: str = '',
    case_ids: list[str] | None = None,
    dry_run: bool = False,
    limit: int | None = None,
    offset: int = 0,
    include_unlinked: bool = False,
) -> dict[str, object]:
    """Re-write TAT cases from Django, resolving the exact case ID first.

    The normal repair scope remains linked cases only. When ``include_unlinked``
    is explicitly selected, cases with no stored Sheet row are also passed to
    ``sync_case_to_sheet``; that helper searches column A and appends only when
    the exact immutable case ID is absent.
    """
    selected_product = str(product_key or '').strip()
    if selected_product and selected_product not in PRODUCTS:
        raise ValueError(f'Unknown TAT product: {selected_product}.')

    selected_case_ids = [str(case_id).strip() for case_id in (case_ids or []) if str(case_id).strip()]
    queryset = operational_tat_cases(
        TatTrackerCase.objects.filter(group_id=str(group_config.group_id), is_deleted=False)
    )
    if selected_product:
        queryset = queryset.filter(product_key=selected_product)
    if selected_case_ids:
        queryset = queryset.filter(case_id__in=selected_case_ids)

    candidate_cases = queryset if include_unlinked else queryset.filter(row_number__gt=0)
    candidate_cases = candidate_cases.order_by('product_key', 'case_id')
    total_candidates = candidate_cases.count()
    selected_offset = max(0, int(offset or 0))
    skipped_unlinked = 0 if include_unlinked else queryset.exclude(row_number__gt=0).count()
    if limit is not None:
        candidate_cases = candidate_cases[selected_offset:selected_offset + max(0, int(limit))]
    elif selected_offset:
        candidate_cases = candidate_cases[selected_offset:]
    candidates = list(candidate_cases)
    result: dict[str, object] = {
        'total_candidates': total_candidates,
        'candidates': len(candidates),
        'synced': 0,
        'skipped_unlinked': skipped_unlinked,
        'failed': [],
        'offset': selected_offset,
        'next_offset': selected_offset + len(candidates) if selected_offset + len(candidates) < total_candidates else None,
    }
    if dry_run:
        return result

    for case in candidates:
        try:
            sync_case_to_sheet(group_config, case)
            result['synced'] = int(result['synced']) + 1
        except Exception as exc:
            logger.exception('TAT repair re-sync failed for %s', case.case_id)
            result['failed'].append({'case_id': case.case_id, 'error': str(exc)})
    return result


def resolve_tat_sheet_columns(product: ProductConfig, headers: list[Any]) -> dict[str, int]:
    normalized_headers = {
        normalize_header(header): index
        for index, header in enumerate(headers, start=1)
        if str(header or '').strip()
    }
    columns: dict[str, int] = {}
    total_col = first_matching_header(normalized_headers, ('TAT Minutes', 'Total TAT Minutes', 'Case TAT Minutes'))
    if total_col:
        columns['total_minutes'] = total_col
    for config in product.stage_tat_columns or STAGE_TAT_COLUMNS.get(product.key, ()):
        columns[config.stage_key] = first_matching_header(normalized_headers, config.aliases) or config.fallback_col
    return columns


def first_matching_header(headers: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for candidate in candidates:
        col = headers.get(normalize_header(candidate))
        if col:
            return col
    return None


def normalize_header(value: Any) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(value or '').strip().lower())


def validate_tracker_identity_headers(headers: list[Any]) -> None:
    if not any(str(header or '').strip() for header in headers):
        return
    expected = ('idnumber', 'phonenumber')
    actual = tuple(normalize_header(headers[index]) if len(headers) > index else '' for index in (2, 3))
    if actual != expected:
        raise ValueError('Tracker sheet row 2 must have ID NUMBER in column C and PHONE NUMBER in column D before cases can be synced.')


def append_case_row(sheet, row_data: list[Any]) -> int:
    result = None
    if hasattr(sheet, 'append_row'):
        result = _tat_sheet_call(
            lambda: sheet.append_row(row_data, value_input_option='USER_ENTERED'),
            description='append TAT case row',
        )
    elif hasattr(sheet, 'append_rows'):
        result = _tat_sheet_call(
            lambda: sheet.append_rows([row_data], value_input_option='USER_ENTERED'),
            description='append TAT case row',
        )
    else:
        row = next_sheet_row(sheet)
        _tat_sheet_call(
            lambda: sheet.update(
                f'A{row}:{column_letter(len(row_data))}{row}',
                [row_data],
                value_input_option='USER_ENTERED',
            ),
            description=f'write new TAT case row {row}',
        )
        return row
    row = row_number_from_update_result(result)
    if row:
        return row
    return next_sheet_row(sheet) - 1


def row_number_from_update_result(result: Any) -> int | None:
    if not isinstance(result, dict):
        return None
    updated_range = str((result.get('updates') or {}).get('updatedRange') or result.get('updatedRange') or '')
    match = re.search(r'![A-Z]+(\d+)(?::[A-Z]+(\d+))?$', updated_range)
    if not match:
        return None
    return int(match.group(1))


def signatures_enabled() -> bool:
    return bool(getattr(settings, 'TAT_TRACKER_SIGNATURES_ENABLED', False))


def should_sync_secondary_sheets(group_config) -> bool:
    workflow = getattr(group_config, 'workflow', None) or {}
    if 'sync_secondary_sheets' in workflow:
        return bool(workflow.get('sync_secondary_sheets'))
    return bool(getattr(settings, 'TAT_TRACKER_SYNC_SECONDARY_SHEETS', False))


def sync_case_index(group_config, case: TatTrackerCase) -> None:
    service = get_sheets_service(sheet_id=group_config.sheet_id, sheet_name='CASE_INDEX')
    if not service.is_available():
        return
    sheet = service._sheet
    rows = sheet.get_all_values()
    target = None
    for idx, row in enumerate(rows[1:], start=2):
        if row and row[0] == case.case_id:
            target = idx
            break
    if not target:
        target = max(len(rows) + 1, 2)
    sheet.update(f'A{target}:K{target}', [[case.case_id, case.sheet_name, case.row_number or '', case.client_name, case.national_id, case.primary_phone, case.branch, case.bro_name, case.status, sheet_datetime(case.stage_values.get('created')), timezone.localtime(timezone.now()).strftime('%d-%b-%Y %H:%M')]], value_input_option='USER_ENTERED')


def sync_audit_log(group_config, case: TatTrackerCase) -> None:
    unsynced = list(case.events.filter(synced_to_sheet=False).order_by('created_at'))
    if not unsynced:
        return
    service = get_sheets_service(sheet_id=group_config.sheet_id, sheet_name='AUDIT LOG')
    if not service.is_available():
        return
    sheet = service._sheet
    existing_count = max(len(sheet.get_all_values()), 1)
    rows = []
    for event in unsynced:
        rows.append([timezone.localtime(event.created_at).strftime('%d-%b-%Y %H:%M'), event.actor_name, case.sheet_name, case.case_id, case.row_number or '', event.stage_label or event.stage_key, event.new_value, '', event.source.upper()])
    if rows:
        start = existing_count + 1
        sheet.update(f'A{start}:I{start + len(rows) - 1}', rows, value_input_option='USER_ENTERED')
        TatTrackerEvent.objects.filter(id__in=[event.id for event in unsynced]).update(synced_to_sheet=True, synced_at=timezone.now(), sync_error='')


def calculated_tat_minutes(case: TatTrackerCase, now=None) -> Decimal | None:
    created = parse_iso_datetime((case.stage_values or {}).get('created'))
    if not created:
        return None
    end = overall_tat_end(case, now=now)
    return minutes_between(created, end)


def calculated_business_tat_minutes(case: TatTrackerCase, now=None) -> Decimal | None:
    """Optional business-hours comparison; official TAT is wall-clock time."""
    created = parse_iso_datetime((case.stage_values or {}).get('created'))
    if not created:
        return None
    return business_minutes_between(created, overall_tat_end(case, now=now))


def calculated_tat_hours(case: TatTrackerCase, now=None) -> Decimal | None:
    minutes = calculated_tat_minutes(case, now=now)
    if minutes is None:
        return None
    return (minutes / Decimal('60')).quantize(Decimal('0.01'))


def calculated_tat_days(case: TatTrackerCase, now=None) -> Decimal | None:
    minutes = calculated_tat_minutes(case, now=now)
    if minutes is None:
        return None
    return (minutes / Decimal('1440')).quantize(Decimal('0.01'))


def overall_tat_end(case: TatTrackerCase, now=None):
    values = case.stage_values or {}
    if case.status in {'Rejected', 'Declined'}:
        return parse_iso_datetime(values.get('decision_ts')) or parse_iso_datetime(values.get('decision')) or now or timezone.now()
    return parse_iso_datetime(values.get('disbursement')) or now or timezone.now()


def minutes_between(start, end) -> Decimal | None:
    if not start or not end:
        return None
    delta = end - start
    seconds = max(Decimal(str(delta.total_seconds())), Decimal('0'))
    return (seconds / Decimal('60')).quantize(Decimal('0.01'))


def stage_tat_minutes(case: TatTrackerCase, stage: StageConfig, now=None) -> Decimal | None:
    product = product_by_key(case.product_key)
    previous = previous_stage_timestamp(case, product, stage)
    if not previous:
        return None
    current = stage_completed_at(case, stage)
    if not current and next_action(case) and next_action(case).key == stage.key:
        current = now or timezone.now()
    return minutes_between(previous, current)


def stage_business_tat_minutes(case: TatTrackerCase, stage: StageConfig, now=None) -> Decimal | None:
    previous = previous_stage_timestamp(case, product_by_key(case.product_key), stage)
    if not previous:
        return None
    current = stage_completed_at(case, stage)
    if not current and next_action(case) and next_action(case).key == stage.key:
        current = now or timezone.now()
    return business_minutes_between(previous, current)


def previous_stage_timestamp(case: TatTrackerCase, product: ProductConfig, stage: StageConfig):
    previous = parse_iso_datetime((case.stage_values or {}).get('created'))
    for current in product.stages:
        if current.key == stage.key:
            return previous
        value = stage_completed_at(case, current)
        if value:
            previous = value
    return previous


def stage_completed_at(case: TatTrackerCase, stage: StageConfig):
    values = case.stage_values or {}
    timestamp = parse_iso_datetime(values.get(stage.auto_timestamp_key)) or parse_iso_datetime(values.get(stage.key))
    if timestamp:
        return timestamp
    if not stage.auto_timestamp_key or not values.get(stage.key) or not case.pk:
        return None
    event = case.events.filter(stage_key=stage.key).order_by('created_at').first()
    return event.created_at if event else None


def tat_targets_for_product(workflow: dict | None, product: ProductConfig) -> dict:
    workflow = workflow or {}
    configured = workflow.get('tat_targets_minutes') or {}
    product_targets = configured.get(product.key) or configured.get(product.sheet_name) or {}
    defaults = DEFAULT_TAT_TARGETS_MINUTES.get(product.key, {})
    return {
        'total': product_targets.get('total', defaults.get('total')),
        'stages': product_targets.get('stages') or defaults.get('stages') or {},
    }



def can_manage_tat_targets(user: dict | None) -> bool:
    """Return whether the staff member may change workflow-wide SLA targets."""
    return _tat_has_capability(user, 'tat.targets.manage')


def _tat_has_capability(user: dict | None, capability: str) -> bool:
    """Use the matrix when present; retain service-level legacy test inputs.

    HTTP callers always receive the explicit ``capabilities`` field from the
    canonical access resolver.  The fallback only preserves existing pure
    service callers that construct a historical ``{'roles': ...}`` payload.
    """
    user = user or {}
    if 'capabilities' in user:
        return capability in set(user.get('capabilities') or [])
    from core.services.workflow_capabilities import capability_definition

    definition = capability_definition('tat_tracker', capability)
    roles = {str(role).strip().upper() for role in user.get('roles') or []}
    return bool(definition and roles.intersection(definition.default_roles))


def _tat_scope_allowed(user: dict | None, capability: str, case: TatTrackerCase | None = None) -> bool:
    if not user:
        return False
    canonical_user = user.get('_canonical_user')
    access = user.get('_access')
    if canonical_user is None or access is None:
        return _tat_has_capability(user, capability)
    from core.services.workflow_access import workflow_access_decision
    return workflow_access_decision(
        canonical_user, 'tat_tracker', capability, access=access, resource=case,
        group_configuration=user.get('_group_configuration'),
    ).allowed


def _tat_scope_allowed_for_values(user: dict | None, capability: str, *, branch: str, product: str) -> bool:
    if not user:
        return False
    canonical_user = user.get('_canonical_user')
    access = user.get('_access')
    if canonical_user is None or access is None:
        return _tat_has_capability(user, capability)
    from core.services.workflow_access import workflow_access_decision
    return workflow_access_decision(
        canonical_user, 'tat_tracker', capability, access=access,
        branch=branch, product=product,
        group_configuration=user.get('_group_configuration'),
    ).allowed


def _scope_tat_queryset(queryset, user: dict, capability: str):
    canonical_user = user.get('_canonical_user')
    access = user.get('_access')
    if canonical_user is None or access is None:
        return queryset
    from core.services.workflow_access import scope_workflow_queryset
    return scope_workflow_queryset(
        queryset, canonical_user, 'tat_tracker', capability, access=access,
        branch_field='branch', product_field='product_key', group_field='group_id',
    )


def tat_target_settings(workflow: dict | None) -> list[dict]:
    """Serialize the configured targets for the administrator Mini App form."""
    settings = []
    for product in configured_products(workflow):
        targets = tat_targets_for_product(workflow, product)
        settings.append({
            'key': product.key,
            'label': product.label,
            'total_minutes': targets.get('total') or '',
            'stages': [
                {
                    'key': stage.key,
                    'label': stage.label,
                    'target_minutes': (targets.get('stages') or {}).get(stage.key) or '',
                }
                for stage in product.stages
            ],
        })
    return settings


def _target_minutes_from_value(value: object, label: str) -> int | None:
    if value in (None, ''):
        return None
    try:
        minutes = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f'{label} must be a number of minutes.') from exc
    if minutes < 0 or minutes > Decimal('5256000'):
        raise ValueError(f'{label} must be between 0 and 5,256,000 minutes.')
    if minutes != minutes.to_integral_value():
        raise ValueError(f'{label} must use whole minutes.')
    return int(minutes)

def normalize_tat_target_settings(workflow: dict | None, payload: object) -> dict:
    """Validate Mini App target minutes and store canonical minute values."""
    submitted = payload if isinstance(payload, dict) else {}
    targets: dict[str, dict] = {}
    for product in configured_products(workflow):
        row = submitted.get(product.key) or {}
        if not isinstance(row, dict):
            raise ValueError(f'{product.label} targets are invalid.')
        product_targets: dict[str, object] = {'stages': {}}
        total = _target_minutes_from_value(row.get('total_minutes'), f'{product.label} total target')
        if total is not None:
            product_targets['total'] = total
        submitted_stages = row.get('stages') or {}
        if not isinstance(submitted_stages, dict):
            raise ValueError(f'{product.label} stage targets are invalid.')
        for stage in product.stages:
            minutes = _target_minutes_from_value(
                submitted_stages.get(stage.key),
                f'{product.label}: {stage.label} target',
            )
            if minutes is not None:
                product_targets['stages'][stage.key] = minutes
        if product_targets.get('total') is not None or product_targets['stages']:
            targets[product.key] = {
                key: value for key, value in product_targets.items()
                if key != 'stages' or value
            }
    return targets


def sync_tat_target_settings_to_sheet(group_config, workflow: dict | None) -> dict:
    """Write configured SLA targets to the Apps Script support tab.

    The tab is created on the first IT target save, so formatting does not
    depend on an administrator remembering to run a separate Apps Script setup.
    """
    if not getattr(group_config, 'sheet_id', ''):
        return {'status': 'not_configured'}
    service = get_sheets_service(sheet_id=group_config.sheet_id, sheet_name='TAT TARGETS')
    sheet = service.get_or_create_worksheet('TAT TARGETS', rows=500, cols=4)
    if sheet is None:
        return {'status': 'unavailable'}
    rows = []
    for product in tat_target_settings(workflow):
        if product['total_minutes']:
            rows.append([product['key'], '__total__', product['total_minutes'], str(NEAR_SLA_RATIO)])
        for stage in product['stages']:
            if stage['target_minutes']:
                rows.append([product['key'], stage['key'], stage['target_minutes'], str(NEAR_SLA_RATIO)])
    try:
        sheet.update('A1:D1', [['Product Key', 'Stage Key', 'Target Minutes', 'Near Ratio']], value_input_option='USER_ENTERED')
        sheet.batch_clear(['A2:D500'])
        if rows:
            sheet.update(f'A2:D{len(rows) + 1}', rows, value_input_option='USER_ENTERED')
        return {'status': 'synced'}
    except Exception as exc:
        logger.warning('TAT target sheet sync failed for group %s: %s', group_config.group_id, exc)
        return {'status': 'failed'}
@transaction.atomic
def update_tat_target_settings(group_config, user: dict, payload: object) -> dict:
    """Persist administrator-managed SLA targets and refresh the group registry."""
    if not can_manage_tat_targets(user):
        raise ValueError('Only IT staff can change SLA targets.')
    from core.models import GroupSheetConfiguration
    from core.services.group_config import GroupRegistry

    config = GroupSheetConfiguration.objects.select_for_update().get(group_id=str(group_config.group_id))
    workflow = dict(config.workflow or {})
    targets = normalize_tat_target_settings(workflow, payload)
    changed = workflow.get('tat_targets_minutes') != targets
    if changed:
        workflow['tat_targets_minutes'] = targets
        config.workflow = workflow
        config.save(update_fields=['workflow', 'updated_at'])
        GroupRegistry.get_instance().reload()
    active_workflow = workflow if changed else config.workflow
    return {
        'changed': changed,
        'targets': tat_target_settings(active_workflow),
        'sheet_sync': sync_tat_target_settings_to_sheet(group_config, active_workflow),
    }
def stage_target_minutes(workflow: dict | None, product: ProductConfig, stage: StageConfig) -> Decimal | None:
    value = (tat_targets_for_product(workflow, product).get('stages') or {}).get(stage.key)
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def snapshot_stage_target(case: TatTrackerCase, workflow: dict | None, product: ProductConfig, stage: StageConfig) -> None:
    """Freeze the target for a stage at entry rather than revising history."""
    snapshots = dict(case.stage_target_snapshots or {})
    if stage.key in snapshots:
        return
    target = stage_target_minutes(workflow, product, stage)
    if target is None:
        return
    snapshots[stage.key] = {
        'target_minutes': str(target),
        'settings_version': int((workflow or {}).get('settings_version') or 1),
        'started_at': timezone.now().isoformat(),
    }
    case.stage_target_snapshots = snapshots


def stage_target_minutes_for_case(case: TatTrackerCase, workflow: dict | None, product: ProductConfig, stage: StageConfig) -> Decimal | None:
    snapshot = (case.stage_target_snapshots or {}).get(stage.key) or {}
    value = snapshot.get('target_minutes')
    if value not in (None, ''):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            pass
    return stage_target_minutes(workflow, product, stage)


def total_target_minutes(workflow: dict | None, product: ProductConfig) -> Decimal | None:
    value = tat_targets_for_product(workflow, product).get('total')
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def sla_status(minutes: Decimal | None, target: Decimal | None) -> str:
    if minutes is None or target is None or target <= 0:
        return ''
    if minutes > target:
        return 'over'
    if minutes >= (target * NEAR_SLA_RATIO):
        return 'near'
    return 'within'


def resolve_case_sheet_row(sheet, case: TatTrackerCase, *, case_ids: list[Any] | None = None) -> int:
    """Return the existing case-ID row, or the append position if absent."""
    if case_ids is None:
        if not hasattr(sheet, 'col_values'):
            # Lightweight test doubles and non-gspread adapters may not
            # expose column reads. In production, gspread provides them and
            # the case ID is always checked before a write.
            return int(case.row_number or next_sheet_row(sheet))
        case_ids = sheet.col_values(1)
    if case.row_number and case.row_number <= len(case_ids):
        current_id = str(case_ids[case.row_number - 1] or '').strip()
        if current_id == str(case.case_id).strip():
            return case.row_number
    for idx, value in enumerate(case_ids, start=1):
        if idx >= 5 and str(value or '').strip() == case.case_id:
            return idx
    return next_sheet_row(sheet, values=case_ids)

def next_sheet_row(sheet, *, values: list[Any] | None = None) -> int:
    values = values if values is not None else sheet.col_values(1)
    for idx in range(len(values), 4, -1):
        if str(values[idx - 1] or '').strip():
            return idx + 1
    return 5


def inspect_tat_sheet_duplicate_case_ids(
    sheet,
    *,
    group_id: str = '',
    data_start_row: int = 5,
) -> list[dict[str, Any]]:
    """Report duplicate case-ID rows without changing the sheet.

    The row with the most populated cells is the proposed keeper. A linked
    Django row is used only as a tie-breaker, so a sparse stale copy cannot
    displace a more complete operational record.
    """
    values = sheet.get_all_values()
    grouped: dict[str, list[dict[str, int]]] = {}
    for row_number, row in enumerate(values[data_start_row - 1:], start=data_start_row):
        case_id = str((row or [''])[0] or '').strip()
        if not case_id:
            continue
        populated = sum(1 for value in row if str(value or '').strip())
        grouped.setdefault(case_id, []).append({
            'row_number': row_number,
            'populated_cells': populated,
        })

    duplicates = []
    for case_id, rows in grouped.items():
        if len(rows) < 2:
            continue
        canonical = None
        if group_id:
            canonical = operational_tat_cases(TatTrackerCase.objects.filter(
                group_id=str(group_id), case_id=case_id, is_deleted=False,
            )).first()
        canonical_row = int(canonical.row_number or 0) if canonical else 0
        keeper = max(
            rows,
            key=lambda item: (
                item['populated_cells'],
                int(item['row_number'] == canonical_row),
                -item['row_number'],
            ),
        )
        duplicates.append({
            'case_id': case_id,
            'rows': rows,
            'keep_row': keeper['row_number'],
            'delete_rows': [item['row_number'] for item in rows if item['row_number'] != keeper['row_number']],
            'canonical_row': canonical_row or None,
            'linked': bool(canonical),
        })
    return sorted(duplicates, key=lambda item: item['case_id'])


def cleanup_tat_sheet_duplicate_case_ids(
    sheet,
    *,
    group_id: str = '',
    group_configuration=None,
    actor: str = '',
    apply: bool = False,
    include_unlinked: bool = False,
    data_start_row: int = 5,
) -> list[dict[str, Any]]:
    """Optionally delete duplicate sheet rows and verify/re-publish survivors.

    ``apply=False`` is deliberately the default because deleting rows in a
    shared Google Sheet is irreversible from Django's perspective. Deletions
    are performed bottom-up, then the live Sheet is read again by immutable
    case ID.  Django row pointers are never derived from assumed row shifts:
    a successful repair must prove one surviving row and re-publish the
    canonical case through the normal safe sync path.
    """
    reports = inspect_tat_sheet_duplicate_case_ids(
        sheet, group_id=group_id, data_start_row=data_start_row,
    )
    if not apply:
        return reports
    if not hasattr(sheet, 'delete_rows'):
        raise RuntimeError('The configured Google Sheet adapter cannot delete duplicate rows.')

    # Delete every extra row globally from the bottom upward. Deleting one
    # duplicate group before another would otherwise shift the second group's
    # recorded row numbers and could remove the wrong customer row.
    actionable = [
        report for report in reports
        if report['linked'] or include_unlinked
    ]
    for report in reports:
        if report not in actionable:
            report['skipped_unlinked'] = True
    all_delete_rows = sorted(
        {row_number for report in actionable for row_number in report['delete_rows']},
        reverse=True,
    )
    for row_number in all_delete_rows:
        sheet.delete_rows(row_number)

    if not hasattr(sheet, 'get_all_values'):
        raise RuntimeError('The configured Google Sheet adapter cannot verify rows after duplicate deletion.')

    verified_rows: dict[str, list[int]] = {}
    try:
        post_delete_values = sheet.get_all_values()
        for row_number, row in enumerate(post_delete_values[data_start_row - 1:], start=data_start_row):
            case_id = str((row or [''])[0] or '').strip()
            if case_id:
                verified_rows.setdefault(case_id, []).append(row_number)
    except Exception as exc:
        logger.exception('Could not re-read TAT Sheet after duplicate-row cleanup')
        for report in actionable:
            report['verification_status'] = 'failed'
            report['cleanup_error'] = 'The Sheet could not be re-read after deleting duplicate rows.'
        raise RuntimeError('Duplicate rows were deleted but the surviving Sheet rows could not be verified.') from exc

    # Callers that operate a live configured register (Admin/management
    # command) must pass the configuration so a verified survivor is
    # immediately re-published. Keeping this explicit prevents a low-level
    # helper from accidentally selecting an unrelated group configuration.
    resolved_config = group_configuration

    linked_cases: dict[str, TatTrackerCase] = {}
    if group_id:
        linked_cases = {
            case.case_id: case
            for case in operational_tat_cases(TatTrackerCase.objects.filter(
                group_id=str(group_id), is_deleted=False, case_id__in=[item['case_id'] for item in actionable],
            ))
        }

    for report in actionable:
        surviving_rows = verified_rows.get(report['case_id'], [])
        if len(surviving_rows) != 1:
            report['verification_status'] = 'failed'
            report['cleanup_error'] = (
                'Expected exactly one surviving row after cleanup; '
                f'found {len(surviving_rows)}.'
            )
            report['resync_status'] = 'not_attempted'
            continue
        report['surviving_row'] = surviving_rows[0]
        report['verification_status'] = 'verified'
        report['resync_status'] = 'not_applicable'

        case = linked_cases.get(report['case_id'])
        if case is not None:
            with transaction.atomic():
                locked_case = TatTrackerCase.objects.select_for_update().get(pk=case.pk)
                locked_case.row_number = surviving_rows[0]
                locked_case.sync_error = ''
                locked_case.save(update_fields=['row_number', 'sync_error', 'updated_at'])
            case = TatTrackerCase.objects.get(pk=case.pk)
            linked_cases[report['case_id']] = case
            if resolved_config is None:
                report['resync_status'] = 'not_requested'
            else:
                try:
                    sync_case_to_sheet(resolved_config, case)
                    report['resync_status'] = 'synced'
                except Exception:
                    logger.exception('Could not re-publish verified TAT case %s after duplicate cleanup', case.case_id)
                    report['resync_status'] = 'failed'
                    report['cleanup_error'] = 'The surviving row was verified, but re-publication failed. Check the case sync error and server logs.'

    # Keep a local append-only record of an external destructive operation and
    # its verification outcome.  This intentionally records case IDs/rows and
    # status, never the spreadsheet row's customer values.
    try:
        from core.models import LiveSheetRecordChange

        for report in actionable:
            failure = report.get('verification_status') != 'verified' or report.get('resync_status') == 'failed'
            case = linked_cases.get(report['case_id'])
            LiveSheetRecordChange.objects.create(
                group_configuration=resolved_config,
                group_id=str(group_id),
                sheet_id=str(getattr(resolved_config, 'sheet_id', '') or ''),
                sheet_tab=str(getattr(case, 'sheet_name', '') or getattr(resolved_config, 'sheet_name', '') or ''),
                row_number=int(report.get('surviving_row') or report.get('keep_row') or data_start_row),
                record_key=report['case_id'],
                action='delete',
                changed_by=str(actor or ''),
                changes={
                    'operation': 'tat_duplicate_row_cleanup',
                    'removed_rows': report['delete_rows'],
                    'surviving_row': report.get('surviving_row'),
                    'verification_status': report.get('verification_status', ''),
                    'resync_status': report.get('resync_status', ''),
                },
                status='failed' if failure else 'success',
                error=str(report.get('cleanup_error') or ''),
            )
    except Exception:
        # A failed audit write must not incorrectly turn a verified external
        # repair into a failed repair. It remains visible in server logs.
        logger.exception('Could not write duplicate-row cleanup audit evidence')
    return reports


def next_action(case: TatTrackerCase) -> StageConfig | None:
    product = product_by_key(case.product_key)
    if case.status in {'Disbursed', 'Rejected', 'Declined'}:
        return None
    for stage in product.stages:
        if not case.stage_values.get(stage.key):
            return stage
    return None


def previous_stages_complete(case: TatTrackerCase, stage: StageConfig) -> bool:
    product = product_by_key(case.product_key)
    for current in product.stages:
        if current.key == stage.key:
            return True
        if not case.stage_values.get(current.key):
            return False
        if signatures_enabled() and current.requires_signature_certificate and not case.approval_certificates.filter(stage_key=current.key, status='signed').exists():
            return False
    return True


def can_user_edit_stage(user: dict, case: TatTrackerCase, stage: StageConfig) -> bool:
    return _tat_scope_allowed(user, f'tat.stage.{stage.key}.update', case)


def can_user_correct_stage(user: dict, case: TatTrackerCase, stage: StageConfig) -> bool:
    """Allow explicit corrections without weakening normal stage sequencing."""
    if stage.requires_signature_certificate:
        return False
    return can_user_edit_stage(user, case, stage)


def can_user_correct_case_details(user: dict, case: TatTrackerCase | None = None) -> bool:
    """Base identity/amount corrections are restricted and audited."""
    return _tat_scope_allowed(user, 'tat.case.correct', case)


def serialize_case_summary(case: TatTrackerCase, user: dict | None = None, next_stage: StageConfig | None = None, workflow: dict | None = None) -> dict:
    next_stage = next_stage or next_action(case)
    product = product_by_key(case.product_key)
    tat_minutes = calculated_tat_minutes(case)
    business_minutes = calculated_business_tat_minutes(case)
    tat_hours = calculated_tat_hours(case) if tat_minutes is not None else None
    tat_days = calculated_tat_days(case) if tat_minutes is not None else None
    total_target = total_target_minutes(workflow, product)
    certificates = {certificate.stage_key: certificate.status for certificate in case.approval_certificates.all()}
    read_only = not is_record_operational(case)
    return {'case_id': case.case_id, 'product': case.product_label or product.label, 'product_key': case.product_key, 'client_name': case.client_name, 'national_id': case.national_id, 'primary_phone': case.primary_phone, 'branch': case.branch, 'bro_name': case.bro_name, 'amount': str(case.amount or ''), 'status': case.status, 'current_stage': case.current_stage, 'workflow_revision': int(case.workflow_revision or 1), 'next_stage': next_stage.label if next_stage and not read_only else '', 'next_stage_key': next_stage.key if next_stage and not read_only else '', 'tat_minutes': str(tat_minutes) if tat_minutes is not None else '', 'wall_clock_minutes': str(tat_minutes) if tat_minutes is not None else '', 'business_minutes': str(business_minutes) if business_minutes is not None else '', 'sla_minutes': str(tat_minutes) if tat_minutes is not None else '', 'tat_hours': str(tat_hours) if tat_hours is not None else '', 'tat_days': str(tat_days) if tat_days is not None else '', 'target_minutes': str(total_target) if total_target is not None else '', 'sla_status': sla_status(tat_minutes, total_target), 'certificate_statuses': certificates, 'updated_at': format_datetime(case.updated_at), 'created_at': format_datetime(case.created_at), 'data_mode': case.data_mode, 'is_pilot': case.data_mode == 'pilot', 'read_only': read_only, 'pilot_cycle_id': str(case.pilot_cycle_id or '')}


def serialize_case_detail(case: TatTrackerCase, user: dict, workflow: dict | None = None) -> dict:
    product = product_by_key(case.product_key)
    read_only = not is_record_operational(case)
    can_correct_details = (not read_only) and can_user_correct_case_details(user, case)
    fields = []
    for stage in product.stages:
        value = case.stage_values.get(stage.key, '')
        editable = (not read_only) and previous_stages_complete(case, stage) and can_user_edit_stage(user, case, stage) and (not value or stage.kind == 'dropdown')
        tat_minutes = stage_tat_minutes(case, stage)
        business_minutes = stage_business_tat_minutes(case, stage)
        target = stage_target_minutes_for_case(case, workflow, product, stage)
        certificate = case.approval_certificates.filter(stage_key=stage.key).first() if stage.requires_signature_certificate else None
        fields.append({'key': stage.key, 'label': stage.label, 'kind': stage.kind, 'value': display_stage_value(stage, value), 'raw_value': str(value or ''), 'editable': editable, 'can_correct': bool(value) and can_user_correct_stage(user, case, stage), 'options': list(stage.options), 'role': stage.role, 'locked_reason': '' if editable or (value and can_user_correct_stage(user, case, stage)) else lock_reason(case, user, stage), 'tat_minutes': str(tat_minutes) if tat_minutes is not None else '', 'wall_clock_minutes': str(tat_minutes) if tat_minutes is not None else '', 'business_minutes': str(business_minutes) if business_minutes is not None else '', 'sla_minutes': str(tat_minutes) if tat_minutes is not None else '', 'target_minutes': str(target) if target is not None else '', 'sla_status': sla_status(tat_minutes, target), 'certificate_status': certificate.status if certificate else ''})
    timeline = tat_case_timeline(case)
    legacy_events = [
        {
            'at': format_datetime(event.created_at),
            'actor': event.actor_name,
            'stage': event.stage_label,
            'value': event.new_value,
            'source': event.source,
        }
        for event in case.events.order_by('-created_at')[:20]
    ]
    requirements = []
    if case.product_version_id:
        requirements = [
            {
                'key': item.key, 'label': item.label, 'description': item.description,
                'type': item.requirement_type, 'stage': item.enforcement_stage,
                'required': item.required,
                'value': (case.product_requirement_evidence or {}).get(item.key),
            }
            for item in case.product_version.requirements.filter(
                active=True,
            ).filter(Q(workflow='') | Q(workflow='tat_tracker'))
        ]
    return {
        'summary': serialize_case_summary(case, user, workflow=workflow),
        'fields': fields,
        'remarks': case.remarks,
        'events': legacy_events,
        'timeline': timeline['entries'],
        'escalation': latest_escalation('tat_tracker', str(case.pk)),
        'can_correct_details': can_correct_details,
        'correction_branches': _allowed_branches(workflow or {}, user) if can_correct_details else [],
        'product_terms': case.product_terms_snapshot,
        'product_quote': case.product_quote_snapshot,
        'product_requirements': requirements,
        'product_custom_attributes': [
            {
                'key': item.key, 'label': item.label, 'type': item.attribute_type,
                'required': item.required, 'help_text': item.help_text,
                'options': item.options, 'value': (case.product_custom_values or {}).get(item.key),
            }
            for item in case.product_version.custom_attributes.all()
            if not item.workflow_visibility or 'tat_tracker' in item.workflow_visibility
        ] if case.product_version_id else [],
        'product_custom_values': case.product_custom_values,
        'product_selected_fee_keys': case.product_selected_fee_keys,
    }


def next_role_alert(group_config, case_data: dict | None) -> dict[str, str]:
    if not case_data:
        return {}
    workflow = getattr(group_config, 'workflow', None) or {}
    if workflow.get('stage_alerts_enabled') is False:
        return {}
    summary = case_data.get('summary') or {}
    next_stage_key = summary.get('next_stage_key') or ''
    if not next_stage_key:
        return {}
    try:
        product = product_by_key(summary.get('product_key') or '')
    except ValueError:
        return {}
    stage = stage_by_key(product, next_stage_key)
    if not stage:
        return {}
    role_label = role_display_name(stage.role)
    return {
        'role': stage.role,
        'role_label': role_label,
        'stage': stage.label,
        'text': (
            f"TAT action needed: {role_label}\n\n"
            f"Case: {summary.get('case_id', '')}\n"
            f"Client: {summary.get('client_name', '')}\n"
            f"Branch: {summary.get('branch', '')}\n"
            f"Next step: {stage.label}\n\n"
            "Please open the TAT Tracker and update this stage when done."
        ),
    }


def role_display_name(role: str) -> str:
    labels = {
        'BRO': 'BRO',
        BUSINESS_ADMIN_ROLE: 'Business Admin',
        'CA': 'Credit Analyst',
        'BM': 'Branch Manager',
        'SECRETARY': 'Secretary',
        'CHAIR': 'Chair',
        'LOAN_APPROVER': 'Loan Approver',
        'FINANCE': 'Finance',
    }
    return labels.get(str(role or '').strip().upper(), str(role or '').strip() or 'Responsible team')


def lock_reason(case: TatTrackerCase, user: dict, stage: StageConfig) -> str:
    if not previous_stages_complete(case, stage):
        return 'Previous stage is not complete.'
    if not can_user_edit_stage(user, case, stage):
        return 'Not assigned to your role.'
    if case.stage_values.get(stage.key) and stage.kind != 'dropdown':
        return 'Already completed.'
    return ''


def public_user(user: dict) -> dict:
    return {
        'name': user.get('name', ''), 'roles': user.get('roles') or [],
        'telegram_id': user.get('telegram_id', ''), 'username': user.get('username', ''),
        'capabilities': user.get('capabilities') or [],
    }


def serialize_product(product: ProductConfig) -> dict:
    payload = {
        'id': product.product_id, 'version_id': product.version_id,
        'key': product.key, 'label': product.label, 'sheet_name': product.sheet_name,
        'min_amount': str(product.min_amount),
        'max_amount': str(product.max_amount) if product.max_amount is not None else '',
    }
    if product.version_id:
        from core.models import ProductVersion
        from core.services.product_catalog import serialize_product_version
        version = ProductVersion.objects.filter(pk=product.version_id).first()
        payload['terms'] = serialize_product_version(version) if version else {}
    else:
        payload['terms'] = {}
    return payload


def _allowed_products(workflow: dict, user: dict) -> list[ProductConfig]:
    products = configured_products(workflow)
    allowed = set(user.get('products') or [])
    upper = {item.upper() for item in allowed}
    if not allowed or 'ALL' in upper or '*' in allowed:
        return products
    return [p for p in products if p.key in allowed or p.sheet_name in allowed]


def _allowed_branches(workflow: dict, user: dict) -> list[str]:
    branches = workflow_branches(workflow)
    allowed = user.get('branches') or []
    upper = {item.upper() for item in allowed}
    if not allowed or 'ALL' in upper or '*' in allowed:
        return branches
    return [branch for branch in branches if branch in allowed]


def product_by_key(key: str) -> ProductConfig:
    normalized = str(key or '').strip().lower().replace('-', '_')
    try:
        return _database_product_by_key(normalized)
    except Exception:
        if normalized not in PRODUCTS:
            raise ValueError('Invalid product.')
        return PRODUCTS[normalized]


def stage_by_key(product: ProductConfig, key: str) -> StageConfig | None:
    return next((stage for stage in product.stages if stage.key == key), None)


def parse_amount(value) -> Decimal:
    try:
        return Decimal(str(value or '').replace(',', '').strip())
    except (InvalidOperation, ValueError):
        raise ValueError('Enter a valid amount.')


def validate_amount(product: ProductConfig, amount: Decimal) -> None:
    if amount < product.min_amount:
        raise ValueError(f'{product.label} amount must be at least KES {product.min_amount:,.0f}.')
    if product.max_amount is not None and amount > product.max_amount:
        raise ValueError(f'{product.label} amount must be at most KES {product.max_amount:,.0f}.')


def display_stage_value(stage: StageConfig, value: Any) -> str:
    if not value:
        return ''
    if stage.kind == 'timestamp':
        return format_datetime(parse_iso_datetime(value))
    if stage.kind == 'dropdown':
        return dropdown_display_value(stage, value)
    return str(value)


def sheet_value_for_stage(stage: StageConfig, value: Any) -> str:
    if not value:
        return ''
    if stage.kind == 'timestamp':
        return sheet_datetime(value)
    if stage.kind == 'dropdown':
        return dropdown_display_value(stage, value)
    return str(value)


def dropdown_display_value(stage: StageConfig, value: Any) -> str:
    raw = str(value or '').strip()
    if stage.key == 'minutes_shared' and parse_iso_datetime(raw):
        return 'Yes'
    if stage.key == 'bro_applied' and parse_iso_datetime(raw):
        return 'Met'
    return raw


def parse_iso_datetime(value: Any):
    if hasattr(value, 'isoformat'):
        return value
    try:
        parsed = timezone.datetime.fromisoformat(str(value))
        if timezone.is_naive(parsed):
            parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
        return parsed
    except Exception:
        return None


def sheet_datetime(value: Any) -> str:
    parsed = parse_iso_datetime(value)
    if not parsed:
        return str(value or '')
    return timezone.localtime(parsed).strftime('%d-%b-%Y %H:%M')


def format_datetime(value) -> str:
    if not value:
        return ''
    if isinstance(value, str):
        value = parse_iso_datetime(value)
    if not value:
        return ''
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.strftime('%d-%b-%Y %H:%M')


def tat_hours_formula(product: ProductConfig, row: int) -> str:
    created_col = column_letter(product.stage_columns['created'])
    disbursement_stage = stage_by_key(product, 'disbursement')
    if not disbursement_stage:
        return ''
    end_col = column_letter(disbursement_stage.column)
    return f'=IF(OR(${created_col}{row}="",${end_col}{row}=""),"",ROUND((${end_col}{row}-${created_col}{row})*24,2))'


def tat_days_formula(product: ProductConfig, row: int) -> str:
    hours_col = column_letter(product.tat_start_col)
    return f'=IF({hours_col}{row}="","",ROUND({hours_col}{row}/24,2))'
def column_letter(index: int) -> str:
    letters = ''
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _telegram_name(user_payload: dict) -> str:
    parts = [str(user_payload.get('first_name') or '').strip(), str(user_payload.get('last_name') or '').strip()]
    return ' '.join(part for part in parts if part).strip()


def _normalize_list(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [item.strip() for item in value.split(',') if item.strip()]
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    return [str(value).strip()] if str(value).strip() else []
