"""Staged import of the Customers Without Loans system export.

The export is an authoritative source for system identity and JBL-side
financial fields, but it must never create a pipeline case by itself. Rows are
therefore matched to existing Jawabu records and held for staff review before
any database write occurs.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from core.models import JawabuCustomer, JawabuFarmerMaster, JawabuFarmerUploadBatch
from core.services.jawabu import is_valid_phone, normalise_phone
from core.services.jawabu_customer_quality import (
    product_quality_message,
    record_customer_phone,
    record_field_provenance,
    resolve_farmer_match,
)
from core.services.jawabu_master import clean_text, row_fingerprint

logger = logging.getLogger(__name__)


REQUIRED_HEADERS = (
    'Customer ID', 'Name', 'Mobile No', 'ID NO', 'Branch',
    'Loan Officer', 'Product Name', 'LGF Balance',
)
DISPLAY_HEADERS = REQUIRED_HEADERS


def _header_key(value: Any) -> str:
    return re.sub(r'[^a-z0-9]+', ' ', str(value or '').strip().lower()).strip()


REQUIRED_HEADER_KEYS = {_header_key(value): value for value in REQUIRED_HEADERS}


def _cell_text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r'\s+', ' ', str(value).strip())


def _normalise_id(value: Any) -> str:
    return re.sub(r'\D', '', _cell_text(value))


def _normalise_name(value: Any) -> str:
    # Keep the comma-separated IMAB value for display. For low-confidence
    # candidate discovery, compare name tokens without relying on first/last
    # order; names remain manual-review candidates and are never auto-matched.
    tokens = re.sub(r'[^a-z0-9]+', ' ', _cell_text(value).casefold()).split()
    return ' '.join(sorted(tokens))


def _normalise_system_row(raw: dict[str, Any], source_row: int) -> dict[str, Any]:
    values = {REQUIRED_HEADER_KEYS[_header_key(key)]: value for key, value in raw.items() if _header_key(key) in REQUIRED_HEADER_KEYS}
    national_id = _normalise_id(values.get('ID NO'))
    customer_no = _normalise_id(values.get('Customer ID'))
    phone_raw = _cell_text(values.get('Mobile No'))
    phone = normalise_phone(phone_raw)
    name = _cell_text(values.get('Name'))
    branch = _cell_text(values.get('Branch'))
    loan_officer = _cell_text(values.get('Loan Officer'))
    product = _cell_text(values.get('Product Name'))
    lgf_raw = _cell_text(values.get('LGF Balance'))
    notes = []
    if not national_id:
        notes.append('Missing ID NO')
    elif not re.fullmatch(r'\d{7,9}', national_id):
        notes.append('ID NO should contain 7-9 digits; confirm this exception before commit')
    if phone_raw and not is_valid_phone(phone):
        notes.append('Mobile No could not be normalized to a valid 254 phone')
    if not customer_no:
        notes.append('Missing Customer ID')
    if not name:
        notes.append('Missing Name')
    product_note = product_quality_message(product)
    if product_note:
        notes.append(product_note)
    lgf_balance = None
    if lgf_raw:
        try:
            lgf_balance = Decimal(re.sub(r'[^0-9.\-]', '', lgf_raw))
            if lgf_balance < 0:
                notes.append('LGF Balance cannot be negative')
        except (InvalidOperation, ValueError):
            notes.append('LGF Balance must be numeric')
    return {
        **{header: _cell_text(values.get(header)) for header in REQUIRED_HEADERS},
        'Customer ID': customer_no,
        'ID NO': national_id,
        'Mobile No': phone,
        'LGF Balance': str(lgf_balance) if lgf_balance is not None else '',
        'Source Row': source_row,
        'Import Status': 'review_needed' if notes else 'ready',
        'Match Basis': '',
        'Matched Farmer ID': '',
        'Matched Customer': '',
        'Match Candidates': [],
        'Cleaning Notes': '; '.join(notes),
        'approved': False,
        'row_fingerprint': row_fingerprint(raw),
    }


def _read_csv(content: bytes) -> tuple[list[tuple[int, dict[str, Any]]], list[str]]:
    text = ''
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252'):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        raise ValueError('Could not read the CSV encoding. Export it as UTF-8 CSV and retry.')
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        raise ValueError('The system export is empty.')
    header_lookup = {_header_key(header): header for header in headers}
    missing = [header for key, header in REQUIRED_HEADER_KEYS.items() if key not in header_lookup]
    if missing:
        raise ValueError('System export is missing required headers: ' + ', '.join(missing))
    rows = []
    for row_number, values in enumerate(reader, start=2):
        if not any(str(value or '').strip() for value in values):
            continue
        padded = list(values) + [''] * max(0, len(headers) - len(values))
        rows.append((row_number, dict(zip(headers, padded))))
    return rows, headers


def _read_xlsx(content: bytes) -> tuple[list[tuple[int, dict[str, Any]]], list[str]]:
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('Could not read the Excel workbook. Save it as .xlsx and retry.') from exc
    try:
        sheet = workbook.active
        header_row_number = None
        headers = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
            candidate = [_cell_text(value) for value in row]
            keys = {_header_key(value) for value in candidate if value}
            if set(REQUIRED_HEADER_KEYS).issubset(keys):
                header_row_number = row_number
                headers = candidate
                break
        if header_row_number is None:
            raise ValueError('System export headers were not found in the first 25 rows.')
        rows = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
            values = list(row)
            if not any(_cell_text(value) for value in values):
                continue
            rows.append((row_number, dict(zip(headers, values))))
        return rows, headers
    finally:
        workbook.close()


def _candidate_snapshot(farmer: JawabuFarmerMaster) -> dict[str, str]:
    return {
        'id': str(farmer.id),
        'customer_name': farmer.customer_name,
        'customer_no': farmer.customer_no,
        'national_id': farmer.national_id,
        'primary_phone': farmer.primary_phone,
    }


def resolve_system_export_row(row: dict[str, Any]) -> dict[str, Any]:
    national_id = str(row.get('ID NO') or '').strip()
    customer_no = str(row.get('Customer ID') or '').strip()
    phone = str(row.get('Mobile No') or '').strip()
    match = resolve_farmer_match(
        national_id=national_id,
        customer_no=customer_no,
        primary_phone=phone,
        name=row.get('Name', ''),
    )
    candidate_ids = list(match.farmer_ids or match.name_candidates)
    candidates = [
        _candidate_snapshot(farmer)
        for farmer in JawabuFarmerMaster.objects.filter(pk__in=candidate_ids)
    ]
    farmer = JawabuFarmerMaster.objects.filter(pk=match.exact_farmer_id).first() if match.exact_farmer_id else None
    if match.conflicts:
        row['Import Status'] = 'review_needed'
        row['approved'] = False
        row['Cleaning Notes'] = '; '.join(filter(None, [row.get('Cleaning Notes', ''), *match.conflicts]))
    elif farmer:
        row['Matched Farmer ID'] = str(farmer.id)
        row['Matched Customer'] = farmer.customer_name
        if row.get('Cleaning Notes'):
            row['Import Status'] = 'review_needed'
            row['approved'] = False
        else:
            row['Import Status'] = 'ready'
            row['approved'] = True
    else:
        row['Import Status'] = 'review_needed'
        row['approved'] = False
        row['Cleaning Notes'] = '; '.join(filter(None, [row.get('Cleaning Notes', ''), 'No exact identity match found']))
    row['Match Basis'] = match.match_basis
    row['Match Candidates'] = candidates
    return row


def parse_system_export(content: bytes, filename: str = '') -> tuple[list[dict[str, Any]], dict[str, Any]]:
    lower = str(filename or '').lower()
    if lower.endswith('.xlsx'):
        source_rows, headers = _read_xlsx(content)
    elif lower.endswith('.csv') or not lower:
        source_rows, headers = _read_csv(content)
    else:
        raise ValueError('The /sysup command only supports .csv or .xlsx system exports.')
    rows = [resolve_system_export_row(_normalise_system_row(raw, row_number)) for row_number, raw in source_rows]
    return rows, {
        'headers': headers,
        'total_rows': len(rows),
        'review_needed': sum(1 for row in rows if row.get('Import Status') == 'review_needed'),
        'skipped_blank': 0,
    }


def create_system_export_review_batch(*, group_id: str, telegram_message_id: str, sender: str, source_filename: str, content: bytes) -> tuple[JawabuFarmerUploadBatch, dict[str, Any]]:
    rows, stats = parse_system_export(content, source_filename)
    batch = JawabuFarmerUploadBatch.objects.create(
        group_id=str(group_id),
        telegram_message_id=str(telegram_message_id or ''),
        sender=str(sender or ''),
        source_filename=str(source_filename or 'system-export.csv'),
        import_kind='system_export',
        total_rows=stats['total_rows'],
        review_needed=stats['review_needed'],
        parsed_rows=rows,
        mapping=[{'source_column': header, 'target': header} for header in DISPLAY_HEADERS],
    )
    return batch, stats


def _mark_review(row: dict[str, Any], message: str) -> None:
    row['Import Status'] = 'review_needed'
    row['approved'] = False
    existing = [part.strip() for part in str(row.get('Cleaning Notes', '') or '').split(';') if part.strip()]
    if message and message not in existing:
        existing.append(message)
    row['Cleaning Notes'] = '; '.join(existing)


def _commit_values(row: dict[str, Any]) -> tuple[dict[str, str | Decimal], list[str]]:
    """Normalize editable review values again before any database write."""
    national_id = _normalise_id(row.get('ID NO'))
    customer_no = _normalise_id(row.get('Customer ID'))
    phone = normalise_phone(_cell_text(row.get('Mobile No')))
    name = clean_text(row.get('Name'))
    branch = clean_text(row.get('Branch'))
    loan_officer = clean_text(row.get('Loan Officer'))
    product = clean_text(row.get('Product Name'))
    lgf_raw = _cell_text(row.get('LGF Balance'))
    errors = []
    if not national_id:
        errors.append('Missing ID NO')
    # A numeric historical/exceptional ID is review-only.  A staff member
    # must make the row approved in the staged review before it reaches this
    # method; preserving it lets the canonical data-quality queue retain the
    # exception instead of encouraging a made-up replacement ID.
    if phone and not is_valid_phone(phone):
        errors.append('Mobile No could not be normalized to a valid 254 phone')
    if not customer_no:
        errors.append('Missing Customer ID')
    if not name:
        errors.append('Missing Name')
    product_error = product_quality_message(product)
    if product_error:
        errors.append(product_error)
    lgf: Decimal | str = ''
    if lgf_raw:
        try:
            lgf = Decimal(re.sub(r'[^0-9.\-]', '', lgf_raw))
            if not lgf.is_finite() or lgf < 0:
                raise InvalidOperation
        except (InvalidOperation, ValueError):
            errors.append('LGF Balance must be numeric and non-negative')
            lgf = ''
    return {
        'national_id': national_id,
        'customer_no': customer_no,
        'phone': phone,
        'name': name,
        'branch': branch,
        'loan_officer': loan_officer,
        'product': product,
        'lgf': lgf,
    }, errors


def _bind_customer_identity(farmer: JawabuFarmerMaster, *, national_id: str, phone: str, customer_no: str) -> None:
    """Keep the canonical JawabuCustomer identity aligned with the farmer row."""
    customer = farmer.customer
    if customer is None:
        identity_filter = Q()
        if national_id:
            identity_filter |= Q(national_id=national_id)
        if phone:
            identity_filter |= Q(primary_phone=phone)
        if customer_no:
            identity_filter |= Q(customer_no=customer_no)
        customer = JawabuCustomer.objects.select_for_update().filter(identity_filter).first() if identity_filter.children else None
        if customer is None:
            customer = JawabuCustomer.objects.create(
                national_id=national_id,
                primary_phone=phone,
                customer_no=customer_no,
                identity_enforced=True,
            )
        farmer.customer = customer
    if national_id and not customer.national_id:
        customer.national_id = national_id
    if phone and not customer.primary_phone:
        customer.primary_phone = phone
    if customer_no:
        customer.customer_no = customer_no
    customer.save(update_fields=['national_id', 'primary_phone', 'customer_no', 'updated_at'])
    record_customer_phone(customer, phone, source='system_export')


def _identifier_belongs_to_a_different_customer(
    farmer: JawabuFarmerMaster, *, field_name: str, value: str,
) -> bool:
    """Reject cross-customer collisions while allowing a customer's extra unit.

    One canonical JawabuCustomer may legitimately have several applications.
    A staff member must still select the intended application for a reviewed
    `/sysup` row, but that selection must not be blocked by its sibling unit.
    """
    if not value:
        return False
    matches = JawabuFarmerMaster.objects.filter(**{field_name: value}).exclude(pk=farmer.pk)
    if farmer.customer_id:
        matches = matches.exclude(customer_id=farmer.customer_id)
    return matches.exists()


@transaction.atomic
def commit_system_export_review_batch(batch: JawabuFarmerUploadBatch, rows: list[dict[str, Any]], *, actor: str = '') -> dict[str, Any]:
    if batch.status == 'committed':
        return {'success': True, 'message': 'This batch has already been committed. No duplicate write was made.', 'committed': 0, 'skipped': 0, 'review_needed': 0, 'errors': []}
    committed = skipped = 0
    errors = []
    remaining = []
    for index, row in enumerate(rows, start=1):
        row = dict(row or {})
        if not row.get('approved'):
            skipped += 1
            remaining.append(row)
            continue
        values, validation_errors = _commit_values(row)
        if validation_errors:
            _mark_review(row, '; '.join(validation_errors))
            errors.append(f"Row {index}: {'; '.join(validation_errors)}")
            remaining.append(row)
            continue
        farmer_id = str(row.get('Matched Farmer ID') or '').strip()
        farmer = JawabuFarmerMaster.objects.select_for_update().filter(pk=farmer_id).first()
        if not farmer:
            _mark_review(row, 'Select a valid existing customer before committing this row.')
            errors.append(f'Row {index}: no existing customer selected.')
            remaining.append(row)
            continue
        national_id = str(values['national_id'])
        phone = str(values['phone'])
        customer_no = str(values['customer_no'])
        conflicts = []
        for label, field, value in (
            ('ID NO', 'national_id', national_id),
            ('Mobile No', 'primary_phone', phone),
            ('Customer ID', 'customer_no', customer_no),
        ):
            existing = str(getattr(farmer, field) or '').strip()
            if value and existing and value != existing:
                conflicts.append(f'{label} differs from the selected customer')
        if _identifier_belongs_to_a_different_customer(
            farmer, field_name='customer_no', value=customer_no,
        ):
            conflicts.append('Customer ID already belongs to another case')
        if _identifier_belongs_to_a_different_customer(
            farmer, field_name='national_id', value=national_id,
        ):
            conflicts.append('ID NO already belongs to another case')
        if _identifier_belongs_to_a_different_customer(
            farmer, field_name='primary_phone', value=phone,
        ):
            conflicts.append('Mobile No already belongs to another case')
        customer_scope = JawabuCustomer.objects.exclude(pk=farmer.customer_id) if farmer.customer_id else JawabuCustomer.objects.all()
        if customer_no and customer_scope.filter(customer_no=customer_no).exists():
            conflicts.append('Customer ID already belongs to another canonical customer')
        if national_id and customer_scope.filter(national_id=national_id).exists():
            conflicts.append('ID NO already belongs to another canonical customer')
        if phone and customer_scope.filter(primary_phone=phone).exists():
            conflicts.append('Mobile No already belongs to another canonical customer')
        if farmer.customer_id:
            customer = farmer.customer
            for label, field, value in (
                ('ID NO', 'national_id', national_id),
                ('Mobile No', 'primary_phone', phone),
                ('Customer ID', 'customer_no', customer_no),
            ):
                existing = str(getattr(customer, field) or '').strip()
                if value and existing and value != existing:
                    conflicts.append(f'{label} differs from the selected canonical customer')
        if conflicts:
            _mark_review(row, '; '.join(conflicts))
            errors.append(f'Row {index}: ' + '; '.join(conflicts))
            remaining.append(row)
            continue
        old_values = {
            'national_id': farmer.national_id,
            'primary_phone': farmer.primary_phone,
            'customer_no': farmer.customer_no,
            'imab_customer_name': farmer.imab_customer_name,
            'system_branch': farmer.system_branch,
            'branch': farmer.branch,
            'system_loan_officer': farmer.system_loan_officer,
            'payment_product': farmer.payment_product,
            'system_deposit_paid_jbl': str(farmer.system_deposit_paid_jbl) if farmer.system_deposit_paid_jbl is not None else '',
        }
        if national_id and not farmer.national_id:
            farmer.national_id = national_id
        if phone and not farmer.primary_phone:
            farmer.primary_phone = phone
        if customer_no:
            farmer.customer_no = customer_no
        _bind_customer_identity(
            farmer,
            national_id=national_id,
            phone=phone,
            customer_no=customer_no,
        )
        exported_name = str(values['name'])
        if exported_name:
            farmer.imab_customer_name = exported_name
        branch = str(values['branch'])
        if branch:
            farmer.system_branch = branch
            farmer.branch = branch
        loan_officer = str(values['loan_officer'])
        if loan_officer:
            farmer.system_loan_officer = loan_officer
        product = str(values['product'])
        if product:
            farmer.payment_product = product
        if values['lgf'] != '':
            farmer.system_deposit_paid_jbl = values['lgf']
        farmer.save()
        from core.services.jawabu_validation import refresh_data_quality_issues
        refresh_data_quality_issues(farmer)
        from core.services.jawabu_case360 import record_pipeline_event
        new_values = {
            'national_id': farmer.national_id,
            'primary_phone': farmer.primary_phone,
            'customer_no': farmer.customer_no,
            'imab_customer_name': farmer.imab_customer_name,
            'system_branch': farmer.system_branch,
            'branch': farmer.branch,
            'system_loan_officer': farmer.system_loan_officer,
            'payment_product': farmer.payment_product,
            'system_deposit_paid_jbl': str(farmer.system_deposit_paid_jbl) if farmer.system_deposit_paid_jbl is not None else '',
        }
        material_changes = {
            field for field, old_value in old_values.items()
            if str(new_values.get(field, '')) != str(old_value or '')
        }
        if material_changes:
            from core.services.jawabu_approvals import invalidate_material_approvals
            invalidate_material_approvals(
                farmer=farmer,
                changed_fields=material_changes,
                reason='A controlled system export changed material customer, branch, product, or financial data.',
            )
        record_pipeline_event(
            farmer,
            action='system_export_updated',
            stage_key='intake',
            actor=actor,
            source='system_export',
            metadata={'source_filename': batch.source_filename, 'source_row': row.get('Source Row'), 'match_basis': row.get('Match Basis', '')},
            old_values=old_values,
            new_values=new_values,
        )
        record_field_provenance(
            farmer,
            old_values=old_values,
            new_values=new_values,
            source='system_export',
            source_reference=batch.source_filename,
            source_row_number=int(row.get('Source Row') or 0) or None,
            actor=actor,
        )
        # The system export is a Django-owned correction/update path. Publish
        # the committed identity and JBL financial values to any configured
        # operational registers, while keeping Sheet failures non-fatal to the
        # canonical database transaction.
        try:
            from core.services.jawabu_pipeline import (
                sync_farmer_to_internal_order_sheet,
                sync_farmer_to_master_sheet,
            )

            sync_farmer_to_master_sheet(farmer)
            sync_farmer_to_internal_order_sheet(farmer)
        except Exception as exc:  # pragma: no cover - publishers already fail safely
            # A transient Drive/Sheets failure must not undo an accepted
            # system-export update; the publisher logs its own failure.
            logger.warning('System-export publication failed for farmer %s: %s', farmer.id, exc, exc_info=True)
        committed += 1
    batch.parsed_rows = remaining
    batch.committed_count += committed
    batch.skipped_count = skipped
    batch.review_needed = len(remaining)
    batch.status = 'committed' if not remaining and not errors else 'pending_review'
    batch.error = '\n'.join(errors[:20])
    if batch.status == 'committed':
        batch.committed_at = timezone.now()
    batch.save()
    return {'success': not errors, 'message': 'System export committed.' if not errors else 'Some rows still need review.', 'committed': committed, 'skipped': skipped, 'review_needed': len(remaining), 'errors': errors[:20]}
