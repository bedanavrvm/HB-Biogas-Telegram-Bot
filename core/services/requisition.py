import re
import copy
import io
import os
import math
import openpyxl
from datetime import date, datetime
from typing import Any
from django.conf import settings
from core.models import JawabuFarmerMaster
from core.services.template_storage import TemplateStorageError, workbook_source_from_template


class RequisitionTemplateError(RuntimeError):
    pass

def clean_deposit_float(val: Any) -> float | int | None:
    if not val:
        return None
    try:
        cleaned = re.sub(r'[^\d.]', '', str(val))
        if not cleaned:
            return None
        if '.' in cleaned:
            return float(cleaned)
        return int(cleaned)
    except ValueError:
        return None


def requisition_deposit_values(farmer: JawabuFarmerMaster) -> tuple[Any, Any]:
    """Return the HBG and JBL deposits for an order row.

    Jawabu records use ``JAWABU`` as the lead-source label, while older
    records used ``JBL``.  Both identify the JBL-funded path.  Prefer the
    explicit IMAB JBL deposit when it exists; otherwise preserve the legacy
    receipt value as the deposit for that source.  Keeping this decision in
    one helper prevents the preview, totals, and workbook from disagreeing.
    """
    if farmer.deposit_paid_hbg is not None:
        canonical_deposit = farmer.deposit_paid_hbg
    elif farmer.payment is not None:
        # Older invoice matches populated only ``payment``. The invoice's
        # Payment column is the operational HBG deposit, so use it before the
        # older pre-invoice receipt field when no canonical value exists.
        canonical_deposit = farmer.payment
    else:
        canonical_deposit = clean_deposit_float(farmer.actual_receipts)
    explicit_jbl = getattr(farmer, 'system_deposit_paid_jbl', None)
    source = str(getattr(farmer, 'lead_source', '') or '').strip().lower()
    is_jbl_source = 'jbl' in source or 'jawabu' in source
    if is_jbl_source:
        return None, explicit_jbl if explicit_jbl is not None else canonical_deposit
    return canonical_deposit, explicit_jbl

def copy_row_formatting(ws: Any, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def _center_written_cell(cell: Any, *, wrap: bool = False) -> None:
    alignment = copy.copy(cell.alignment)
    cell.alignment = openpyxl.styles.Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=alignment.text_rotation,
        wrap_text=wrap or alignment.wrap_text,
        shrink_to_fit=alignment.shrink_to_fit,
        indent=alignment.indent,
    )


def _write_system_value(
    ws: Any, row: int, column: int, value: Any, *,
    style_from: Any = None, bold: bool = False, wrap: bool = False,
) -> None:
    cell = ws.cell(row=row, column=column, value=value)
    if style_from is not None:
        cell.font = copy.copy(style_from.font)
    if bold:
        base_font = copy.copy(cell.font)
        base_font.bold = True
        cell.font = base_font
    _center_written_cell(cell, wrap=wrap)


def requisition_location_text(farmer: Any) -> str:
    constituency = str(getattr(farmer, 'sub_county', '') or '').strip()
    village = str(getattr(farmer, 'village', '') or '').strip()
    return ' - '.join(part for part in (constituency, village) if part)

def generate_requisition_excel(farmers: list[JawabuFarmerMaster], order_number: str, requisition_date: date) -> bytes:
    import os
    from django.conf import settings
    from core.models import RequisitionTemplate

    active_template = RequisitionTemplate.objects.filter(
        is_active=True,
    ).order_by('-updated_at', '-created_at').first()
    fallback_path = os.path.join(settings.BASE_DIR, 'requisition', 'JBL_Requisition_Form_184.xlsx')
    try:
        template_source = workbook_source_from_template(active_template, fallback_path=fallback_path)
    except TemplateStorageError as exc:
        raise RequisitionTemplateError(
            'No requisition Excel template is available. Upload the template in '
            'Django Admin > Requisition templates and confirm it was stored in Google Drive.'
        ) from exc

    from core.services.template_validation import (
        UnsafeTemplateError, requisition_worksheet, template_source_bytes,
        validate_template_bytes,
    )
    template_bytes = template_source_bytes(template_source)
    try:
        validate_template_bytes(template_bytes, 'requisition')
    except UnsafeTemplateError as exc:
        raise RequisitionTemplateError(str(exc)) from exc
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    try:
        ws = requisition_worksheet(wb)
    except UnsafeTemplateError as exc:
        raise RequisitionTemplateError(str(exc)) from exc
    wb.active = wb.index(ws)
    
    # 1. Search for date and order ref cell placeholders in rows 1 to 7
    date_cell = None
    order_ref_cell = None
    for r in range(1, 8):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            val = str(cell.value or "").strip()
            val_clean = val.upper().replace(" ", "")
            if "DATE:" in val_clean:
                date_cell = cell
            elif "ORDERNO:" in val_clean or "BATCH/ORDERREF:" in val_clean or "ORDERREF:" in val_clean:
                order_ref_cell = cell

    date_str = requisition_date.strftime('%d-%B-%Y') if isinstance(requisition_date, (date, datetime)) else str(requisition_date)
    if date_cell:
        date_cell.value = str(date_cell.value or '').split(':')[0] + ':'
        _write_system_value(ws, date_cell.row, date_cell.column + 1, date_str, style_from=date_cell, bold=True)
    else:
        ws['A4'] = "Date:"
        _write_system_value(ws, 4, 2, date_str, style_from=ws['A4'], bold=True)

    if order_ref_cell:
        order_ref_cell.value = str(order_ref_cell.value or '').split(':')[0] + ':'
        _write_system_value(ws, order_ref_cell.row, order_ref_cell.column + 1, order_number, style_from=order_ref_cell, bold=True)
    else:
        ws['H4'] = "Order No:"
        _write_system_value(ws, 4, 9, order_number, style_from=ws['H4'], bold=True)
    
    # 2. Find main header row
    header_row_idx = None
    for r in range(1, 40):
        for c in range(1, 20):
            val = ws.cell(row=r, column=c).value
            if val and "NAME OF THE CUSTOMER" in str(val).upper():
                header_row_idx = r
                break
        if header_row_idx:
            break

    if not header_row_idx:
        header_row_idx = 7

    # 3. Map columns dynamically using word boundaries
    col_mapping = {}
    
    def get_words(text):
        return set(re.findall(r'\b\w+\b', text.upper()))
        
    for c in range(1, 16):
        h_val = ws.cell(row=header_row_idx, column=c).value
        sub_val = ws.cell(row=header_row_idx + 1, column=c).value
        
        h_text = str(h_val or "").strip()
        sub_text = str(sub_val or "").strip()
        
        # Combined words
        words = get_words(h_text) | get_words(sub_text)
        
        if h_text.upper() in ("NO.", "NO"):
            col_mapping["no"] = c
        elif "NAME" in words and "CUSTOMER" in words:
            col_mapping["name"] = c
        elif "CONTACT" in words or "PHONE" in words:
            col_mapping["phone"] = c
        elif "ID" in words and "PAID" not in words:
            col_mapping["id"] = c
        elif "CREDIT" in words:
            col_mapping["credit"] = c
        elif "CALLUP" in words or ("CALL" in words and "COMMENT" in words):
            col_mapping["callup"] = c
        elif "COUNTY" in words:
            col_mapping["county"] = c
        elif "LANDMARK" in words or "LOCATION" in words:
            col_mapping["landmark"] = c
        elif "HBG" in words:
            col_mapping["hbg"] = c
        elif "JBL" in words:
            col_mapping["jbl"] = c
        elif "SALES" in words:
            col_mapping["sales"] = c

    col_no = col_mapping.get("no", 1)
    col_name = col_mapping.get("name", 2)
    col_phone = col_mapping.get("phone", 3)
    col_id = col_mapping.get("id", 4)
    col_credit = col_mapping.get("credit", 5)
    col_callup = col_mapping.get("callup", 6)
    col_county = col_mapping.get("county", 7)
    col_landmark = col_mapping.get("landmark", 8)
    col_hbg = col_mapping.get("hbg", 9)
    col_jbl = col_mapping.get("jbl", 10)
    col_sales = col_mapping.get("sales", 11)

    # 4. Dynamically count pre-filled data rows in template
    first_data_row = None
    last_data_row = None
    for r in range(header_row_idx + 2, 50):
        val = ws.cell(row=r, column=col_no).value
        is_num = False
        if val is not None:
            try:
                int(str(val).strip())
                is_num = True
            except ValueError:
                pass
                
        if is_num:
            if first_data_row is None:
                first_data_row = r
            last_data_row = r
        else:
            if first_data_row is not None:
                break
                
    if first_data_row is None:
        first_data_row = 10
        last_data_row = 14
        
    template_data_rows = last_data_row - first_data_row + 1
    N = len(farmers)
    
    totals_row_idx = last_data_row + 1
    preserved_images = list(getattr(ws, '_images', []) or [])
    
    # 5. Adjust row count dynamically
    if N > template_data_rows:
        insert_count = N - template_data_rows
        ws.insert_rows(totals_row_idx, insert_count)
        for r in range(totals_row_idx, totals_row_idx + insert_count):
            copy_row_formatting(ws, first_data_row, r)
    elif N < template_data_rows:
        delete_count = template_data_rows - N
        ws.delete_rows(first_data_row + N, delete_count)
        
    # 6. Write the data
    for idx, farmer in enumerate(farmers):
        r = first_data_row + idx
        _write_system_value(ws, r, col_no, idx + 1)  # NO.
        _write_system_value(ws, r, col_name, farmer.customer_name, wrap=True)  # NAME OF THE CUSTOMER
        _write_system_value(ws, r, col_phone, farmer.primary_phone)  # CONTACT NO.
        _write_system_value(ws, r, col_id, farmer.national_id)  # ID NO.
        _write_system_value(ws, r, col_credit, farmer.credit_decision)  # CREDIT ANALYSIS
        callup_comment = str(getattr(farmer, 'final_decision_comment', '') or '').strip()
        _write_system_value(ws, r, col_callup, callup_comment, wrap=True)  # HEAD OF URBAN CALLUP COMMENT
        _write_system_value(ws, r, col_county, farmer.county, wrap=True)  # COUNTY
        location_text = requisition_location_text(farmer)
        _write_system_value(ws, r, col_landmark, location_text, wrap=True)  # LOCATION & NEAREST LANDMARK
        
        hbg_deposit, jbl_deposit = requisition_deposit_values(farmer)
        _write_system_value(ws, r, col_hbg, hbg_deposit if hbg_deposit is not None else "")  # HBG
        _write_system_value(ws, r, col_jbl, jbl_deposit if jbl_deposit is not None else "")  # JBL
        ws.cell(row=r, column=col_hbg).number_format = '0'
        ws.cell(row=r, column=col_jbl).number_format = '0'
            
        _write_system_value(ws, r, col_sales, farmer.hb_sales_person, wrap=True)  # HB SALES PERSON

        line_count = max(
            1,
            math.ceil(len(str(farmer.customer_name or '')) / 24),
            math.ceil(len(str(farmer.county or '')) / 14),
            math.ceil(len(location_text) / 30),
            math.ceil(len(callup_comment) / 24),
            math.ceil(len(str(farmer.hb_sales_person or '')) / 20),
        )
        current_height = float(ws.row_dimensions[r].height or 18)
        ws.row_dimensions[r].height = max(current_height, min(60, 15 * line_count))

    # 7. Remove the template totals row; staff requested no total-customer/deposit summary row.
    new_totals_row = first_data_row + N
    ws.delete_rows(new_totals_row, 1)
    if preserved_images:
        ws._images = preserved_images
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
