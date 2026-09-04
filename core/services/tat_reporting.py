"""Scoped, read-only TAT reporting and audited exports.

Timing semantics depend on the current workflow contract: an authoritative
stage stamp completes that stage and starts the next one. There is therefore
no separate pickup/handoff-lag measure. If the workflow later introduces a
distinct claimed or in-progress state, duration and handoff reporting must be
revisited against that new persisted timestamp rather than inferred here.
Admin corrections currently replace an authoritative completed stamp
immediately; there is no separate unresolved-correction lifecycle. Correction
rate therefore uses distinct completed case-stage actions as its denominator
and counts one or many correction events for that action once in its numerator.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_CEILING
from io import BytesIO
from math import ceil, sqrt

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from core.models import (
    AccessGrant, GroupSheetConfiguration, TatTrackerCase, WorkflowTatDailyMetric,
    WorkflowTatMetricRebuildRequest, TatResponsibilityAssignment,
)
from core.services.tat_presentation import presentation_settings
from core.services.tat_tracker import (
    TAT_COMPLETED_STATUSES, calculated_tat_seconds, minutes_between, next_action,
    overall_tat_end, parse_iso_datetime, product_for_case, stage_completed_at,
    stage_target_minutes_for_case, stage_tat_minutes,
)
from core.services.workflow_data_mode import operational_tat_cases


SORT_FIELDS = {
    'case_id', 'client_name', 'group', 'branch', 'product_label', 'status', 'current_stage',
    'responsible_role', 'created_at', 'finished_at', 'elapsed_minutes',
    'target_minutes', 'variance_minutes', 'sla_state',
}
TERMINAL = set(TAT_COMPLETED_STATUSES)
CHART_DIMENSIONS = frozenset({'stage', 'role', 'branch', 'product'})
CHART_METRICS = frozenset({
    'workload', 'sla_state', 'duration', 'target_usage', 'sla_met',
    'correction_rate', 'load_per_assignee',
})
HEATMAP_PAIRS = {
    'stage_branch': ('stage', 'branch'),
    'product_stage': ('product', 'stage'),
    'role_branch': ('role', 'branch'),
}
HEATMAP_METRICS = frozenset({'workload', 'sla_met', 'duration', 'target_usage'})
TARGET_REVIEW_SIGNAL_POLICY = {
    'confidence_level': 0.95,
    'wilson_z': 1.959963984540054,
    'systemic_min_samples': 20,
    'systemic_min_over_percent': 70,
    'systemic_wilson_lower_bound': 50,
    'cohort_min_samples': 10,
    'cohort_min_over_percent': 60,
    'cohort_coverage_percent': 60,
    'localized_min_samples': 10,
    'localized_min_over_percent': 70,
    'localized_wilson_lower_bound': 50,
}
# A localized signal deliberately accepts fewer samples than the systemic
# signal: it prompts review of one directly inspectable scope and never makes
# an organization-wide claim or changes configuration.


def _scope_query(actor):
    if actor.is_active and actor.is_superuser:
        return Q()
    grants = AccessGrant.objects.filter(user=actor, workflow='tat_tracker', active=True).select_related('group_configuration')
    query = Q(pk__in=[])
    for grant in grants:
        part = Q()
        scoped = False
        if grant.group_configuration_id:
            part &= Q(group_id=str(grant.group_configuration.group_id))
            scoped = True
        if grant.branch:
            part &= Q(branch__iexact=grant.branch)
            scoped = True
        if grant.product:
            part &= Q(product_key__iexact=grant.product)
            scoped = True
        if not scoped:
            return Q()
        query |= part
    return query


def scoped_cases(actor):
    return operational_tat_cases(
        TatTrackerCase.objects.filter(is_deleted=False).filter(_scope_query(actor))
    ).select_related('product', 'product_version').prefetch_related('events')


def _parse_date(value, *, default):
    try:
        return date.fromisoformat(str(value or ''))
    except ValueError:
        return default


def _casefold_distinct_labels(values):
    """Return clean, deterministically sorted labels without casing duplicates."""
    labels = {}
    for value in values:
        candidate = ' '.join(str(value or '').split())
        if not candidate:
            continue
        key = candidate.casefold()
        existing = labels.get(key)
        # Prefer normal display casing over legacy all-upper/all-lower values.
        if existing is None or (
            (existing.isupper() or existing.islower())
            and not (candidate.isupper() or candidate.islower())
        ):
            labels[key] = candidate
    return sorted(labels.values(), key=str.casefold)


def _iso_local_date(value):
    parsed = datetime.fromisoformat(value)
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return timezone.localdate(parsed)


def _filters(payload):
    today = timezone.localdate()
    date_to = _parse_date(payload.get('date_to'), default=today)
    date_from = _parse_date(payload.get('date_from'), default=date_to - timedelta(days=29))
    if date_from > date_to:
        raise ValueError('Start date must be on or before end date.')
    if (date_to - date_from).days > 3652:
        raise ValueError('Choose a reporting period of 10 years or less.')
    granularity = str(payload.get('granularity') or 'month').lower()
    if granularity not in {'day', 'week', 'month', 'year'}:
        raise ValueError('Choose Day, Week, Month, or Year grouping.')
    bucket_estimate = {
        'day': (date_to - date_from).days + 1,
        'week': ceil(((date_to - date_from).days + 1) / 7),
        'month': (date_to.year - date_from.year) * 12 + date_to.month - date_from.month + 1,
        'year': date_to.year - date_from.year + 1,
    }[granularity]
    if bucket_estimate > 366:
        raise ValueError('This grouping would create more than 366 chart points. Choose a coarser time grouping.')
    chart_dimension = str(payload.get('chart_dimension') or 'stage').strip().lower()
    chart_metric = str(payload.get('chart_metric') or 'workload').strip().lower()
    heatmap_pair = str(payload.get('heatmap_pair') or 'stage_branch').strip().lower()
    heatmap_metric = str(payload.get('heatmap_metric') or 'sla_met').strip().lower()
    if chart_dimension not in CHART_DIMENSIONS:
        raise ValueError('Choose Stage, Role, Branch, or Product for the chart dimension.')
    if chart_metric not in CHART_METRICS:
        raise ValueError('The selected comparison metric is not supported.')
    if heatmap_pair not in HEATMAP_PAIRS:
        raise ValueError('The selected heatmap comparison is not supported.')
    if heatmap_metric not in HEATMAP_METRICS:
        raise ValueError('The selected heatmap metric is not supported.')
    return {
        'view': 'performance' if payload.get('view') == 'performance' else 'current',
        'group': str(payload.get('group') or '').strip(),
        'branch': str(payload.get('branch') or '').strip(),
        'product': str(payload.get('product') or '').strip(),
        'stage': str(payload.get('stage') or '').strip(),
        'role': str(payload.get('role') or '').strip().upper(),
        'status': str(payload.get('status') or '').strip(),
        'sla_state': str(payload.get('sla_state') or '').strip(),
        'search': str(payload.get('search') or '').strip(),
        'date_from': date_from, 'date_to': date_to, 'granularity': granularity,
        'chart_dimension': chart_dimension, 'chart_metric': chart_metric,
        'heatmap_pair': heatmap_pair, 'heatmap_metric': heatmap_metric,
    }


def _filtered_cases(actor, filters):
    qs = scoped_cases(actor)
    if filters['group']:
        qs = qs.filter(group_id=filters['group'])
    if filters['branch']:
        qs = qs.filter(branch__iexact=filters['branch'])
    if filters['product']:
        qs = qs.filter(product_key__iexact=filters['product'])
    if filters['status']:
        qs = qs.filter(status__iexact=filters['status'])
    if filters['search']:
        term = filters['search']
        qs = qs.filter(Q(case_id__icontains=term) | Q(client_name__icontains=term) | Q(branch__icontains=term) | Q(product_label__icontains=term))
    return list(qs)


def _case_target(case, product, stage=None, *, config=None):
    if stage is not None:
        config = config or GroupSheetConfiguration.objects.filter(group_id=case.group_id).first()
        return stage_target_minutes_for_case(case, (config.workflow if config else {}), product, stage)
    values = []
    for item in (case.stage_target_snapshots or {}).values():
        try:
            value = Decimal(str((item or {}).get('target_minutes')))
        except Exception:
            continue
        if value > 0:
            values.append(value)
    return sum(values, Decimal('0')) if values else None


def _case_row(case, *, include_people=False, now=None):
    now = now or timezone.now()
    product = product_for_case(case)
    config = GroupSheetConfiguration.objects.filter(group_id=case.group_id).only('display_name', 'workflow').first()
    stage = next_action(case)
    display_stage = stage
    if display_stage is None:
        completed_stages = [item for item in product.stages if stage_completed_at(case, item)]
        display_stage = completed_stages[-1] if completed_stages else None
    finished_at = overall_tat_end(case, now=case.updated_at) if case.status in TERMINAL else None
    if stage:
        elapsed = stage_tat_minutes(case, stage, now=now)
        target = _case_target(case, product, stage, config=config)
        role = stage.role
        stage_label = stage.label
    else:
        seconds = calculated_tat_seconds(case, now=finished_at or now)
        elapsed = Decimal(seconds) / Decimal('60') if seconds is not None else None
        target = _case_target(case, product, config=config)
        role = display_stage.role if display_stage else ''
        stage_label = display_stage.label if display_stage else (case.current_stage or ('Finished' if case.status in TERMINAL else ''))
    variance = elapsed - target if elapsed is not None and target is not None else None
    if target is None or target <= 0 or elapsed is None:
        sla_state = 'target_unavailable'
    elif elapsed > target:
        sla_state = 'overdue'
    else:
        ratio = Decimal(presentation_settings()['near_target_percent']) / Decimal('100')
        sla_state = 'near_target' if elapsed >= target * ratio else 'within_target'
    row = {
        'case_id': case.case_id, 'client_name': case.client_name,
        'group': (config.display_name if config else '') or 'TAT Tracker',
        'branch': case.branch or '', 'product_label': case.product_label or case.product_key,
        'status': case.status, 'current_stage': stage_label, 'responsible_role': role,
        'current_stage_key': display_stage.key if display_stage else str(case.current_stage or ''),
        'created_at': case.created_at.isoformat(),
        'finished_at': finished_at.isoformat() if finished_at else '',
        'elapsed_minutes': float(elapsed) if elapsed is not None else None,
        'target_minutes': float(target) if target is not None else None,
        'variance_minutes': float(variance) if variance is not None else None,
        'sla_state': sla_state,
        '_group_id': str(case.group_id), '_product_key': str(case.product_key or ''),
    }
    if include_people:
        if stage:
            assignment = TatResponsibilityAssignment.objects.filter(
                group_configuration__group_id=case.group_id, active=True,
                branch__iexact=case.branch, role__iexact=stage.role,
            ).filter(Q(product_key='') | Q(product_key__iexact=case.product_key)).filter(
                Q(stage_key='') | Q(stage_key=stage.key),
            ).select_related('primary_user').order_by('-stage_key', '-product_key').first()
            if assignment and assignment.primary_user:
                row['responsible_person'] = assignment.primary_user.get_full_name() or assignment.primary_user.get_username()
            else:
                row['responsible_person'] = ''
        else:
            stage_key = display_stage.key if display_stage else case.current_stage
            event = next((event for event in case.events.all() if event.stage_key == stage_key), None)
            row['responsible_person'] = event.actor_name if event else ''
    return row


def _eligible_rows(actor, filters, *, include_people=False):
    rows = []
    action_filtered = bool(
        filters['view'] == 'performance'
        and (filters['stage'] or filters['role'] or filters['sla_state'])
    )
    for case in _filtered_cases(actor, filters):
        row = _case_row(case, include_people=include_people)
        if filters['view'] == 'current' and case.status in TERMINAL:
            continue
        if filters['view'] == 'performance':
            if not action_filtered:
                if case.status not in TERMINAL or not row['finished_at']:
                    continue
                finished_date = _iso_local_date(row['finished_at'])
                if not filters['date_from'] <= finished_date <= filters['date_to']:
                    continue
        if filters['view'] == 'performance' and (filters['stage'] or filters['role']):
            samples = _stage_samples([case], filters, include_people=include_people)
            if not samples:
                continue
            sample = samples[0]
            row.update(current_stage_key=sample['stage_key'], current_stage=sample['stage'], responsible_role=sample['role'], elapsed_minutes=sample['elapsed_minutes'], target_minutes=sample['target_minutes'], variance_minutes=sample['variance_minutes'], sla_state=sample['sla_state'])
            if include_people:
                row['responsible_person'] = sample['person']
        elif action_filtered:
            samples = _stage_samples([case], filters, include_people=include_people)
            if not samples:
                continue
            sample = samples[0]
            row.update(current_stage_key=sample['stage_key'], current_stage=sample['stage'], responsible_role=sample['role'], elapsed_minutes=sample['elapsed_minutes'], target_minutes=sample['target_minutes'], variance_minutes=sample['variance_minutes'], sla_state=sample['sla_state'])
            if include_people:
                row['responsible_person'] = sample['person']
        else:
            if filters['stage'] and filters['stage'].casefold() != str(row['current_stage_key']).casefold():
                continue
            if filters['role'] and row['responsible_role'].upper() != filters['role']:
                continue
        if filters['sla_state'] and row['sla_state'] != filters['sla_state']:
            continue
        rows.append(row)
    return rows


def _stage_samples(cases, filters, *, include_people=False):
    samples = []
    for case in cases:
        try:
            product = product_for_case(case)
        except ValueError:
            continue
        config = GroupSheetConfiguration.objects.filter(group_id=case.group_id).first()
        workflow = config.workflow if config else {}
        for stage in product.stages:
            if filters['stage'] and stage.key.casefold() != filters['stage'].casefold():
                continue
            if filters['role'] and stage.role.upper() != filters['role']:
                continue
            completed = stage_completed_at(case, stage)
            if not completed:
                continue
            completed_date = timezone.localdate(completed)
            if not filters['date_from'] <= completed_date <= filters['date_to']:
                continue
            elapsed = stage_tat_minutes(case, stage, now=completed)
            target = stage_target_minutes_for_case(case, workflow, product, stage)
            variance = elapsed - target if elapsed is not None and target is not None else None
            if not target or target <= 0 or elapsed is None:
                sla_state = 'target_unavailable'
            elif elapsed > target:
                sla_state = 'overdue'
            else:
                ratio = Decimal(presentation_settings()['near_target_percent']) / Decimal('100')
                sla_state = 'near_target' if elapsed >= target * ratio else 'within_target'
            if filters['sla_state'] and sla_state != filters['sla_state']:
                continue
            event = next((item for item in case.events.all() if item.stage_key == stage.key), None)
            samples.append({
                'case_id': case.case_id, 'stage_key': stage.key, 'stage': stage.label,
                'role': stage.role, 'person': event.actor_name if include_people and event else '',
                'group_id': str(case.group_id), 'branch': str(case.branch or ''),
                'product_key': str(case.product_key or ''),
                'product': str(case.product_label or case.product_key or ''),
                'corrected': any(
                    item.stage_key == stage.key and item.source == 'admin_correction'
                    for item in case.events.all()
                ),
                'completed_at': completed.isoformat(),
                'elapsed_minutes': float(elapsed) if elapsed is not None else None,
                'target_minutes': float(target) if target is not None else None,
                'variance_minutes': float(variance) if variance is not None else None,
                'sla_state': sla_state,
            })
    return samples


def _percentile(values, percentile):
    values = sorted(Decimal(str(value)) for value in values if value is not None)
    if not values:
        return None
    index = max(0, ceil(len(values) * percentile) - 1)
    return float(values[index].quantize(Decimal('0.01')))


def _bucket_label(value, granularity):
    if granularity == 'day':
        return value.isoformat()
    if granularity == 'week':
        return (value - timedelta(days=value.weekday())).isoformat()
    if granularity == 'year':
        return date(value.year, 1, 1).isoformat()
    return date(value.year, value.month, 1).isoformat()


def _chart_payload(
    chart_id, title, basis, subtitle, labels, series, *, applied_filters,
    unavailable_filters=None, sample_count=0, excluded_count=0, exclusion_reason='',
    extras=None,
):
    payload = {
        'id': chart_id,
        'title': title,
        'basis': basis,
        'subtitle': subtitle,
        'applied_filters': sorted(set(applied_filters)),
        'unavailable_filters': sorted(set(unavailable_filters or [])),
        'sample_count': sample_count,
        'excluded_count': excluded_count,
        'exclusion_reason': exclusion_reason,
        'labels': list(labels),
        'series': list(series),
    }
    if extras:
        payload.update(extras)
    return payload


def _active_filter_names(filters, *, include_dates=True):
    names = [
        key for key in ('group', 'branch', 'product', 'stage', 'role', 'status', 'sla_state', 'search')
        if filters.get(key)
    ]
    if include_dates:
        names.extend(['date_range', 'granularity'])
    return names


_REPORT_SCOPE_FILTERS = (
    'search', 'branch', 'product', 'stage', 'role', 'status', 'sla_state',
)
_REPORT_DATE_FILTERS = ('date_from', 'date_to')
_REPORT_ALL_CONTROLS = _REPORT_SCOPE_FILTERS + _REPORT_DATE_FILTERS + (
    'granularity', 'chart_dimension', 'chart_metric', 'heatmap_pair', 'heatmap_metric',
)
_REPORT_CONTROL_ONLY_REASONS = {
    'chart_dimension': 'This control only configures Operational Comparison.',
    'chart_metric': 'This control only configures Operational Comparison.',
    'heatmap_pair': 'This control only configures the Operational Heatmap.',
    'heatmap_metric': 'This control only configures the Operational Heatmap.',
}


def _filter_guidance(
    *, applicable_filters, chart_controls=(), basis_changing_filters=(),
    unavailable_reasons=None, filter_notes=None,
):
    """Describe report-control behaviour using stable frontend field names."""
    applicable_filters = list(dict.fromkeys(applicable_filters))
    chart_controls = list(dict.fromkeys(chart_controls))
    basis_changing_filters = list(dict.fromkeys(basis_changing_filters))
    unavailable_reasons = dict(unavailable_reasons or {})
    used = set(applicable_filters) | set(chart_controls) | set(basis_changing_filters)
    for key in _REPORT_ALL_CONTROLS:
        if key not in used:
            unavailable_reasons.setdefault(
                key,
                _REPORT_CONTROL_ONLY_REASONS.get(key, 'This filter does not affect this insight.'),
            )
    return {
        'applicable_filters': applicable_filters,
        'chart_controls': chart_controls,
        'basis_changing_filters': basis_changing_filters,
        'unavailable_filters': [
            {'key': key, 'reason': reason}
            for key, reason in unavailable_reasons.items()
        ],
        'filter_notes': dict(filter_notes or {}),
    }


def _attach_report_filter_guidance(common, charts, filters):
    """Attach the authoritative filter-to-insight contract to every slide."""
    scope = list(_REPORT_SCOPE_FILTERS)
    dated_scope = scope + list(_REPORT_DATE_FILTERS)
    current_only_reasons = {
        'date_from': 'This insight shows the current workload, not a historical period.',
        'date_to': 'This insight shows the current workload, not a historical period.',
        'granularity': 'Time grouping only affects insights plotted over time.',
    }
    trend = charts.get('trend')
    if trend:
        trend['filter_guidance'] = _filter_guidance(
            applicable_filters=dated_scope,
            chart_controls=('granularity',),
            basis_changing_filters=(
                ('search', 'stage', 'role', 'status', 'sla_state')
                if filters['view'] == 'current' else ()
            ),
            filter_notes={
                key: 'Selecting this changes the current workload trend to completed actions over time.'
                for key in ('search', 'stage', 'role', 'status', 'sla_state')
            } if filters['view'] == 'current' else None,
        )
    backlog = charts.get('backlog_age')
    if backlog:
        backlog['filter_guidance'] = _filter_guidance(
            applicable_filters=scope,
            unavailable_reasons=current_only_reasons,
        )
    for key in ('sla_compliance', 'tat_percentiles', 'stage_target'):
        if charts.get(key):
            charts[key]['filter_guidance'] = _filter_guidance(
                applicable_filters=dated_scope,
                chart_controls=('granularity',) if key != 'stage_target' else (),
                unavailable_reasons=(
                    {'granularity': 'This comparison is grouped by stage rather than time.'}
                    if key == 'stage_target' else None
                ),
            )

    explorer = charts.get('explorer')
    if explorer:
        sample_based = filters['view'] == 'performance' or filters['chart_metric'] in {
            'duration', 'target_usage', 'sla_met', 'correction_rate',
        }
        explorer['filter_guidance'] = _filter_guidance(
            applicable_filters=dated_scope if sample_based else scope,
            chart_controls=('chart_dimension', 'chart_metric'),
            unavailable_reasons=None if sample_based else current_only_reasons,
        )

    heatmap = common.get('heatmap')
    if heatmap:
        sample_based = filters['view'] == 'performance' or filters['heatmap_metric'] != 'workload'
        heatmap['filter_guidance'] = _filter_guidance(
            applicable_filters=dated_scope if sample_based else scope,
            chart_controls=('heatmap_pair', 'heatmap_metric'),
            unavailable_reasons=None if sample_based else current_only_reasons,
        )

    signals = common.get('target_review_signals')
    if signals:
        signals['filter_guidance'] = _filter_guidance(
            applicable_filters=dated_scope,
            unavailable_reasons={
                'granularity': 'Review signals use the selected period but are not grouped over time.',
            },
            filter_notes={
                'branch': 'Changes the selected-scope result; the systemic baseline remains organization-wide.',
                'product': 'Changes the selected-scope result; the systemic baseline remains organization-wide.',
            },
        )
    oldest = common.get('oldest_cases')
    if oldest:
        oldest['filter_guidance'] = _filter_guidance(
            applicable_filters=scope,
            unavailable_reasons=current_only_reasons,
        )


def _series(key, label, values):
    return {'key': key, 'label': label, 'values': list(values)}


def _breakdown_chart(chart_id, title, basis, subtitle, rows, *, applied_filters, unavailable_filters=None):
    return _chart_payload(
        chart_id, title, basis, subtitle,
        [row['label'] for row in rows],
        [_series('count', 'Actions' if 'action' in basis else 'Cases', [row['count'] for row in rows])],
        applied_filters=applied_filters,
        unavailable_filters=unavailable_filters,
        sample_count=sum(row['count'] for row in rows),
    )


def _dimension_value(item, dimension, *, sample=False):
    keys = {
        'stage': 'stage' if sample else 'current_stage',
        'role': 'role' if sample else 'responsible_role',
        'branch': 'branch',
        'product': 'product' if sample else 'product_label',
    }
    return str(item.get(keys[dimension]) or 'Unassigned')


def _active_assignment_counts(rows, dimension):
    now = timezone.now()
    assignments = list(TatResponsibilityAssignment.objects.filter(
        active=True, effective_from__lte=now,
        group_configuration__group_id__in={row['_group_id'] for row in rows},
    ).filter(Q(effective_until__isnull=True) | Q(effective_until__gt=now)).select_related('group_configuration'))
    grouped_rows = defaultdict(list)
    for row in rows:
        grouped_rows[_dimension_value(row, dimension)].append(row)
    result = {}
    for label, cohort in grouped_rows.items():
        users = set()
        for assignment in assignments:
            for row in cohort:
                if str(assignment.group_configuration.group_id) != row['_group_id']:
                    continue
                if assignment.branch.casefold() != str(row.get('branch') or '').casefold():
                    continue
                if assignment.role.casefold() != str(row.get('responsible_role') or '').casefold():
                    continue
                if assignment.product_key and assignment.product_key.casefold() != row['_product_key'].casefold():
                    continue
                if assignment.stage_key and assignment.stage_key.casefold() != str(row.get('current_stage_key') or '').casefold():
                    continue
                users.add(assignment.primary_user_id)
                break
        result[label] = len(users)
    return result


def _comparison_explorer(rows, samples, filters):
    dimension = filters['chart_dimension']
    metric = filters['chart_metric']
    applied = _active_filter_names(filters)
    if metric == 'load_per_assignee' and filters['view'] == 'performance':
        return _chart_payload(
            'explorer', f'Cases per Configured Assignee by {dimension.title()}',
            'current_cases_per_distinct_configured_primary_assignee',
            'This is a current responsibility-coverage measure and is not available for historical performance.',
            [], [], applied_filters=applied, unavailable_filters=['performance_view'],
        )
    source_is_samples = filters['view'] == 'performance' or metric in {
        'duration', 'target_usage', 'sla_met', 'correction_rate',
    }
    if not source_is_samples:
        applied = _active_filter_names(filters, include_dates=False)
    source = samples if source_is_samples else rows
    grouped = defaultdict(list)
    for item in source:
        grouped[_dimension_value(item, dimension, sample=source_is_samples)].append(item)
    labels = sorted(grouped)
    excluded = 0
    extras = {'dimension': dimension, 'metric': metric}
    if metric == 'workload':
        values = [len(grouped[label]) for label in labels]
        basis = 'completed_stage_actions' if source_is_samples else 'current_workload'
        series = [_series('count', 'Actions' if source_is_samples else 'Cases', values)]
        subtitle = f"Showing {'completed actions' if source_is_samples else 'current cases'} by {dimension}."
    elif metric == 'sla_state':
        states = ('within_target', 'near_target', 'overdue', 'target_unavailable')
        state_labels = ('Within Target', 'Near Target', 'Overdue', 'Target Unavailable')
        series = [
            _series(state, label, [sum(item.get('sla_state') == state for item in grouped[key]) for key in labels])
            for state, label in zip(states, state_labels)
        ]
        basis = 'completed_stage_sla_state' if source_is_samples else 'current_sla_state'
        subtitle = f'Showing SLA-state composition by {dimension}.'
    elif metric == 'duration':
        series = [
            _series('median_minutes', 'Median', [_percentile([item.get('elapsed_minutes') for item in grouped[label]], .5) for label in labels]),
            _series('p90_minutes', 'P90', [_percentile([item.get('elapsed_minutes') for item in grouped[label]], .9) for label in labels]),
        ]
        extras['iqr_minutes'] = {
            label: {
                'q1': _percentile([item.get('elapsed_minutes') for item in grouped[label]], .25),
                'q3': _percentile([item.get('elapsed_minutes') for item in grouped[label]], .75),
            } for label in labels
        }
        basis = 'completed_stage_wall_clock_duration'
        subtitle = f'Showing exact completed-stage duration by {dimension}; IQR is available in details.'
    elif metric == 'target_usage':
        ratios = {}
        for label in labels:
            ratios[label] = []
            for item in grouped[label]:
                elapsed = item.get('elapsed_minutes'); target = item.get('target_minutes')
                if elapsed is None or target is None or target <= 0:
                    excluded += 1
                    continue
                ratios[label].append(float(elapsed) * 100 / float(target))
        labels = [label for label in labels if ratios[label]]
        series = [
            _series('median_percent', 'Median % of Target', [_percentile(ratios[label], .5) for label in labels]),
            _series('p90_percent', 'P90 % of Target', [_percentile(ratios[label], .9) for label in labels]),
        ]
        basis = 'completed_actions_percent_of_frozen_target'
        subtitle = f'Each action is normalized against its own frozen target before comparison by {dimension}.'
        extras.update(axis_title='% of target', reference_line=100)
    elif metric == 'sla_met':
        values = []
        for label in labels:
            valid = [item for item in grouped[label] if item.get('sla_state') != 'target_unavailable']
            excluded += len(grouped[label]) - len(valid)
            values.append(round(sum(item.get('sla_state') != 'overdue' for item in valid) * 100 / len(valid), 1) if valid else None)
        series = [_series('sla_met_percent', 'SLA Met %', values)]
        basis = 'completed_stage_sla_compliance'
        subtitle = f'Showing completed actions within target by {dimension}.'
        extras['axis_title'] = 'SLA met %'
    elif metric == 'correction_rate':
        series = [_series('correction_percent', 'Recorded Correction %', [
            round(sum(bool(item.get('corrected')) for item in grouped[label]) * 100 / len(grouped[label]), 1)
            if grouped[label] else None for label in labels
        ])]
        basis = 'distinct_completed_actions_with_recorded_admin_correction'
        subtitle = 'Each completed case-stage action counts once; repeated corrections do not increase the numerator.'
        extras['axis_title'] = 'Recorded correction %'
    else:
        assignment_counts = _active_assignment_counts(rows, dimension)
        values = [round(len(grouped[label]) / assignment_counts.get(label, 0), 2) if assignment_counts.get(label, 0) else None for label in labels]
        unassigned = sum(len(grouped[label]) for label in labels if not assignment_counts.get(label, 0))
        series = [_series('cases_per_assignee', 'Cases per Configured Assignee', values)]
        basis = 'current_cases_per_distinct_configured_primary_assignee'
        subtitle = 'Configured responsibility coverage only; this does not measure attendance or individual productivity.'
        extras.update(assignee_counts=assignment_counts, unassigned_case_count=unassigned)
    return _chart_payload(
        'explorer', f"{metric.replace('_', ' ').title()} by {dimension.title()}", basis, subtitle,
        labels, series, applied_filters=applied,
        unavailable_filters=(['date_range', 'granularity'] if not source_is_samples else []),
        sample_count=len(source), excluded_count=excluded,
        exclusion_reason='Target unavailable', extras=extras,
    )


def _heatmap_payload(rows, samples, filters):
    row_dimension, column_dimension = HEATMAP_PAIRS[filters['heatmap_pair']]
    metric = filters['heatmap_metric']
    sample_based = filters['view'] == 'performance' or metric != 'workload'
    source = samples if sample_based else rows
    grouped = defaultdict(list)
    for item in source:
        grouped[(
            _dimension_value(item, row_dimension, sample=sample_based),
            _dimension_value(item, column_dimension, sample=sample_based),
        )].append(item)
    row_labels = sorted({key[0] for key in grouped})
    column_labels = sorted({key[1] for key in grouped})
    cells = []
    total_excluded = 0
    for row_label in row_labels:
        for column_label in column_labels:
            cohort = grouped.get((row_label, column_label), [])
            excluded = 0
            if metric == 'workload':
                value = len(cohort)
            elif metric == 'duration':
                value = _percentile([item.get('elapsed_minutes') for item in cohort], .5)
            elif metric == 'target_usage':
                ratios = []
                for item in cohort:
                    elapsed = item.get('elapsed_minutes'); target = item.get('target_minutes')
                    if elapsed is None or target is None or target <= 0:
                        excluded += 1
                    else:
                        ratios.append(float(elapsed) * 100 / float(target))
                value = _percentile(ratios, .5)
            else:
                valid = [item for item in cohort if item.get('sla_state') != 'target_unavailable']
                excluded = len(cohort) - len(valid)
                value = round(sum(item.get('sla_state') != 'overdue' for item in valid) * 100 / len(valid), 1) if valid else None
            total_excluded += excluded
            cells.append({
                'row': row_label, 'column': column_label, 'value': value,
                'sample_count': len(cohort) - excluded, 'excluded_count': excluded,
            })
    return {
        'id': 'heatmap', 'title': f'{row_dimension.title()} × {column_dimension.title()}',
        'basis': ('completed_stage_actions' if sample_based else 'current_workload'),
        'subtitle': f"Showing {metric.replace('_', ' ')} across two operational dimensions.",
        'applied_filters': _active_filter_names(filters, include_dates=sample_based),
        'unavailable_filters': ['date_range', 'granularity'] if not sample_based else [],
        'sample_count': len(source), 'excluded_count': total_excluded,
        'exclusion_reason': 'Target unavailable', 'metric': metric,
        'row_dimension': row_dimension, 'column_dimension': column_dimension,
        'rows': row_labels, 'columns': column_labels, 'cells': cells,
    }


def _wilson_lower_percent(over_count, sample_count):
    if sample_count <= 0:
        return None
    z = TARGET_REVIEW_SIGNAL_POLICY['wilson_z']
    proportion = over_count / sample_count
    denominator = 1 + (z * z / sample_count)
    centre = proportion + (z * z / (2 * sample_count))
    margin = z * sqrt((proportion * (1 - proportion) / sample_count) + (z * z / (4 * sample_count * sample_count)))
    return round(((centre - margin) / denominator) * 100, 1)


def _target_stats(samples):
    valid = [item for item in samples if item.get('target_minutes') is not None and item.get('target_minutes') > 0]
    states = Counter(item.get('sla_state') for item in valid)
    count = len(valid); over = states['overdue']
    ratios = [float(item['elapsed_minutes']) * 100 / float(item['target_minutes']) for item in valid if item.get('elapsed_minutes') is not None]
    return {
        'valid_samples': count, 'within_count': states['within_target'],
        'near_count': states['near_target'], 'over_count': over,
        'within_percent': round(states['within_target'] * 100 / count, 1) if count else None,
        'near_percent': round(states['near_target'] * 100 / count, 1) if count else None,
        'over_percent': round(over * 100 / count, 1) if count else None,
        'target_unavailable_count': len(samples) - count,
        'wilson_lower_bound': _wilson_lower_percent(over, count),
        'median_percent_of_target': _percentile(ratios, .5),
        'p90_percent_of_target': _percentile(ratios, .9),
    }


def _target_review_signals(actor, filters, selected_samples):
    baseline_filters = dict(filters)
    for key in ('branch', 'product', 'search', 'status', 'role', 'sla_state'):
        baseline_filters[key] = ''
    baseline_cases = _filtered_cases(actor, baseline_filters)
    baseline_samples = _stage_samples(baseline_cases, baseline_filters, include_people=False)
    selected_by_key = defaultdict(list)
    for sample in selected_samples:
        selected_by_key[(sample['group_id'], sample['stage_key'])].append(sample)
    grouped = defaultdict(list)
    for sample in baseline_samples:
        grouped[(sample['group_id'], sample['stage_key'], sample['stage'])].append(sample)
    group_labels = dict(GroupSheetConfiguration.objects.filter(
        group_id__in={sample['group_id'] for sample in baseline_samples},
    ).values_list('group_id', 'display_name'))
    policy = TARGET_REVIEW_SIGNAL_POLICY
    results = []
    for (group_id, stage_key, stage_label), samples in sorted(grouped.items()):
        baseline = _target_stats(samples)
        cohorts = {}
        breadth_ok = False
        qualifying_details = []
        for dimension, field in (('branch', 'branch'), ('product', 'product')):
            dimension_groups = defaultdict(list)
            for sample in samples:
                dimension_groups[str(sample.get(field) or 'Unassigned')].append(sample)
            details = []
            for label, cohort_samples in sorted(dimension_groups.items()):
                stats = _target_stats(cohort_samples)
                qualifies = bool(
                    stats['valid_samples'] >= policy['cohort_min_samples']
                    and (stats['over_percent'] or 0) >= policy['cohort_min_over_percent']
                )
                details.append({'label': label, **stats, 'qualifies': qualifies})
            qualifying = [item for item in details if item['qualifies']]
            coverage = round(sum(item['valid_samples'] for item in qualifying) * 100 / baseline['valid_samples'], 1) if baseline['valid_samples'] else 0
            cohorts[dimension] = {'items': details, 'qualifying_count': len(qualifying), 'coverage_percent': coverage}
            if len(qualifying) >= 2 and coverage >= policy['cohort_coverage_percent']:
                breadth_ok = True
                qualifying_details.append(dimension)
        systemic = bool(
            baseline['valid_samples'] >= policy['systemic_min_samples']
            and (baseline['over_percent'] or 0) >= policy['systemic_min_over_percent']
            and (baseline['wilson_lower_bound'] or 0) > policy['systemic_wilson_lower_bound']
            and breadth_ok
        )
        localized = []
        if not systemic:
            for dimension, detail in cohorts.items():
                for cohort in detail['items']:
                    if (
                        cohort['valid_samples'] >= policy['localized_min_samples']
                        and (cohort['over_percent'] or 0) >= policy['localized_min_over_percent']
                        and (cohort['wilson_lower_bound'] or 0) > policy['localized_wilson_lower_bound']
                    ):
                        localized.append({'dimension': dimension, **cohort})
        selected = _target_stats(selected_by_key.get((group_id, stage_key), []))
        narrowed = bool(filters['branch'] or filters['product'])
        if narrowed and selected['valid_samples'] >= policy['localized_min_samples'] and (selected['over_percent'] or 0) >= policy['localized_min_over_percent'] and (selected['wilson_lower_bound'] or 0) > policy['localized_wilson_lower_bound']:
            classification = 'selected_scope_high'
            message = 'High exceedance in selected scope. Review this stage before drawing an organization-wide conclusion.'
        elif systemic:
            classification = 'review_recommended'
            message = 'Review recommended: most cases exceed target on this stage across multiple areas. Check whether the target is realistic or whether a shared process issue is causing delays.'
        elif localized:
            classification = 'localized_delay'
            first = localized[0]
            message = f"Localized delay signal: {first['label']} frequently exceeds target on this stage; other areas do not show the same pattern."
        else:
            classification = 'none'; message = ''
        results.append({
            'group_id': group_id, 'group': group_labels.get(group_id) or 'TAT Tracker',
            'stage_key': stage_key, 'stage': stage_label,
            'signal_scope': 'authorized_group_date_baseline',
            'signal_policy': {key: value for key, value in policy.items() if key != 'wilson_z'},
            'baseline': baseline, 'selected_scope': selected,
            'baseline_systemic': systemic,
            'qualifying_cohorts': qualifying_details, 'cohorts': cohorts,
            'classification': classification, 'message': message,
        })
    return results


def report_summary(actor, payload, *, include_people=False):
    filters = _filters(payload)
    all_cases = _filtered_cases(actor, filters)
    rows = _eligible_rows(actor, filters, include_people=include_people)
    scope_cases = list(scoped_cases(actor))
    stages = set(); roles = set()
    for case in scope_cases:
        try:
            for stage in product_for_case(case).stages:
                stages.add((stage.key, stage.label)); roles.add(stage.role)
        except ValueError:
            continue
    group_names = dict(GroupSheetConfiguration.objects.filter(
        group_id__in={case.group_id for case in scope_cases},
    ).values_list('group_id', 'display_name'))
    options = {
        'groups': sorted({
            (case.group_id, group_names.get(case.group_id) or 'TAT Tracker')
            for case in scope_cases
        }),
        'branches': _casefold_distinct_labels(case.branch for case in scope_cases),
        'products': sorted({(case.product_key, case.product_label or case.product_key) for case in scope_cases}),
        'stages': sorted(stages), 'roles': sorted(role for role in roles if role),
    }
    # Explorer, heatmap, target-review and correction-rate insights all use
    # the same exact timestamp-backed stage observations. Committing one stage
    # stamp starts the next stage; there is no inferred pickup timestamp.
    stage_samples = _stage_samples(all_cases, filters, include_people=include_people)
    breakdown_rows = rows
    breakdown_basis = 'current_workload'
    if filters['view'] == 'performance':
        if stage_samples:
            breakdown_basis = 'completed_stage_actions'
        else:
            created_cases = [
                case for case in all_cases
                if filters['date_from'] <= timezone.localdate(case.created_at) <= filters['date_to']
            ]
            breakdown_rows = [
                row for row in (_case_row(case, include_people=include_people) for case in created_cases)
                if (not filters['stage'] or row['current_stage_key'].casefold() == filters['stage'].casefold())
                and (not filters['role'] or row['responsible_role'].upper() == filters['role'])
                and (not filters['sla_state'] or row['sla_state'] == filters['sla_state'])
            ]
            breakdown_basis = 'created_cases_current_stage'
    breakdown_samples = stage_samples if filters['view'] == 'performance' else []
    by_stage = Counter(
        (item['stage'] for item in breakdown_samples)
        if breakdown_samples else (row['current_stage'] or 'Unassigned' for row in breakdown_rows)
    )
    by_role = Counter(
        (item['role'] or 'Unassigned' for item in breakdown_samples)
        if breakdown_samples else (row['responsible_role'] or 'Unassigned' for row in breakdown_rows)
    )
    latest_metrics = WorkflowTatDailyMetric.objects.filter(
        workflow='tat_tracker', metric_grain='current_leaf',
    ).filter(_metric_scope_q(actor))
    latest_metrics = _filter_metric_queryset(latest_metrics, filters)
    latest = latest_metrics.order_by('-metric_date').first()
    earliest = latest_metrics.order_by('metric_date').first()
    scoped_case_ids = scoped_cases(actor).values_list('pk', flat=True)
    scoped_rebuilds = WorkflowTatMetricRebuildRequest.objects.filter(
        Q(case__isnull=True) | Q(case_id__in=scoped_case_ids),
    )
    pending_rebuilds = scoped_rebuilds.filter(status__in=['pending', 'processing']).count()
    failed_rebuilds = scoped_rebuilds.filter(status='failed', attempts__gte=3).count()
    common = {
        'view': filters['view'], 'filters': options,
        'by_stage': [{'label': key, 'count': value} for key, value in by_stage.most_common()],
        'by_role': [{'label': key, 'count': value} for key, value in by_role.most_common()],
        'breakdown_basis': breakdown_basis,
        'freshness': {
            'latest_snapshot': latest.metric_date.isoformat() if latest else '',
            'earliest_snapshot': earliest.metric_date.isoformat() if earliest else '',
            'pending_rebuilds': pending_rebuilds,
            'failed_rebuilds': failed_rebuilds,
            'near_target_percent': presentation_settings()['near_target_percent'],
            'presentation_revision': presentation_settings()['revision'],
        },
    }
    charts = {}
    active_filters = _active_filter_names(filters)
    live_filters = _active_filter_names(filters, include_dates=False)
    live_unavailable = ['date_range', 'granularity']
    if filters['view'] == 'current':
        states = Counter(row['sla_state'] for row in rows)
        common['metrics'] = {
            'active': len(rows), 'within_target': states['within_target'],
            'near_target': states['near_target'],
            'overdue': states['overdue'],
            'stalled': sum(row['status'] == 'Stalled' for row in rows),
            'target_unavailable': states['target_unavailable'],
        }
        action_trend = bool(filters['search'] or filters['status'] or filters['stage'] or filters['role'] or filters['sla_state'])
        if action_trend:
            buckets = defaultdict(int)
            for sample in stage_samples:
                completed_date = _iso_local_date(sample['completed_at'])
                buckets[_bucket_label(completed_date, filters['granularity'])] += 1
            common['trend'] = [
                {'label': label, 'completed_actions': count}
                for label, count in sorted(buckets.items())
            ]
            common['trend_notice'] = 'No completed actions match these filters.'
            charts['trend'] = _chart_payload(
                'trend', 'Completed Actions over Time', 'completed_stage_actions',
                'Showing completed stage actions because case-level filters cannot be applied to aggregate workload snapshots.',
                [item['label'] for item in common['trend']],
                [_series('completed_actions', 'Completed Actions', [item['completed_actions'] for item in common['trend']])],
                applied_filters=active_filters, sample_count=len(stage_samples),
            )
        else:
            daily = WorkflowTatDailyMetric.objects.filter(
                workflow='tat_tracker', metric_grain='current_leaf',
                metric_date__range=(filters['date_from'], filters['date_to']),
            ).filter(_metric_scope_q(actor))
            daily = _filter_metric_queryset(daily, filters)
            daily_totals = defaultdict(lambda: Counter())
            for item in daily:
                daily_totals[item.metric_date]['active'] += item.active_count
                daily_totals[item.metric_date]['near_target'] += item.near_target_count
                daily_totals[item.metric_date]['overdue'] += item.overdue_count
            # Workload is a point-in-time measure. For coarser grouping use
            # the latest available snapshot in each bucket, never a sum of
            # daily balances (which would inflate the apparent queue).
            trend = {}
            for metric_date, values in sorted(daily_totals.items()):
                trend[_bucket_label(metric_date, filters['granularity'])] = values
            common['trend'] = [{'label': key, **dict(value)} for key, value in sorted(trend.items())]
            charts['trend'] = _chart_payload(
                'trend', 'Workload over Time', 'daily_point_in_time_snapshots',
                'Showing the latest reliable point-in-time workload snapshot in each period.',
                [item['label'] for item in common['trend']],
                [
                    _series('active', 'Active', [item.get('active', 0) for item in common['trend']]),
                    _series('near_target', 'Near Target', [item.get('near_target', 0) for item in common['trend']]),
                    _series('overdue', 'Overdue', [item.get('overdue', 0) for item in common['trend']]),
                ],
                applied_filters=active_filters, sample_count=sum(item.get('active', 0) for item in common['trend']),
            )

        backlog = Counter()
        for row in rows:
            minutes = row.get('elapsed_minutes')
            if minutes is None:
                continue
            days = max(0, float(minutes) / 1440)
            bucket = '0–1 day' if days < 1 else ('1–3 days' if days < 3 else ('3–7 days' if days < 7 else '7+ days'))
            backlog[bucket] += 1
        backlog_labels = ['0–1 day', '1–3 days', '3–7 days', '7+ days']
        charts['backlog_age'] = _chart_payload(
            'backlog_age', 'Current-stage Backlog Age', 'current_stage_wall_clock_age',
            'Showing how long active cases have remained in their current stage.',
            backlog_labels, [_series('cases', 'Cases', [backlog[label] for label in backlog_labels])],
            applied_filters=live_filters, unavailable_filters=live_unavailable,
            sample_count=sum(backlog.values()),
        )
    else:
        terminal_rows = rows
        action_filtered = bool(filters['stage'] or filters['role'] or filters['sla_state'])
        metric_rows = stage_samples if action_filtered else terminal_rows
        elapsed = [row['elapsed_minutes'] for row in metric_rows]
        valid = [row for row in metric_rows if row['sla_state'] != 'target_unavailable']
        met = sum(row['sla_state'] != 'overdue' for row in valid)
        outcomes = Counter(row['status'] for row in terminal_rows)
        created_cases = [case for case in all_cases if filters['date_from'] <= timezone.localdate(case.created_at) <= filters['date_to']]
        created = len({sample['case_id'] for sample in stage_samples}) if action_filtered else len(created_cases)
        common['metrics'] = {
            'created': created, 'finished': len(stage_samples) if action_filtered else len(terminal_rows), 'disbursed': outcomes['Disbursed'],
            'rejected': outcomes['Rejected'], 'declined': outcomes['Declined'],
            'sla_met': met, 'sla_sample': len(valid),
            'sla_met_percent': round((met * 100 / len(valid)), 1) if valid else None,
            'median_tat_minutes': _percentile(elapsed, .5), 'p90_tat_minutes': _percentile(elapsed, .9),
            'target_unavailable': len(metric_rows) - len(valid),
        }
        common['metric_basis'] = 'completed_stage_actions' if action_filtered else 'finished_cases'
        trend = defaultdict(lambda: Counter())
        if action_filtered:
            for sample in stage_samples:
                completed_date = _iso_local_date(sample['completed_at'])
                trend[_bucket_label(completed_date, filters['granularity'])]['completed_actions'] += 1
        else:
            for case in created_cases:
                trend[_bucket_label(timezone.localdate(case.created_at), filters['granularity'])]['created'] += 1
            for row in terminal_rows:
                finished_date = _iso_local_date(row['finished_at'])
                bucket = trend[_bucket_label(finished_date, filters['granularity'])]
                bucket['finished'] += 1
                bucket['disbursed'] += int(row['status'] == 'Disbursed')
                bucket['rejected'] += int(row['status'] == 'Rejected')
                bucket['declined'] += int(row['status'] == 'Declined')
                if row['sla_state'] != 'target_unavailable':
                    bucket['sla_sample'] += 1
                    bucket['sla_met'] += int(row['sla_state'] != 'overdue')
        common['trend'] = [
            {'label': key, **dict(value), 'sla_met_percent': round(value['sla_met'] * 100 / value['sla_sample'], 1) if value['sla_sample'] else None}
            for key, value in sorted(trend.items())
        ]
        if action_filtered:
            charts['trend'] = _chart_payload(
                'trend', 'Completed Actions over Time', 'completed_stage_actions',
                'Showing completed actions that match the selected Stage, Role, and SLA filters.',
                [item['label'] for item in common['trend']],
                [_series('completed_actions', 'Completed Actions', [item.get('completed_actions', 0) for item in common['trend']])],
                applied_filters=active_filters, sample_count=len(stage_samples),
            )
        else:
            charts['trend'] = _chart_payload(
                'trend', 'Cases and Outcomes over Time', 'case_creation_and_terminal_outcomes',
                'Showing case creation and final workflow outcomes for the selected period.',
                [item['label'] for item in common['trend']],
                [
                    _series('created', 'Created', [item.get('created', 0) for item in common['trend']]),
                    _series('finished', 'Finished', [item.get('finished', 0) for item in common['trend']]),
                    _series('disbursed', 'Disbursed', [item.get('disbursed', 0) for item in common['trend']]),
                    _series('rejected', 'Rejected', [item.get('rejected', 0) for item in common['trend']]),
                    _series('declined', 'Declined', [item.get('declined', 0) for item in common['trend']]),
                ],
                applied_filters=active_filters, sample_count=len(created_cases) + len(terminal_rows),
            )

        sla_source = stage_samples if action_filtered else terminal_rows
        sla_buckets = defaultdict(lambda: Counter())
        duration_buckets = defaultdict(list)
        for item in sla_source:
            timestamp = item.get('completed_at') if action_filtered else item.get('finished_at')
            if not timestamp:
                continue
            label = _bucket_label(_iso_local_date(timestamp), filters['granularity'])
            if item.get('sla_state') != 'target_unavailable':
                sla_buckets[label]['sample'] += 1
                sla_buckets[label]['met'] += int(item.get('sla_state') != 'overdue')
            if item.get('elapsed_minutes') is not None:
                duration_buckets[label].append(item['elapsed_minutes'])
        time_labels = sorted(set(sla_buckets) | set(duration_buckets))
        charts['sla_compliance'] = _chart_payload(
            'sla_compliance', 'SLA Compliance over Time',
            'completed_stage_actions' if action_filtered else 'finished_cases',
            'Showing the percentage of valid samples completed within their frozen target.',
            time_labels,
            [_series('sla_met_percent', 'SLA Met %', [
                round(sla_buckets[label]['met'] * 100 / sla_buckets[label]['sample'], 1)
                if sla_buckets[label]['sample'] else None for label in time_labels
            ])],
            applied_filters=active_filters,
            sample_count=sum(value['sample'] for value in sla_buckets.values()),
            excluded_count=sum(item.get('sla_state') == 'target_unavailable' for item in sla_source),
            exclusion_reason='Target unavailable',
        )
        charts['tat_percentiles'] = _chart_payload(
            'tat_percentiles', 'Median and P90 TAT over Time',
            'completed_stage_duration' if action_filtered else 'finished_case_duration',
            'Showing exact wall-clock duration percentiles; daily percentile rows are not re-aggregated.',
            time_labels,
            [
                _series('median_minutes', 'Median', [_percentile(duration_buckets[label], .5) for label in time_labels]),
                _series('p90_minutes', 'P90', [_percentile(duration_buckets[label], .9) for label in time_labels]),
            ],
            applied_filters=active_filters,
            sample_count=sum(len(values) for values in duration_buckets.values()),
        )

        target_groups = defaultdict(list)
        target_versions = defaultdict(set)
        target_unavailable = 0
        for sample in stage_samples:
            elapsed_minutes = sample.get('elapsed_minutes')
            target_minutes = sample.get('target_minutes')
            if elapsed_minutes is None or target_minutes is None or target_minutes <= 0:
                target_unavailable += 1
                continue
            target_groups[sample['stage']].append(float(elapsed_minutes) * 100 / float(target_minutes))
            target_versions[sample['stage']].add(round(float(target_minutes) / 1440, 2))
        target_labels = sorted(target_groups, key=lambda label: _percentile(target_groups[label], .9) or 0, reverse=True)
        target_details = []
        if filters['product']:
            for label in target_labels:
                matching = [item for item in stage_samples if item['stage'] == label and item.get('elapsed_minutes') is not None and item.get('target_minutes')]
                targets = sorted(target_versions[label])
                target_details.append({
                    'label': label,
                    'median_days': round((_percentile([item['elapsed_minutes'] for item in matching], .5) or 0) / 1440, 2),
                    'p90_days': round((_percentile([item['elapsed_minutes'] for item in matching], .9) or 0) / 1440, 2),
                    'target_days': targets[0] if len(targets) == 1 else None,
                    'target_versions_days': targets,
                })
        charts['stage_target'] = _chart_payload(
            'stage_target', 'Stage Performance Against Target', 'completed_actions_percent_of_frozen_target',
            'Each action is compared with its own frozen target before median and P90 percentages are calculated.',
            target_labels,
            [
                _series('median_percent', 'Median % of Target', [_percentile(target_groups[label], .5) for label in target_labels]),
                _series('p90_percent', 'P90 % of Target', [_percentile(target_groups[label], .9) for label in target_labels]),
            ],
            applied_filters=active_filters, sample_count=sum(len(values) for values in target_groups.values()),
            excluded_count=target_unavailable, exclusion_reason='Target unavailable',
            extras={'axis_title': '% of target', 'reference_line': 100, 'single_product_details': target_details},
        )

    stage_basis = breakdown_basis
    stage_title = 'Completed Actions by Stage' if stage_basis == 'completed_stage_actions' else ('Created Cases by Current Stage' if stage_basis == 'created_cases_current_stage' else 'Cases by Current Stage')
    role_title = 'Completed Actions by Role' if stage_basis == 'completed_stage_actions' else ('Created Cases by Current Role' if stage_basis == 'created_cases_current_stage' else 'Cases by Responsible Role')
    basis_subtitle = {
        'completed_stage_actions': 'Showing timestamped completed stage actions in the selected period.',
        'created_cases_current_stage': 'No completed actions matched; showing created cases by their current stage and role.',
        'current_workload': 'Showing the current active workload.',
    }[stage_basis]
    breakdown_unavailable = live_unavailable if filters['view'] == 'current' else []
    charts['stage'] = _breakdown_chart('stage', stage_title, stage_basis, basis_subtitle, common['by_stage'], applied_filters=live_filters if filters['view'] == 'current' else active_filters, unavailable_filters=breakdown_unavailable)
    charts['role'] = _breakdown_chart('role', role_title, stage_basis, basis_subtitle, common['by_role'], applied_filters=live_filters if filters['view'] == 'current' else active_filters, unavailable_filters=breakdown_unavailable)
    if include_people:
        common['by_person'] = [
            {'label': key, 'count': value}
            for key, value in Counter(
                (item.get('person') or 'Unassigned' for item in breakdown_samples)
                if breakdown_samples else (row.get('responsible_person') or 'Unassigned' for row in breakdown_rows)
            ).most_common()
        ]
        person_basis = 'completed_action_performer' if stage_samples and filters['view'] == 'performance' else 'current_responsibility_assignment'
        charts['person'] = _breakdown_chart(
            'person', 'Completed Actions by Person' if person_basis == 'completed_action_performer' else 'Cases by Responsible Person',
            person_basis,
            'Showing recorded action performers.' if person_basis == 'completed_action_performer' else 'Showing current responsibility assignments.',
            common['by_person'], applied_filters=live_filters if filters['view'] == 'current' else active_filters,
            unavailable_filters=breakdown_unavailable,
        )
    charts['explorer'] = _comparison_explorer(rows, stage_samples, filters)
    common['heatmap'] = _heatmap_payload(rows, stage_samples, filters)
    target_signals = _target_review_signals(actor, filters, stage_samples)
    common['target_review_signals'] = {
        'basis': 'completed_actions_against_each_frozen_target',
        'subtitle': 'Statistical review prompts use an authorization-scoped unfiltered branch and product baseline; they never change targets.',
        'applied_filters': active_filters,
        'unavailable_filters': [],
        'sample_count': sum(item['selected_scope']['valid_samples'] for item in target_signals),
        'excluded_count': sum(item['selected_scope']['target_unavailable_count'] for item in target_signals),
        'exclusion_reason': 'Target unavailable',
        'items': target_signals,
    }

    current_filters = dict(filters)
    current_filters['view'] = 'current'
    oldest_rows = sorted(
        _eligible_rows(actor, current_filters, include_people=include_people),
        key=lambda item: item.get('elapsed_minutes') if item.get('elapsed_minutes') is not None else -1,
        reverse=True,
    )[:10]
    oldest_fields = (
        'case_id', 'client_name', 'branch', 'product_label', 'current_stage',
        'responsible_role', 'elapsed_minutes', 'target_minutes', 'sla_state',
    )
    common['oldest_cases'] = {
        'basis': 'current_active_stage_wall_clock_age',
        'subtitle': 'Showing the ten oldest active cases. Date range and time grouping do not apply to a current-workload ranking.',
        'applied_filters': _active_filter_names(filters, include_dates=False),
        'unavailable_filters': ['date_range', 'granularity'],
        'sample_count': len(oldest_rows), 'excluded_count': 0, 'exclusion_reason': '',
        'items': [{key: row.get(key) for key in oldest_fields} for row in oldest_rows],
    }
    _attach_report_filter_guidance(common, charts, filters)
    common['charts'] = charts
    return common


def _metric_scope_q(actor):
    from core.models import WORKFLOW_DATA_MODE_PILOT, WORKFLOW_DATA_MODE_PRODUCTION
    from core.services.workflow_data_mode import WORKFLOW_TAT, mode_snapshot
    snapshot = mode_snapshot(WORKFLOW_TAT)
    operational = Q(data_mode=WORKFLOW_DATA_MODE_PRODUCTION)
    if snapshot.mode == WORKFLOW_DATA_MODE_PILOT:
        operational |= Q(data_mode=WORKFLOW_DATA_MODE_PILOT, pilot_cycle_id=snapshot.pilot_cycle_id)
    if actor.is_active and actor.is_superuser:
        return operational
    query = Q(pk__in=[])
    for grant in AccessGrant.objects.filter(user=actor, workflow='tat_tracker', active=True).select_related('group_configuration'):
        part = Q()
        scoped = False
        if grant.group_configuration_id:
            part &= Q(group_id=str(grant.group_configuration.group_id))
            scoped = True
        if grant.branch:
            part &= Q(branch__iexact=grant.branch)
            scoped = True
        if grant.product:
            part &= Q(product_key__iexact=grant.product)
            scoped = True
        if not scoped:
            return operational
        query |= part
    return operational & query


def _filter_metric_queryset(qs, filters):
    if filters['group']:
        qs = qs.filter(group_id=filters['group'])
    if filters['branch']:
        qs = qs.filter(branch__iexact=filters['branch'])
    if filters['product']:
        qs = qs.filter(product_key__iexact=filters['product'])
    if filters['stage']:
        qs = qs.filter(stage_key__iexact=filters['stage'])
    if filters['role']:
        qs = qs.filter(responsible_role__iexact=filters['role'])
    return qs


def report_cases(actor, payload, *, include_people=False):
    filters = _filters(payload)
    rows = _eligible_rows(actor, filters, include_people=include_people)
    sort = str(payload.get('sort') or '-created_at')
    descending = sort.startswith('-')
    key = sort.lstrip('-')
    if key not in SORT_FIELDS:
        raise ValueError('This report column cannot be sorted.')
    rows.sort(key=lambda row: (row.get(key) is None, row.get(key) or ''), reverse=descending)
    try:
        page = max(1, int(payload.get('page') or 1))
        page_size = max(1, min(100, int(payload.get('page_size') or 25)))
    except (TypeError, ValueError):
        raise ValueError('Page and page size must be valid numbers.')
    start = (page - 1) * page_size
    public_rows = [
        {key: value for key, value in row.items() if not key.startswith('_')}
        for row in rows[start:start + page_size]
    ]
    return {'results': public_rows, 'count': len(rows), 'page': page, 'page_size': page_size}


def export_report_xlsx(actor, payload, *, include_people=False, request_id=''):
    # Exports are allowed to exceed the API page cap, while retaining a hard operational limit.
    filters = _filters(payload)
    rows = _eligible_rows(actor, filters, include_people=include_people)
    if len(rows) > 10000:
        raise ValueError('More than 10,000 cases match. Narrow the filters before downloading.')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    def xlsx_safe(value):
        return "'" + value if isinstance(value, str) and value[:1] in {'=', '+', '-', '@'} else value

    headers = ['Reference', 'Customer', 'TAT Group', 'Branch', 'Product', 'Status', 'Stage', 'Responsible Role', 'Created', 'Finished', 'Elapsed Minutes', 'Target Minutes', 'Variance Minutes', 'SLA State']
    keys = ['case_id', 'client_name', 'group', 'branch', 'product_label', 'status', 'current_stage', 'responsible_role', 'created_at', 'finished_at', 'elapsed_minutes', 'target_minutes', 'variance_minutes', 'sla_state']
    if include_people:
        headers.append('Responsible Person'); keys.append('responsible_person')
    workbook = Workbook(); sheet = workbook.active; sheet.title = 'TAT Report'; sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='245B8A')
    for row in rows:
        values = []
        for key in keys:
            value = row.get(key, '')
            value = xlsx_safe(value)
            if key in {'created_at', 'finished_at'} and value:
                value = datetime.fromisoformat(value).strftime('%d-%m-%y %H:%M')
            values.append(value)
        sheet.append(values)
    # The existing audited export also carries the two selected aggregate
    # comparisons. It does not introduce another export endpoint or expose
    # case-level fields beyond the established TAT Report sheet.
    stage_samples = _stage_samples(_filtered_cases(actor, filters), filters, include_people=False)
    explorer = _comparison_explorer(rows, stage_samples, filters)
    insight_sheet = workbook.create_sheet('Selected Insight')
    insight_sheet.append([explorer['title']])
    insight_sheet.append(['Basis', explorer['basis']])
    insight_sheet.append(['Description', explorer['subtitle']])
    insight_sheet.append([])
    insight_sheet.append(['Dimension', *[item['label'] for item in explorer.get('series', [])]])
    for index, label in enumerate(explorer.get('labels', [])):
        insight_sheet.append([
            xlsx_safe(label),
            *[item.get('values', [])[index] if index < len(item.get('values', [])) else None for item in explorer.get('series', [])],
        ])
    for cell in insight_sheet[5]:
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='245B8A')

    heatmap = _heatmap_payload(rows, stage_samples, filters)
    heatmap_sheet = workbook.create_sheet('Selected Heatmap')
    heatmap_sheet.append([heatmap['title']])
    heatmap_sheet.append(['Basis', heatmap['basis']])
    heatmap_sheet.append(['Metric', heatmap['metric']])
    heatmap_sheet.append([])
    heatmap_sheet.append([heatmap['row_dimension'], *[xlsx_safe(value) for value in heatmap['columns']]])
    lookup = {(item['row'], item['column']): item for item in heatmap['cells']}
    for row_label in heatmap['rows']:
        values = [lookup.get((row_label, column), {}).get('value') for column in heatmap['columns']]
        heatmap_sheet.append([xlsx_safe(row_label), *values])
    for cell in heatmap_sheet[5]:
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='245B8A')
    output = BytesIO(); workbook.save(output)
    from core.services.compliance_audit import record_event
    record_event(
        workflow='tat_tracker', action='tat.report.exported', category='data_export', origin='human',
        subject_type='tat_report', subject_id=str(filters['view']), actor=actor, authority_user=actor,
        deduplication_key=f'tat-report-export:{actor.pk}:{request_id}', before_values={}, after_values={},
        metadata={'view': filters['view'], 'filters': {key: str(value) for key, value in filters.items()}, 'fields': headers, 'row_count': len(rows)},
        sensitive=False,
    )
    return output.getvalue(), len(rows)


def enqueue_metric_rebuild(case, correction_revision, affected_dates):
    valid = [item for item in affected_dates if item]
    start = min(valid) if valid else timezone.localdate()
    end = timezone.localdate()
    return WorkflowTatMetricRebuildRequest.objects.get_or_create(
        request_key=f'case:{case.pk}:{correction_revision}',
        defaults={'case': case, 'correction_revision': correction_revision, 'date_from': start, 'date_to': end, 'next_date': start},
    )[0]


@transaction.atomic
def replace_metric_date(metric_date):
    from core.services.workflow_sla import collect_tat_daily_metrics, record_tat_daily_metrics
    metrics = collect_tat_daily_metrics(metric_date=metric_date)
    WorkflowTatDailyMetric.objects.filter(workflow='tat_tracker', metric_date=metric_date).delete()
    return record_tat_daily_metrics(metrics, metric_date=metric_date)


def process_metric_rebuilds(*, max_days=31):
    processed = 0
    while processed < max_days:
        with transaction.atomic():
            request = WorkflowTatMetricRebuildRequest.objects.select_for_update().filter(
                Q(status__in=['pending', 'processing']) | Q(status='failed', attempts__lt=3),
                next_date__lte=models_f('date_to'),
            ).first()
            if not request:
                break
            request.status = 'processing'; request.save(update_fields=['status', 'updated_at'])
            metric_date = request.next_date
        try:
            replace_metric_date(metric_date)
        except Exception as exc:
            WorkflowTatMetricRebuildRequest.objects.filter(pk=request.pk).update(
                status='failed', attempts=models_f('attempts') + 1,
                last_error=str(exc)[:500],
            )
            continue
        processed += 1
        next_date = metric_date + timedelta(days=1)
        updates = {'next_date': next_date, 'last_error': '', 'attempts': 0}
        if next_date > request.date_to:
            updates.update(status='complete', completed_at=timezone.now())
        else:
            updates['status'] = 'pending'
        WorkflowTatMetricRebuildRequest.objects.filter(pk=request.pk).update(**updates)
    return processed


def models_f(field):
    from django.db.models import F
    return F(field)
