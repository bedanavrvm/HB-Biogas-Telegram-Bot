from unittest.mock import MagicMock, patch
from datetime import timedelta
from decimal import Decimal
from io import BytesIO, StringIO
import hashlib
import hmac
import json
import time
from pathlib import Path
from urllib.parse import urlencode

import openpyxl
from django.contrib import admin
from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import connection
from django.test import RequestFactory, TestCase, override_settings
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone

from core.models import AccessGrant, BusinessCalendarHoliday, GroupSheetConfiguration, LiveSheetRecordChange, SheetRegisterContract, SheetSyncAuditSnapshot, TatActionTask, TatActionTaskRecipient, TatConfigurationEvent, TatEscalationRule, TatPresentationSettings, TatRepairJob, TatTrackerApprovalCertificate, TatTrackerCase, TatTrackerEvent, UserMiniAppPreference, UserProfile, WorkflowConfigurationChangeRequest, WorkflowRoleCapability, WorkflowSlaEscalation, WorkflowTatMetricRebuildRequest
from core.api.views import _dispatch_tat_approval_certificate, _process_telegram_message, tat_tracker_identity_context, tat_tracker_settings
from core.services.group_config import GroupConfig, GroupRegistry
from core.services.tat_tracker import (
    _TAT_HEADER_CACHE,
    apply_side_effects,
    bootstrap,
    build_tat_tracker_url,
    calculated_tat_days,
    calculated_tat_hours,
    calculated_tat_minutes,
    canonical_tat_status,
    create_tat_start_param,
    decode_tat_start_param,
    get_case_detail,
    product_by_key,
    parse_tat_batch_rows,
    parse_tat_batch_file,
    parse_iso_datetime,
    process_tat_batch_upload,
    process_tat_batch_file,
    previous_stages_complete,
    stage_by_key,
    stage_tat_minutes,
    stage_target_minutes_for_case,
    tat_days_formula,
    can_manage_tat_targets,
    normalize_tat_target_settings,
    update_tat_target_settings,
    tat_hours_formula,
    create_case,
    is_tat_tracker_workflow,
    home_data,
    next_role_alert,
    overall_tat_running,
    staff_user_for_payload,
    sync_case_to_sheet,
    sync_tat_batch_created_cases,
    resync_tat_tracker_cases,
    serialize_product,
    inspect_tat_sheet_duplicate_case_ids,
    cleanup_tat_sheet_duplicate_case_ids,
    _tat_sheet_call,
    search_cases,
    soft_delete_tat_case,
    sync_tat_target_settings_to_sheet,
    tat_batch_format_message,
    tat_case_identity_context,
    tat_reporting_status,
    validate_tracker_identity_headers,
    update_case,
    workflow_branches,
)
from core.services.workflow_transitions import WorkflowRevisionConflict
from core.services.workflow_sla import collect_sla_candidates, collect_tat_daily_metrics, record_sla_candidates
from core.services.miniapp_settings import create_tat_configuration_request, preference_payload, review_tat_configuration_request, update_preference
from core.services.tat_presentation import update_presentation_settings
from core.services.tat_reporting import export_report_xlsx, report_cases, report_summary
from core.services.workflow_capabilities import default_enabled_capability_keys
from core.services.sync_governance import assert_registered_schema_before_publish, audit_sheet_register


@override_settings(SECURE_SSL_REDIRECT=False)
class TatTrackerWorkflowTest(TestCase):
    def setUp(self):
        _TAT_HEADER_CACHE.clear()
        GroupRegistry._instance = None
        self.config = GroupSheetConfiguration.objects.create(
            group_id='-100tat',
            display_name='TAT Test',
            sheet_id='sheet123',
            sheet_name='TRACKER-Business',
            tat_sheet_projection_enabled=True,
            workflow={
                'type': 'tat_tracker',
                'products': ['business', 'logbook'],
                'branches': ['Nakuru', 'Embu'],
            },
        )
        User = get_user_model()
        self.bro_user = User.objects.create_user(username='bro-user', first_name='BRO', last_name='User', is_active=True)
        self.bro_user.set_unusable_password()
        self.bro_user.save(update_fields=['password'])
        UserProfile.objects.create(user=self.bro_user, telegram_id='111', telegram_username='bro_user')
        AccessGrant.objects.create(
            user=self.bro_user, workflow='tat_tracker', role='BRO',
            branch='Nakuru', product='business', group_configuration=self.config,
        )
        self.it_user = User.objects.create_user(username='it-user', first_name='IT', last_name='User', is_active=True)
        self.it_user.set_unusable_password()
        self.it_user.save(update_fields=['password'])
        UserProfile.objects.create(user=self.it_user, telegram_id='444', telegram_username='it_user')
        AccessGrant.objects.create(
            user=self.it_user, workflow='tat_tracker', role='IT',
            branch='Nakuru', product='business', group_configuration=self.config,
        )
        self.admin_user = User.objects.create_user(username='admin-user', first_name='Admin', last_name='User', is_active=True)
        self.admin_user.set_unusable_password()
        self.admin_user.save(update_fields=['password'])
        UserProfile.objects.create(user=self.admin_user, telegram_id='222', telegram_username='admin_user')
        AccessGrant.objects.create(
            user=self.admin_user, workflow='tat_tracker', role='BUSINESS_ADMIN',
            branch='Nakuru', product='business', group_configuration=self.config,
        )

    def signed_init_data(self, telegram_id='111', username='bro_user'):
        pairs = {
            'auth_date': str(int(time.time())),
            'user': json.dumps({'id': int(telegram_id), 'username': username}),
        }
        check = '\n'.join(f'{key}={value}' for key, value in sorted(pairs.items()))
        secret = hmac.new(b'WebAppData', b'test-bot-token', hashlib.sha256).digest()
        pairs['hash'] = hmac.new(secret, check.encode('utf-8'), hashlib.sha256).hexdigest()
        return urlencode(pairs)

    def test_detects_tat_tracker_workflow(self):
        self.assertTrue(is_tat_tracker_workflow(self.config))

    def test_new_tat_group_defaults_to_django_only_register(self):
        group = GroupSheetConfiguration.objects.create(
            group_id='-100django-only', display_name='Django only', sheet_id='',
            sheet_name='', workflow={'type': 'tat_tracker', 'branches': ['Nakuru']},
        )

        self.assertFalse(group.tat_sheet_projection_enabled)
        group.full_clean()

    def test_product_amount_limits_match_current_tat_policy(self):
        self.assertEqual(product_by_key('logbook').max_amount, Decimal('700000'))
        self.assertEqual(product_by_key('mjengo').min_amount, Decimal('10000'))
        self.assertEqual(product_by_key('mjengo').max_amount, Decimal('500000'))
        self.assertEqual(product_by_key('micro_asset').min_amount, Decimal('10000'))
        self.assertEqual(
            {
                key: value
                for key, value in serialize_product(product_by_key('logbook')).items()
                if key in {'min_amount', 'max_amount'}
            },
            {'min_amount': '50000.00', 'max_amount': '700000.00'},
        )
        self.assertEqual(serialize_product(product_by_key('business'))['max_amount'], '')

    def test_overdue_tat_stage_records_one_pending_follow_up_per_day(self):
        # SLA time is measured only during the official Nairobi business
        # calendar.  Keep this test inside a weekday business window rather
        # than coupling it to the wall-clock time at which the test runs.
        now = timezone.datetime(2026, 7, 27, 10, 0, tzinfo=timezone.get_current_timezone())
        self.config.workflow['tat_targets_minutes'] = {
            'business': {'stages': {'mpesa_to_admin': 1}},
        }
        self.config.save(update_fields=['workflow', 'updated_at'])
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-SLA',
            product_key='business',
            product_label='Business',
            client_name='SLA Client',
            branch='Nakuru',
            status='Active',
            stage_values={'created': (now - timedelta(minutes=5)).isoformat()},
        )

        candidates = collect_sla_candidates(workflow='tat_tracker', now=now)
        candidate = next(item for item in candidates if item.subject_id == str(case.pk))
        self.assertEqual(candidate.stage_key, 'mpesa_to_admin')
        self.assertGreater(candidate.overdue_minutes, 0)

        records, created = record_sla_candidates(candidates, today=timezone.localdate(now))
        _, repeated_created = record_sla_candidates(candidates, today=timezone.localdate(now))

        self.assertEqual(created, 1)
        self.assertEqual(repeated_created, 0)
        self.assertEqual(len(records), 1)
        self.assertEqual(WorkflowSlaEscalation.objects.filter(subject_id=str(case.pk)).count(), 1)

    def test_home_lists_paginate_independently(self):
        for index in range(30):
            TatTrackerCase.objects.create(
                group_id=self.config.group_id,
                case_id=f'JBL-BS-2026-{index:03d}',
                product_key='business',
                product_label='Business',
                client_name=f'Client {index}',
                branch='Nakuru',
                status='Active',
                stage_values={'created': timezone.now().isoformat()},
            )
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        first_page = home_data(self.config, user, queue='role', page=1)
        second_page = home_data(self.config, user, queue='role', page=2)
        legacy_page = home_data(self.config, user, action_offset=10, recent_offset=10, page_size=10)

        self.assertEqual(len(first_page['items']), 10)
        self.assertEqual(first_page['pagination']['total'], 30)
        self.assertEqual(first_page['pagination']['pages'], 3)
        self.assertEqual(len(second_page['items']), 10)
        self.assertEqual(second_page['pagination']['page'], 2)
        self.assertEqual(second_page['pagination']['offset'], 10)
        self.assertEqual(len(legacy_page['action_required']), 10)
        self.assertEqual(first_page['pagination']['action_required']['total'], 30)
        self.assertTrue(first_page['pagination']['action_required']['has_more'])
        self.assertEqual(len(legacy_page['recent']), 10)

        oversized_page = home_data(self.config, user, queue='all', page=1, page_size=50)
        self.assertEqual(len(oversized_page['items']), 10)
        self.assertEqual(oversized_page['pagination']['page_size'], 10)

    def test_home_metrics_follow_filters_and_deduplicate_stalled_cases(self):
        self.config.workflow['tat_targets_minutes'] = {
            'business': {'stages': {'mpesa_to_admin': 30}},
        }
        self.config.save(update_fields=['workflow', 'updated_at'])
        now = timezone.now()
        cases = []
        for index, status in enumerate(['Active', 'Stalled', 'Active', 'Disbursed', 'Rejected', 'Declined']):
            cases.append(TatTrackerCase.objects.create(
                group_id=self.config.group_id,
                case_id=f'JBL-BS-METRIC-{index}',
                product_key='business', product_label='Business',
                client_name=f'Metric Client {index}', branch='Nakuru', status=status,
                stage_values={'created': (now - timedelta(minutes=120 if index in {1, 2} else 5)).isoformat()},
            ))
        task = TatActionTask.objects.create(
            case=cases[0], group_configuration=self.config,
            stage_key='mpesa_to_admin', stage_label='M-Pesa to Admin',
            responsible_role='BRO', case_revision=cases[0].workflow_revision,
        )
        TatActionTaskRecipient.objects.create(
            task=task, user=self.bro_user, kind=TatActionTaskRecipient.KIND_PRIMARY,
        )
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        result = home_data(self.config, user, queue='assigned', product_key='business', branch='Nakuru')

        self.assertEqual(result['metrics'], {
            'assigned': 1, 'role': 3, 'total': 6, 'completed': 3, 'stalled': 2,
        })
        self.assertEqual([item['case_id'] for item in result['items']], [cases[0].case_id])

    def test_bootstrap_includes_the_same_queue_contract_as_home(self):
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-BOOTSTRAP-001',
            product_key='business', product_label='Business',
            client_name='Bootstrap Client', branch='Nakuru', status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )

        data = bootstrap(self.config, {'id': 111, 'username': 'bro_user'})

        self.assertEqual(data['queue'], 'role')
        self.assertEqual(data['metrics']['total'], 1)
        self.assertEqual(data['visibility']['code'], 'cases_visible')
        self.assertEqual([item['case_id'] for item in data['items']], [case.case_id])

    def test_home_treats_blank_access_scope_as_all_branches_products_and_group(self):
        AccessGrant.objects.filter(user=self.bro_user, workflow='tat_tracker').delete()
        AccessGrant.objects.create(
            user=self.bro_user, workflow='tat_tracker', role='BRO',
            branch='', product='', group_configuration=None,
        )
        visible_case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-GLOBAL-SCOPE-001',
            product_key='logbook', product_label='Logbook',
            client_name='Global Scope Client', branch='Embu', status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        result = home_data(self.config, user, queue='all')

        self.assertEqual(result['visibility']['scoped_total'], 1)
        self.assertEqual(result['metrics']['total'], 1)
        self.assertEqual([item['case_id'] for item in result['items']], [visible_case.case_id])

    def test_home_explains_when_access_scope_excludes_existing_cases(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-OUTSIDE-SCOPE-001',
            product_key='logbook', product_label='Logbook',
            client_name='Outside Scope Client', branch='Embu', status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        result = home_data(self.config, user, queue='all')

        self.assertEqual(result['metrics']['total'], 0)
        self.assertEqual(result['visibility']['stored_group_total'], 1)
        self.assertEqual(result['visibility']['operational_group_total'], 1)
        self.assertEqual(result['visibility']['scoped_total'], 0)
        self.assertEqual(result['visibility']['code'], 'access_scope_excludes_cases')

    def test_home_explains_when_active_filters_hide_accessible_cases(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-FILTERED-001',
            product_key='business', product_label='Business',
            client_name='Filtered Client', branch='Nakuru', status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        result = home_data(
            self.config, user, queue='all', statuses=['Deferred'],
        )

        self.assertEqual(result['metrics']['total'], 0)
        self.assertEqual(result['visibility']['scoped_total'], 1)
        self.assertEqual(result['visibility']['code'], 'filters_exclude_cases')

    def test_home_supports_authorized_multi_select_filters(self):
        User = get_user_model()
        superuser = User.objects.create_superuser(
            username='tat-filter-root', email='tat-filter@example.invalid', password='unused-password',
        )
        UserProfile.objects.create(user=superuser, telegram_id='991', telegram_username='tat_filter_root')
        cases = [
            TatTrackerCase.objects.create(
                group_id=self.config.group_id, case_id='JBL-MULTI-001',
                product_key='business', product_label='Business', client_name='Business Active',
                branch='Nakuru', status='Active', stage_values={'created': timezone.now().isoformat()},
            ),
            TatTrackerCase.objects.create(
                group_id=self.config.group_id, case_id='JBL-MULTI-002',
                product_key='logbook', product_label='Logbook', client_name='Logbook Deferred',
                branch='Embu', status='Deferred', stage_values={'created': timezone.now().isoformat()},
            ),
            TatTrackerCase.objects.create(
                group_id=self.config.group_id, case_id='JBL-MULTI-003',
                product_key='logbook', product_label='Logbook Rejected', client_name='Excluded Status',
                branch='Embu', status='Rejected', stage_values={'created': timezone.now().isoformat()},
            ),
        ]
        user = staff_user_for_payload(self.config, {'id': 991, 'username': 'tat_filter_root'})

        result = home_data(
            self.config, user, queue='all',
            product_keys=['business', 'logbook'], branches=['Nakuru', 'Embu'],
            statuses=['Active', 'Deferred'],
        )

        self.assertEqual(result['metrics']['total'], 3)
        self.assertEqual({item['case_id'] for item in result['items']}, {case.case_id for case in cases})

    def test_home_rejects_multi_select_values_outside_authorized_scope(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='JBL-MULTI-SCOPE',
            product_key='business', product_label='Business', client_name='Scoped Client',
            branch='Nakuru', status='Active', stage_values={'created': timezone.now().isoformat()},
        )
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        result = home_data(
            self.config, user, queue='all', product_keys=['business', 'logbook'], branches=['Nakuru'],
        )

        self.assertEqual(result['metrics']['total'], 0)
        self.assertEqual(result['items'], [])

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token')
    def test_home_api_accepts_multi_select_filter_arrays(self):
        for suffix, status in [('ACTIVE', 'Active'), ('REJECTED', 'Rejected'), ('DEFERRED', 'Deferred')]:
            TatTrackerCase.objects.create(
                group_id=self.config.group_id, case_id=f'JBL-MULTI-API-{suffix}',
                product_key='business', product_label='Business', client_name=f'{status} Client',
                branch='Nakuru', status=status, stage_values={'created': timezone.now().isoformat()},
            )

        response = self.client.post(
            reverse('tat_tracker_home'),
            data=json.dumps({
                'group_id': self.config.group_id,
                'init_data': self.signed_init_data(),
                'queue': 'all',
                'product_keys': ['business'],
                'branches': ['Nakuru'],
                'statuses': ['Active', 'Rejected'],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()['data']
        self.assertEqual(payload['metrics']['total'], 3)
        self.assertEqual(
            {item['case_id'] for item in payload['items']},
            {'JBL-MULTI-API-ACTIVE', 'JBL-MULTI-API-REJECTED', 'JBL-MULTI-API-DEFERRED'},
        )

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token')
    def test_home_api_caps_queue_pages_at_ten_cases(self):
        for index in range(12):
            TatTrackerCase.objects.create(
                group_id=self.config.group_id,
                case_id=f'JBL-PAGE-CAP-{index:02d}',
                product_key='business', product_label='Business',
                client_name=f'Page Client {index}', branch='Nakuru', status='Active',
                stage_values={'created': timezone.now().isoformat()},
            )

        response = self.client.post(
            reverse('tat_tracker_home'),
            data=json.dumps({
                'group_id': self.config.group_id,
                'init_data': self.signed_init_data(),
                'queue': 'all',
                'page': 1,
                'page_size': 50,
                'client_request_id': 'tat-page-cap-request',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()['data']
        self.assertEqual(len(payload['items']), 10)
        self.assertEqual(payload['pagination']['page_size'], 10)
        self.assertEqual(payload['pagination']['pages'], 2)

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token')
    def test_home_fragment_renders_recent_cases(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-001',
            product_key='business',
            product_label='Business',
            client_name='Fragment Client',
            branch='Nakuru',
            status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )

        response = self.client.post(
            reverse('tat_tracker_home_fragment'),
            {
                'group_id': self.config.group_id,
                'init_data': self.signed_init_data(),
                'list': 'recent',
                'product_key': 'business',
                'branch': 'Nakuru',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'tat_tracker/partials/case_list.html')
        self.assertContains(response, 'Fragment Client')
        self.assertContains(response, 'htmx-tat-case-card')

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token')
    def test_search_fragment_renders_matching_cases(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-002',
            product_key='business',
            product_label='Business',
            client_name='Searchable Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )

        response = self.client.post(
            reverse('tat_tracker_search_fragment'),
            {
                'group_id': self.config.group_id,
                'init_data': self.signed_init_data(),
                'query': 'Searchable',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Searchable Client')
        self.assertContains(response, 'JBL-BS-2026-002')

    def test_home_data_filters_by_product_and_branch(self):
        cases = [
            ('JBL-BS-2026-001', 'business', 'Business', 'Nakuru', 'Business Nakuru'),
            ('JBL-BS-2026-002', 'business', 'Business', 'Embu', 'Business Embu'),
            ('JBL-LB-2026-001', 'logbook', 'Logbook', 'Nakuru', 'Logbook Nakuru'),
        ]
        for case_id, product_key, product_label, branch, client_name in cases:
            TatTrackerCase.objects.create(
                group_id=self.config.group_id,
                case_id=case_id,
                product_key=product_key,
                product_label=product_label,
                client_name=client_name,
                branch=branch,
                status='Active',
                stage_values={'created': timezone.now().isoformat()},
            )
        user = {
            'name': 'IT User',
            'roles': ['IT'],
            'branches': ['Nakuru', 'Embu'],
            'products': ['business', 'logbook'],
        }

        filtered = home_data(self.config, user, product_key='business', branch='Nakuru')

        self.assertEqual(filtered['pagination']['recent']['total'], 1)
        self.assertEqual(filtered['recent'][0]['case_id'], 'JBL-BS-2026-001')
        self.assertTrue(all(item['product_key'] == 'business' for item in filtered['recent']))
        self.assertTrue(all(item['branch'] == 'Nakuru' for item in filtered['recent']))

    def test_soft_deleted_cases_are_hidden_from_mini_app_lists(self):
        deleted_case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-DEL',
            product_key='business',
            product_label='Business',
            client_name='Deleted Client',
            branch='Nakuru',
            status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        active_case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-ACT',
            product_key='business',
            product_label='Business',
            client_name='Active Client',
            branch='Nakuru',
            status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        changed = soft_delete_tat_case(
            deleted_case,
            actor_name='Admin User',
            actor_role='BUSINESS_ADMIN',
            reason='Duplicate test data cleanup.',
        )

        self.assertTrue(changed)
        deleted_case.refresh_from_db()
        self.assertTrue(deleted_case.is_deleted)
        self.assertEqual(deleted_case.deleted_by, 'Admin User')
        self.assertEqual(
            TatTrackerEvent.objects.filter(case=deleted_case, stage_key='deleted').count(),
            1,
        )

        home = home_data(self.config, user)
        home_ids = {item['case_id'] for item in home['recent']}
        self.assertNotIn(deleted_case.case_id, home_ids)
        self.assertIn(active_case.case_id, home_ids)

        search_results = search_cases(self.config, user, 'Deleted')
        self.assertEqual(search_results, [])
        with self.assertRaises(TatTrackerCase.DoesNotExist):
            get_case_detail(self.config, user, deleted_case.case_id)

    def test_tat_mini_app_sends_queue_filters_with_home_pagination(self):
        source = Path('core/static/miniapp/tat_tracker.js').read_text(encoding='utf-8')
        template = Path('core/templates/tat_tracker/app.html').read_text(encoding='utf-8')

        self.assertIn('id="queueProductFilters"', template)
        self.assertIn('id="queueBranchFilters"', template)
        self.assertIn('id="queueStatusFilters"', template)
        self.assertIn('class="filter-checkbox-grid"', template)
        self.assertIn("miniapp/utils.js", template)
        self.assertIn("product_keys: checkedFilterValues('queueProductFilters')", source)
        self.assertIn("branches: checkedFilterValues('queueBranchFilters')", source)
        self.assertIn("statuses: checkedFilterValues('queueStatusFilters')", source)
        self.assertIn("api('/api/tat-tracker/home/', homePayload({", source)
        self.assertIn('queue: requestedQueue', source)
        self.assertIn("queue: state.homeQueue", source)
        self.assertIn("page: state.homePages[state.homeQueue] || 1", source)
        self.assertIn('utils.fetchJson(path', source)

    def test_dm_task_launch_is_consumed_and_queue_navigation_does_not_reload_it(self):
        source = Path('core/static/miniapp/tat_tracker.js').read_text(encoding='utf-8')
        template = Path('core/templates/tat_tracker/app.html').read_text(encoding='utf-8')

        self.assertIn('function consumeTaskLaunchUrl()', source)
        self.assertIn("const keys = ['startapp', 'start_param', 'tgWebAppStartParam'];", source)
        self.assertIn("window.history.replaceState(window.history.state, '', url.toString())", source)
        self.assertIn("$('backBtn').addEventListener('click', returnToQueue)", source)
        self.assertIn("refresh({ background: true }).catch(() => {})", source)
        self.assertIn('Assigned to me', template)
        self.assertIn('data-home-queue="role"', template)
        self.assertIn('miniapp/tat_tracker.js', template)
        self.assertIn("miniapp/tat_tracker.js' %}?v=76", template)

    def test_compact_home_has_filter_sheet_metrics_and_explicit_pagination(self):
        source = Path('core/static/miniapp/tat_tracker.js').read_text(encoding='utf-8')
        stylesheet = Path('core/static/miniapp/tat_tracker.css').read_text(encoding='utf-8')
        template = Path('core/templates/tat_tracker/app.html').read_text(encoding='utf-8')
        case_list_template = Path('core/templates/tat_tracker/partials/case_list.html').read_text(encoding='utf-8')

        for label in ['Current Workload', 'Period Performance', 'Near Target']:
            self.assertIn(label, template)
        self.assertIn('Stalled (Overdue)', source)
        self.assertIn('queueFilterOverlay', template)
        self.assertIn('queuePreviousBtn', template)
        self.assertIn('queueNextBtn', template)
        self.assertIn('function applyQueueFilters()', source)
        self.assertIn('function trapFilterSheetFocus(event)', source)
        self.assertNotIn('show_business_hours_time', source)
        self.assertNotIn('preferenceBusinessHours', template)
        self.assertNotIn('holidaySettingsForm', template)
        self.assertNotIn('name="tenor"', source)
        self.assertIn('Object.assign({ page_size: 10 }', source)
        self.assertIn('class="form-grid new-case-grid"', template)
        self.assertIn('class="new-case-full">Client Name', template)
        self.assertIn('.new-case-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }', stylesheet)
        self.assertIn('.target-input-grid {', stylesheet)
        self.assertIn('grid-template-columns: repeat(2, minmax(0, 1fr));', stylesheet)
        self.assertIn('class="case-number"', source)
        self.assertIn('class="case-side"', source)
        self.assertIn('Number(state.home.pagination.offset || 0)', source)
        self.assertIn('class="case-number">#{{ forloop.counter }}', case_list_template)
        self.assertIn('class="case-side"', case_list_template)
        self.assertIn('.tat-sheet-overlay', stylesheet)
        self.assertIn('class="notice-close tat-sheet-close"', template)
        self.assertIn('grid-template-columns: minmax(0, 1fr) 44px', stylesheet)
        self.assertIn("miniapp/tat_tracker.css' %}?v=51", template)
        self.assertIn('id="appHeader" class="app-top"', template)
        self.assertIn('class="refresh-label"', template)
        self.assertIn('function bindCollapsingHeader()', source)
        self.assertIn("header.classList.add('header-hidden')", source)
        self.assertIn('bindCollapsingHeader();', source)
        self.assertIn('.app-top.header-hidden', stylesheet)
        self.assertIn('@media (prefers-reduced-motion: reduce)', stylesheet)
        self.assertIn('id="homeQueueTabs"', template)
        self.assertIn('id="workspaceTabs"', template)
        self.assertIn('id="dashboardView"', template)
        self.assertIn('id="casesWorkspaceBtn"', template)
        self.assertIn('id="dashboardWorkspaceBtn"', template)
        self.assertIn('id="tatReportMetrics"', template)
        self.assertIn("show('dashboard');", source)
        self.assertIn('.home-queue-tabs {', stylesheet)
        self.assertIn('grid-template-columns: repeat(3, minmax(0, 1fr));', stylesheet)
        self.assertIn('grid-template-columns: minmax(0, 1fr) auto minmax(0, 1fr);', stylesheet)
        self.assertIn('.case-primary {', stylesheet)
        self.assertIn('gap: 7px;', stylesheet)
        self.assertIn('text-align: left;', stylesheet)
        self.assertIn('function renderHomeQueueSelection(queue, loading)', source)
        self.assertIn('renderHomeQueueSelection(queue, true)', source)
        self.assertIn('refresh({ requestedQueue: queue, forceHomeRender: true })', source)
        self.assertIn('queue: requestedQueue', source)
        self.assertIn('if (forceHomeRender || queueRenderIsSafe())', source)
        self.assertIn("if (state.currentView === 'queue')", source)
        self.assertIn('.queue-pagination', stylesheet)
        self.assertNotIn('.tat-business-time', stylesheet)
        self.assertIn('.check-label[hidden]', stylesheet)

    def test_tat_report_ui_reuses_vendored_grid_and_compact_correction_actions(self):
        source = Path('core/static/miniapp/tat_tracker.js').read_text(encoding='utf-8')
        stylesheet = Path('core/static/miniapp/tat_tracker.css').read_text(encoding='utf-8')
        template = Path('core/templates/tat_tracker/app.html').read_text(encoding='utf-8')

        self.assertIn('vendor-ag-grid-community-36.1.0.min.js', template)
        self.assertIn('vendor-chartjs-4.5.1.umd.min.js', template)
        self.assertIn('data-required-capability="tat.reports.view"', template)
        self.assertIn("actionWrap.classList.add('correction-open')", source)
        self.assertIn('.stage-action-wrap.correction-open { grid-template-columns:repeat(2,minmax(0,1fr)); }', stylesheet)
        self.assertIn('state.report.abortController?.abort()', source)
        self.assertIn("suppressMovableColumns: touch", source)
        self.assertIn('function formatTatDateTime(value)', source)
        self.assertIn('data-date-display="date_from"', template)
        self.assertIn("headerName: '#', colId: 'row_number', pinned: 'left', lockPinned: true", source)
        self.assertIn("{ headerName: 'Reference', field: 'case_id', width: 112, minWidth: 92", source)
        self.assertNotIn("field: 'case_id', pinned: 'left'", source)
        self.assertIn('id="tatBacklogChart"', template)
        self.assertIn('id="tatSlaChart"', template)
        self.assertIn('id="tatPercentilesChart"', template)
        self.assertIn('id="tatTargetChart"', template)
        self.assertIn('id="tatExplorerChart"', template)
        self.assertIn('id="tatProgressionChart"', template)
        self.assertIn('data-report-key="case_progression"', template)
        self.assertIn('data-tat-chart-key="case_progression" data-tat-chart-type="line"', template)
        self.assertIn('data-tat-chart-key="backlog_age" data-tat-chart-type="pie"', template)
        self.assertIn('id="tatHeatmap"', template)
        self.assertIn('id="tatTargetSignals"', template)
        self.assertIn('id="tatOldestCases"', template)
        self.assertIn('data-chart-display="carousel"', template)
        self.assertIn("localStorage.setItem('tat-report-chart-display'", source)
        self.assertIn("recordTatCarouselGesture('gesture_started')", source)
        self.assertIn("if (event.target.closest?.('.tat-heatmap')) return;", source)
        self.assertIn('function tatChartExplanation(payload)', source)
        self.assertIn('Question: ${payload.question}', source)
        self.assertIn('How to read it: ${payload.interpretation}', source)
        self.assertIn("return parts.filter(Boolean).join('\\n');", source)
        self.assertIn('.tat-report-charts .chart-basis{white-space:pre-line}', stylesheet)
        self.assertIn('data-heat-row', source)
        self.assertIn("summary.metric_basis || ''", source)
        self.assertIn("text: payload.axis_title || '% of target'", source)
        self.assertIn("'branch', 'product', 'stage', 'role', 'status', 'sla_state'", source)
        self.assertIn("'date_from', 'date_to', 'granularity'", source)
        self.assertIn('No target performance data is available for this selection.', source)
        self.assertIn('<option value="">All Branches</option>', template)
        self.assertNotIn('<option value="">Any Branch</option>', template)
        self.assertIn('id="tatReportFilterOverlay"', template)
        self.assertIn('id="openTatReportFiltersBtn"', template)
        self.assertIn('data-report-key="trend"', template)
        self.assertIn('data-report-filter="branch"', template)
        self.assertNotIn('data-report-filter="group"', template)
        self.assertIn('class="native-date-icon"', template)
        self.assertIn('id="tatReportLoadingState"', template)
        self.assertIn('id="tatReportSheetLoading"', template)
        self.assertIn('id="tatReportInitialLoading"', template)
        self.assertIn('id="tatReportInitialRetry"', template)
        self.assertIn('<option>Active</option><option>Stalled</option><option>Declined</option><option>Disbursed</option>', template)
        self.assertNotIn('<option>Pending Docs</option>', template)
        self.assertNotIn('<option>Deferred</option>', template)
        self.assertNotIn('<option>Rejected</option>', template)
        self.assertIn('function bindReportDatePickers()', source)
        self.assertIn("typeof input.showPicker !== 'function'", source)
        self.assertIn('function setTatReportLoading(loading)', source)
        self.assertIn("setTatReportInitialState('ready')", source)
        self.assertIn('function compactTatReportLabel(value)', source)
        self.assertIn("['active', 'Active', ''], ['within_target', 'Within Target', 'good']", source)
        self.assertIn("contextualReportError('The TAT report could not be updated'", source)
        self.assertIn("'/api/tat-tracker/update/': 'Saving the case update'", source)
        self.assertIn('throw contextualTatApiError(path, error)', source)
        self.assertIn('function syncReportFilterGuidance()', source)
        self.assertIn('filter_guidance', source)
        self.assertIn('not affect ${insight.title}', source)
        self.assertIn('.tat-report-charts .chart-empty[hidden]{display:none!important}', stylesheet)
        self.assertIn('.tat-report-filter-bar{', stylesheet)
        self.assertIn('label[data-guidance="unavailable"]', stylesheet)
        self.assertIn('.tat-report-pagination span{position:static;grid-column:2', stylesheet)
        self.assertIn('.tat-report-loading,.tat-report-sheet-loading{', stylesheet)
        self.assertIn('.tat-report-initial-loading{position:fixed', stylesheet)
        self.assertIn('.report-metrics{display:grid;grid-template-columns:repeat(5,minmax(0,1fr))', stylesheet)
        self.assertIn('.tat-report-charts .tat-chart-body{position:relative;height:300px}', stylesheet)
        self.assertIn('.tat-report-charts .tat-chart-body{height:270px}', stylesheet)
        self.assertIn('@media(min-width:701px){.tat-report-charts{width:max(100%,50vw)', stylesheet)
        self.assertIn('.tat-insight-chart-toggle button.active{', stylesheet)
        self.assertIn('.tat-report-grid .ag-cell-value{display:block;max-width:100%;overflow:hidden;text-overflow:ellipsis', stylesheet)
        self.assertIn('function stageTatColumns(stages)', source)
        self.assertIn("colId: `stage_tat__${stage.key}`", source)
        self.assertIn("syncTatReportStageColumns(table.stage_columns || [])", source)
        self.assertIn('.tat-report-grid .tat-stage-tat-within_target', stylesheet)
        self.assertIn('.tat-report-grid .tat-stage-tat-near_target', stylesheet)
        self.assertIn('.tat-report-grid .tat-stage-tat-overdue', stylesheet)
        self.assertIn('Press and hold any table cell to copy its displayed value.', template)
        self.assertIn('function bindTatReportCellCopyHold()', source)
        self.assertIn("navigator.clipboard.writeText(value)", source)
        self.assertIn('timer: setTimeout(async () => {', source)
        self.assertIn('.tat-report-grid .ag-cell.tat-cell-copy-holding{', stylesheet)
        self.assertIn('.tat-report-grid .ag-cell.tat-cell-copied{', stylesheet)
        self.assertIn("const statusColors = { active: '#3390ec', stalled: '#ef9b36', declined: '#e45858', disbursed: '#23a67a' }", source)
        self.assertIn('const categoryGridlineLimit = 12;', source)
        self.assertIn('const numericGridlineLimit = 14;', source)
        self.assertIn('maxTicksLimit: horizontal ? numericGridlineLimit : categoryGridlineLimit', source)
        self.assertIn('maxTicksLimit: horizontal ? categoryGridlineLimit : numericGridlineLimit', source)
        self.assertIn('.tat-report-grid .tat-status-active{', stylesheet)
        self.assertIn('.tat-report-grid .tat-status-stalled{', stylesheet)
        self.assertIn('.tat-report-grid .tat-status-declined{', stylesheet)
        self.assertIn('.tat-report-grid .tat-status-disbursed{', stylesheet)
        self.assertIn('.tat-report-filters{display:grid;grid-template-columns:repeat(3,minmax(0,1fr))', stylesheet)
        self.assertIn('@media(max-width:700px){.tat-report-filters{grid-template-columns:repeat(3,minmax(0,1fr))}', stylesheet)
        self.assertIn('.report-metrics{grid-template-columns:repeat(5,minmax(0,1fr));gap:4px}', stylesheet)
        self.assertIn("response_mode: 'focused_v1'", source)
        self.assertIn('const TAT_REPORT_CHART_TYPES = {', source)
        self.assertIn("localStorage.setItem('tat-report-chart-types'", source)
        self.assertIn('function loadTatReportInsight(insight)', source)
        self.assertIn('Date.now() - cached.storedAt < 60000', source)
        self.assertIn('filterRevision: 0, insightCache: new Map()', source)
        self.assertIn('.tat-report-charts article.insight-loading{', stylesheet)
        self.assertIn('.tat-report-charts .tat-heatmap{height:300px}', stylesheet)
        self.assertIn('.tat-report-charts .tat-heatmap{height:260px}', stylesheet)
        self.assertIn('.tat-heatmap td{min-width:64px;padding:1px', stylesheet)

    def test_tat_reporting_is_scoped_allowlisted_and_page_size_capped(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-REPORT-1', product_key='business',
            product_label='Business', client_name='VISIBLE CLIENT', national_id='12345678',
            primary_phone='254700000000', branch='Nakuru', status='Active',
            stage_values={'created': timezone.now().isoformat()},
            stage_target_snapshots={'mpesa_to_admin': {'target_minutes': '60'}},
        )
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-REPORT-2', product_key='business',
            product_label='Business', client_name='OUTSIDE CLIENT', national_id='87654321',
            primary_phone='254711111111', branch='Embu', status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        data = report_cases(self.bro_user, {'view': 'current', 'page_size': 999})
        self.assertEqual(data['page_size'], 100)
        self.assertEqual(data['calculation_path'], 'database_paginated')
        self.assertEqual([row['case_id'] for row in data['results']], ['TAT-REPORT-1'])
        self.assertNotIn('national_id', data['results'][0])
        self.assertNotIn('primary_phone', data['results'][0])
        self.assertNotIn('stage_values', data['results'][0])
        self.assertEqual(data['results'][0]['group'], 'TAT Test')
        summary = report_summary(self.bro_user, {'view': 'current'})
        self.assertEqual(summary['metrics']['active'], 1)
        self.assertEqual(summary['metrics']['within_target'], 1)
        self.assertEqual(summary['filters']['groups'], [(self.config.group_id, 'TAT Test')])
        for payload in [
            *summary['charts'].values(), summary['heatmap'],
            summary['target_review_signals'], summary['oldest_cases'],
        ]:
            self.assertTrue(payload['question'])
            self.assertTrue(payload['interpretation'])
            self.assertTrue(payload['scope_note'].startswith('Scope:'))
            self.assertTrue(payload['subtitle'].startswith('Question:'))
            self.assertIn('How to read it:', payload['subtitle'])
        trend_guidance = summary['charts']['trend']['filter_guidance']
        self.assertIn('branch', trend_guidance['applicable_filters'])
        self.assertIn('stage', trend_guidance['basis_changing_filters'])
        self.assertIn('granularity', trend_guidance['chart_controls'])
        described_controls = (
            set(trend_guidance['applicable_filters'])
            | set(trend_guidance['basis_changing_filters'])
            | set(trend_guidance['chart_controls'])
            | {item['key'] for item in trend_guidance['unavailable_filters']}
        )
        self.assertEqual(described_controls, {
            'search', 'branch', 'product', 'stage', 'role', 'status', 'sla_state',
            'date_from', 'date_to', 'granularity', 'chart_dimension', 'chart_metric',
            'heatmap_pair', 'heatmap_metric',
        })
        backlog_guidance = summary['charts']['backlog_age']['filter_guidance']
        backlog_unavailable = {item['key']: item['reason'] for item in backlog_guidance['unavailable_filters']}
        self.assertIn('date_from', backlog_unavailable)
        self.assertIn('chart_metric', backlog_unavailable)
        self.assertIn('current workload', backlog_unavailable['date_from'])
        performance = report_summary(self.bro_user, {
            'view': 'performance',
            'date_from': timezone.localdate().isoformat(),
            'date_to': timezone.localdate().isoformat(),
        })
        self.assertEqual(performance['breakdown_basis'], 'created_cases_current_stage')
        self.assertTrue(performance['by_stage'])
        self.assertTrue(performance['by_role'])
        self.assertEqual(
            performance['charts']['explorer']['filter_guidance']['chart_controls'],
            ['chart_dimension', 'chart_metric'],
        )
        self.assertEqual(
            performance['heatmap']['filter_guidance']['chart_controls'],
            ['heatmap_pair', 'heatmap_metric'],
        )
        for payload in [
            *performance['charts'].values(), performance['heatmap'],
            performance['target_review_signals'], performance['oldest_cases'],
        ]:
            self.assertTrue(payload['question'])
            self.assertTrue(payload['interpretation'])
            self.assertIn('Scope:', payload['subtitle'])

        filtered = report_summary(self.bro_user, {
            'view': 'current', 'branch': 'Nakuru', 'product': 'business',
            'search': 'VISIBLE', 'chart_dimension': 'branch',
            'chart_metric': 'target_usage', 'heatmap_pair': 'stage_branch',
            'heatmap_metric': 'target_usage',
        })
        self.assertIn('branch Nakuru', filtered['charts']['explorer']['scope_note'])
        self.assertIn('product Business', filtered['charts']['explorer']['scope_note'])
        self.assertIn('search "VISIBLE"', filtered['charts']['explorer']['scope_note'])
        self.assertIn('100% line', filtered['charts']['explorer']['interpretation'])
        self.assertIn('above 100%', filtered['heatmap']['interpretation'])

        comparison_meanings = {
            'workload': 'most current cases',
            'sla_state': 'within target, near target, overdue',
            'duration': 'take longest',
            'target_usage': 'largest share',
            'sla_met': 'most consistently',
            'correction_rate': 'recorded corrections',
            'load_per_assignee': 'configured assignee',
        }
        for metric, expected in comparison_meanings.items():
            focused = report_summary(self.bro_user, {
                'view': 'current', 'response_mode': 'focused_v1',
                'insight': 'explorer', 'chart_dimension': 'branch',
                'chart_metric': metric,
            })
            explanation = focused['charts']['explorer']
            self.assertIn(expected, explanation['question'])
            self.assertNotIn('point-in-time', explanation['subtitle'])
            self.assertNotIn('IQR', explanation['subtitle'])

        heatmap_meanings = {
            'workload': 'most current cases',
            'duration': 'take longest',
            'target_usage': 'use or exceed',
            'sla_met': 'meet their targets most often',
        }
        for metric, expected in heatmap_meanings.items():
            focused = report_summary(self.bro_user, {
                'view': 'current', 'response_mode': 'focused_v1',
                'insight': 'heatmap', 'heatmap_pair': 'stage_branch',
                'heatmap_metric': metric,
            })
            self.assertIn(expected, focused['heatmap']['question'])

    def test_tat_report_focused_response_returns_only_requested_insight(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-FOCUSED-1', product_key='business',
            product_label='Business', client_name='FOCUSED CLIENT', branch='Nakuru', status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )

        focused = report_summary(self.bro_user, {
            'view': 'current', 'response_mode': 'focused_v1', 'insight': 'backlog_age',
            'include_overview': False, 'include_options': False,
        })

        self.assertEqual(focused['response_mode'], 'focused_v1')
        self.assertEqual(focused['insight'], 'backlog_age')
        self.assertEqual(set(focused['charts']), {'backlog_age'})
        self.assertNotIn('metrics', focused)
        self.assertNotIn('filters', focused)
        self.assertNotIn('freshness', focused)
        self.assertNotIn('heatmap', focused)
        self.assertNotIn('target_review_signals', focused)
        with self.assertRaisesMessage(ValueError, 'insight is not supported'):
            report_summary(self.bro_user, {
                'response_mode': 'focused_v1', 'insight': 'made_up_chart',
            })

    def test_tat_report_fast_table_query_count_is_bounded(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-QUERY-00', product_key='business',
            product_label='Business', client_name='QUERY CLIENT 00', branch='Nakuru', status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        with CaptureQueriesContext(connection) as small:
            small_result = report_cases(self.bro_user, {'view': 'current', 'page_size': 25})
        TatTrackerCase.objects.bulk_create([
            TatTrackerCase(
                group_id=self.config.group_id, case_id=f'TAT-QUERY-{index:02d}', product_key='business',
                product_label='Business', client_name=f'QUERY CLIENT {index:02d}', branch='Nakuru',
                status='Active', stage_values={'created': timezone.now().isoformat()},
            )
            for index in range(1, 21)
        ])
        with CaptureQueriesContext(connection) as large:
            large_result = report_cases(self.bro_user, {'view': 'current', 'page_size': 25})

        self.assertEqual(small_result['calculation_path'], 'database_paginated')
        self.assertEqual(large_result['calculation_path'], 'database_paginated')
        self.assertLessEqual(len(large), 20)
        self.assertLessEqual(len(large) - len(small), 2)

    def test_tat_report_branch_options_deduplicate_case_and_whitespace_variants(self):
        user = get_user_model().objects.create_superuser(username='branch-report-root', password='unused')
        for index, branch in enumerate(('Corporate', 'CORPORATE', '  corporate  '), start=1):
            TatTrackerCase.objects.create(
                group_id=self.config.group_id,
                case_id=f'TAT-BRANCH-OPTION-{index}',
                product_key='business',
                product_label='Business',
                client_name=f'BRANCH CLIENT {index}',
                branch=branch,
                status='Active',
                stage_values={'created': timezone.now().isoformat()},
            )

        summary = report_summary(user, {'view': 'current'})

        self.assertEqual(summary['filters']['branches'], ['Corporate'])

    def test_tat_reporting_exposes_chart_basis_backlog_and_frozen_target_percentages(self):
        now = timezone.now()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-REPORT-CHARTS', product_key='business',
            product_label='Business', client_name='CHART CLIENT', branch='Nakuru', status='Active',
            stage_values={
                'created': (now - timedelta(hours=2)).isoformat(),
                'mpesa_to_admin': (now - timedelta(hours=1)).isoformat(),
            },
            stage_target_snapshots={'mpesa_to_admin': {'target_minutes': '30', 'settings_version': 1}},
        )
        current = report_summary(self.bro_user, {'view': 'current'})
        self.assertEqual(current['charts']['backlog_age']['basis'], 'current_stage_wall_clock_age')
        self.assertIn('% of target', report_summary(self.bro_user, {
            'view': 'performance',
            'product': 'business',
            'date_from': timezone.localdate().isoformat(),
            'date_to': timezone.localdate().isoformat(),
        })['charts']['stage_target']['axis_title'])
        performance = report_summary(self.bro_user, {
            'view': 'performance', 'stage': 'mpesa_to_admin',
            'date_from': timezone.localdate().isoformat(),
            'date_to': timezone.localdate().isoformat(),
        })
        self.assertEqual(performance['charts']['trend']['basis'], 'completed_stage_actions')
        self.assertEqual(performance['charts']['trend']['series'][0]['key'], 'completed_actions')
        target = performance['charts']['stage_target']
        self.assertEqual(target['basis'], 'completed_actions_percent_of_frozen_target')
        self.assertEqual(target['series'][0]['values'], [200.0])
        self.assertEqual(target['reference_line'], 100)
        self.assertEqual(target['excluded_count'], 0)
        self.assertEqual(case.case_id, 'TAT-REPORT-CHARTS')

    def test_tat_reporting_explorer_heatmap_and_target_signals_use_exact_samples(self):
        User = get_user_model()
        super_user = User.objects.create_superuser(username='report-super', password='unused')
        now = timezone.now()
        cases = []
        for index in range(20):
            branch = 'Nakuru' if index < 10 else 'Embu'
            case = TatTrackerCase.objects.create(
                group_id=self.config.group_id, case_id=f'TAT-SIGNAL-{index:02d}',
                product_key='business', product_label='Business', client_name=f'CLIENT {index}',
                branch=branch, status='Active',
                stage_values={
                    'created': (now - timedelta(hours=2)).isoformat(),
                    'mpesa_to_admin': now.isoformat(),
                },
                stage_target_snapshots={
                    'mpesa_to_admin': {'target_minutes': '30', 'settings_version': 1},
                },
            )
            cases.append(case)
        for revision in range(2):
            TatTrackerEvent.objects.create(
                case=cases[0], group_id=self.config.group_id,
                stage_key='mpesa_to_admin', stage_label='MPESA sent to Admin',
                source='admin_correction', request_id=f'correction-{revision}',
            )

        payload = {
            'view': 'performance', 'chart_dimension': 'branch',
            'chart_metric': 'correction_rate', 'heatmap_pair': 'stage_branch',
            'heatmap_metric': 'target_usage',
            'date_from': timezone.localdate().isoformat(),
            'date_to': timezone.localdate().isoformat(),
        }
        report = report_summary(super_user, payload)
        explorer = report['charts']['explorer']
        self.assertEqual(explorer['basis'], 'distinct_completed_actions_with_recorded_admin_correction')
        self.assertEqual(dict(zip(explorer['labels'], explorer['series'][0]['values'])), {'Embu': 0.0, 'Nakuru': 10.0})
        self.assertEqual(report['heatmap']['metric'], 'target_usage')
        self.assertEqual(report['heatmap']['sample_count'], 20)
        signal = next(item for item in report['target_review_signals']['items'] if item['stage_key'] == 'mpesa_to_admin')
        self.assertEqual(signal['baseline']['valid_samples'], 20)
        self.assertEqual(signal['baseline']['over_percent'], 100.0)
        self.assertTrue(signal['baseline_systemic'])
        self.assertEqual(signal['classification'], 'review_recommended')

        narrowed = report_summary(super_user, {**payload, 'branch': 'Nakuru'})
        narrowed_signal = next(item for item in narrowed['target_review_signals']['items'] if item['stage_key'] == 'mpesa_to_admin')
        self.assertEqual(narrowed_signal['baseline']['valid_samples'], 20)
        self.assertEqual(narrowed_signal['selected_scope']['valid_samples'], 10)
        self.assertEqual(narrowed_signal['classification'], 'selected_scope_high')

    def test_tat_reporting_heatmap_uses_configured_stage_order_on_either_axis(self):
        now = timezone.now()
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-HEATMAP-SEQUENCE',
            product_key='business', product_label='Business', client_name='SEQUENCED CLIENT',
            branch='Nakuru', status='Active',
            stage_values={
                'created': (now - timedelta(hours=4)).isoformat(),
                'mpesa_to_admin': (now - timedelta(hours=3)).isoformat(),
                'mpesa_verified': (now - timedelta(hours=2)).isoformat(),
                'ca_analysis_sent': (now - timedelta(hours=1)).isoformat(),
            },
        )
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-HEATMAP-FIRST-STAGE',
            product_key='business', product_label='Business', client_name='FIRST STAGE CLIENT',
            branch='Nakuru', status='Active',
            stage_values={'created': (now - timedelta(minutes=30)).isoformat()},
        )

        stage_rows = report_summary(self.bro_user, {
            'view': 'current', 'heatmap_pair': 'stage_branch',
        })['heatmap']['rows']
        self.assertEqual(stage_rows[:3], [
            'MPESA sent to Admin',
            'MPESA verified by Business Admin and sent to CA',
            'Credit analysis sent',
        ])

        stage_columns = report_summary(self.bro_user, {
            'view': 'current', 'heatmap_pair': 'product_stage',
        })['heatmap']['columns']
        self.assertEqual(stage_columns[:3], stage_rows[:3])

        current_stage_rows = report_summary(self.bro_user, {
            'view': 'current', 'heatmap_pair': 'stage_branch',
            'heatmap_metric': 'workload',
        })['heatmap']['rows']
        self.assertEqual(current_stage_rows, ['MPESA sent to Admin', 'BRO response to CA'])

    def test_tat_report_rows_do_not_expose_internal_scope_keys(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-PUBLIC-ROW', product_key='business',
            product_label='Business', client_name='PUBLIC ROW', branch='Nakuru', status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        row = report_cases(self.bro_user, {'view': 'current'})['results'][0]
        self.assertFalse(any(key.startswith('_') for key in row))
        with self.assertRaisesMessage(ValueError, 'chart dimension'):
            report_summary(self.bro_user, {'chart_dimension': 'customer_name'})
        with self.assertRaisesMessage(ValueError, 'heatmap comparison'):
            report_summary(self.bro_user, {'heatmap_pair': 'stage_customer'})

    def test_tat_report_rows_include_each_stage_duration_and_frozen_target_state(self):
        now = timezone.now()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-STAGE-COLUMNS', product_key='business',
            product_label='Business', client_name='STAGE CLIENT', branch='Nakuru', status='Active',
            stage_values={
                'created': (now - timedelta(minutes=190)).isoformat(),
                'mpesa_to_admin': (now - timedelta(minutes=160)).isoformat(),
                'mpesa_verified': (now - timedelta(minutes=105)).isoformat(),
                'ca_analysis_sent': (now - timedelta(minutes=35)).isoformat(),
            },
            stage_target_snapshots={
                'mpesa_to_admin': {'target_minutes': '60', 'settings_version': 1},
                'mpesa_verified': {'target_minutes': '60', 'settings_version': 1},
                'ca_analysis_sent': {'target_minutes': '60', 'settings_version': 1},
            },
        )

        report = report_cases(self.bro_user, {'view': 'current'})
        self.assertEqual(report['results'][0]['case_id'], case.case_id)
        self.assertEqual(
            [item['key'] for item in report['stage_columns'][:3]],
            ['mpesa_to_admin', 'mpesa_verified', 'ca_analysis_sent'],
        )
        stage_tat = report['results'][0]['stage_tat']
        self.assertEqual(stage_tat['mpesa_to_admin']['minutes'], 30.0)
        self.assertEqual(stage_tat['mpesa_to_admin']['sla_state'], 'within_target')
        self.assertEqual(stage_tat['mpesa_verified']['minutes'], 55.0)
        self.assertEqual(stage_tat['mpesa_verified']['sla_state'], 'near_target')
        self.assertEqual(stage_tat['ca_analysis_sent']['minutes'], 70.0)
        self.assertEqual(stage_tat['ca_analysis_sent']['sla_state'], 'overdue')
        self.assertTrue(stage_tat['ca_analysis_sent']['completed'])
        self.assertTrue(stage_tat['bro_response']['active'])

    def test_tat_report_case_progression_requires_one_case_and_orders_stage_times(self):
        now = timezone.now()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-PROGRESSION-1', product_key='business',
            product_label='Business', client_name='PROGRESSION CLIENT', branch='Nakuru', status='Active',
            stage_values={
                'created': (now - timedelta(minutes=190)).isoformat(),
                'mpesa_to_admin': (now - timedelta(minutes=160)).isoformat(),
                'mpesa_verified': (now - timedelta(minutes=105)).isoformat(),
                'ca_analysis_sent': (now - timedelta(minutes=35)).isoformat(),
            },
            stage_target_snapshots={
                'mpesa_to_admin': {'target_minutes': '60', 'settings_version': 1},
                'mpesa_verified': {'target_minutes': '60', 'settings_version': 1},
                'ca_analysis_sent': {'target_minutes': '60', 'settings_version': 1},
                'bro_response': {'target_minutes': '45', 'settings_version': 1},
            },
        )

        focused = report_summary(self.bro_user, {
            'view': 'current', 'response_mode': 'focused_v1', 'insight': 'case_progression',
            'include_overview': False, 'include_options': False,
            'search': case.case_id, 'branch': 'Nakuru', 'product': 'business',
        })

        self.assertEqual(set(focused['charts']), {'case_progression'})
        progression = focused['charts']['case_progression']
        self.assertEqual(progression['selection_state'], 'single')
        self.assertEqual(progression['case_id'], case.case_id)
        self.assertEqual(
            progression['labels'][:4],
            [
                'MPESA sent to Admin',
                'MPESA verified by Business Admin and sent to CA',
                'Credit analysis sent',
                'BRO response to CA',
            ],
        )
        self.assertEqual(progression['series'][0]['values'][:3], [30.0, 55.0, 70.0])
        self.assertEqual(progression['series'][1]['values'][:4], [60.0, 60.0, 60.0, 45.0])
        self.assertEqual(progression['stage_statuses'][:4], ['completed', 'completed', 'completed', 'current'])
        self.assertIsNotNone(progression['series'][0]['values'][3])
        guidance = progression['filter_guidance']
        self.assertIn('search', guidance['applicable_filters'])
        self.assertIn('branch', guidance['applicable_filters'])
        self.assertIn('product', guidance['applicable_filters'])
        self.assertIn('stage', {item['key'] for item in guidance['unavailable_filters']})

        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-PROGRESSION-2', product_key='business',
            product_label='Business', client_name='PROGRESSION CLIENT', branch='Nakuru', status='Active',
            stage_values={'created': now.isoformat()},
        )
        ambiguous = report_summary(self.bro_user, {
            'view': 'current', 'response_mode': 'focused_v1', 'insight': 'case_progression',
            'include_overview': False, 'include_options': False,
            'search': 'PROGRESSION CLIENT', 'branch': 'Nakuru', 'product': 'business',
        })['charts']['case_progression']
        self.assertEqual(ambiguous['selection_state'], 'multiple')
        self.assertEqual(ambiguous['matched_case_count'], 2)
        self.assertIn('exact case reference', ambiguous['selection_message'])
        self.assertEqual(ambiguous['labels'], [])

    def test_tat_reports_use_four_exclusive_statuses_and_merge_negative_outcomes(self):
        now = timezone.now()
        cases = [
            TatTrackerCase.objects.create(
                group_id=self.config.group_id, case_id='TAT-STATUS-ACTIVE', product_key='business',
                product_label='Business', client_name='ACTIVE CLIENT', branch='Nakuru', status='Active',
                stage_values={'created': (now - timedelta(minutes=5)).isoformat()},
                stage_target_snapshots={'mpesa_to_admin': {'target_minutes': '30'}},
            ),
            TatTrackerCase.objects.create(
                group_id=self.config.group_id, case_id='TAT-STATUS-STALLED', product_key='business',
                product_label='Business', client_name='STALLED CLIENT', branch='Nakuru', status='Active',
                stage_values={'created': (now - timedelta(minutes=120)).isoformat()},
                stage_target_snapshots={'mpesa_to_admin': {'target_minutes': '30'}},
            ),
        ]
        for raw_status in ('Rejected', 'Deferred'):
            cases.append(TatTrackerCase.objects.create(
                group_id=self.config.group_id, case_id=f'TAT-STATUS-{raw_status.upper()}',
                product_key='business', product_label='Business', client_name=f'{raw_status} CLIENT',
                branch='Nakuru', status=raw_status,
                stage_values={
                    'created': (now - timedelta(minutes=180)).isoformat(),
                    'decision': raw_status,
                    'decision_ts': (now - timedelta(minutes=60)).isoformat(),
                },
            ))
        cases.append(TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-STATUS-DISBURSED', product_key='business',
            product_label='Business', client_name='DISBURSED CLIENT', branch='Nakuru', status='Disbursed',
            stage_values={
                'created': (now - timedelta(minutes=180)).isoformat(),
                'disbursement': (now - timedelta(minutes=30)).isoformat(),
            },
        ))

        current = report_cases(self.bro_user, {'view': 'current'})
        self.assertEqual(
            {row['case_id']: row['status'] for row in current['results']},
            {'TAT-STATUS-ACTIVE': 'Active', 'TAT-STATUS-STALLED': 'Stalled'},
        )
        stalled = report_cases(self.bro_user, {'view': 'current', 'status': 'Stalled'})
        self.assertEqual([row['case_id'] for row in stalled['results']], ['TAT-STATUS-STALLED'])

        period = {
            'view': 'performance',
            'date_from': timezone.localdate(now).isoformat(),
            'date_to': timezone.localdate(now).isoformat(),
        }
        performance = report_summary(self.bro_user, period)
        self.assertEqual(performance['metrics']['declined'], 2)
        self.assertEqual(performance['metrics']['disbursed'], 1)
        trend_keys = {series['key'] for series in performance['charts']['trend']['series']}
        self.assertIn('declined', trend_keys)
        self.assertIn('disbursed', trend_keys)
        self.assertNotIn('rejected', trend_keys)

        legacy_filter = report_cases(self.bro_user, {**period, 'status': 'Deferred'})
        self.assertEqual(legacy_filter['count'], 2)
        self.assertEqual({row['status'] for row in legacy_filter['results']}, {'Declined'})

    def test_tat_report_export_reuses_audited_workbook_for_selected_insights(self):
        now = timezone.now()
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-EXPORT-INSIGHT', product_key='business',
            product_label='Business', client_name='EXPORT CLIENT', branch='Nakuru', status='Active',
            stage_values={
                'created': (now - timedelta(minutes=190)).isoformat(),
                'mpesa_to_admin': (now - timedelta(minutes=160)).isoformat(),
                'mpesa_verified': (now - timedelta(minutes=105)).isoformat(),
                'ca_analysis_sent': (now - timedelta(minutes=35)).isoformat(),
            },
            stage_target_snapshots={
                'mpesa_to_admin': {'target_minutes': '60', 'settings_version': 1},
                'mpesa_verified': {'target_minutes': '60', 'settings_version': 1},
                'ca_analysis_sent': {'target_minutes': '60', 'settings_version': 1},
            },
        )
        content, count = export_report_xlsx(
            self.bro_user,
            {'view': 'current', 'chart_dimension': 'branch', 'heatmap_pair': 'stage_branch'},
            request_id='tat-export-selected-insight',
        )
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(count, 1)
        self.assertEqual(workbook.sheetnames, ['TAT Report', 'Selected Insight', 'Selected Heatmap'])
        self.assertIn('Branch', workbook['Selected Insight']['A1'].value)
        report_sheet = workbook['TAT Report']
        headers = [cell.value for cell in report_sheet[1]]
        expected = {
            'MPESA sent to Admin TAT (minutes)': (30, 'C6EFCE'),
            'MPESA verified by Business Admin and sent to CA TAT (minutes)': (55, 'FFEB9C'),
            'Credit analysis sent TAT (minutes)': (70, 'FFC7CE'),
        }
        for header, (minutes, color) in expected.items():
            column = headers.index(header) + 1
            cell = report_sheet.cell(row=2, column=column)
            self.assertEqual(cell.value, minutes)
            self.assertTrue(cell.fill.fgColor.rgb.endswith(color))
            self.assertIsNotNone(report_sheet.cell(row=1, column=column).comment)

    def test_daily_stage_duration_samples_only_belong_to_completion_date(self):
        now = timezone.now()
        completed = now - timedelta(days=1, hours=1)
        created = completed - timedelta(hours=1)
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-DAILY-COMPLETION', product_key='business',
            product_label='Business', client_name='DAILY CLIENT', branch='Nakuru', status='Active',
            stage_values={'created': created.isoformat(), 'mpesa_to_admin': completed.isoformat()},
            stage_target_snapshots={'mpesa_to_admin': {'target_minutes': '90', 'settings_version': 1}},
        )
        TatTrackerCase.objects.filter(pk=case.pk).update(created_at=created)
        completion_date = timezone.localdate(completed)
        completed_metrics = collect_tat_daily_metrics(metric_date=completion_date, now=now)
        later_metrics = collect_tat_daily_metrics(metric_date=timezone.localdate(now), now=now)
        matching_completed = [
            item for item in completed_metrics
            if item['workflow'] == 'tat_tracker' and item['metric_grain'] == 'stage_completion_leaf'
            and item['stage_key'] == 'mpesa_to_admin'
        ]
        matching_later = [
            item for item in later_metrics
            if item['workflow'] == 'tat_tracker' and item['metric_grain'] == 'stage_completion_leaf'
            and item['stage_key'] == 'mpesa_to_admin'
        ]
        self.assertEqual(matching_completed[0]['completed_count'], 1)
        self.assertEqual(matching_completed[0]['sample_count'], 1)
        self.assertEqual(matching_later, [])

    def test_snapshot_command_previous_day_is_explicit_and_exclusive(self):
        output = StringIO()
        with patch(
            'core.management.commands.snapshot_workflow_tat.collect_tat_daily_metrics',
            return_value=[],
        ) as collect:
            call_command('snapshot_workflow_tat', '--previous-day', stdout=output)
        self.assertEqual(collect.call_args.kwargs['metric_date'], timezone.localdate() - timedelta(days=1))
        with self.assertRaises(CommandError):
            call_command('snapshot_workflow_tat', '--previous-day', '--date', timezone.localdate().isoformat())

    def test_tat_report_capabilities_and_correction_rebuild_request(self):
        self.assertIn('tat.reports.view', default_enabled_capability_keys('tat_tracker', 'BRO'))
        self.assertNotIn('tat.reports.people.view', default_enabled_capability_keys('tat_tracker', 'BRO'))
        self.assertIn('tat.reports.people.view', default_enabled_capability_keys('tat_tracker', 'IT'))
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-CORRECTION-REPORT', product_key='business',
            product_label='Business', client_name='CORRECTION CLIENT', branch='Nakuru',
            status='Active', stage_values={'created': (timezone.now() - timedelta(days=2)).isoformat(), 'mpesa_to_admin': (timezone.now() - timedelta(days=1)).isoformat()},
        )
        user = staff_user_for_payload(self.config, {'id': 444, 'username': 'it_user'})
        update_case(
            self.config, user, case.case_id,
            [{'field': 'mpesa_to_admin', 'value': timezone.now().isoformat(), 'correction': True}],
            expected_revision=case.workflow_revision, request_id='report-correction-1',
        )
        self.assertEqual(WorkflowTatMetricRebuildRequest.objects.filter(case=case).count(), 1)

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token')
    def test_tat_report_endpoints_enforce_auth_capability_and_sort_allowlist(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id, case_id='TAT-API-REPORT', product_key='business',
            product_label='Business', client_name='API CLIENT', branch='Nakuru', status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        payload = {'group_id': self.config.group_id, 'init_data': self.signed_init_data(), 'view': 'current'}
        headers = {'X-MiniApp-Message-Contract': '2', 'X-Request-ID': 'tat-report-api-1', 'Idempotency-Key': 'tat-report-api-1'}
        payload['client_request_id'] = 'tat-report-api-1'
        response = self.client.post(reverse('tat_tracker_reports_cases'), data=json.dumps(payload), content_type='application/json', headers=headers)
        self.assertEqual(response.status_code, 200, response.content.decode())
        row = response.json()['data']['results'][0]
        self.assertNotIn('responsible_person', row)
        bad_headers = {'X-MiniApp-Message-Contract': '2', 'X-Request-ID': 'tat-report-api-2', 'Idempotency-Key': 'tat-report-api-2'}
        bad_sort = self.client.post(reverse('tat_tracker_reports_cases'), data=json.dumps({**payload, 'client_request_id': 'tat-report-api-2', 'sort': 'national_id'}), content_type='application/json', headers=bad_headers)
        self.assertEqual(bad_sort.status_code, 400)
        self.assertEqual(bad_sort.json()['code'], 'tat_report_invalid_filter')
        self.assertEqual(bad_sort.json()['message'], 'This report column cannot be sorted.')
        self.assertNotIn('We could not understand that request', bad_sort.json()['message'])
        self.assertEqual(bad_sort.json()['request_id'], 'tat-report-api-2')
        WorkflowRoleCapability.objects.filter(
            workflow='tat_tracker', role='BRO', capability_key='tat.reports.view',
        ).update(effect=WorkflowRoleCapability.EFFECT_DENY)
        denied_headers = {'X-MiniApp-Message-Contract': '2', 'X-Request-ID': 'tat-report-api-3', 'Idempotency-Key': 'tat-report-api-3'}
        denied = self.client.post(reverse('tat_tracker_reports_summary'), data=json.dumps({**payload, 'client_request_id': 'tat-report-api-3'}), content_type='application/json', headers=denied_headers)
        self.assertEqual(denied.status_code, 403)

    def test_queue_polling_uses_shared_visibility_runtime_and_health_feedback(self):
        source = Path('core/static/miniapp/tat_tracker.js').read_text(encoding='utf-8')
        runtime = Path('core/static/miniapp/runtime.js').read_text(encoding='utf-8')
        diagnostics = Path('core/static/miniapp/diagnostics.js').read_text(encoding='utf-8')
        template = Path('core/templates/tat_tracker/app.html').read_text(encoding='utf-8')

        self.assertIn('createVisibleInterval', runtime)
        self.assertIn('sharedRuntime.createVisibleInterval', diagnostics)
        self.assertIn("cancelled ? 'cancelled'", diagnostics)
        self.assertIn("xhr.__miniappDiagnosticAborted ? 'cancelled'", diagnostics)
        self.assertIn("refresh({ background: true, periodic: true })", source)
        self.assertIn('if (periodic && state.homeRequestsInFlight > 0)', source)
        self.assertIn('requestNumber !== state.homeRequestNumber', source)
        self.assertIn('state.pendingHome = nextHome', source)
        self.assertIn('refreshDetailBackground()', source)
        self.assertIn('state.pendingDetail = result.data', source)
        self.assertIn('Couldn’t refresh — showing data from', source)
        self.assertIn('id="queueFreshness"', template)

    def test_live_tat_counter_uses_server_anchor_and_monotonic_clock(self):
        source = Path('core/static/miniapp/tat_tracker.js').read_text(encoding='utf-8')
        runtime = Path('core/static/miniapp/runtime.js').read_text(encoding='utf-8')

        self.assertIn('data-elapsed-seconds', source)
        self.assertIn('data-calculated-at', source)
        self.assertIn('Updated from server', source)
        self.assertIn('performance.now()', runtime)
        self.assertIn('createServerClock', runtime)

    def test_responsibility_editor_hides_redundant_related_object_controls(self):
        from core.admin import TatResponsibilityBackupInline
        from core.models import TatResponsibilityAssignment

        root = get_user_model().objects.create_superuser(
            username='responsibility-root', email='root@example.invalid',
            password='test-password',
        )
        request = RequestFactory().get(reverse(
            'admin:core_tatresponsibilityassignment_add',
        ))
        request.user = root
        model_admin = admin.site._registry[TatResponsibilityAssignment]
        form = model_admin.get_form(request)(initial={
            'group_configuration': self.config.pk, 'branch': 'Nakuru',
            'role': 'BRO', 'product_key': 'business',
        })

        for field_name in ('group_configuration', 'primary_user'):
            widget = form.fields[field_name].widget
            self.assertFalse(widget.can_add_related)
            self.assertFalse(widget.can_change_related)
            self.assertFalse(widget.can_delete_related)
            self.assertFalse(widget.can_view_related)
        self.assertEqual(TatResponsibilityBackupInline.extra, 0)

    def test_responsibility_workspace_keeps_stage_overrides_advanced(self):
        template = Path(
            'core/templates/admin/core/tatresponsibilityassignment/change_list.html',
        ).read_text(encoding='utf-8')

        self.assertIn('<details class="rounded border', template)
        self.assertIn('Advanced stage overrides', template)
        self.assertIn('Default role rosters already cover normal routing.', template)

    def test_compact_cards_have_a_distinct_queue_hierarchy(self):
        source = Path('core/static/miniapp/tat_tracker.js').read_text(encoding='utf-8')
        stylesheet = Path('core/static/miniapp/tat_tracker.css').read_text(encoding='utf-8')
        template = Path('core/templates/tat_tracker/app.html').read_text(encoding='utf-8')

        self.assertIn('Compact queue preview on. Save my settings to keep it.', source)
        self.assertIn('applyPersonalPreference(result.data', source)
        self.assertIn('body.compact-cards .case-identifiers', stylesheet)
        self.assertIn('body.compact-cards .case-time', stylesheet)
        self.assertIn('Open a case for identifiers and timestamps.', template)

    @patch('core.services.tat_tracker.sync_tat_target_settings_to_sheet', return_value={'status': 'unavailable'})
    def test_it_can_save_stage_targets_in_minutes(self, sync_targets):
        user = {'roles': ['IT'], 'name': 'IT User'}

        result = update_tat_target_settings(self.config, user, {
            'business': {
                'total_minutes': '1440',
                'stages': {'mpesa_to_admin': '30'},
            },
            'logbook': {'total_minutes': '', 'stages': {}},
        })

        self.config.refresh_from_db()
        targets = self.config.workflow['tat_targets_minutes']['business']
        self.assertTrue(result['changed'])
        self.assertEqual(targets['total'], 1440)
        self.assertEqual(targets['stages']['mpesa_to_admin'], 30)
        sync_targets.assert_called_once()

    def test_admin_cannot_save_tat_targets(self):
        user = staff_user_for_payload(self.config, {'id': 222, 'username': 'admin_user'})

        self.assertFalse(can_manage_tat_targets(user))
        with self.assertRaisesRegex(ValueError, 'Only IT'):
            update_tat_target_settings(self.config, user, {})

    @override_settings(TELEGRAM_BOT_TOKEN='test-bot-token')
    def test_settings_endpoint_resolves_runtime_group_config_to_database_row(self):
        """The registry object is not a ForeignKey value for settings rows."""
        request = RequestFactory().post(
            reverse('tat_tracker_settings'),
            data=json.dumps({
                'group_id': self.config.group_id,
                'init_data': self.signed_init_data(),
            }),
            content_type='application/json',
        )

        response = tat_tracker_settings(request)

        self.assertEqual(response.status_code, 200, response.content.decode())
        payload = json.loads(response.content)
        self.assertTrue(payload['ok'])
        self.assertIn('configuration', payload['data'])
        self.assertEqual(payload['data']['account']['workflow'], 'tat_tracker')
        self.assertTrue(payload['data']['account']['roles'])
        self.assertEqual(payload['data']['configuration']['settings_version'], 1)

    def test_preference_and_tat_target_proposal_require_independent_approval(self):
        it_actor = staff_user_for_payload(self.config, {'id': 444, 'username': 'it_user'})
        admin_actor = staff_user_for_payload(self.config, {'id': 222, 'username': 'admin_user'})
        preference = update_preference(self.it_user, 'tat_tracker', {
            'default_screen': 'home', 'default_filters': {'branch': 'Nakuru'},
            'compact_cards': True, 'show_business_hours_time': False, 'alert_mode': 'quiet',
        })
        self.assertEqual(preference['alert_mode'], 'quiet')
        self.assertFalse(preference['show_business_hours_time'])
        self.assertTrue(UserMiniAppPreference.objects.filter(user=self.it_user, workflow='tat_tracker').exists())

        proposal = create_tat_configuration_request(self.config, it_actor,
            setting_key='tat_targets', reason='Align the stage SLA with the revised operating plan.',
            proposed={'business': {'total_minutes': '1440', 'stages': {'mpesa_to_admin': '30'}}, 'logbook': {'total_minutes': '', 'stages': {}}},
            request_id='tat-settings-proposal-1')
        self.config.refresh_from_db()
        self.assertEqual(proposal.status, WorkflowConfigurationChangeRequest.STATUS_PENDING)
        self.assertNotIn('tat_targets_minutes', self.config.workflow)
        self_approver = dict(it_actor)
        self_approver['roles'] = list(self_approver.get('roles') or []) + ['BUSINESS_ADMIN']
        self_approver['capabilities'] = list(self_approver.get('capabilities') or []) + ['tat.settings.targets.approve']
        with self.assertRaisesRegex(PermissionError, 'different authorised Business Admin'):
            review_tat_configuration_request(str(proposal.pk), self_approver, approve=True)

        reviewed = review_tat_configuration_request(str(proposal.pk), admin_actor, approve=True)
        self.config.refresh_from_db()
        self.assertEqual(reviewed.status, WorkflowConfigurationChangeRequest.STATUS_APPROVED)
        self.assertEqual(self.config.workflow['tat_targets_minutes']['business']['stages']['mpesa_to_admin'], 30)

    def test_stage_target_is_frozen_when_case_enters_stage(self):
        self.config.workflow['tat_targets_minutes'] = {'business': {'stages': {'mpesa_to_admin': 30}}}
        self.config.save(update_fields=['workflow', 'updated_at'])
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        detail = create_case(self.config, user, {
            'product_key': 'business', 'client_name': 'Snapshot Client', 'national_id': '12345678',
            'primary_phone': '0712345678', 'branch': 'Nakuru', 'bro_name': 'BRO User', 'amount': '10000',
            'creation_intent': 'new_loan', '_defer_sheet_sync': True,
        })
        case = TatTrackerCase.objects.get(case_id=detail['summary']['case_id'])
        self.config.workflow['tat_targets_minutes']['business']['stages']['mpesa_to_admin'] = 90
        self.config.save(update_fields=['workflow', 'updated_at'])
        stage = product_by_key('business').stages[0]
        self.assertEqual(stage_target_minutes_for_case(case, self.config.workflow, product_by_key('business'), stage), Decimal('30'))

    def test_tat_calendar_proposal_preserves_history_and_requires_authorised_review(self):
        historic_date = timezone.localdate() - timedelta(days=1)
        future_date = timezone.localdate() + timedelta(days=14)
        BusinessCalendarHoliday.objects.create(date=historic_date, name='Previous operational holiday')
        it_actor = staff_user_for_payload(self.config, {'id': 444, 'username': 'it_user'})
        admin_actor = staff_user_for_payload(self.config, {'id': 222, 'username': 'admin_user'})
        bro_actor = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        with self.assertRaisesRegex(PermissionError, 'cannot propose'):
            create_tat_configuration_request(
                self.config, bro_actor, setting_key='business_calendar',
                reason='Add public holiday for the next review cycle.',
                proposed={'holidays': []}, request_id='tat-calendar-denied',
            )
        with self.assertRaisesRegex(ValueError, 'Historical holidays cannot be changed'):
            create_tat_configuration_request(
                self.config, it_actor, setting_key='business_calendar',
                reason='Attempt to alter a historical calendar result.',
                proposed={'holidays': [{'date': historic_date.isoformat(), 'name': 'Changed history', 'active': True}]},
                request_id='tat-calendar-history',
            )

        request = create_tat_configuration_request(
            self.config, it_actor, setting_key='business_calendar',
            reason='Add public holiday for the next review cycle.',
            proposed={'holidays': [{'date': future_date.isoformat(), 'name': 'Future public holiday', 'active': True}]},
            request_id='tat-calendar-allowed',
        )
        self.assertIn(historic_date.isoformat(), {item['date'] for item in request.proposed_snapshot['holidays']})
        review_tat_configuration_request(str(request.pk), admin_actor, approve=True)
        self.assertTrue(BusinessCalendarHoliday.objects.get(date=future_date).active)
        self.assertEqual(BusinessCalendarHoliday.objects.get(date=historic_date).name, 'Previous operational holiday')

    def test_global_business_time_switch_is_superuser_only_audited_and_enforced(self):
        root = get_user_model().objects.create_superuser(
            username='tat-presentation-root', email='root@example.invalid', password='test-password',
        )
        row = TatPresentationSettings.objects.get(singleton=1)
        with self.assertRaisesRegex(PermissionError, 'active Superuser'):
            update_presentation_settings(
                actor=self.it_user, business_time_visible=False,
                reason='Hide the optional comparison during this rollout.',
                expected_revision=row.revision,
            )

        updated = update_presentation_settings(
            actor=root, business_time_visible=False,
            reason='Hide the optional comparison during this rollout.',
            expected_revision=row.revision,
        )

        self.assertFalse(updated.business_time_enabled)
        self.assertEqual(updated.revision, row.revision + 1)
        self.assertTrue(TatConfigurationEvent.objects.filter(
            action='tat.presentation.business_time.changed', actor=root,
        ).exists())
        with self.assertRaisesRegex(ValueError, 'global TAT presentation policy'):
            update_preference(self.it_user, 'tat_tracker', {
                'default_screen': 'home', 'compact_cards': False,
                'show_business_hours_time': True, 'alert_mode': 'immediate',
            })
        it_actor = staff_user_for_payload(self.config, {'id': 444, 'username': 'it_user'})
        with self.assertRaisesRegex(ValueError, 'unavailable'):
            create_tat_configuration_request(
                self.config, it_actor, setting_key='business_calendar',
                reason='Add a future public holiday for operational reporting.',
                proposed={'holidays': []}, request_id='calendar-disabled',
            )
        from core.services.miniapp_settings import tat_settings_payload
        settings_payload = tat_settings_payload(self.config, it_actor)
        self.assertEqual(settings_payload['holidays'], {})
        self.assertFalse(settings_payload['cards']['business_calendar']['can_propose'])

    def test_business_time_switch_cannot_hide_pending_calendar_proposal(self):
        root = get_user_model().objects.create_superuser(
            username='tat-pending-root', email='pending@example.invalid', password='test-password',
        )
        it_actor = staff_user_for_payload(self.config, {'id': 444, 'username': 'it_user'})
        future_date = timezone.localdate() + timedelta(days=21)
        create_tat_configuration_request(
            self.config, it_actor, setting_key='business_calendar',
            reason='Add a future public holiday before the next planning cycle.',
            proposed={'holidays': [{
                'date': future_date.isoformat(), 'name': 'Planned holiday', 'active': True,
            }]}, request_id='calendar-pending-toggle',
        )
        row = TatPresentationSettings.objects.get(singleton=1)

        with self.assertRaisesRegex(ValueError, 'Resolve the pending'):
            update_presentation_settings(
                actor=root, business_time_visible=False,
                reason='Hide business time after calendar governance is resolved.',
                expected_revision=row.revision,
            )

    def test_tat_escalation_proposal_is_scope_validated_and_applied_once_approved(self):
        it_actor = staff_user_for_payload(self.config, {'id': 444, 'username': 'it_user'})
        admin_actor = staff_user_for_payload(self.config, {'id': 222, 'username': 'admin_user'})
        with self.assertRaisesRegex(ValueError, 'configured branch'):
            create_tat_configuration_request(
                self.config, it_actor, setting_key='tat_escalation',
                reason='Route overdue cases through a controlled escalation path.',
                proposed={'rules': [{'threshold_percent': 100, 'routing_role': 'MANAGEMENT', 'branch': 'Unknown branch'}]},
                request_id='tat-escalation-invalid-scope',
            )

        request = create_tat_configuration_request(
            self.config, it_actor, setting_key='tat_escalation',
            reason='Route overdue cases through a controlled escalation path.',
            proposed={'rules': [
                {'threshold_percent': 100, 'routing_role': 'RESPONSIBLE_ROLE', 'branch': ''},
                {'threshold_percent': 150, 'routing_role': 'BRANCH_MANAGER', 'branch': 'Nakuru'},
            ]},
            request_id='tat-escalation-allowed',
        )
        review_tat_configuration_request(str(request.pk), admin_actor, approve=True)
        rules = TatEscalationRule.objects.filter(group_configuration=self.config, active=True)
        self.assertEqual(rules.count(), 2)
        self.assertTrue(rules.filter(branch='Nakuru', routing_role='BRANCH_MANAGER').exists())

    def test_personal_preference_rejects_ambiguous_boolean_text(self):
        with self.assertRaisesRegex(ValueError, 'Compact cards'):
            update_preference(self.it_user, 'tat_tracker', {
                'default_screen': 'home', 'compact_cards': 'sometimes', 'alert_mode': 'immediate',
            })

    def test_reading_default_personal_preference_does_not_create_a_row(self):
        payload = preference_payload(self.it_user, 'tat_tracker')

        self.assertEqual(payload['default_screen'], '')
        self.assertTrue(payload['show_business_hours_time'])
        self.assertFalse(UserMiniAppPreference.objects.filter(user=self.it_user, workflow='tat_tracker').exists())

    def test_target_minutes_must_be_whole_number(self):
        with self.assertRaisesRegex(ValueError, 'whole minutes'):
            normalize_tat_target_settings(self.config.workflow, {
                'business': {'total_minutes': '0.01', 'stages': {}},
                'logbook': {'total_minutes': '', 'stages': {}},
            })

    @patch('core.services.tat_tracker.get_sheets_service')
    def test_target_sync_creates_missing_support_tab(self, get_service):
        sheet = MagicMock()
        get_service.return_value.get_or_create_worksheet.return_value = sheet

        result = sync_tat_target_settings_to_sheet(self.config, {
            'products': ['business'],
            'tat_targets_minutes': {'business': {'total': 1440, 'stages': {'mpesa_to_admin': 30}}},
        })

        self.assertEqual(result['status'], 'synced')
        get_service.return_value.get_or_create_worksheet.assert_called_once_with('TAT TARGETS', rows=500, cols=4)
        sheet.batch_clear.assert_called_once_with(['A2:D500'])

    @override_settings(TAT_TRACKER_SIGNATURES_ENABLED=True)
    def test_sme_bm_certificate_blocks_the_next_stage_until_signed(self):
        User = get_user_model()
        signer = User.objects.create_user(username='bm-user', first_name='BM', last_name='User', is_active=True)
        signer.set_unusable_password()
        signer.save(update_fields=['password'])
        UserProfile.objects.create(
            user=signer, telegram_id='333', telegram_username='bm_user',
            signing_national_id='12345678', signing_phone_number='+254700000001',
        )
        AccessGrant.objects.create(
            user=signer, workflow='tat_tracker', role='BM',
            branch='Nakuru', product='business', group_configuration=self.config,
        )
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-2026-APPROVAL',
            product_key='business',
            product_label='Business',
            client_name='Approval Client',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={
                'created': timezone.now().isoformat(),
                'mpesa_to_admin': timezone.now().isoformat(),
                'mpesa_verified': timezone.now().isoformat(),
                'ca_analysis_sent': timezone.now().isoformat(),
                'bro_response': timezone.now().isoformat(),
                'bm_response': timezone.now().isoformat(),
            },
        )
        event = TatTrackerEvent.objects.create(case=case, group_id=case.group_id, stage_key='bm_response')
        certificate = TatTrackerApprovalCertificate.objects.create(
            case=case,
            event=event,
            staff_user=signer,
            signer_name='BM User',
            signer_telegram_id='333',
            signer_national_id='12345678',
            signer_phone_number='+254700000001',
            stage_key='bm_response',
            external_reference='TAT-test-bm-response-v1',
        )
        next_stage = stage_by_key(product_by_key('business'), 'bro_applied')

        self.assertFalse(previous_stages_complete(case, next_stage))

        certificate.status = 'signed'
        certificate.save(update_fields=['status'])

        self.assertTrue(previous_stages_complete(case, next_stage))

    @override_settings(TAT_TRACKER_SIGNATURES_ENABLED=False)
    @patch('core.models.TatTrackerApprovalCertificate.objects.filter')
    def test_signature_dispatch_is_disabled_by_default(self, certificate_filter):
        _dispatch_tat_approval_certificate('JBL-BS-2026-001', {'telegram_id': '333'})

        certificate_filter.assert_not_called()
    def test_sme_bm_certificate_does_not_block_when_signatures_are_disabled(self):
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-2026-SIGNATURES-OFF',
            product_key='business',
            product_label='Business',
            client_name='Approval Client',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={
                'created': timezone.now().isoformat(),
                'mpesa_to_admin': timezone.now().isoformat(),
                'mpesa_verified': timezone.now().isoformat(),
                'ca_analysis_sent': timezone.now().isoformat(),
                'bro_response': timezone.now().isoformat(),
                'bm_response': timezone.now().isoformat(),
            },
        )

        self.assertTrue(previous_stages_complete(case, stage_by_key(product_by_key('business'), 'bro_applied')))
    @override_settings(APP_BASE_URL='https://example.test')
    def test_builds_secure_tracker_url(self):
        url = build_tat_tracker_url(self.config.group_id)
        self.assertIn('https://example.test/tat-tracker/', url)
        self.assertIn('group_id=-100tat', url)
        self.assertIn('token=', url)


    @override_settings(APP_BASE_URL='https://example.test', TELEGRAM_BOT_USERNAME='testbot', TAT_TRACKER_MINI_APP_SHORT_NAME='tattracker')
    def test_tat_command_routes_to_mini_app_button(self):
        GroupRegistry._instance = None
        result = _process_telegram_message({
            'message_id': 900,
            'chat': {'id': self.config.group_id, 'type': 'supergroup', 'title': 'TAT Test'},
            'from': {'id': 111, 'first_name': 'BRO', 'last_name': 'User', 'username': 'bro_user'},
            'text': '@testbot /tat',
            'date': 1783920000,
        })

        self.assertEqual(result['status'], 'command')
        self.assertIn('TAT Tracker', result['reply_text'])
        button = result['reply_markup']['inline_keyboard'][0][0]
        self.assertEqual(button['text'], 'Open TAT Tracker Mini App')
        self.assertIn('url', button)
        self.assertNotIn('web_app', button)
        self.assertTrue(button['url'].startswith('https://t.me/testbot/tattracker?startapp='))

    @override_settings(TELEGRAM_BOT_USERNAME='testbot')
    def test_tatbatch_command_returns_batch_format(self):
        GroupRegistry._instance = None
        result = _process_telegram_message({
            'message_id': 901,
            'chat': {'id': self.config.group_id, 'type': 'supergroup', 'title': 'TAT Test'},
            'from': {'id': 111, 'first_name': 'BRO', 'last_name': 'User', 'username': 'bro_user'},
            'text': '@testbot /tatbatch',
            'date': 1783920000,
        })

        self.assertEqual(result['status'], 'command')
        self.assertIn('Attach an Excel .xlsx or CSV file', result['reply_text'])
        self.assertIn('Product, Client Name, National ID, Phone, Branch, Amount', result['reply_text'])

    def test_parse_tat_batch_rows_accepts_pipe_rows(self):
        rows = parse_tat_batch_rows(
            "product | client name | national id | phone | branch | amount\n"
            "business | Mary Wanjiku | 12345678 | 254712345678 | Nakuru | 25000"
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['payload']['product_key'], 'business')
        self.assertEqual(rows[0]['payload']['client_name'], 'Mary Wanjiku')

    def test_parse_tat_batch_csv_accepts_required_headers(self):
        rows = parse_tat_batch_file(
            'tat_batch.csv',
            (
                "Product,Client Name,National ID,Phone,Branch,Amount\n"
                "business,Mary Wanjiku,12345678,254712345678,Nakuru,25000\n"
            ).encode('utf-8'),
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['line_number'], 2)
        self.assertEqual(rows[0]['payload']['product_key'], 'business')
        self.assertEqual(rows[0]['payload']['primary_phone'], '254712345678')

    def test_parse_tat_batch_xlsx_accepts_required_headers(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Product', 'Client Name', 'National ID', 'Phone', 'Branch', 'Amount'])
        sheet.append(['business', 'Mary Wanjiku', '12345678', '254712345678', 'Nakuru', '25000'])
        stream = BytesIO()
        workbook.save(stream)

        rows = parse_tat_batch_file('tat_batch.xlsx', stream.getvalue())

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['line_number'], 2)
        self.assertEqual(rows[0]['payload']['client_name'], 'Mary Wanjiku')

    @patch('core.services.tat_tracker.sync_tat_batch_created_cases', return_value={'synced': 1, 'failed': []})
    def test_bro_can_upload_tat_batch_csv_file(self, sync_mock):
        result = process_tat_batch_file(
            self.config,
            filename='tat_batch.csv',
            content=(
                "Product,Client Name,National ID,Phone,Branch,Amount\n"
                "business,Mary Wanjiku,12345678,254712345678,Nakuru,25000\n"
            ).encode('utf-8'),
            user_payload={'id': 111, 'username': 'bro_user'},
            telegram_message_id='csv-1',
            sender='BRO User',
        )

        self.assertEqual(result['status'], 'tat_batch_processed')
        self.assertEqual(result['created'], 1)
        self.assertEqual(TatTrackerCase.objects.get().client_name, 'MARY WANJIKU')
        sync_mock.assert_called_once()

    @override_settings(TELEGRAM_BOT_USERNAME='testbot')
    @patch('core.services.tat_tracker.sync_tat_batch_created_cases', return_value={'synced': 1, 'failed': []})
    def test_bro_can_upload_tat_batch_with_batch_command(self, sync_mock):
        GroupRegistry._instance = None
        result = _process_telegram_message({
            'message_id': 902,
            'chat': {'id': self.config.group_id, 'type': 'supergroup', 'title': 'TAT Test'},
            'from': {'id': 111, 'first_name': 'BRO', 'last_name': 'User', 'username': 'bro_user'},
            'text': (
                '@testbot /batch\n'
                'business | Mary Wanjiku | 12345678 | 254712345678 | Nakuru | 25000'
            ),
            'date': 1783920000,
        })

        self.assertEqual(result['status'], 'tat_batch_processed')
        self.assertEqual(result['created'], 1)
        case = TatTrackerCase.objects.get(client_name='MARY WANJIKU')
        self.assertEqual(case.bro_name, 'BRO User')
        self.assertEqual(case.create_request_id, 'tat-batch:-100tat:902:1')
        sync_mock.assert_called_once()

    @patch('core.services.tat_tracker.sync_tat_batch_created_cases', return_value={'synced': 1, 'failed': []})
    def test_tat_batch_retry_is_idempotent(self, sync_mock):
        payload = (
            "business | Mary Wanjiku | 12345678 | 254712345678 | Nakuru | 25000"
        )

        first = process_tat_batch_upload(
            self.config,
            payload,
            user_payload={'id': 111, 'username': 'bro_user'},
            telegram_message_id='retry-1',
            sender='BRO User',
        )
        second = process_tat_batch_upload(
            self.config,
            payload,
            user_payload={'id': 111, 'username': 'bro_user'},
            telegram_message_id='retry-1',
            sender='BRO User',
        )

        self.assertEqual(first['created'], 1)
        self.assertEqual(second['duplicates'], 1)
        self.assertEqual(TatTrackerCase.objects.filter(client_name='MARY WANJIKU').count(), 1)

    def test_non_bro_cannot_upload_tat_batch(self):
        result = process_tat_batch_upload(
            self.config,
            "business | Mary Wanjiku | 12345678 | 254712345678 | Nakuru | 25000",
            user_payload={'id': 222, 'username': 'admin_user'},
            telegram_message_id='not-bro',
            sender='Admin User',
        )

        self.assertEqual(result['status'], 'command')
        self.assertIn('Only configured BRO users', result['reply_text'])

    @override_settings(STORAGES={
        'default': {'BACKEND': 'django.core.files.storage.FileSystemStorage'},
        'staticfiles': {'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage'},
    })
    def test_tat_app_preserves_signed_token_from_start_param(self):
        start_param = create_tat_start_param(self.config.group_id)
        token = decode_tat_start_param(start_param)['token']

        response = self.client.get('/tat-tracker/', {'tgWebAppStartParam': start_param})

        self.assertEqual(response.status_code, 200)
        html = response.content.decode('utf-8')
        self.assertIn(f'data-token="{token}"', html)
        self.assertNotIn('\\u003A', html)



    def test_workflow_branches_replace_stale_default_branch_list(self):
        stale_workflow = {
            'branches': ['Corporate', 'Thika Road', 'East Nairobi', 'West Nairobi', 'Nakuru', 'Embu', 'Limuru'],
        }

        self.assertEqual(
            workflow_branches(stale_workflow),
            ['Corporate', 'East Nairobi', 'West Nairobi', 'Thika Road', 'Limuru', 'Embu', 'Nakuru', 'Biogas Unit', 'Eco Conserve'],
        )

    @override_settings(TAT_TRACKER_BRANCH_CHOICES='Biogas Unit, Muranga, Thika Road')
    def test_workflow_branches_use_tat_env_override(self):
        workflow = {
            'branches': ['Nakuru', 'Embu'],
        }

        self.assertEqual(workflow_branches(workflow), ['Biogas Unit', 'Muranga', 'Thika Road'])

    @override_settings(TAT_TRACKER_BRANCH_CHOICES='Biogas Unit,Nakuru,Muranga')
    def test_bootstrap_filters_tat_env_branches_by_staff_access(self):
        data = bootstrap(self.config, {'id': 111, 'username': 'bro_user'})

        self.assertEqual(data['branches'], ['Nakuru'])
        self.assertEqual(data['bro_names'], ['BRO User'])
        self.assertEqual(data['bro_names'], ['BRO User'])

    def test_bootstrap_exposes_users_tagged_with_bro_role(self):
        data = bootstrap(self.config, {'id': 111, 'username': 'bro_user'})

        self.assertEqual(data['default_bro_user_id'], self.bro_user.pk)
        self.assertEqual(
            data['bro_users'],
            [{
                'id': self.bro_user.pk,
                'name': 'BRO User',
                'username': 'bro-user',
                'telegram_username': 'bro_user',
            }],
        )

    def test_bootstrap_lists_active_bros_across_scopes_for_every_creator(self):
        User = get_user_model()
        other_group = GroupSheetConfiguration.objects.create(
            group_id='-100tat-other', sheet_id='other-sheet', sheet_name='TAT Other',
            workflow={'type': 'tat_tracker'},
        )
        other_bro = User.objects.create_user(
            username='other-bro', first_name='Other', last_name='BRO', is_active=True,
        )
        UserProfile.objects.create(user=other_bro, telegram_id='333', telegram_username='other_bro')
        AccessGrant.objects.create(
            user=other_bro, workflow='tat_tracker', role='BRO',
            group_configuration=other_group,
        )
        inactive_bro = User.objects.create_user(
            username='inactive-bro', first_name='Inactive', last_name='BRO', is_active=False,
        )
        AccessGrant.objects.create(
            user=inactive_bro, workflow='tat_tracker', role='BRO',
            group_configuration=other_group,
        )

        bro_data = bootstrap(self.config, {'id': 111, 'username': 'bro_user'})
        admin_data = bootstrap(self.config, {'id': 222, 'username': 'admin_user'})

        self.assertEqual(
            {item['name'] for item in bro_data['bro_users']},
            {'BRO User', 'Other BRO'},
        )
        self.assertEqual(
            {item['name'] for item in admin_data['bro_users']},
            {'BRO User', 'Other BRO'},
        )
        self.assertEqual(bro_data['default_bro_user_id'], self.bro_user.pk)
        self.assertIsNone(admin_data['default_bro_user_id'])

    def test_tat_formula_helpers_match_tracker_columns(self):
        business = product_by_key('business')
        logbook = product_by_key('logbook')
        mjengo = product_by_key('mjengo')

        self.assertEqual(tat_hours_formula(business, 5), '=IF(OR($H5="",$R5=""),"",ROUND(($R5-$H5)*24,2))')
        self.assertEqual(tat_days_formula(business, 5), '=IF(U5="","",ROUND(U5/24,2))')
        self.assertEqual(tat_hours_formula(logbook, 5), '=IF(OR($H5="",$Z5=""),"",ROUND(($Z5-$H5)*24,2))')
        self.assertEqual(tat_days_formula(logbook, 5), '=IF(AC5="","",ROUND(AC5/24,2))')
        self.assertEqual(tat_hours_formula(mjengo, 5), '=IF(OR($H5="",$Y5=""),"",ROUND(($Y5-$H5)*24,2))')
        self.assertEqual(tat_days_formula(mjengo, 5), '=IF(AB5="","",ROUND(AB5/24,2))')
    def test_tat_access_comes_from_canonical_user_grants(self):
        group_config = type('GroupConfigLike', (), self.config.as_group_config_kwargs())()
        user = staff_user_for_payload(group_config, {'id': 111, 'username': 'bro_user'})

        self.assertTrue(user['authorized'])
        self.assertEqual(user['name'], 'BRO User')
        self.assertEqual(user['roles'], ['BRO'])
        self.assertEqual(user['branches'], ['Nakuru'])
        self.assertEqual(user['products'], ['business'])

    def test_group_admin_form_exposes_tat_targets_from_workflow(self):
        self.config.workflow.setdefault('tat_targets_minutes', {}).setdefault(
            'business',
            {'total': 20160, 'stages': {}},
        )['stages'] = {
            'mpesa_to_admin': 45,
            'ca_analysis_sent': 180,
        }
        self.config.save()

        from core.admin import GroupSheetConfigurationAdminForm

        form = GroupSheetConfigurationAdminForm(instance=self.config)

        self.assertIn('tat_target_business_total', form.fields)
        self.assertIn('tat_target_business_mpesa_to_admin', form.fields)
        self.assertEqual(form['tat_target_business_total'].value(), 20160)
        self.assertEqual(form['tat_target_business_mpesa_to_admin'].value(), 45)
        self.assertEqual(form['tat_target_business_ca_analysis_sent'].value(), 180)

    def test_group_admin_form_saves_tat_targets_into_generated_workflow(self):
        from core.admin import GroupSheetConfigurationAdminForm

        data = {
            'workflow_preset': 'tat_tracker',
            'group_id': self.config.group_id,
            'display_name': self.config.display_name,
            'enabled': 'on',
            'sheet_id': self.config.sheet_id,
            'sheet_name': self.config.sheet_name,
            'sheet_schema': '{}',
            'workflow': json.dumps(self.config.workflow),
            'parser_rules': '{}',
            'metadata': '{}',
            'tat_target_business_total': '1440',
            'tat_target_business_mpesa_to_admin': '30',
            'tat_target_business_ca_analysis_sent': '120',
        }

        form = GroupSheetConfigurationAdminForm(data=data, instance=self.config)

        self.assertTrue(form.is_valid(), form.errors)
        workflow = form.generated_workflow()
        self.assertEqual(workflow['tat_targets_minutes']['business']['total'], 1440)
        self.assertEqual(
            workflow['tat_targets_minutes']['business']['stages']['mpesa_to_admin'],
            30,
        )
        self.assertEqual(
            workflow['tat_targets_minutes']['business']['stages']['ca_analysis_sent'],
            120,
        )

    def test_group_admin_form_preserves_existing_tat_targets_when_fields_blank(self):
        self.config.workflow.setdefault('tat_targets_minutes', {}).setdefault(
            'business',
            {'total': 20160, 'stages': {}},
        )['stages'] = {
            'mpesa_to_admin': 45,
            'ca_analysis_sent': 180,
        }
        self.config.save()

        from core.admin import GroupSheetConfigurationAdminForm

        data = {
            'workflow_preset': 'tat_tracker',
            'group_id': self.config.group_id,
            'display_name': self.config.display_name,
            'enabled': 'on',
            'sheet_id': self.config.sheet_id,
            'sheet_name': self.config.sheet_name,
            'sheet_schema': '{}',
            'workflow': json.dumps(self.config.workflow),
            'parser_rules': '{}',
            'metadata': '{}',
        }

        form = GroupSheetConfigurationAdminForm(data=data, instance=self.config)

        self.assertTrue(form.is_valid(), form.errors)
        workflow = form.generated_workflow()
        self.assertEqual(
            workflow['tat_targets_minutes']['business']['stages']['mpesa_to_admin'],
            45,
        )
        self.assertEqual(
            workflow['tat_targets_minutes']['business']['stages']['ca_analysis_sent'],
            180,
        )

    def test_group_admin_form_preserves_existing_tat_workflow_settings(self):
        self.config.workflow.update({
            'products': ['business'],
            'branches': ['Muranga', 'Thika Road'],
            'alert_next_role': False,
        })
        self.config.save()

        from core.admin import GroupSheetConfigurationAdminForm

        data = {
            'workflow_preset': 'tat_tracker',
            'group_id': self.config.group_id,
            'display_name': self.config.display_name,
            'enabled': 'on',
            'sheet_id': self.config.sheet_id,
            'sheet_name': self.config.sheet_name,
            'sheet_schema': '{}',
            'workflow': json.dumps(self.config.workflow),
            'parser_rules': '{}',
            'metadata': '{}',
            'tat_target_business_total': '1440',
        }

        form = GroupSheetConfigurationAdminForm(data=data, instance=self.config)

        self.assertTrue(form.is_valid(), form.errors)
        workflow = form.generated_workflow()
        self.assertEqual(workflow['products'], ['business'])
        self.assertEqual(workflow['branches'], ['Muranga', 'Thika Road'])
        self.assertIs(workflow['alert_next_role'], False)
        self.assertEqual(workflow['tat_targets_minutes']['business']['total'], 1440)

    def test_group_admin_form_loads_tat_targets_even_when_preset_is_manual(self):
        self.config.workflow = {
            'type': 'custom_tat_tracker',
            'tat_targets_minutes': {
                'business': {
                    'total': 1440,
                    'stages': {'mpesa_to_admin': 30},
                },
            },
        }
        self.config.save()

        from core.admin import GroupSheetConfigurationAdminForm

        form = GroupSheetConfigurationAdminForm(instance=self.config)

        self.assertEqual(form['tat_target_business_total'].value(), 1440)
        self.assertEqual(form['tat_target_business_mpesa_to_admin'].value(), 30)

    def test_group_admin_manual_tat_workflow_merges_gui_target_fields(self):
        self.config.workflow = {
            'type': 'tat_tracker',
            'products': ['business'],
            'branches': ['Nakuru'],
            'tat_targets_minutes': {
                'business': {
                    'total': 20160,
                    'stages': {'mpesa_to_admin': 45},
                },
            },
        }
        self.config.save()

        from core.admin import GroupSheetConfigurationAdminForm

        data = {
            'workflow_preset': 'manual',
            'group_id': self.config.group_id,
            'display_name': self.config.display_name,
            'enabled': 'on',
            'sheet_id': self.config.sheet_id,
            'sheet_name': self.config.sheet_name,
            'sheet_schema': '{}',
            'workflow': json.dumps(self.config.workflow),
            'parser_rules': '{}',
            'metadata': '{}',
            'tat_target_business_total': '1440',
            'tat_target_business_ca_analysis_sent': '120',
        }

        form = GroupSheetConfigurationAdminForm(data=data, instance=self.config)

        self.assertTrue(form.is_valid(), form.errors)
        workflow = form.generated_workflow()
        self.assertEqual(workflow['products'], ['business'])
        self.assertEqual(workflow['branches'], ['Nakuru'])
        self.assertEqual(workflow['tat_targets_minutes']['business']['total'], 1440)
        self.assertEqual(
            workflow['tat_targets_minutes']['business']['stages']['mpesa_to_admin'],
            45,
        )
        self.assertEqual(
            workflow['tat_targets_minutes']['business']['stages']['ca_analysis_sent'],
            120,
        )

    def test_group_admin_change_form_accepts_tat_target_fieldset_fields(self):
        request = RequestFactory().get('/admin/core/groupsheetconfiguration/2/change/')
        request.user = get_user_model().objects.create_superuser(
            username='admin',
            email='admin@example.test',
            password='password',
        )

        model_admin = admin.site._registry[GroupSheetConfiguration]
        form_class = model_admin.get_form(request, self.config)

        self.assertIn('tat_target_business_total', form_class.base_fields)
        self.assertIn('tat_target_logbook_ca_analysis_sent', form_class.base_fields)
    def test_staff_user_matches_canonical_telegram_id(self):
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'someone_else'})
        self.assertTrue(user['authorized'])
        self.assertEqual(user['name'], 'BRO User')
        self.assertEqual(user['roles'], ['BRO'])

    @staticmethod
    def mark_case_synced(_group_config, case):
        case.row_number = case.row_number or 5
        case.sheet_name = case.sheet_name or 'TRACKER-Business'
        case.sync_error = ''
        case.save(update_fields=['row_number', 'sheet_name', 'sync_error', 'updated_at'])

    @override_settings(TAT_TRACKER_SYNC_SECONDARY_SHEETS=True)
    def test_sync_case_to_sheet_writes_django_calculated_tat_values(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []

            def row_values(self, _row):
                return [''] * 20

            def update(self, a1_range, values, value_input_option=None):
                self.updates.append((a1_range, values, value_input_option))

        class FakeService:
            def __init__(self, sheet):
                self._sheet = sheet

            def is_available(self):
                return True

        sheet = FakeSheet()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-001',
            product_key='business',
            product_label='Business',
            client_name='Test Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={
                'created': timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 0)).isoformat(),
                'disbursement': timezone.make_aware(timezone.datetime(2026, 7, 15, 14, 0)).isoformat(),
            },
            status='Active',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService(sheet)), \
             patch('core.services.tat_tracker.sync_case_index'), \
             patch('core.services.tat_tracker.sync_audit_log'):
            sync_case_to_sheet(self.config, case)

        self.assertEqual(sheet.updates[0][0], 'A5:AE5')
        self.assertEqual(len(sheet.updates[0][1][0]), 31)
        self.assertEqual(sheet.updates[0][1][0][2], '12345678')
        self.assertEqual(sheet.updates[0][1][0][3], '254712345678')
        self.assertEqual(sheet.updates[0][1][0][20], 30.0)
        self.assertEqual(sheet.updates[0][1][0][21], 1.25)
        self.assertFalse(any(str(value).startswith('=IF(') for value in sheet.updates[0][1][0]))

    @override_settings(GOOGLE_SHEETS_MAX_RETRIES=2, TAT_REPAIR_RETRY_BASE_SECONDS=0.25)
    @patch('core.services.tat_tracker.time.sleep')
    def test_tat_sheet_call_retries_google_quota_errors(self, sleep):
        operation = MagicMock(side_effect=[Exception('429 rate limit'), 'ok'])

        result = _tat_sheet_call(operation, description='test repair write')

        self.assertEqual(result, 'ok')
        self.assertEqual(operation.call_count, 2)
        sleep.assert_called_once_with(0.25)

    def test_sync_case_to_sheet_keeps_register_approval_tat_numeric(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []

            def row_values(self, row):
                if row == 2:
                    return [''] * 31
                values = [''] * 31
                values[29] = 'legacy TAT value'
                return values

            def update(self, a1_range, values, value_input_option=None):
                self.updates.append((a1_range, values, value_input_option))

        class FakeService:
            def __init__(self, sheet):
                self._sheet = sheet

            def is_available(self):
                return True

        registered_at = timezone.make_aware(timezone.datetime(2026, 7, 15, 9, 0))
        approved_at = timezone.make_aware(timezone.datetime(2026, 7, 15, 10, 0))
        sheet = FakeSheet()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-REGISTER-TAT',
            product_key='business',
            product_label='Business',
            client_name='Approval Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={
                'created': registered_at.isoformat(),
                'disbursement_register': '10:00am',
                'register_ts': registered_at.isoformat(),
                'register_approved': 'Approved',
            },
            status='Active',
        )
        event = TatTrackerEvent.objects.create(
            case=case,
            group_id=case.group_id,
            actor_name='Loan Approver',
            stage_key='register_approved',
            stage_label='Register approved',
            old_value='',
            new_value='Approved',
            source='mini_app',
        )
        TatTrackerEvent.objects.filter(pk=event.pk).update(created_at=approved_at)

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService(sheet)):
            sync_case_to_sheet(self.config, case)

        self.assertEqual(sheet.updates[0][1][0][29], 60.0)
        event.refresh_from_db()
        self.assertTrue(event.synced_to_sheet)
        self.assertIsNotNone(event.synced_at)
        self.assertEqual(event.sheet_name, 'TRACKER-Business')
        self.assertEqual(event.row_number, 5)
        self.assertEqual(event.sync_error, '')

    def test_primary_sheet_failure_keeps_event_unsynced_with_error(self):
        class UnavailableService:
            def is_available(self):
                return False

        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-2026-SYNC-FAIL',
            product_key='business',
            product_label='Business',
            client_name='Sync Failure Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )
        event = TatTrackerEvent.objects.create(
            case=case,
            group_id=case.group_id,
            stage_key='created',
            stage_label='Case Created',
            new_value='Created',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=UnavailableService()):
            with self.assertRaisesRegex(RuntimeError, 'Google Sheets service unavailable'):
                sync_case_to_sheet(self.config, case)

        event.refresh_from_db()
        self.assertFalse(event.synced_to_sheet)
        self.assertIsNone(event.synced_at)
        self.assertEqual(event.sync_error, 'Google Sheets service unavailable.')

    def test_sync_case_to_sheet_writes_mjengo_dropdown_values_not_stage_timestamps(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []

            def row_values(self, row):
                if row == 2:
                    return [''] * 43
                return []

            def update(self, a1_range, values, value_input_option=None):
                self.updates.append((a1_range, values, value_input_option))

        class FakeService:
            def __init__(self, sheet):
                self._sheet = sheet

            def is_available(self):
                return True

        sheet = FakeSheet()
        now = timezone.now().isoformat()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-MJENGO',
            row_number=5,
            case_id='JBL-MJ-2026-SHEET-DROPDOWN',
            product_key='mjengo',
            product_label='Mjengo',
            client_name='Sheet Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='100000',
            stage_values={
                'created': now,
                'mpesa_to_admin': now,
                'mpesa_verified': now,
                'ca_analysis_sent': now,
                'bro_response': now,
                'bm_tat_request': now,
                'tat_scheduled': now,
                'tat_held': now,
                'decision': 'Approved',
                'decision_ts': now,
                'minutes_shared': 'Yes',
                'minutes_shared_ts': now,
                'sanctions': 'Met',
                'sanctions_ts': now,
                'bro_applied': 'Met',
                'bro_applied_ts': now,
            },
            status='Active',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService(sheet)):
            sync_case_to_sheet(self.config, case)

        row = sheet.updates[0][1][0]
        self.assertEqual(row[17], 'Yes')
        self.assertEqual(row[20], 'Met')

    def test_sync_case_to_sheet_maps_legacy_mjengo_stage_timestamps_to_dropdown_values(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []

            def row_values(self, row):
                if row == 2:
                    return [''] * 43
                return []

            def update(self, a1_range, values, value_input_option=None):
                self.updates.append((a1_range, values, value_input_option))

        class FakeService:
            def __init__(self, sheet):
                self._sheet = sheet

            def is_available(self):
                return True

        sheet = FakeSheet()
        now = timezone.now().isoformat()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-MJENGO',
            row_number=5,
            case_id='JBL-MJ-2026-LEGACY-DROPDOWN',
            product_key='mjengo',
            product_label='Mjengo',
            client_name='Legacy Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='100000',
            stage_values={
                'created': now,
                'decision': 'Approved',
                'decision_ts': now,
                'minutes_shared': now,
                'sanctions': 'Met',
                'sanctions_ts': now,
                'bro_applied': now,
            },
            status='Active',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService(sheet)):
            sync_case_to_sheet(self.config, case)

        row = sheet.updates[0][1][0]
        self.assertEqual(row[17], 'Yes')
        self.assertEqual(row[20], 'Met')

    def test_completed_dropdowns_use_done_timeline_indicators(self):
        source = Path('core/static/miniapp/tat_tracker.js').read_text(encoding='utf-8')

        self.assertIn("'stage-row' + (hasValue ? ' done' : field.editable ? ' editable' : ' locked')", source)
        self.assertIn('if (hasValue) {\n        indicatorHtml = `<span class="indicator-icon check-done">', source)

    def test_sync_case_to_sheet_skips_secondary_sheets_by_default(self):
        class FakeSheet:
            def row_values(self, _row):
                return [''] * 20

            def update(self, *_args, **_kwargs):
                return None

        class FakeService:
            _sheet = FakeSheet()

            def is_available(self):
                return True

        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-003',
            product_key='business',
            product_label='Business',
            client_name='Test Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService()), \
             patch('core.services.tat_tracker.sync_case_index') as index_mock, \
             patch('core.services.tat_tracker.sync_audit_log') as audit_mock:
            sync_case_to_sheet(self.config, case)

        index_mock.assert_not_called()
        audit_mock.assert_not_called()

    def test_sync_case_to_sheet_appends_new_rows_after_case_id_scan(self):
        class FakeSheet:
            def __init__(self):
                self.appended = []
                self.row_values_calls = []
                self.col_values_called = False

            def row_values(self, row):
                self.row_values_calls.append(row)
                return [''] * 20

            def col_values(self, _col):
                self.col_values_called = True
                return ['Case ID']

            def append_row(self, row, value_input_option=None):
                self.appended.append((row, value_input_option))
                return {'updates': {'updatedRange': 'TRACKER-Business!A6:AC6'}}

        class FakeService:
            def __init__(self, sheet):
                self._sheet = sheet

            def is_available(self):
                return True

        sheet = FakeSheet()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-2026-005',
            product_key='business',
            product_label='Business',
            client_name='Test Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService(sheet)):
            sync_case_to_sheet(self.config, case)

        self.assertEqual(case.row_number, 6)
        self.assertEqual(len(sheet.appended), 1)
        self.assertEqual(sheet.row_values_calls, [2])
        self.assertTrue(sheet.col_values_called)

    def test_sync_case_to_sheet_follows_case_id_when_stored_row_is_stale(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []
                self.appended = []

            def row_values(self, row):
                if row == 2:
                    headers = [''] * 31
                    headers[2] = 'ID NUMBER'
                    headers[3] = 'PHONE NUMBER'
                    return headers
                return ['JBL-BS-2026-006'] + ['existing'] * 30

            def col_values(self, _col):
                return ['Case ID', '', '', '', 'OTHER-CASE', '', 'JBL-BS-2026-006']

            def update(self, a1_range, values, value_input_option=None):
                self.updates.append((a1_range, values, value_input_option))

            def append_row(self, row, value_input_option=None):
                self.appended.append((row, value_input_option))
                return {'updates': {'updatedRange': 'TRACKER-Business!A8:AE8'}}

        class FakeService:
            def __init__(self, sheet):
                self._sheet = sheet

            def is_available(self):
                return True

        sheet = FakeSheet()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-006',
            product_key='business',
            product_label='Business',
            client_name='Recovered Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService(sheet)):
            sync_case_to_sheet(self.config, case)

        case.refresh_from_db()
        self.assertEqual(case.row_number, 7)
        self.assertEqual(sheet.updates[0][0], 'A7:AE7')
        self.assertEqual(sheet.appended, [])

    def test_duplicate_case_rows_keep_the_most_populated_row(self):
        class FakeSheet:
            def __init__(self):
                self.deleted = []
                self.rows = [
                    ['title'],
                    ['headers'],
                    [],
                    [],
                    ['JBL-BS-2026-007', 'client', 'id'],
                    ['JBL-BS-2026-007', 'client', 'id', 'phone', 'branch', 'status'],
                    [],
                    ['JBL-BS-2026-008', 'client'],
                    ['JBL-BS-2026-008', 'client', 'id', 'phone'],
                ]

            def get_all_values(self):
                return [list(row) for row in self.rows]

            def delete_rows(self, row):
                self.deleted.append(row)
                self.rows.pop(row - 1)

        sheet = FakeSheet()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-007',
            product_key='business',
            product_label='Business',
            client_name='Duplicate Client',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )
        other_case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=8,
            case_id='JBL-BS-2026-008',
            product_key='business',
            product_label='Business',
            client_name='Second Duplicate Client',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )

        report = inspect_tat_sheet_duplicate_case_ids(sheet, group_id=self.config.group_id)
        self.assertEqual(len(report), 2)
        self.assertEqual(report[0]['keep_row'], 6)
        self.assertEqual(report[0]['delete_rows'], [5])
        self.assertEqual(report[1]['keep_row'], 9)
        self.assertEqual(report[1]['delete_rows'], [8])

        repaired = cleanup_tat_sheet_duplicate_case_ids(sheet, group_id=self.config.group_id, apply=True)
        case.refresh_from_db()
        other_case.refresh_from_db()
        self.assertEqual(sheet.deleted, [8, 5])
        self.assertEqual(case.row_number, 5)
        self.assertEqual(other_case.row_number, 7)
        self.assertEqual([item['verification_status'] for item in repaired], ['verified', 'verified'])
        self.assertEqual(LiveSheetRecordChange.objects.filter(status='success').count(), 2)

    def test_duplicate_cleanup_does_not_guess_a_row_when_post_delete_verification_fails(self):
        class FakeSheet:
            def get_all_values(self):
                return [
                    ['title'], ['headers'], [], [],
                    ['JBL-BS-2026-VERIFY', 'sparse'],
                    ['JBL-BS-2026-VERIFY', 'complete', 'id'],
                ]

            def delete_rows(self, row):
                # Simulate a provider-side failure that accepted the API call
                # but left the duplicate data present.
                del row

        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-VERIFY',
            product_key='business',
            product_label='Business',
            client_name='Verification Client',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )

        reports = cleanup_tat_sheet_duplicate_case_ids(FakeSheet(), group_id=self.config.group_id, apply=True)
        case.refresh_from_db()
        self.assertEqual(reports[0]['verification_status'], 'failed')
        self.assertEqual(case.row_number, 5)
        audit = LiveSheetRecordChange.objects.get(record_key=case.case_id)
        self.assertEqual(audit.status, 'failed')

    def test_duplicate_cleanup_republishes_only_after_immutable_id_verification(self):
        class FakeSheet:
            def __init__(self):
                self.rows = [
                    ['title'], ['headers'], [], [],
                    ['JBL-BS-2026-REPUBLISH', 'sparse'],
                    ['JBL-BS-2026-REPUBLISH', 'complete', 'id'],
                ]

            def get_all_values(self):
                return [list(row) for row in self.rows]

            def delete_rows(self, row):
                self.rows.pop(row - 1)

        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-REPUBLISH',
            product_key='business',
            product_label='Business',
            client_name='Republish Client',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )

        with patch('core.services.tat_tracker.sync_case_to_sheet') as republish:
            reports = cleanup_tat_sheet_duplicate_case_ids(
                FakeSheet(), group_id=self.config.group_id,
                group_configuration=self.config, actor='test-admin', apply=True,
            )
        case.refresh_from_db()
        self.assertEqual(reports[0]['verification_status'], 'verified')
        self.assertEqual(reports[0]['resync_status'], 'synced')
        self.assertEqual(case.row_number, 5)
        republish.assert_called_once()
        self.assertEqual(republish.call_args.args[1].row_number, 5)

    def test_register_audit_flags_schema_and_row_divergence_without_customer_values(self):
        class FakeSheet:
            def row_values(self, row):
                if row != 2:
                    raise AssertionError(f'Unexpected header row {row}')
                return ['Case ID', 'ID NUMBER', 'PHONE NUMBER']

            def get_all_values(self):
                return [
                    ['title'],
                    ['Case ID', 'ID NUMBER', 'PHONE NUMBER'],
                    [], [],
                    ['JBL-BS-2026-AUDIT', '87654321', '254712345678'],
                ]

        class FakeService:
            _sheet = FakeSheet()

            def is_available(self):
                return True

        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-AUDIT',
            product_key='business',
            product_label='Business',
            client_name='Audit Client',
            national_id='87654321',
            primary_phone='254712345678',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )
        contract = SheetRegisterContract.objects.create(
            group_configuration=self.config,
            register_key='tat_business',
            sheet_name='TRACKER-Business',
            subject_type=SheetRegisterContract.SUBJECT_TAT_CASE,
            header_row=2,
            data_start_row=5,
            row_key_header='Case ID',
            expected_headers=['Case ID', 'ID NUMBER', 'PHONE NUMBER'],
            field_ownership={
                'Case ID': {'owner': 'immutable', 'model_field': 'case_id'},
                'ID NUMBER': {'owner': 'immutable', 'model_field': 'national_id', 'comparison': 'digits'},
                'PHONE NUMBER': {'owner': 'immutable', 'model_field': 'primary_phone', 'comparison': 'digits'},
            },
        )

        with patch('core.services.sync_governance.get_sheets_service', return_value=FakeService()):
            healthy = audit_sheet_register(contract, checked_by='test-admin')
        self.assertEqual(healthy['status'], 'healthy')
        self.assertEqual(SheetSyncAuditSnapshot.objects.get().rows_checked, 1)

        case.row_number = 7
        case.save(update_fields=['row_number', 'updated_at'])
        with patch('core.services.sync_governance.get_sheets_service', return_value=FakeService()):
            divergent = audit_sheet_register(contract, checked_by='test-admin')
        self.assertEqual(divergent['status'], 'divergence')
        self.assertEqual(divergent['discrepancies'][0]['kind'], 'row_pointer_mismatch')
        self.assertNotIn('87654321', str(divergent['discrepancies']))

    def test_tat_sync_refuses_a_registered_sheet_schema_that_has_drifted(self):
        class FakeSheet:
            def row_values(self, row):
                if row == 2:
                    return ['Changed Case ID', '', 'ID NUMBER', 'PHONE NUMBER']
                raise AssertionError('Schema guard should stop the sync before any case row is read.')

        class FakeService:
            _sheet = FakeSheet()

            def is_available(self):
                return True

        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-2026-SCHEMA',
            product_key='business',
            product_label='Business',
            client_name='Schema Client',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )
        SheetRegisterContract.objects.create(
            group_configuration=self.config,
            register_key='tat_business_schema',
            sheet_name='TRACKER-Business',
            subject_type=SheetRegisterContract.SUBJECT_TAT_CASE,
            header_row=2,
            data_start_row=5,
            row_key_header='Case ID',
            expected_headers=['Case ID', 'ID NUMBER', 'PHONE NUMBER'],
            field_ownership={
                'Case ID': {'owner': 'immutable', 'model_field': 'case_id'},
                'ID NUMBER': {'owner': 'immutable', 'model_field': 'national_id'},
                'PHONE NUMBER': {'owner': 'immutable', 'model_field': 'primary_phone'},
            },
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService()):
            with self.assertLogs('core.services.tat_tracker', level='ERROR'):
                with self.assertRaisesMessage(ValueError, 'blocks publication'):
                    sync_case_to_sheet(self.config, case)
        case.refresh_from_db()
        self.assertIn('blocks publication', case.sync_error)

    def test_registered_schema_accepts_runtime_group_config_for_the_same_group(self):
        """Mini App sync passes GroupConfig, not the admin model instance."""
        SheetRegisterContract.objects.create(
            group_configuration=self.config,
            register_key='tat_business_runtime_config',
            sheet_name='TRACKER-Business',
            subject_type=SheetRegisterContract.SUBJECT_TAT_CASE,
            header_row=2,
            data_start_row=5,
            row_key_header='Case ID',
            expected_headers=['Case ID', 'ID NUMBER'],
            field_ownership={
                'Case ID': {'owner': 'immutable', 'model_field': 'case_id'},
                'ID NUMBER': {'owner': 'immutable', 'model_field': 'national_id'},
            },
        )

        runtime_config = GroupConfig(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name=self.config.sheet_name,
            workflow=self.config.workflow,
        )

        assert_registered_schema_before_publish(
            runtime_config,
            'TRACKER-Business',
            ['Case ID', 'ID NUMBER'],
        )

    def test_sync_tat_batch_created_cases_appends_same_product_in_one_sheet_write(self):
        class FakeSheet:
            def __init__(self):
                self.row_values_calls = []
                self.appended_rows = []

            def row_values(self, row):
                self.row_values_calls.append(row)
                headers = [''] * 31
                headers[2] = 'ID NUMBER'
                headers[3] = 'PHONE NUMBER'
                return headers

            def append_rows(self, rows, value_input_option=None):
                self.appended_rows.append((rows, value_input_option))
                return {'updates': {'updatedRange': 'TRACKER-Business!A5:AE6'}}

        class FakeService:
            def __init__(self, sheet):
                self._sheet = sheet

            def is_available(self):
                return True

        sheet = FakeSheet()
        created_at = timezone.now().isoformat()
        case_one = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-2026-101',
            product_key='business',
            product_label='Business',
            client_name='First Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={'created': created_at},
            status='Active',
        )
        case_two = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-2026-102',
            product_key='business',
            product_label='Business',
            client_name='Second Client',
            national_id='22345678',
            primary_phone='254722345678',
            branch='Embu',
            bro_name='BRO User',
            amount='20000',
            stage_values={'created': created_at},
            status='Active',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService(sheet)):
            result = sync_tat_batch_created_cases(self.config, [case_one, case_two])

        self.assertEqual(result, {'synced': 2, 'failed': []})
        self.assertEqual(sheet.row_values_calls, [2])
        self.assertEqual(len(sheet.appended_rows), 1)
        self.assertEqual(sheet.appended_rows[0][1], 'USER_ENTERED')
        self.assertEqual(len(sheet.appended_rows[0][0]), 2)
        case_one.refresh_from_db()
        case_two.refresh_from_db()
        self.assertEqual(case_one.row_number, 5)
        self.assertEqual(case_two.row_number, 6)

    def test_sync_case_to_sheet_prefers_stage_tat_headers_over_fixed_lag_columns(self):
        class FakeSheet:
            def __init__(self):
                self.updates = []

            def row_values(self, row):
                if row == 2:
                    headers = [''] * 34
                    headers[2] = 'ID NUMBER'
                    headers[3] = 'PHONE NUMBER'
                    headers[33] = 'MPESA sent to Admin TAT Minutes'
                    return headers
                return [''] * 34

            def update(self, a1_range, values, value_input_option=None):
                self.updates.append((a1_range, values, value_input_option))

        class FakeService:
            def __init__(self, sheet):
                self._sheet = sheet

            def is_available(self):
                return True

        sheet = FakeSheet()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-009',
            product_key='business',
            product_label='Business',
            client_name='Header Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={
                'created': timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 0)).isoformat(),
                'mpesa_to_admin': timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 25)).isoformat(),
            },
            status='Active',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService(sheet)):
            sync_case_to_sheet(self.config, case)

        self.assertEqual(sheet.updates[0][0], 'A5:AH5')
        self.assertEqual(sheet.updates[0][1][0][33], 25.0)

    @override_settings(TAT_TRACKER_SYNC_SECONDARY_SHEETS=False)
    def test_secondary_sheet_override_syncs_index_but_not_unused_audit_log(self):
        class FakeSheet:
            def row_values(self, _row):
                return [''] * 20

            def update(self, *_args, **_kwargs):
                return None

        class FakeService:
            _sheet = FakeSheet()

            def is_available(self):
                return True

        self.config.workflow['sync_secondary_sheets'] = True
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-004',
            product_key='business',
            product_label='Business',
            client_name='Test Client',
            national_id='12345678',
            primary_phone='254712345678',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={'created': timezone.now().isoformat()},
            status='Active',
        )

        with patch('core.services.tat_tracker.get_sheets_service', return_value=FakeService()), \
             patch('core.services.tat_tracker.sync_case_index') as index_mock, \
             patch('core.services.tat_tracker.sync_audit_log') as audit_mock:
            sync_case_to_sheet(self.config, case)

        index_mock.assert_called_once_with(self.config, case)
        audit_mock.assert_not_called()

    def test_calculated_tat_values_use_aware_datetimes_and_ongoing_now(self):
        case = TatTrackerCase(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-002',
            product_key='business',
            client_name='Ongoing Client',
            stage_values={'created': timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 0)).isoformat()},
        )
        now = timezone.make_aware(timezone.datetime(2026, 7, 14, 20, 0))

        self.assertEqual(calculated_tat_minutes(case, now=now), Decimal('720.00'))
        self.assertEqual(calculated_tat_hours(case, now=now), Decimal('12.00'))
        self.assertEqual(calculated_tat_days(case, now=now), Decimal('0.50'))

    def test_rejected_tat_ends_at_decision_timestamp(self):
        case = TatTrackerCase(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-006',
            product_key='business',
            client_name='Rejected Client',
            status='Rejected',
            stage_values={
                'created': timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 0)).isoformat(),
                'decision': 'Rejected',
                'decision_ts': timezone.make_aware(timezone.datetime(2026, 7, 14, 10, 30)).isoformat(),
            },
        )
        now = timezone.make_aware(timezone.datetime(2026, 7, 15, 8, 0))

        self.assertEqual(calculated_tat_minutes(case, now=now), Decimal('150.00'))

    def test_negative_decisions_share_declined_status_and_deferred_stops_tat(self):
        created = timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 0))
        decided = timezone.make_aware(timezone.datetime(2026, 7, 14, 10, 30))
        later = timezone.make_aware(timezone.datetime(2026, 7, 15, 8, 0))
        product = product_by_key('mjengo')
        decision_stage = stage_by_key(product, 'decision')

        for legacy_status in ('Rejected', 'Declined', 'Deferred'):
            case = TatTrackerCase(
                group_id=self.config.group_id,
                case_id=f'JBL-BS-NEGATIVE-{legacy_status}',
                product_key='business',
                client_name='Negative Outcome Client',
                status=legacy_status,
                stage_values={
                    'created': created.isoformat(),
                    'decision': legacy_status,
                    'decision_ts': decided.isoformat(),
                },
            )
            self.assertEqual(canonical_tat_status(case.status), 'Declined')
            self.assertEqual(tat_reporting_status(case, workflow=self.config.workflow, now=later), 'Declined')
            self.assertFalse(overall_tat_running(case))
            self.assertEqual(calculated_tat_minutes(case, now=later), Decimal('150.00'))

        new_case = TatTrackerCase(
            group_id=self.config.group_id, case_id='JBL-BS-NEGATIVE-NEW',
            product_key='mjengo', client_name='New Negative Outcome',
            status='Active', stage_values={'created': created.isoformat()},
        )
        apply_side_effects(new_case, product, decision_stage, 'Deferred')
        self.assertEqual(new_case.status, 'Declined')
        apply_side_effects(new_case, product, decision_stage, 'Approved')
        self.assertEqual(new_case.status, 'Active')

    def test_stage_tat_minutes_use_previous_stage_and_current_pending_stage(self):
        product = product_by_key('business')
        case = TatTrackerCase(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-007',
            product_key='business',
            client_name='Stage Client',
            stage_values={
                'created': timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 0)).isoformat(),
                'mpesa_to_admin': timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 45)).isoformat(),
            },
        )
        pending_now = timezone.make_aware(timezone.datetime(2026, 7, 14, 9, 30))

        self.assertEqual(stage_tat_minutes(case, product.stages[0]), Decimal('45.00'))
        self.assertEqual(stage_tat_minutes(case, product.stages[1], now=pending_now), Decimal('45.00'))

    def test_detail_payload_includes_stage_tat_and_sla_status(self):
        self.config.workflow['tat_targets_minutes'] = {
            'business': {'total': 120, 'stages': {'mpesa_to_admin': 60, 'mpesa_verified': 30}}
        }
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-2026-008',
            product_key='business',
            product_label='Business',
            client_name='Target Client',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={
                'created': timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 0)).isoformat(),
                'mpesa_to_admin': timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 50)).isoformat(),
            },
            status='Active',
        )

        from core.services.tat_tracker import serialize_case_detail
        detail = serialize_case_detail(case, user, workflow=self.config.workflow)

        self.assertEqual(detail['summary']['target_minutes'], '120')
        self.assertIsInstance(detail['summary']['elapsed_seconds'], int)
        self.assertTrue(detail['summary']['running'])
        self.assertEqual(detail['summary']['target_seconds'], 7200)
        self.assertIn('server_now', detail['summary'])
        self.assertEqual(detail['fields'][0]['tat_minutes'], '50.00')
        self.assertEqual(detail['fields'][0]['elapsed_seconds'], 3000)
        self.assertFalse(detail['fields'][0]['running'])
        self.assertEqual(detail['fields'][0]['target_seconds'], 3600)
        self.assertEqual(detail['fields'][0]['target_minutes'], '60')
        self.assertEqual(detail['fields'][0]['sla_status'], 'near')
        self.assertEqual(detail['fields'][1]['target_minutes'], '30')

        hidden = serialize_case_detail(
            case, user, workflow=self.config.workflow, include_business_time=False,
        )
        self.assertFalse(hidden['business_time_enabled'])
        self.assertNotIn('business_minutes', hidden['summary'])
        self.assertTrue(all('business_minutes' not in field for field in hidden['fields']))

    def test_terminal_case_official_counter_is_fixed(self):
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        created_at = timezone.make_aware(timezone.datetime(2026, 7, 14, 8, 0))
        completed_at = timezone.make_aware(timezone.datetime(2026, 7, 14, 10, 0))
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id, sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business', case_id='JBL-BS-2026-FIXED',
            product_key='business', product_label='Business', client_name='Fixed Client',
            branch='Nakuru', status='Disbursed',
            stage_values={'created': created_at.isoformat()},
        )
        TatTrackerCase.objects.filter(pk=case.pk).update(updated_at=completed_at)
        case.refresh_from_db()

        from core.services.tat_tracker import serialize_case_detail
        detail = serialize_case_detail(case, user, workflow=self.config.workflow)

        self.assertFalse(detail['summary']['running'])
        self.assertEqual(detail['summary']['elapsed_seconds'], 7200)

    def test_next_role_alert_targets_pending_stage_role(self):
        data = {
            'summary': {
                'case_id': 'JBL-BS-2026-001',
                'product_key': 'business',
                'client_name': 'Test Client',
                'national_id': '12345678',
                'primary_phone': '0712345678',
                'branch': 'Nakuru',
                'next_stage_key': 'mpesa_to_admin',
            }
        }

        alert = next_role_alert(self.config, data)

        self.assertEqual(alert['role'], 'BRO')
        self.assertIn('⏰ Action Required', alert['text'])
        self.assertIn('Assigned to: BRO', alert['text'])
        self.assertIn('Reference: JBL-BS-2026-001', alert['text'])
        self.assertIn('Next Step: MPESA sent to Admin', alert['text'])
        self.assertIn('confirm the required action', alert['text'])

    def test_next_role_alert_can_be_disabled_in_workflow(self):
        self.config.workflow['stage_alerts_enabled'] = False
        data = {'summary': {'product_key': 'business', 'next_stage_key': 'mpesa_to_admin'}}

        self.assertEqual(next_role_alert(self.config, data), {})

    @patch('core.api.views._post_telegram_reply')
    @patch('core.services.tat_tracker.sync_case_to_sheet')
    @patch('core.services.tat_tracker.validate_tat_telegram_webapp_init_data')
    def test_create_endpoint_alerts_next_stage_role(self, mock_auth, sync_mock, mock_reply):
        mock_auth.return_value = (True, '', {'id': 111, 'username': 'bro_user'})
        sync_mock.side_effect = self.mark_case_synced
        GroupRegistry._instance = None

        response = self.client.post(
            '/api/tat-tracker/create/',
            data=json.dumps({
                'group_id': self.config.group_id,
                'init_data': 'mock',
                'product_key': 'business',
                'branch': 'Nakuru',
                'client_name': 'Test Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
                'bro_name': 'BRO User',
                'amount': '10000',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        mock_reply.assert_called_once()
        self.assertIn('⏰ Action Required', mock_reply.call_args.kwargs['text'])
        self.assertIn('Assigned to: BRO', mock_reply.call_args.kwargs['text'])

    @patch('core.services.tat_tracker.validate_tat_telegram_webapp_init_data')
    def test_create_endpoint_returns_specific_invalid_bro_message(self, mock_auth):
        mock_auth.return_value = (True, '', {'id': 111, 'username': 'bro_user'})
        GroupRegistry._instance = None

        response = self.client.post(
            '/api/tat-tracker/create/',
            data=json.dumps({
                'group_id': self.config.group_id,
                'init_data': 'mock',
                'product_key': 'business',
                'branch': 'Nakuru',
                'client_name': 'Test Client',
                'national_id': '12345678',
                'primary_phone': '0712345678',
                'bro_user_id': '999999',
                'amount': '10000',
                'client_request_id': 'invalid-bro-create-123',
            }),
            content_type='application/json',
            HTTP_IDEMPOTENCY_KEY='invalid-bro-create-123',
            HTTP_X_REQUEST_ID='invalid-bro-create-123',
            HTTP_X_MINIAPP_MESSAGE_CONTRACT='2',
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['code'], 'tat_create_invalid_bro')
        self.assertIn('active TAT BRO', response.json()['message'])
        self.assertNotIn('error', response.json())

    def test_tracker_identifier_headers_are_required_when_headers_exist(self):
        with self.assertRaisesRegex(ValueError, 'ID NUMBER'):
            validate_tracker_identity_headers(['Case ID', 'Client Name', 'Branch', 'BRO Name'])

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_normalizes_and_stores_customer_identifiers(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        detail = create_case(self.config, user, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Test Client',
            'national_id': '12 345 678',
            'primary_phone': '+254 712 345 678',
            'bro_name': 'BRO User',
            'amount': '10000',
        })

        case = TatTrackerCase.objects.get(case_id=detail['summary']['case_id'])
        self.assertEqual(case.national_id, '12345678')
        self.assertEqual(case.primary_phone, '254712345678')
        self.assertEqual(detail['summary']['national_id'], '12345678')
        self.assertEqual(detail['summary']['primary_phone'], '254712345678')
        self.assertEqual(search_cases(self.config, user, '0712345678')[0]['case_id'], case.case_id)

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_resolves_bro_user_id_to_canonical_name(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        User = get_user_model()
        other_group = GroupSheetConfiguration.objects.create(
            group_id='-100tat-assignment', sheet_id='assignment-sheet',
            sheet_name='TAT Assignment', workflow={'type': 'tat_tracker'},
        )
        selected_bro = User.objects.create_user(
            username='selected-bro', first_name='Selected', last_name='BRO', is_active=True,
        )
        AccessGrant.objects.create(
            user=selected_bro, workflow='tat_tracker', role='BRO',
            group_configuration=other_group,
        )
        user = staff_user_for_payload(self.config, {'id': 222, 'username': 'admin_user'})

        detail = create_case(self.config, user, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Assigned Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'bro_user_id': selected_bro.pk,
            'bro_name': 'Stale Browser Name',
            'amount': '10000',
        })

        case = TatTrackerCase.objects.get(case_id=detail['summary']['case_id'])
        self.assertEqual(case.bro_name, 'Selected BRO')

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_rejects_unknown_or_inactive_bro_user_id(self, sync_mock):
        User = get_user_model()
        inactive_bro = User.objects.create_user(
            username='inactive-selected-bro', first_name='Inactive', last_name='Selected',
            is_active=False,
        )
        AccessGrant.objects.create(
            user=inactive_bro, workflow='tat_tracker', role='BRO',
            group_configuration=self.config,
        )
        user = staff_user_for_payload(self.config, {'id': 222, 'username': 'admin_user'})
        payload = {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Invalid Assignment',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'amount': '10000',
        }

        for bro_user_id in (inactive_bro.pk, '999999'):
            with self.subTest(bro_user_id=bro_user_id):
                with self.assertRaisesRegex(ValueError, 'active TAT BRO'):
                    create_case(self.config, user, {**payload, 'bro_user_id': bro_user_id})

        self.assertFalse(TatTrackerCase.objects.exists())
        sync_mock.assert_not_called()

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_defaults_active_bro_creator_to_self(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        detail = create_case(self.config, user, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Default Assignment',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'amount': '10000',
        })

        case = TatTrackerCase.objects.get(case_id=detail['summary']['case_id'])
        self.assertEqual(case.bro_name, 'BRO User')

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_requires_non_bro_creator_to_select_bro(self, sync_mock):
        user = staff_user_for_payload(self.config, {'id': 222, 'username': 'admin_user'})

        with self.assertRaisesRegex(ValueError, 'Select a BRO'):
            create_case(self.config, user, {
                'product_key': 'business',
                'branch': 'Nakuru',
                'client_name': 'Missing Assignment',
                'national_id': '12345678',
                'primary_phone': '0712345678',
                'amount': '10000',
            })

        self.assertFalse(TatTrackerCase.objects.exists())
        sync_mock.assert_not_called()

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_rejects_invalid_customer_identifiers(self, sync_mock):
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        payload = {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Test Client',
            'national_id': '1234',
            'primary_phone': '0712345678',
            'bro_name': 'BRO User',
            'amount': '10000',
        }

        with self.assertRaisesRegex(ValueError, 'ID number'):
            create_case(self.config, user, payload)
        self.assertFalse(sync_mock.called)
    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_assigns_sequential_case_id(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        first = create_case(self.config, user, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Test Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'bro_name': 'BRO User',
            'amount': '10000',
        })
        second = create_case(self.config, user, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Second Client',
            'national_id': '87654321',
            'primary_phone': '0712345679',
            'bro_name': 'BRO User',
            'amount': '10000',
        })
        self.assertEqual(first['summary']['case_id'], 'JBL-BS-2026-001')
        self.assertEqual(second['summary']['case_id'], 'JBL-BS-2026-002')
        self.assertEqual(TatTrackerCase.objects.count(), 2)
        self.assertEqual(sync_mock.call_count, 2)

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_retry_with_same_request_id_returns_existing_case(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        payload = {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Test Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'bro_user_id': self.bro_user.pk,
            'bro_name': 'Stale Browser Name',
            'amount': '10000',
            'client_request_id': 'req-123',
        }

        first = create_case(self.config, user, payload)
        second = create_case(self.config, user, payload)

        self.assertEqual(first['summary']['case_id'], second['summary']['case_id'])
        self.assertEqual(TatTrackerCase.objects.count(), 1)
        self.assertEqual(TatTrackerCase.objects.get().bro_name, 'BRO User')
        self.assertEqual(sync_mock.call_count, 1)

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_retry_does_not_resync_existing_unsynced_case(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        payload = {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Test Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'bro_name': 'BRO User',
            'amount': '10000',
            'client_request_id': 'req-unsynced-retry',
        }

        first = create_case(self.config, user, payload)
        case = TatTrackerCase.objects.get(case_id=first['summary']['case_id'])
        case.row_number = None
        case.sync_error = 'response lost after sheet append'
        case.save(update_fields=['row_number', 'sync_error', 'updated_at'])

        second = create_case(self.config, user, payload)

        self.assertEqual(first['summary']['case_id'], second['summary']['case_id'])
        self.assertEqual(TatTrackerCase.objects.count(), 1)
        self.assertEqual(sync_mock.call_count, 1)

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_create_case_rolls_back_when_primary_sheet_sync_fails(self, sync_mock):
        sync_mock.side_effect = RuntimeError('Primary sheet write failed')
        user = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        with self.assertRaises(RuntimeError):
            create_case(self.config, user, {
                'product_key': 'business',
                'branch': 'Nakuru',
                'client_name': 'Test Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
                'bro_name': 'BRO User',
                'amount': '10000',
                'client_request_id': 'req-fail',
            })

        self.assertEqual(TatTrackerCase.objects.count(), 0)

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_stage_updates_are_role_and_sequence_controlled(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        bro = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        admin = staff_user_for_payload(self.config, {'id': 222, 'username': 'admin_user'})
        detail = create_case(self.config, bro, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Test Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'bro_name': 'BRO User',
            'amount': '10000',
        })
        case_id = detail['summary']['case_id']

        with self.assertRaises(ValueError):
            update_case(self.config, admin, case_id, [{'field': 'mpesa_verified', 'value': 'STAMP'}])

        update_case(self.config, bro, case_id, [{'field': 'mpesa_to_admin', 'value': 'STAMP'}])
        updated = update_case(self.config, admin, case_id, [{'field': 'mpesa_verified', 'value': 'STAMP'}])
        self.assertEqual(updated['summary']['next_stage_key'], 'ca_analysis_sent')

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_tat_update_uses_revision_receipts_for_conflicts_and_retries(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        bro = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        detail = create_case(self.config, bro, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Revision Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'bro_name': 'BRO User',
            'amount': '10000',
        })
        case = TatTrackerCase.objects.get(case_id=detail['summary']['case_id'])
        expected_revision = case.workflow_revision

        update_case(
            self.config,
            bro,
            case.case_id,
            [{'field': 'mpesa_to_admin', 'value': 'STAMP'}],
            expected_revision=expected_revision,
            request_id='tat-revision-1',
        )
        case.refresh_from_db()
        receipt = case.events.get(request_id='tat-revision-1')
        self.assertEqual(case.workflow_revision, expected_revision + 1)
        self.assertEqual(receipt.actor_user, self.bro_user)
        self.assertEqual(receipt.authority_user, self.bro_user)
        self.assertEqual(receipt.revision_before, expected_revision)
        self.assertEqual(receipt.revision_after, expected_revision + 1)

        replay = update_case(
            self.config,
            bro,
            case.case_id,
            [{'field': 'mpesa_to_admin', 'value': 'STAMP'}],
            expected_revision=expected_revision,
            request_id='tat-revision-1',
        )
        self.assertEqual(replay['summary']['workflow_revision'], expected_revision + 1)
        self.assertEqual(case.events.filter(request_id='tat-revision-1').count(), 1)

        with self.assertRaises(WorkflowRevisionConflict):
            update_case(
                self.config,
                bro,
                case.case_id,
                [{'field': 'mpesa_to_admin', 'value': 'STAMP'}],
                expected_revision=expected_revision,
                request_id='tat-revision-stale',
            )

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_admin_case_detail_correction_is_normalized_and_audited(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        bro = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        admin = staff_user_for_payload(self.config, {'id': 222, 'username': 'admin_user'})
        detail = create_case(self.config, bro, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Original Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'bro_name': 'BRO User',
            'amount': '10000',
        })
        original_case = TatTrackerCase.objects.get(case_id=detail['summary']['case_id'])
        original_pk = original_case.pk
        case_count_before_correction = TatTrackerCase.objects.filter(group_id=self.config.group_id, is_deleted=False).count()

        updated = update_case(self.config, admin, detail['summary']['case_id'], [
            {'field': 'client_name', 'value': 'Corrected Client', 'correction': True},
            {'field': 'national_id', 'value': '87 654 321', 'correction': True},
            {'field': 'primary_phone', 'value': '+254 712 345 679', 'correction': True},
        ])

        case = TatTrackerCase.objects.get(case_id=detail['summary']['case_id'])
        self.assertEqual(case.pk, original_pk)
        self.assertEqual(
            TatTrackerCase.objects.filter(group_id=self.config.group_id, is_deleted=False).count(),
            case_count_before_correction,
        )
        self.assertEqual(case.client_name, 'CORRECTED CLIENT')
        self.assertEqual(case.national_id, '87654321')
        self.assertEqual(case.primary_phone, '254712345679')
        self.assertEqual(updated['summary']['client_name'], 'CORRECTED CLIENT')
        self.assertEqual(case.events.filter(stage_key='case_details', source='admin_correction').count(), 3)

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_bro_can_correct_scoped_case_details_but_not_completed_stages(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        bro = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        detail = create_case(self.config, bro, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Original Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'bro_name': 'BRO User',
            'amount': '10000',
        })
        case_id = detail['summary']['case_id']
        update_case(self.config, bro, case_id, [
            {'field': 'client_name', 'value': 'BRO Corrected Client', 'correction': True},
        ])
        update_case(self.config, bro, case_id, [
            {'field': 'mpesa_to_admin', 'value': 'STAMP'},
        ])

        with self.assertRaisesMessage(ValueError, 'Your role cannot correct MPESA sent to Admin'):
            update_case(self.config, bro, case_id, [{
                'field': 'mpesa_to_admin',
                'value': '2026-07-20T10:30',
                'correction': True,
            }])

        case = TatTrackerCase.objects.get(case_id=case_id)
        serialized = get_case_detail(self.config, bro, case_id)
        completed = next(field for field in serialized['fields'] if field['key'] == 'mpesa_to_admin')
        self.assertEqual(case.client_name, 'BRO CORRECTED CLIENT')
        self.assertTrue(serialized['can_correct_details'])
        self.assertFalse(completed['can_correct'])

    def test_create_case_rejects_correction_shaped_payload(self):
        bro = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})

        with self.assertRaisesMessage(ValueError, 'Case corrections must use the existing case update action'):
            create_case(self.config, bro, {
                'case_id': 'JBL-BS-2026-001',
                'workflow_revision': 1,
                'updates': [{'field': 'client_name', 'value': 'Incorrect route'}],
                'creation_intent': 'new_loan',
            })

    def test_identity_context_lists_prior_loans_without_treating_them_as_duplicates(self):
        first = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-IDENTITY-1',
            product_key='business', product_label='Business',
            client_name='Repeat Client', national_id='12345678', primary_phone='254712345678',
            branch='Nakuru', status='Active',
        )
        second = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            case_id='JBL-BS-2026-IDENTITY-2',
            product_key='logbook', product_label='Logbook',
            client_name='Repeat Client', national_id='12345678', primary_phone='254712345678',
            branch='Nakuru', status='Active',
        )

        context = tat_case_identity_context(self.config, '12345678', '0712345678')

        self.assertEqual({item['case_id'] for item in context['matches']}, {first.case_id, second.case_id})
        self.assertEqual(context['matched_on'], ['National ID', 'primary phone'])

    def test_identity_context_endpoint_allows_case_creators_and_denies_others(self):
        request = RequestFactory().post(
            '/api/tat-tracker/identity-context/',
            data=json.dumps({'national_id': '12345678'}),
            content_type='application/json',
        )
        allowed_user = {'authorized': True, 'capabilities': ['tat.case.create']}
        with patch('core.api.views._tat_context', return_value=(self.config.group_id, self.config, {}, allowed_user, None)):
            allowed = tat_tracker_identity_context(request)
        self.assertEqual(allowed.status_code, 200)

        denied_request = RequestFactory().post(
            '/api/tat-tracker/identity-context/',
            data=json.dumps({'national_id': '12345678'}),
            content_type='application/json',
        )
        denied_user = {'authorized': True, 'capabilities': []}
        with patch('core.api.views._tat_context', return_value=(self.config.group_id, self.config, {}, denied_user, None)):
            denied = tat_tracker_identity_context(denied_request)
        self.assertEqual(denied.status_code, 403)

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_completed_timestamp_correction_preserves_audit_history(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        bro = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        it_user = staff_user_for_payload(self.config, {'id': 444, 'username': 'it_user'})
        detail = create_case(self.config, bro, {
            'product_key': 'business',
            'branch': 'Nakuru',
            'client_name': 'Timestamp Client',
            'national_id': '12345678',
            'primary_phone': '0712345678',
            'bro_name': 'BRO User',
            'amount': '10000',
        })
        case_id = detail['summary']['case_id']
        update_case(self.config, bro, case_id, [{'field': 'mpesa_to_admin', 'value': 'STAMP'}])

        update_case(self.config, it_user, case_id, [{
            'field': 'mpesa_to_admin',
            'value': '2026-07-20T10:30',
            'correction': True,
        }])

        case = TatTrackerCase.objects.get(case_id=case_id)
        corrected = parse_iso_datetime(case.stage_values['mpesa_to_admin'])
        self.assertEqual(corrected.strftime('%Y-%m-%d %H:%M'), '2026-07-20 10:30')
        event = case.events.filter(stage_key='mpesa_to_admin').order_by('-created_at').first()
        self.assertEqual(event.source, 'admin_correction')
        self.assertIn('(Correction)', event.stage_label)
        self.assertEqual(event.actor_user, self.it_user)

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_superuser_can_correct_completed_stage_without_business_role_grant(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        bro = staff_user_for_payload(self.config, {'id': 111, 'username': 'bro_user'})
        root = get_user_model().objects.create_superuser(
            username='tat-root', email='tat-root@example.test', password='password',
        )
        UserProfile.objects.create(user=root, telegram_id='999', telegram_username='tat_root')
        root_payload = staff_user_for_payload(self.config, {'id': 999, 'username': 'tat_root'})
        detail = create_case(self.config, bro, {
            'product_key': 'business', 'branch': 'Nakuru', 'client_name': 'Root correction',
            'national_id': '12345678', 'primary_phone': '0712345678',
            'bro_name': 'BRO User', 'amount': '10000',
        })
        case_id = detail['summary']['case_id']
        update_case(self.config, bro, case_id, [{'field': 'mpesa_to_admin', 'value': 'STAMP'}])

        update_case(self.config, root_payload, case_id, [{
            'field': 'mpesa_to_admin', 'value': '2026-07-21T11:45', 'correction': True,
        }])

        case = TatTrackerCase.objects.get(case_id=case_id)
        self.assertEqual(
            parse_iso_datetime(case.stage_values['mpesa_to_admin']).strftime('%Y-%m-%d %H:%M'),
            '2026-07-21 11:45',
        )

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_assigned_role_can_change_a_dropdown_value_and_audit_the_change(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        sanctions_timestamp = timezone.make_aware(timezone.datetime(2026, 7, 18, 10, 0)).isoformat()
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-MJENGO',
            row_number=5,
            case_id='JBL-MJ-2026-EDIT-DROPDOWN',
            product_key='mjengo',
            product_label='Mjengo',
            client_name='Dropdown Client',
            branch='Nakuru',
            bro_name='BRO User',
            amount='100000',
            stage_values={
                'created': timezone.now().isoformat(),
                'mpesa_to_admin': timezone.now().isoformat(),
                'mpesa_verified': timezone.now().isoformat(),
                'ca_analysis_sent': timezone.now().isoformat(),
                'bro_response': timezone.now().isoformat(),
                'bm_tat_request': timezone.now().isoformat(),
                'tat_scheduled': timezone.now().isoformat(),
                'tat_held': timezone.now().isoformat(),
                'decision': 'Approved',
                'decision_ts': timezone.now().isoformat(),
                'minutes_shared': 'Yes',
                'minutes_shared_ts': timezone.now().isoformat(),
                'sanctions': 'Pending',
                'sanctions_ts': sanctions_timestamp,
            },
        )
        loan_approver = {
            'name': 'Loan Approver',
            'telegram_id': '444',
            'roles': ['LOAN_APPROVER'],
            'branches': ['Nakuru'],
            'products': ['mjengo'],
        }

        detail = update_case(
            self.config,
            loan_approver,
            case.case_id,
            [{'field': 'sanctions', 'value': 'Met'}],
        )

        case.refresh_from_db()
        sanctions_field = next(field for field in detail['fields'] if field['key'] == 'sanctions')
        event = case.events.get(stage_key='sanctions')
        self.assertTrue(sanctions_field['editable'])
        self.assertEqual(case.stage_values['sanctions'], 'Met')
        self.assertEqual(case.stage_values['sanctions_ts'], sanctions_timestamp)
        self.assertEqual(event.old_value, 'Pending')
        self.assertEqual(event.new_value, 'Met')

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_mjengo_minutes_shared_writes_dropdown_value_and_internal_timestamp(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-MJENGO',
            row_number=5,
            case_id='JBL-MJ-2026-MINUTES-DROPDOWN',
            product_key='mjengo',
            product_label='Mjengo',
            client_name='Minutes Client',
            branch='Nakuru',
            bro_name='BRO User',
            amount='100000',
            stage_values={
                'created': timezone.now().isoformat(),
                'mpesa_to_admin': timezone.now().isoformat(),
                'mpesa_verified': timezone.now().isoformat(),
                'ca_analysis_sent': timezone.now().isoformat(),
                'bro_response': timezone.now().isoformat(),
                'bm_tat_request': timezone.now().isoformat(),
                'tat_scheduled': timezone.now().isoformat(),
                'tat_held': timezone.now().isoformat(),
                'decision': 'Approved',
                'decision_ts': timezone.now().isoformat(),
            },
        )
        secretary = {
            'name': 'Secretary',
            'telegram_id': '333',
            'roles': ['SECRETARY'],
            'branches': ['Nakuru'],
            'products': ['mjengo'],
        }

        update_case(self.config, secretary, case.case_id, [{'field': 'minutes_shared', 'value': 'Yes'}])

        case.refresh_from_db()
        self.assertEqual(case.stage_values['minutes_shared'], 'Yes')
        self.assertTrue(parse_iso_datetime(case.stage_values['minutes_shared_ts']))
        self.assertIsNotNone(stage_tat_minutes(case, stage_by_key(product_by_key('mjengo'), 'minutes_shared')))

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_mjengo_bro_applied_writes_dropdown_value_and_internal_timestamp(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-MJENGO',
            row_number=5,
            case_id='JBL-MJ-2026-BRO-APPLIED-DROPDOWN',
            product_key='mjengo',
            product_label='Mjengo',
            client_name='Applied Client',
            branch='Nakuru',
            bro_name='BRO User',
            amount='100000',
            stage_values={
                'created': timezone.now().isoformat(),
                'mpesa_to_admin': timezone.now().isoformat(),
                'mpesa_verified': timezone.now().isoformat(),
                'ca_analysis_sent': timezone.now().isoformat(),
                'bro_response': timezone.now().isoformat(),
                'bm_tat_request': timezone.now().isoformat(),
                'tat_scheduled': timezone.now().isoformat(),
                'tat_held': timezone.now().isoformat(),
                'decision': 'Approved',
                'decision_ts': timezone.now().isoformat(),
                'minutes_shared': 'Yes',
                'minutes_shared_ts': timezone.now().isoformat(),
                'sanctions': 'Met',
                'sanctions_ts': timezone.now().isoformat(),
            },
        )
        bro = {
            'name': 'BRO User',
            'telegram_id': '111',
            'roles': ['BRO'],
            'branches': ['Nakuru'],
            'products': ['mjengo'],
        }

        update_case(self.config, bro, case.case_id, [{'field': 'bro_applied', 'value': 'Met'}])

        case.refresh_from_db()
        self.assertEqual(case.stage_values['bro_applied'], 'Met')
        self.assertTrue(parse_iso_datetime(case.stage_values['bro_applied_ts']))
        self.assertIsNotNone(stage_tat_minutes(case, stage_by_key(product_by_key('mjengo'), 'bro_applied')))

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_register_approval_records_a_completion_timestamp_for_its_tat(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-REGISTER-APPROVAL',
            product_key='business',
            product_label='Business',
            client_name='Approval Client',
            branch='Nakuru',
            bro_name='BRO User',
            amount='10000',
            stage_values={
                'created': timezone.now().isoformat(),
                'mpesa_to_admin': timezone.now().isoformat(),
                'mpesa_verified': timezone.now().isoformat(),
                'ca_analysis_sent': timezone.now().isoformat(),
                'bro_response': timezone.now().isoformat(),
                'bm_response': timezone.now().isoformat(),
                'bro_applied': timezone.now().isoformat(),
                'disbursement_register': '10:00am',
                'register_ts': timezone.now().isoformat(),
            },
        )
        loan_approver = {
            'name': 'Loan Approver',
            'telegram_id': '444',
            'roles': ['LOAN_APPROVER'],
            'branches': ['Nakuru'],
            'products': ['business'],
        }

        update_case(
            self.config,
            loan_approver,
            case.case_id,
            [{'field': 'register_approved', 'value': 'Approved'}],
        )

        case.refresh_from_db()
        self.assertEqual(case.stage_values['register_approved'], 'Approved')
        self.assertTrue(parse_iso_datetime(case.stage_values['register_approved_ts']))
        self.assertIsNotNone(stage_tat_minutes(case, stage_by_key(product_by_key('business'), 'register_approved')))

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_changing_a_decision_dropdown_reopens_a_rejected_case(self, sync_mock):
        sync_mock.side_effect = self.mark_case_synced
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-MJENGO',
            row_number=5,
            case_id='JBL-MJ-2026-REOPEN-DROPDOWN',
            product_key='mjengo',
            product_label='Mjengo',
            client_name='Decision Client',
            branch='Nakuru',
            bro_name='BRO User',
            amount='100000',
            status='Rejected',
            stage_values={
                'created': timezone.now().isoformat(),
                'mpesa_to_admin': timezone.now().isoformat(),
                'mpesa_verified': timezone.now().isoformat(),
                'ca_analysis_sent': timezone.now().isoformat(),
                'bro_response': timezone.now().isoformat(),
                'bm_tat_request': timezone.now().isoformat(),
                'tat_scheduled': timezone.now().isoformat(),
                'tat_held': timezone.now().isoformat(),
                'decision': 'Rejected',
                'decision_ts': timezone.now().isoformat(),
            },
        )
        chair = {
            'name': 'Chair User',
            'telegram_id': '555',
            'roles': ['CHAIR'],
            'branches': ['Nakuru'],
            'products': ['mjengo'],
        }

        detail = update_case(self.config, chair, case.case_id, [{'field': 'decision', 'value': 'Approved'}])

        case.refresh_from_db()
        self.assertEqual(case.status, 'Active')
        self.assertEqual(detail['summary']['next_stage_key'], 'minutes_shared')
        self.assertEqual(case.events.get(stage_key='decision').old_value, 'Rejected')


class TatTrackerRepairTest(TestCase):
    def setUp(self):
        self.config = GroupSheetConfiguration.objects.create(
            group_id='-100tat-repair',
            display_name='TAT Repair Test',
            sheet_id='sheet-repair',
            sheet_name='TRACKER-Business',
            workflow={'type': 'tat_tracker', 'products': ['business', 'logbook']},
        )
        self.repairable_case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-2026-REPAIR',
            product_key='business',
            product_label='Business',
            client_name='Repairable Client',
            branch='Nakuru',
            status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-LOGBOOK',
            row_number=6,
            case_id='JBL-LB-2026-REPAIR',
            product_key='logbook',
            product_label='Logbook',
            client_name='Other Product',
            branch='Nakuru',
            status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )
        TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-2026-UNLINKED',
            product_key='business',
            product_label='Business',
            client_name='Unlinked Client',
            branch='Nakuru',
            status='Active',
            stage_values={'created': timezone.now().isoformat()},
        )

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_repair_resync_limits_to_linked_cases_and_selected_product(self, sync_case):
        result = resync_tat_tracker_cases(self.config, product_key='business')

        self.assertEqual(result, {
            'total_candidates': 1,
            'candidates': 1,
            'synced': 1,
            'skipped_unlinked': 1,
            'failed': [],
            'offset': 0,
            'next_offset': None,
        })
        sync_case.assert_called_once_with(self.config, self.repairable_case)

    @patch('core.services.tat_tracker.sync_case_to_sheet')
    def test_repair_dry_run_does_not_write_to_google_sheets(self, sync_case):
        result = resync_tat_tracker_cases(self.config, dry_run=True)

        self.assertEqual(result['candidates'], 2)
        self.assertEqual(result['synced'], 0)
        self.assertEqual(result['skipped_unlinked'], 1)
        sync_case.assert_not_called()

    @patch('core.services.tat_tracker.resync_tat_tracker_cases')
    def test_background_repair_checkpoints_each_case(self, resync):
        from core.services.tat_repair_jobs import create_repair_job, run_repair_job

        resync.return_value = {'synced': 1, 'failed': []}
        job = create_repair_job(self.config, product_key='business', requested_by='admin')

        run_repair_job(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, 'completed')
        self.assertEqual(job.cursor, 1)
        self.assertEqual(job.synced_cases, 1)
        resync.assert_called_once_with(
            self.config,
            product_key='business',
            case_ids=[self.repairable_case.case_id],
            dry_run=False,
            limit=None,
            offset=0,
            include_unlinked=True,
        )

    @patch('core.services.tat_tracker.resync_tat_tracker_cases')
    def test_background_repair_records_failure_and_completes(self, resync):
        from core.services.tat_repair_jobs import create_repair_job, run_repair_job

        resync.return_value = {'synced': 0, 'failed': [{'case_id': self.repairable_case.case_id, 'error': 'quota'}]}
        job = create_repair_job(self.config, product_key='business')

        run_repair_job(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, 'completed_with_errors')
        self.assertEqual(job.cursor, 1)
        self.assertEqual(len(job.failures), 1)

    def test_apps_script_contains_an_explicit_formula_only_repair(self):
        source = (Path(__file__).resolve().parent.parent / 'tat_tracker_apps_script.gs').read_text(encoding='utf-8')

        self.assertIn("'Remove legacy TAT formulas (safe)'", source)
        self.assertIn('function clearLegacyTatFormulas()', source)
        self.assertIn('range.getFormulas()', source)
        self.assertIn('getRangeList(formulaCells).clearContent()', source)

    @patch('core.management.commands.resync_tat_tracker_cases.resync_tat_tracker_cases')
    def test_repair_command_passes_dry_run_without_writing(self, resync):
        resync.return_value = {'candidates': 2, 'synced': 0, 'skipped_unlinked': 1, 'failed': []}
        output = StringIO()

        call_command(
            'resync_tat_tracker_cases',
            f'--group-id={self.config.group_id}',
            '--product',
            'business',
            '--dry-run',
            stdout=output,
        )

        resync.assert_called_once_with(
            self.config,
            product_key='business',
            case_ids=[],
            dry_run=True,
            limit=None,
            offset=0,
        )
        self.assertIn("'synced': 0", output.getvalue())


@override_settings(SECURE_SSL_REDIRECT=False)
class TatTrackerRepairAdminTest(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username='repair-admin',
            email='repair-admin@example.test',
            password='password',
        )
        self.config = GroupSheetConfiguration.objects.create(
            group_id='-100tat-admin-repair',
            display_name='TAT Admin Repair',
            sheet_id='sheet-admin-repair',
            sheet_name='TRACKER-Business',
            workflow={'type': 'tat_tracker', 'products': ['business']},
        )
        self.url = reverse('admin:core_groupsheetconfiguration_tat_repair', args=[self.config.pk])
        self.duplicates_url = reverse('admin:core_groupsheetconfiguration_tat_duplicates', args=[self.config.pk])
        self.client.force_login(self.user)

    @patch('core.services.sheets.get_sheets_service')
    def test_duplicate_page_previews_case_id_rows_without_writing(self, get_service):
        fake_sheet = MagicMock()
        fake_sheet.get_all_values.return_value = [
            ['Case ID', 'Client'],
            [],
            [],
            [],
            ['JBL-BS-DUPLICATE', 'Old'],
            ['JBL-BS-DUPLICATE', 'New', 'More data'],
        ]
        service = MagicMock(is_available=MagicMock(return_value=True), _sheet=fake_sheet)
        get_service.return_value = service

        response = self.client.get(self.duplicates_url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'JBL-BS-DUPLICATE')
        self.assertContains(response, 'Preview duplicate rows')
        self.assertContains(response, 'Clean duplicate rows')
        fake_sheet.delete_rows.assert_not_called()

    @patch('core.services.sheets.get_sheets_service')
    def test_duplicate_page_requires_preview_and_typed_confirmation(self, get_service):
        fake_sheet = MagicMock()
        fake_sheet.get_all_values.return_value = [
            ['Case ID', 'Client'],
            [],
            [],
            [],
            ['JBL-BS-DUPLICATE', 'Old'],
            ['JBL-BS-DUPLICATE', 'New', 'More data'],
        ]
        service = MagicMock(is_available=MagicMock(return_value=True), _sheet=fake_sheet)
        get_service.return_value = service
        TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-DUPLICATE',
            product_key='business',
            product_label='Business',
            client_name='Duplicate Client',
            branch='Nakuru',
        )

        self.client.get(self.duplicates_url + '?product=business')
        response = self.client.post(self.duplicates_url, {
            'action': 'clean',
            'product': 'business',
            'confirm': 'not the phrase',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Type CLEAN DUPLICATES exactly')
        fake_sheet.delete_rows.assert_not_called()

    @patch('core.services.sheets.get_sheets_service')
    def test_duplicate_page_cleans_only_after_preview_and_confirmation(self, get_service):
        fake_sheet = MagicMock()
        fake_sheet.get_all_values.return_value = [
            ['Case ID', 'Client'],
            [],
            [],
            [],
            ['JBL-BS-DUPLICATE', 'Old'],
            ['JBL-BS-DUPLICATE', 'New', 'More data'],
        ]
        service = MagicMock(is_available=MagicMock(return_value=True), _sheet=fake_sheet)
        get_service.return_value = service
        case = TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            row_number=5,
            case_id='JBL-BS-DUPLICATE',
            product_key='business',
            product_label='Business',
            client_name='Duplicate Client',
            branch='Nakuru',
        )

        preview = [{
            'case_id': 'JBL-BS-DUPLICATE', 'rows': [], 'keep_row': 6,
            'delete_rows': [5], 'canonical_row': 5, 'linked': True,
        }]
        repaired = [{
            **preview[0], 'surviving_row': 5,
            'verification_status': 'verified', 'resync_status': 'synced',
        }]
        with patch('core.admin.cleanup_tat_sheet_duplicate_case_ids') as cleanup:
            # The Admin performs a new preview immediately before applying;
            # model a separately verified/re-published cleanup result.
            cleanup.side_effect = [preview, preview, repaired]
            self.client.get(self.duplicates_url + '?product=business')
            response = self.client.post(self.duplicates_url, {
                'action': 'clean',
                'product': 'business',
                'confirm': 'CLEAN DUPLICATES',
            })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(cleanup.call_count, 3)
        self.assertEqual(cleanup.call_args.kwargs['group_configuration'], self.config)
        case.refresh_from_db()
        self.assertEqual(case.row_number, 5)

    @patch('core.admin.resync_tat_tracker_cases')
    def test_repair_page_previews_a_bounded_batch_without_sheet_writes(self, resync):
        resync.return_value = {
            'total_candidates': 30,
            'candidates': 25,
            'synced': 0,
            'skipped_unlinked': 2,
            'failed': [],
            'offset': 0,
            'next_offset': 25,
        }

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Preview only')
        self.assertContains(response, 'Type REPAIR')
        resync.assert_called_once_with(
            self.config,
            dry_run=True,
            limit=25,
            offset=0,
            product_key='',
            include_unlinked=False,
        )

    @patch('core.admin.resync_tat_tracker_cases')
    def test_repair_page_requires_confirmation_before_writing(self, resync):
        response = self.client.post(self.url, {'confirm': 'no'})

        self.assertEqual(response.status_code, 200)
        resync.assert_not_called()

    @patch('core.admin.resync_tat_tracker_cases')
    def test_repair_page_queues_durable_job_after_typed_confirmation(self, resync):
        resync.return_value = {
            'total_candidates': 1,
            'candidates': 1,
            'synced': 0,
            'skipped_unlinked': 0,
            'failed': [],
            'offset': 0,
            'next_offset': None,
        }
        self.client.get(self.url + '?product=business')
        resync.reset_mock()
        response = self.client.post(self.url, {'confirm': 'REPAIR', 'product': 'business', 'offset': '0'})

        self.assertEqual(response.status_code, 302)
        resync.assert_not_called()
        job = TatRepairJob.objects.get()
        self.assertEqual(job.product_key, 'business')
        self.assertEqual(job.status, 'queued')
        self.assertIn(f'job={job.id}', response['Location'])

    @patch('core.admin.resync_tat_tracker_cases')
    def test_repair_page_can_include_cases_without_stored_sheet_rows(self, resync):
        resync.return_value = {
            'total_candidates': 1,
            'candidates': 1,
            'synced': 0,
            'skipped_unlinked': 0,
            'failed': [],
            'offset': 0,
            'next_offset': None,
        }
        TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-MISSING-ROW',
            product_key='business',
            product_label='Business',
            client_name='Missing Row Client',
            branch='Nakuru',
        )

        self.client.get(self.url + '?product=business&include_unlinked=1')
        resync.assert_called_once_with(
            self.config,
            dry_run=True,
            limit=25,
            offset=0,
            product_key='business',
            include_unlinked=True,
        )
        resync.reset_mock()
        response = self.client.post(self.url, {
            'confirm': 'REPAIR',
            'product': 'business',
            'offset': '0',
            'include_unlinked': '1',
        })

        self.assertEqual(response.status_code, 302)
        job = TatRepairJob.objects.get()
        self.assertEqual(job.case_ids, ['JBL-BS-MISSING-ROW'])
        self.assertEqual(job.skipped_unlinked, 0)

    def test_repair_page_renders_recorded_failure_details(self):
        job = TatRepairJob.objects.create(
            group_configuration=self.config,
            case_ids=['JBL-BS-FAILED'],
            total_cases=1,
            cursor=1,
            status='completed_with_errors',
            failures=[{
                'case_id': 'JBL-BS-FAILED',
                'error': 'Google Sheets quota exceeded; retry later.',
            }],
        )

        response = self.client.get(f'{self.url}?job={job.id}')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'JBL-BS-FAILED')
        self.assertContains(response, 'Google Sheets quota exceeded; retry later.')

    def test_repair_page_can_retry_only_recorded_failures(self):
        TatTrackerCase.objects.create(
            group_id=self.config.group_id,
            sheet_id=self.config.sheet_id,
            sheet_name='TRACKER-Business',
            case_id='JBL-BS-FAILED',
            product_key='business',
            product_label='Business',
            client_name='Failed Client',
            branch='Nakuru',
        )
        job = TatRepairJob.objects.create(
            group_configuration=self.config,
            case_ids=['JBL-BS-FAILED'],
            total_cases=1,
            cursor=1,
            status='completed_with_errors',
            failures=[{'case_id': 'JBL-BS-FAILED', 'error': 'quota'}],
        )

        response = self.client.post(self.url, {
            'action': 'retry_failures',
            'job_id': str(job.id),
            'confirm': 'RETRY FAILED',
        })

        self.assertEqual(response.status_code, 302)
        retry_job = TatRepairJob.objects.exclude(pk=job.pk).get()
        self.assertEqual(retry_job.case_ids, ['JBL-BS-FAILED'])
        self.assertEqual(retry_job.status, 'queued')

    @patch('core.admin.resync_tat_tracker_cases')
    def test_repair_page_rejects_a_write_without_matching_preview(self, resync):
        response = self.client.post(self.url, {'confirm': 'REPAIR', 'product': 'business', 'offset': '0'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Preview this exact batch')
        resync.assert_not_called()
