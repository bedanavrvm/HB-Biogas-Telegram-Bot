"""
Tests for the biogas telegram bot system.

Run with: python manage.py test
"""
import json
import zipfile
from io import BytesIO, StringIO
from datetime import datetime, timedelta
from decimal import Decimal
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.contrib.auth import get_user_model
from django.utils import timezone
from unittest.mock import patch, MagicMock

from core.models import (
    CaseUpdate,
    FcaImportRecord,
    GroupSheetConfiguration,
    JawabuFarmerMaster,
    JawabuFarmerUploadBatch,
    JawabuVisitRecord,
    LiveSheetRecordChange,
    OrderApprovalUpdate,
    MediaAttachment,
    RawMessage,
    ProcessedMessage,
    ParsedMessage,
)
from core.services.deduplication import generate_message_hash, is_duplicate
from core.services.parser import (
    analyze_whatsapp_export,
    parse_message,
    split_batch_message,
)
from core.services.sheets import GoogleSheetsService, batch_append_messages
from core.services.sheet_sync import sync_sheet_to_backend
from core.services.storage import (
    MessageRejectedError,
    bulk_resync_to_sheets,
    process_and_store_message,
)


def create_parsed_case(
    message_id: str,
    group_id: str = '-100123',
    customer_name: str = 'Jane Doe',
    customer_phone: str = '0712345678',
    customer_id: str = 'CUST-1',
    description: str = 'No gas supply',
    complaint_status: str = '',
    branch_region: str = '',
    complaint_category: str = '',
    risk_level: str = '',
    synced_to_sheets: bool = False,
    sheet_id: str = '',
    sheet_name: str = '',
    last_sync_error: str = '',
    processed_status: str = 'success',
    created_at=None,
) -> ParsedMessage:
    """Create a parsed case with its raw/processed parents for tests."""
    raw = RawMessage.objects.create(
        telegram_message_id=message_id,
        source_telegram_message_id=message_id,
        sender='Agent',
        content=description,
    )
    processed = ProcessedMessage.objects.create(
        message_hash=generate_message_hash('Agent', f'{message_id}:{description}'),
        raw_message=raw,
        status=processed_status,
    )
    parsed = ParsedMessage.objects.create(
        processed_message=processed,
        message_id=message_id,
        timestamp=created_at or timezone.now(),
        customer_name=customer_name,
        customer_phone=customer_phone,
        customer_id=customer_id,
        branch_region=branch_region,
        complaint_category=complaint_category,
        complaint_status=complaint_status,
        complaint_description=description,
        risk_level=risk_level,
        raw_message=description,
        group_id=group_id,
        sheet_id=sheet_id,
        sheet_name=sheet_name,
        source='telegram bot',
        synced_to_sheets=synced_to_sheets,
        last_sync_error=last_sync_error,
    )
    if created_at:
        ParsedMessage.objects.filter(pk=parsed.pk).update(created_at=created_at)
        parsed.refresh_from_db()
    return parsed


class DeduplicationServiceTest(TestCase):
    """Test the deduplication service."""
    
    def test_generate_message_hash_deterministic(self):
        """Same inputs should produce same hash."""
        hash1 = generate_message_hash("John", "Sold 3 bread 50 each")
        hash2 = generate_message_hash("John", "Sold 3 bread 50 each")
        self.assertEqual(hash1, hash2)


class CaseUpdateServiceTest(TestCase):
    """Test chat-driven case status updates."""

    def _patch_sheet_success(self):
        group_config = MagicMock(sheet_id='sheet_123', sheet_name='Complaints')
        registry = MagicMock()
        registry.get_group.return_value = group_config
        sheet = MagicMock()
        sheet.update_case_row.return_value = True
        return (
            patch('core.services.case_updates.GroupRegistry.get_instance', return_value=registry),
            patch('core.services.case_updates.get_sheets_service', return_value=sheet),
            sheet,
        )

    def test_parse_case_update_resolved(self):
        from core.services.case_updates import parse_case_update

        result = parse_case_update('Status: resolved - jiko relocated')

        self.assertTrue(result.is_update)
        self.assertEqual(result.new_status, 'Closed')
        self.assertEqual(result.resolution_text, 'jiko relocated')

    def test_parse_case_update_uses_note_as_resolution_text(self):
        from core.services.case_updates import parse_case_update

        result = parse_case_update(
            'Status: resolved\nNOTE: Jiko relocated and customer confirmed'
        )

        self.assertTrue(result.is_update)
        self.assertEqual(result.new_status, 'Closed')
        self.assertEqual(
            result.resolution_text,
            'Jiko relocated and customer confirmed',
        )

    def test_reply_status_update_updates_case_and_sheet(self):
        from core.services.case_updates import handle_case_status_reply

        case = create_parsed_case(
            'MSG_UPDATE_1',
            complaint_status='Open',
            synced_to_sheets=True,
        )
        case.processed_message.raw_message.telegram_message_id = '777'
        case.processed_message.raw_message.source_telegram_message_id = '777'
        case.processed_message.raw_message.save()

        registry_patch, sheet_patch, sheet = self._patch_sheet_success()
        with registry_patch, sheet_patch:
            result = handle_case_status_reply(
                group_id='-100123',
                reply_to_telegram_message_id='777',
                update_telegram_message_id='888',
                sender='Peter',
                content='Status: resolved - jiko relocated successfully',
            )

        case.refresh_from_db()
        self.assertEqual(result['status'], 'command')
        self.assertIn('Case updated', result['reply_text'])
        self.assertIn('Case ID: MSG_UPDATE_1', result['reply_text'])
        self.assertEqual(case.complaint_status, 'Closed')
        self.assertIsNotNone(case.date_resolved)
        self.assertEqual(case.resolution_details, 'jiko relocated successfully')
        update = CaseUpdate.objects.get(parsed_message=case)
        self.assertEqual(update.sync_status, 'success')
        sheet.update_case_row.assert_called_once()
        args, _ = sheet.update_case_row.call_args
        self.assertEqual(args[0], 'MSG_UPDATE_1')
        self.assertEqual(args[1]['status'], 'Closed')
        self.assertEqual(args[1]['resolution_details'], 'jiko relocated successfully')
        self.assertIn('date_resolved', args[1])

    def test_reply_status_update_writes_note_to_resolution_details(self):
        from core.services.case_updates import handle_case_status_reply

        case = create_parsed_case('MSG_UPDATE_NOTE', complaint_status='Open')
        case.processed_message.raw_message.telegram_message_id = '771'
        case.processed_message.raw_message.source_telegram_message_id = '771'
        case.processed_message.raw_message.save()

        registry_patch, sheet_patch, sheet = self._patch_sheet_success()
        with registry_patch, sheet_patch:
            result = handle_case_status_reply(
                group_id='-100123',
                reply_to_telegram_message_id='771',
                update_telegram_message_id='772',
                sender='Peter',
                content=(
                    'Status: resolved\n'
                    'NOTE: Jiko relocated and customer confirmed'
                ),
            )

        case.refresh_from_db()
        self.assertEqual(result['status'], 'command')
        self.assertEqual(case.complaint_status, 'Closed')
        self.assertEqual(
            case.resolution_details,
            'Jiko relocated and customer confirmed',
        )
        self.assertNotIn('NOTE:', case.resolution_details)
        args, _ = sheet.update_case_row.call_args
        self.assertEqual(
            args[1]['resolution_details'],
            'Jiko relocated and customer confirmed',
        )
        self.assertNotIn('NOTE:', args[1]['resolution_details'])

    def test_reply_status_update_does_not_update_db_when_sheet_fails(self):
        from core.services.case_updates import handle_case_status_reply

        case = create_parsed_case('MSG_UPDATE_FAIL', complaint_status='Open')
        case.processed_message.raw_message.telegram_message_id = '700'
        case.processed_message.raw_message.source_telegram_message_id = '700'
        case.processed_message.raw_message.save()

        group_config = MagicMock(sheet_id='sheet_123', sheet_name='Complaints')
        registry = MagicMock()
        registry.get_group.return_value = group_config
        sheet = MagicMock()
        sheet.update_case_row.return_value = False

        with patch('core.services.case_updates.GroupRegistry.get_instance', return_value=registry):
            with patch('core.services.case_updates.get_sheets_service', return_value=sheet):
                result = handle_case_status_reply(
                    group_id='-100123',
                    reply_to_telegram_message_id='700',
                    update_telegram_message_id='701',
                    sender='Peter',
                    content='Status: resolved - fixed',
                )

        case.refresh_from_db()
        self.assertEqual(case.complaint_status, 'Open')
        self.assertIsNone(case.date_resolved)
        self.assertIn('not update the register', result['reply_text'])
        update = CaseUpdate.objects.get(parsed_message=case)
        self.assertEqual(update.sync_status, 'failed')

    def test_reply_status_update_requires_case_id_for_batch_ambiguity(self):
        from core.services.case_updates import handle_case_status_reply

        first = create_parsed_case('MSG_BATCH_A', customer_name='Alice')
        second = create_parsed_case('MSG_BATCH_B', customer_name='Bob')
        first.processed_message.raw_message.telegram_message_id = '900_0'
        first.processed_message.raw_message.source_telegram_message_id = '900'
        first.processed_message.raw_message.batch_index = 0
        first.processed_message.raw_message.save()
        second.processed_message.raw_message.telegram_message_id = '900_1'
        second.processed_message.raw_message.source_telegram_message_id = '900'
        second.processed_message.raw_message.batch_index = 1
        second.processed_message.raw_message.save()

        result = handle_case_status_reply(
            group_id='-100123',
            reply_to_telegram_message_id='900',
            update_telegram_message_id='901',
            sender='Peter',
            content='Status: resolved',
        )

        self.assertEqual(result['status'], 'command')
        self.assertIn('/update MSG_ID', result['reply_text'])
        self.assertEqual(CaseUpdate.objects.count(), 0)

    def test_reply_to_bot_confirmation_uses_case_id_in_quoted_text(self):
        from core.services.case_updates import handle_case_status_reply

        case = create_parsed_case('MSG_CONFIRM_1', complaint_status='Open')
        registry_patch, sheet_patch, _sheet = self._patch_sheet_success()

        with registry_patch, sheet_patch:
            result = handle_case_status_reply(
                group_id='-100123',
                reply_to_telegram_message_id='999',
                update_telegram_message_id='1000',
                sender='Peter',
                content='Status: resolved - jiko relocated',
                reply_to_text=(
                    'OK. Message received and saved successfully\n'
                    'Case ID: MSG_CONFIRM_1\n'
                    'Captured: Customer Name, Customer Phone'
                ),
            )

        case.refresh_from_db()
        self.assertEqual(result['status'], 'command')
        self.assertEqual(case.complaint_status, 'Closed')
        self.assertIn('jiko relocated', case.resolution_details)

    def test_reply_to_original_case_can_match_from_quoted_case_text(self):
        from core.services.case_updates import handle_case_status_reply

        case = create_parsed_case(
            'MSG_ORIGINAL_TEXT',
            customer_name='Henry mwenda',
            customer_phone='0720809218',
            customer_id='24289449',
            description='Requesting for a jiko relocation',
            complaint_status='Open',
        )
        registry_patch, sheet_patch, _sheet = self._patch_sheet_success()

        with registry_patch, sheet_patch:
            result = handle_case_status_reply(
                group_id='-100123',
                reply_to_telegram_message_id='123',
                update_telegram_message_id='124',
                sender='Peter',
                content='Status: resolved - jiko relocated',
                reply_to_text=(
                    '@hb_biogas_cases_bot Henry  mwenda\n'
                    '24289449\n'
                    '0720809218/0726011961\n\n'
                    'Requesting for a jiko relocation'
                ),
            )

        case.refresh_from_db()
        self.assertEqual(result['status'], 'command')
        self.assertEqual(case.complaint_status, 'Closed')
        self.assertIn('jiko relocated', case.resolution_details)

    def test_explicit_update_command_updates_case(self):
        from core.services.commands import handle_bot_command

        case = create_parsed_case('MSG_CMD_1', complaint_status='Open')
        registry_patch, sheet_patch, _sheet = self._patch_sheet_success()

        with registry_patch, sheet_patch:
            result = handle_bot_command(
                '/update MSG_CMD_1 Status: scheduled for Thursday',
                '-100123',
                sender='Peter',
                telegram_message_id='901',
            )

        case.refresh_from_db()
        self.assertEqual(result['status'], 'command')
        self.assertEqual(case.complaint_status, 'In Progress')
        self.assertIn('scheduled for Thursday', case.resolution_details)


class DeduplicationHashTest(TestCase):
    """Additional deduplication hash tests."""

    def test_generate_message_hash_different_content(self):
        """Different content should produce different hash."""
        hash1 = generate_message_hash("John", "Sold 3 bread 50 each")
        hash2 = generate_message_hash("John", "Sold 5 milk 100 each")
        self.assertNotEqual(hash1, hash2)
    
    def test_generate_message_hash_case_insensitive(self):
        """Hash should be case-insensitive for content."""
        hash1 = generate_message_hash("John", "SOLD 3 BREAD 50 EACH")
        hash2 = generate_message_hash("john", "sold 3 bread 50 each")
        self.assertEqual(hash1, hash2)
    
    def test_is_duplicate_new_message(self):
        """New message should not be duplicate."""
        msg_hash = generate_message_hash("Test", "Unique content 12345")
        self.assertFalse(is_duplicate(msg_hash))
    
    def test_is_duplicate_after_processing(self):
        """Message should be duplicate after processing."""
        raw_message = RawMessage.objects.create(
            telegram_message_id='test_123',
            sender='Test User',
            content='Test content for dedup',
        )
        msg_hash = generate_message_hash('Test User', 'Test content for dedup')
        
        ProcessedMessage.objects.create(
            message_hash=msg_hash,
            raw_message=raw_message,
            status='success',
        )
        
        self.assertTrue(is_duplicate(msg_hash))

    def test_is_duplicate_ignores_failed_status(self):
        """Failed processing attempts should not be considered duplicates."""
        raw_message = RawMessage.objects.create(
            telegram_message_id='test_456',
            sender='Test User',
            content='Test content for dedup fail',
        )
        msg_hash = generate_message_hash('Test User', 'Test content for dedup fail')
        ProcessedMessage.objects.create(
            message_hash=msg_hash,
            raw_message=raw_message,
            status='failed',
        )
        self.assertFalse(is_duplicate(msg_hash))


class ParserServiceTest(TestCase):
    """Test the message parser."""
    
    def test_parse_sold_pattern(self):
        """Test parsing 'Sold X item Y each' pattern."""
        result = parse_message("Sold 3 bread 50 each to John", sender="Seller")
        
        self.assertEqual(result.quantity, Decimal('3'))
        self.assertEqual(result.item, 'bread')
        self.assertEqual(result.price, Decimal('50'))
        self.assertGreater(result.confidence, 0.5)
    
    def test_parse_paid_pattern(self):
        """Test parsing 'X paid Y for Z item' pattern."""
        result = parse_message("John paid 200 for 4 milk", sender="John")
        
        self.assertEqual(result.price, Decimal('200'))
        self.assertEqual(result.quantity, Decimal('4'))
        self.assertEqual(result.item, 'milk')
    
    def test_parse_gps_url(self):
        """Test GPS URL extraction."""
        content = "📍 https://maps.app.goo.gl/abc123 Sold 2 bags maize"
        result = parse_message(content, sender="Seller")
        
        self.assertIn('https://maps.app.goo.gl/abc123', result.gps_link)
        self.assertEqual(result.item, 'bags maize')
    
    def test_parse_image_flag(self):
        """Test image flag is set correctly."""
        result = parse_message("Sold 3 bread", has_image=True)
        self.assertTrue(result.image_flag)
    
    def test_parse_empty_message(self):
        """Test empty message handling."""
        result = parse_message("")
        self.assertEqual(result.confidence, 0.0)
        self.assertIn("Empty message content", result.warnings)
    
    def test_split_batch_message(self):
        """Test splitting batch forwarded messages."""
        batch_content = """[14/03/2026, 10:30:15] John: Sold 3 bread 50 each
[14/03/2026, 10:31:20] Mary: John paid 200 for 4 milk"""
        
        messages = split_batch_message(batch_content)
        
        self.assertEqual(len(messages), 2)
        self.assertEqual(messages[0]['sender'], 'John')
        self.assertEqual(messages[1]['sender'], 'Mary')

    def test_analyze_whatsapp_export_handles_dash_format_and_continuations(self):
        """WhatsApp .txt exports should be split without saving system lines."""
        export_text = """23/05/2026, 12:46 - Messages and calls are end-to-end encrypted.
23/05/2026, 12:47 - Alice Agent: CUSTOMER COMPLAIN
NAME: Jane Doe
TEL: 0712345678
ID: A123
COUNTY: KISUMU
NATURE OF THE PROBLEM: No gas supply
23/05/2026, 12:48 - Bob Agent: Normal group chat"""

        analysis = analyze_whatsapp_export(export_text)

        self.assertEqual(analysis['format'], 'whatsapp_dash_export')
        self.assertEqual(analysis['system_lines'], 1)
        self.assertEqual(analysis['entry_count'], 2)
        self.assertEqual(analysis['entries'][0]['sender'], 'Alice Agent')
        self.assertIn('NAME: Jane Doe', analysis['entries'][0]['content'])
        self.assertEqual(analysis['entries'][1]['content'], 'Normal group chat')
        self.assertIsNotNone(analysis['entries'][0]['received_at'])

    def test_analyze_whatsapp_export_handles_bracketed_format(self):
        """Bracketed WhatsApp exports should also be accepted."""
        export_text = """[23/05/2026 12:46] Alice Agent: CUSTOMER COMPLAIN
NAME: Jane Doe
TEL: 0712345678
ID: A123
COUNTY: KISUMU
NATURE OF THE PROBLEM: No gas supply"""

        analysis = analyze_whatsapp_export(export_text)

        self.assertEqual(analysis['format'], 'whatsapp_bracketed_export')
        self.assertEqual(analysis['entry_count'], 1)
        self.assertEqual(analysis['entries'][0]['sender'], 'Alice Agent')

    def test_split_multiple_complaint_cases(self):
        """One message with repeated complaint headers should split into cases."""
        batch_content = """*CUSTOMER COMPLAIN*
NAME: Jane Doe
TEL: 0712345678
ID: A123
NATURE OF THE PROBLEM: No gas supply
*CUSTOMER COMPLAIN: The system has stopped producing gas

*CUSTOMER COMPLAIN*
NAME: John Smith
TEL: 0798765432
ID: B456
NATURE OF THE PROBLEM: Gas leakage
*CUSTOMER COMPLAIN: Gas smell around the digester"""

        messages = split_batch_message(batch_content)

        self.assertEqual(len(messages), 2)
        self.assertEqual(messages[0]['sender'], 'Jane Doe')
        self.assertEqual(messages[1]['sender'], 'John Smith')
        self.assertIn('The system has stopped producing gas', messages[0]['content'])
        self.assertIn('Gas smell around the digester', messages[1]['content'])

    def test_split_complaint_ignores_description_only_fragment(self):
        """A complaint header without identifiers should not become its own case."""
        batch_content = """*CUSTOMER COMPLAIN*
NAME: Jane Doe
TEL: 0712345678
ID: A123
 NATURE OF COMPLAIN: No gas supply
*CUSTOMER COMPLAIN*
The system has stopped producing gas"""

        messages = split_batch_message(batch_content)

        self.assertEqual(len(messages), 1)

    def test_split_multiple_complaint_cases_merges_description_fragments(self):
        """Description-only complaint fragments belong to the previous case."""
        batch_content = """*CUSTOMER COMPLAIN*
NAME: Jane Doe
TEL: 0712345678
ID: A123
 NATURE OF COMPLAIN: No gas supply
*CUSTOMER COMPLAIN*
The system has stopped producing gas

*CUSTOMER COMPLAIN*
NAME: John Smith
TEL: 0798765432
ID: B456
NATURE OF THE COMPLAINT: Gas leakage
*CUSTOMER COMPLAIN*
Gas smell around the digester"""

        messages = split_batch_message(batch_content)

        self.assertEqual(len(messages), 2)
        self.assertEqual(messages[0]['sender'], 'Jane Doe')
        self.assertEqual(messages[1]['sender'], 'John Smith')
        self.assertIn('The system has stopped producing gas', messages[0]['content'])
        self.assertIn('Gas smell around the digester', messages[1]['content'])

    def test_split_unlabeled_complaint_cases(self):
        """Unlabeled complaint blocks should still split when complete."""
        batch_content = """*CUSTOMER COMPLAIN*
Jane Doe
0712345678
A123
No gas supply at home

*CUSTOMER COMPLAIN*
John Smith
0798765432
B456
Gas leaking around the digester"""

        messages = split_batch_message(batch_content)

        self.assertEqual(len(messages), 2)
        self.assertIn('Jane Doe', messages[0]['content'])
        self.assertIn('John Smith', messages[1]['content'])
    
    def test_intent_detection_sale(self):
        """Test intent detection for sale messages."""
        from core.services.parser import detect_message_intent, MessageIntent
        
        result = detect_message_intent("Sold 3 bread 50 each to John")
        self.assertEqual(result, MessageIntent.SALE)
        
        result = detect_message_intent("Delivered 2 bags maize")
        self.assertEqual(result, MessageIntent.SALE)


class JawabuWorkflowServiceTest(TestCase):
    """Tests for the Jawabu HomeBiogas WhatsApp import workflow."""

    def setUp(self):
        self.group = GroupSheetConfiguration.objects.create(
            group_id='-100jawabu',
            display_name='Jawabu HomeBiogas',
            enabled=True,
            sheet_id='sheet_jawabu',
            sheet_name='Jawabu Visits',
            workflow={'type': 'jawabu_homebiogas', 'header_row': 1, 'import_start_date': ''},
        )

    def test_extract_jawabu_fields_uses_national_id_and_primary_phone(self):
        from core.services.jawabu import extract_jawabu_fields, jawabu_duplicate_key

        content = """IMG-20260316-WA0005.jpg (file attached)
Latitude: S 0°33'14.148"
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031/0785116424
Case deferred, client has seasonal income."""

        fields = extract_jawabu_fields(content, 'Alex Kairu', timezone.now())

        self.assertEqual(fields['customer_name'], 'MARY NJERI NJIHIA')
        self.assertEqual(fields['national_id'], '1382654')
        self.assertEqual(fields['primary_phone'], '254720570031')
        self.assertEqual(fields['secondary_phone'], '254785116424')
        self.assertEqual(fields['county'], 'EMBU')
        self.assertEqual(fields['latitude'], '-0.55393')
        self.assertEqual(fields['longitude'], '37.526935')
        self.assertEqual(fields['decision'], 'DEFERRED')
        self.assertEqual(
            fields['duplicate_key'],
            jawabu_duplicate_key('1382654', '0720570031'),
        )

    def test_whatsapp_zip_export_text_is_extracted(self):
        from core.api.views import _extract_whatsapp_export_text

        zip_buffer = BytesIO()
        export_text = "3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)\n0720570031"
        with zipfile.ZipFile(zip_buffer, 'w') as archive:
            archive.writestr('WhatsApp Chat with Jawabu-HomeBiogas.txt', export_text)
            archive.writestr('IMG-20260316-WA0005.jpg', b'image-bytes')

        extracted = _extract_whatsapp_export_text(
            zip_buffer.getvalue(),
            'WhatsApp Chat with Jawabu-HomeBiogas.zip',
            5,
        )

        self.assertEqual(extracted, export_text)

    def test_whatsapp_export_parses_mm_dd_yy_and_ampm_timestamps(self):
        from core.services.parser import analyze_whatsapp_export

        export = "4/13/26, 11:49 AM - Dickson JBL: IMG-20260413-WA0042.jpg (file attached)\n0711651414"

        analysis = analyze_whatsapp_export(export)

        self.assertEqual(len(analysis['entries']), 1)
        received_at = timezone.localtime(analysis['entries'][0]['received_at'])
        self.assertEqual(received_at.strftime('%d-%b-%Y %H:%M'), '13-Apr-2026 11:49')

    def test_jawabu_parser_uses_location_date_and_avoids_image_date_as_id(self):
        from core.services.jawabu import extract_jawabu_fields

        content = """IMG-20260410-WA0052.jpg (file attached)
Latitude: S 0?36'8.10036"
Longitude: E 36?57'50.45004"
https://www.google.com/maps/search/?api=1&query=-0.6022501,36.9640139

Altitude: -
Location read date: 10/04/2026 14:20
Location source: Fused

Street: -Kaganjo
City: -Kiriani
State: -Kaganjo
County: - Muranga
Beatrice Nyachui Wachira
0727769644
Id. 8840023"""
        received_at = timezone.make_aware(
            datetime(2026, 4, 10, 14, 29),
            timezone.get_current_timezone(),
        )

        fields = extract_jawabu_fields(content, 'John Muiruri JBL Biogas', received_at)

        self.assertEqual(fields['visit_date'], '10-Apr-2026')
        self.assertEqual(fields['whatsapp_message_at'], '10-Apr-2026 14:29')
        self.assertEqual(fields['customer_name'], 'BEATRICE NYACHUI WACHIRA')
        self.assertEqual(fields['national_id'], '8840023')
        self.assertEqual(fields['primary_phone'], '254727769644')
        self.assertEqual(fields['county'], 'MURANGA')
        self.assertEqual(fields['sub_county'], 'KIRIANI')
        self.assertEqual(fields['landmark'], 'KAGANJO')
        self.assertEqual(fields['latitude'], '-0.6022501')
        self.assertEqual(fields['longitude'], '36.9640139')

    def test_jawabu_parser_uses_whatsapp_date_to_choose_mm_dd_location_date(self):
        from core.services.jawabu import extract_jawabu_fields

        content = """IMG-20260410-WA0052.jpg (file attached)
Latitude: S 0?36'8.10036"
Longitude: E 36?57'50.45004"
https://www.google.com/maps/search/?api=1&query=-0.6022501,36.9640139
Location read date: 04/10/2026 14:20
State: -Kaganjo
County: - Muranga
Beatrice Nyachui Wachira
0727769644
Id. 8840023"""
        received_at = timezone.make_aware(
            datetime(2026, 4, 10, 14, 29),
            timezone.get_current_timezone(),
        )

        fields = extract_jawabu_fields(content, 'John Muiruri JBL Biogas', received_at)

        self.assertEqual(fields['visit_date'], '10-Apr-2026')
        self.assertEqual(fields['whatsapp_message_at'], '10-Apr-2026 14:29')

    def test_jawabu_parser_never_uses_location_date_after_whatsapp_time(self):
        from core.services.jawabu import extract_jawabu_fields

        content = """IMG-20260410-WA0052.jpg (file attached)
Latitude: S 0?36'8.10036"
Longitude: E 36?57'50.45004"
https://www.google.com/maps/search/?api=1&query=-0.6022501,36.9640139
Location read date: 06/04/2026 14:20
State: -Kaganjo
County: - Muranga
Beatrice Nyachui Wachira
0727769644
Id. 8840023"""
        received_at = timezone.make_aware(
            datetime(2026, 4, 10, 14, 29),
            timezone.get_current_timezone(),
        )

        fields = extract_jawabu_fields(content, 'John Muiruri JBL Biogas', received_at)

        self.assertEqual(fields['visit_date'], '06-Apr-2026')
        self.assertEqual(fields['whatsapp_message_at'], '10-Apr-2026 14:29')

    def test_jawabu_parser_ignores_unambiguous_future_location_date(self):
        from core.services.jawabu import extract_jawabu_fields

        content = """IMG-20260410-WA0052.jpg (file attached)
Latitude: S 0?36'8.10036"
Longitude: E 36?57'50.45004"
https://www.google.com/maps/search/?api=1&query=-0.6022501,36.9640139
Date: Jun 4, 2026 02:20 PM
State: -Kaganjo
County: - Muranga
Beatrice Nyachui Wachira
0727769644
Id. 8840023"""
        received_at = timezone.make_aware(
            datetime(2026, 4, 10, 14, 29),
            timezone.get_current_timezone(),
        )

        fields = extract_jawabu_fields(content, 'John Muiruri JBL Biogas', received_at)

        self.assertEqual(fields['visit_date'], '10-Apr-2026')
        self.assertEqual(fields['whatsapp_message_at'], '10-Apr-2026 14:29')

    def test_jawabu_parser_uses_message_date_when_location_date_missing(self):
        from core.services.jawabu import extract_jawabu_fields

        content = """IMG-20260413-WA0042.jpg (file attached)
Latitude: S 0?25'19.75116"
Longitude: E 37?41'22.09632"
https://www.google.com/maps/search/?api=1&query=-0.4221531,37.6894712
State: Tharaka-Nithi County
Country: Kenya
Dinah karimi
10970922
0711651414"""
        received_at = timezone.make_aware(
            datetime(2026, 4, 13, 11, 49),
            timezone.get_current_timezone(),
        )

        fields = extract_jawabu_fields(content, 'Dickson JBL', received_at)

        self.assertEqual(fields['visit_date'], '13-Apr-2026')
        self.assertEqual(fields['whatsapp_message_at'], '13-Apr-2026 11:49')
        self.assertEqual(fields['national_id'], '10970922')
        self.assertEqual(fields['primary_phone'], '254711651414')
        self.assertEqual(fields['county'], 'THARAKA-NITHI')

    def test_jawabu_parser_fills_message_time_from_location_date_if_header_missing(self):
        from core.services.jawabu import extract_jawabu_fields

        content = """IMG-20260410-WA0048.jpg (file attached)
Latitude: -0.726058
Longitude: 37.081925
Address: C72, Kenya
Date: Mar 25, 2026 12:45:46 PM
Google Maps: https://maps.google.com/maps?q=-0.7260583%2C37.081925
Street: Kwambari
City:Kiharu
County:Muranga
Dan Irungu Kamau
Id. 5919286
0722429932
Opted for cash"""

        fields = extract_jawabu_fields(content, 'John Muiruri JBL Biogas', None)

        self.assertEqual(fields['visit_date'], '25-Mar-2026')
        self.assertEqual(fields['whatsapp_message_at'], '25-Mar-2026 12:45')
        self.assertEqual(fields['national_id'], '5919286')
        self.assertEqual(fields['primary_phone'], '254722429932')
        self.assertEqual(fields['latitude'], '-0.7260583')
        self.assertEqual(fields['longitude'], '37.081925')

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_appends_unique_records_to_sheet(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        fake_sheet = FakeJawabuSheet(list(JAWABU_FIELD_HEADERS.values()))
        mock_service.return_value = FakeJawabuService(fake_sheet)

        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='900',
            sender='Importer',
        )

        self.assertEqual(result['status'], 'jawabu_batch_processed')
        self.assertEqual(result['imported'], 1)
        self.assertEqual(result['duplicate_review'], 0)
        self.assertEqual(len(fake_sheet.appended_rows), 1)
        self.assertEqual(fake_sheet.row_values_calls, 1)
        self.assertEqual(fake_sheet.get_all_values_calls, 1)
        self.assertEqual(fake_sheet.append_rows_calls, 1)
        record = JawabuVisitRecord.objects.get()
        self.assertEqual(record.import_status, 'imported')
        self.assertEqual(record.duplicate_status, 'unique')
        self.assertEqual(record.national_id, '1382654')
        self.assertEqual(record.primary_phone, '254720570031')
        status_index = list(JAWABU_FIELD_HEADERS.values()).index('Import Status')
        duplicate_index = list(JAWABU_FIELD_HEADERS.values()).index('Duplicate Status')
        self.assertEqual(fake_sheet.appended_rows[0][status_index], 'Imported')
        self.assertEqual(fake_sheet.appended_rows[0][duplicate_index], 'Unique')

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_accepts_name_and_national_id_without_phone(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        fake_sheet = FakeJawabuSheet(list(JAWABU_FIELD_HEADERS.values()))
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri Njihia
1382654"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='902',
            sender='Importer',
        )

        self.assertEqual(result['imported'], 1)
        record = JawabuVisitRecord.objects.get()
        self.assertEqual(record.import_status, 'imported')
        self.assertEqual(record.parsed_fields['customer_name'], 'MARY NJERI NJIHIA')
        self.assertEqual(record.national_id, '1382654')
        self.assertEqual(record.primary_phone, '')

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_accepts_name_and_primary_phone_without_id(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        fake_sheet = FakeJawabuSheet(list(JAWABU_FIELD_HEADERS.values()))
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: <Media omitted>
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri Njihia
0720570031"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='903',
            sender='Importer',
        )

        self.assertEqual(result['imported'], 1)
        record = JawabuVisitRecord.objects.get()
        self.assertEqual(record.import_status, 'imported')
        self.assertEqual(record.parsed_fields['customer_name'], 'MARY NJERI NJIHIA')
        self.assertEqual(record.national_id, '')
        self.assertEqual(record.primary_phone, '254720570031')

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_rejects_identifier_without_name(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        fake_sheet = FakeJawabuSheet(list(JAWABU_FIELD_HEADERS.values()))
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
1382654
0720570031"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='904',
            sender='Importer',
        )

        self.assertEqual(result['imported'], 0)
        self.assertEqual(result['rejected'], 1)
        self.assertEqual(result['rejections'][0]['missing_fields'], ['Customer Name'])
        self.assertEqual(len(fake_sheet.appended_rows), 1)
        status_index = list(JAWABU_FIELD_HEADERS.values()).index('Import Status')
        notes_index = list(JAWABU_FIELD_HEADERS.values()).index('Review Notes')
        self.assertEqual(fake_sheet.appended_rows[0][status_index], 'Rejected')
        self.assertIn('Missing required field', fake_sheet.appended_rows[0][notes_index])
        record = JawabuVisitRecord.objects.get()
        self.assertEqual(record.import_status, 'rejected')

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_batch_appends_unique_records_once(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        fake_sheet = FakeJawabuSheet(list(JAWABU_FIELD_HEADERS.values()))
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri Njihia
1382654
3/16/26, 11:47 - Alex Kairu: <Media omitted>
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Meru County
Country: Kenya
John Mwangi
0720570031"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='905',
            sender='Importer',
        )

        self.assertEqual(result['imported'], 2)
        self.assertEqual(len(fake_sheet.appended_rows), 2)
        self.assertEqual(fake_sheet.row_values_calls, 1)
        self.assertEqual(fake_sheet.get_all_values_calls, 1)
        self.assertEqual(fake_sheet.append_rows_calls, 1)

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_skips_messages_before_configured_start_date(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        self.group.workflow = {'type': 'jawabu_homebiogas', 'header_row': 1}
        self.group.save(update_fields=['workflow'])
        fake_sheet = FakeJawabuSheet(list(JAWABU_FIELD_HEADERS.values()))
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """4/30/26, 11:46 AM - Alex Kairu: IMG-20260430-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
April Customer
1382654
0720570031
5/1/26, 09:15 AM - Alex Kairu: IMG-20260501-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
May Customer
8840023
0727769644"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='906',
            sender='Importer',
        )

        self.assertEqual(result['export_messages'], 2)
        self.assertEqual(result['skipped_before_start'], 1)
        self.assertEqual(result['imported'], 1)
        self.assertEqual(len(fake_sheet.appended_rows), 1)
        record = JawabuVisitRecord.objects.get()
        self.assertEqual(record.national_id, '8840023')
        self.assertEqual(record.primary_phone, '254727769644')

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_skips_messages_at_or_before_latest_whatsapp_time(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        headers = list(JAWABU_FIELD_HEADERS.values())
        fake_sheet = FakeJawabuSheet(headers)
        mock_service.return_value = FakeJawabuService(fake_sheet)
        first_export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031"""

        first = process_jawabu_batch_export(
            group_config=self.group,
            export_text=first_export,
            telegram_message_id='911',
            sender='Importer',
        )

        self.assertEqual(first['imported'], 1)
        second_export = first_export + """
3/16/26, 11:50 - Alex Kairu: IMG-20260316-WA0006.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55394,37.526936
State: Embu County
Country: Kenya
Jane Wanjiku
7654321
0720570032"""

        second = process_jawabu_batch_export(
            group_config=self.group,
            export_text=second_export,
            telegram_message_id='912',
            sender='Importer',
        )

        self.assertEqual(second['skipped_already_processed'], 1)
        self.assertEqual(second['latest_processed_at'], '16-Mar-2026 11:46')
        self.assertEqual(second['imported'], 1)
        self.assertEqual(second['duplicate_review'], 0)
        self.assertEqual(JawabuVisitRecord.objects.count(), 2)
        self.assertEqual(len(fake_sheet.appended_rows), 2)
        name_index = headers.index('Customer Name')
        self.assertEqual(fake_sheet.appended_rows[-1][name_index], 'JANE WANJIKU')

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_uses_configured_start_when_db_was_reset(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        headers = list(JAWABU_FIELD_HEADERS.values())
        fake_sheet = FakeJawabuSheet(headers)
        existing_row = ['' for _ in headers]
        existing_row[headers.index('WhatsApp Message Time')] = '16-Mar-2026 11:46'
        fake_sheet.appended_rows.append(existing_row)
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031
3/16/26, 11:50 - Alex Kairu: IMG-20260316-WA0006.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55394,37.526936
State: Embu County
Country: Kenya
Jane Wanjiku
7654321
0720570032"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='913',
            sender='Importer',
        )

        self.assertEqual(result['skipped_already_processed'], 0)
        self.assertEqual(result['latest_processed_at'], '')
        self.assertEqual(result['imported'], 2)
        self.assertEqual(result['duplicate_review'], 0)
        self.assertEqual(JawabuVisitRecord.objects.count(), 2)

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_ignores_stale_db_duplicates_removed_from_sheet(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        JawabuVisitRecord.objects.create(
            group_id=self.group.group_id,
            sheet_id=self.group.sheet_id,
            sheet_tab=self.group.sheet_name,
            telegram_message_id='old_jawabu_1',
            source_telegram_message_id='old',
            whatsapp_message_index=1,
            sender='Old Importer',
            national_id='1382654',
            primary_phone='254720570031',
            duplicate_key='ID:1382654|PHONE:254720570031',
            duplicate_status='unique',
            import_status='imported',
            parsed_fields={'duplicate_key': 'ID:1382654|PHONE:254720570031'},
            raw_text='old row removed from sheet',
        )
        fake_sheet = FakeJawabuSheet(list(JAWABU_FIELD_HEADERS.values()))
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='907',
            sender='Importer',
        )

        self.assertEqual(result['imported'], 1)
        self.assertEqual(result['duplicate_review'], 0)
        self.assertEqual(JawabuVisitRecord.objects.count(), 1)
        record = JawabuVisitRecord.objects.get()
        self.assertEqual(record.import_status, 'imported')
        self.assertEqual(record.telegram_message_id, '907_jawabu_0')

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_flags_duplicates_still_present_on_sheet(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        headers = list(JAWABU_FIELD_HEADERS.values())
        fake_sheet = FakeJawabuSheet(headers)
        row = ['' for _ in headers]
        row[headers.index('Duplicate Key')] = 'ID:1382654|PHONE:254720570031'
        fake_sheet.appended_rows.append(row)
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='908',
            sender='Importer',
        )

        self.assertEqual(result['imported'], 0)
        self.assertEqual(result['duplicate_review'], 1)
        status_index = headers.index('Import Status')
        self.assertEqual(fake_sheet.appended_rows[-1][status_index], 'Duplicate Review')


    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_consolidates_duplicate_media_for_same_entry(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        headers = list(JAWABU_FIELD_HEADERS.values())
        fake_sheet = FakeJawabuSheet(headers)
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031
3/16/26, 11:47 - Alex Kairu: IMG-20260316-WA0006.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='901',
            sender='Importer',
        )

        self.assertEqual(result['imported'], 1)
        self.assertEqual(result['duplicate_review'], 0)
        self.assertEqual(result['consolidated'], 1)
        self.assertEqual(len(fake_sheet.appended_rows), 1)
        media_index = headers.index('Media Filenames')
        self.assertIn('IMG-20260316-WA0005.jpg', fake_sheet.appended_rows[0][media_index])
        self.assertIn('IMG-20260316-WA0006.jpg', fake_sheet.appended_rows[0][media_index])
        record = JawabuVisitRecord.objects.get()
        self.assertEqual(record.import_status, 'imported')
        self.assertIn('IMG-20260316-WA0006.jpg', record.parsed_fields['media_filenames'])

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_consolidates_complementary_same_id_records(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        fake_sheet = FakeJawabuSheet(list(JAWABU_FIELD_HEADERS.values()))
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
3/16/26, 11:47 - Alex Kairu: IMG-20260316-WA0006.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='909',
            sender='Importer',
        )

        self.assertEqual(result['imported'], 1)
        self.assertEqual(result['duplicate_review'], 0)
        self.assertEqual(result['consolidated'], 1)
        record = JawabuVisitRecord.objects.get()
        self.assertEqual(record.primary_phone, '254720570031')

    @patch('core.services.jawabu.get_sheets_service')
    def test_jawabu_import_keeps_conflicting_same_id_records_for_review(self, mock_service):
        from core.services.jawabu import JAWABU_FIELD_HEADERS, process_jawabu_batch_export

        headers = list(JAWABU_FIELD_HEADERS.values())
        fake_sheet = FakeJawabuSheet(headers)
        mock_service.return_value = FakeJawabuService(fake_sheet)
        export = """3/16/26, 11:46 - Alex Kairu: IMG-20260316-WA0005.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0720570031
3/16/26, 11:47 - Alex Kairu: IMG-20260316-WA0006.jpg (file attached)
https://www.google.com/maps/search/?api=1&query=-0.55393,37.526935
State: Embu County
Country: Kenya
Mary Njeri njihia
1382654
0711111111"""

        result = process_jawabu_batch_export(
            group_config=self.group,
            export_text=export,
            telegram_message_id='910',
            sender='Importer',
        )

        self.assertEqual(result['imported'], 0)
        self.assertEqual(result['duplicate_review'], 2)
        self.assertEqual(result['consolidated'], 0)
        status_index = headers.index('Import Status')
        self.assertEqual(
            [row[status_index] for row in fake_sheet.appended_rows],
            ['Duplicate Review', 'Duplicate Review'],
        )


    def test_jawabu_workflow_farmers_csv_confirmed_master_mapping(self):
        from core.services.jawabu_master import import_jawabu_farmers_csv

        csv_file = StringIO(
            "Full Name,ID NUMBER,HBG Hub,Mobile,Phone,Actual Receipts,Sign Date,Sign Date,Created Date,HBG Contract Name\n"
            "David Mugambi [23215888],,Embu,+254721997481,+254704408281,5000,01/05/2026,24/06/2026,30/06/2026,HBGC-14560\n"
        )

        result = import_jawabu_farmers_csv(csv_file, source_name='farmers.csv')

        self.assertEqual(result.created, 1)
        self.assertEqual(result.review_needed, 0)
        farmer = JawabuFarmerMaster.objects.get()
        self.assertEqual(farmer.customer_name, 'DAVID MUGAMBI')
        self.assertEqual(farmer.national_id, '23215888')
        self.assertEqual(farmer.primary_phone, '254721997481')
        self.assertEqual(farmer.secondary_phone, '254704408281')
        self.assertEqual(farmer.county, 'EMBU')
        self.assertEqual(farmer.actual_receipts, '5000')
        self.assertEqual(farmer.sign_date, '24-June-2026')
        self.assertEqual(farmer.created_date, '')
        self.assertEqual(farmer.hbg_contract_name, '')
        self.assertEqual(farmer.external_id, '')
        self.assertEqual(farmer.status, 'active')
    def test_jawabu_farmup_review_batch_commits_edited_rows(self):
        from core.services.jawabu_master import (
            commit_farmup_review_batch,
            create_farmup_review_batch,
        )

        csv_text = (
            "Full Name,ID NUMBER,HBG Hub,Mobile,Phone,Actual Receipts,Sign Date,Sign Date,Created Date,HBG Contract Name\n"
            "David Mugambi [23215888],,Embu,+254721997481,+254704408281,5000,01/05/2026,24/06/2026,30/06/2026,HBGC-14560\n"
        )
        batch, stats = create_farmup_review_batch(
            group_id=self.group.group_id,
            telegram_message_id='farmup_1',
            sender='Uploader',
            source_filename='farmers.csv',
            csv_text=csv_text,
        )
        rows = list(batch.parsed_rows)
        rows[0]['County'] = 'Meru'
        rows[0]['approved'] = True

        result = commit_farmup_review_batch(batch, rows)

        self.assertTrue(result['success'])
        self.assertEqual(stats['total_rows'], 1)
        self.assertEqual(result['committed'], 1)
        self.assertEqual(JawabuFarmerUploadBatch.objects.get(pk=batch.pk).status, 'committed')
        farmer = JawabuFarmerMaster.objects.get()
        self.assertEqual(farmer.customer_name, 'DAVID MUGAMBI')
        self.assertEqual(farmer.national_id, '23215888')
        self.assertEqual(farmer.primary_phone, '254721997481')
        self.assertEqual(farmer.county, 'MERU')
        self.assertEqual(farmer.sign_date, '24-June-2026')
        self.assertEqual(farmer.actual_receipts, '5000')

    def test_farmup_master_sheet_writer_appends_system_columns_at_far_right(self):
        from core.services.jawabu_master import (
            MASTER_SYSTEM_HEADERS,
            ensure_master_system_headers,
            write_rows_to_master_sheet,
        )

        class FakeSpreadsheet:
            def __init__(self):
                self.requests = []

            def batch_update(self, payload):
                self.requests.append(payload)

        class FakeSheet:
            def __init__(self):
                self.id = 123
                self.spreadsheet = FakeSpreadsheet()
                self.col_count = 6
                self.values = [
                    [],
                    [],
                    ['No.', 'Customer Name', 'National ID', 'Primary Phone', 'County', 'Deposit Paid to HB'],
                    [],
                ]

            def row_values(self, row):
                return list(self.values[row - 1])

            def add_cols(self, count):
                self.col_count += count

            def update_cell(self, row, col, value):
                while len(self.values) < row:
                    self.values.append([])
                while len(self.values[row - 1]) < col:
                    self.values[row - 1].append('')
                self.values[row - 1][col - 1] = value

            def get_all_values(self):
                return [list(row) for row in self.values]

            def update(self, range_name, rows, value_input_option=None):
                row_number = int(range_name.split(':', 1)[0][1:])
                while len(self.values) < row_number:
                    self.values.append([])
                self.values[row_number - 1] = list(rows[0])

        batch = JawabuFarmerUploadBatch.objects.create(
            group_id=self.group.group_id,
            sender='Reviewer',
            source_filename='farmers.csv',
            total_rows=1,
        )
        sheet = FakeSheet()
        headers = ensure_master_system_headers(sheet, 3)
        result = write_rows_to_master_sheet(
            sheet=sheet,
            headers=headers,
            data_start_row=5,
            batch=batch,
            cleaned_rows=[{
                'customer_name': 'DAVID MUGAMBI',
                'national_id': '23215888',
                'primary_phone': '254721997481',
                'county': 'EMBU',
                'actual_receipts': '5000',
                'duplicate_key': '23215888|254721997481',
                'source_row_number': 2,
            }],
        )

        self.assertEqual(result['created'], 1)
        self.assertEqual(result['errors'], [])
        self.assertEqual(headers[-len(MASTER_SYSTEM_HEADERS):], MASTER_SYSTEM_HEADERS)
        self.assertEqual(sheet.values[4][1], 'DAVID MUGAMBI')
        self.assertEqual(sheet.values[4][5], '5000')
        self.assertEqual(sheet.values[4][headers.index('Import Status')], 'created')
        self.assertTrue(sheet.spreadsheet.requests)

    def test_farmup_master_sheet_writer_flags_conflicts_without_overwriting(self):
        from core.services.jawabu_master import ensure_master_system_headers, write_rows_to_master_sheet

        class FakeSpreadsheet:
            def batch_update(self, payload):
                pass

        class FakeSheet:
            id = 123
            spreadsheet = FakeSpreadsheet()
            col_count = 10

            def __init__(self):
                self.values = [
                    [],
                    [],
                    ['No.', 'Customer Name', 'National ID', 'Primary Phone', 'County', 'Duplicate Key'],
                    [],
                    ['1', 'DAVID MUGAMBI', '23215888', '254721997481', 'EMBU', '23215888|254721997481'],
                ]

            def row_values(self, row):
                return list(self.values[row - 1])

            def add_cols(self, count):
                self.col_count += count

            def update_cell(self, row, col, value):
                while len(self.values[row - 1]) < col:
                    self.values[row - 1].append('')
                self.values[row - 1][col - 1] = value

            def get_all_values(self):
                return [list(row) for row in self.values]

            def update(self, range_name, rows, value_input_option=None):
                row_number = int(range_name.split(':', 1)[0][1:])
                self.values[row_number - 1] = list(rows[0])

        batch = JawabuFarmerUploadBatch.objects.create(
            group_id=self.group.group_id,
            sender='Reviewer',
            source_filename='farmers.csv',
            total_rows=1,
        )
        sheet = FakeSheet()
        headers = ensure_master_system_headers(sheet, 3)
        result = write_rows_to_master_sheet(
            sheet=sheet,
            headers=headers,
            data_start_row=5,
            batch=batch,
            cleaned_rows=[{
                'customer_name': 'DAVID MUGAMBI',
                'national_id': '23215888',
                'primary_phone': '254721997481',
                'county': 'MERU',
                'duplicate_key': '23215888|254721997481',
                'source_row_number': 2,
            }],
        )

        self.assertEqual(result['updated'], 1)
        self.assertEqual(result['conflicts'], 1)
        self.assertEqual(sheet.values[4][4], 'EMBU')
        self.assertEqual(sheet.values[4][headers.index('Import Status')], 'updated_with_conflict')
        self.assertIn('County', sheet.values[4][headers.index('Review Notes')])
class FcaWorkflowServiceTest(TestCase):
    """Tests for FCA Excel batch imports."""

    def setUp(self):
        self.group = GroupSheetConfiguration.objects.create(
            group_id='-100order',
            display_name='Order Approval',
            enabled=True,
            sheet_id='sheet_orders',
            sheet_name='Orders',
            workflow={
                'type': 'order_approval',
                'header_row': 2,
                'create_sheet_name': 'Orders',
                'record_id_prefix': 'JBL',
            },
        )

    def fca_workbook_bytes(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'FCA'
        sheet['C2'] = 'FIELD CONTROL VISIT APPROVAL FORM'
        sheet['H2'] = 'Date: 23rd June 2026'
        sheet['H4'] = 'HUB... Kiambu'
        sheet.append([])
        sheet.append([])
        sheet.append([
            '', 'CUSTOMER NAME', 'CONTACTS', 'LOCATION', 'HB STAFF',
            'DEPOSIT', 'APPROVAL BASIS', '',
        ])
        sheet.append([
            1, 'Samuel Ndungu', '0724733556/0711111111', 'Ngenda',
            'Nathan', '5000 HB', '', 'opted cash',
        ])
        sheet.append([
            2, 'Jane Wanjiku', '724000111', 'Kieni',
            'Mary', '5000 HB', '', 'Approved dairy farmer',
        ])
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def test_extract_fca_workbook_records_gets_date_comment_and_cash_decision(self):
        from core.services.fca import extract_fca_workbook_records

        records = extract_fca_workbook_records(
            'FCA Kiambu 23rd June 2026 after visit.xlsx',
            self.fca_workbook_bytes(),
        )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0].fields['fca_visit_date'], '23-Jun-2026')
        self.assertEqual(records[0].fields['customer_name'], 'SAMUEL NDUNGU')
        self.assertEqual(records[0].fields['primary_phone'], '254724733556')
        self.assertEqual(records[0].fields['secondary_phone'], '254711111111')
        self.assertEqual(records[0].fields['fca_comment'], 'opted cash')
        self.assertEqual(records[0].fields['fca_decision'], 'Cash')
        self.assertEqual(records[1].fields['fca_decision'], 'Approved')

    @patch('core.services.fca.get_sheets_service')
    def test_process_fca_batch_files_appends_rows_to_order_sheet(self, mock_service):
        from core.services.fca import FCA_FIELD_HEADERS, process_fca_batch_files

        headers = list(FCA_FIELD_HEADERS.values())
        fake_sheet = FakeFcaSheet(headers)
        mock_service.return_value = FakeJawabuService(fake_sheet)

        result = process_fca_batch_files(
            group_config=self.group,
            files=[('fca.xlsx', self.fca_workbook_bytes())],
            telegram_message_id='501',
            sender='Importer',
        )

        self.assertEqual(result['status'], 'fca_batch_processed')
        self.assertEqual(result['processed'], 2)
        self.assertEqual(result['imported'], 2)
        self.assertEqual(result['cash'], 1)
        self.assertEqual(FcaImportRecord.objects.count(), 2)
        self.assertEqual(fake_sheet.append_rows_calls, 1)
        decision_index = headers.index('FCA DECISION')
        self.assertEqual(fake_sheet.appended_rows[0][decision_index], 'Cash')
        record_id_index = headers.index('ORDER RECORD ID')
        self.assertEqual(fake_sheet.appended_rows[0][record_id_index], 'JBL-1')
        self.assertEqual(fake_sheet.appended_rows[1][record_id_index], 'JBL-2')


class FakeFcaSheet:
    def __init__(self, headers):
        self.headers = headers
        self.appended_rows = []
        self.row_values_calls = 0
        self.get_all_values_calls = 0
        self.append_rows_calls = 0

    def row_values(self, row_number):
        self.row_values_calls += 1
        return self.headers if row_number == 2 else []

    def get_all_values(self):
        self.get_all_values_calls += 1
        return [[]] + [self.headers] + self.appended_rows

    def append_rows(self, rows, value_input_option='USER_ENTERED'):
        self.append_rows_calls += 1
        first_row = len(self.appended_rows) + 3
        self.appended_rows.extend(rows)
        last_row = first_row + len(rows) - 1
        return {'updates': {'updatedRange': f"'Orders'!A{first_row}:P{last_row}"}}



class FakeJawabuSheet:
    def __init__(self, headers):
        self.headers = headers
        self.appended_rows = []
        self.row_values_calls = 0
        self.get_all_values_calls = 0
        self.append_rows_calls = 0

    def row_values(self, row_number):
        self.row_values_calls += 1
        return self.headers if row_number == 1 else []

    def get_all_values(self):
        self.get_all_values_calls += 1
        return [self.headers] + self.appended_rows

    def append_rows(self, rows, value_input_option='USER_ENTERED'):
        self.append_rows_calls += 1
        first_row = len(self.appended_rows) + 2
        self.appended_rows.extend(rows)
        last_row = first_row + len(rows) - 1
        return {'updates': {'updatedRange': f"'Jawabu Visits'!A{first_row}:V{last_row}"}}

    def append_row(self, row, value_input_option='USER_ENTERED'):
        first_row = len(self.appended_rows) + 2
        self.appended_rows.append(row)
        return {'updates': {'updatedRange': f"'Jawabu Visits'!A{first_row}:V{first_row}"}}


class FakeJawabuService:
    def __init__(self, sheet):
        self._sheet = sheet

    def is_available(self):
        return True


class ParserServiceContinuedTest(TestCase):
    def test_intent_detection_purchase(self):
        """Test intent detection for purchase messages."""
        from core.services.parser import detect_message_intent, MessageIntent
        
        result = detect_message_intent("John bought 3 bags maize @ 100")
        self.assertEqual(result, MessageIntent.PURCHASE)
        
        result = detect_message_intent("Mary purchased 5 chicken")
        self.assertEqual(result, MessageIntent.PURCHASE)
    
    def test_intent_detection_payment(self):
        """Test intent detection for payment messages."""
        from core.services.parser import detect_message_intent, MessageIntent
        
        result = detect_message_intent("John paid 200 for 4 milk")
        self.assertEqual(result, MessageIntent.PAYMENT)
        
        result = detect_message_intent("Mary sent 500")
        self.assertEqual(result, MessageIntent.PAYMENT)
    
    def test_intent_detection_location(self):
        """Test intent detection for location messages."""
        from core.services.parser import detect_message_intent, MessageIntent
        
        result = detect_message_intent("📍 https://maps.app.goo.gl/abc123 Location update")
        self.assertEqual(result, MessageIntent.LOCATION)
        
        result = detect_message_intent("Location: https://goo.gl/maps/xyz")
        self.assertEqual(result, MessageIntent.LOCATION)
    
    def test_intent_detection_status(self):
        """Test intent detection for status messages."""
        from core.services.parser import detect_message_intent, MessageIntent
        
        result = detect_message_intent("Order arrived at destination")
        self.assertEqual(result, MessageIntent.STATUS)
        
        result = detect_message_intent("Product is ready for pickup")
        self.assertEqual(result, MessageIntent.STATUS)
    
    def test_intent_detection_complaint(self):
        """Test intent detection for complaint messages."""
        from core.services.parser import detect_message_intent, MessageIntent
        
        result = detect_message_intent("*CUSTOMER COMPLAIN* NAME: John Doe TEL: 0712345678 ID: A12345")
        self.assertEqual(result, MessageIntent.COMPLAINT)
        
        result = detect_message_intent("CUSTOMER COMPLAIN NATURE OF THE PROBLEM: No gas supply")
        self.assertEqual(result, MessageIntent.COMPLAINT)
    
    def test_parse_complaint_transaction(self):
        """Test parsing structured complaint/case report."""
        from core.services.parser import MessageIntent

        content = (
            "*CUSTOMER COMPLAIN*\n"
            "NAME: John Doe\n"
            "TEL: 0712345678\n"
            "ID: A12345\n"
            "NATURE OF THE PROBLEM: No gas supply at home\n"
            "Please assist urgently."
        )

        result = parse_message(content, sender="Agent")
        self.assertEqual(result.intent, MessageIntent.COMPLAINT)
        self.assertEqual(result.customer_name, 'John Doe')
        self.assertEqual(result.customer_phone, '254712345678')
        self.assertEqual(result.customer_id, 'A12345')
        self.assertIn('No gas supply', result.problem_description)
        self.assertGreater(result.confidence, 0.0)

    def test_parse_complaint_fields_on_one_line_have_clear_boundaries(self):
        """Adjacent labels should not be swallowed into earlier fields."""
        from core.services.parser import MessageIntent

        content = (
            "CUSTOMER COMPLAIN NAME: John Doe TEL: 0712345678 "
            "ID: A12345 COUNTY: Muranga NATURE OF COMPLAIN: No gas supply"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.intent, MessageIntent.COMPLAINT)
        self.assertEqual(result.customer_name, 'John Doe')
        self.assertEqual(result.customer_phone, '254712345678')
        self.assertEqual(result.customer_id, 'A12345')
        self.assertEqual(result.branch_region, 'MURANGA')
        self.assertEqual(result.problem_description, 'No gas supply')
        self.assertEqual(result.confidence, 1.0)

    def test_parse_complaint_county_label_can_appear_after_description(self):
        """County should map to the Branch / Region sheet field from any label position."""
        content = (
            "CUSTOMER COMPLAIN NAME: John Doe TEL: 0712345678 "
            "ID: A12345 NATURE OF COMPLAIN: No gas supply COUNTY: Kiambu"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.branch_region, 'KIAMBU')
        self.assertEqual(result.problem_description, 'No gas supply')
        self.assertEqual(result.confidence, 1.0)

    def test_parse_complaint_without_customer_id_is_partial(self):
        """A complaint with phone but no ID is importable for manual review."""
        content = (
            "CUSTOMER COMPLAIN NAME: John Doe TEL: 0712345678 "
            "COUNTY: Muranga NATURE OF COMPLAIN: No gas supply"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.customer_phone, '254712345678')
        self.assertEqual(result.customer_id, '')
        self.assertLess(result.confidence, 1.0)
        self.assertIn(
            'Customer ID / Account missing; saved with Status: Review Needed.',
            result.warnings,
        )
        self.assertNotIn(
            'Missing required complaint field(s): Customer ID / Account',
            result.warnings,
        )

    def test_parse_complaint_without_phone_is_partial(self):
        """A complaint with ID but no phone is importable for manual review."""
        content = (
            "CUSTOMER COMPLAIN NAME: John Doe ID: A12345 COUNTY: Muranga "
            "NATURE OF COMPLAIN: No gas supply"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.customer_id, 'A12345')
        self.assertEqual(result.customer_phone, '')
        self.assertLess(result.confidence, 1.0)
        self.assertIn(
            'Phone Number missing; saved with Status: Review Needed.',
            result.warnings,
        )
        self.assertNotIn(
            'Missing required complaint field(s): Phone Number',
            result.warnings,
        )

    def test_parse_complaint_without_any_identifier_is_partial(self):
        """A complaint needs at least phone or ID to identify the customer."""
        content = (
            "CUSTOMER COMPLAIN NAME: John Doe COUNTY: Muranga "
            "NATURE OF COMPLAIN: No gas supply"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.customer_phone, '')
        self.assertEqual(result.customer_id, '')
        self.assertIn(
            'Missing required complaint field(s): Phone Number or Customer ID / Account',
            result.warnings,
        )

    def test_parse_complaint_without_county_is_complete(self):
        """County is optional, but still extracted into Branch / Region when present."""
        content = (
            "CUSTOMER COMPLAIN NAME: John Doe TEL: 0712345678 "
            "ID: A12345 NATURE OF COMPLAIN: No gas supply"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.branch_region, '')
        self.assertEqual(result.customer_phone, '254712345678')
        self.assertEqual(result.customer_id, 'A12345')
        self.assertNotIn(
            'Missing required complaint field(s): County (Branch / Region)',
            result.warnings,
        )

    def test_parse_complaint_fields_without_label_separators(self):
        """Labels without punctuation should still define clean field bounds."""
        from core.services.parser import MessageIntent

        content = (
            "CUSTOMER COMPLAIN NAME John Doe TEL NO 0712345678 "
            "ID A12345 NATURE OF COMPLAIN No gas supply at home"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.intent, MessageIntent.COMPLAINT)
        self.assertEqual(result.customer_name, 'John Doe')
        self.assertEqual(result.customer_phone, '254712345678')
        self.assertEqual(result.customer_id, 'A12345')
        self.assertEqual(result.problem_description, 'No gas supply at home')
        self.assertNotIn('NAME', result.customer_name.upper())
        self.assertNotIn('CUSTOMER COMPLAIN', result.problem_description.upper())

    def test_parse_multiline_complaint_fields_without_separators(self):
        """Line breaks should work as separators, but not be required."""
        content = (
            "CUSTOMER COMPLAIN\n"
            "NAME Jane Doe\n"
            "TEL NO 0712345678\n"
            "ID A12345\n"
            "NATURE OF COMPLAIN No gas supply at home"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.customer_name, 'Jane Doe')
        self.assertEqual(result.customer_phone, '254712345678')
        self.assertEqual(result.customer_id, 'A12345')
        self.assertEqual(result.problem_description, 'No gas supply at home')

    def test_parse_complaint_description_does_not_swallow_following_labels(self):
        """Problem text should stop before later structured fields."""
        content = (
            "CUSTOMER COMPLAIN NATURE OF COMPLAINT: Burner not working "
            "NAME: Alice Smith TEL: +254712345678 ID: CUST_100"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.customer_name, 'Alice Smith')
        self.assertEqual(result.customer_phone, '254712345678')
        self.assertEqual(result.customer_id, 'CUST_100')
        self.assertEqual(result.problem_description, 'Burner not working')

    def test_parse_unlabeled_comma_name_and_numeric_id(self):
        """Unlabeled case blocks should allow comma-form names and numeric IDs."""
        content = (
            "CUSTOMER COMPLAIN\n"
            "NYAMU , ROSE RUGURU\n"
            "0721552446\n"
            "11598558\n"
            "Less cooking hrs than expected"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.customer_name, 'NYAMU, ROSE RUGURU')
        self.assertEqual(result.customer_phone, '254721552446')
        self.assertEqual(result.customer_id, '11598558')
        self.assertEqual(result.problem_description, 'Less cooking hrs than expected')

    def test_parse_of_phone_sentence_name(self):
        """A name before 'of phone:' should be captured as the customer name."""
        from core.services.parser import MessageIntent

        content = (
            "Joseph Mbaabu of phone:0714953414 is requesting for an agronomy training"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.intent, MessageIntent.COMPLAINT)
        self.assertEqual(result.customer_name, 'Joseph Mbaabu')
        self.assertEqual(result.customer_phone, '254714953414')
        self.assertEqual(
            result.problem_description,
            'is requesting for an agronomy training',
        )

    def test_parse_subject_sentence_name_and_normalises_phone(self):
        """Sentence-subject names are used only inside complaint parsing."""
        from core.services.parser import MessageIntent

        content = (
            "Francis Kaihura Kuria is requesting installation\n"
            "254797963674"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.intent, MessageIntent.COMPLAINT)
        self.assertEqual(result.customer_name, 'Francis Kaihura Kuria')
        self.assertEqual(result.customer_phone, '254797963674')
        self.assertEqual(result.problem_description, 'is requesting installation')

    def test_parse_status_description_does_not_return_full_message(self):
        """If there is no problem keyword, keep only the non-identity remainder."""
        content = (
            "CUSTOMER COMPLAIN NAME Doreen Gaceri\n"
            "Phone:0718077338\n"
            "ready for installation"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.customer_name, 'Doreen Gaceri')
        self.assertEqual(result.customer_phone, '254718077338')
        self.assertEqual(result.problem_description, 'ready for installation')
        self.assertNotIn('CUSTOMER COMPLAIN', result.problem_description)
        self.assertNotIn('Phone', result.problem_description)

    def test_parse_unlabeled_bot_tagged_relocation_case(self):
        """Bot-tagged unlabeled blocks should parse identity plus request text."""
        from core.services.parser import MessageIntent

        content = (
            "@hb_biogas_cases_bot Henry  mwenda\n"
            "24289449\n"
            "0720809218/0726011961\n"
            "\n"
            "Requesting for a jiko relocation"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.intent, MessageIntent.COMPLAINT)
        self.assertEqual(result.customer_name, 'Henry mwenda')
        self.assertEqual(result.customer_id, '24289449')
        self.assertEqual(result.customer_phone, '254720809218')
        self.assertEqual(result.problem_description, 'Requesting for a jiko relocation')

    def test_parse_complaint_description_excludes_trailing_awareness_mentions(self):
        """Trailing @mentions are notification tags, not complaint text."""
        content = (
            "CUSTOMER COMPLAIN\n"
            "Henry mwenda\n"
            "24289449\n"
            "0720809218/0726011961\n"
            "\n"
            "Requesting for a jiko relocation @area_manager"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.customer_name, 'Henry mwenda')
        self.assertEqual(result.problem_description, 'Requesting for a jiko relocation')
        self.assertNotIn('@area_manager', result.problem_description)

    def test_parse_complaint_description_excludes_spaced_display_mention(self):
        """Display-name mentions such as '@~Eunny K' should be stripped fully."""
        content = (
            "@hb_biogas_cases_bot Henry  mwenda\n"
            "24289449\n"
            "0720809218/0726011961\n"
            "\n"
            "Requesting for a jiko relocation @~Eunny K"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.customer_name, 'Henry mwenda')
        self.assertEqual(result.problem_description, 'Requesting for a jiko relocation')
        self.assertNotIn('Eunny', result.problem_description)
        self.assertNotIn('K', result.problem_description)

    def test_parse_complaint_description_excludes_final_mention_line(self):
        """A final line containing only mentions should not enter the description."""
        content = (
            "CUSTOMER COMPLAIN NAME Jane Doe TEL 0712345678 "
            "NATURE OF COMPLAIN No gas supply at home\n"
            "@technician @supervisor"
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.problem_description, 'No gas supply at home')
        self.assertNotIn('@technician', result.problem_description)

    def test_parse_unlabeled_complaint_transaction(self):
        """Plain complaint blocks should infer identifiers and description."""
        from core.services.parser import MessageIntent

        content = (
            "CUSTOMER COMPLAIN\n"
            "John Doe\n"
            "0712345678\n"
            "A12345\n"
            "County: Embu\n"
            "No gas supply at home. Please assist urgently."
        )

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.intent, MessageIntent.COMPLAINT)
        self.assertEqual(result.customer_name, 'John Doe')
        self.assertEqual(result.customer_phone, '254712345678')
        self.assertEqual(result.customer_id, 'A12345')
        self.assertEqual(result.branch_region, 'EMBU')
        self.assertEqual(
            result.problem_description,
            'No gas supply at home. Please assist urgently.',
        )
        self.assertEqual(result.confidence, 1.0)

    def test_parse_unlabeled_complaint_without_header(self):
        """Phone plus complaint language should detect an unlabeled complaint."""
        from core.services.parser import MessageIntent

        content = "Jane Doe 0798765432 B456 in Muranga biogas is leaking near the valve"

        result = parse_message(content, sender="Agent")

        self.assertEqual(result.intent, MessageIntent.COMPLAINT)
        self.assertEqual(result.customer_name, 'Jane Doe')
        self.assertEqual(result.customer_phone, '254798765432')
        self.assertEqual(result.customer_id, 'B456')
        self.assertEqual(result.branch_region, 'MURANGA')
        self.assertEqual(result.problem_description, 'biogas is leaking near the valve')
        self.assertEqual(result.confidence, 1.0)
    
    def test_parse_with_intent_sale(self):
        """Test parsing with SALE intent."""
        from core.services.parser import MessageIntent
        
        result = parse_message("Sold 3 bread 50 each to John", sender="Seller")
        
        self.assertEqual(result.intent, MessageIntent.SALE)
        self.assertEqual(result.quantity, Decimal('3'))
        self.assertEqual(result.item, 'bread')
        self.assertEqual(result.price, Decimal('50'))
        self.assertGreater(result.confidence, 0.5)
    
    def test_parse_with_intent_purchase(self):
        """Test parsing with PURCHASE intent."""
        from core.services.parser import MessageIntent
        
        result = parse_message("John bought 3 bags maize @ 100", sender="John")
        
        self.assertEqual(result.intent, MessageIntent.PURCHASE)
        self.assertEqual(result.quantity, Decimal('3'))
        self.assertEqual(result.item, 'bags maize')
        self.assertEqual(result.price, Decimal('100'))
    
    def test_parse_with_intent_payment(self):
        """Test parsing with PAYMENT intent."""
        from core.services.parser import MessageIntent
        
        result = parse_message("John paid 200 for 4 milk", sender="John")
        
        self.assertEqual(result.intent, MessageIntent.PAYMENT)
        self.assertEqual(result.price, Decimal('200'))
        self.assertEqual(result.quantity, Decimal('4'))
        self.assertEqual(result.item, 'milk')
    
    def test_parse_with_intent_location(self):
        """Test parsing with SALE intent (transaction with location info)."""
        from core.services.parser import MessageIntent
        
        content = "📍 https://maps.app.goo.gl/abc123 Sold 2 bags maize 150 each"
        result = parse_message(content, sender="Seller")
        
        self.assertEqual(result.intent, MessageIntent.SALE)
        self.assertIn('https://maps.app.goo.gl/abc123', result.gps_link)
        self.assertEqual(result.quantity, Decimal('2'))
        self.assertEqual(result.item, 'bags maize')
        self.assertEqual(result.price, Decimal('150'))
    
    def test_parse_diverse_formats(self):
        """Test parsing various real-world WhatsApp message formats."""
        test_cases = [
            # Standard formats
            ("Sold 5 maize 20 each to Peter", {'qty': '5', 'item': 'maize', 'price': '20'}),
            ("John paid 1000 for 10 bags", {'price': '1000', 'qty': '10', 'item': 'bags'}),
            ("Bought 2 chicken @ 300 each", {'qty': '2', 'item': 'chicken', 'price': '300'}),
            
            # With currency
            ("Sold 3 bread KSH 150 each", {'qty': '3', 'item': 'bread', 'price': '150'}),
            ("Paid 2000 total for fertilizer", {'price': '2000', 'item': 'fertilizer'}),
            
            # With GPS
            ("📍 https://maps.app.goo.gl/xyz Sold 4 milk 50 each", {'gps': True, 'qty': '4', 'item': 'milk', 'price': '50'}),
            
            # Different verbs
            ("Delivered 6 eggs 10 each", {'qty': '6', 'item': 'eggs', 'price': '10'}),
            ("Purchased 1 bag cement @ 500", {'qty': '1', 'item': 'bag cement', 'price': '500'}),
            ("Transferred 750 for 3 tomatoes", {'price': '750', 'qty': '3', 'item': 'tomatoes'}),
            
            # Edge cases
            ("Sold bread", {'item': 'bread'}),  # Missing quantity/price
            ("Paid 100", {'price': '100'}),  # Missing item/quantity
            ("Bought maize", {'item': 'maize'}),  # Missing quantity/price
        ]
        
        for content, expected in test_cases:
            with self.subTest(content=content):
                result = parse_message(content)
                
                if 'qty' in expected:
                    self.assertEqual(result.quantity, Decimal(expected['qty']))
                if 'item' in expected:
                    self.assertEqual(result.item, expected['item'])
                if 'price' in expected:
                    self.assertEqual(result.price, Decimal(expected['price']))
                if 'gps' in expected and expected['gps']:
                    self.assertTrue(result.gps_link)
                
                # Should have reasonable confidence for valid transactions
                if any(key in expected for key in ['qty', 'item', 'price']):
                    self.assertGreater(result.confidence, 0.0)
    
    def test_parse_timestamp_formats(self):
        """Test parsing various timestamp formats."""
        from datetime import datetime
        
        test_cases = [
            ("[14/03/2026, 10:30:15] Sold 3 bread", "%d/%m/%Y %H:%M:%S"),
            ("14/03/2026 10:30 Sold maize", "%d/%m/%Y %H:%M"),
            ("[15-03-2026, 14:45:30] Paid 200", "%d-%m-%Y %H:%M:%S"),
        ]
        
        for content, fmt in test_cases:
            with self.subTest(content=content):
                result = parse_message(content)
                self.assertIsNotNone(result.timestamp)
                # Extract the expected time string from content
                import re
                match = re.search(r'\[?(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})[\s,]+(\d{1,2}:\d{2}(?::\d{2})?)\]?', content)
                if match:
                    expected_time_str = f"{match.group(1)} {match.group(2)}"
                    expected = datetime.strptime(expected_time_str, fmt)
                    self.assertEqual(result.timestamp.replace(tzinfo=None), expected)
    
    def test_parse_sender_extraction(self):
        """Test sender name extraction from various patterns."""
        test_cases = [
            ("Sold 3 bread to John", "John"),
            ("John paid 200 for milk", "John"),
            ("Mary bought 5 eggs", "Mary"),
            ("Delivered goods to Customer Peter", "Customer Peter"),
        ]
        
        for content, expected_sender in test_cases:
            with self.subTest(content=content):
                result = parse_message(content)
                self.assertEqual(result.sender, expected_sender)


class StorageServiceTest(TestCase):
    """Test the storage service."""
    
    @patch('core.services.storage.append_parsed_message_to_sheet')
    def test_process_and_store_new_message(self, mock_sheet):
        """Test processing and storing a new message."""
        mock_sheet.return_value = True
        
        parsed = process_and_store_message(
            telegram_message_id='test_001',
            content='Sold 3 bread 50 each to John',
            sender='Seller',
            received_at=timezone.now(),
        )
        
        self.assertIsNotNone(parsed)
        self.assertEqual(parsed.item, 'bread')
        self.assertEqual(parsed.quantity, Decimal('3'))
        
        # Verify raw message stored
        self.assertTrue(
            RawMessage.objects.filter(telegram_message_id='test_001').exists()
        )
        
        # Verify processed message stored
        self.assertTrue(
            ProcessedMessage.objects.filter(
                raw_message__telegram_message_id='test_001'
            ).exists()
        )

    @patch('core.services.sheets.append_parsed_message_to_sheet')
    def test_process_and_store_persists_sheet_identity(self, mock_sheet):
        """Stored cases should retain the sheet they were routed to."""
        mock_sheet.return_value = True

        parsed = process_and_store_message(
            telegram_message_id='test_sheet_identity',
            content='Sold 3 bread 50 each to John',
            sender='Seller',
            received_at=timezone.now(),
            group_id='-100123',
            sheet_id='sheet_123',
            sheet_name='Cases',
        )

        self.assertIsNotNone(parsed)
        self.assertEqual(parsed.group_id, '-100123')
        self.assertEqual(parsed.sheet_id, 'sheet_123')
        self.assertEqual(parsed.sheet_name, 'Cases')
        self.assertEqual(mock_sheet.call_args.kwargs['sheet_id'], 'sheet_123')
        self.assertEqual(mock_sheet.call_args.kwargs['sheet_name'], 'Cases')

    @patch('core.services.sheets.append_parsed_message_to_sheet')
    def test_process_and_store_can_defer_sheet_sync(self, mock_sheet):
        """Batch imports should store locally without one sheet write per row."""
        parsed = process_and_store_message(
            telegram_message_id='test_defer_sheet_sync',
            content=(
                'CUSTOMER COMPLAINT\n'
                'NAME: Jane Doe\n'
                'TEL: 0712345678\n'
                'ID: A12345\n'
                'COUNTY: KISUMU\n'
                'NATURE OF THE PROBLEM: No gas supply'
            ),
            sender='Agent',
            received_at=timezone.now(),
            group_id='-100123',
            sheet_id='sheet_123',
            sheet_name='Cases',
            defer_sheet_sync=True,
        )

        self.assertIsNotNone(parsed)
        self.assertEqual(parsed.group_id, '-100123')
        self.assertFalse(parsed.synced_to_sheets)
        self.assertEqual(getattr(parsed, '_processing_status'), 'success')
        mock_sheet.assert_not_called()

    @patch('core.services.sheets.append_parsed_message_to_sheet')
    def test_process_and_store_missing_id_sets_review_needed_status(self, mock_sheet):
        """Phone-only complaints should be saved for manual ID review."""
        parsed = process_and_store_message(
            telegram_message_id='missing_id_review',
            content=(
                'CUSTOMER COMPLAINT\n'
                'NAME: Jane Doe\n'
                'TEL: 0712345678\n'
                'NATURE OF THE PROBLEM: No gas supply'
            ),
            sender='Agent',
            received_at=timezone.now(),
            group_id='-100123',
            sheet_id='sheet_123',
            sheet_name='Cases',
            defer_sheet_sync=True,
        )

        self.assertEqual(parsed.customer_phone, '254712345678')
        self.assertEqual(parsed.customer_id, '')
        self.assertEqual(parsed.complaint_status, 'Review Needed')
        self.assertIn('Customer ID / Account missing', getattr(parsed, '_processing_warnings')[0])
        mock_sheet.assert_not_called()

    @patch('core.services.sheets.append_parsed_message_to_sheet')
    def test_process_and_store_missing_phone_sets_review_needed_status(self, mock_sheet):
        """ID-only complaints should be saved for manual phone review."""
        parsed = process_and_store_message(
            telegram_message_id='missing_phone_review',
            content=(
                'CUSTOMER COMPLAINT\n'
                'NAME: Jane Doe\n'
                'ID: A12345\n'
                'NATURE OF THE PROBLEM: No gas supply'
            ),
            sender='Agent',
            received_at=timezone.now(),
            group_id='-100123',
            sheet_id='sheet_123',
            sheet_name='Cases',
            defer_sheet_sync=True,
        )

        self.assertEqual(parsed.customer_phone, '')
        self.assertEqual(parsed.customer_id, 'A12345')
        self.assertEqual(parsed.complaint_status, 'Review Needed')
        self.assertIn('Phone Number missing', getattr(parsed, '_processing_warnings')[0])
        mock_sheet.assert_not_called()

    @patch('core.services.sheets.append_parsed_message_to_sheet')
    def test_process_and_store_rejects_incomplete_complaint_atomically(self, mock_sheet):
        """Incomplete complaint intake should not leave raw, processed, or parsed rows."""
        with self.assertRaises(MessageRejectedError) as context:
            process_and_store_message(
                telegram_message_id='reject_missing_identifier',
                content=(
                    'CUSTOMER COMPLAINT\n'
                    'NAME: Jane Doe\n'
                    'NATURE OF THE PROBLEM: No gas supply'
                ),
                sender='Agent',
                received_at=timezone.now(),
                group_id='-100123',
                sheet_id='sheet_123',
                sheet_name='Complaints',
            )

        self.assertIn(
            'Phone Number or Customer ID / Account',
            context.exception.missing_fields,
        )
        self.assertEqual(RawMessage.objects.count(), 0)
        self.assertEqual(ProcessedMessage.objects.count(), 0)
        self.assertEqual(ParsedMessage.objects.count(), 0)
        mock_sheet.assert_not_called()
    
    @patch('core.services.storage.append_parsed_message_to_sheet')
    def test_process_duplicate_message(self, mock_sheet):
        """Test that duplicate messages return None."""
        mock_sheet.return_value = True
        
        # Store first message
        result1 = process_and_store_message(
            telegram_message_id='dup_001',
            content='Sold 3 bread 50 each',
            sender='Seller',
            received_at=timezone.now(),
        )
        self.assertIsNotNone(result1)
        
        # Try to store duplicate
        result2 = process_and_store_message(
            telegram_message_id='dup_002',
            content='Sold 3 bread 50 each',
            sender='Seller',
            received_at=timezone.now(),
        )
        self.assertIsNone(result2)

    @patch('core.services.storage.append_parsed_message_to_sheet')
    def test_process_and_store_message_sheet_sync_failure(self, mock_sheet):
        """Sheet sync failures should mark message processing as partial."""
        mock_sheet.return_value = False

        parsed = process_and_store_message(
            telegram_message_id='test_sync_fail',
            content='Sold 3 bread 50 each to John',
            sender='Seller',
            received_at=timezone.now(),
        )

        self.assertIsNotNone(parsed)
        self.assertEqual(getattr(parsed, '_processing_status', None), 'partial')
        self.assertEqual(getattr(parsed, '_processing_error', None), 'Google Sheets sync failed')

    @patch('core.services.sheets.get_sheets_service')
    def test_append_parsed_message_increments_attempts_on_failure(self, mock_service):
        """A failed sheet sync should increment sync_attempts and record the error."""
        raw_message = RawMessage.objects.create(
            telegram_message_id='fail_sheet',
            sender='Test',
            content='Test content',
        )
        msg_hash = generate_message_hash('Test', 'Test content')
        processed = ProcessedMessage.objects.create(
            message_hash=msg_hash,
            raw_message=raw_message,
        )
        parsed = ParsedMessage.objects.create(
            processed_message=processed,
            message_id='MSG_FAIL_1',
            timestamp=timezone.now(),
            sender='John',
            raw_message='Sold 3 bread 50 each',
            item='bread',
            quantity=Decimal('3'),
            price=Decimal('50'),
            gps_link='',
            image_flag=False,
            source='whatsapp_telegram',
        )

        mock_service.return_value = MagicMock()
        mock_service.return_value.append_row.side_effect = Exception('Sheet unavailable')

        from core.services.sheets import append_parsed_message_to_sheet
        success = append_parsed_message_to_sheet(parsed)

        self.assertFalse(success)
        parsed.refresh_from_db()
        self.assertEqual(parsed.sync_attempts, 1)
        self.assertIn('Sheet unavailable', parsed.last_sync_error)

    def test_bulk_resync_to_sheets_skips_max_attempts(self):
        """Resync should ignore messages that have reached max retry attempts."""
        raw_message = RawMessage.objects.create(
            telegram_message_id='resync_001',
            sender='Test',
            content='Test content',
        )
        msg_hash = generate_message_hash('Test', 'Test content')
        processed = ProcessedMessage.objects.create(
            message_hash=msg_hash,
            raw_message=raw_message,
        )
        ParsedMessage.objects.create(
            processed_message=processed,
            message_id='MSG_RESYNC_1',
            timestamp=timezone.now(),
            sender='John',
            raw_message='Sold 3 bread 50 each',
            item='bread',
            quantity=Decimal('3'),
            price=Decimal('50'),
            gps_link='',
            image_flag=False,
            source='whatsapp_telegram',
            sync_attempts=5,
        )

        result = bulk_resync_to_sheets(limit=10, max_attempts=5)
        self.assertEqual(result['attempted'], 0)
        self.assertEqual(result['success_count'], 0)
        self.assertEqual(result['failed_count'], 0)
        self.assertIn('No eligible unsynced messages', result['errors'])

    def test_batch_append_messages_updates_synced_status(self):
        """Batch sheet sync should update only successfully appended messages."""
        raw_message = RawMessage.objects.create(
            telegram_message_id='test_sheet_batch',
            sender='Test',
            content='Test content',
        )
        msg_hash = generate_message_hash('Test', 'Test content')
        processed = ProcessedMessage.objects.create(
            message_hash=msg_hash,
            raw_message=raw_message,
        )
        parsed1 = ParsedMessage.objects.create(
            processed_message=processed,
            message_id='MSG_BATCH_1',
            timestamp=timezone.now(),
            sender='John',
            raw_message='Sold 3 bread 50 each',
            item='bread',
            quantity=Decimal('3'),
            price=Decimal('50'),
            gps_link='',
            image_flag=False,
            source='whatsapp_telegram',
        )
        parsed2 = ParsedMessage.objects.create(
            processed_message=processed,
            message_id='MSG_BATCH_2',
            timestamp=timezone.now(),
            sender='Mary',
            raw_message='Paid 200 for 4 milk',
            item='milk',
            quantity=Decimal('4'),
            price=Decimal('200'),
            gps_link='',
            image_flag=False,
            source='whatsapp_telegram',
        )

        service = GoogleSheetsService()
        service._initialized = True
        service._sheet = MagicMock()
        service._sheet.row_values.return_value = service.SHEET_COLUMNS
        service._sheet.col_values.return_value = ['message_id', 'MSG_BATCH_1']
        service._sheet.get_all_values.return_value = [
            service.SHEET_COLUMNS,
            ['', 'MSG_BATCH_1'] + [''] * 19,
        ]
        service._sheet.update.return_value = None

        with patch('core.services.sheets.get_sheets_service', return_value=service), \
             patch('core.services.sheets.GoogleSheetsService.is_available', return_value=True):
            result = batch_append_messages([parsed1, parsed2])

        self.assertEqual(result['success_count'], 2)
        self.assertEqual(result['failed_count'], 0)
        self.assertIn('MSG_BATCH_1', result['synced_message_ids'])
        self.assertIn('MSG_BATCH_2', result['synced_message_ids'])

        refreshed = ParsedMessage.objects.filter(message_id__in=['MSG_BATCH_1', 'MSG_BATCH_2'])
        self.assertTrue(all(msg.synced_to_sheets for msg in refreshed))

    @patch('core.services.sheet_sync.get_sheets_service')
    def test_sync_sheet_to_backend_mirrors_sheet_rows(self, mock_service):
        """Sheet sync should create, update, and delete local case rows."""
        existing = create_parsed_case(
            'MSG_KEEP',
            customer_name='Old Name',
            complaint_status='Open',
        )
        original_processed_id = existing.processed_message_id
        original_raw = existing.processed_message.raw_message
        original_raw.telegram_message_id = '777'
        original_raw.source_telegram_message_id = '777'
        original_raw.save(update_fields=['telegram_message_id', 'source_telegram_message_id'])
        deleted = create_parsed_case('MSG_DELETE')
        deleted_processed_id = deleted.processed_message_id
        deleted_raw_id = deleted.processed_message.raw_message_id

        service = MagicMock()
        service.is_available.return_value = True
        service.validate_sheet_structure.return_value = (True, '')
        service.fetch_rows.return_value = [
            {
                'row_number': 2,
                'values': {
                    'complaint id': 'COMP-1',
                    'message_id': 'MSG_KEEP',
                    'date reported': '29/04/2026',
                    'customer name': 'Updated Name',
                    'customer id / account': 'ACC-1',
                    'phone number': '0700000001',
                    'jbl reported by': 'Sheet User',
                    'branch / region': 'Nairobi',
                    'complaint category': 'No gas',
                    'complaint description': 'Edited in sheet',
                    'raw_message': 'Edited raw',
                    'gps_link': '',
                    'image_flag': 'TRUE',
                    'source': 'telegram bot',
                    'loan status': 'Active',
                    'loan at risk': 'No',
                    'risk level': 'Low',
                    'status': 'Closed',
                    'resolution details': 'Fixed',
                    'date resolved': '30/04/2026',
                    'days open': '1',
                },
            },
            {
                'row_number': 3,
                'values': {
                    'complaint id': 'COMP-2',
                    'message_id': 'MSG_NEW',
                    'date reported': '30/04/2026',
                    'customer name': 'New Customer',
                    'customer id / account': 'ACC-2',
                    'phone number': '0700000002',
                    'jbl reported by': 'Sheet User',
                    'complaint description': 'Created from sheet',
                    'status': 'Open',
                },
            },
        ]
        mock_service.return_value = service

        result = sync_sheet_to_backend(
            group_id='-100123',
            sheet_id='sheet_123',
            sheet_name='Cases',
            delete_missing=True,
        )

        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['created_count'], 1)
        self.assertEqual(result['updated_count'], 1)
        self.assertEqual(result['deleted_count'], 1)
        self.assertFalse(ParsedMessage.objects.filter(message_id='MSG_DELETE').exists())
        self.assertFalse(ProcessedMessage.objects.filter(pk=deleted_processed_id).exists())
        self.assertFalse(RawMessage.objects.filter(pk=deleted_raw_id).exists())

        existing.refresh_from_db()
        self.assertEqual(existing.customer_name, 'Updated Name')
        self.assertEqual(existing.complaint_status, 'Closed')
        self.assertEqual(existing.resolution_details, 'Fixed')
        self.assertEqual(existing.sheet_id, 'sheet_123')
        self.assertEqual(existing.sheet_name, 'Cases')
        self.assertTrue(existing.synced_to_sheets)
        self.assertTrue(existing.image_flag)
        self.assertEqual(existing.processed_message_id, original_processed_id)
        self.assertEqual(
            existing.processed_message.raw_message.telegram_message_id,
            '777',
        )

        created = ParsedMessage.objects.get(message_id='MSG_NEW')
        self.assertEqual(created.customer_name, 'New Customer')
        self.assertEqual(created.complaint_description, 'Created from sheet')
        self.assertEqual(created.group_id, '-100123')
        self.assertEqual(created.sheet_id, 'sheet_123')
        self.assertEqual(created.sheet_name, 'Cases')

    @patch('core.services.sheet_sync.get_sheets_service')
    def test_sync_sheet_to_backend_clear_sheet_removes_group_cases_and_dedupe(self, mock_service):
        """When the sheet is cleared, backend cases and stale dedupe rows are cleared too."""
        first = create_parsed_case('MSG_CLEAR_1', group_id='-100clear')
        second = create_parsed_case('MSG_CLEAR_2', group_id='-100clear')
        other = create_parsed_case('MSG_OTHER_GROUP', group_id='-100other')
        first_processed_id = first.processed_message_id
        second_processed_id = second.processed_message_id
        first_raw_id = first.processed_message.raw_message_id
        second_raw_id = second.processed_message.raw_message_id
        stale_raw = RawMessage.objects.create(
            telegram_message_id='STALE_RAW',
            source_telegram_message_id='STALE_RAW',
            sender='Agent',
            content='Previously deleted sheet row',
        )
        stale_processed = ProcessedMessage.objects.create(
            message_hash='stale-success-hash',
            raw_message=stale_raw,
            status='success',
        )

        service = MagicMock()
        service.is_available.return_value = True
        service.validate_sheet_structure.return_value = (True, '')
        service.fetch_rows.return_value = []
        mock_service.return_value = service

        result = sync_sheet_to_backend(
            group_id='-100clear',
            sheet_id='sheet_123',
            sheet_name='Cases',
            delete_missing=True,
        )

        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['row_count'], 0)
        self.assertEqual(result['deleted_count'], 2)
        self.assertEqual(result['backend_count'], 0)
        self.assertFalse(ParsedMessage.objects.filter(group_id='-100clear').exists())
        self.assertFalse(ProcessedMessage.objects.filter(pk__in=[first_processed_id, second_processed_id]).exists())
        self.assertFalse(RawMessage.objects.filter(pk__in=[first_raw_id, second_raw_id]).exists())
        self.assertFalse(ProcessedMessage.objects.filter(pk=stale_processed.pk).exists())
        self.assertFalse(RawMessage.objects.filter(pk=stale_raw.pk).exists())
        self.assertTrue(ParsedMessage.objects.filter(pk=other.pk).exists())
        self.assertTrue(ProcessedMessage.objects.filter(pk=other.processed_message_id).exists())
        self.assertTrue(RawMessage.objects.filter(pk=other.processed_message.raw_message_id).exists())


class LiveSheetRecordServiceTest(TestCase):
    """Test safe live worksheet row operations used by Django admin."""

    def setUp(self):
        self.config = GroupSheetConfiguration.objects.create(
            group_id='-100live',
            display_name='Live Cases',
            sheet_id='sheet_live',
            sheet_name='Complaints',
            workflow={'type': 'case'},
            sheet_schema={
                'header_row': 2,
                'columns': ['Complaint ID', 'message_id', 'Customer Name'],
                'formula_fields': ['complaint_id'],
                'field_headers': {
                    'complaint_id': 'Complaint ID',
                    'message_id': 'message_id',
                    'customer_name': 'Customer Name',
                },
            },
        )

    @patch('core.services.live_sheet_records.get_sheets_service')
    def test_load_live_sheet_table_preserves_headers_and_formula_cells(self, mock_service):
        from core.services.live_sheet_records import load_live_sheet_table

        sheet = MagicMock()
        sheet.get_all_values.side_effect = [
            [
                ['TITLE'],
                ['Complaint ID', 'message_id', 'Customer Name'],
                ['CMP-1', 'MSG_1', 'Jane Doe'],
            ],
            [
                ['TITLE'],
                ['Complaint ID', 'message_id', 'Customer Name'],
                ['=ROW()-2', 'MSG_1', 'Jane Doe'],
            ],
        ]
        service = MagicMock()
        service.is_available.return_value = True
        service._sheet = sheet
        mock_service.return_value = service

        table = load_live_sheet_table(self.config)

        self.assertEqual(
            table['headers'],
            ['Complaint ID', 'message_id', 'Customer Name'],
        )
        self.assertEqual(table['rows'][0]['row_number'], 3)
        self.assertEqual(table['rows'][0]['record_key'], 'MSG_1')
        self.assertTrue(table['rows'][0]['cells'][0]['is_readonly'])
        self.assertTrue(table['rows'][0]['cells'][1]['is_readonly'])
        self.assertFalse(table['rows'][0]['cells'][2]['is_readonly'])

    @patch('core.services.live_sheet_records.get_sheets_service')
    def test_load_live_sheet_table_uses_workflow_header_row_for_case_group(
        self,
        mock_service,
    ):
        """The live viewer should use the configured workflow header row."""
        from core.services.live_sheet_records import load_live_sheet_table

        self.config.workflow = {'type': 'case', 'header_row': 3}
        self.config.sheet_schema = {}
        self.config.save(update_fields=['workflow', 'sheet_schema'])

        sheet = MagicMock()
        sheet.get_all_values.side_effect = [
            [
                ['TITLE'],
                ['SUBTITLE'],
                ['Complaint ID', 'message_id', 'Customer Name'],
                ['CMP-1', 'MSG_1', 'Jane Doe'],
            ],
            [
                ['TITLE'],
                ['SUBTITLE'],
                ['Complaint ID', 'message_id', 'Customer Name'],
                ['=ROW()-3', 'MSG_1', 'Jane Doe'],
            ],
        ]
        service = MagicMock()
        service.is_available.return_value = True
        service._sheet = sheet
        mock_service.return_value = service

        table = load_live_sheet_table(self.config)

        self.assertEqual(table['header_row'], 3)
        self.assertEqual(
            table['headers'],
            ['Complaint ID', 'message_id', 'Customer Name'],
        )
        self.assertEqual(table['rows'][0]['row_number'], 4)

    @patch('core.services.live_sheet_records.get_sheets_service')
    def test_update_live_sheet_row_batches_only_changed_non_formula_cells(self, mock_service):
        from core.services.live_sheet_records import update_live_sheet_row

        sheet = MagicMock()
        sheet.get_all_values.side_effect = [
            [
                ['TITLE'],
                ['Complaint ID', 'message_id', 'Customer Name'],
                ['CMP-1', 'MSG_1', 'Jane Doe'],
            ],
            [
                ['TITLE'],
                ['Complaint ID', 'message_id', 'Customer Name'],
                ['=ROW()-2', 'MSG_1', 'Jane Doe'],
            ],
        ]
        service = MagicMock()
        service.is_available.return_value = True
        service._sheet = sheet
        mock_service.return_value = service

        result = update_live_sheet_row(
            self.config,
            'Complaints',
            3,
            {
                0: 'DO NOT CHANGE FORMULA',
                1: 'MSG_1',
                2: 'Jane Smith',
            },
        )

        self.assertTrue(result['changed'])
        self.assertEqual(
            result['changes'],
            {'Customer Name': {'old': 'Jane Doe', 'new': 'Jane Smith'}},
        )
        sheet.batch_update.assert_called_once_with(
            [{'range': 'C3', 'values': [['Jane Smith']]}],
            raw=True,
        )

    @patch('core.services.live_sheet_records.get_sheets_service')
    def test_delete_live_sheet_row_deletes_the_selected_live_row(self, mock_service):
        from core.services.live_sheet_records import delete_live_sheet_row

        sheet = MagicMock()
        sheet.get_all_values.side_effect = [
            [
                ['TITLE'],
                ['Complaint ID', 'message_id', 'Customer Name'],
                ['CMP-1', 'MSG_1', 'Jane Doe'],
            ],
            [
                ['TITLE'],
                ['Complaint ID', 'message_id', 'Customer Name'],
                ['=ROW()-2', 'MSG_1', 'Jane Doe'],
            ],
        ]
        service = MagicMock()
        service.is_available.return_value = True
        service._sheet = sheet
        mock_service.return_value = service

        result = delete_live_sheet_row(self.config, 'Complaints', 3)

        self.assertEqual(result['record_key'], 'MSG_1')
        sheet.delete_rows.assert_called_once_with(3)


class GroupResetServiceTest(TestCase):
    def test_reset_group_data_deletes_only_selected_group_records(self):
        from core.services.group_reset import group_data_counts, reset_group_data

        target = create_parsed_case('MSG_RESET_1', group_id='-100reset')
        other = create_parsed_case('MSG_KEEP_1', group_id='-100keep')
        CaseUpdate.objects.create(
            parsed_message=target,
            group_id='-100reset',
            raw_update_text='STATUS: Closed',
        )
        OrderApprovalUpdate.objects.create(
            group_id='-100reset',
            sheet_id='sheet_order',
            id_number='12345678',
        )
        MediaAttachment.objects.create(
            group_id='-100reset',
            telegram_file_id='file-reset',
            business_key_type='id_number',
            business_key_value='12345678',
        )
        JawabuVisitRecord.objects.create(
            group_id='-100reset',
            sheet_id='sheet_jawabu',
            national_id='12345678',
            primary_phone='254712345678',
        )
        LiveSheetRecordChange.objects.create(
            group_id='-100reset',
            sheet_id='sheet_order',
            sheet_tab='Orders',
            row_number=2,
            action='update',
            status='success',
        )

        before = group_data_counts('-100reset')
        result = reset_group_data('-100reset')

        self.assertGreater(before['parsed_messages'], 0)
        self.assertEqual(sum(result['after'].values()), 0)
        self.assertFalse(ParsedMessage.objects.filter(group_id='-100reset').exists())
        self.assertFalse(CaseUpdate.objects.filter(group_id='-100reset').exists())
        self.assertFalse(OrderApprovalUpdate.objects.filter(group_id='-100reset').exists())
        self.assertFalse(MediaAttachment.objects.filter(group_id='-100reset').exists())
        self.assertFalse(JawabuVisitRecord.objects.filter(group_id='-100reset').exists())
        self.assertFalse(LiveSheetRecordChange.objects.filter(group_id='-100reset').exists())
        self.assertFalse(ProcessedMessage.objects.filter(pk=target.processed_message_id).exists())
        self.assertFalse(RawMessage.objects.filter(pk=target.processed_message.raw_message_id).exists())

        self.assertTrue(ParsedMessage.objects.filter(pk=other.pk).exists())
        self.assertTrue(ProcessedMessage.objects.filter(pk=other.processed_message_id).exists())
        self.assertTrue(RawMessage.objects.filter(pk=other.processed_message.raw_message_id).exists())


class GroupConfigurationServiceTest(TestCase):
    """Test admin-managed group routing configuration."""

    def tearDown(self):
        from core.services.group_config import GroupRegistry
        GroupRegistry._instance = None
        super().tearDown()

    @override_settings(GROUP_MAPPING={}, GOOGLE_SHEET_ID='')
    def test_admin_group_configuration_routes_group_to_sheet(self):
        """A group can be configured from the admin UI instead of env JSON."""
        from core.services.group_config import GroupRegistry

        GroupSheetConfiguration.objects.create(
            group_id='-100999',
            display_name='Support Team',
            sheet_id='admin_sheet_123',
            sheet_name='Support Cases',
            sheet_schema={
                'field_headers': {
                    'message_id': 'Backend ID',
                    'customer_name': 'Client',
                },
            },
            workflow={'status_values': ['Open', 'In Progress', 'Closed']},
            parser_rules={'bot_username': 'hb_biogas_cases_bot'},
        )

        GroupRegistry._instance = None
        config = GroupRegistry.get_instance().get_group('-100999')

        self.assertIsNotNone(config)
        self.assertEqual(config.sheet_id, 'admin_sheet_123')
        self.assertEqual(config.sheet_name, 'Support Cases')
        self.assertEqual(config.sheet_schema.header('message_id'), 'Backend ID')
        self.assertEqual(config.workflow['status_values'][0], 'Open')
        self.assertEqual(config.parser_rules['bot_username'], 'hb_biogas_cases_bot')

    @override_settings(
        GROUP_MAPPING={
            '-100999': {
                'sheet_id': 'env_sheet',
                'sheet_name': 'Env Cases',
            },
        },
        GOOGLE_SHEET_ID='',
    )
    def test_admin_group_configuration_overrides_settings_mapping(self):
        """Admin UI config should win over env config for the same group."""
        from core.services.group_config import GroupRegistry

        GroupSheetConfiguration.objects.create(
            group_id='-100999',
            display_name='Admin Team',
            sheet_id='admin_sheet',
            sheet_name='Admin Cases',
        )

        GroupRegistry._instance = None
        config = GroupRegistry.get_instance().get_group('-100999')

        self.assertEqual(config.sheet_id, 'admin_sheet')
        self.assertEqual(config.sheet_name, 'Admin Cases')

    @override_settings(
        STORAGES={
            'default': {
                'BACKEND': 'django.core.files.storage.FileSystemStorage',
            },
            'staticfiles': {
                'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage',
            },
        },
    )
    def test_group_configuration_admin_changelist_renders(self):
        """Admin changelist should render without template context copy errors."""
        user = get_user_model().objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='password',
        )
        GroupSheetConfiguration.objects.create(
            group_id='-100777',
            display_name='Admin Render Test',
            sheet_id='sheet_777',
            sheet_name='Cases',
            workflow={'type': 'case'},
        )
        GroupSheetConfiguration.objects.create(
            group_id='-100778',
            display_name='Order Admin Render Test',
            sheet_id='sheet_778',
            sheet_name='Orders',
            workflow={'type': 'order_approval'},
        )
        self.client.force_login(user)

        response = self.client.get('/admin/core/groupsheetconfiguration/')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Admin Render Test')
        self.assertContains(response, 'Order Admin Render Test')
        self.assertContains(response, 'Open live sheet records', count=2)
        self.assertContains(response, 'View complaint cases')
        self.assertContains(response, 'View order update audit')
        self.assertContains(response, 'View media audit', count=2)
        self.assertContains(response, 'group_id__exact=-100777')
        self.assertContains(response, 'sheet_id__exact=sheet_777')

    @override_settings(
        STORAGES={
            'default': {
                'BACKEND': 'django.core.files.storage.FileSystemStorage',
            },
            'staticfiles': {
                'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage',
            },
        },
    )
    @patch('core.services.live_sheet_records.load_live_sheet_table')
    def test_live_sheet_records_admin_view_renders_actual_headers(
        self,
        mock_load_table,
    ):
        """Live records view should display the worksheet header order and values."""
        user = get_user_model().objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='password',
        )
        config = GroupSheetConfiguration.objects.create(
            group_id='-100779',
            display_name='Live Admin Test',
            sheet_id='sheet_779',
            sheet_name='Complaints',
            workflow={'type': 'case'},
        )
        mock_load_table.return_value = {
            'sheet_tab': 'Complaints',
            'header_row': 2,
            'headers': ['Complaint ID', 'Customer Name'],
            'rows': [{
                'row_number': 3,
                'record_key': 'CMP-1',
                'values': ['CMP-1', 'Jane Doe'],
                'formula_indexes': [0],
                'cells': [
                    {
                        'index': 0,
                        'header': 'Complaint ID',
                        'value': 'CMP-1',
                        'is_formula': True,
                    },
                    {
                        'index': 1,
                        'header': 'Customer Name',
                        'value': 'Jane Doe',
                        'is_formula': False,
                    },
                ],
            }],
            'row_count': 1,
            'formula_indexes': [0],
            'workflow_type': 'case',
        }
        self.client.force_login(user)

        response = self.client.get(
            f'/admin/core/groupsheetconfiguration/{config.pk}/live-records/'
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Complaint ID')
        self.assertContains(response, 'Customer Name')
        self.assertContains(response, 'Jane Doe')
        self.assertContains(response, 'Edit')

    @override_settings(
        STORAGES={
            'default': {
                'BACKEND': 'django.core.files.storage.FileSystemStorage',
            },
            'staticfiles': {
                'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage',
            },
        },
    )
    @patch('core.admin.GroupSheetConfigurationAdmin._sync_case_mirror')
    @patch('core.services.live_sheet_records.load_live_sheet_table')
    @patch('core.services.live_sheet_records.update_live_sheet_row')
    def test_live_sheet_records_admin_update_creates_audit_record(
        self,
        mock_update_row,
        mock_load_table,
        mock_sync_mirror,
    ):
        """Admin live-row saves should be applied through the sheet service and audited."""
        user = get_user_model().objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='password',
        )
        config = GroupSheetConfiguration.objects.create(
            group_id='-100780',
            display_name='Live Update Test',
            sheet_id='sheet_780',
            sheet_name='Complaints',
            workflow={'type': 'case'},
        )
        mock_update_row.return_value = {
            'changed': True,
            'changes': {
                'Customer Name': {'old': 'Jane Doe', 'new': 'Jane Smith'},
            },
            'record_key': 'MSG_1',
            'sheet_tab': 'Complaints',
            'row_number': 3,
        }
        mock_load_table.return_value = {
            'sheet_tab': 'Complaints',
            'header_row': 2,
            'headers': [],
            'rows': [],
            'row_count': 0,
            'formula_indexes': [],
            'workflow_type': 'case',
        }
        mock_sync_mirror.return_value = {'status': 'success'}
        self.client.force_login(user)

        response = self.client.post(
            f'/admin/core/groupsheetconfiguration/{config.pk}/live-records/',
            {
                'action': 'update',
                'sheet_tab': 'Complaints',
                'row_number': '3',
                'col_1': 'MSG_1',
                'col_2': 'Jane Smith',
            },
        )

        self.assertEqual(response.status_code, 302)
        mock_update_row.assert_called_once()
        mock_sync_mirror.assert_called_once_with(config)
        audit = LiveSheetRecordChange.objects.get()
        self.assertEqual(audit.action, 'update')
        self.assertEqual(audit.record_key, 'MSG_1')
        self.assertEqual(audit.changed_by, 'admin')
        self.assertEqual(audit.status, 'success')

    def test_sheet_mirror_and_audit_admins_are_read_only(self):
        """Backend mirror and audit tables should not allow direct admin edits."""
        from django.contrib.admin.sites import AdminSite
        from core.admin import (
            LiveSheetRecordChangeAdmin,
            OrderApprovalUpdateAdmin,
            ParsedMessageAdmin,
        )

        site = AdminSite()
        request = MagicMock()

        for model, admin_class in [
            (ParsedMessage, ParsedMessageAdmin),
            (OrderApprovalUpdate, OrderApprovalUpdateAdmin),
            (LiveSheetRecordChange, LiveSheetRecordChangeAdmin),
        ]:
            model_admin = admin_class(model, site)
            self.assertFalse(model_admin.has_add_permission(request))
            self.assertFalse(model_admin.has_delete_permission(request))
            self.assertEqual(
                set(model_admin.get_readonly_fields(request)),
                {field.name for field in model._meta.fields},
            )

    def test_group_configuration_admin_form_generates_order_approval_workflow(self):
        """Order approval preset should avoid hand-written workflow JSON."""
        from core.admin import GroupSheetConfigurationAdminForm

        form = GroupSheetConfigurationAdminForm(data={
            'enabled': 'on',
            'group_id': '-100222',
            'display_name': 'Order Approval',
            'sheet_id': 'sheet_order_123',
            'sheet_name': 'Orders',
            'sheet_schema': '{}',
            'workflow': '{}',
            'parser_rules': '{}',
            'metadata': '{}',
            'workflow_preset': 'order_approval',
            'order_approval_search_tabs': 'Orders',
            'order_approval_match_field': 'id_number',
            'order_approval_media_field': 'media_urls',
            'order_approval_header_row': '3',
            'order_approval_media_root_folder': 'BRO Order Approvals',
        })

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(
            form.generated_workflow(),
            {
                'type': 'order_approval',
                'match_field': 'id_number',
                'search_sheet_names': ['Orders'],
                'create_sheet_name': 'Orders',
                'media_field': 'media_urls',
                'record_id_prefix': 'JBL',
                'header_row': 3,
                'media_root_folder': 'BRO Order Approvals',
            },
        )

    def test_group_configuration_admin_form_generates_case_workflow(self):
        """Case preset should make the existing complaints workflow explicit."""
        from core.admin import GroupSheetConfigurationAdminForm

        form = GroupSheetConfigurationAdminForm(data={
            'enabled': 'on',
            'group_id': '-100111',
            'display_name': 'Cases',
            'sheet_id': 'sheet_case_123',
            'sheet_name': 'Complaints Register',
            'sheet_schema': '{}',
            'workflow': '{}',
            'parser_rules': '{}',
            'metadata': '{}',
            'workflow_preset': 'case',
        })

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.generated_workflow(), {'type': 'case', 'header_row': 2})

    def test_group_configuration_admin_form_generates_jawabu_workflow(self):
        """Jawabu preset should expose the configurable import start date."""
        from core.admin import GroupSheetConfigurationAdminForm

        form = GroupSheetConfigurationAdminForm(data={
            'enabled': 'on',
            'group_id': '-100333',
            'display_name': 'Jawabu',
            'sheet_id': 'sheet_jawabu_123',
            'sheet_name': 'Jawabu Visits',
            'sheet_schema': '{}',
            'workflow': '{}',
            'parser_rules': '{}',
            'metadata': '{}',
            'workflow_preset': 'jawabu_homebiogas',
            'jawabu_import_start_date': '2026-05-01',
        })

        self.assertTrue(form.is_valid(), form.errors)
        workflow = form.generated_workflow()
        self.assertEqual(workflow['type'], 'jawabu_homebiogas')
        self.assertEqual(workflow['header_row'], 1)
        self.assertEqual(workflow['import_start_date'], '2026-05-01')

    def test_workflow_presets_define_order_approval_defaults(self):
        """Future workflow additions should follow the shared preset contract."""
        from core.services.workflow_presets import (
            build_workflow_from_preset,
            defaults_for_preset,
            preset_choices,
            preset_for_workflow,
        )

        choices = dict(preset_choices())
        self.assertIn('case', choices)
        self.assertEqual(choices['case'], 'Case / Complaints')
        self.assertIn('order_approval', choices)
        self.assertEqual(choices['order_approval'], 'Order Approval')

        case_defaults = defaults_for_preset('case')
        self.assertEqual(case_defaults['sheet_name'], 'Complaints Register')
        self.assertEqual(case_defaults['workflow'], {'type': 'case', 'header_row': 2})
        self.assertEqual(build_workflow_from_preset('case'), {'type': 'case', 'header_row': 2})
        self.assertEqual(preset_for_workflow({}), 'case')

        defaults = defaults_for_preset('order_approval')
        self.assertEqual(defaults['sheet_name'], 'Orders')
        self.assertEqual(defaults['sheet_schema'], {})
        self.assertEqual(defaults['parser_rules'], {})

        workflow = build_workflow_from_preset('order_approval')
        self.assertEqual(workflow['type'], 'order_approval')
        self.assertEqual(
            workflow['search_sheet_names'],
            ['Orders'],
        )
        self.assertEqual(workflow['header_row'], 2)
        self.assertEqual(workflow['media_root_folder'], '')
        self.assertEqual(workflow['record_id_prefix'], 'JBL')

    def test_workflow_preset_overrides_order_approval_tabs(self):
        """Admin can override preset tabs without editing raw JSON."""
        from core.services.workflow_presets import build_workflow_from_preset

        workflow = build_workflow_from_preset(
            'order_approval',
            overrides={
                'search_sheet_names': ['Pending', '190'],
                'match_field': 'id_number',
                'media_field': 'media_urls',
                'header_row': 3,
                'media_root_folder': 'BRO Order Approvals',
            },
        )

        self.assertEqual(workflow['search_sheet_names'], ['Pending', '190'])
        self.assertEqual(workflow['create_sheet_name'], 'Pending')
        self.assertEqual(workflow['header_row'], 3)
        self.assertEqual(workflow['media_root_folder'], 'BRO Order Approvals')

    def test_workflow_preset_uses_first_overridden_tab_for_order_creation(self):
        """New order approval rows are created in the first configured tab."""
        from core.services.workflow_presets import build_workflow_from_preset

        workflow = build_workflow_from_preset(
            'order_approval',
            overrides={
                'search_sheet_names': ['Sheet', 'Archive'],
                'match_field': 'id_number',
                'media_field': 'media_urls',
            },
        )

        self.assertEqual(workflow['search_sheet_names'], ['Sheet', 'Archive'])
        self.assertEqual(workflow['create_sheet_name'], 'Sheet')


class SheetAnalyzerServiceTest(TestCase):
    """Test Google Sheet analysis and schema suggestion."""

    def _mock_sheet_service(self):
        service = MagicMock()
        service.is_available.return_value = True
        service._sheet_id = 'sheet_123'
        service._sheet_name = 'Cases'
        service._api_initialized = True
        service._sheet.get_all_values.return_value = [
            [
                'Complaint ID', 'Backend ID', 'Reported On', 'Client',
                'Mobile', 'Case State', 'Fix Notes', 'Days Open',
            ],
            [
                '=ROW()-1', 'MSG_001', '11/05/2026', 'Jane Doe',
                '0712345678', 'Open', '', '=TODAY()-C2',
            ],
            [
                '=ROW()-1', 'MSG_002', '12/05/2026', 'John Doe',
                '254712345678', 'Closed', 'Fixed pipe', '=TODAY()-C3',
            ],
        ]
        service._sheet.get.return_value = [
            [
                'Complaint ID', 'Backend ID', 'Reported On', 'Client',
                'Mobile', 'Case State', 'Fix Notes', 'Days Open',
            ],
            [
                '=ROW()-1', 'MSG_001', '11/05/2026', 'Jane Doe',
                '0712345678', 'Open', '', '=TODAY()-C2',
            ],
        ]

        metadata_get = MagicMock()
        metadata_get.execute.return_value = {
            'sheets': [
                {
                    'properties': {'title': 'Cases'},
                    'data': [
                        {
                            'rowData': [
                                {'values': [{} for _ in range(8)]},
                                {
                                    'values': [
                                        {}, {}, {}, {}, {},
                                        {
                                            'dataValidation': {
                                                'condition': {
                                                    'type': 'ONE_OF_LIST',
                                                    'values': [
                                                        {'userEnteredValue': 'Open'},
                                                        {'userEnteredValue': 'In Progress'},
                                                        {'userEnteredValue': 'Closed'},
                                                    ],
                                                },
                                            },
                                        },
                                    ],
                                },
                            ],
                        },
                    ],
                },
            ],
        }
        service._sheets_api_service.spreadsheets.return_value.get.return_value = metadata_get
        return service

    @patch('core.services.sheet_analyzer.get_sheets_service')
    def test_analyze_google_sheet_suggests_schema_and_dropdowns(self, mock_service):
        from core.services.sheet_analyzer import analyze_google_sheet

        mock_service.return_value = self._mock_sheet_service()

        result = analyze_google_sheet('sheet_123', 'Cases')

        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['row_count'], 2)
        schema = result['suggested_schema']
        self.assertEqual(schema['field_headers']['message_id'], 'Backend ID')
        self.assertEqual(schema['field_headers']['customer_name'], 'Client')
        self.assertEqual(schema['field_headers']['customer_phone'], 'Mobile')
        self.assertEqual(schema['field_headers']['status'], 'Case State')
        self.assertIn('complaint_id', schema['formula_fields'])
        self.assertIn('days_open', schema['formula_fields'])
        self.assertEqual(
            result['workflow']['dropdown_values']['status'],
            ['Open', 'In Progress', 'Closed'],
        )

    @patch('core.services.sheet_analyzer.list_google_sheet_worksheets')
    @patch('core.services.sheet_analyzer.get_sheets_service')
    def test_analyze_google_sheet_unavailable_result_has_template_defaults(
        self,
        mock_service,
        mock_worksheets,
    ):
        from core.services.sheet_analyzer import analyze_google_sheet

        service = MagicMock()
        service.is_available.return_value = False
        mock_service.return_value = service
        mock_worksheets.return_value = (['Orders'], '')

        result = analyze_google_sheet(
            'sheet_123',
            'Orders',
            workflow={'header_row': 2},
        )

        self.assertEqual(result['status'], 'error')
        self.assertEqual(result['worksheet_titles'], ['Orders'])
        self.assertEqual(result['row_count'], 0)
        self.assertEqual(result['data_row_count'], 0)
        self.assertEqual(result['header_row'], 2)
        self.assertEqual(result['sample_size'], 0)
        self.assertEqual(result['headers'], [])
        self.assertEqual(result['columns'], [])
        self.assertEqual(result['warnings'], [])

    @patch('core.services.sheet_analyzer.get_sheets_service')
    def test_analyze_google_sheet_uses_configured_order_header_row(self, mock_service):
        from core.services.sheet_analyzer import analyze_google_sheet

        service = MagicMock()
        service.is_available.return_value = True
        service._sheet_id = 'sheet_123'
        service._sheet_name = 'Orders'
        service._api_initialized = False
        service._sheet.get_all_values.return_value = [
            ['ORDER APPROVAL FORM'],
            ['DATE VISITED', 'CUSTOMER NAME', 'ID NUMBER', 'Media URLs'],
            ['25-May-2026', 'Jane Doe', '113650221', ''],
        ]
        service._sheet.get.return_value = [
            ['DATE VISITED', 'CUSTOMER NAME', 'ID NUMBER', 'Media URLs'],
            ['25-May-2026', 'Jane Doe', '113650221', ''],
        ]
        mock_service.return_value = service

        result = analyze_google_sheet(
            'sheet_123',
            'Orders',
            workflow={'type': 'order_approval', 'header_row': 2},
        )

        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['header_row'], 2)
        self.assertEqual(result['headers'][2], 'ID NUMBER')
        self.assertEqual(result['warnings'], [])

    @patch('core.services.sheet_analyzer.get_sheets_service')
    def test_analyze_google_sheet_custom_workflow_has_no_complaint_warnings(self, mock_service):
        from core.services.sheet_analyzer import analyze_google_sheet

        service = MagicMock()
        service.is_available.return_value = True
        service._sheet_id = 'sheet_123'
        service._sheet_name = 'Any'
        service._api_initialized = False
        service._sheet.get_all_values.return_value = [
            ['External ID', 'Name', 'Status'],
            ['A1', 'Jane Doe', 'Open'],
        ]
        service._sheet.get.return_value = [
            ['External ID', 'Name', 'Status'],
            ['A1', 'Jane Doe', 'Open'],
        ]
        mock_service.return_value = service

        result = analyze_google_sheet('sheet_123', 'Any', workflow={})

        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['warnings'], [])

    def test_apply_analysis_to_config_saves_schema_workflow_and_metadata(self):
        from core.services.sheet_analyzer import apply_analysis_to_config

        config = GroupSheetConfiguration.objects.create(
            group_id='-100555',
            sheet_id='sheet_123',
            sheet_name='Cases',
        )
        analysis = {
            'suggested_schema': {
                'columns': ['Backend ID', 'Client'],
                'field_headers': {
                    'message_id': 'Backend ID',
                    'customer_name': 'Client',
                },
            },
            'workflow': {
                'dropdown_values': {
                    'status': ['Open', 'Closed'],
                },
            },
            'row_count': 10,
            'sample_size': 5,
            'columns': [{'header': 'Backend ID'}],
            'warnings': [],
        }

        apply_analysis_to_config(config, analysis)

        config.refresh_from_db()
        self.assertEqual(config.sheet_schema['field_headers']['message_id'], 'Backend ID')
        self.assertEqual(config.workflow['dropdown_values']['status'], ['Open', 'Closed'])
        self.assertEqual(config.metadata['sheet_analysis']['row_count'], 10)


class BotCommandServiceTest(TestCase):
    """Test database-backed Telegram bot commands."""

    def test_last_command_returns_latest_group_cases(self):
        """The /last command should return recent cases for the current group."""
        now = timezone.now()
        create_parsed_case(
            'MSG_OLD',
            customer_name='Old Customer',
            description='Older issue',
            created_at=now - timedelta(minutes=10),
        )
        create_parsed_case(
            'MSG_NEW',
            customer_name='New Customer',
            description='Newest issue',
            created_at=now,
        )
        create_parsed_case(
            'MSG_OTHER_GROUP',
            group_id='-999',
            customer_name='Other Group',
            description='Should not appear',
            created_at=now + timedelta(minutes=1),
        )

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/last 2', '-100123')

        self.assertEqual(result['status'], 'command')
        self.assertIn('Latest 2 case(s):', result['reply_text'])
        self.assertIn('NEW CUSTOMER', result['reply_text'])
        self.assertIn('Newest issue', result['reply_text'])
        self.assertIn('OLD CUSTOMER', result['reply_text'])
        self.assertIn('Older issue', result['reply_text'])
        self.assertNotIn('MSG_NEW', result['reply_text'])
        self.assertNotIn('MSG_OLD', result['reply_text'])
        self.assertNotIn('MSG_OTHER_GROUP', result['reply_text'])

    def test_last_command_hides_imported_from_sheets_placeholder(self):
        """List output should not show internal placeholders as problem text."""
        create_parsed_case(
            'MSG_IMPORTED',
            customer_name='Sheet Customer',
            customer_phone='0711111111',
            customer_id='ACC-1',
            description='',
        )
        ParsedMessage.objects.filter(message_id='MSG_IMPORTED').update(
            complaint_description='',
            raw_message='Imported from Google Sheets',
        )

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/last 1', '-100123')

        self.assertEqual(result['status'], 'command')
        self.assertIn('SHEET CUSTOMER', result['reply_text'])
        self.assertIn('Tel: 0711111111', result['reply_text'])
        self.assertIn('Customer ID: ACC-1', result['reply_text'])
        self.assertIn('Problem: no problem description recorded', result['reply_text'])
        self.assertNotIn('Imported from Google Sheets', result['reply_text'])
        self.assertNotIn('MSG_IMPORTED', result['reply_text'])

    def test_last_command_empty_group(self):
        """The command should return a useful empty-state response."""
        from core.services.commands import handle_bot_command

        result = handle_bot_command('last 5', '-100123')

        self.assertEqual(result['status'], 'command')
        self.assertEqual(result['reply_text'], 'No cases found for this group yet.')

    def test_non_command_returns_none(self):
        """Ordinary complaint text should continue to the parser."""
        from core.services.commands import handle_bot_command

        self.assertIsNone(
            handle_bot_command('CUSTOMER COMPLAIN: no gas supply', '-100123')
        )

    def test_case_command_returns_detail(self):
        """The /case command should return a full case view."""
        create_parsed_case(
            'MSG_DETAIL',
            customer_name='Jane Doe',
            customer_phone='0700000000',
            customer_id='ACC-123',
            description='Detailed case description',
            complaint_status='Open',
            synced_to_sheets=True,
        )

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/case MSG_DETAIL', '-100123')

        self.assertEqual(result['status'], 'command')
        self.assertIn('Case MSG_DETAIL', result['reply_text'])
        self.assertIn('Customer: JANE DOE', result['reply_text'])
        self.assertIn('Phone: 0700000000', result['reply_text'])
        self.assertIn('Customer ID: ACC-123', result['reply_text'])
        self.assertIn('Synced: yes', result['reply_text'])

    def test_search_command_matches_customer_and_description(self):
        """Search should scan customer fields and complaint text."""
        create_parsed_case(
            'MSG_LEAK',
            customer_name='Leak Customer',
            customer_phone='0711111111',
            description='Gas leakage near the pipe',
        )
        create_parsed_case(
            'MSG_OTHER',
            customer_name='Other Customer',
            customer_phone='0722222222',
            description='No matching text',
        )

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/search leakage', '-100123')

        self.assertEqual(result['status'], 'command')
        self.assertIn('LEAK CUSTOMER', result['reply_text'])
        self.assertIn('Gas leakage near the pipe', result['reply_text'])
        self.assertNotIn('MSG_LEAK', result['reply_text'])
        self.assertNotIn('OTHER CUSTOMER', result['reply_text'])

    def test_today_command_returns_todays_cases(self):
        """The /today command should show only cases created today."""
        now = timezone.now()
        create_parsed_case('MSG_TODAY', customer_name='Today Customer', created_at=now)
        create_parsed_case('MSG_YESTERDAY', customer_name='Yesterday Customer', created_at=now - timedelta(days=1))

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/today', '-100123')

        self.assertEqual(result['status'], 'command')
        self.assertIn('TODAY CUSTOMER', result['reply_text'])
        self.assertNotIn('YESTERDAY CUSTOMER', result['reply_text'])
        self.assertNotIn('MSG_TODAY', result['reply_text'])

    def test_unsynced_command_returns_unsynced_cases(self):
        """The /unsynced command should show only unsynced rows."""
        create_parsed_case(
            'MSG_UNSYNCED',
            customer_name='Unsynced Customer',
            synced_to_sheets=False,
            last_sync_error='Sheet unavailable',
        )
        create_parsed_case('MSG_SYNCED', customer_name='Saved Customer', synced_to_sheets=True)

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/unsynced 5', '-100123')

        self.assertEqual(result['status'], 'command')
        self.assertIn('UNSYNCED CUSTOMER', result['reply_text'])
        self.assertIn('Sheet unavailable', result['reply_text'])
        self.assertNotIn('SAVED CUSTOMER', result['reply_text'])
        self.assertNotIn('MSG_UNSYNCED', result['reply_text'])

    @override_settings(
        GOOGLE_SHEET_ID='sheet_123',
        GOOGLE_SHEET_TAB_NAME='Cases',
        GROUP_MAPPING={},
    )
    def test_group_command_returns_routing(self):
        """The /group command should show current group routing."""
        from core.services.commands import handle_bot_command
        from core.services.group_config import GroupRegistry

        GroupRegistry._instance = None
        result = handle_bot_command('/group', '-100123')

        self.assertEqual(result['status'], 'command')
        self.assertIn('Group: -100123', result['reply_text'])
        self.assertIn('Sheet ID: sheet_123', result['reply_text'])
        self.assertIn('Sheet tab: Cases', result['reply_text'])
        GroupRegistry._instance = None

    @override_settings(
        GOOGLE_SHEET_ID='sheet_123',
        GOOGLE_SHEET_TAB_NAME='Cases',
        GROUP_MAPPING={},
    )
    def test_health_command_returns_db_and_counts(self):
        """The /health command should return DB status and group counts."""
        from core.services.commands import handle_bot_command
        from core.services.group_config import GroupRegistry

        GroupRegistry._instance = None
        create_parsed_case('MSG_HEALTH', synced_to_sheets=False)

        result = handle_bot_command('/health', '-100123')

        self.assertEqual(result['status'], 'command')
        self.assertIn('Database: ok', result['reply_text'])
        self.assertIn('Group: configured', result['reply_text'])
        self.assertIn('Workflow: case', result['reply_text'])
        self.assertIn('Cases: 1', result['reply_text'])
        self.assertIn('Unsynced: 1', result['reply_text'])
        self.assertIn('Telegram token:', result['reply_text'])
        GroupRegistry._instance = None

    @override_settings(
        GOOGLE_SHEET_ID='',
        GOOGLE_SHEET_TAB_NAME='Cases',
        GROUP_MAPPING={},
        MEDIA_STORAGE_PROVIDER='google_drive',
        MEDIA_MAX_FILE_SIZE_MB=20,
        ORDER_APPROVAL_MAX_TOTAL_UPLOAD_MB=60,
    )
    def test_health_command_returns_order_workflow_diagnostics(self):
        """The /health command should include order workflow counters."""
        from core.models import GroupSheetConfiguration, MediaAttachment, OrderApprovalUpdate
        from core.services.commands import handle_bot_command
        from core.services.group_config import GroupRegistry

        GroupSheetConfiguration.objects.create(
            group_id='-100order',
            display_name='Order group',
            enabled=True,
            sheet_id='sheet_order',
            sheet_name='Orders',
            workflow={
                'type': 'order_approval',
                'search_sheet_names': ['Orders'],
                'header_row': 2,
            },
        )
        OrderApprovalUpdate.objects.create(
            group_id='-100order',
            sheet_id='sheet_order',
            id_number='113650221',
            update_status='failed',
        )
        MediaAttachment.objects.create(
            group_id='-100order',
            telegram_message_id='1',
            telegram_file_id='file_1',
            upload_status='failed',
        )
        GroupRegistry._instance = None

        result = handle_bot_command('/health', '-100order')

        self.assertEqual(result['status'], 'command')
        self.assertIn('Workflow: order_approval', result['reply_text'])
        self.assertIn('Order workflow', result['reply_text'])
        self.assertIn('Failed updates: 1', result['reply_text'])
        self.assertIn('Failed media: 1', result['reply_text'])
        self.assertIn('Order tabs: Orders', result['reply_text'])
        self.assertIn('Max upload total: 60 MB', result['reply_text'])
        self.assertIn('Image previews: off', result['reply_text'])
        GroupRegistry._instance = None

    def test_status_filter_commands(self):
        """Open, pending, and closed commands should filter by status."""
        create_parsed_case('MSG_OPEN', customer_name='Open Customer', complaint_status='Open')
        create_parsed_case('MSG_PENDING', customer_name='Pending Customer', complaint_status='')
        create_parsed_case('MSG_CLOSED', customer_name='Closed Customer', complaint_status='Closed')

        from core.services.commands import handle_bot_command

        open_result = handle_bot_command('/open 10', '-100123')
        pending_result = handle_bot_command('/pending 10', '-100123')
        closed_result = handle_bot_command('/closed 10', '-100123')

        self.assertIn('OPEN CUSTOMER', open_result['reply_text'])
        self.assertIn('PENDING CUSTOMER', open_result['reply_text'])
        self.assertNotIn('CLOSED CUSTOMER', open_result['reply_text'])
        self.assertIn('PENDING CUSTOMER', pending_result['reply_text'])
        self.assertNotIn('OPEN CUSTOMER', pending_result['reply_text'])
        self.assertIn('CLOSED CUSTOMER', closed_result['reply_text'])
        self.assertNotIn('OPEN CUSTOMER', closed_result['reply_text'])

    def test_stale_command_returns_old_not_closed_cases(self):
        """Stale should show old cases that are not closed."""
        now = timezone.now()
        create_parsed_case(
            'MSG_STALE',
            customer_name='Stale Customer',
            complaint_status='Open',
            created_at=now - timedelta(days=10),
        )
        create_parsed_case(
            'MSG_CLOSED_OLD',
            customer_name='Closed Old Customer',
            complaint_status='Closed',
            created_at=now - timedelta(days=10),
        )
        create_parsed_case(
            'MSG_RECENT',
            customer_name='Recent Customer',
            complaint_status='Open',
            created_at=now,
        )

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/stale 7', '-100123')

        self.assertIn('STALE CUSTOMER', result['reply_text'])
        self.assertIn('Age:', result['reply_text'])
        self.assertNotIn('CLOSED OLD CUSTOMER', result['reply_text'])
        self.assertNotIn('RECENT CUSTOMER', result['reply_text'])

    def test_errors_command_returns_sync_errors(self):
        """Errors should show cases with non-empty last_sync_error."""
        create_parsed_case(
            'MSG_ERROR',
            customer_name='Error Customer',
            last_sync_error='Google quota exceeded',
        )
        create_parsed_case('MSG_NO_ERROR', customer_name='No Error Customer', last_sync_error='')

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/errors 10', '-100123')

        self.assertIn('ERROR CUSTOMER', result['reply_text'])
        self.assertIn('Google quota exceeded', result['reply_text'])
        self.assertNotIn('NO ERROR CUSTOMER', result['reply_text'])

    def test_summary_today_command_returns_counts(self):
        """Summary today should count status and sync state for today."""
        now = timezone.now()
        create_parsed_case(
            'MSG_OPEN_SUMMARY',
            complaint_status='Open',
            synced_to_sheets=False,
            last_sync_error='Sheet unavailable',
            created_at=now,
        )
        create_parsed_case(
            'MSG_CLOSED_SUMMARY',
            complaint_status='Closed',
            synced_to_sheets=True,
            created_at=now,
        )
        create_parsed_case(
            'MSG_OLD_SUMMARY',
            complaint_status='Open',
            created_at=now - timedelta(days=2),
        )

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/summary today', '-100123')

        self.assertIn('Summary for today', result['reply_text'])
        self.assertIn('Total: 2', result['reply_text'])
        self.assertIn('Open/not closed: 1', result['reply_text'])
        self.assertIn('Closed: 1', result['reply_text'])
        self.assertIn('Unsynced: 1', result['reply_text'])
        self.assertIn('Sync errors: 1', result['reply_text'])

    def test_summary_week_command_includes_week_cases(self):
        """Summary week should include cases since the start of the local week."""
        now = timezone.now()
        create_parsed_case('MSG_WEEK_1', created_at=now)
        create_parsed_case('MSG_WEEK_2', created_at=now - timedelta(days=1))

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/summary week', '-100123')

        self.assertIn('Summary for this week', result['reply_text'])
        self.assertIn('Total:', result['reply_text'])

    def test_week_command_returns_this_weeks_cases(self):
        """The /week command should list cases created this week."""
        now = timezone.now()
        start_of_week = timezone.localdate() - timedelta(days=timezone.localdate().weekday())
        last_week = timezone.make_aware(
            datetime.combine(start_of_week - timedelta(days=1), datetime.min.time())
        )
        create_parsed_case('MSG_THIS_WEEK', customer_name='This Week Customer', created_at=now)
        create_parsed_case('MSG_LAST_WEEK', customer_name='Last Week Customer', created_at=last_week)

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/week', '-100123')

        self.assertIn("This week's cases", result['reply_text'])
        self.assertIn('THIS WEEK CUSTOMER', result['reply_text'])
        self.assertNotIn('LAST WEEK CUSTOMER', result['reply_text'])

    def test_phone_and_id_commands_lookup_cases(self):
        """Phone and customer ID lookup commands should search specific fields."""
        create_parsed_case(
            'MSG_PHONE',
            customer_name='Lookup Customer',
            customer_phone='254712345000',
            customer_id='ACC-123',
        )
        create_parsed_case(
            'MSG_OTHER_LOOKUP',
            customer_name='Other Lookup Customer',
            customer_phone='0799999999',
            customer_id='ACC-999',
        )

        from core.services.commands import handle_bot_command

        phone_result = handle_bot_command('/phone 0712345', '-100123')
        phone_254_result = handle_bot_command('/phone 254712345000', '-100123')
        id_result = handle_bot_command('/id ACC-123', '-100123')

        self.assertIn('LOOKUP CUSTOMER', phone_result['reply_text'])
        self.assertNotIn('OTHER LOOKUP CUSTOMER', phone_result['reply_text'])
        self.assertIn('LOOKUP CUSTOMER', phone_254_result['reply_text'])
        self.assertNotIn('OTHER LOOKUP CUSTOMER', phone_254_result['reply_text'])
        self.assertIn('LOOKUP CUSTOMER', id_result['reply_text'])
        self.assertNotIn('OTHER LOOKUP CUSTOMER', id_result['reply_text'])

    def test_missing_command_returns_cases_missing_requested_field(self):
        """Missing should filter by the requested blank field."""
        create_parsed_case('MSG_MISSING_PHONE', customer_name='Missing Phone Customer', customer_phone='')
        create_parsed_case('MSG_HAS_PHONE', customer_name='Has Phone Customer', customer_phone='0712345678')

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/missing phone 10', '-100123')

        self.assertIn('missing phone number', result['reply_text'])
        self.assertIn('MISSING PHONE CUSTOMER', result['reply_text'])
        self.assertNotIn('HAS PHONE CUSTOMER', result['reply_text'])

    def test_lowconfidence_command_returns_partial_or_incomplete_cases(self):
        """Low-confidence should include partial processing and incomplete cases."""
        create_parsed_case(
            'MSG_PARTIAL_CASE',
            customer_name='Partial Customer',
            processed_status='partial',
        )
        create_parsed_case(
            'MSG_INCOMPLETE_CASE',
            customer_name='',
        )
        create_parsed_case('MSG_COMPLETE_CASE', customer_name='Complete Customer')

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/lowconfidence 10', '-100123')

        self.assertIn('PARTIAL CUSTOMER', result['reply_text'])
        self.assertIn('partial processing', result['reply_text'])
        self.assertIn('UNKNOWN', result['reply_text'])
        self.assertIn('missing name', result['reply_text'])
        self.assertNotIn('COMPLETE CUSTOMER', result['reply_text'])

    def test_risk_command_filters_by_risk_level(self):
        """Risk should return cases matching the requested level."""
        create_parsed_case('MSG_HIGH_RISK', customer_name='High Risk Customer', risk_level='High')
        create_parsed_case('MSG_LOW_RISK', customer_name='Low Risk Customer', risk_level='Low')

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/risk high 10', '-100123')

        self.assertIn('HIGH RISK CUSTOMER', result['reply_text'])
        self.assertNotIn('LOW RISK CUSTOMER', result['reply_text'])

    def test_duplicates_command_reports_repeated_phone_or_customer_id(self):
        """Duplicates should group repeated phone numbers and customer IDs."""
        create_parsed_case(
            'MSG_DUP_1',
            customer_phone='0712000000',
            customer_id='ACC-DUP',
        )
        create_parsed_case(
            'MSG_DUP_2',
            customer_phone='0712000000',
            customer_id='ACC-DUP',
        )
        create_parsed_case(
            'MSG_UNIQUE',
            customer_phone='0799000000',
            customer_id='ACC-UNIQUE',
        )

        from core.services.commands import handle_bot_command

        result = handle_bot_command('/duplicates 30', '-100123')

        self.assertIn('Duplicate hints', result['reply_text'])
        self.assertIn('phone 0712000000: 2 case(s)', result['reply_text'])
        self.assertIn('customer ID ACC-DUP: 2 case(s)', result['reply_text'])
        self.assertNotIn('ACC-UNIQUE', result['reply_text'])

    def test_top_commands_report_regions_and_issues(self):
        """Top commands should aggregate non-blank region and issue fields."""
        create_parsed_case(
            'MSG_REGION_1',
            branch_region='Nairobi',
            complaint_category='Leak',
        )
        create_parsed_case(
            'MSG_REGION_2',
            branch_region='Nairobi',
            complaint_category='Leak',
        )
        create_parsed_case(
            'MSG_REGION_3',
            branch_region='Mombasa',
            complaint_category='No gas',
        )

        from core.services.commands import handle_bot_command

        regions_result = handle_bot_command('/top regions 7', '-100123')
        issues_result = handle_bot_command('/top issues 7', '-100123')

        self.assertIn('Top regions', regions_result['reply_text'])
        self.assertIn('Nairobi: 2', regions_result['reply_text'])
        self.assertIn('Mombasa: 1', regions_result['reply_text'])
        self.assertIn('Top issues', issues_result['reply_text'])
        self.assertIn('Leak: 2', issues_result['reply_text'])
        self.assertIn('No gas: 1', issues_result['reply_text'])

    def test_help_command_lists_useful_commands(self):
        """Help should expose the useful command set."""
        from core.services.commands import handle_bot_command

        result = handle_bot_command('/help', '-100123')

        self.assertIn('/phone 0712345678', result['reply_text'])
        self.assertIn('/missing phone 10', result['reply_text'])
        self.assertIn('/duplicates 30', result['reply_text'])
        self.assertIn('/top regions 7', result['reply_text'])
        self.assertIn('/batch', result['reply_text'])

    def test_help_command_uses_order_approval_group_commands(self):
        """Order approval groups should not see complaint/case commands."""
        from core.services.commands import handle_bot_command
        from core.services.group_config import GroupRegistry

        GroupSheetConfiguration.objects.create(
            group_id='-100order',
            display_name='Order group',
            enabled=True,
            sheet_id='sheet_order',
            sheet_name='Orders',
            workflow={'type': 'order_approval'},
        )
        GroupRegistry._instance = None

        result = handle_bot_command('/help', '-100order')

        self.assertIn('/order - Open the order approval form', result['reply_text'])
        self.assertIn('/form - Open the order approval form', result['reply_text'])
        self.assertIn('/group - Show this chat', result['reply_text'])
        self.assertNotIn('/last 5', result['reply_text'])
        self.assertNotIn('/case MSG_ID', result['reply_text'])

    def test_help_command_uses_jawabu_group_commands(self):
        """Jawabu groups should only show Jawabu import and shared commands."""
        from core.services.commands import handle_bot_command
        from core.services.group_config import GroupRegistry

        GroupSheetConfiguration.objects.create(
            group_id='-100jawabu',
            display_name='Jawabu group',
            enabled=True,
            sheet_id='sheet_jawabu',
            sheet_name='Jawabu Visits',
            workflow={'type': 'jawabu_homebiogas'},
        )
        GroupRegistry._instance = None

        result = handle_bot_command('/help', '-100jawabu')

        self.assertIn('/batch - Import a Jawabu WhatsApp export', result['reply_text'])
        self.assertIn('/group - Show this chat', result['reply_text'])
        self.assertNotIn('/order - Open', result['reply_text'])
        self.assertNotIn('/case MSG_ID', result['reply_text'])


class TelegramCommandMenuTest(TestCase):
    """Test native Telegram command autocomplete sync."""

    @override_settings(TELEGRAM_BOT_TOKEN='token', API_REQUEST_TIMEOUT=5)
    @patch('core.management.commands.sync_telegram_commands.requests.post')
    def test_sync_telegram_commands_sets_workflow_specific_group_scopes(self, mock_post):
        response = MagicMock()
        response.status_code = 200
        response.json.return_value = {'ok': True, 'result': True}
        mock_post.return_value = response
        GroupSheetConfiguration.objects.create(
            group_id='-100order',
            display_name='Order group',
            enabled=True,
            sheet_id='sheet_order',
            sheet_name='Orders',
            workflow={'type': 'order_approval'},
        )
        GroupSheetConfiguration.objects.create(
            group_id='-100cases',
            display_name='Case group',
            enabled=True,
            sheet_id='sheet_cases',
            sheet_name='Cases',
            workflow={},
        )
        GroupSheetConfiguration.objects.create(
            group_id='-100jawabu',
            display_name='Jawabu group',
            enabled=True,
            sheet_id='sheet_jawabu',
            sheet_name='Jawabu Visits',
            workflow={'type': 'jawabu_homebiogas'},
        )

        output = StringIO()
        call_command('sync_telegram_commands', stdout=output)

        set_calls = [
            call for call in mock_post.call_args_list
            if call.args[0].endswith('/setMyCommands')
        ]
        delete_calls = [
            call for call in mock_post.call_args_list
            if call.args[0].endswith('/deleteMyCommands')
        ]
        payloads = [call.kwargs['json'] for call in set_calls]
        scopes = [payload['scope'] for payload in payloads]
        self.assertIn({'type': 'all_private_chats'}, scopes)
        self.assertNotIn({'type': 'all_group_chats'}, scopes)
        self.assertIn({'type': 'chat', 'chat_id': '-100order'}, scopes)
        self.assertIn({'type': 'chat', 'chat_id': '-100cases'}, scopes)
        self.assertIn({'type': 'chat', 'chat_id': '-100jawabu'}, scopes)
        self.assertEqual(
            delete_calls[0].kwargs['json']['scope'],
            {'type': 'all_group_chats'},
        )
        order_payload = next(
            payload for payload in payloads
            if payload['scope'] == {'type': 'chat', 'chat_id': '-100order'}
        )
        order_commands = [item['command'] for item in order_payload['commands']]
        self.assertIn('order', order_commands)
        self.assertIn('form', order_commands)
        self.assertIn('group', order_commands)
        self.assertNotIn('last', order_commands)
        self.assertNotIn('case', order_commands)
        case_payload = next(
            payload for payload in payloads
            if payload['scope'] == {'type': 'chat', 'chat_id': '-100cases'}
        )
        case_commands = [item['command'] for item in case_payload['commands']]
        self.assertIn('last', case_commands)
        self.assertIn('case', case_commands)
        self.assertIn('group', case_commands)
        jawabu_payload = next(
            payload for payload in payloads
            if payload['scope'] == {'type': 'chat', 'chat_id': '-100jawabu'}
        )
        jawabu_commands = [item['command'] for item in jawabu_payload['commands']]
        self.assertIn('batch', jawabu_commands)
        self.assertIn('group', jawabu_commands)
        self.assertNotIn('order', jawabu_commands)
        self.assertNotIn('case', jawabu_commands)
        self.assertNotIn('order', case_commands)

    def test_sync_telegram_commands_dry_run_lists_group_scope_without_token(self):
        GroupSheetConfiguration.objects.create(
            group_id='-100order',
            display_name='Order group',
            enabled=True,
            sheet_id='sheet_order',
            sheet_name='Orders',
            workflow={'type': 'order_approval'},
        )

        output = StringIO()
        call_command('sync_telegram_commands', '--dry-run', '--group-id=-100order', stdout=output)

        text = output.getvalue()
        self.assertIn('Would sync chat -100order', text)
        self.assertIn('/order', text)
        self.assertIn('/group', text)

    @override_settings(TELEGRAM_BOT_TOKEN='token', API_REQUEST_TIMEOUT=5)
    @patch('core.management.commands.sync_telegram_commands.requests.post')
    def test_sync_telegram_commands_updates_migrated_group_id(self, mock_post):
        failed_response = MagicMock()
        failed_response.status_code = 400
        failed_response.text = (
            '{"ok":false,"error_code":400,'
            '"description":"Bad Request: group chat was upgraded to a supergroup chat",'
            '"parameters":{"migrate_to_chat_id":-1003817885962}}'
        )
        failed_response.json.return_value = {
            'ok': False,
            'error_code': 400,
            'description': 'Bad Request: group chat was upgraded to a supergroup chat',
            'parameters': {'migrate_to_chat_id': -1003817885962},
        }
        success_response = MagicMock()
        success_response.status_code = 200
        success_response.json.return_value = {'ok': True, 'result': True}
        mock_post.side_effect = [failed_response, success_response]
        config = GroupSheetConfiguration.objects.create(
            group_id='-5259879581',
            display_name='Order group',
            enabled=True,
            sheet_id='sheet_order',
            sheet_name='Orders',
            workflow={'type': 'order_approval'},
        )

        output = StringIO()
        call_command(
            'sync_telegram_commands',
            '--group-id=-5259879581',
            stdout=output,
        )

        config.refresh_from_db()
        self.assertEqual(config.group_id, '-1003817885962')
        self.assertEqual(config.metadata['migrated_from_chat_id'], '-5259879581')
        payloads = [call.kwargs['json'] for call in mock_post.call_args_list]
        self.assertEqual(payloads[0]['scope'], {'type': 'chat', 'chat_id': '-5259879581'})
        self.assertEqual(payloads[1]['scope'], {'type': 'chat', 'chat_id': '-1003817885962'})
        self.assertIn(
            'Updated migrated Telegram group -5259879581 -> -1003817885962',
            output.getvalue(),
        )


@override_settings(TELEGRAM_WEBHOOK_SECRET=None, DEBUG=True)
class TelegramWebhookViewTest(TestCase):
    """Test the Telegram webhook endpoint."""
    
    @patch('core.api.views._process_telegram_message')
    def test_webhook_receives_message(self, mock_process):
        """Test webhook accepts message."""
        mock_process.return_value = {
            'status': 'success',
            'message_id': 'MSG_TEST',
        }
        
        payload = {
            'update_id': 123456,
            'message': {
                'message_id': 789,
                'from': {'id': 123, 'first_name': 'Test'},
                'chat': {'id': -100123, 'type': 'group'},
                'date': 1711123456,
                'text': 'Sold 3 bread 50 each',
            }
        }
        
        response = self.client.post(
            '/api/webhook/telegram/',
            data=json.dumps(payload),
            content_type='application/json'
        )
        
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['status'], 'success')

    @patch('core.api.views._send_telegram_reply')
    @patch('core.api.views._process_telegram_message')
    def test_webhook_returns_partial_response_on_sync_failure(self, mock_process, mock_reply):
        """Webhook should return partial status when Google Sheets sync fails."""
        mock_process.return_value = {
            'status': 'partial',
            'message_id': 'MSG_PARTIAL',
            'error': 'Google Sheets sync failed',
        }

        payload = {
            'update_id': 123456,
            'message': {
                'message_id': 789,
                'from': {'id': 123, 'first_name': 'Test'},
                'chat': {'id': -100123, 'type': 'group'},
                'date': 1711123456,
                'text': 'Sold 3 bread 50 each',
            }
        }

        response = self.client.post(
            '/api/webhook/telegram/',
            data=json.dumps(payload),
            content_type='application/json'
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['status'], 'partial')
        self.assertIn('warnings', data)
        self.assertEqual(data['warnings'][0], 'Google Sheets sync failed')
    
    def test_webhook_rejects_get(self):
        """Test webhook only accepts POST."""
        response = self.client.get('/api/webhook/telegram/')
        self.assertEqual(response.status_code, 405)
    
    def test_health_check(self):
        """Test health check endpoint."""
        response = self.client.get('/api/health/')
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['status'], 'success')

    @override_settings(TELEGRAM_WEBHOOK_SECRET='', DEBUG=False)
    def test_webhook_requires_secret_in_production(self):
        payload = {
            'update_id': 123456,
            'message': {
                'message_id': 789,
                'from': {'id': 123, 'first_name': 'Test'},
                'chat': {'id': -100123, 'type': 'group'},
                'date': 1711123456,
                'text': 'Sold 3 bread 50 each',
            }
        }

        response = self.client.post(
            '/api/webhook/telegram/',
            data=json.dumps(payload),
            content_type='application/json'
        )

        self.assertEqual(response.status_code, 503)
        self.assertEqual(response.json()['code'], 'WEBHOOK_SECRET_REQUIRED')

    @override_settings(TELEGRAM_WEBHOOK_SECRET='expected-secret', DEBUG=False)
    def test_webhook_rejects_invalid_secret(self):
        payload = {
            'update_id': 123456,
            'message': {
                'message_id': 789,
                'from': {'id': 123, 'first_name': 'Test'},
                'chat': {'id': -100123, 'type': 'group'},
                'date': 1711123456,
                'text': 'Sold 3 bread 50 each',
            }
        }

        response = self.client.post(
            '/api/webhook/telegram/',
            data=json.dumps(payload),
            content_type='application/json',
            HTTP_X_TELEGRAM_BOT_API_SECRET_TOKEN='wrong-secret',
        )

        self.assertEqual(response.status_code, 401)

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.api.views._process_single_message')
    def test_telegram_message_ignored_when_bot_not_tagged(self, mock_process):
        """Webhook messages should not process or sync unless the bot is tagged."""
        from core.api.views import _process_telegram_message

        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': 'CUSTOMER COMPLAIN: no gas supply',
        })

        self.assertEqual(result['status'], 'ignored')
        mock_process.assert_not_called()

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.services.case_updates.handle_case_status_reply')
    @patch('core.api.views._process_single_message')
    def test_telegram_status_reply_can_be_untagged(
        self,
        mock_process,
        mock_update,
    ):
        """Untagged Status replies to case messages should route as updates."""
        from core.api.views import _process_telegram_message

        mock_update.return_value = {
            'status': 'command',
            'reply_text': 'OK. Case updated.',
        }

        result = _process_telegram_message({
            'message_id': 456,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': 'Status: resolved - repaired',
            'reply_to_message': {'message_id': 123},
        })

        self.assertEqual(result['status'], 'command')
        mock_update.assert_called_once()
        self.assertEqual(
            mock_update.call_args.kwargs['reply_to_telegram_message_id'],
            '123',
        )
        mock_process.assert_not_called()

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.api.views._process_single_message')
    def test_telegram_status_without_reply_returns_help(self, mock_process):
        """Tagged Status text without reply context should not create a case."""
        from core.api.views import _process_telegram_message

        result = _process_telegram_message({
            'message_id': 456,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': '@biogas_bot Status: resolved - repaired',
        })

        self.assertEqual(result['status'], 'command')
        self.assertIn('/update MSG_ID', result['reply_text'])
        mock_process.assert_not_called()

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.api.views._process_single_message')
    def test_telegram_message_processes_tagged_content_only(self, mock_process):
        """The bot mention is stripped before parsing real message content."""
        from core.api.views import _process_telegram_message

        mock_process.return_value = {'status': 'success', 'message_id': 'MSG_TEST'}
        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': '@biogas_bot CUSTOMER COMPLAIN: no gas supply',
        })

        self.assertEqual(result['status'], 'success')
        self.assertEqual(
            mock_process.call_args.kwargs['content'],
            'CUSTOMER COMPLAIN: no gas supply',
        )

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.api.views._process_single_message')
    def test_order_approval_message_in_case_group_returns_config_help(self, mock_process):
        """Order-style messages should not fall through to complaint sheet sync."""
        from core.api.views import _process_telegram_message
        from core.services.group_config import GroupRegistry

        GroupSheetConfiguration.objects.create(
            group_id='-100123',
            display_name='Misconfigured order group',
            enabled=True,
            sheet_id='sheet_order',
            sheet_name='Orders',
            workflow={},
        )
        GroupRegistry._instance = None

        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'supergroup'},
            'date': 1711123456,
            'text': '@biogas_bot\nID: 113650221\nCUSTOMER NAME: Jane Doe',
        })

        self.assertEqual(result['status'], 'command')
        self.assertIn('not configured for Order Approval', result['reply_text'])
        self.assertIn('Workflow preset to Order Approval', result['reply_text'])
        mock_process.assert_not_called()

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.api.views._process_single_message')
    def test_telegram_last_command_does_not_process_complaint(self, mock_process):
        """Commands should return a command response without storing/syncing."""
        from core.api.views import _process_telegram_message

        create_parsed_case(
            'MSG_RECENT',
            group_id='-100123',
            customer_name='Recent Customer',
            description='Recent issue',
        )

        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': '@biogas_bot /last 1',
        })

        self.assertEqual(result['status'], 'command')
        self.assertIn('RECENT CUSTOMER', result['reply_text'])
        self.assertIn('Recent issue', result['reply_text'])
        self.assertNotIn('MSG_RECENT', result['reply_text'])
        mock_process.assert_not_called()

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.api.views._process_single_message')
    def test_telegram_message_processes_multiple_tagged_complaint_cases(
        self,
        mock_process,
    ):
        """Multiple complaint cases in one tagged message should process separately."""
        from core.api.views import _process_telegram_message

        mock_process.side_effect = [
            {'status': 'success', 'message_id': 'MSG_1'},
            {'status': 'success', 'message_id': 'MSG_2'},
        ]
        payload_text = """@biogas_bot
*CUSTOMER COMPLAIN*
NAME: Jane Doe
TEL: 0712345678
ID: A123
NATURE OF THE PROBLEM: No gas supply
*CUSTOMER COMPLAIN: The system has stopped producing gas

*CUSTOMER COMPLAIN*
NAME: John Smith
TEL: 0798765432
ID: B456
NATURE OF THE PROBLEM: Gas leakage
*CUSTOMER COMPLAIN: Gas smell around the digester"""

        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': payload_text,
        })

        self.assertEqual(result['status'], 'batch_processed')
        self.assertEqual(result['total'], 2)
        self.assertEqual(result['success'], 2)
        self.assertEqual(mock_process.call_count, 2)
        self.assertEqual(
            mock_process.call_args_list[0].kwargs['telegram_message_id'],
            '123_0',
        )
        self.assertEqual(
            mock_process.call_args_list[1].kwargs['telegram_message_id'],
            '123_1',
        )
        self.assertEqual(
            mock_process.call_args_list[0].kwargs['source_telegram_message_id'],
            '123',
        )
        self.assertEqual(mock_process.call_args_list[0].kwargs['batch_index'], 0)
        self.assertEqual(mock_process.call_args_list[1].kwargs['batch_index'], 1)
        self.assertIn('Jane Doe', mock_process.call_args_list[0].kwargs['content'])
        self.assertIn('John Smith', mock_process.call_args_list[1].kwargs['content'])

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.api.views._process_single_message')
    def test_telegram_batch_result_counts_rejected_cases(self, mock_process):
        """Batch summaries should include rejected complaint blocks."""
        from core.api.views import _process_telegram_message

        mock_process.side_effect = [
            {
                'status': 'rejected',
                'missing_fields': ['Phone Number or Customer ID / Account'],
                'captured_fields': {'Customer Name': 'Jane Doe'},
            },
            {'status': 'success', 'message_id': 'MSG_2'},
        ]
        payload_text = """@biogas_bot
CUSTOMER COMPLAINT
NAME: Jane Doe
TEL: 0712345678
ID: A123
NATURE OF THE PROBLEM: No gas supply

CUSTOMER COMPLAINT
NAME: John Smith
TEL: 0798765432
ID: B456
COUNTY: Kisumu
NATURE OF THE PROBLEM: Gas leakage"""

        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': payload_text,
        })

        self.assertEqual(result['status'], 'batch_processed')
        self.assertEqual(result['total'], 2)
        self.assertEqual(result['success'], 1)
        self.assertEqual(result['rejected'], 1)

    @override_settings(
        TELEGRAM_BOT_USERNAME='biogas_bot',
        WHATSAPP_BATCH_MAX_MESSAGES=50,
    )
    @patch('core.api.views._sync_case_sheet_for_batch')
    @patch('core.api.views._process_single_message')
    def test_telegram_batch_command_processes_whatsapp_export(self, mock_process, mock_sync):
        """The /batch command should process complaint entries from exports only."""
        from core.api.views import _process_telegram_message

        mock_process.side_effect = [
            {'status': 'success', 'message_id': 'MSG_1'},
            {'status': 'duplicate', 'message_id': '123_wa_2'},
        ]
        mock_sync.side_effect = [
            {'status': 'success', 'row_count': 4, 'backend_count': 4, 'errors': []},
            {'status': 'success', 'row_count': 5, 'backend_count': 5, 'errors': []},
        ]
        payload_text = """@biogas_bot /batch
23/05/2026, 12:46 - System: Normal group chat
23/05/2026, 12:47 - Alice Agent: CUSTOMER COMPLAIN
NAME: Jane Doe
TEL: 0712345678
ID: A123
COUNTY: KISUMU
NATURE OF THE PROBLEM: No gas supply
23/05/2026, 12:48 - Bob Agent: CUSTOMER COMPLAIN
NAME: John Smith
TEL: 0798765432
ID: B456
COUNTY: NAKURU
NATURE OF THE PROBLEM: Gas leakage"""

        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': payload_text,
        })

        self.assertEqual(result['status'], 'batch_processed')
        self.assertEqual(result['source'], 'whatsapp_export')
        self.assertEqual(result['export_messages'], 3)
        self.assertEqual(result['skipped_non_complaint'], 1)
        self.assertEqual(result['total'], 2)
        self.assertEqual(result['success'], 1)
        self.assertEqual(result['duplicates'], 1)
        self.assertEqual(mock_process.call_count, 2)
        self.assertEqual(
            mock_process.call_args_list[0].kwargs['telegram_message_id'],
            '123_wa_1',
        )
        self.assertEqual(
            mock_process.call_args_list[0].kwargs['source'],
            'whatsapp_export',
        )
        self.assertEqual(
            mock_process.call_args_list[0].kwargs['source_telegram_message_id'],
            '123',
        )
        self.assertEqual(mock_process.call_args_list[0].kwargs['batch_index'], 1)
        self.assertEqual(mock_process.call_args_list[0].kwargs['sender'], 'Alice Agent')
        self.assertIs(mock_process.call_args_list[0].kwargs['sync_after_success'], False)
        self.assertIs(mock_process.call_args_list[0].kwargs['defer_sheet_sync'], True)
        self.assertEqual(mock_sync.call_count, 2)
        self.assertEqual(mock_sync.call_args_list[0].args, ('-100123',))
        self.assertEqual(mock_sync.call_args_list[0].kwargs, {'delete_missing': True})
        self.assertEqual(mock_sync.call_args_list[1].args, ('-100123',))
        self.assertEqual(mock_sync.call_args_list[1].kwargs, {'delete_missing': False})
        self.assertEqual(result['sheet_sync_before']['row_count'], 4)
        self.assertEqual(result['sheet_sync_after']['row_count'], 5)
        self.assertEqual(
            timezone.localtime(mock_process.call_args_list[0].kwargs['received_at']).strftime('%d/%m/%Y %H:%M'),
            '23/05/2026 12:47',
        )

    @override_settings(
        TELEGRAM_BOT_USERNAME='biogas_bot',
        WHATSAPP_BATCH_ASYNC_THRESHOLD=1,
    )
    @patch('core.api.views._start_case_batch_background_import')
    @patch('core.api.views._process_single_message')
    def test_large_telegram_batch_command_starts_background_import(
        self,
        mock_process,
        mock_start_background,
    ):
        """Large WhatsApp exports should not block the webhook worker."""
        from core.api.views import _process_telegram_message

        payload_text = """@biogas_bot /batch
23/05/2026, 12:47 - Alice Agent: CUSTOMER COMPLAIN
NAME: Jane Doe
TEL: 0712345678
ID: A123
COUNTY: KISUMU
NATURE OF THE PROBLEM: No gas supply
23/05/2026, 12:48 - Bob Agent: CUSTOMER COMPLAIN
NAME: John Smith
TEL: 0798765432
ID: B456
COUNTY: NAKURU
NATURE OF THE PROBLEM: Gas leakage"""

        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': payload_text,
        })

        self.assertEqual(result['status'], 'command')
        self.assertIn('WhatsApp batch import started', result['reply_text'])
        self.assertIn('Export messages found: 2', result['reply_text'])
        mock_start_background.assert_called_once()
        mock_process.assert_not_called()

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.api.views._process_single_message')
    def test_telegram_batch_command_requires_export_content(self, mock_process):
        """A bare /batch command should explain how to attach a WhatsApp export."""
        from core.api.views import _process_telegram_message

        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'text': '@biogas_bot /batch',
        })

        self.assertEqual(result['status'], 'command')
        self.assertIn('WhatsApp export', result['reply_text'])
        mock_process.assert_not_called()

    @override_settings(TELEGRAM_BOT_USERNAME='biogas_bot')
    @patch('core.api.views._sync_case_sheet_for_batch')
    @patch('core.api.views._download_telegram_text_document')
    @patch('core.api.views._process_single_message')
    def test_telegram_batch_command_prefers_attached_export(
        self,
        mock_process,
        mock_download,
        mock_sync,
    ):
        """Caption notes after /batch should not hide the attached export file."""
        from core.api.views import _process_telegram_message

        mock_download.return_value = (
            """23/05/2026, 12:47 - Alice Agent: CUSTOMER COMPLAIN
NAME: Jane Doe
TEL: 0712345678
ID: A123
COUNTY: KISUMU
NATURE OF THE PROBLEM: No gas supply""",
            '',
        )
        mock_process.return_value = {'status': 'success', 'message_id': 'MSG_1'}
        mock_sync.return_value = {'status': 'success', 'row_count': 1, 'backend_count': 1, 'errors': []}

        result = _process_telegram_message({
            'message_id': 123,
            'from': {'first_name': 'Test'},
            'chat': {'id': -100123, 'type': 'group'},
            'date': 1711123456,
            'caption': '@biogas_bot /batch please process this export',
            'document': {
                'file_id': 'file_123',
                'file_name': 'WhatsApp Chat.txt',
                'mime_type': 'text/plain',
                'file_size': 500,
            },
        })

        self.assertEqual(result['status'], 'batch_processed')
        self.assertEqual(result['success'], 1)
        self.assertEqual(mock_sync.call_count, 2)
        mock_download.assert_called_once()
        mock_process.assert_called_once()
        self.assertIs(mock_process.call_args.kwargs['sync_after_success'], False)
        self.assertIs(mock_process.call_args.kwargs['defer_sheet_sync'], True)

    @override_settings(TELEGRAM_BOT_TOKEN='token')
    @patch('core.api.views.requests.post')
    def test_telegram_reply_uses_plain_ascii_status_text(self, mock_post):
        """Telegram replies should not contain mojibake status prefixes."""
        from core.api.views import _send_telegram_reply

        _send_telegram_reply(
            {
                'message_id': 123,
                'chat': {'id': -100123},
            },
            {
                'status': 'success',
                'message_id': 'MSG_REPLY_1',
                'captured_fields': {
                    'Sender': 'Agent',
                    'Customer Name': 'Jane',
                    'Phone Number': '254712345678',
                    'Complaint Description': 'No gas supply',
                },
            },
        )

        text = mock_post.call_args.kwargs['data']['text']
        self.assertIn('OK. Message received and saved successfully', text)
        self.assertIn('Case ID: MSG_REPLY_1', text)
        self.assertIn('(use this for /update)', text)
        self.assertIn('Captured:\n', text)
        self.assertIn('Sender: Agent', text)
        self.assertIn('Customer Name: Jane', text)
        self.assertIn('Phone Number: 254712345678', text)
        self.assertIn('Complaint Description: No gas supply', text)
        self.assertNotIn('Captured: Customer Name', text)
        self.assertTrue(text.isascii())

    @override_settings(TELEGRAM_BOT_TOKEN='token')
    @patch('core.api.views.requests.post')
    def test_telegram_reply_shows_partial_warning_details(self, mock_post):
        """Partial complaint replies should tell staff which fields are missing."""
        from core.api.views import _send_telegram_reply

        _send_telegram_reply(
            {
                'message_id': 123,
                'chat': {'id': -100123},
            },
            {
                'status': 'partial',
                'message_id': 'MSG_REPLY_PARTIAL',
                'captured_fields': {
                    'Customer Name': 'Jane',
                    'Phone Number': '254712345678',
                    'Complaint Description': 'No gas supply',
                },
                'warnings': [
                    'Missing required complaint field(s): Customer ID / Account',
                ],
            },
        )

        text = mock_post.call_args.kwargs['data']['text']
        self.assertIn('Warning: Message partially processed', text)
        self.assertIn('Warnings:\n', text)
        self.assertIn('Customer ID / Account', text)
        self.assertIn('Case ID: MSG_REPLY_PARTIAL', text)

    @override_settings(TELEGRAM_BOT_TOKEN='token')
    @patch('core.api.views.requests.post')
    def test_telegram_reply_shows_rejected_complaint_details(self, mock_post):
        """Rejected complaint replies should be clear and should not show a Case ID."""
        from core.api.views import _send_telegram_reply

        _send_telegram_reply(
            {
                'message_id': 123,
                'chat': {'id': -100123},
            },
            {
                'status': 'rejected',
                'missing_fields': ['Phone Number or Customer ID / Account'],
                'captured_fields': {
                    'Customer Name': 'Jane',
                    'Phone Number': '254712345678',
                    'Customer ID': 'A12345',
                    'Complaint Description': 'No gas supply',
                },
            },
        )

        text = mock_post.call_args.kwargs['data']['text']
        self.assertIn('Rejected. Complaint was not saved', text)
        self.assertIn('Missing required fields:', text)
        self.assertIn('Phone Number or Customer ID / Account', text)
        self.assertIn('Required complaint fields:', text)
        self.assertIn('Customer Name: Jane', text)
        self.assertNotIn('Case ID:', text)

    @override_settings(TELEGRAM_BOT_TOKEN='token')
    @patch('core.api.views.requests.post')
    def test_telegram_reply_shows_batch_rejection_summary(self, mock_post):
        """Batch replies should list why any complaint block was rejected."""
        from core.api.views import _send_telegram_reply

        _send_telegram_reply(
            {
                'message_id': 123,
                'chat': {'id': -100123},
            },
            {
                'status': 'batch_processed',
                'total': 2,
                'success': 1,
                'rejected': 1,
                'duplicates': 0,
                'results': [
                    {
                        'status': 'rejected',
                        'missing_fields': ['Phone Number or Customer ID / Account'],
                    },
                    {'status': 'success', 'message_id': 'MSG_2'},
                ],
            },
        )

        text = mock_post.call_args.kwargs['data']['text']
        self.assertIn('Batch processed: 1/2 messages saved.', text)
        self.assertIn('Rejected: 1', text)
        self.assertIn('Missing: Phone Number or Customer ID / Account', text)
        self.assertIn('Each complaint must include NAME, TEL or ID, and NATURE OF THE PROBLEM', text)

    @override_settings(TELEGRAM_BOT_TOKEN='token')
    @patch('core.api.views.requests.post')
    def test_telegram_reply_shows_whatsapp_batch_summary(self, mock_post):
        """WhatsApp batch replies should include export analysis and duplicate counts."""
        from core.api.views import _send_telegram_reply

        _send_telegram_reply(
            {
                'message_id': 123,
                'chat': {'id': -100123},
            },
            {
                'status': 'batch_processed',
                'source': 'whatsapp_export',
                'export_messages': 4,
                'total': 2,
                'success': 1,
                'partial': 1,
                'rejected': 1,
                'duplicates': 1,
                'skipped_non_complaint': 2,
                'system_lines': 1,
                'sheet_sync_before': {
                    'status': 'success',
                    'row_count': 10,
                    'backend_count': 10,
                    'errors': [],
                },
                'sheet_sync_after': {
                    'status': 'success',
                    'row_count': 12,
                    'backend_count': 12,
                    'errors': [],
                },
                'results': [
                    {
                        'status': 'rejected',
                        'missing_fields': ['Phone Number'],
                    },
                ],
            },
        )

        text = mock_post.call_args.kwargs['data']['text']
        self.assertIn('WhatsApp batch processed', text)
        self.assertIn('Export messages found: 4', text)
        self.assertIn('Complaint entries processed: 2', text)
        self.assertIn('Saved: 2', text)
        self.assertIn('Skipped non-complaint chat messages: 2', text)
        self.assertIn('Skipped WhatsApp system lines: 1', text)
        self.assertIn('Duplicates skipped: 1', text)
        self.assertIn('Saved with sync warnings: 1', text)
        self.assertIn('Sheet sync before import: success (10 sheet rows, 10 backend cases)', text)
        self.assertIn('Sheet sync after import: success (12 sheet rows, 12 backend cases)', text)

    @override_settings(TELEGRAM_BOT_TOKEN='token')
    @patch('core.api.views.requests.post')
    def test_telegram_reply_sends_command_text(self, mock_post):
        """Command responses should be sent directly to Telegram."""
        from core.api.views import _send_telegram_reply

        _send_telegram_reply(
            {
                'message_id': 123,
                'chat': {'id': -100123},
            },
            {
                'status': 'command',
                'reply_text': 'Latest 1 case(s):\n1. MSG_RECENT',
            },
        )

        text = mock_post.call_args.kwargs['data']['text']
        self.assertEqual(text, 'Latest 1 case(s):\n1. MSG_RECENT')

    @patch('core.services.storage.process_and_store_message')
    @patch('core.services.group_config.GroupRegistry.get_instance')
    def test_process_single_message_passes_group_sheet_name(
        self,
        mock_registry_get_instance,
        mock_process_store,
    ):
        """Group-specific worksheet tabs should be forwarded to storage."""
        from core.api.views import _process_single_message

        registry = MagicMock()
        registry.get_group.return_value = MagicMock(
            sheet_id='sheet_123',
            sheet_name='Support Tickets',
            sheet_schema_config={'field_headers': {'message_id': 'Backend ID'}},
        )
        mock_registry_get_instance.return_value = registry

        parsed = MagicMock()
        parsed.message_id = 'MSG_TEST'
        parsed._processing_status = 'success'
        mock_process_store.return_value = parsed

        result = _process_single_message(
            telegram_message_id='123',
            content='CUSTOMER COMPLAIN: no gas',
            sender='Agent',
            has_image=False,
            received_at=timezone.now(),
            group_id='-100123',
        )

        self.assertEqual(result['status'], 'success')
        self.assertEqual(
            mock_process_store.call_args.kwargs['sheet_name'],
            'Support Tickets',
        )
        self.assertEqual(mock_process_store.call_args.kwargs['sheet_id'], 'sheet_123')
        self.assertEqual(
            mock_process_store.call_args.kwargs['sheet_schema'],
            {'field_headers': {'message_id': 'Backend ID'}},
        )



    def test_jawabu_farmers_csv_import_cleans_and_creates_master_record(self):
        from core.services.jawabu_master import import_jawabu_farmers_csv

        csv_file = StringIO(
            "Farmer Name,ID Number,Phone,Alternative Phone,County,Sub County,Branch\n"
            "Mary Njeri,1382654,0720570031,0785116424,Embu County,Manyatta,Embu\n"
        )

        result = import_jawabu_farmers_csv(csv_file, source_name='farmers.csv')

        self.assertEqual(result.created, 1)
        self.assertEqual(result.updated, 0)
        farmer = JawabuFarmerMaster.objects.get()
        self.assertEqual(farmer.customer_name, 'MARY NJERI')
        self.assertEqual(farmer.national_id, '1382654')
        self.assertEqual(farmer.primary_phone, '254720570031')
        self.assertEqual(farmer.secondary_phone, '254785116424')
        self.assertEqual(farmer.county, 'EMBU')
        self.assertEqual(farmer.sub_county, 'MANYATTA')
        self.assertEqual(farmer.branch, 'EMBU')
        self.assertEqual(farmer.status, 'active')
        self.assertIn('ID:1382654', farmer.duplicate_key)

    def test_jawabu_farmers_csv_import_upserts_existing_duplicate_key(self):
        from core.services.jawabu_master import import_jawabu_farmers_csv

        first = StringIO(
            "Name,National ID,Mobile,County\n"
            "Mary Njeri,1382654,0720570031,Embu\n"
        )
        second = StringIO(
            "Name,National ID,Mobile,County,Sub County\n"
            "Mary Njeri,1382654,0720570031,Muranga,Kiharu\n"
        )

        import_jawabu_farmers_csv(first, source_name='first.csv')
        result = import_jawabu_farmers_csv(second, source_name='second.csv')

        self.assertEqual(result.created, 0)
        self.assertEqual(result.updated, 1)
        self.assertEqual(JawabuFarmerMaster.objects.count(), 1)
        farmer = JawabuFarmerMaster.objects.get()
        self.assertEqual(farmer.county, 'MURANGA')
        self.assertEqual(farmer.sub_county, 'KIHARU')
        self.assertEqual(farmer.source_name, 'second.csv')

    def test_jawabu_farmers_csv_import_keeps_incomplete_rows_for_review(self):
        from core.services.jawabu_master import import_jawabu_farmers_csv

        csv_file = StringIO(
            "Farmer Name,County,Location\n"
            "Unknown Farmer,Kisumu,Near market\n"
        )

        result = import_jawabu_farmers_csv(csv_file, source_name='review.csv')

        self.assertEqual(result.created, 1)
        self.assertEqual(result.review_needed, 1)
        farmer = JawabuFarmerMaster.objects.get()
        self.assertEqual(farmer.customer_name, 'UNKNOWN FARMER')
        self.assertEqual(farmer.status, 'review_needed')
        self.assertIn('Missing National ID and primary phone', farmer.cleaning_notes)

    def test_jawabu_farmers_csv_import_applies_confirmed_master_mapping(self):
        from core.services.jawabu_master import import_jawabu_farmers_csv

        csv_file = StringIO(
            "Full Name,ID NUMBER,HBG Hub,Mobile,Phone,Actual Receipts,Sign Date,Sign Date,Created Date,HBG Contract Name\n"
            "David Mugambi [23215888],,Embu,+254721997481,+254704408281,5000,01/05/2026,24/06/2026,30/06/2026,HBGC-14560\n"
        )

        result = import_jawabu_farmers_csv(csv_file, source_name='farmers.csv')

        self.assertEqual(result.created, 1)
        self.assertEqual(result.review_needed, 0)
        farmer = JawabuFarmerMaster.objects.get()
        self.assertEqual(farmer.customer_name, 'DAVID MUGAMBI')
        self.assertEqual(farmer.national_id, '23215888')
        self.assertEqual(farmer.primary_phone, '254721997481')
        self.assertEqual(farmer.secondary_phone, '254704408281')
        self.assertEqual(farmer.county, 'EMBU')
        self.assertEqual(farmer.actual_receipts, '5000')
        self.assertEqual(farmer.sign_date, '24-June-2026')
        self.assertEqual(farmer.created_date, '')
        self.assertEqual(farmer.hbg_contract_name, '')
        self.assertEqual(farmer.external_id, '')
        self.assertEqual(farmer.status, 'active')
class ParsedMessageModelTest(TestCase):
    """Test the ParsedMessage model."""
    
    def test_to_sheet_row(self):
        """Test conversion to Google Sheet row format."""
        raw_message = RawMessage.objects.create(
            telegram_message_id='test_sheet',
            sender='Test',
            content='Test content',
        )
        msg_hash = generate_message_hash('Test', 'Test content')
        processed = ProcessedMessage.objects.create(
            message_hash=msg_hash,
            raw_message=raw_message,
        )
        
        parsed = ParsedMessage.objects.create(
            processed_message=processed,
            message_id='MSG_TEST_SHEET',
            timestamp=timezone.now(),
            sender='John',
            raw_message='Sold 3 bread 50 each',
            customer_name='Jane Doe',
            customer_phone='0712345678',
            customer_id='A12345',
            branch_region='Nairobi',
            complaint_category='System Underperformance',
            complaint_description='No gas supply',
            gps_link='https://maps.example.com/xyz',
            image_flag=True,
            source='whatsapp_telegram',
        )
        
        row = parsed.to_sheet_row()
        
        # Verify 21-column production schema
        self.assertEqual(len(row), 21, f"Expected 21 columns, got {len(row)}")
        
        # [0-1] System/control fields
        self.assertEqual(row[0], '', "Complaint ID (bot leaves blank for FORMULA field)")
        self.assertEqual(row[1], 'MSG_TEST_SHEET', "message_id dedup key")
        
        # [2-9] Bot intake fields
        self.assertIsInstance(row[2], str)  # Date Reported
        self.assertEqual(row[3], 'JANE DOE', "Customer Name (CAPITALIZED by bot)")
        self.assertEqual(row[4], 'A12345', "Customer ID")
        self.assertEqual(row[5], '254712345678', "Phone Number")
        self.assertEqual(row[6], 'John', "JBL Reported By uses the message sender")
        self.assertEqual(row[7], 'Nairobi', "Branch / Region")
        self.assertEqual(row[8], 'System Underperformance', "Complaint Category")
        self.assertEqual(row[9], 'No gas supply', "Complaint Description")
        
        # [10-13] Raw data / Audit trail
        self.assertEqual(row[10], 'Sold 3 bread 50 each', "raw_message")
        self.assertEqual(row[11], 'https://maps.example.com/xyz', "gps_link")
        self.assertEqual(row[12], 'TRUE', "image_flag as TRUE")
        self.assertEqual(row[13], 'whatsapp_telegram', "source")
        
        # [14-20] Human workflow fields (should be empty for bot-generated row)
        self.assertEqual(row[14], '', "Loan Status (human)")
        self.assertEqual(row[15], '', "Loan at Risk (human)")
        self.assertEqual(row[16], '', "Risk Level (human)")
        self.assertEqual(row[17], '', "Status (human)")
        self.assertEqual(row[18], '', "Resolution Details (human)")
        self.assertEqual(row[19], '', "Date Resolved (human)")
        self.assertEqual(row[20], '', "Days Open (formula - should be empty)")

