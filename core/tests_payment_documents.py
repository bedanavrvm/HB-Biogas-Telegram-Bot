from datetime import date, timedelta
from decimal import Decimal
import io
import json
import tempfile
from unittest.mock import patch

from django.core.files import File
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.models import (
    InvoiceUploadBatch,
    JawabuFarmerMaster,
    ParsedInvoice,
    ParsedInvoiceEvent,
    PaymentDocument,
    PaymentDocumentTemplate,
    Product,
    ProductAlias,
    RequisitionBatch,
)
from core.services.invoice_parser import ingest_invoice_upload_batch
from core.services.payment_documents import (
    create_payment_document,
    generate_payment_workbook,
    payment_readiness,
    payment_template_layout,
)


@override_settings(
    ALLOWED_HOSTS=['testserver'],
    PORTAL_WEBAPP_REQUIRE_TELEGRAM_AUTH=False,
    SECURE_SSL_REDIRECT=False,
)
class InvoicePoolAndPaymentDocumentTests(TestCase):
    def setUp(self):
        # The payment fixture uses a historical external product spelling.
        # Register it explicitly so payment readiness exercises the canonical
        # catalogue path instead of an unresolved mapping issue.
        ProductAlias.objects.get_or_create(
            product=Product.objects.get(code='biogas'),
            normalized_alias='biogas_premium',
            defaults={'alias': 'BIOGAS PREMIUM'},
        )

    def farmer(self, **overrides):
        data = {
            'customer_name': 'Mary Wanjiku',
            'national_id': '12345678',
            'primary_phone': '254712345678',
            'secondary_phone': '254700000001',
            'branch': 'Nakuru',
            'jbl_officer': 'Officer Jane',
            'final_decision': 'Approved',
            'customer_no': '15357',
            'imab_customer_name': 'MARY WANJIKU',
            'system_branch': 'Nakuru Branch',
            'system_loan_officer': 'Officer Jane',
            'system_deposit_paid_jbl': Decimal('0'),
            'requisition_date': date(2026, 7, 23),
            'order_number': 'ORDER-001',
            'invoice_number': '9505',
            'invoice_date': date(2026, 7, 20),
            'invoice_amount': Decimal('54000'),
            'discount': Decimal('4500'),
            'payment': Decimal('6000'),
            'balance_due': Decimal('43500'),
            'actual_receipts': '6000',
            'lead_source': 'HomeBiogas',
            'repayment_date': '10TH',
            'repayment_tenor': '6',
            'payment_product': 'BIOGAS PREMIUM',
        }
        data.update(overrides)
        return JawabuFarmerMaster.objects.create(**data)

    def invoice_batch(self, farmer=None):
        batch = InvoiceUploadBatch.objects.create(
            original_filename='invoices.pdf',
            content_type='application/pdf',
            size=100,
            uploaded_by='Tester',
            drive_file_id='drive-pdf',
            drive_url='https://drive.test/invoices',
            status='parsed',
            total_pages=1,
            total_parsed=1,
            unmatched_count=1,
        )
        ParsedInvoice.objects.create(
            batch=batch,
            page=1,
            invoice_no='9505',
            invoice_date=date(2026, 7, 20),
            customer_name='Mary Wanjiku',
            customer_id='12345678',
            customer_phone='254712345678',
            invoice_amount=Decimal('54000'),
            discount=Decimal('4500'),
            payment=Decimal('6000'),
            balance_due=Decimal('43500'),
            status='matched' if farmer else 'unmatched',
            matched_farmer=farmer,
            matched_order_number=farmer.order_number if farmer else '',
        )
        return batch

    @patch('core.services.invoice_parser.parse_invoice_pdf_bytes')
    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_invoice_pool_upload_stores_drive_batch_and_parsed_rows(self, storage, parse_pdf):
        storage.return_value.upload.return_value = ('drive-id', 'https://drive.test/pdf')
        parse_pdf.return_value = ([
            {
                'page': 1,
                'invoice_no': '9505',
                'invoice_date': '20/07/2026',
                'customer_name': 'Mary Wanjiku',
                'customer_id': '12345678',
                'customer_phone': '254712345678',
                'invoice_amount': '54,000.00',
                'total_after_discount': '49,500.00',
                'discount': '4,500.00',
                'payment': '6,000.00',
                'balance_due': '43,500.00',
                'balance_due_check': 'OK',
            }
        ], 1)

        batch = ingest_invoice_upload_batch(
            pdf_bytes=b'%PDF-1.4',
            filename='invoices.pdf',
            uploaded_by='Tester',
        )

        self.assertEqual(batch.drive_file_id, 'drive-id')
        self.assertEqual(batch.total_parsed, 1)
        self.assertEqual(batch.unmatched_count, 1)
        parsed = batch.invoices.get()
        self.assertEqual(parsed.invoice_no, '9505')
        self.assertEqual(parsed.status, 'draft')
        event = parsed.events.get()
        self.assertEqual(event.action, 'parsed')
        self.assertEqual(event.actor, 'Tester')
        self.assertEqual(event.metadata['drive_file_id'], 'drive-id')

    @patch('core.services.invoice_parser.parse_invoice_pdf_bytes')
    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_invoice_upload_request_id_is_idempotent(self, storage, parse_pdf):
        storage.return_value.upload.return_value = ('drive-id', 'https://drive.test/pdf')
        parse_pdf.return_value = ([], 0)

        first = ingest_invoice_upload_batch(
            pdf_bytes=b'%PDF-1.4',
            filename='invoices.pdf',
            uploaded_by='Tester',
            order_number='ORDER-001',
            client_request_id='retry-123',
        )
        second = ingest_invoice_upload_batch(
            pdf_bytes=b'%PDF-1.4-retried',
            filename='invoices.pdf',
            uploaded_by='Tester',
            order_number='ORDER-001',
            client_request_id='retry-123',
        )

        self.assertEqual(first.pk, second.pk)
        self.assertEqual(InvoiceUploadBatch.objects.count(), 1)
        storage.return_value.upload.assert_called_once()
        parse_pdf.assert_called_once()

    @patch('core.services.invoice_parser.parse_invoice_pdf_bytes')
    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_invoice_upload_request_id_cannot_be_reused_for_another_order(self, storage, parse_pdf):
        storage.return_value.upload.return_value = ('drive-id', 'https://drive.test/pdf')
        parse_pdf.return_value = ([], 0)

        ingest_invoice_upload_batch(
            pdf_bytes=b'%PDF-1.4',
            filename='invoices.pdf',
            order_number='ORDER-001',
            client_request_id='retry-123',
        )

        with self.assertRaisesMessage(ValueError, 'already used for another order'):
            ingest_invoice_upload_batch(
                pdf_bytes=b'%PDF-1.4',
                filename='invoices.pdf',
                order_number='ORDER-002',
                client_request_id='retry-123',
            )

    @patch('core.services.invoice_parser.parse_invoice_pdf_bytes')
    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_invoice_pool_upload_accepts_multiple_pdfs(self, storage, parse_pdf):
        storage.return_value.upload.side_effect = [
            ('drive-id-1', 'https://drive.test/pdf-1'),
            ('drive-id-2', 'https://drive.test/pdf-2'),
        ]
        parse_pdf.side_effect = [
            ([{
                'page': 1,
                'invoice_no': '9505',
                'invoice_date': '20/07/2026',
                'customer_name': 'Mary Wanjiku',
                'customer_id': '12345678',
                'customer_phone': '254712345678',
                'invoice_amount': '54,000.00',
                'payment': '6,000.00',
                'balance_due': '43,500.00',
            }], 1),
            ([{
                'page': 1,
                'invoice_no': '9506',
                'invoice_date': '21/07/2026',
                'customer_name': 'John Kamau',
                'customer_id': '87654321',
                'customer_phone': '254722222222',
                'invoice_amount': '89,900.00',
                'payment': '10,000.00',
                'balance_due': '79,900.00',
            }], 1),
        ]

        response = self.client.post(reverse('portal_invoice_pool_upload'), {
            'file': [
                SimpleUploadedFile('invoice-1.pdf', b'%PDF-1.4 one', content_type='application/pdf'),
                SimpleUploadedFile('invoice-2.pdf', b'%PDF-1.4 two', content_type='application/pdf'),
            ],
        })
        data = response.json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(data['ok'])
        self.assertEqual(data['total_uploaded'], 2)
        self.assertEqual(data['total_parsed'], 2)
        self.assertEqual(data['unmatched_count'], 2)
        self.assertEqual(len(data['invoice_batch_ids']), 2)
        self.assertEqual(InvoiceUploadBatch.objects.count(), 2)
        self.assertEqual(ParsedInvoice.objects.count(), 2)

    def test_invoice_pool_endpoint_lists_batches_and_invoices_with_filters(self):
        farmer = self.farmer(order_number='ORDER-MATCHED')
        batch = self.invoice_batch(farmer)
        unmatched_batch = self.invoice_batch()

        response = self.client.get(reverse('portal_invoice_pool'), {'status': 'matched', 'search': '9505'})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['summary']['batch_count'], 2)
        self.assertEqual(data['summary']['invoice_count'], 2)
        self.assertEqual(data['summary']['matched_count'], 1)
        self.assertEqual(data['summary']['unmatched_count'], 1)
        self.assertEqual(len(data['invoices']), 1)
        self.assertEqual(data['invoices'][0]['status'], 'matched')
        self.assertEqual(data['invoices'][0]['matched_order_number'], 'ORDER-MATCHED')
        self.assertEqual(data['invoices'][0]['payment_readiness']['ready_count'], 1)
        self.assertEqual(data['invoices'][0]['payment_readiness']['blocked_count'], 0)
        self.assertEqual(data['invoices'][0]['duplicate_count'], 1)
        batch_ids = {item['id'] for item in data['batches']}
        self.assertIn(str(batch.id), batch_ids)
        self.assertIn(str(unmatched_batch.id), batch_ids)

    def test_invoice_pool_review_filter_returns_duplicates(self):
        farmer = self.farmer(order_number='ORDER-MATCHED')
        batch = self.invoice_batch(farmer)
        duplicate_batch = self.invoice_batch()

        response = self.client.get(reverse('portal_invoice_pool'), {'review': 'duplicates'})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['ok'])
        invoice_ids = {item['id'] for item in data['invoices']}
        self.assertEqual(invoice_ids, {
            str(batch.invoices.get().id),
            str(duplicate_batch.invoices.get().id),
        })

    def test_invoice_pool_review_filter_returns_payment_ready_and_blocked(self):
        ready_farmer = self.farmer(order_number='ORDER-READY')
        self.invoice_batch(ready_farmer)
        blocked_farmer = self.farmer(
            order_number='ORDER-BLOCKED',
            customer_name='Blocked Customer',
            national_id='88887777',
            primary_phone='254722222222',
            repayment_date='',
            repayment_tenor='',
        )
        self.invoice_batch(blocked_farmer)

        ready_response = self.client.get(reverse('portal_invoice_pool'), {'review': 'payment_ready'})
        blocked_response = self.client.get(reverse('portal_invoice_pool'), {'review': 'payment_blocked'})

        self.assertEqual(ready_response.status_code, 200)
        self.assertEqual(blocked_response.status_code, 200)
        ready_orders = {item['matched_order_number'] for item in ready_response.json()['invoices']}
        blocked_orders = {item['matched_order_number'] for item in blocked_response.json()['invoices']}
        self.assertIn('ORDER-READY', ready_orders)
        self.assertNotIn('ORDER-BLOCKED', ready_orders)
        self.assertIn('ORDER-BLOCKED', blocked_orders)
        self.assertNotIn('ORDER-READY', blocked_orders)

    def test_invoice_farmer_candidate_search(self):
        farmer = self.farmer(customer_name='Searchable Client', national_id='99887766')

        response = self.client.get(reverse('portal_invoice_farmer_candidates'), {'search': '99887766'})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['farmers'][0]['id'], str(farmer.id))
        self.assertTrue(data['farmers'][0]['has_invoice'])
        self.assertIn('Existing invoice', data['farmers'][0]['invoice_conflict_label'])

    def test_invoice_farmer_candidates_are_ranked_from_invoice_identity(self):
        id_match = self.farmer(customer_name='Different Name', national_id='11112222', primary_phone='254700000100')
        phone_match = self.farmer(customer_name='Other Name', national_id='33334444', primary_phone='254799888777')
        batch = self.invoice_batch()
        invoice = batch.invoices.get()
        invoice.customer_id = '11112222'
        invoice.customer_phone = '0799888777'
        invoice.save(update_fields=['customer_id', 'customer_phone', 'updated_at'])

        response = self.client.get(reverse('portal_invoice_farmer_candidates'), {
            'invoice_id': str(invoice.id),
            'search': 'different',
        })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['farmers'][0]['id'], str(id_match.id))
        self.assertIn('ID match', data['farmers'][0]['match_reasons'])
        self.assertEqual(data['farmers'][0]['match_tier'], 'likely')
        self.assertEqual(data['candidate_scope'], 'operational')
        ids = {farmer['id'] for farmer in data['farmers']}
        self.assertIn(str(phone_match.id), ids)

    def test_historical_applicant_search_is_explicit_and_labelled(self):
        farmer = self.farmer(
            customer_name='Historical Applicant', national_id='44556677', status='inactive',
        )
        JawabuFarmerMaster.objects.filter(pk=farmer.pk).update(
            updated_at=timezone.now() - timedelta(days=500),
        )

        operational = self.client.get(
            reverse('portal_invoice_farmer_candidates'), {'search': '44556677'},
        ).json()
        historical = self.client.get(
            reverse('portal_invoice_farmer_candidates'), {'search': '44556677', 'scope': 'historical'},
        ).json()

        self.assertEqual(operational['farmers'], [])
        self.assertTrue(historical['historical_search'])
        self.assertEqual(historical['farmers'][0]['id'], str(farmer.id))
        self.assertEqual(historical['farmers'][0]['candidate_scope'], 'historical')

    def test_invoice_detail_exposes_audit_events_and_duplicates(self):
        farmer = self.farmer(order_number='ORDER-MATCHED')
        batch = self.invoice_batch(farmer)
        invoice = batch.invoices.get()
        duplicate_batch = self.invoice_batch()
        duplicate = duplicate_batch.invoices.get()
        duplicate.customer_id = invoice.customer_id
        duplicate.customer_phone = invoice.customer_phone
        duplicate.save(update_fields=['customer_id', 'customer_phone', 'updated_at'])
        ParsedInvoiceEvent.objects.create(
            invoice=invoice,
            action='matched',
            actor='Tester',
            note='Confirmed from invoice detail test',
        )

        response = self.client.get(reverse('portal_invoice_detail', args=[str(invoice.id)]))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['invoice']['id'], str(invoice.id))
        self.assertEqual(data['batch']['id'], str(batch.id))
        self.assertEqual(data['source_pdf_url'], batch.drive_url)
        self.assertEqual(data['events'][0]['action'], 'matched')
        duplicate_ids = {item['id'] for item in data['duplicates']}
        self.assertIn(str(duplicate.id), duplicate_ids)
        reasons = data['duplicates'][0]['duplicate_reasons']
        self.assertTrue({'Same invoice no', 'Same ID', 'Same phone'} & set(reasons))

    @patch('core.services.invoice_parser.reserve_farmer_publication')
    def test_manual_invoice_match_endpoint_links_invoice_to_farmer(self, mock_reserve_publication):
        farmer = self.farmer(
            order_number='ORDER-MANUAL',
            invoice_number='',
            invoice_date=None,
            invoice_amount=None,
            discount=None,
            payment=None,
            balance_due=None,
        )
        batch = self.invoice_batch()
        invoice = batch.invoices.get()

        response = self.client.post(
            reverse('portal_invoice_match', args=[str(invoice.id)]),
            data=json.dumps({'farmer_id': str(farmer.id), 'note': 'Verified by phone'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        invoice.refresh_from_db()
        farmer.refresh_from_db()
        self.assertEqual(invoice.status, 'matched')
        self.assertEqual(invoice.matched_farmer_id, farmer.id)
        self.assertEqual(invoice.matched_order_number, 'ORDER-MANUAL')
        self.assertEqual(farmer.invoice_number, '9505')
        self.assertEqual(farmer.balance_due, Decimal('43500'))
        self.assertIn('Verified by phone', invoice.review_notes)
        event = invoice.events.filter(action='matched').latest('created_at')
        self.assertEqual(event.note, 'Verified by phone')
        self.assertEqual(event.metadata['order_number'], 'ORDER-MANUAL')
        # Invoice reconciliation commits canonical Django state first; the
        # register publication is a durable follow-up rather than a request
        # blocking Google Sheets write.
        mock_reserve_publication.assert_called_once()
        self.assertEqual(mock_reserve_publication.call_args.args[0].pk, farmer.pk)
        self.assertEqual(
            mock_reserve_publication.call_args.kwargs['request_id'],
            f'invoice-match:{invoice.id}:{farmer.id}',
        )

    @patch('core.services.invoice_parser.reserve_farmer_publication')
    def test_manual_invoice_unmatch_endpoint_clears_linked_farmer_invoice_fields(self, mock_reserve_publication):
        farmer = self.farmer(order_number='ORDER-MATCHED')
        batch = self.invoice_batch(farmer)
        invoice = batch.invoices.get()

        response = self.client.post(
            reverse('portal_invoice_unmatch', args=[str(invoice.id)]),
            data=json.dumps({'note': 'Wrong household'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        invoice.refresh_from_db()
        farmer.refresh_from_db()
        self.assertEqual(invoice.status, 'unmatched')
        self.assertIsNone(invoice.matched_farmer)
        self.assertEqual(invoice.matched_order_number, '')
        self.assertEqual(farmer.invoice_number, '')
        self.assertIsNone(farmer.balance_due)
        self.assertIn('Wrong household', invoice.review_notes)
        self.assertTrue(invoice.events.filter(action='unmatched', note='Wrong household').exists())
        mock_reserve_publication.assert_called_once()
        self.assertEqual(mock_reserve_publication.call_args.args[0].pk, farmer.pk)
        self.assertEqual(
            mock_reserve_publication.call_args.kwargs['request_id'],
            f'invoice-unmatch:{invoice.id}:{farmer.id}',
        )

    def test_manual_invoice_ignore_endpoint_marks_invoice_ignored(self):
        batch = self.invoice_batch()
        invoice = batch.invoices.get()

        response = self.client.post(
            reverse('portal_invoice_ignore', args=[str(invoice.id)]),
            data=json.dumps({'note': 'Duplicate PDF page'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        invoice.refresh_from_db()
        batch.refresh_from_db()
        self.assertEqual(invoice.status, 'ignored')
        self.assertIn('Duplicate PDF page', invoice.review_notes)
        self.assertTrue(invoice.events.filter(action='ignored', note='Duplicate PDF page').exists())
        self.assertEqual(batch.unmatched_count, 0)

    def test_manual_invoice_restore_endpoint_moves_ignored_invoice_back_to_unmatched(self):
        batch = self.invoice_batch()
        invoice = batch.invoices.get()
        invoice.status = 'ignored'
        invoice.review_notes = 'Ignored earlier'
        invoice.save(update_fields=['status', 'review_notes', 'updated_at'])

        response = self.client.post(
            reverse('portal_invoice_restore', args=[str(invoice.id)]),
            data=json.dumps({'note': 'Needs review again'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        invoice.refresh_from_db()
        batch.refresh_from_db()
        self.assertEqual(invoice.status, 'unmatched')
        self.assertIn('Needs review again', invoice.review_notes)
        self.assertTrue(invoice.events.filter(action='restored', note='Needs review again').exists())
        self.assertEqual(batch.unmatched_count, 1)

    def test_bulk_invoice_ignore_skips_matched_invoices(self):
        matched_farmer = self.farmer(order_number='ORDER-MATCHED')
        matched_batch = self.invoice_batch(matched_farmer)
        matched_invoice = matched_batch.invoices.get()
        review_batch = self.invoice_batch()
        review_invoice = review_batch.invoices.get()

        response = self.client.post(
            reverse('portal_invoice_bulk_action'),
            data=json.dumps({
                'action': 'ignore',
                'invoice_ids': [str(matched_invoice.id), str(review_invoice.id)],
                'note': 'Batch cleanup',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['changed_count'], 1)
        self.assertEqual(data['skipped_count'], 1)
        matched_invoice.refresh_from_db()
        review_invoice.refresh_from_db()
        self.assertEqual(matched_invoice.status, 'matched')
        self.assertEqual(review_invoice.status, 'ignored')
        self.assertTrue(review_invoice.events.filter(action='ignored', note='Batch cleanup').exists())

    def test_bulk_invoice_restore_only_restores_ignored_invoices(self):
        ignored_batch = self.invoice_batch()
        ignored_invoice = ignored_batch.invoices.get()
        ignored_invoice.status = 'ignored'
        ignored_invoice.save(update_fields=['status', 'updated_at'])
        unmatched_batch = self.invoice_batch()
        unmatched_invoice = unmatched_batch.invoices.get()

        response = self.client.post(
            reverse('portal_invoice_bulk_action'),
            data=json.dumps({
                'action': 'restore',
                'invoice_ids': [str(ignored_invoice.id), str(unmatched_invoice.id)],
                'note': 'Back to review',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['changed_count'], 1)
        self.assertEqual(data['skipped_count'], 1)
        ignored_invoice.refresh_from_db()
        unmatched_invoice.refresh_from_db()
        self.assertEqual(ignored_invoice.status, 'unmatched')
        self.assertEqual(unmatched_invoice.status, 'unmatched')
        self.assertTrue(ignored_invoice.events.filter(action='restored', note='Back to review').exists())

    def test_payment_template_layout_uses_visible_sheet_when_config_is_stale(self):
        workbook = load_workbook('requisition/HB_PAYMENT__89__7__machine_ready (1).xlsx')
        layout = payment_template_layout(workbook)

        self.assertEqual(layout.sheet_name, '#89')
        self.assertEqual(layout.header_row, 7)
        self.assertEqual(layout.data_start_row, 8)
        self.assertEqual(layout.totals_row, 12)
        self.assertEqual(layout.columns['cust_no'], 5)
        self.assertIn('header_row config=5 visible=7', layout.config_warnings)

    def test_same_named_payment_template_keeps_only_latest_active(self):
        first = PaymentDocumentTemplate.objects.create(name='HB Payment Document', is_active=True)
        latest = PaymentDocumentTemplate.objects.create(name='hb payment document', is_active=True)

        first.refresh_from_db()
        latest.refresh_from_db()
        self.assertFalse(first.is_active)
        self.assertTrue(latest.is_active)

    def test_payment_workbook_generation_uses_active_admin_uploaded_template(self):
        farmer = self.farmer(order_number='ORDER-UPLOADED')
        self.invoice_batch(farmer)

        with tempfile.TemporaryDirectory() as media_root:
            with self.settings(MEDIA_ROOT=media_root):
                template = PaymentDocumentTemplate.objects.create(
                    name='Uploaded payment template',
                    is_active=True,
                )
                with open('requisition/HB_PAYMENT__89__7__machine_ready (1).xlsx', 'rb') as handle:
                    template.file.save('uploaded_payment_template.xlsx', File(handle), save=True)

                xlsx, summary = generate_payment_workbook('ORDER-UPLOADED', '91')

        self.assertTrue(xlsx)
        self.assertEqual(summary['ready_count'], 1)

    def test_payment_workbook_generation_uses_drive_backed_template_when_local_file_is_missing(self):
        farmer = self.farmer(order_number='ORDER-DRIVE-TEMPLATE')
        self.invoice_batch(farmer)
        template = PaymentDocumentTemplate.objects.create(
            name='Drive payment template',
            is_active=True,
            drive_file_id='drive-template-id',
        )
        template_bytes = open('requisition/HB_PAYMENT__89__7__machine_ready (1).xlsx', 'rb').read()

        with patch(
            'core.services.payment_documents.workbook_source_from_template',
            return_value=io.BytesIO(template_bytes),
        ) as source:
            xlsx, summary = generate_payment_workbook('ORDER-DRIVE-TEMPLATE', '92')

        source.assert_called_once()
        self.assertTrue(xlsx)
        self.assertEqual(summary['ready_count'], 1)

    def test_payment_readiness_blocks_missing_repayment_terms(self):
        farmer = self.farmer(repayment_date='', repayment_tenor='')
        self.invoice_batch(farmer)

        readiness = payment_readiness('ORDER-001')

        self.assertEqual(readiness['ready_count'], 0)
        self.assertEqual(readiness['blocked_count'], 1)
        self.assertIn('Repayment Dates', readiness['blocked'][0]['missing'])
        self.assertIn('Tenor', readiness['blocked'][0]['missing'])

    def test_payment_readiness_blocks_missing_invoice_balance_due(self):
        farmer = self.farmer(balance_due=None)
        self.invoice_batch(farmer)

        readiness = payment_readiness('ORDER-001')

        self.assertEqual(readiness['ready_count'], 0)
        self.assertIn('Balance Due', readiness['blocked'][0]['missing'])

    def test_payment_col_is_separate_from_order_call_up_comment(self):
        farmer = self.farmer(final_decision_comment='Approved for order after customer call.')
        self.invoice_batch(farmer)

        readiness = payment_readiness('ORDER-001')

        self.assertEqual(readiness['ready'][0]['row']['call_up_comments'], '')
        self.assertEqual(
            readiness['ready'][0]['order_call_up_comments'],
            'Approved for order after customer call.',
        )

        xlsx, _summary = generate_payment_workbook('ORDER-001', '106')
        workbook = load_workbook(io.BytesIO(xlsx), data_only=False)
        layout = payment_template_layout(workbook)
        ws = workbook['#106']
        self.assertIn(ws.cell(row=layout.data_start_row, column=layout.columns['call_up_comments']).value, (None, ''))

    def test_payment_workbook_generation_uses_ready_farmer_and_preserves_signatures(self):
        farmer = self.farmer()
        self.invoice_batch(farmer)

        xlsx, summary = generate_payment_workbook('ORDER-001', '105')
        path = 'tmp_payment_output.xlsx'
        self.addCleanup(lambda: __import__('pathlib').Path(path).exists() and __import__('pathlib').Path(path).unlink())
        with open(path, 'wb') as handle:
            handle.write(xlsx)
        workbook = load_workbook(path, data_only=False)
        ws = workbook['#105']

        self.assertEqual(ws['H4'].value, '#105')
        self.assertEqual(summary['payment_number'], '105')

        self.assertEqual(summary['ready_count'], 1)
        self.assertEqual(ws['C8'].value.date(), date(2026, 7, 23))
        self.assertEqual(ws['D8'].value, 'ORDER-001')
        self.assertEqual(ws['E8'].value, '15357')
        self.assertEqual(ws['G8'].value, 'MARY WANJIKU')
        self.assertEqual(ws['H8'].value, 'Mary Wanjiku')
        self.assertEqual(ws['K8'].value, 'Nakuru Branch')
        self.assertEqual(ws['L8'].value, 'Officer Jane')
        self.assertEqual(ws['M8'].value, 43500)
        self.assertEqual(ws['M8'].number_format, '0')
        self.assertIn(ws['N8'].value, (None, ''))
        self.assertEqual(ws['O8'].value, 4500)
        self.assertEqual(ws['O8'].number_format, '0')
        self.assertEqual(ws['P8'].value, 6000)
        self.assertIsNone(ws['R8'].value)
        self.assertEqual(ws['S8'].value, '10TH')
        self.assertEqual(ws['T8'].value, '6')
        self.assertEqual(ws['U8'].value, 'BIOGAS PREMIUM')
        self.assertIn(ws.cell(row=summary['totals_row'], column=14).value, (None, ''))
        prepared_rows = [row for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=3).value == 'PREPARED BY:']
        self.assertTrue(prepared_rows)
        self.assertGreater(prepared_rows[0], summary['totals_row'])

    def test_payment_preview_data_returns_rows_without_generating_workbook(self):
        farmer = self.farmer()
        self.invoice_batch(farmer)

        with patch('core.services.payment_documents.generate_payment_workbook') as generate:
            response = self.client.get(reverse('portal_payment_preview_data', args=['ORDER-001']), {'payment_number': '89'})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['preview']['rows'][0]['cust_no'], '15357')
        self.assertEqual(data['preview']['rows'][0]['hb_invoice_amount'], '43500.00')
        self.assertIsNone(data['preview']['rows'][0]['expected_invoice_amount'])
        self.assertEqual(data['preview']['payment_number'], '89')
        self.assertNotIn('workbook_preview', data['preview'])
        generate.assert_not_called()

    def test_payment_preview_requires_numeric_payment_number(self):
        self.farmer()

        missing = self.client.get(reverse('portal_payment_preview_data', args=['ORDER-001']))
        invalid = self.client.get(
            reverse('portal_payment_preview_data', args=['ORDER-001']),
            {'payment_number': 'PAY/89'},
        )

        self.assertEqual(missing.status_code, 400)
        self.assertEqual(missing.json()['error'], 'Payment number is required.')
        self.assertEqual(invalid.status_code, 400)
        self.assertIn('digits only', invalid.json()['error'])

    def test_document_history_lists_and_opens_final_payment_snapshot(self):
        doc = PaymentDocument.objects.create(
            order_number='ORDER-001', payment_number='89', status='final', row_count=1,
            filename='HB_Payment_89_ORDER-001_final_v1.xlsx',
            validation_summary={'preview_rows': [{
                'name': 'Mary Wanjiku', 'hb_invoice_amount': '43500.00',
                'discount': '4500.00', 'deposit_paid_hbg': '6000.00',
                'deposit_paid_jbl': None,
            }]},
        )

        history = self.client.get(reverse('portal_document_history'), {'kind': 'payments'})
        detail = self.client.get(reverse('portal_payment_document_detail', args=[str(doc.id)]))

        self.assertEqual(history.status_code, 200)
        self.assertEqual(history.json()['documents'][0]['payment_number'], '89')
        self.assertEqual(history.json()['documents'][0]['version'], 1)
        self.assertEqual(history.json()['documents'][0]['filename'], 'HB_Payment_89_ORDER-001_final_v1.xlsx')
        self.assertEqual(
            history.json()['documents'][0]['workbook_generated_at'],
            history.json()['documents'][0]['generated_at'],
        )
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(detail.json()['preview']['rows'][0]['hb_invoice_amount'], '43500.00')
        self.assertEqual(detail.json()['preview']['totals']['hb_invoice_amount'], '43500.00')

    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_payment_document_history_can_regenerate_a_new_review_snapshot(self, storage):
        farmer = self.farmer()
        self.invoice_batch(farmer)
        source = PaymentDocument.objects.create(
            order_number='ORDER-001', payment_number='89', status='final', version=3,
            row_count=1, farmer_ids=[str(farmer.id)],
            filename='HB_Payment_89_ORDER-001_final_v3.xlsx',
            call_up_comments='Legacy batch comment',
            case_call_up_comments={str(farmer.id): 'Existing case comment'},
            validation_summary={'preview_rows': [{'farmer_id': str(farmer.id)}]},
        )
        storage.return_value.upload.return_value = ('review-xlsx', 'https://drive.test/regenerated-payment')

        response = self.client.post(
            reverse('portal_payment_document_regenerate', args=[str(source.id)]),
            data=json.dumps({}), content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['ok'])
        self.assertTrue(data['requires_head_rural_review'])
        regenerated = PaymentDocument.objects.get(pk=data['document']['id'])
        self.assertEqual(regenerated.status, 'pending_review')
        self.assertEqual(regenerated.version, 4)
        self.assertEqual(regenerated.farmer_ids, [str(farmer.id)])
        self.assertEqual(regenerated.case_call_up_comments, {str(farmer.id): 'Existing case comment'})
        self.assertEqual(regenerated.drive_url, 'https://drive.test/regenerated-payment')

    def test_regenerating_pending_payment_review_is_idempotent(self):
        farmer = self.farmer()
        review = PaymentDocument.objects.create(
            order_number='ORDER-001', payment_number='89', status='pending_review',
            version=2, row_count=1, farmer_ids=[str(farmer.id)],
            filename='HB_Payment_89_ORDER-001_review_v2.xlsx',
        )

        response = self.client.post(
            reverse('portal_payment_document_regenerate', args=[str(review.id)]),
            data=json.dumps({}), content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['idempotent_replay'])
        self.assertEqual(response.json()['document']['id'], str(review.id))
        self.assertEqual(PaymentDocument.objects.filter(order_number='ORDER-001').count(), 1)

    def test_payment_finalize_replays_existing_pending_review(self):
        farmer = self.farmer()
        review = PaymentDocument.objects.create(
            order_number='ORDER-001', payment_number='89', status='pending_review',
            version=2, row_count=1, farmer_ids=[str(farmer.id)],
            filename='HB_Payment_89_ORDER-001_review_v2.xlsx',
        )

        response = self.client.post(
            reverse('portal_payment_document_finalize', args=['ORDER-001']),
            data=json.dumps({'payment_number': '89'}), content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['idempotent_replay'])
        self.assertEqual(response.json()['document']['id'], str(review.id))
        self.assertEqual(PaymentDocument.objects.filter(order_number='ORDER-001').count(), 1)

    def test_batch_detail_uses_live_invoice_count_over_stored_snapshot(self):
        farmer = self.farmer()
        RequisitionBatch.objects.create(
            order_number='ORDER-001',
            requisition_date=date(2026, 7, 23),
            farmer_ids=[str(farmer.id)],
            farmer_count=1,
            invoice_summary={
                'invoiced_count': 0,
                'pending_invoice_count': 1,
                'last_invoice_upload_status': 'success',
                'last_invoice_upload_error': '',
                'invoice_batch_id': 'upload-001',
            },
        )

        response = self.client.get(reverse('portal_requisition_batch_detail', args=['ORDER-001']))

        self.assertEqual(response.status_code, 200)
        summary = response.json()['batch']['invoice_summary']
        self.assertEqual(summary['invoiced_count'], 1)
        self.assertEqual(summary['pending_invoice_count'], 0)
        self.assertEqual(summary['last_invoice_upload_status'], 'success')
        self.assertEqual(summary['invoice_batch_id'], 'upload-001')

    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_payment_preview_endpoint_returns_drive_document(self, storage):
        farmer = self.farmer()
        self.invoice_batch(farmer)
        storage.return_value.upload.return_value = ('drive-xlsx', 'https://drive.test/payment')

        response = self.client.post(
            reverse('portal_payment_document_preview', args=['ORDER-001']),
            data=json.dumps({'payment_number': '89'}), content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['document']['drive_url'], 'https://drive.test/payment')
        self.assertEqual(data['document']['sync_status'], 'succeeded')
        self.assertEqual(data['document']['payment_number'], '89')
        self.assertEqual(PaymentDocument.objects.get().status, 'preview')

    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_payment_generation_creates_review_and_head_rural_approval_creates_final(self, storage):
        farmer = self.farmer()
        self.invoice_batch(farmer)
        storage.return_value.upload.side_effect = [
            ('review-xlsx', 'https://drive.test/review'),
            ('final-xlsx', 'https://drive.test/final'),
        ]

        review = create_payment_document(
            'ORDER-001', '89', actor='Operations', status='pending_review',
        )
        self.assertEqual(review.status, 'pending_review')
        self.assertEqual(review.call_up_comments, '')

        response = self.client.post(
            reverse('portal_payment_document_approve', args=[str(review.id)]),
            data=json.dumps({'call_up_comments': 'Approved for call up on 04-July-2026.'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        final = PaymentDocument.objects.get(status='final')
        review.refresh_from_db()
        self.assertEqual(final.version, 2)
        self.assertEqual(final.drive_url, 'https://drive.test/final')
        self.assertEqual(final.call_up_comments, 'Approved for call up on 04-July-2026.')
        self.assertEqual(review.status, 'reviewed')
        self.assertEqual(review.reviewed_by, '')
        self.assertTrue(farmer.pipeline_events.filter(action='payment_finalized').exists())

        replay = self.client.post(
            reverse('portal_payment_document_approve', args=[str(review.id)]),
            data=json.dumps({'call_up_comments': 'Same retry'}),
            content_type='application/json',
        )
        self.assertEqual(replay.status_code, 200)
        self.assertEqual(replay.json()['document']['id'], str(final.id))
        self.assertEqual(PaymentDocument.objects.filter(status='final').count(), 1)

        xlsx, _summary = generate_payment_workbook(
            'ORDER-001', '89', farmer_ids=[str(farmer.id)],
            call_up_comments=final.call_up_comments,
        )
        workbook = load_workbook(io.BytesIO(xlsx), data_only=False)
        ws = workbook['#89']
        layout = payment_template_layout(workbook)
        self.assertEqual(ws.cell(row=layout.data_start_row, column=layout.columns['call_up_comments']).value, final.call_up_comments)

    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_payment_approval_persists_a_comment_for_each_case(self, storage):
        farmer = self.farmer(final_decision_comment='Order-stage comment')
        self.invoice_batch(farmer)
        storage.return_value.upload.side_effect = [
            ('review-xlsx', 'https://drive.test/review'),
            ('final-xlsx', 'https://drive.test/final'),
        ]
        review = create_payment_document('ORDER-001', '107', status='pending_review')
        comment = 'HOR payment review: release after balance confirmation.'
        response = self.client.post(
            reverse('portal_payment_document_approve', args=[str(review.id)]),
            data=json.dumps({'case_call_up_comments': {str(farmer.id): comment}}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        final = PaymentDocument.objects.get(status='final')
        self.assertEqual(final.case_call_up_comments, {str(farmer.id): comment})
        self.assertEqual(final.call_up_comments, '')
        case_comment = farmer.case_comments.get(stage_key='payment')
        self.assertEqual(case_comment.role_label, 'Head of Rural')
        self.assertEqual(case_comment.comment, comment)
        workbook = load_workbook(io.BytesIO(generate_payment_workbook(
            'ORDER-001', '107', farmer_ids=[str(farmer.id)],
            case_call_up_comments={str(farmer.id): comment},
        )[0]), data_only=False)
        layout = payment_template_layout(workbook)
        self.assertEqual(
            workbook['#107'].cell(row=layout.data_start_row, column=layout.columns['call_up_comments']).value,
            comment,
        )

    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_selected_payment_submission_is_not_final(self, storage):
        farmer = self.farmer()
        self.invoice_batch(farmer)
        storage.return_value.upload.return_value = ('review-xlsx', 'https://drive.test/review')

        response = self.client.post(
            reverse('portal_payment_selection'),
            data=json.dumps({
                'payment_number': '90',
                'farmer_ids': [str(farmer.id)],
                'final': True,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['requires_head_rural_review'])
        self.assertEqual(response.json()['document']['status'], 'pending_review')
        self.assertFalse(PaymentDocument.objects.filter(status='final').exists())

    def test_payment_history_includes_pending_review_for_head_of_rural(self):
        farmer = self.farmer()
        review = PaymentDocument.objects.create(
            order_number='ORDER-001',
            payment_number='91',
            status='pending_review',
            row_count=1,
            farmer_ids=[str(farmer.id)],
            filename='HB_Payment_91_ORDER-001_review_v1.xlsx',
            validation_summary={'preview_rows': [{'name': farmer.customer_name}]},
        )

        history = self.client.get(reverse('portal_document_history'), {'kind': 'payments'})
        detail = self.client.get(reverse('portal_payment_document_detail', args=[str(review.id)]))

        self.assertEqual(history.status_code, 200)
        self.assertEqual(history.json()['documents'][0]['status'], 'pending_review')
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(detail.json()['document']['status'], 'pending_review')

    def test_pending_payment_detail_does_not_expose_legacy_order_comment_as_payment_col(self):
        farmer = self.farmer(final_decision_comment='Order-stage comment')
        review = PaymentDocument.objects.create(
            order_number='ORDER-001',
            payment_number='92',
            status='pending_review',
            row_count=1,
            farmer_ids=[str(farmer.id)],
            validation_summary={'preview_rows': [{
                'farmer_id': str(farmer.id),
                'name': farmer.customer_name,
                'call_up_comments': 'Order-stage comment',
            }]},
            case_call_up_comments={},
        )

        detail = self.client.get(reverse('portal_payment_document_detail', args=[str(review.id)]))

        self.assertEqual(detail.status_code, 200)
        self.assertEqual(detail.json()['preview']['rows'][0]['call_up_comments'], '')

    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_payment_preview_endpoint_does_not_require_csrf_cookie(self, storage):
        farmer = self.farmer()
        self.invoice_batch(farmer)
        storage.return_value.upload.return_value = ('drive-xlsx', 'https://drive.test/payment')
        csrf_client = Client(enforce_csrf_checks=True)

        response = csrf_client.post(
            reverse('portal_payment_document_preview', args=['ORDER-001']),
            data=json.dumps({'payment_number': '89'}), content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])

    def test_payment_drive_failure_leaves_retryable_failed_document(self):
        farmer = self.farmer()
        self.invoice_batch(farmer)

        with patch(
            'core.services.payment_documents._upload_payment_workbook',
            side_effect=RuntimeError('simulated Drive outage'),
        ):
            with self.assertRaises(RuntimeError):
                create_payment_document('ORDER-001', '89', actor='Tester', final=False)

        failed = PaymentDocument.objects.get(order_number='ORDER-001')
        self.assertEqual(failed.status, 'failed')
        self.assertEqual(failed.error, 'Drive upload failed; retry required.')
        self.assertEqual(failed.drive_sync_attempts, 1)
        self.assertIsNotNone(failed.drive_next_retry_at)
        self.assertFalse(failed.drive_url)

    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_failed_payment_document_can_be_retried_without_losing_audit_row(self, storage):
        farmer = self.farmer()
        self.invoice_batch(farmer)
        failed = PaymentDocument.objects.create(
            order_number='ORDER-001', payment_number='89', status='failed', version=2,
            row_count=1, farmer_ids=[str(farmer.id)],
            filename='HB_Payment_89_ORDER-001_preview_v2.xlsx',
            error='Drive upload failed; retry required.',
            validation_summary={'artifact_status': 'preview', 'preview_rows': [{'farmer_id': str(farmer.id)}]},
        )
        storage.return_value.upload.return_value = ('retry-payment', 'https://drive.test/retry-payment')

        response = self.client.post(
            reverse('portal_payment_document_regenerate', args=[str(failed.id)]),
            data=json.dumps({}), content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        replacement = PaymentDocument.objects.get(pk=response.json()['document']['id'])
        self.assertNotEqual(replacement.id, failed.id)
        self.assertEqual(replacement.status, 'preview')
        self.assertEqual(replacement.drive_url, 'https://drive.test/retry-payment')
        self.assertTrue(PaymentDocument.objects.filter(pk=failed.id, status='failed').exists())

    @patch('core.services.order_approval.GoogleDriveMediaStorage')
    def test_reconciliation_replaces_failed_payment_without_requeueing_old_row(self, storage):
        farmer = self.farmer()
        self.invoice_batch(farmer)
        failed = PaymentDocument.objects.create(
            order_number='ORDER-001', payment_number='89', status='failed', version=2,
            row_count=1, farmer_ids=[str(farmer.id)],
            filename='HB_Payment_89_ORDER-001_preview_v2.xlsx',
            error='Drive upload failed; retry required.',
            validation_summary={'artifact_status': 'preview', 'preview_rows': [{'farmer_id': str(farmer.id)}]},
        )
        storage.return_value.upload.return_value = ('retry-payment-2', 'https://drive.test/retry-payment-2')

        from core.services.portal_reconciliation import retry_payment_document
        result = retry_payment_document(failed, actor='system:test')

        self.assertTrue(result['ok'])
        failed.refresh_from_db()
        self.assertFalse(failed.error)
        self.assertIsNone(failed.drive_next_retry_at)
        self.assertIn('reconciliation_note', failed.validation_summary)

    @patch('core.services.payment_documents._upload_payment_workbook')
    def test_repeated_payment_preview_creates_versioned_local_documents(self, upload):
        farmer = self.farmer()
        self.invoice_batch(farmer)
        upload.return_value = ('drive-xlsx', 'https://drive.test/payment')

        first = create_payment_document('ORDER-001', '89', actor='Tester', final=False)
        second = create_payment_document('ORDER-001', '89', actor='Tester', final=False)

        self.assertNotEqual(first.id, second.id)
        self.assertEqual(first.version, 1)
        self.assertEqual(second.version, 2)
        self.assertTrue(second.filename.endswith('_preview_v2.xlsx'))
        self.assertEqual(PaymentDocument.objects.filter(order_number='ORDER-001').count(), 2)

    def test_payment_preview_endpoint_returns_readiness_when_blocked(self):
        self.farmer(repayment_date='')

        response = self.client.post(
            reverse('portal_payment_document_preview', args=['ORDER-001']),
            data=json.dumps({'payment_number': '89'}), content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        data = response.json()
        self.assertFalse(data['ok'])
        self.assertEqual(data['readiness']['blocked_count'], 1)

    def test_individual_requisition_assignment_cannot_set_payment_product(self):
        farmer = self.farmer(order_number='', repayment_date='', repayment_tenor='', payment_product='')

        response = self.client.post(
            reverse('portal_assign_order', args=[str(farmer.id)]),
            data=json.dumps({
                'workflow_revision': farmer.workflow_revision,
                'order_number': 'ORDER-009',
                'requisition_date': '2026-07-23',
                'repayment_date': '15TH',
                'repayment_tenor': '9',
                'payment_product': 'BIOGAS',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 410)
        farmer.refresh_from_db()
        self.assertEqual(farmer.order_number, '')
        self.assertEqual(farmer.repayment_date, '')
        self.assertEqual(farmer.repayment_tenor, '')
        self.assertEqual(farmer.payment_product, '')
