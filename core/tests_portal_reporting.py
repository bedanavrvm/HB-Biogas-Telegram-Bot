"""Focused regression tests for the IT-only controlled Portal reports."""
from __future__ import annotations

from io import BytesIO, StringIO
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook

from core.models import ComplianceAuditEvent, JawabuFarmerMaster, PortalReportChart
from core.services.portal_reporting import (
    PortalReportingError,
    catalogue_payload,
    create_definition,
    export_xlsx,
    preview_chart,
    run_definition,
    validate_charts,
    validate_configuration,
)
from core.services.reporting_relationships import portal_relationship_summary
from core.services.workflow_capabilities import has_capability


class PortalReportingTests(TestCase):
    """Reports remain bounded, scoped, auditable, and free of Google effects."""

    def setUp(self):
        self.it_user = get_user_model().objects.create_user(
            username='portal-reporting-it', password='not-used', is_active=True,
        )
        self.emb_u_case = JawabuFarmerMaster.objects.create(
            customer_name='Embu reporting case', national_id='10000001',
            primary_phone='254700000001', branch='EMBU', county='EMBU',
            deposit_paid_hbg=Decimal('5000.00'), status='active',
        )
        JawabuFarmerMaster.objects.create(
            customer_name='Nakuru reporting case', national_id='10000002',
            primary_phone='254700000002', branch='NAKURU', county='NAKURU',
            deposit_paid_hbg=Decimal('7500.00'), status='active',
        )
        self.configuration = {
            'fields': ['customer_name', 'branch', 'deposit_paid_hbg'],
            'filters': [],
            'ordering': {'field': 'customer_name', 'direction': 'asc'},
        }

    def test_it_capability_is_allowed_and_jbl_officer_is_explicitly_denied(self):
        """Policy has a happy path and a denied path for the sensitive module."""
        self.assertTrue(has_capability(
            self.it_user, 'jawabu_portal', 'portal.reports.view',
            access={'roles': ['IT']},
        ))
        self.assertFalse(has_capability(
            self.it_user, 'jawabu_portal', 'portal.reports.manage',
            access={'roles': ['JBL_OFFICER']},
        ))

    def test_live_run_is_branch_scoped_and_audited_definition_is_idempotent(self):
        definition, replayed = create_definition(
            payload={
                'title': 'Branch deposits',
                'configuration': self.configuration,
                'charts': [{
                    'chart_type': 'bar', 'dimension_field': 'branch',
                    'aggregation': 'count', 'metric_field': '', 'date_bucket': '',
                }],
            },
            actor=self.it_user,
            request_id='portal-report-create-1',
        )
        self.assertFalse(replayed)
        duplicate, replayed = create_definition(
            payload={'title': 'Changed retry title', 'configuration': self.configuration, 'charts': []},
            actor=self.it_user,
            request_id='portal-report-create-1',
        )
        self.assertTrue(replayed)
        self.assertEqual(duplicate.pk, definition.pk)
        self.assertEqual(definition.charts.count(), 1)
        self.assertTrue(ComplianceAuditEvent.objects.filter(
            action='portal.report.created', subject_id=str(definition.pk),
        ).exists())

        result = run_definition(
            definition=definition,
            user=self.it_user,
            access={'roles': ['IT'], 'branches': ['EMBU']},
        )
        self.assertEqual(result['total_rows'], 1)
        self.assertEqual(result['rows'][0]['customer_name'], self.emb_u_case.customer_name)
        self.assertEqual(result['charts'][0]['labels'], ['EMBU'])
        self.assertEqual(result['charts'][0]['values'], [1.0])

    def test_chart_catalogue_is_curated_and_preview_is_branch_scoped_without_an_audit_write(self):
        catalogue = catalogue_payload()
        fields = {
            item['key']: item
            for category in catalogue['categories']
            for item in category['fields']
        }
        self.assertTrue(fields['branch']['chart_dimension'])
        self.assertTrue(fields['deposit_paid_hbg']['chart_metric'])
        self.assertFalse(fields['customer_name']['chart_dimension'])
        self.assertFalse(fields['village']['chart_dimension'])

        audit_count = ComplianceAuditEvent.objects.count()
        preview = preview_chart(
            configuration=self.configuration,
            chart={
                'chart_type': 'bar', 'dimension_field': 'branch',
                'aggregation': 'count', 'metric_field': '', 'date_bucket': '',
            },
            user=self.it_user,
            access={'roles': ['IT'], 'branches': ['EMBU']},
        )

        self.assertEqual(preview['labels'], ['EMBU'])
        self.assertEqual(preview['values'], [1.0])
        self.assertEqual(preview['type'], 'bar')
        self.assertEqual(ComplianceAuditEvent.objects.count(), audit_count)

    def test_chart_rules_reject_unapproved_dimensions_and_incomplete_date_grouping(self):
        with self.assertRaises(PortalReportingError):
            validate_charts([{
                'chart_type': 'bar', 'dimension_field': 'customer_name',
                'aggregation': 'count', 'metric_field': '', 'date_bucket': '',
            }], selected_fields=['customer_name'])
        with self.assertRaises(PortalReportingError):
            validate_charts([{
                'chart_type': 'line', 'dimension_field': 'created_at',
                'aggregation': 'count', 'metric_field': '', 'date_bucket': '',
            }], selected_fields=['created_at'])
        with self.assertRaises(PortalReportingError):
            validate_charts([{
                'chart_type': 'bar', 'dimension_field': 'branch',
                'aggregation': 'count', 'metric_field': 'deposit_paid_hbg', 'date_bucket': '',
            }], selected_fields=['branch', 'deposit_paid_hbg'])

    def test_doughnut_preview_falls_back_to_a_bar_for_many_categories(self):
        for index in range(3, 10):
            JawabuFarmerMaster.objects.create(
                customer_name=f'Branch reporting case {index}', national_id=f'1000000{index}',
                primary_phone=f'2547000000{index:02d}', branch=f'BRANCH {index}', county='OTHER',
                deposit_paid_hbg=Decimal('1000.00'), status='active',
            )
        preview = preview_chart(
            configuration=self.configuration,
            chart={
                'chart_type': 'doughnut', 'dimension_field': 'branch',
                'aggregation': 'count', 'metric_field': '', 'date_bucket': '',
            },
            user=self.it_user,
            access={'roles': ['IT'], 'branches': []},
        )
        self.assertEqual(preview['requested_type'], 'doughnut')
        self.assertEqual(preview['type'], 'bar')
        self.assertIn('More than eight categories', preview['notice'])

    def test_catalogue_rejects_sensitive_or_arbitrary_fields_and_xlsx_is_local(self):
        with self.assertRaises(PortalReportingError):
            validate_configuration({
                'fields': ['comments'],
                'filters': [],
                'ordering': {'field': 'comments', 'direction': 'asc'},
            })

        definition, _ = create_definition(
            payload={'title': 'Local export', 'configuration': self.configuration, 'charts': []},
            actor=self.it_user,
            request_id='portal-report-create-2',
        )
        workbook = load_workbook(BytesIO(export_xlsx(
            definition=definition,
            user=self.it_user,
            access={'roles': ['IT'], 'branches': ['EMBU']},
        )))
        self.assertEqual(workbook['Data']['A1'].value, 'Customer name')
        self.assertEqual(workbook['Data']['A2'].value, self.emb_u_case.customer_name)
        self.assertEqual(workbook['Data'].max_row, 2)

    def test_live_date_filter_and_numeric_ordering_use_the_current_case_rows(self):
        lower_deposit = JawabuFarmerMaster.objects.create(
            customer_name='Second Embu reporting case', national_id='10000003',
            primary_phone='254700000003', branch='EMBU', county='EMBU',
            deposit_paid_hbg=Decimal('1200.00'), status='active',
        )
        definition, _ = create_definition(
            payload={
                'title': 'Today Embu deposits',
                'configuration': {
                    'fields': ['customer_name', 'created_at', 'deposit_paid_hbg'],
                    'filters': [{
                        'field': 'created_at', 'operator': 'equals',
                        'value': timezone.localdate().isoformat(),
                    }],
                    'ordering': {'field': 'deposit_paid_hbg', 'direction': 'desc'},
                },
                'charts': [],
            },
            actor=self.it_user,
            request_id='portal-report-filter-order-1',
        )

        result = run_definition(
            definition=definition,
            user=self.it_user,
            access={'roles': ['IT'], 'branches': ['EMBU']},
        )

        self.assertEqual(result['total_rows'], 2)
        self.assertEqual(
            [row['customer_name'] for row in result['rows']],
            [self.emb_u_case.customer_name, lower_deposit.customer_name],
        )

    def test_one_invalid_historical_chart_does_not_block_rows_or_other_charts(self):
        definition, _ = create_definition(
            payload={'title': 'Chart recovery', 'configuration': self.configuration, 'charts': []},
            actor=self.it_user,
            request_id='portal-report-chart-recovery-1',
        )
        PortalReportChart.objects.create(
            definition=definition,
            chart_type='bar', dimension_field='removed_field',
            metric_field='', aggregation='count', date_bucket='', position=0,
        )

        result = run_definition(
            definition=definition,
            user=self.it_user,
            access={'roles': ['IT'], 'branches': ['EMBU']},
        )

        self.assertEqual(result['total_rows'], 1)
        self.assertEqual(result['charts'][0]['error'], 'This chart configuration needs review. The report rows are still available.')

    def test_relationship_inventory_keeps_other_workflows_non_joinable(self):
        summary = portal_relationship_summary()
        self.assertIn('core.TatTrackerCase', summary['unlinked_identity_only'])
        self.assertIn('core.SpinCreditRequest', summary['unlinked_identity_only'])
        self.assertIn('core.CaseUpdate', summary['unlinked_identity_only'])
        output = StringIO()
        call_command('inspect_reporting_relationships', '--json', stdout=output)
        self.assertIn('core.JawabuFarmerMaster', output.getvalue())

    @override_settings(PORTAL_WEBAPP_REQUIRE_TELEGRAM_AUTH=False)
    def test_catalogue_endpoint_is_wired_without_a_google_side_effect(self):
        response = self.client.get('/api/portal/reports/catalogue/')
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['ok'])
        all_keys = {
            field['key']
            for category in payload['catalogue']['categories']
            for field in category['fields']
        }
        self.assertIn('customer_name', all_keys)
        self.assertNotIn('comments', all_keys)

    @override_settings(PORTAL_WEBAPP_REQUIRE_TELEGRAM_AUTH=False)
    def test_chart_preview_endpoint_returns_only_a_scoped_aggregate_without_an_audit_write(self):
        audit_count = ComplianceAuditEvent.objects.count()
        response = self.client.post(
            '/api/portal/reports/preview/',
            data={
                'configuration': self.configuration,
                'chart': {
                    'chart_type': 'bar', 'dimension_field': 'branch',
                    'aggregation': 'count', 'metric_field': '', 'date_bucket': '',
                },
            },
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['ok'])
        self.assertEqual(payload['preview']['labels'], ['EMBU', 'NAKURU'])
        self.assertNotIn('rows', payload['preview'])
        self.assertEqual(ComplianceAuditEvent.objects.count(), audit_count)
