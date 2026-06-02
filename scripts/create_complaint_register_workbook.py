"""Create a polished, bot-compatible complaint register workbook.

The generated workbook mirrors the Order Approval sheet pattern:

- Row 1 is a visual title banner.
- Row 2 is the bot-readable header row.
- Row 3+ contains case data.

The script imports existing rows from COMPLAINT MANAGEMENT REGISTER(1).xlsx by
default, then writes a redesigned workbook with dropdowns, validation,
conditional formatting, dashboard formulas, Staff, Dropdown Options, and Legend
tabs.
"""
from __future__ import annotations

import argparse
from collections import OrderedDict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:  # pragma: no cover - exercised manually.
    raise SystemExit(
        "openpyxl is required for this script. Install it with: pip install openpyxl"
    ) from exc


DEFAULT_INPUT = "COMPLAINT MANAGEMENT REGISTER(1).xlsx"
DEFAULT_OUTPUT = "COMPLAINT_MANAGEMENT_REGISTER_V2.xlsx"
REGISTER_SHEET = "Complaints Register"
SUMMARY_SHEET = "Complaint Summary"
OPTIONS_SHEET = "Dropdown Options"
STAFF_SHEET = "Staff"
LEGEND_SHEET = "Legend"
TITLE = "COMPLAINT MANAGEMENT REGISTER - CUSTOMER CASES"
PRE_FORMATTED_ROWS = 500

HEADERS = [
    "Complaint ID",
    "message_id",
    "Date Reported",
    "Customer Name",
    "Customer ID / Account",
    "Phone Number",
    "JBL Reported By",
    "Branch / Region",
    "Complaint Category",
    "Complaint Description",
    "raw_message",
    "gps_link",
    "image_flag",
    "source",
    "Loan Status",
    "Loan at Risk",
    "Risk Level",
    "Status",
    "Resolution Details",
    "Date Resolved",
    "Days Open",
]

REQUIRED_HEADERS = [
    "Date Reported",
    "Customer Name",
    "Customer ID / Account",
    "Phone Number",
    "Branch / Region",
    "Complaint Description",
    "Status",
]

OPTIONS_HEADERS = [
    "Branch / Region",
    "JBL Reported By",
    "Complaint Category",
    "Status",
    "Loan Status",
    "Risk Level",
    "Source",
    "Image Flag",
]

KENYA_COUNTIES = [
    "BARINGO", "BOMET", "BUNGOMA", "BUSIA", "ELGEYO MARAKWET", "EMBU",
    "GARISSA", "HOMA BAY", "ISIOLO", "KAJIADO", "KAKAMEGA", "KERICHO",
    "KIAMBU", "KILIFI", "KIRINYAGA", "KISII", "KISUMU", "KITUI", "KWALE",
    "LAIKIPIA", "LAMU", "MACHAKOS", "MAKUENI", "MANDERA", "MARSABIT",
    "MERU", "MIGORI", "MOMBASA", "MURANGA", "NAIROBI", "NAKURU", "NANDI",
    "NAROK", "NYAMIRA", "NYANDARUA", "NYERI", "SAMBURU", "SIAYA",
    "TAITA TAVETA", "TANA RIVER", "THARAKA NITHI", "TRANS NZOIA",
    "TURKANA", "UASIN GISHU", "VIHIGA", "WAJIR", "WEST POKOT",
]

DEFAULT_OPTIONS = {
    "Branch / Region": KENYA_COUNTIES,
    "JBL Reported By": ["JACKSON NJOROGE", "DICKSON MWANGI"],
    "Complaint Category": [
        "System Underperformance",
        "System Damage(Tear/Burst)",
        "Bag Leakage",
        "Blockage Inlet/Oulet",
        "Relocation",
        "Other",
    ],
    "Status": ["Open", "In Progress", "Waiting for Customer", "Resolved", "Closed"],
    "Loan Status": ["Performing", "Non Performing", "Cleared", "Unknown"],
    "Risk Level": ["Low", "Moderate", "High", "Critical"],
    "Source": ["telegram bot", "google sheets", "manual"],
    "Image Flag": ["TRUE", "FALSE"],
}

HEADER_GROUPS = {
    "Complaint ID": "system",
    "message_id": "system",
    "Date Reported": "intake",
    "Customer Name": "customer",
    "Customer ID / Account": "customer",
    "Phone Number": "customer",
    "JBL Reported By": "staff",
    "Branch / Region": "staff",
    "Complaint Category": "complaint",
    "Complaint Description": "complaint",
    "raw_message": "system",
    "gps_link": "system",
    "image_flag": "system",
    "source": "system",
    "Loan Status": "risk",
    "Loan at Risk": "risk",
    "Risk Level": "risk",
    "Status": "workflow",
    "Resolution Details": "workflow",
    "Date Resolved": "workflow",
    "Days Open": "system",
}

GROUP_COLOURS = {
    "system": ("37474F", "ECEFF1"),
    "intake": ("1A7744", "E8F5E9"),
    "customer": ("1565C0", "E3F2FD"),
    "staff": ("00695C", "E0F2F1"),
    "complaint": ("6A1B9A", "F3E5F5"),
    "risk": ("E65100", "FBE9E7"),
    "workflow": ("B71C1C", "FFEBEE"),
}

COLUMN_WIDTHS = {
    "Complaint ID": 16,
    "message_id": 24,
    "Date Reported": 15,
    "Customer Name": 28,
    "Customer ID / Account": 20,
    "Phone Number": 18,
    "JBL Reported By": 22,
    "Branch / Region": 18,
    "Complaint Category": 26,
    "Complaint Description": 42,
    "raw_message": 42,
    "gps_link": 36,
    "image_flag": 12,
    "source": 16,
    "Loan Status": 18,
    "Loan at Risk": 16,
    "Risk Level": 14,
    "Status": 20,
    "Resolution Details": 38,
    "Date Resolved": 15,
    "Days Open": 12,
}


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Create a redesigned complaint register workbook.",
    )
    parser.add_argument("-i", "--input", default=DEFAULT_INPUT)
    parser.add_argument("-o", "--output", default=DEFAULT_OUTPUT)
    parser.add_argument("--blank", action="store_true", help="Do not import existing rows.")
    args = parser.parse_args()

    source = Path(args.input)
    output = Path(args.output)
    rows = [] if args.blank else load_source_rows(source)
    options = build_options(rows)
    create_workbook(output, rows, options)
    print(f"Created {output.resolve()}")


def load_source_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Input workbook not found: {path}")

    wb = load_workbook(path, data_only=False)
    if REGISTER_SHEET not in wb.sheetnames:
        raise SystemExit(f"Input workbook has no {REGISTER_SHEET!r} sheet.")

    ws = wb[REGISTER_SHEET]
    source_headers = [
        str(ws.cell(1, col).value or "").strip()
        for col in range(1, ws.max_column + 1)
    ]
    header_indexes = {
        normalize(header): idx + 1
        for idx, header in enumerate(source_headers)
        if header
    }

    rows: list[dict[str, Any]] = []
    for row_index in range(2, ws.max_row + 1):
        row = {}
        for header in HEADERS:
            source_col = header_indexes.get(normalize(header))
            row[header] = ws.cell(row_index, source_col).value if source_col else None
        if is_real_case_row(row):
            normalize_imported_row(row)
            rows.append(row)
    return rows


def is_real_case_row(row: dict[str, Any]) -> bool:
    for header in (
        "message_id",
        "Date Reported",
        "Customer Name",
        "Phone Number",
        "Complaint Description",
        "Status",
    ):
        value = row.get(header)
        if value and not looks_like_formula(value):
            return True
    return False


def normalize_imported_row(row: dict[str, Any]) -> None:
    for header in ("Customer Name", "JBL Reported By", "Branch / Region"):
        value = clean_text(row.get(header))
        row[header] = value.upper() if value else None

    row["Phone Number"] = normalize_phone(row.get("Phone Number"))
    row["Status"] = clean_status(row.get("Status")) or "Open"
    row["source"] = clean_text(row.get("source")) or "google sheets"
    row["image_flag"] = clean_image_flag(row.get("image_flag"))

    for header in ("Complaint ID", "Days Open"):
        if looks_like_formula(row.get(header)):
            row[header] = None


def build_options(rows: list[dict[str, Any]]) -> dict[str, list[str]]:
    options = {key: list(values) for key, values in DEFAULT_OPTIONS.items()}
    for row in rows:
        append_option(options, "JBL Reported By", row.get("JBL Reported By"))
        append_option(options, "Complaint Category", row.get("Complaint Category"))
        append_option(options, "Status", row.get("Status"))
        append_option(options, "Loan Status", row.get("Loan Status"))
        append_option(options, "Risk Level", row.get("Risk Level"))
        append_option(options, "Source", row.get("source"))
        append_option(options, "Image Flag", row.get("image_flag"))
    return {
        key: sorted(unique_nonempty(values), key=lambda value: value.upper())
        for key, values in options.items()
    }


def append_option(options: dict[str, list[str]], key: str, value: Any) -> None:
    value = clean_text(value)
    if not value or looks_like_formula(value):
        return
    options.setdefault(key, []).append(value.upper() if key in {"Branch / Region", "JBL Reported By"} else value)


def create_workbook(output: Path, source_rows: list[dict[str, Any]], options: dict[str, list[str]]) -> None:
    wb = Workbook()
    register = wb.active
    register.title = REGISTER_SHEET
    summary = wb.create_sheet(SUMMARY_SHEET)
    options_sheet = wb.create_sheet(OPTIONS_SHEET)
    staff = wb.create_sheet(STAFF_SHEET)
    legend = wb.create_sheet(LEGEND_SHEET)

    build_register_sheet(register, source_rows)
    build_options_sheet(options_sheet, options)
    build_staff_sheet(staff)
    build_summary_sheet(summary)
    build_legend_sheet(legend)
    wb.active = wb.sheetnames.index(REGISTER_SHEET)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def build_register_sheet(ws, source_rows: list[dict[str, Any]]) -> None:
    last_col = len(HEADERS)
    ws.sheet_properties.tabColor = "0D47A1"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = ws.cell(1, 1, TITLE)
    title.fill = PatternFill("solid", fgColor="0D47A1")
    title.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(2, col, header)
        header_fill, data_fill = GROUP_COLOURS[HEADER_GROUPS.get(header, "system")]
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS.get(header, 18)
    ws.row_dimensions[2].height = 36

    total_rows = max(PRE_FORMATTED_ROWS, len(source_rows) + 50)
    for row_offset in range(total_rows):
        row_number = row_offset + 3
        source = source_rows[row_offset] if row_offset < len(source_rows) else {}
        write_register_row(ws, row_number, source)

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}2"
    apply_validations(ws, total_rows)
    apply_conditional_formatting(ws, total_rows)
    ws.sheet_view.showGridLines = False


def write_register_row(ws, row_number: int, source: dict[str, Any]) -> None:
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row_number, col)
        _, data_fill = GROUP_COLOURS[HEADER_GROUPS.get(header, "system")]
        cell.fill = PatternFill("solid", fgColor=data_fill)
        cell.font = Font(name="Arial", size=10, color="111827")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border()

        if header == "Complaint ID":
            cell.value = f'=IF(C{row_number}="","","CMP"&TEXT(ROW()-2,"000"))'
        elif header == "Days Open":
            cell.value = f'=IF(R{row_number}="Closed","",IF(C{row_number}="","",TODAY()-C{row_number}))'
        else:
            value = source.get(header)
            cell.value = "" if looks_like_formula(value) else value

        if header in {"Date Reported", "Date Resolved"}:
            cell.number_format = "dd-mmm-yyyy"
        if header == "Loan at Risk":
            cell.number_format = '#,##0'


def build_options_sheet(ws, options: dict[str, list[str]]) -> None:
    ws.sheet_properties.tabColor = "37474F"
    for col, header in enumerate(OPTIONS_HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = PatternFill("solid", fgColor="37474F")
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 24
        values = options.get(header, [])
        for row, value in enumerate(values, start=2):
            ws.cell(row, col, value)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OPTIONS_HEADERS))}1"


def build_staff_sheet(ws) -> None:
    headers = ["Name", "Email", "Role", "Branch", "Notify On", "Editable Columns", "Active"]
    samples = [
        ["IT Admin", "it@example.com", "IT", "All", "all", "All", "Yes"],
        ["Manager", "manager@example.com", "Manager", "All", "stale_digest,status_closed,status_resolved", "Resolution,Risk", "Yes"],
        ["Field Staff", "staff@example.com", "BRO", "MURANGA", "stale_digest", "Resolution", "Yes"],
    ]
    ws.sheet_properties.tabColor = "00695C"
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = PatternFill("solid", fgColor="00695C")
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 24
    for row_index, row in enumerate(samples, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row_index, col, value)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:G1"


def build_summary_sheet(ws) -> None:
    ws.sheet_properties.tabColor = "1A7744"
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "Complaint Dashboard"
    title.fill = PatternFill("solid", fgColor="1A7744")
    title.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center")

    metrics = [
        ("Total Complaints", '=COUNTA(\'Complaints Register\'!D3:D1000)'),
        ("Open", '=COUNTIF(\'Complaints Register\'!R:R,"Open")'),
        ("In Progress", '=COUNTIF(\'Complaints Register\'!R:R,"In Progress")'),
        ("Waiting for Customer", '=COUNTIF(\'Complaints Register\'!R:R,"Waiting for Customer")'),
        ("Resolved", '=COUNTIF(\'Complaints Register\'!R:R,"Resolved")'),
        ("Closed", '=COUNTIF(\'Complaints Register\'!R:R,"Closed")'),
        ("High/Critical Risk", '=COUNTIF(\'Complaints Register\'!Q:Q,"High")+COUNTIF(\'Complaints Register\'!Q:Q,"Critical")'),
        ("Average Days Open", '=IFERROR(AVERAGE(\'Complaints Register\'!U3:U1000),0)'),
    ]
    for row, (label, formula) in enumerate(metrics, start=3):
        ws.cell(row, 1, label)
        ws.cell(row, 2, formula)
        ws.cell(row, 1).font = Font(name="Arial", bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def build_legend_sheet(ws) -> None:
    ws.sheet_properties.tabColor = "6A1B9A"
    rows = [
        ["COLUMN GROUP GUIDE", "", ""],
        ["Group", "Columns included", "Filled by"],
        ["System / Bot", "Complaint ID, message_id, raw_message, gps_link, image_flag, source, Days Open", "Bot / formulas"],
        ["Intake", "Date Reported", "Bot / staff"],
        ["Customer", "Customer Name, Customer ID / Account, Phone Number", "Bot / staff"],
        ["Staff", "JBL Reported By, Branch / Region", "Bot / staff"],
        ["Complaint", "Complaint Category, Complaint Description", "Bot / staff"],
        ["Risk", "Loan Status, Loan at Risk, Risk Level", "Back-office"],
        ["Workflow", "Status, Resolution Details, Date Resolved", "Back-office / support"],
        ["Bot note", "The bot reads row 2. Configure workflow.header_row = 2 in Django Admin.", ""],
    ]
    for row_index, row in enumerate(rows, start=1):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row_index, col, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_index in {1, 2}:
                cell.fill = PatternFill("solid", fgColor="6A1B9A")
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 28
    ws.freeze_panes = "A3"


def apply_validations(ws, data_rows: int) -> None:
    first_row, last_row = 3, data_rows + 2
    option_map = {
        "Branch / Region": 1,
        "JBL Reported By": 2,
        "Complaint Category": 3,
        "Status": 4,
        "Loan Status": 5,
        "Risk Level": 6,
        "source": 7,
        "image_flag": 8,
    }
    for header, options_col in option_map.items():
        add_list_validation(
            ws,
            header,
            f"'{OPTIONS_SHEET}'!${get_column_letter(options_col)}$2:${get_column_letter(options_col)}$500",
            first_row,
            last_row,
        )
    add_phone_validation(ws, "Phone Number", first_row, last_row)
    add_date_validation(ws, "Date Reported", first_row, last_row)
    add_date_validation(ws, "Date Resolved", first_row, last_row)
    add_nonnegative_validation(ws, "Loan at Risk", first_row, last_row)
    for header in ("Customer Name", "JBL Reported By", "Branch / Region"):
        add_uppercase_validation(ws, header, first_row, last_row)


def apply_conditional_formatting(ws, data_rows: int) -> None:
    last_row = data_rows + 2
    last_col = get_column_letter(len(HEADERS))
    data_range = f"A3:{last_col}{last_row}"
    status_col = get_column_letter(HEADERS.index("Status") + 1)
    risk_col = get_column_letter(HEADERS.index("Risk Level") + 1)
    rules = [
        (f'${status_col}3="Closed"', "E8F5E9", "1B5E20"),
        (f'${status_col}3="Resolved"', "E8F5E9", "1B5E20"),
        (f'${status_col}3="Open"', "FFF8E1", "E65100"),
        (f'${risk_col}3="High"', "FFEBEE", "B71C1C"),
        (f'${risk_col}3="Critical"', "FCE4EC", "880E4F"),
    ]
    for formula, fill, font in rules:
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[formula],
                fill=PatternFill("solid", fgColor=fill),
                font=Font(color=font),
            ),
        )


def add_list_validation(ws, header: str, formula: str, first_row: int, last_row: int) -> None:
    col = header_col(header)
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = f"Choose a valid {header} from Dropdown Options."
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}{first_row}:{get_column_letter(col)}{last_row}")


def add_phone_validation(ws, header: str, first_row: int, last_row: int) -> None:
    col = get_column_letter(header_col(header))
    formula = f'=OR({col}{first_row}="",AND(ISNUMBER(--{col}{first_row}),LEN({col}{first_row})=12,LEFT({col}{first_row},3)="254"))'
    dv = DataValidation(type="custom", formula1=formula, allow_blank=True)
    dv.error = "Use 254XXXXXXXXX format, for example 254740614990."
    dv.errorTitle = "Invalid phone"
    ws.add_data_validation(dv)
    dv.add(f"{col}{first_row}:{col}{last_row}")


def add_date_validation(ws, header: str, first_row: int, last_row: int) -> None:
    col = get_column_letter(header_col(header))
    dv = DataValidation(type="date", operator="between", formula1="1/1/2020", formula2="12/31/2099", allow_blank=True)
    dv.error = f"{header} must be a valid date."
    dv.errorTitle = "Invalid date"
    ws.add_data_validation(dv)
    dv.add(f"{col}{first_row}:{col}{last_row}")


def add_nonnegative_validation(ws, header: str, first_row: int, last_row: int) -> None:
    col = get_column_letter(header_col(header))
    dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv.error = f"{header} must be zero or greater."
    dv.errorTitle = "Invalid amount"
    ws.add_data_validation(dv)
    dv.add(f"{col}{first_row}:{col}{last_row}")


def add_uppercase_validation(ws, header: str, first_row: int, last_row: int) -> None:
    col = get_column_letter(header_col(header))
    formula = f'=OR({col}{first_row}="",EXACT({col}{first_row},UPPER({col}{first_row})))'
    dv = DataValidation(type="custom", formula1=formula, allow_blank=True)
    dv.error = f"{header} should be uppercase."
    dv.errorTitle = "Use uppercase"
    ws.add_data_validation(dv)
    dv.add(f"{col}{first_row}:{col}{last_row}")


def header_col(header: str) -> int:
    return HEADERS.index(header) + 1


def thin_border() -> Border:
    side = Side(style="thin", color="DDDDDD")
    return Border(left=side, right=side, top=side, bottom=side)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize(value: Any) -> str:
    return clean_text(value).lower()


def looks_like_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def normalize_phone(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    raw = clean_text(value)
    if looks_like_formula(raw):
        return ""
    digits = "".join(ch for ch in raw if ch.isdigit())
    if not digits:
        return ""
    if digits.startswith("0") and len(digits) == 10:
        return "254" + digits[1:]
    if (digits.startswith("7") or digits.startswith("1")) and len(digits) == 9:
        return "254" + digits
    return digits


def clean_status(value: Any) -> str:
    text = clean_text(value)
    lookup = {
        "waiting for customer": "Waiting for Customer",
        "in progress": "In Progress",
        "closed": "Closed",
        "resolved": "Resolved",
        "open": "Open",
    }
    return lookup.get(text.lower(), text)


def clean_image_flag(value: Any) -> str:
    text = clean_text(value).upper()
    if text in {"TRUE", "YES", "1"}:
        return "TRUE"
    if text in {"FALSE", "NO", "0"}:
        return "FALSE"
    return ""


def unique_nonempty(values: list[Any]) -> list[str]:
    ordered = OrderedDict()
    for value in values:
        text = clean_text(value)
        if text and not looks_like_formula(text):
            ordered[text] = None
    return list(ordered.keys())


if __name__ == "__main__":
    main()
