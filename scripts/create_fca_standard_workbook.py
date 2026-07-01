"""Create a simple standardized FCA workbook for staff use.

This workbook is intentionally not automated with Apps Script. It provides two
flat, filterable sheets that BRO/FCA staff can fill manually:

- FCA Visits: approval/appraisal visit outcomes
- FCA Collections: collection follow-up outcomes
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = "FCA_Simple_Staff_Template.xlsx"
MAX_ROWS = 1000
HEADER_ROW = 2
FIRST_DATA_ROW = 3

VISIT_HEADERS = [
    "FCA VISIT DATE",
    "CUSTOMER NAME",
    "PHONE",
    "COUNTY / HUB",
    "LOCATION / LANDMARK",
    "STAFF",
    "DEPOSIT",
    "APPROVAL BASIS",
    "COMMENT",
    "DECISION",
]

COLLECTION_HEADERS = [
    "FCA VISIT DATE",
    "CUSTOMER NAME",
    "PHONE",
    "COUNTY / HUB",
    "LOCATION / LANDMARK",
    "OFFICER",
    "ARREARS",
    "AMOUNT PAID",
    "COMMENT",
    "OUTCOME",
    "NEXT COMMITMENT DATE",
]

VISIT_WIDTHS = {
    "FCA VISIT DATE": 16,
    "CUSTOMER NAME": 28,
    "PHONE": 20,
    "COUNTY / HUB": 18,
    "LOCATION / LANDMARK": 34,
    "STAFF": 22,
    "DEPOSIT": 14,
    "APPROVAL BASIS": 30,
    "COMMENT": 44,
    "DECISION": 16,
}

COLLECTION_WIDTHS = {
    "FCA VISIT DATE": 16,
    "CUSTOMER NAME": 28,
    "PHONE": 20,
    "COUNTY / HUB": 18,
    "LOCATION / LANDMARK": 34,
    "OFFICER": 22,
    "ARREARS": 14,
    "AMOUNT PAID": 14,
    "COMMENT": 44,
    "OUTCOME": 22,
    "NEXT COMMITMENT DATE": 22,
}

VISIT_DECISIONS = ["Approved", "Rejected", "Deferred", "Cash", "Under Review"]
COLLECTION_OUTCOMES = [
    "Paid",
    "Part Paid",
    "PTP",
    "Demand Issued",
    "Disconnected",
    "Reconnect After Payment",
    "Not Available",
    "Not Visited",
    "Decommission Recommended",
    "Under Review",
]

COUNTIES = [
    "Baringo", "Bomet", "Bungoma", "Busia", "Elgeyo-Marakwet", "Embu",
    "Garissa", "Homa Bay", "Isiolo", "Kajiado", "Kakamega", "Kericho",
    "Kiambu", "Kilifi", "Kirinyaga", "Kisii", "Kisumu", "Kitui", "Kwale",
    "Laikipia", "Lamu", "Machakos", "Makueni", "Mandera", "Marsabit",
    "Meru", "Migori", "Mombasa", "Murang'a", "Nairobi", "Nakuru",
    "Nandi", "Narok", "Nyamira", "Nyandarua", "Nyeri", "Samburu",
    "Siaya", "Taita-Taveta", "Tana River", "Tharaka-Nithi", "Trans Nzoia",
    "Turkana", "Uasin Gishu", "Vihiga", "Wajir", "West Pokot",
]

HEADER_GROUPS = {
    "identity": {"CUSTOMER NAME", "PHONE"},
    "location": {"COUNTY / HUB", "LOCATION / LANDMARK"},
    "staff": {"STAFF", "OFFICER"},
    "money": {"DEPOSIT", "ARREARS", "AMOUNT PAID"},
    "comment": {"APPROVAL BASIS", "COMMENT"},
    "decision": {"DECISION", "OUTCOME", "NEXT COMMITMENT DATE"},
}

COLORS = {
    "title": "1F4E3D",
    "identity": "1F4E79",
    "location": "60497A",
    "staff": "0F766E",
    "money": "B45309",
    "comment": "9F1239",
    "decision": "44337A",
    "default": "334155",
    "white": "FFFFFF",
    "light_green": "E7F5E8",
    "light_red": "FCE8E8",
    "light_yellow": "FFF6D6",
    "light_gray": "F8FAFC",
}


def main() -> None:
    workbook = Workbook()
    visit_sheet = workbook.active
    visit_sheet.title = "FCA Visits"
    setup_sheet(
        visit_sheet,
        title="FCA VISITS - APPROVAL / APPRAISAL",
        headers=VISIT_HEADERS,
        widths=VISIT_WIDTHS,
    )
    setup_sheet(
        workbook.create_sheet("FCA Collections"),
        title="FCA COLLECTIONS - FOLLOW UP / ARREARS",
        headers=COLLECTION_HEADERS,
        widths=COLLECTION_WIDTHS,
    )
    workbook.save(OUTPUT)
    print(f"Created {Path(OUTPUT).resolve()}")


def setup_sheet(sheet, title: str, headers: list[str], widths: dict[str, int]) -> None:
    last_col = get_column_letter(len(headers))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.fill = PatternFill("solid", fgColor=COLORS["title"])
    title_cell.font = Font(color=COLORS["white"], bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26

    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=index, value=header)
        group = header_group(header)
        cell.fill = PatternFill("solid", fgColor=COLORS[group])
        cell.font = Font(color=COLORS["white"], bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(header, 18)

    for row in range(FIRST_DATA_ROW, MAX_ROWS + 1):
        fill = PatternFill("solid", fgColor=COLORS["light_gray"] if row % 2 == 0 else "FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border(light=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{MAX_ROWS}"
    sheet.sheet_view.showGridLines = False

    add_validations(sheet, headers)
    add_conditional_formatting(sheet, headers)
    add_column_formats(sheet, headers)


def header_group(header: str) -> str:
    for group, group_headers in HEADER_GROUPS.items():
        if header in group_headers:
            return group
    return "default"


def thin_border(light: bool = False) -> Border:
    color = "E2E8F0" if light else "CBD5E1"
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def add_validations(sheet, headers: list[str]) -> None:
    add_list_validation(sheet, headers, "DECISION", VISIT_DECISIONS)
    add_list_validation(sheet, headers, "OUTCOME", COLLECTION_OUTCOMES)
    add_list_validation(sheet, headers, "COUNTY / HUB", COUNTIES)

    for header in ("FCA VISIT DATE", "NEXT COMMITMENT DATE"):
        if header in headers:
            col = get_column_letter(headers.index(header) + 1)
            validation = DataValidation(
                type="date",
                operator="between",
                formula1="DATE(2020,1,1)",
                formula2="DATE(2035,12,31)",
                allow_blank=True,
            )
            validation.error = "Use a valid date, for example 25-May-2026."
            validation.errorTitle = "Invalid date"
            sheet.add_data_validation(validation)
            validation.add(f"{col}{FIRST_DATA_ROW}:{col}{MAX_ROWS}")

    for header in ("PHONE",):
        if header in headers:
            col = get_column_letter(headers.index(header) + 1)
            validation = DataValidation(
                type="textLength",
                operator="between",
                formula1="9",
                formula2="15",
                allow_blank=True,
            )
            validation.prompt = "Use 254 format where possible, for example 254712345678."
            validation.error = "Phone number should be in 254 format where possible."
            validation.errorTitle = "Check phone number"
            sheet.add_data_validation(validation)
            validation.add(f"{col}{FIRST_DATA_ROW}:{col}{MAX_ROWS}")


def add_list_validation(sheet, headers: list[str], header: str, values: list[str]) -> None:
    if header not in headers:
        return
    col = get_column_letter(headers.index(header) + 1)
    formula = '"' + ','.join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = f"Choose a valid {header.lower()} from the dropdown."
    validation.errorTitle = "Invalid selection"
    sheet.add_data_validation(validation)
    validation.add(f"{col}{FIRST_DATA_ROW}:{col}{MAX_ROWS}")


def add_conditional_formatting(sheet, headers: list[str]) -> None:
    if "DECISION" in headers:
        col = get_column_letter(headers.index("DECISION") + 1)
        row_range = f"A{FIRST_DATA_ROW}:{get_column_letter(len(headers))}{MAX_ROWS}"
        add_row_rule(sheet, row_range, f'${col}{FIRST_DATA_ROW}="Approved"', COLORS["light_green"])
        add_row_rule(sheet, row_range, f'${col}{FIRST_DATA_ROW}="Rejected"', COLORS["light_red"])
        add_row_rule(sheet, row_range, f'${col}{FIRST_DATA_ROW}="Cash"', COLORS["light_yellow"])

    if "OUTCOME" in headers:
        col = get_column_letter(headers.index("OUTCOME") + 1)
        row_range = f"A{FIRST_DATA_ROW}:{get_column_letter(len(headers))}{MAX_ROWS}"
        add_row_rule(sheet, row_range, f'${col}{FIRST_DATA_ROW}="Paid"', COLORS["light_green"])
        add_row_rule(sheet, row_range, f'${col}{FIRST_DATA_ROW}="Disconnected"', COLORS["light_red"])
        add_row_rule(sheet, row_range, f'${col}{FIRST_DATA_ROW}="PTP"', COLORS["light_yellow"])


def add_row_rule(sheet, cell_range: str, formula: str, color: str) -> None:
    sheet.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)),
    )


def add_column_formats(sheet, headers: list[str]) -> None:
    for header in ("FCA VISIT DATE", "NEXT COMMITMENT DATE"):
        if header in headers:
            col = headers.index(header) + 1
            for row in range(FIRST_DATA_ROW, MAX_ROWS + 1):
                sheet.cell(row=row, column=col).number_format = "dd-mmm-yyyy"
    for header in ("DEPOSIT", "ARREARS", "AMOUNT PAID"):
        if header in headers:
            col = headers.index(header) + 1
            for row in range(FIRST_DATA_ROW, MAX_ROWS + 1):
                sheet.cell(row=row, column=col).number_format = '#,##0'


if __name__ == "__main__":
    main()
