"""Extract HomeBiogas invoice PDF details into an Excel review file.

This is a local verification utility. It uses the same parser functions as the
portal invoice upload flow, but it does not update the database or Google
Sheets.

Examples:
    python scripts/extract_invoices_to_excel.py "INVOICING/docmaster/SOFY/#076.pdf"
    python scripts/extract_invoices_to_excel.py "INVOICING/docmaster/SOFY" -o artifacts/invoice_extract.xlsx
    python scripts/extract_invoices_to_excel.py "INVOICING" --recursive
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

import django
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from core.services.invoice_parser import (  # noqa: E402
    PdfReader,
    _invoice_segments_from_text,
    clean_amount,
    parse_invoice_date,
    parse_invoice_text,
)


HEADERS = [
    "Source File",
    "Source Page",
    "Status",
    "Reason",
    "Invoice No",
    "Invoice Date Raw",
    "Invoice Date",
    "Customer Name",
    "Customer Phone",
    "National ID",
    "Invoice Amount",
    "Total After Discount",
    "Discount",
    "Payment",
    "Balance Due",
    "Calculated Balance Due",
    "Balance Check",
    "Balance Difference",
    "Text Preview",
]


def _format_date(value: str) -> str:
    parsed = parse_invoice_date(value)
    return parsed.strftime("%d-%B-%Y") if parsed else ""


def _format_amount(value: str) -> str:
    parsed = clean_amount(value)
    return str(parsed) if parsed is not None else ""


def _pdf_paths(source: Path, recursive: bool) -> list[Path]:
    if source.is_file():
        return [source] if source.suffix.lower() == ".pdf" else []
    pattern = "**/*.pdf" if recursive else "*.pdf"
    return sorted(source.glob(pattern))


def _preview(text: str, limit: int = 500) -> str:
    return " ".join((text or "").split())[:limit]


def _invoice_rows_from_page(pdf_path: Path, page_number: int, text: str) -> list[dict]:
    parsed = parse_invoice_text(text, page_number)
    if parsed:
        return [_success_row(pdf_path, parsed, text)]

    rows = []
    for index, segment in enumerate(_invoice_segments_from_text(text), start=1):
        segment_page = page_number if len(rows) == 0 else f"{page_number}.{index}"
        parsed_segment = parse_invoice_text(segment, page_number)
        if parsed_segment:
            parsed_segment["page"] = segment_page
            rows.append(_success_row(pdf_path, parsed_segment, segment))

    if rows:
        return rows

    return [{
        "Source File": str(pdf_path),
        "Source Page": page_number,
        "Status": "Not parsed",
        "Reason": "No valid HomeBiogas invoice found on this page.",
        "Text Preview": _preview(text),
    }]


def _success_row(pdf_path: Path, parsed: dict, text: str) -> dict:
    return {
        "Source File": str(pdf_path),
        "Source Page": parsed.get("page", ""),
        "Status": "Parsed",
        "Reason": "",
        "Invoice No": parsed.get("invoice_no", ""),
        "Invoice Date Raw": parsed.get("invoice_date", ""),
        "Invoice Date": _format_date(parsed.get("invoice_date", "")),
        "Customer Name": parsed.get("customer_name", ""),
        "Customer Phone": parsed.get("customer_phone", ""),
        "National ID": parsed.get("customer_id", ""),
        "Invoice Amount": _format_amount(parsed.get("invoice_amount", "")),
        "Total After Discount": _format_amount(parsed.get("total_after_discount", "")),
        "Discount": _format_amount(parsed.get("discount", "")),
        "Payment": _format_amount(parsed.get("payment", "")),
        "Balance Due": _format_amount(parsed.get("balance_due", "")),
        "Calculated Balance Due": _format_amount(parsed.get("calculated_balance_due", "")),
        "Balance Check": parsed.get("balance_due_check", ""),
        "Balance Difference": _format_amount(parsed.get("balance_due_difference", "")),
        "Text Preview": _preview(text),
    }


def extract_invoices(paths: list[Path]) -> list[dict]:
    rows = []
    for pdf_path in paths:
        try:
            reader = PdfReader(str(pdf_path))
        except Exception as exc:
            rows.append({
                "Source File": str(pdf_path),
                "Source Page": "",
                "Status": "Error",
                "Reason": f"Could not open PDF: {exc}",
            })
            continue

        for page_number, page in enumerate(reader.pages, start=1):
            try:
                text = page.extract_text() or ""
            except Exception as exc:
                rows.append({
                    "Source File": str(pdf_path),
                    "Source Page": page_number,
                    "Status": "Error",
                    "Reason": f"Could not extract page text: {exc}",
                })
                continue
            rows.extend(_invoice_rows_from_page(pdf_path, page_number, text))
    return rows


def write_excel(rows: list[dict], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Extract"

    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_col = HEADERS.index("Status") + 1
    balance_check_col = HEADERS.index("Balance Check") + 1
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=status_col).value
        fill = PatternFill("solid", fgColor="E2F0D9" if status == "Parsed" else "FCE4D6")
        ws.cell(row=row, column=status_col).fill = fill
        balance_check = ws.cell(row=row, column=balance_check_col).value
        if balance_check == "OK":
            ws.cell(row=row, column=balance_check_col).fill = PatternFill("solid", fgColor="E2F0D9")
        elif balance_check:
            ws.cell(row=row, column=balance_check_col).fill = PatternFill("solid", fgColor="FCE4D6")
        ws.cell(row=row, column=HEADERS.index("Text Preview") + 1).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "Source File": 45,
        "Source Page": 12,
        "Status": 12,
        "Reason": 38,
        "Invoice No": 16,
        "Invoice Date Raw": 18,
        "Invoice Date": 20,
        "Customer Name": 28,
        "Customer Phone": 18,
        "National ID": 16,
        "Invoice Amount": 16,
        "Total After Discount": 20,
        "Discount": 14,
        "Payment": 14,
        "Balance Due": 16,
        "Calculated Balance Due": 22,
        "Balance Check": 18,
        "Balance Difference": 18,
        "Text Preview": 80,
    }
    for index, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(header, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("Summary")
    parsed_count = sum(1 for row in rows if row.get("Status") == "Parsed")
    summary.append(["Generated At", datetime.now().strftime("%d-%B-%Y %H:%M:%S")])
    summary.append(["Rows", len(rows)])
    summary.append(["Parsed", parsed_count])
    summary.append(["Not Parsed / Errors", len(rows) - parsed_count])
    for cell in summary["A"]:
        cell.font = Font(bold=True)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 30

    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract invoice PDF fields into an Excel review workbook.")
    parser.add_argument("source", help="PDF file or directory containing PDF files.")
    parser.add_argument("-o", "--output", default="artifacts/invoice_extract.xlsx", help="Output .xlsx path.")
    parser.add_argument("--recursive", action="store_true", help="Search directories recursively.")
    args = parser.parse_args()

    source = Path(args.source)
    output = Path(args.output)
    pdfs = _pdf_paths(source, args.recursive)
    if not pdfs:
        print(f"No PDF files found at: {source}")
        return 2

    rows = extract_invoices(pdfs)
    write_excel(rows, output)
    parsed_count = sum(1 for row in rows if row.get("Status") == "Parsed")
    print(f"PDF files scanned: {len(pdfs)}")
    print(f"Rows written: {len(rows)}")
    print(f"Parsed invoices: {parsed_count}")
    print(f"Review workbook: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
